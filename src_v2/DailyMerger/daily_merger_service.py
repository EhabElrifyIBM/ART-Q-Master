"""
Daily Case Merger — Service Layer
==================================

Pure business logic for merging up to 30 daily Active Cases workbooks.

File naming convention:  Active Cases PA MM-DD.xlsx
Sheet conventions:
  - Header row  : row 1
  - Data rows   : row 2 onwards
  - Case number : column A (index 0) for handler + companies sheets
  - Case number : column L (index 11) for Chat Agent's Cases sheet

Processing order per daily file:
  1. All XXXX's Cases sheets (alphabetical, excluding Chat Agent's Cases)
  2. Chat Agent's Cases
  3. Companies

Deduplication: global across all sheets and all days.
  Later day always overwrites earlier day for the same case number.

ZIP support:
  When accept_zip=True (default), .zip files that contain exactly one .xlsx
  workbook matching the Active Cases PA MM-DD.xlsx naming convention are
  accepted.  The workbook is extracted to a temporary file for processing
  and deleted after the merge completes.
"""

from __future__ import annotations

import io
import re
import tempfile
import zipfile
from collections import defaultdict
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Types
# ---------------------------------------------------------------------------

ProgressCallback = Optional[Callable[[int, str], None]]

# Sentinel used when a row's key cell is empty
_EMPTY_KEY = object()

# Column index (0-based) for case numbers
_CASE_COL_HANDLER = 0    # Column A
_CASE_COL_CHAT    = 11   # Column L
_CASE_COL_COMPANY = 0    # Column A

# Sheet-name patterns
_CHAT_SHEET      = "Chat Agent's Cases"
_COMPANY_SHEET   = "Companies"
_HANDLER_SUFFIX  = "'s Cases"
_ALL_CASES_SHEET = "All Cases"      # produced by a prior merge — treated as handler data


def _report(cb: ProgressCallback, percent: int, msg: str, level: str = "INFO") -> None:
    """Fire progress callback with an optional severity level prefix.

    The *level* prefix is stripped from the progress-bar message but preserved
    in the log line that the worker thread emits (see DailyMergeWorker).
    Levels: INFO (default) | WARN | ERROR | DETAIL
    """
    if cb:
        try:
            cb(percent, f"[{level}] {msg}")
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class DailyFile:
    """Metadata for a single daily workbook."""

    path: Path
    date_str: str          # "MM-DD" extracted from filename
    month: str             # "MM"
    day: int               # numeric day for chronological sort
    sheet_names: List[str] = field(default_factory=list)

    # Derived
    handler_sheets: List[str] = field(default_factory=list)
    has_chat: bool = False
    has_companies: bool = False

    # ZIP source info
    source_zip: Optional[Path] = None     # set when extracted from a .zip
    temp_path:  Optional[Path] = None     # temp file path used during merge
    zip_xlsx_count: int = 0               # total .xlsx files found inside the zip


@dataclass
class ValidationResult:
    """Result of the file validation / scan step."""

    daily_files: List[DailyFile]         # sorted chronologically
    months_found: Dict[str, int]         # month → file count
    handler_names: List[str]             # union across all files (sorted)
    warnings: List[str]                  # missing-sheet warnings per day
    skipped_months: List[str]            # months skipped in the loaded range (e.g. ["02"])
    is_valid: bool = True
    error: str = ""


@dataclass
class MergeConfig:
    """Configuration passed to DailyMergerService.merge()."""

    file_paths: List[Path]
    output_path: Path
    accept_zip: bool = True          # whether to expand .zip files
    relaxed_filename: bool = False   # skip strict "Active Cases PA MM-DD" check


@dataclass
class MergeStats:
    """Statistics produced after a successful merge."""

    files_processed: int = 0
    all_cases_count: int = 0
    chat_cases_count: int = 0
    companies_count: int = 0
    skipped_empty_keys: int = 0
    zip_files_expanded: int = 0


@dataclass
class MergeResult:
    """Result of the merge operation."""

    success: bool
    message: str
    output_path: Optional[Path] = None
    stats: MergeStats = field(default_factory=MergeStats)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _parse_date_from_filename(filename: str) -> Optional[Tuple[str, str, int]]:
    """
    Extract (month_str, date_str, day_int) from a filename like
    "Active Cases PA 01-15.xlsx".

    Returns None if the pattern is not found.
    """
    # Match MM-DD anywhere in the stem
    match = re.search(r"(\d{1,2})-(\d{1,2})", Path(filename).stem)
    if not match:
        return None
    month_str = match.group(1).zfill(2)
    day_str   = match.group(2).zfill(2)
    date_str  = f"{month_str}-{day_str}"
    return month_str, date_str, int(day_str)


def _parse_month_label_from_filename(filename: str) -> Optional[Tuple[str, str, str]]:
    """
    Extract a representative sort key from a monthly merged filename like
    "Merged Cases 01-05 to 01-30.xlsx" or "Merged Cases 03-02 to 03-31.xlsx".

    Returns (sort_key, start_date_str, label) where sort_key is "MM-DD" of the
    first date found (used for chronological ordering), or None if no date found.

    The caller uses this as a drop-in replacement for _parse_date_from_filename
    when working with monthly files.
    """
    stem = Path(filename).stem
    # Find all MM-DD occurrences in the stem
    dates = re.findall(r"(\d{1,2})-(\d{1,2})", stem)
    if not dates:
        return None
    # Use the first date as the sort anchor (start of month range)
    m1, d1 = dates[0]
    month_str = m1.zfill(2)
    day_str   = d1.zfill(2)
    date_str  = f"{month_str}-{day_str}"
    # Human label: use both dates if range present
    if len(dates) >= 2:
        m2, d2 = dates[-1]
        label = f"{month_str}-{day_str} → {m2.zfill(2)}-{d2.zfill(2)}"
    else:
        label = date_str
    return month_str, date_str, label


def _classify_sheets(
    sheet_names: List[str],
) -> Tuple[List[str], bool, bool]:
    """
    Classify sheet names into (handler_sheets, has_chat, has_companies).

    Handler sheets: ends with "'s Cases" but is NOT "Chat Agent's Cases",
    OR is the special pre-merged "All Cases" sheet produced by a prior run.
    Both types use column A as the case-number key and feed all_cases_map.
    """
    handler_sheets = []
    has_chat = False
    has_companies = False

    for name in sheet_names:
        if name == _CHAT_SHEET:
            has_chat = True
        elif name == _COMPANY_SHEET:
            has_companies = True
        elif name == _ALL_CASES_SHEET:
            # Pre-merged "All Cases" sheet — treat exactly like a handler sheet
            handler_sheets.append(name)
        elif name.endswith(_HANDLER_SUFFIX):
            handler_sheets.append(name)

    handler_sheets.sort()
    return handler_sheets, has_chat, has_companies


def _get_row_values(ws: Worksheet, row_idx: int) -> List[Any]:
    """Return all cell values for a given 1-based row index."""
    return [cell.value for cell in ws[row_idx]]


def _get_row_cells(ws: Worksheet, row_idx: int):
    """Return all cells for a given 1-based row index."""
    return list(ws[row_idx])


def _count_xlsx_in_zip(zip_path: Path) -> int:
    """
    Return the total number of .xlsx files inside a ZIP archive
    (excluding macOS artefacts).  Returns 0 on any error.
    """
    try:
        with zipfile.ZipFile(str(zip_path), "r") as zf:
            return sum(
                1 for name in zf.namelist()
                if name.lower().endswith(".xlsx")
                and not name.startswith("__MACOSX")
                and not Path(name).name.startswith(".")
            )
    except Exception:
        return 0


def _extract_xlsx_from_zip(zip_path: Path) -> Optional[Tuple[Path, str, int]]:
    """
    Look inside a .zip file for a single .xlsx whose name matches the
    Active Cases PA MM-DD pattern.

    Returns (temp_path, original_xlsx_name, total_xlsx_count) on success,
    None on failure.
    The caller is responsible for deleting temp_path when done.
    """
    if not zipfile.is_zipfile(str(zip_path)):
        return None

    try:
        with zipfile.ZipFile(str(zip_path), "r") as zf:
            # Find all .xlsx members (ignore macOS __MACOSX artefacts)
            candidates = [
                name for name in zf.namelist()
                if name.lower().endswith(".xlsx")
                and not name.startswith("__MACOSX")
                and not Path(name).name.startswith(".")
            ]
            total_count = len(candidates)
            if not candidates:
                return None

            # Prefer files matching the expected naming pattern; fall back to
            # any single xlsx if exactly one is present.
            matched = [c for c in candidates if _parse_date_from_filename(Path(c).name)]
            chosen = matched[0] if matched else (candidates[0] if len(candidates) == 1 else None)
            if chosen is None:
                return None

            xlsx_name = Path(chosen).name
            data = zf.read(chosen)

        # Write to a uniquely-named temp file (auto-deleted by caller)
        tmp = tempfile.NamedTemporaryFile(
            suffix=".xlsx", prefix="daily_merger_", delete=False
        )
        tmp.write(data)
        tmp.close()
        return Path(tmp.name), xlsx_name, total_count

    except Exception:
        return None


def _detect_skipped_months(months_found: Dict[str, int]) -> List[str]:
    """
    Given a dict of month strings that ARE present (e.g. {"01":5, "03":10}),
    return any months that fall within the min→max range but are absent.

    E.g. if months_found = {"01": 5, "03": 10}  →  returns ["02"]
    """
    if len(months_found) < 2:
        return []
    month_ints = sorted(int(m) for m in months_found)
    skipped = []
    for m in range(month_ints[0] + 1, month_ints[-1]):
        if m not in month_ints:
            skipped.append(f"{m:02d}")
    return skipped


def _is_zip_file(path: Path) -> bool:
    return path.suffix.lower() == ".zip" and path.exists()


def _fast_sheet_names(xlsx_path: Path) -> Optional[List[str]]:
    """
    Return sheet names from an .xlsx file WITHOUT loading cell data.

    An .xlsx is a ZIP archive — we read only  xl/workbook.xml  which is tiny
    (~2-5 KB) regardless of workbook size.  This is ~100x faster than
    openpyxl.load_workbook() on large files.

    Returns None on any error (caller should fall back to openpyxl).
    """
    import xml.etree.ElementTree as ET
    _NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            if "xl/workbook.xml" not in zf.namelist():
                return None
            xml_bytes = zf.read("xl/workbook.xml")
        root = ET.fromstring(xml_bytes)
        sheets = []
        for sheet in root.iter(f"{{{_NS}}}sheet"):
            name = sheet.get("name")
            if name:
                sheets.append(name)
        return sheets if sheets else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class DailyMergerService:
    """
    Merges multiple daily Active Cases workbooks into one deduplicated output.

    Accepts both plain .xlsx files and .zip archives containing a single
    .xlsx workbook when accept_zip=True (the default).

    Usage::

        svc = DailyMergerService()
        validation = svc.validate_files([Path("Active Cases PA 01-01.xlsx"), ...])
        if validation.is_valid:
            config = MergeConfig(file_paths=[...], output_path=Path("output.xlsx"))
            result = svc.merge(config, progress_callback=my_cb)
    """

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def validate_files(
        self,
        paths: List[Path],
        accept_zip: bool = True,
        relaxed_filename: bool = False,
    ) -> ValidationResult:
        """
        Scan each workbook's sheet names (light-weight — no full data load)
        and return a ValidationResult with metadata, month groupings, and
        any per-day warnings about missing expected sheets.

        Args:
            paths:             List of .xlsx or .zip file paths.
            accept_zip:        When True (default), .zip files are expanded.
            relaxed_filename:  When True, any .xlsx is accepted regardless of
                               filename — date is still parsed from the name
                               but a parse failure is a warning, not an error.
                               Used for monthly merged files.
        """
        if not paths:
            return ValidationResult(
                daily_files=[],
                months_found={},
                handler_names=[],
                warnings=[],
                skipped_months=[],
                is_valid=False,
                error="No files provided.",
            )

        daily_files: List[DailyFile] = []
        parse_errors: List[str] = []
        # Track temp files created during validation so we can clean them up
        _temp_files: List[Path] = []

        for p in paths:
            p = Path(p)
            if not p.exists():
                parse_errors.append(f"File not found: {p.name}")
                continue

            source_zip: Optional[Path] = None
            effective_path = p
            zip_xlsx_count = 0  # only non-zero for ZIP sources

            if p.suffix.lower() == ".zip":
                if not accept_zip:
                    parse_errors.append(f"ZIP files not enabled: {p.name}")
                    continue
                result = _extract_xlsx_from_zip(p)
                if result is None:
                    parse_errors.append(
                        f"No matching .xlsx found inside ZIP: {p.name}"
                    )
                    continue
                effective_path, xlsx_name, zip_xlsx_count = result
                source_zip = p
                _temp_files.append(effective_path)
                # Use the inner xlsx name for date parsing
                p_for_parsing = xlsx_name
            else:
                p_for_parsing = p.name

            if effective_path.suffix.lower() not in (".xlsx", ".xls"):
                parse_errors.append(f"Not an Excel file: {p.name}")
                _cleanup_temp(effective_path, source_zip)
                continue

            parsed = _parse_date_from_filename(p_for_parsing)
            if not parsed:
                if relaxed_filename:
                    # Try the monthly range pattern "Merged Cases MM-DD to MM-DD"
                    monthly = _parse_month_label_from_filename(p_for_parsing)
                    if monthly:
                        month_str, date_str, _ = monthly
                        day_int = int(date_str.split("-")[1])
                        parsed = (month_str, date_str, day_int)
                    else:
                        # No date at all — use a placeholder so file is included
                        month_str  = "00"
                        date_str   = "00-00"
                        day_int    = 0
                        parsed = (month_str, date_str, day_int)
                else:
                    parse_errors.append(
                        f"Cannot parse date from filename: {p.name} "
                        f"(expected format: Active Cases PA MM-DD.xlsx)"
                    )
                    _cleanup_temp(effective_path, source_zip)
                    continue

            month_str, date_str, day_int = parsed

            # Read sheet names only — use the fast XML reader first.
            # _fast_sheet_names() reads only xl/workbook.xml (~2-5 KB) from the
            # OOXML ZIP without parsing any cell data: typically < 5 ms per file.
            try:
                sheet_names = _fast_sheet_names(effective_path)
                if sheet_names is None:
                    # Rare fallback for malformed files
                    wb = openpyxl.load_workbook(
                        str(effective_path), read_only=True, data_only=True
                    )
                    sheet_names = wb.sheetnames
                    wb.close()
            except Exception as exc:
                parse_errors.append(f"Cannot open {p.name}: {exc}")
                _cleanup_temp(effective_path, source_zip)
                continue

            handler_sheets, has_chat, has_companies = _classify_sheets(sheet_names)

            daily_files.append(DailyFile(
                path=effective_path,
                date_str=date_str,
                month=month_str,
                day=day_int,
                sheet_names=sheet_names,
                handler_sheets=handler_sheets,
                has_chat=has_chat,
                has_companies=has_companies,
                source_zip=source_zip,
                zip_xlsx_count=zip_xlsx_count if source_zip else 0,
            ))

        if parse_errors:
            return ValidationResult(
                daily_files=[],
                months_found={},
                handler_names=[],
                warnings=parse_errors,
                skipped_months=[],
                is_valid=False,
                error="; ".join(parse_errors),
            )

        # Sort chronologically: by month first, then by day
        daily_files.sort(key=lambda f: (int(f.month), f.day))

        # Group by month
        months_found: Dict[str, int] = defaultdict(int)
        for df in daily_files:
            months_found[df.month] += 1

        # Union of all handler names across all files
        all_handler_names: set = set()
        for df in daily_files:
            for sh in df.handler_sheets:
                # Strip "'s Cases" suffix to get handler name
                handler_name = sh[: -len(_HANDLER_SUFFIX)]
                all_handler_names.add(handler_name)

        # Determine the expected set of handler sheets (union across all files)
        expected_handlers = sorted(all_handler_names)
        expected_handler_sheets = [f"{h}{_HANDLER_SUFFIX}" for h in expected_handlers]

        # Detect skipped months in the loaded range
        skipped_months = _detect_skipped_months(dict(months_found))

        # Build per-day warnings for missing sheets
        warnings: List[str] = []
        for df in daily_files:
            missing: List[str] = []
            for sh in expected_handler_sheets:
                if sh not in df.handler_sheets:
                    missing.append(sh)
            if not df.has_chat:
                missing.append(_CHAT_SHEET)
            if not df.has_companies:
                missing.append(_COMPANY_SHEET)
            if missing:
                warnings.append(
                    f"Day {df.date_str} missing: {', '.join(missing)}"
                )

        return ValidationResult(
            daily_files=daily_files,
            months_found=dict(months_found),
            handler_names=expected_handlers,
            warnings=warnings,
            skipped_months=skipped_months,
            is_valid=True,
        )

    def merge(
        self,
        config: MergeConfig,
        progress_callback: ProgressCallback = None,
    ) -> MergeResult:
        """
        Validate, load, deduplicate, and write the merged output workbook.

        Processing order per file (chronological — oldest first):
          handler sheets (alphabetical) → Chat Agent's Cases → Companies

        Deduplication: upsert by case number — later day always wins.
        """
        cb = progress_callback

        _report(cb, 0, "Validating files…")
        validation = self.validate_files(config.file_paths, accept_zip=config.accept_zip)
        if not validation.is_valid:
            return MergeResult(success=False, message=validation.error)

        daily_files = validation.daily_files
        total_files = len(daily_files)
        if total_files == 0:
            return MergeResult(success=False, message="No valid daily files to merge.")

        stats = MergeStats()
        stats.zip_files_expanded = sum(1 for df in daily_files if df.source_zip is not None)

        # ------------------------------------------------------------------
        # In-memory maps: case_number (str) → list of cell values (row data)
        # We also store the header row once per sheet type.
        # ------------------------------------------------------------------
        all_cases_map:   Dict[str, List[Any]] = {}
        chat_map:        Dict[str, List[Any]] = {}
        companies_map:   Dict[str, List[Any]] = {}

        # Header rows (captured from the first file that has the sheet)
        all_cases_header:   Optional[List[Any]] = None
        chat_header:        Optional[List[Any]] = None
        companies_header:   Optional[List[Any]] = None

        # ------------------------------------------------------------------
        # Process each daily file in chronological order
        # ------------------------------------------------------------------
        for file_idx, df in enumerate(daily_files):
            base_pct = int(file_idx / total_files * 85)  # 0–85% for loading
            label = df.source_zip.name if df.source_zip else df.path.name

            # ── File banner ───────────────────────────────────────────
            _report(
                cb, base_pct,
                f"── Day {file_idx + 1}/{total_files}  [{df.date_str}]  {label}",
            )

            if df.source_zip:
                _report(cb, base_pct, f"Extracted from ZIP: {df.source_zip.name}", "DETAIL")

            try:
                wb = openpyxl.load_workbook(str(df.path), data_only=True, read_only=True)
            except Exception as exc:
                _report(cb, base_pct, f"Cannot open {label}: {exc}", "ERROR")
                _cleanup_temp_files(daily_files)
                return MergeResult(
                    success=False,
                    message=f"Cannot open {label}: {exc}",
                )

            _report(
                cb, base_pct,
                f"Opened workbook — sheets found: {', '.join(wb.sheetnames) or '(none)'}",
                "DETAIL",
            )

            # ── Handler sheets ────────────────────────────────────────
            if df.handler_sheets:
                _report(
                    cb, base_pct,
                    f"Reading {len(df.handler_sheets)} handler sheet(s): "
                    + ", ".join(df.handler_sheets),
                    "DETAIL",
                )

            for sh_name in df.handler_sheets:
                if sh_name not in wb.sheetnames:
                    _report(cb, base_pct, f"Sheet not found (skipped): {sh_name}", "WARN")
                    continue
                ws = wb[sh_name]
                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    _report(cb, base_pct, f"Sheet is empty (skipped): {sh_name}", "WARN")
                    continue

                # Row 1 (index 0) = header
                if all_cases_header is None:
                    all_cases_header = list(rows[0])

                # Data from row 2 onwards (index 1+)
                new_rows = 0
                updated_rows = 0
                for row in rows[1:]:
                    key = _extract_key(row, _CASE_COL_HANDLER)
                    if key is None:
                        stats.skipped_empty_keys += 1
                        continue
                    if key in all_cases_map:
                        updated_rows += 1
                    else:
                        new_rows += 1
                    all_cases_map[key] = list(row)

                _report(
                    cb, base_pct,
                    f"  {sh_name}: {len(rows) - 1} data row(s) read"
                    + (f" — {new_rows} new, {updated_rows} overwritten" if updated_rows else ""),
                    "DETAIL",
                )

            # ── Chat Agent's Cases ────────────────────────────────────
            if df.has_chat and _CHAT_SHEET in wb.sheetnames:
                ws = wb[_CHAT_SHEET]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    if chat_header is None:
                        chat_header = list(rows[0])
                    new_rows = 0
                    updated_rows = 0
                    for row in rows[1:]:
                        key = _extract_key(row, _CASE_COL_CHAT)
                        if key is None:
                            stats.skipped_empty_keys += 1
                            continue
                        if key in chat_map:
                            updated_rows += 1
                        else:
                            new_rows += 1
                        chat_map[key] = list(row)
                    _report(
                        cb, base_pct,
                        f"  Chat Agent's Cases: {len(rows) - 1} data row(s) read"
                        + (f" — {new_rows} new, {updated_rows} overwritten" if updated_rows else ""),
                        "DETAIL",
                    )
            elif df.has_chat:
                _report(cb, base_pct, f"Chat Agent's Cases sheet missing in {label}", "WARN")

            # ── Companies ─────────────────────────────────────────────
            if df.has_companies and _COMPANY_SHEET in wb.sheetnames:
                ws = wb[_COMPANY_SHEET]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    if companies_header is None:
                        companies_header = list(rows[0])
                    new_rows = 0
                    updated_rows = 0
                    for row in rows[1:]:
                        key = _extract_key(row, _CASE_COL_COMPANY)
                        if key is None:
                            stats.skipped_empty_keys += 1
                            continue
                        if key in companies_map:
                            updated_rows += 1
                        else:
                            new_rows += 1
                        companies_map[key] = list(row)
                    _report(
                        cb, base_pct,
                        f"  Companies: {len(rows) - 1} data row(s) read"
                        + (f" — {new_rows} new, {updated_rows} overwritten" if updated_rows else ""),
                        "DETAIL",
                    )
            elif df.has_companies:
                _report(cb, base_pct, f"Companies sheet missing in {label}", "WARN")

            wb.close()
            stats.files_processed += 1
            _report(
                cb, base_pct,
                f"Day {df.date_str} done — running totals: "
                f"{len(all_cases_map)} cases · {len(chat_map)} chat · {len(companies_map)} companies",
            )

        # ------------------------------------------------------------------
        # Write output workbook
        # ------------------------------------------------------------------
        # Clean up any temp files extracted from ZIPs (do this after all
        # workbooks are closed but before we report completion)
        _cleanup_temp_files(daily_files)
        _report(cb, 87, "── Writing output workbook…")
        _report(cb, 87,
                f"Deduplication complete — "
                f"{len(all_cases_map)} unique cases · "
                f"{len(chat_map)} unique chat · "
                f"{len(companies_map)} unique companies",
                "DETAIL")

        try:
            wb_out = openpyxl.Workbook()
            wb_out.remove(wb_out.active)  # remove default empty sheet

            # ── All Cases ────────────────────────────────────────────
            _report(cb, 89, "Writing sheet: All Cases…", "DETAIL")
            ws_all = wb_out.create_sheet("All Cases")
            if all_cases_header:
                ws_all.append(all_cases_header)
            for row_data in all_cases_map.values():
                ws_all.append(_clean_row(row_data))
            stats.all_cases_count = len(all_cases_map)
            _report(cb, 91, f"✔ All Cases: {stats.all_cases_count} unique rows written")

            # ── Chat Agent's Cases ───────────────────────────────────
            _report(cb, 93, "Writing sheet: Chat Agent's Cases…", "DETAIL")
            ws_chat = wb_out.create_sheet(_CHAT_SHEET)
            if chat_header:
                ws_chat.append(chat_header)
            for row_data in chat_map.values():
                ws_chat.append(_clean_row(row_data))
            stats.chat_cases_count = len(chat_map)
            _report(cb, 95, f"✔ Chat Agent's Cases: {stats.chat_cases_count} unique rows written")

            # ── Companies ────────────────────────────────────────────
            _report(cb, 96, "Writing sheet: Companies…", "DETAIL")
            ws_comp = wb_out.create_sheet(_COMPANY_SHEET)
            if companies_header:
                ws_comp.append(companies_header)
            for row_data in companies_map.values():
                ws_comp.append(_clean_row(row_data))
            stats.companies_count = len(companies_map)
            _report(cb, 98, f"✔ Companies: {stats.companies_count} unique rows written")

            # Save
            output_path = Path(config.output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb_out.save(str(output_path))
            wb_out.close()

        except Exception as exc:
            return MergeResult(success=False, message=f"Failed to write output: {exc}")

        _report(cb, 100, "Merge complete.")
        total_unique = stats.all_cases_count + stats.chat_cases_count + stats.companies_count
        zip_note = (
            f", {stats.zip_files_expanded} from ZIP"
            if stats.zip_files_expanded else ""
        )
        return MergeResult(
            success=True,
            message=(
                f"Merged {stats.files_processed} daily files{zip_note} → "
                f"{total_unique} unique rows "
                f"({stats.all_cases_count} cases, "
                f"{stats.chat_cases_count} chat, "
                f"{stats.companies_count} companies)"
            ),
            output_path=output_path,
            stats=stats,
        )


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _cleanup_temp(path: Path, source_zip: Optional[Path]) -> None:
    """Delete a temp file if it was extracted from a ZIP."""
    if source_zip is not None and path.exists():
        try:
            path.unlink()
        except Exception:
            pass


def _cleanup_temp_files(daily_files: List["DailyFile"]) -> None:
    """Delete all temp files that were extracted from ZIPs."""
    for df in daily_files:
        _cleanup_temp(df.path, df.source_zip)


def _extract_key(row: tuple, col_index: int) -> Optional[str]:
    """
    Extract the case-number key from a row tuple at *col_index*.

    Returns None for empty / whitespace-only cells so those rows are skipped.
    """
    if col_index >= len(row):
        return None
    val = row[col_index]
    if val is None:
        return None
    key = str(val).strip()
    return key if key else None


def _clean_row(row_data: List[Any]) -> List[Any]:
    """Return a clean copy of a row, converting None to empty string for safety."""
    return [v if v is not None else "" for v in row_data]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
__all__ = [
    "DailyFile",
    "ValidationResult",
    "MergeConfig",
    "MergeResult",
    "MergeStats",
    "DailyMergerService",
    # Fast utilities (used by UI for table scan)
    "_fast_sheet_names",
    "_classify_sheets",
    "_parse_date_from_filename",
    "_parse_month_label_from_filename",
    "_extract_xlsx_from_zip",
]

# Made with Bob
