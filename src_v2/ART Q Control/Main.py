import platform
import subprocess
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from datetime import datetime, timedelta
import tkinter as tk
from openpyxl import load_workbook
import time
import json
import os
import pandas as pd
import pyautogui
import traceback
import ctypes
from PyQt5.QtWidgets import (
    QApplication, QDialog, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QGridLayout, QInputDialog, QWidget, QCheckBox, QScrollArea, QProgressBar, QFrame
)
from PyQt5.QtCore import Qt, QPropertyAnimation, QPoint, QTimer, QRect, pyqtSignal
from PyQt5.QtGui import QIcon, QCursor
import sys
from config_loader import init_config
from pathlib import Path


# ============================================================================
# CONFIGURATION INITIALIZATION - RUNS FIRST
# ============================================================================
print("[INFO] Initializing configuration system...")
CONFIG_MANAGER = init_config()

# Load configuration values - NO FALLBACKS, strict validation
try:
    AGENT_NAME = CONFIG_MANAGER.get_value('agent_settings', 'agent_name')
    DIALER_USERNAME = CONFIG_MANAGER.get_value('agent_settings', 'user_id')
    DIALER_PASSWORD = CONFIG_MANAGER.get_value('agent_settings', 'password')
    DIALER_PLACE_ID = CONFIG_MANAGER.get_value('agent_settings', 'place_id')
    
    EXCEL_BASE_PATH = CONFIG_MANAGER.get_value('file_paths', 'excel_base_path')
    CACHE_DIRECTORY = CONFIG_MANAGER.get_value('file_paths', 'cache_directory')
    
    EXCEL_SHEET_NAME = CONFIG_MANAGER.get_value('crm_settings', 'excel_sheet_name')
    
    START_TIME_STR = CONFIG_MANAGER.get_value('execution_settings', 'start_time')
    END_TIME_STR = CONFIG_MANAGER.get_value('execution_settings', 'end_time')
    REFRESH_INTERVAL = CONFIG_MANAGER.get_value('execution_settings', 'refresh_interval')
    
    # Parse time strings (format: "HH:MM")
    START_HOUR, START_MINUTE = map(int, START_TIME_STR.split(':'))
    END_HOUR, END_MINUTE = map(int, END_TIME_STR.split(':'))
    
    print(f"[INFO] Configuration loaded for agent: {AGENT_NAME}")
    
except Exception as e:
    print(f"[CRITICAL ERROR] Failed to load configuration: {e}")
    sys.exit(1)

# ============================================================================
# HARDCODED SETTINGS (Do not change)
# ============================================================================
DIALER_URL = "https://104.232.254.43/ui/ad/v1/index.html"
EMAIL_DOMAIN = "na_thinkcare@lenovo.com"
EMAIL_DOMAIN_NAME = "NA Think Care"

# ============================================================================
# HELPER FUNCTION: Safe Excel Sheet Update (Preserves Other Sheets)
# ============================================================================
def update_excel_sheet_safely(excel_path, dataframe, sheet_name):
    """
    Update a specific sheet in an Excel workbook while preserving all other sheets.
    
    Args:
        excel_path: Path to the Excel file
        dataframe: Pandas DataFrame with updated data
        sheet_name: Name of the sheet to update
    """
    try:
        # Load existing workbook to preserve other sheets
        workbook = load_workbook(excel_path)
        
        # If sheet exists, delete it so we can recreate it with new data
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        
        # Close workbook without saving
        workbook.close()
        
        # Use pandas ExcelWriter to write the specific sheet while keeping others
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"[INFO] Excel sheet '{sheet_name}' updated successfully. Other sheets preserved.")
        
    except FileNotFoundError:
        # File doesn't exist yet, create it normally
        dataframe.to_excel(excel_path, sheet_name=sheet_name, index=False)
        print(f"[INFO] Excel file created: {excel_path}")
    except Exception as e:
        print(f"[ERROR] Failed to update Excel sheet: {e}")
        # Fallback to normal write if safe update fails
        dataframe.to_excel(excel_path, sheet_name=sheet_name, index=False)

def update_cache_file(cache_path, dataframe, sheet_name):
    """
    Fast update to cache file (much smaller, no other sheets to preserve)
    
    Args:
        cache_path: Path to the cache Excel file
        dataframe: Pandas DataFrame with working data
        sheet_name: Name of the sheet to update
    """
    try:
        dataframe.to_excel(cache_path, sheet_name=sheet_name, index=False)
        print(f"[DEBUG] Cache file updated: {len(dataframe)} rows")
    except Exception as e:
        print(f"[ERROR] Failed to update cache file: {e}")
        traceback.print_exc()

# ============================================================================
# WELCOME DIALOG
# ============================================================================
def convert_to_12hour_format(time_str):
    """Convert HH:MM (24-hour) to HH:MM AM/PM format"""
    try:
        hour, minute = map(int, time_str.split(':'))
        am_pm = "AM" if hour < 12 else "PM"
        hour_12 = hour if hour <= 12 else hour - 12
        if hour_12 == 0:
            hour_12 = 12
        return f"{hour_12:02d}:{minute:02d} {am_pm}"
    except:
        return time_str

def show_welcome_dialog():
    """
    Display welcome dialog with config information and options to start or update config.
    Returns: (proceed, skip_timer)
      - (True, False) = Start with timer
      - (True, True) = Start and skip timer
      - (False, False) = Update config
    """
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    dialog = QDialog()
    dialog.setWindowTitle("Welcome - ART Automation")
    dialog.setGeometry(100, 100, 700, 500)
    dialog.setStyleSheet("""
        QDialog {
            background-color: white;
        }
    """)
    
    layout = QVBoxLayout()
    layout.setContentsMargins(20, 20, 20, 20)
    layout.setSpacing(15)
    
    # Title
    title = QLabel("=== ART AUTOMATION SYSTEM ===")
    title.setStyleSheet("font-weight: bold; font-size: 16px; color: #1976D2;")
    layout.addWidget(title)
    
    # Configuration Information
    config_info = QLabel(
        f"<b>Current Configuration:</b><br><br>"
        f"<table cellpadding='8'>"
        f"<tr><td><b>Agent Name:</b></td><td>{AGENT_NAME}</td></tr>"
        f"<tr><td><b>User ID:</b></td><td>{DIALER_USERNAME}</td></tr>"
        f"<tr><td><b>Excel Base Path:</b></td><td>{EXCEL_BASE_PATH}</td></tr>"
        f"<tr><td><b>Cache Directory:</b></td><td>{CACHE_DIRECTORY}</td></tr>"
        f"<tr><td><b>Sheet Name:</b></td><td>{EXCEL_SHEET_NAME}</td></tr>"
        f"<tr><td><b>Start Time:</b></td><td>{convert_to_12hour_format(START_TIME_STR)}</td></tr>"
        f"<tr><td><b>End Time:</b></td><td>{convert_to_12hour_format(END_TIME_STR)}</td></tr>"
        f"<tr><td><b>Refresh Interval:</b></td><td>{REFRESH_INTERVAL} cases</td></tr>"
        f"</table><br>"
        f"<i>Please verify the above information is correct before proceeding.</i>"
    )
    config_info.setStyleSheet("""
        QLabel {
            background-color: #F5F5F5;
            color: #333333;
            padding: 15px;
            border-radius: 5px;
            border: 1px solid #CCCCCC;
        }
    """)
    config_info.setWordWrap(True)
    layout.addWidget(config_info)
    
    # Buttons
    button_layout = QVBoxLayout()
    button_layout.setSpacing(10)
    
    start_btn = QPushButton("✓ Start Process (with Timer)")
    start_btn.setStyleSheet("""
        QPushButton {
            background-color: #4CAF50;
            color: white;
            padding: 12px;
            font-weight: bold;
            border-radius: 5px;
            border: none;
            font-size: 15px;
        }
        QPushButton:hover {
            background-color: #45a049;
        }
        QPushButton:pressed {
            background-color: #3d8b40;
        }
    """)
    start_btn.clicked.connect(lambda: dialog.done(1))
    button_layout.addWidget(start_btn)
    
    skip_timer_btn = QPushButton("⏭ Start Now (Skip Timer)")
    skip_timer_btn.setStyleSheet("""
        QPushButton {
            background-color: #FF9800;
            color: white;
            padding: 12px;
            font-weight: bold;
            border-radius: 5px;
            border: none;
            font-size: 15px;
        }
        QPushButton:hover {
            background-color: #FB8C00;
        }
        QPushButton:pressed {
            background-color: #F57C00;
        }
    """)
    skip_timer_btn.clicked.connect(lambda: dialog.done(3))
    button_layout.addWidget(skip_timer_btn)
    
    update_btn = QPushButton("⚙ Update Configuration")
    update_btn.setStyleSheet("""
        QPushButton {
            background-color: #2196F3;
            color: white;
            padding: 12px;
            font-weight: bold;
            border-radius: 5px;
            border: none;
            font-size: 15px;
        }
        QPushButton:hover {
            background-color: #0b7dda;
        }
        QPushButton:pressed {
            background-color: #0a68c4;
        }
    """)
    update_btn.clicked.connect(lambda: dialog.done(2))
    button_layout.addWidget(update_btn)
    
    main_menu_btn = QPushButton("☰ Main Menu")
    main_menu_btn.setStyleSheet("""
        QPushButton {
            background-color: #607D8B;
            color: white;
            padding: 12px;
            font-weight: bold;
            border-radius: 5px;
            border: none;
            font-size: 15px;
        }
        QPushButton:hover {
            background-color: #455A64;
        }
        QPushButton:pressed {
            background-color: #37474F;
        }
    """)
    main_menu_btn.clicked.connect(lambda: dialog.done(4))
    button_layout.addWidget(main_menu_btn)
    
    layout.addLayout(button_layout)
    
    # Support Mode Checkbox
    from PyQt5.QtWidgets import QCheckBox
    support_checkbox = QCheckBox("🤝 Supporting another agent")
    support_checkbox.setStyleSheet("font-size: 15px; padding: 8px; color: #161616;")
    layout.addWidget(support_checkbox)

    # Footer
    footer_label = QLabel(
        '<span style="color:#161616; font-size: 15px;">Developed by: Ehab Elrify | Adam Maged <br>'
        'Email: <a href="mailto:ehab.elrify@ibm.com" style="color:#0f62fe;">ehab.elrify@ibm.com</a> | '
        '<a href="mailto:abdelrahman.maged@ibm.com" style="color:#0f62fe;">abdelrahman.maged@ibm.com</a><br>'
        'Assurance Resolution Team</span>'
    )
    footer_label.setAlignment(Qt.AlignCenter)
    footer_label.setOpenExternalLinks(True)
    footer_label.setStyleSheet("padding-top: 10px;")
    layout.addWidget(footer_label)
    
    dialog.setLayout(layout)
    result = dialog.exec_()
    
    # Check if support mode was enabled
    support_agent = None
    if support_checkbox.isChecked() and result in [1, 3]:
        # Prompt for the supported agent's name
        from PyQt5.QtWidgets import QInputDialog
        agent_name, ok = QInputDialog.getText(
            None, "Support Mode", 
            "Enter the name of the agent you are supporting:\n(This will be used for the sheet name: '[AgentName]'s Cases')")
        if ok and agent_name.strip():
            support_agent = agent_name.strip()
            print(f"[INFO] Support mode enabled for agent: {support_agent}")
        else:
            support_agent = None  # Cancelled or empty
    
    if result == 2:  # Update config
        # Import and show config dialog
        from config_loader import create_config_setup_dialog
        config_dialog = create_config_setup_dialog(CONFIG_MANAGER)
        config_dialog.exec_()
        # After config update, show welcome again
        return show_welcome_dialog()
    elif result == 3:  # Start and skip timer
        return (True, True, support_agent)
    elif result == 1:  # Start with timer
        return (True, False, support_agent)
    elif result == 4: # Main Menu
        return ("MAIN_MENU", False, None)
    else:  # Cancelled
        return (False, False, None)

# ============================================================================
# FILE SEARCH POPUP - Shows when file is not found with retry countdown
# ============================================================================
def show_file_search_popup(excel_path, retry_interval_seconds=10):
    """
    Show a popup when the Excel file is not found.
    Features: live countdown, status indicator, Abort button, Manual file button.
    
    Returns:
        - ("FOUND", excel_path) = File was found at configured path
        - ("MANUAL", manual_path) = User selected a manual file (one-time use)
        - ("ABORT", None) = User clicked abort
    """
    from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog
    from PyQt5.QtCore import QTimer, Qt
    
    result = {"action": None, "path": None}
    
    class FileSearchDialog(QDialog):
        def __init__(self, file_path, retry_seconds):
            super().__init__()
            self.file_path = file_path
            self.retry_seconds = retry_seconds
            self.countdown = retry_seconds
            self.found = False
            
            self.setWindowTitle("Searching for File")
            self.resize(500, 200)
            self.setModal(True)
            
            layout = QVBoxLayout(self)
            layout.setContentsMargins(20, 20, 20, 20)
            layout.setSpacing(15)
            
            # Status label
            self.status_label = QLabel("🔍 Searching for Excel file...")
            self.status_label.setStyleSheet("font-weight: bold; font-size: 15px; color: #0F62FE;")
            self.status_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(self.status_label)
            
            # File path label
            path_label = QLabel(f"Path: {file_path}")
            path_label.setStyleSheet("font-size: 12px; color: #525252;")
            path_label.setWordWrap(True)
            layout.addWidget(path_label)
            
            # Countdown label
            self.countdown_label = QLabel(f"Next check in: {self.countdown} seconds")
            self.countdown_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #161616;")
            self.countdown_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(self.countdown_label)
            
            # Status indicator
            self.found_label = QLabel("Status: ❌ Not Yet Found")
            self.found_label.setStyleSheet("font-size: 15px; color: #DA1E28;")
            self.found_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(self.found_label)
            
            # Buttons
            btn_layout = QHBoxLayout()
            
            abort_btn = QPushButton("⛔ Abort")
            abort_btn.setStyleSheet("""
                QPushButton {
                    background-color: #DA1E28;
                    color: white;
                    font-weight: bold;
                    padding: 12px 24px;
                    border-radius: 5px;
                    font-size: 15px;
                }
                QPushButton:hover { background-color: #A2191F; }
            """)
            abort_btn.clicked.connect(self.on_abort)
            btn_layout.addWidget(abort_btn)
            
            manual_btn = QPushButton("📂 Load File Manually")
            manual_btn.setStyleSheet("""
                QPushButton {
                    background-color: #0F62FE;
                    color: white;
                    font-weight: bold;
                    padding: 12px 24px;
                    border-radius: 5px;
                    font-size: 15px;
                }
                QPushButton:hover { background-color: #0043CE; }
            """)
            manual_btn.clicked.connect(self.on_manual)
            btn_layout.addWidget(manual_btn)
            
            layout.addLayout(btn_layout)
            
            # Timer for countdown and file check
            self.timer = QTimer(self)
            self.timer.timeout.connect(self.tick)
            self.timer.start(1000)  # Every 1 second
            
            # Initial check
            self.check_file()
        
        def check_file(self):
            import os
            if os.path.exists(self.file_path):
                self.found = True
                self.found_label.setText("Status: ✅ Found!")
                self.found_label.setStyleSheet("font-size: 15px; color: #24A148; font-weight: bold;")
                self.status_label.setText("✅ Excel file found!")
                self.status_label.setStyleSheet("font-weight: bold; font-size: 15px; color: #24A148;")
                self.countdown_label.setText("Proceeding in 2 seconds...")
                self.timer.stop()
                # Wait 2 seconds then proceed
                QTimer.singleShot(2000, self.on_found)
        
        def tick(self):
            self.countdown -= 1
            if self.countdown <= 0:
                self.countdown = self.retry_seconds
                self.check_file()
            self.countdown_label.setText(f"Next check in: {self.countdown} seconds")
        
        def on_found(self):
            result["action"] = "FOUND"
            result["path"] = self.file_path
            self.accept()
        
        def on_abort(self):
            self.timer.stop()
            result["action"] = "ABORT"
            result["path"] = None
            self.accept()
        
        def on_manual(self):
            self.timer.stop()
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)"
            )
            if file_path:
                result["action"] = "MANUAL"
                result["path"] = file_path
                self.accept()
            else:
                # User cancelled file dialog, resume countdown
                self.timer.start(1000)
    
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    dialog = FileSearchDialog(excel_path, retry_interval_seconds)
    dialog.exec_()
    
    return result["action"], result["path"]

# ============================================================================
# COMPANIES PROCESS - Process company cases before case reviewer
# ============================================================================

class CompaniesProcessDialog(QDialog):
    """Dialog asking if user wants to process company cases first"""
    def __init__(self, total_cases, distinct_emails):
        super().__init__()
        self.setWindowTitle("Companies Process")
        self.setFixedSize(400, 180)
        self.result = "SKIP"
        
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        
        # Title
        title = QLabel("🏢 Company Cases Available")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #0f62fe;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Stats
        stats = QLabel(f"Total Cases: {total_cases}  |  Companies (Emails): {distinct_emails}")
        stats.setStyleSheet("font-size: 15px; color: #161616;")
        stats.setAlignment(Qt.AlignCenter)
        layout.addWidget(stats)
        
        # Description
        desc = QLabel("Process company cases grouped by email before reviewing individual cases?")
        desc.setStyleSheet("font-size: 13px; color: #525252;")
        desc.setWordWrap(True)
        desc.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        yes_btn = QPushButton("✅ Yes, Process Companies")
        yes_btn.setStyleSheet("""
            QPushButton {
                background-color: #0f62fe;
                color: white;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 15px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #0353e9; }
        """)
        yes_btn.clicked.connect(self.on_yes)
        btn_layout.addWidget(yes_btn)
        
        skip_btn = QPushButton("⏭️ Skip to Case Reviewer")
        skip_btn.setStyleSheet("""
            QPushButton {
                background-color: #e0e0e0;
                color: #161616;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 15px;
            }
            QPushButton:hover { background-color: #cacaca; }
        """)
        skip_btn.clicked.connect(self.on_skip)
        btn_layout.addWidget(skip_btn)
        
        layout.addLayout(btn_layout)
    
    def on_yes(self):
        self.result = "YES"
        self.accept()
    
    def on_skip(self):
        self.result = "SKIP"
        self.accept()

def load_companies_for_handler(cache_file, handler_name, sheet_name="Companies"):
    """Load Companies sheet cases assigned to handler, grouped by email"""
    try:
        df = pd.read_excel(cache_file, sheet_name=sheet_name)
        
        # Find columns
        email_col = find_column_case_insensitive(df, 'Email')
        assigned_col = find_column_case_insensitive(df, 'Assigned To')
        case_col = find_column_case_insensitive(df, 'Case Number')
        serial_col = find_column_case_insensitive(df, 'Serial Number')
        mtm_col = find_column_case_insensitive(df, 'Product ID (MTM)') or find_column_case_insensitive(df, 'Product ID')
        status_col = find_column_case_insensitive(df, 'Status')
        
        if not email_col or not assigned_col or not case_col:
            print("[WARN] Companies sheet missing required columns")
            return {}, df
        
        # Filter for handler's cases only
        handler_first = handler_name.split()[0] if handler_name else ""
        df_handler = df[df[assigned_col].astype(str).str.strip() == handler_first].copy()
        
        if df_handler.empty:
            return {}, df
        
        # Group by email
        grouped = {}
        for idx, row in df_handler.iterrows():
            email = str(row.get(email_col, '')).strip().lower()
            if not email or email == 'nan':
                continue
            
            case_num = str(row.get(case_col, '')).strip()
            serial = str(row.get(serial_col, '')).strip() if serial_col else ''
            mtm = str(row.get(mtm_col, '')).strip() if mtm_col else ''
            status = str(row.get(status_col, '')).strip().lower() if status_col else ''
            
            if email not in grouped:
                grouped[email] = {'cases': [], 'df_indices': []}
            
            grouped[email]['cases'].append({
                'case_number': case_num,
                'serial': serial if serial and serial.lower() != 'nan' else '',
                'mtm': mtm if mtm and mtm.lower() != 'nan' else '',
                'status': status,
                'row_idx': idx
            })
            grouped[email]['df_indices'].append(idx)
        
        return grouped, df
    except Exception as e:
        print(f"[ERROR] Failed to load companies for handler: {e}")
        return {}, pd.DataFrame()

def show_companies_email_confirmation(email, cases_info, email_body):
    """Show email preview with confirmation - must confirm to proceed"""
    from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QTextEdit, QHBoxLayout
    
    class EmailConfirmDialog(QDialog):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Confirm Email - Companies Batch")
            self.setFixedSize(600, 500)
            self.confirmed = False
            
            layout = QVBoxLayout(self)
            
            # Header
            header = QLabel(f"📧 Sending Email to: {email}")
            header.setStyleSheet("font-size: 16px; font-weight: bold; color: #0f62fe;")
            layout.addWidget(header)
            
            # Cases list
            cases_label = QLabel(f"Cases in this batch: {len(cases_info)}")
            cases_label.setStyleSheet("font-size: 13px; color: #525252;")
            layout.addWidget(cases_label)
            
            # Email preview
            preview = QTextEdit()
            preview.setPlainText(email_body)
            preview.setReadOnly(True)
            preview.setStyleSheet("font-size: 13px; background: #f4f4f4;")
            layout.addWidget(preview)
            
            # Confirm button only (no cancel option per user request)
            confirm_btn = QPushButton("✅ Confirm and Send Email")
            confirm_btn.setStyleSheet("""
                QPushButton {
                    background-color: #24a148;
                    color: white;
                    border-radius: 5px;
                    padding: 12px 24px;
                    font-size: 15px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #1c8a3c; }
            """)
            confirm_btn.clicked.connect(self.on_confirm)
            layout.addWidget(confirm_btn)
        
        def on_confirm(self):
            self.confirmed = True
            self.accept()
    
    dialog = EmailConfirmDialog()
    dialog.exec_()
    return dialog.confirmed

def build_companies_email_body(cases_info, agent_name):
    """Build email body with Case Number | Serial/MTM format"""
    devices_list = []
    for case in cases_info:
        case_num = case.get('case_number', '')
        serial = case.get('serial', '')
        mtm = case.get('mtm', '')
        
        if serial:
            devices_list.append(f"{case_num} | {serial}")
        elif mtm:
            devices_list.append(f"{case_num} | {mtm}")
        else:
            devices_list.append(f"{case_num} | (No Serial/MTM)")
    
    devices_str = "\n".join(devices_list)
    
    email_body = (
        " Hello All\n\n"
        "I hope you are doing well.\n\n"
        "Regarding the devices:\n\n"
        f"{devices_str}\n\n"
        "I am contacting you to verify that the devices are functioning properly "
        "and performing to your expectations.\n\n"
        "Thank you for choosing Lenovo Services.\n\n"
        "We appreciate your business and are here to support you.\n\n"
        "Thanks and Regards,\n"
        f"{agent_name}\n"
        "NA Lenovo PC Assurance Resolution Team"
    )
    return email_body

def run_companies_process(driver, cache_file, agent_name, sheet_name="Companies"):
    """
    Main function to process all company cases grouped by email.
    
    FLOW:
    1. Gather case numbers from cache file
    2. For each case: case_search_and_open_no_edit (without Edit)
    3. Use solution_provided_check_and_skip to check status
    4. Go to first solution provided case
    5. Use original case_search_and_open (with Edit)
    6. Build company email (confirm)
    7. Send email
    8. Perform call
    9. Leave note with closing code on first case
    10. Register on excel sheet
    11. Go to all other cases for same email - leave note (call performed on first case)
    12. Press save on each
    13. Go to next email group
    """
    print("[INFO] === Starting Companies Process ===")
    
    grouped, df_companies = load_companies_for_handler(cache_file, agent_name, sheet_name)
    
    if not grouped:
        print("[INFO] No company cases to process for this handler")
        return
    
    print(f"[INFO] Found {len(grouped)} distinct companies (emails) to process")
    
    today_str = datetime.now().strftime("%b %d, %Y")
    
    for email, data in grouped.items():
        cases = data['cases']
        
        print(f"\n[INFO] ========== Processing Company Email: {email} ({len(cases)} cases) ==========")
        
        # =====================================================================
        # STEP 1-3: Go through each case and check Solution Provided status
        # =====================================================================
        solution_provided_cases = []
        
        for case in cases:
            case_num = case['case_number']
            print(f"\n[INFO] Step 2: Opening case {case_num} (no edit)...")
            
            try:
                # Step 2: Open case WITHOUT pressing Edit
                case_search_and_open_no_edit(driver, case_num)
                time.sleep(2)
                
                # Step 3: Check if solution provided using solution_provided_check_and_skip
                print(f"[INFO] Step 3: Checking Solution Provided status...")
                is_solution_provided = solution_provided_check_and_skip(driver, case_num, df_companies, cache_file)
                
                if is_solution_provided:
                    solution_provided_cases.append({
                        'case_number': case_num,
                        'serial': case.get('serial', ''),
                        'mtm': case.get('mtm', ''),
                        'row_idx': case.get('row_idx')
                    })
                    print(f"[INFO] ✓ Case {case_num}: Solution Provided - added to batch")
                else:
                    print(f"[INFO] ✗ Case {case_num}: NOT Solution Provided - skipping")
                    
            except Exception as e:
                print(f"[WARN] Failed to check case {case_num}: {e}")
                continue
        
        # =====================================================================
        # Check if we have any solution provided cases
        # =====================================================================
        if not solution_provided_cases:
            print(f"\n[INFO] No Solution Provided cases found for {email} - skipping company")
            continue
        
        print(f"\n[INFO] Found {len(solution_provided_cases)} Solution Provided cases for email batch")
        
        # =====================================================================
        # STEP 4-5: Go to first solution provided case with Edit
        # =====================================================================
        first_case = solution_provided_cases[0]['case_number']
        
        try:
            print(f"\n[INFO] Step 4-5: Opening first case {first_case} WITH Edit...")
            case_search_and_open(driver, first_case)
            time.sleep(2)
            
            # =====================================================================
            # STEP 6: Build email and confirm
            # =====================================================================
            print(f"[INFO] Step 6: Building email for confirmation...")
            email_body = build_companies_email_body(solution_provided_cases, agent_name)
            confirmed = show_companies_email_confirmation(email, solution_provided_cases, email_body)
            
            if not confirmed:
                print("[INFO] Email not confirmed - skipping this company")
                continue
            
            # =====================================================================
            # STEP 7: Send email
            # =====================================================================
            print("[INFO] Step 7: Sending companies batch email...")
            send_Email(driver, email_body)
            time.sleep(2)
            
            # =====================================================================
            # STEP 8: Perform call
            # =====================================================================
            print("[INFO] Step 8: Performing call flow...")
            perform_call_flow(driver)
            
            # =====================================================================
            # STEP 9: Get call closing code and leave note on first case
            # =====================================================================
            call_code, add_note = get_call_closing_code()
            print(f"[INFO] Step 9: Call Closing Code: {call_code}")
            
            # Build note for first case
            serials_str = "\n".join([f"{c['case_number']} | {c['serial'] or c['mtm']}" for c in solution_provided_cases])
            first_case_note = (
                f"Date: {today_str}\n"
                "Queue: ART Project - Follow up\n"
                f"Agent: {agent_name}\n"
                f"Action: Sent Company Email with devices:\n{serials_str}\n"
                f"Call Closing Code: {call_code}\n"
                " \n ------------------------"
            )
            
            print(f"[INFO] Leaving note on first case {first_case}...")
            add_Case_Note(driver, first_case, first_case_note)
            
            # =====================================================================
            # STEP 10-11: Update Excel and go to other cases to leave notes
            # =====================================================================
            print("[INFO] Step 10: Updating Companies sheet in Excel...")
            
            for i, case in enumerate(solution_provided_cases):
                idx = case.get('row_idx')
                if idx is not None:
                    df_companies.at[idx, 'Action 2'] = 'Sent Email'
                    df_companies.at[idx, 'Action 3'] = 'Called the Customer'
                    df_companies.at[idx, 'Final Action'] = call_code if call_code else 'Called'
                    df_companies.at[idx, 'Status'] = 'closed'
            
            # Save to cache immediately
            with pd.ExcelWriter(cache_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_companies.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"[INFO] Updated {len(solution_provided_cases)} cases in Companies sheet")
            
            # =====================================================================
            # STEP 12-13: Go to other cases and leave note that call was on first case
            # =====================================================================
            if len(solution_provided_cases) > 1:
                print(f"\n[INFO] Step 11-12: Leaving notes on other {len(solution_provided_cases)-1} cases...")
                
                for case in solution_provided_cases[1:]:  # Skip first case already processed
                    case_num = case['case_number']
                    print(f"[INFO] Processing case {case_num}...")
                    
                    try:
                        # Open case with Edit
                        case_search_and_open(driver, case_num)
                        time.sleep(2)
                        
                        # Leave note referencing first case
                        other_case_note = (
                            f"Date: {today_str}\n"
                            "Queue: ART Project - Follow up\n"
                            f"Agent: {agent_name}\n"
                            f"Action: Company Bulk Email sent and Call performed on Case Number: {first_case}\n"
                            f"Call Closing Code: {call_code}\n"
                            " \n ------------------------"
                        )
                        
                        add_Case_Note(driver, case_num, other_case_note)
                        
                        # Step 13: Press Save
                        print(f"[INFO] Step 13: Saving case {case_num}...")
                        click_safe(
                            driver,
                            By.XPATH,
                            "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckIn.Command') and contains(@id,'-button')]",
                            timeout=2,
                            retries=2
                        )
                        time.sleep(2)
                        
                    except Exception as e:
                        print(f"[WARN] Failed to process other case {case_num}: {e}")
                        continue
            
            print(f"\n[INFO] ✓ Completed processing company: {email}")
            
        except Exception as e:
            print(f"[ERROR] Failed to process company {email}: {e}")
            continue
    
    print("\n[INFO] === Companies Process Complete ===")

# ============================================================================
# ART Team Variables
# ============================================================================
SMSText = ("Hello {CX_Name} , \n\nPlease reply with 1, 2, or 3 regarding your recent Lenovo Service for serial number: {case_number} . \n\n1. Issue Resolved\n2. Need Assistance\n3. Stop Text Messages\n\nIf there is no reply, we will contact you by phone within 24 hours between 8:00 AM - 6:00 PM local time.")
today_str = datetime.now().strftime("%b %d, %Y")
CaseNote = f"Date: {today_str}\nQueue: ART Project - Follow up \nAgent: {AGENT_NAME} \nAction: Sent SMS  // Sent Email \n \n ------------------------"
CaseEmailOnSite_Depot = (
    " Hello {CX_Name}\n\n"
    "I hope you are doing well.\n\n"
    "I hope your device of Serial Number: {serial_val} is performing well after its recent repair.\n\n"
    "To ensure everything is working as expected, please reply with one of the following options:\n\n"
    "1 for Issue Resolved — Everything is working properly.\n"
    "2 for Need Assistance — Please include details of any ongoing issues.\n"
    "3 for STOP — To stop receiving follow-up messages.\n\n"
    "If we do not receive a response, a member of our team may follow up with a phone call on the next business day (Monday—Friday) between\n"
    "8:00 AM and 6:00 PM local time.\n\n"
    "Thank you for choosing Lenovo Services.\n\n"
    "We appreciate your business and are here to support you.\n\n"
    "Thanks and Regards,\n"
    "{AGENT_NAME}\n"
    "NA Lenovo PC Assurance Resolution Team"
)
CaseEmailCRU = (
    " Hello {CX_Name}\n\n"
    "I hope you are doing well.\n\n"
    "Regarding the device of Serial number:  {serial_val}\n\n"
    "I am contacting you to verify that the Lenovo replacement part is functioning properly and performing to your expectations.\n\n"
    "To ensure everything is working as expected, please reply with one of the following options:\n\n"
    "1 for Issue Resolved — Everything is working properly.\n"
    "2 for Need Assistance — Please include details of any ongoing issues.\n"
    "3 for STOP — To stop receiving follow-up messages.\n\n"
    "If we do not receive a response, a member of our team may follow up with a phone call on the next business day (Monday—Friday) between\n"
    "8:00 AM and 6:00 PM local time.\n\n"
    "Thank you for choosing Lenovo Services.\n\n"
    "We appreciate your business and are here to support you.\n\n"
    "Thanks and Regards,\n"
    "{AGENT_NAME}\n"
    "NA Lenovo PC Assurance Resolution Team"
)


# ART Chrome profile setup
chrome_options = Options()
user_data_dir = os.environ.get('CHROME_USER_DATA_DIR') or os.path.join(os.path.expanduser('~'), '@')
profile_dir = os.environ.get('CHROME_PROFILE_DIR') or 'Default'
if user_data_dir and os.path.exists(user_data_dir):
    try:
        chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
        chrome_options.add_argument(f'--profile-directory={profile_dir}')
        print(f"[INFO] Using Chrome profile: {user_data_dir}")
    except Exception as e:
        print(f"[WARN] Failed to set Chrome profile: ")
else:
    # Create profile directory if it doesn't exist
    try:
        os.makedirs(user_data_dir, exist_ok=True)
        chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
        chrome_options.add_argument(f'--profile-directory={profile_dir}')
        print(f"[INFO] Created new Chrome profile at: {user_data_dir}")
    except Exception as e:
        print(f"[WARN] Failed to create Chrome profile: ")


def _ensure_dir(path):
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        pass

def safe_find(driver, by, locator, timeout=20, clickable=False, retries=3, poll=0.5):
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            wait = WebDriverWait(driver, timeout)
            if clickable:
                el = wait.until(EC.element_to_be_clickable((by, locator)))
            else:
                el = wait.until(EC.presence_of_element_located((by, locator)))
            return el
        except Exception as e:
            last_exc = e
            print(f"[WARN] safe_find attempt {attempt}/{retries} for {locator} failed: ")
            time.sleep(poll)

    _ensure_dir('errors')
    safe_name = locator.replace('/', '_').replace(' ', '_')[:100]
    img_path = os.path.join('errors', f'failure_{safe_name}.png')
    try:
        driver.save_screenshot(img_path)
        print(f"[ERROR] Element not found: {locator}. Screenshot saved to {img_path}")
    except Exception:
        print(f"[ERROR] Element not found: {locator}. Also failed to save screenshot.")
    traceback.print_exception(type(last_exc), last_exc, last_exc.__traceback__)
    return None

def click_safe(driver, by, locator, timeout=20, retries=3, poll=0.5):
    el = safe_find(driver, by, locator, timeout=timeout, clickable=True, retries=retries, poll=poll)
    if not el:
        return False
    try:
        el.click()
        return True
    except Exception as e:
        print(f"[ERROR] click_safe failed on {locator}: ")
        return False

def send_keys_safe(driver, by, locator, value, timeout=20, retries=3, poll=0.5, enter=False):
    el = safe_find(driver, by, locator, timeout=timeout, clickable=True, retries=retries, poll=poll)
    if not el:
        return False
    try:
        try:
            el.clear()
        except Exception:
            pass
        el.send_keys(value)
        if enter:
            el.send_keys(Keys.ENTER)
        return True
    except Exception as e:
        print(f"[ERROR] send_keys_safe failed on {locator}: ")
        return False

def enable_windows_inhibit():
    global _inhibit_active
    try:
        if platform.system().lower().startswith('win'):
            ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS | ES_SYSTEM_REQUIRED | ES_DISPLAY_REQUIRED)
            _inhibit_active = True
            print("[INFO] Windows sleep/display inhibit enabled")
    except Exception as e:
        print(f"[WARN] Failed to enable Windows inhibit: ")

def perform_call_flow(driver):
    try:
        # Read mobile phone
        mobile_input = safe_find(driver, By.XPATH, "//input[@type='tel' and @aria-label='Mobile Phone']", timeout=1, retries=7)
        if mobile_input:
            mobile_value = (mobile_input.get_attribute('value') or '').strip()
            if not mobile_value:
                # fallback to title if value not populated
                mobile_value = (mobile_input.get_attribute('title') or '').strip()
        else:
            mobile_value = ''

        # Prepare dial number: replace leading '+' with '9'
        if mobile_value:
            number_core = mobile_value.lstrip('+')
            dialed_number = f"9{number_core}"
        else:
            print("[WARN] No mobile phone value found to dial")
            return False
        
        # Switch to the dialer tab (should be at index 0)
        handles = driver.window_handles
        if len(handles) > 0:
            dialer_handle = handles[0]
            try:
                driver.switch_to.window(dialer_handle)
                print("[INFO] Switched to dialer window for call")
            except Exception as e:
                print(f"[WARN] Failed to switch to dialer: {e}")
                return False
        else:
            print("[ERROR] No dialer window found")
            return False

        # Click mark done and dial the number
        click_safe(driver, By.XPATH, '//span[contains(@class,"wwe-sprite-mark-done")]', timeout=1, retries=1)
        send_keys_safe(driver, By.ID, 'wweTeamCommunicatorDialerField', dialed_number, timeout=3, retries=5, poll=0.5, enter=True)
        
        print(f"[INFO] Dialed number: {dialed_number}")
        
        # Wait for user to complete call
        time.sleep(5)
        
        # Switch back to CRM using robust window switching
        print("[INFO] Switching back to CRM window after call...")
        success = switch_to_crm_window(driver, max_retries=5)
        
        if not success:
            print("[ERROR] Failed to switch back to CRM window after call")
            return False
        
        print("[SUCCESS] Successfully returned to CRM window")
        return True
        
    except Exception as e:
        print(f"[WARN] perform_call_flow encountered an error: {e}")
        traceback.print_exc()
        return False

def disable_windows_inhibit():
    global _inhibit_active
    try:
        if _inhibit_active and platform.system().lower().startswith('win'):
            # clear the previous flags by calling with ES_CONTINUOUS
            ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS)
            _inhibit_active = False
            print("[INFO] Windows sleep/display inhibit disabled")
    except Exception as e:
        print(f"[WARN] Failed to disable Windows inhibit: ")

def get_case_closing_code(case_number, cases_completed_count, total_in_progress_count=None, case_status="in_progress"):
    """
    Opens a case reviewer dialog for the current case.
    Creates a fresh dialog for each case to avoid garbage collection issues.
    Returns the selected closing code and whether to add a note.
    """
    print(f"[DEBUG] get_case_closing_code called: case={case_number}, completed={cases_completed_count}, total={total_in_progress_count}, status={case_status}")
    
    class CaseReviewerDialog(QDialog):
        def __init__(self, case_num, cases_completed, total_count, status, on_break=False):
            super().__init__()
            print(f"[DEBUG] CaseReviewerDialog created with: case_num={case_num}, completed={cases_completed}, total={total_count}, status={status}, on_break={on_break}")
            self.setWindowTitle("Case Reviewer")
            self.resize(600, 750)  # Taller window to fit error log section
            self.selected_code = None
            self.add_note = False
            self.case_status = status
            self.is_on_break = on_break  # Initialize from parameter
            self.setModal(True)
            
            # Main layout
            main_layout = QVBoxLayout(self)
            main_layout.setContentsMargins(10, 10, 10, 10)  # Reduced from 15
            main_layout.setSpacing(8)  # Reduced from 10
            
            # ========== TOP PANEL: Case Info & Progress ==========
            info_layout = QVBoxLayout()
            info_layout.setSpacing(5)
            
            self.case_info_label = QLabel()
            self.case_info_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #0F62FE;")
            info_layout.addWidget(self.case_info_label)
            
            self.progress_bar = QProgressBar()
            self.progress_bar.setMinimum(0)
            self.progress_bar.setMaximum(100)
            self.progress_bar.setTextVisible(True)
            self.progress_bar.setStyleSheet("""
                QProgressBar {
                    border: 2px solid #525252;
                    border-radius: 5px;
                    text-align: center;
                    height: 20px;
                    background-color: #262626;
                    color: #F4F4F4;
                    font-size: 13px;
                    font-weight: bold;
                }
                QProgressBar::chunk {
                    background-color: #24A148;
                }
            """)
            info_layout.addWidget(self.progress_bar)
            
            main_layout.addLayout(info_layout)
            main_layout.addSpacing(10)
            
            # ========== SCROLLABLE BUTTON SECTIONS ==========
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll_widget = QWidget()
            scroll_layout = QVBoxLayout(scroll_widget)
            scroll_layout.setSpacing(10)  # Reduced from 15
            scroll_layout.setContentsMargins(5, 5, 5, 5)  # Add margins
            
            # ===== SECTION A: Issue Resolution =====
            scroll_layout.addWidget(self._create_section_header("Issue Resolution"))
            sec_a = QGridLayout()
            sec_a.setSpacing(6)  # Reduced from 8
            sec_a.setColumnStretch(0, 1)
            sec_a.setColumnStretch(1, 1)
            
            sec_a_buttons = [
                ("Resolved", "Issue Resolved"),
                ("Issue Not Fixed", "Issue Not Fixed"),
                ("Not Reached", "Customer Not Reached"),
                ("Not Yet Tested", "Machine Not Yet Tested"),
                ("DND", "DND"),
                ("Escalated", "Escalated"),
                ("Still in Review", "Need Third Action"),
            ]
            
            for idx, (label_text, code) in enumerate(sec_a_buttons):
                btn = self._create_button(label_text, code, "#E3F2FD")
                sec_a.addWidget(btn, idx // 2, idx % 2)
            
            scroll_layout.addLayout(sec_a)
            scroll_layout.addSpacing(3)  # Reduced from 5
            
            # ===== SECTION B: Communication Follow-up =====
            scroll_layout.addWidget(self._create_section_header("Communication Follow-up"))
            sec_b = QGridLayout()
            sec_b.setSpacing(6)  # Reduced from 8
            sec_b.setColumnStretch(0, 1)
            sec_b.setColumnStretch(1, 1)
            
            sec_b_buttons = [
                ("Send SMS", "Send SMS"),
                ("Send Email", "Send Email"),
                ("Send SMS + Email", "Send SMS and Email"),
                ("Call Customer", "Call the Customer"),
            ]
            
            for idx, (label_text, code) in enumerate(sec_b_buttons):
                btn = self._create_button(label_text, code, "#F3E5F5")
                sec_b.addWidget(btn, idx // 2, idx % 2)
            
            scroll_layout.addLayout(sec_b)
            scroll_layout.addSpacing(5)
            
            # ===== SECTION C: Case Management =====
            scroll_layout.addWidget(self._create_section_header("Case Management"))
            sec_c = QGridLayout()
            sec_c.setSpacing(8)
            sec_c.setColumnStretch(0, 1)
            sec_c.setColumnStretch(1, 1)
            
            skip_btn = self._create_button("Skip the Case", "Skipped", "#FFEBEE")
            sec_c.addWidget(skip_btn, 0, 0)
            
            prev_case_btn = self._create_button("⬅ Previous Case", "PREVIOUS_CASE", "#FFF9C4")
            sec_c.addWidget(prev_case_btn, 0, 1)
            
            other_btn = self._create_button("Custom Code", "custom", "#E0E0E0")
            other_btn.clicked.disconnect()
            other_btn.clicked.connect(self.ask_other_code)
            sec_c.addWidget(other_btn, 1, 0)
            
            scroll_layout.addLayout(sec_c)
            scroll_layout.addStretch()
            
            scroll.setWidget(scroll_widget)
            main_layout.addWidget(scroll)
            
            # ========== ERROR LOG SECTION ==========
            from PyQt5.QtWidgets import QTextEdit, QGroupBox
            error_group = QGroupBox("Recent Errors")
            error_group.setStyleSheet("""
                QGroupBox {
                    font-weight: bold;
                    font-size: 12px;
                    color: #FA4D56;
                    border: 1px solid #525252;
                    border-radius: 4px;
                    margin-top: 8px;
                    padding-top: 8px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px;
                }
            """)
            error_layout = QVBoxLayout(error_group)
            self.error_log = QTextEdit()
            self.error_log.setReadOnly(True)
            self.error_log.setMaximumHeight(60)
            self.error_log.setStyleSheet("""
                QTextEdit {
                    background-color: #1C1C1C;
                    border: none;
                    color: #FA4D56;
                    font-size: 11px;
                    padding: 4px;
                }
            """)
            self.error_log.setPlaceholderText("No errors")
            error_layout.addWidget(self.error_log)
            main_layout.addWidget(error_group)
            
            # ========== BOTTOM PANEL: Options ==========
            bottom_layout = QVBoxLayout()
            bottom_layout.setSpacing(8)
            
            self.add_note_checkbox = QCheckBox("✓ Add Case Note")
            self.add_note_checkbox.setStyleSheet("font-size: 14px; padding: 6px; color: #161616;")
            bottom_layout.addWidget(self.add_note_checkbox)
            
            # Take a Break / Back on Duty button (toggles based on is_on_break state)
            self.take_break_btn = QPushButton()
            if self.is_on_break:
                # Currently on break - show "Back on Duty"
                self.take_break_btn.setText("🟢 Back on Duty")
                self.take_break_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #24A148;
                        color: #F4F4F4;
                        border: none;
                        border-radius: 4px;
                        font-weight: bold;
                        padding: 10px;
                        font-size: 14px;
                    }
                    QPushButton:hover {
                        background-color: #1E8E3E;
                    }
                """)
            else:
                # Currently working - show "Take a Break"
                self.take_break_btn.setText("☕ Take a Break")
                self.take_break_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #795548;
                        color: #F4F4F4;
                        border: none;
                        border-radius: 4px;
                        font-weight: bold;
                        padding: 10px;
                        font-size: 14px;
                    }
                    QPushButton:hover {
                        background-color: #5D4037;
                    }
                """)
            self.take_break_btn.clicked.connect(self.toggle_break_status)
            bottom_layout.addWidget(self.take_break_btn)
            
            # Close/Exit button
            close_btn = QPushButton("⛔ Close & Exit Application")
            close_btn.setStyleSheet("""
                QPushButton {
                    background-color: #DA1E28;
                    color: #F4F4F4;
                    border: 1px solid #A2191F;
                    border-radius: 4px;
                    font-weight: bold;
                    padding: 10px;
                    font-size: 14px;
                }
                QPushButton:hover {
                    background-color: #A2191F;
                }
                QPushButton:pressed {
                    background-color: #750E13;
                }
            """)
            close_btn.clicked.connect(lambda: self.on_button_clicked("CLOSE_APPLICATION"))
            bottom_layout.addWidget(close_btn)
            
            # Add bottom layout to main layout
            main_layout.addLayout(bottom_layout)
            
            self.setLayout(main_layout)
            self.update_case_info(case_num, cases_completed, total_count)
        
        def _create_section_header(self, title):
            """Create a section header label with IBM Design styling"""
            header = QLabel(f"▼ {title}")
            header.setStyleSheet("""
                font-weight: bold;
                font-size: 14px;
                color: #F4F4F4;
                background-color: #0F62FE;
                padding: 8px 12px;
                border-radius: 4px;
                margin-top: 5px;
            """)
            return header
        
        def _create_button(self, label_text, code, bg_color):
            """Create a styled button with IBM Design"""
            btn = QPushButton(label_text)
            btn.setMinimumHeight(44)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {bg_color};
                    border: 1px solid #525252;
                    border-radius: 5px;
                    font-weight: bold;
                    color: #161616;
                    padding: 10px 12px;
                    font-size: 15px;
                }}
                QPushButton:hover {{
                    border: 2px solid #0F62FE;
                    background-color: {bg_color};
                }}
                QPushButton:pressed {{
                    background-color: #78A9FF;
                }}
            """)
            btn.clicked.connect(lambda: self.on_button_clicked(code))
            return btn
        
        def update_case_info(self, case_num, cases_completed, total_count):
            """Update the case info and progress bar"""
            try:
                print(f"[DEBUG] update_case_info: cases_completed={cases_completed}, total_count={total_count}, type(total)={type(total_count)}")
                status_display = self.case_status.title() if self.case_status else "Unknown"
                if total_count and total_count > 0:
                    # cases_completed is the count of completed cases (0-indexed, so +1 for display)
                    display_num = cases_completed + 1
                    percentage = int(((cases_completed + 1) / total_count) * 100)
                    print(f"[DEBUG] Progress calculation: {display_num}/{total_count} = {percentage}% (raw: {(cases_completed + 1) / total_count * 100})")
                    self.case_info_label.setText(f"Case: {case_num}  |  Status: {status_display}  |  Progress: {display_num} of {total_count} ({percentage}%)")
                    self.progress_bar.setValue(percentage)
                    self.progress_bar.setFormat(f"{display_num}/{total_count} ({percentage}%)")
                else:
                    print(f"[DEBUG] Total count is 0 or None: {total_count}")
                    self.case_info_label.setText(f"Case: {case_num}  |  Status: {status_display}")
                    self.progress_bar.setValue(0)
            except Exception as e:
                print(f"[WARN] Error updating case info: {e}")
        
        def log_error(self, message):
            """Append an error message to the error display"""
            from datetime import datetime
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.error_log.append(f"⚠ [{timestamp}] {message}")
            # Keep only last 5 errors to save space
            if self.error_log.document().blockCount() > 5:
                cursor = self.error_log.textCursor()
                cursor.movePosition(cursor.Start)
                cursor.movePosition(cursor.Down, cursor.KeepAnchor)
                cursor.removeSelectedText()
            # Scroll to bottom
            self.error_log.verticalScrollBar().setValue(
                self.error_log.verticalScrollBar().maximum()
            )
        
        def toggle_break_status(self):
            """Toggle between Take a Break and Back on Duty"""
            if not self.is_on_break:
                # Currently working -> going on break
                self.is_on_break = True
                self.take_break_btn.setText("🟢 Back on Duty")
                self.take_break_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #24A148;
                        color: #F4F4F4;
                        border: none;
                        border-radius: 4px;
                        font-weight: bold;
                        padding: 10px;
                        font-size: 14px;
                    }
                    QPushButton:hover {
                        background-color: #1E8E3E;
                    }
                """)
                self.selected_code = "TAKE_A_BREAK"
                self.accept()  # Close dialog to trigger break
            else:
                # Currently on break -> back on duty
                self.is_on_break = False
                self.take_break_btn.setText("☕ Take a Break")
                self.take_break_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #795548;
                        color: #F4F4F4;
                        border: none;
                        border-radius: 4px;
                        font-weight: bold;
                        padding: 10px;
                        font-size: 14px;
                    }
                    QPushButton:hover {
                        background-color: #5D4037;
                    }
                """)
                self.selected_code = "BACK_ON_DUTY"
                self.accept()  # Close dialog to resume
        
        def on_button_clicked(self, code):
            """Handle button click - close dialog with result"""
            self.selected_code = code
            self.add_note = bool(self.add_note_checkbox.isChecked())
            print(f"[DEBUG] Button clicked: code={code}, add_note={self.add_note}")
            self.accept()  # Close dialog and return success
        
        def ask_other_code(self):
            """Ask for custom closing code with Final Action and Status options"""
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QRadioButton, QButtonGroup, QPushButton, QGroupBox
            
            custom_dialog = QDialog(self)
            custom_dialog.setWindowTitle("Custom Closing Code")
            custom_dialog.resize(400, 280)
            
            layout = QVBoxLayout(custom_dialog)
            layout.setSpacing(15)
            
            # Custom code text input
            text_label = QLabel("Enter custom closing code (for notes):")
            text_label.setStyleSheet("font-weight: bold;")
            layout.addWidget(text_label)
            
            text_input = QLineEdit()
            text_input.setPlaceholderText("Custom code text...")
            text_input.setMinimumHeight(30)
            layout.addWidget(text_input)
            
            # Radio buttons section
            radio_layout = QHBoxLayout()
            
            # Final Action column
            final_action_group = QGroupBox("Final Action")
            final_action_layout = QVBoxLayout(final_action_group)
            self.final_action_btn_group = QButtonGroup(custom_dialog)
            
            reviewed_radio = QRadioButton("Reviewed")
            reviewed_radio.setChecked(True)
            not_reached_radio = QRadioButton("Not Reached")
            
            self.final_action_btn_group.addButton(reviewed_radio, 0)
            self.final_action_btn_group.addButton(not_reached_radio, 1)
            
            final_action_layout.addWidget(reviewed_radio)
            final_action_layout.addWidget(not_reached_radio)
            radio_layout.addWidget(final_action_group)
            
            # Status column
            status_group = QGroupBox("Status")
            status_layout = QVBoxLayout(status_group)
            self.status_btn_group = QButtonGroup(custom_dialog)
            
            skipped_radio = QRadioButton("Skipped")
            skipped_radio.setChecked(True)
            closed_radio = QRadioButton("Closed")
            
            self.status_btn_group.addButton(skipped_radio, 0)
            self.status_btn_group.addButton(closed_radio, 1)
            
            status_layout.addWidget(skipped_radio)
            status_layout.addWidget(closed_radio)
            radio_layout.addWidget(status_group)
            
            layout.addLayout(radio_layout)
            
            # OK/Cancel buttons
            btn_layout = QHBoxLayout()
            ok_btn = QPushButton("OK")
            ok_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px 20px;")
            cancel_btn = QPushButton("Cancel")
            cancel_btn.setStyleSheet("background-color: #9E9E9E; color: white; padding: 8px 20px;")
            
            ok_btn.clicked.connect(custom_dialog.accept)
            cancel_btn.clicked.connect(custom_dialog.reject)
            
            btn_layout.addWidget(ok_btn)
            btn_layout.addWidget(cancel_btn)
            layout.addLayout(btn_layout)
            
            if custom_dialog.exec_() == QDialog.Accepted:
                custom_text = text_input.text().strip()
                
                # Get selected Final Action
                final_action_id = self.final_action_btn_group.checkedId()
                final_action = "Reviewed" if final_action_id == 0 else "Not Reached"
                
                # Get selected Status
                status_id = self.status_btn_group.checkedId()
                status = "Skipped" if status_id == 0 else "Closed"
                
                # Store the selections for later use
                self.custom_final_action = final_action
                self.custom_status = status
                
                # Use the custom text as the code (for notes), but also pass final action and status
                if custom_text:
                    # Format: "custom_text|final_action|status" to be parsed later
                    result_code = f"CUSTOM|{custom_text}|{final_action}|{status}"
                    self.on_button_clicked(result_code)
                else:
                    # No custom text, just use the selected final action
                    result_code = f"CUSTOM||{final_action}|{status}"
                    self.on_button_clicked(result_code)
    
    # Get or create QApplication
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    # Create fresh dialog for this case (pass global is_on_break state)
    global is_on_break
    dialog = CaseReviewerDialog(case_number, cases_completed_count, total_in_progress_count, case_status, on_break=is_on_break)
    
    # Show dialog modally and wait for result
    print(f"[INFO] Waiting for case reviewer input for case: {case_number}")
    result = dialog.exec_()
    
    # Check if user selected something
    # In PyQt6, QDialog.Accepted is now QDialog.DialogCode.Accepted (value = 1)
    if result == 1 and dialog.selected_code:  # 1 = QDialog.DialogCode.Accepted
        result_code = dialog.selected_code
        result_note = dialog.add_note
        print(f"[INFO] Case {case_number} closing code selected: {result_code}")
    else:
        # User closed dialog without selecting - default to "New"
        result_code = "New"
        result_note = False
        print(f"[WARN] Case {case_number} dialog closed without selection, defaulting to: {result_code}")
    
    # Clean up
    dialog.close()
    dialog.deleteLater()
    
    return result_code, result_note

def close_case_reviewer():
    """Cleanup function - no longer needed with fresh dialog approach"""
    pass


# ============================================================================
# NEW CASES PROGRESS POPUP - Floating progress widget with slide animation
# ============================================================================
class NewCasesProgressPopup(QWidget):
    """
    A floating progress popup that appears during new case processing.
    Features: slide-in animation on hover, draggable, pause/continue controls.
    """
    # Signals for communication with main loop
    pause_signal = pyqtSignal()
    continue_signal = pyqtSignal()
    quit_after_current_signal = pyqtSignal()
    abort_now_signal = pyqtSignal()
    take_a_break_signal = pyqtSignal()
    back_on_duty_signal = pyqtSignal()
    
    def __init__(self, total_cases=0, parent=None):
        super().__init__(parent)
        self.total_cases = total_cases
        self.cases_done = 0
        self.is_paused = False
        self.quit_after_current = False
        self.abort_now = False
        self.take_a_break = False
        self.is_expanded = False
        self.drag_position = None
        
        # Window flags: frameless, always on top, tool window
        self.setWindowFlags(
            Qt.FramelessWindowHint | 
            Qt.WindowStaysOnTopHint | 
            Qt.Tool
        )
        self.setAttribute(Qt.WA_TranslucentBackground, False)
        
        # Dimensions
        self.collapsed_width = 25
        self.expanded_width = 280
        self.popup_height = 350
        
        self.setup_ui()
        self.position_on_screen()
        
        # Animation
        self.slide_animation = QPropertyAnimation(self, b"geometry")
        self.slide_animation.setDuration(200)
        
        # Hover timer for auto-collapse
        self.collapse_timer = QTimer()
        self.collapse_timer.setSingleShot(True)
        self.collapse_timer.timeout.connect(self.collapse)
    
    def setup_ui(self):
        """Setup the UI components with IBM Design colors"""
        self.setFixedSize(self.expanded_width, self.popup_height)
        
        # IBM Design Colors
        IBM_GREY_90 = "#161616"  # Background
        IBM_BLUE_60 = "#0F62FE"  # Accent/Border
        IBM_WHITE = "#F4F4F4"   # Primary text
        IBM_GREY_30 = "#C6C6C6" # Secondary text
        IBM_GREY_50 = "#525252" # Muted elements
        IBM_RED_50 = "#FA4D56"  # Errors
        IBM_GREEN_50 = "#24A148" # Success/Progress
        
        # Main container
        self.container = QFrame(self)
        self.container.setGeometry(0, 0, self.expanded_width, self.popup_height)
        self.container.setStyleSheet(f"""
            QFrame {{
                background-color: {IBM_GREY_90};
                border: 2px solid {IBM_BLUE_60};
                border-radius: 10px;
            }}
        """)
        
        layout = QVBoxLayout(self.container)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(6)
        
        # Header (configurable for different modes)
        self.header = QLabel("📋 New Cases Progress")
        self.header.setStyleSheet(f"""
            font-weight: bold;
            font-size: 16px;
            color: {IBM_BLUE_60};
            padding: 4px;
        """)
        self.header.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.header)
        
        # Progress label
        self.progress_label = QLabel("0 of 0 (0%)")
        self.progress_label.setStyleSheet(f"""
            font-size: 14px;
            font-weight: bold;
            color: {IBM_WHITE};
            padding: 2px;
        """)
        self.progress_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.progress_label)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                border: 2px solid {IBM_GREY_50};
                border-radius: 5px;
                text-align: center;
                height: 24px;
                background-color: {IBM_GREY_90};
                color: {IBM_WHITE};
                font-size: 12px;
                font-weight: bold;
            }}
            QProgressBar::chunk {{
                background-color: {IBM_GREEN_50};
                border-radius: 3px;
            }}
        """)
        layout.addWidget(self.progress_bar)
        
        # Status label
        self.status_label = QLabel("Processing...")
        self.status_label.setStyleSheet(f"""
            font-size: 13px;
            color: {IBM_GREY_30};
            padding: 2px;
        """)
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)
        
        layout.addSpacing(6)
        
        # Control buttons with IBM-styled colors
        # Pause button
        self.pause_btn = QPushButton("⏸ Pause")
        self.pause_btn.setStyleSheet(self._button_style("#FF9800", "#E68900"))
        self.pause_btn.clicked.connect(self.on_pause_clicked)
        layout.addWidget(self.pause_btn)
        
        # Continue button (hidden initially)
        self.continue_btn = QPushButton("▶ Continue")
        self.continue_btn.setStyleSheet(self._button_style(IBM_GREEN_50, "#1E8E3E"))
        self.continue_btn.clicked.connect(self.on_continue_clicked)
        self.continue_btn.hide()
        layout.addWidget(self.continue_btn)
        
        # Quit after current case button
        self.quit_after_btn = QPushButton("🚪 Quit After Current")
        self.quit_after_btn.setStyleSheet(self._button_style("#8A3FFC", "#7A35E0"))
        self.quit_after_btn.clicked.connect(self.on_quit_after_current)
        layout.addWidget(self.quit_after_btn)
        
        # Abort now button
        self.abort_btn = QPushButton("⛔ Abort Now")
        self.abort_btn.setStyleSheet(self._button_style(IBM_RED_50, "#DA1E28"))
        self.abort_btn.clicked.connect(self.on_abort_now)
        layout.addWidget(self.abort_btn)
        
        # Take a break button
        self.break_btn = QPushButton("☕ Take a Break")
        self.break_btn.setStyleSheet(self._button_style("#795548", "#5D4037"))
        self.break_btn.clicked.connect(self.on_take_a_break)
        self.break_btn.setToolTip("Pause after current case and set dialer to Break status")
        layout.addWidget(self.break_btn)
        
        # Back on Duty button (hidden initially)
        self.back_on_duty_btn = QPushButton("🟢 Back on Duty")
        self.back_on_duty_btn.setStyleSheet(self._button_style("#009688", "#00796B"))
        self.back_on_duty_btn.clicked.connect(self.on_back_on_duty)
        self.back_on_duty_btn.setToolTip("Set dialer to Not Ready (Chat) and resume processing")
        self.back_on_duty_btn.hide()
        layout.addWidget(self.back_on_duty_btn)
        
        # Error log display
        from PyQt5.QtWidgets import QTextEdit
        self.error_log = QTextEdit()
        self.error_log.setReadOnly(True)
        self.error_log.setMaximumHeight(70)
        self.error_log.setStyleSheet(f"""
            QTextEdit {{
                background-color: #1C1C1C;
                border: 1px solid {IBM_GREY_50};
                border-radius: 4px;
                color: {IBM_RED_50};
                font-size: 11px;
                padding: 4px;
            }}
        """)
        self.error_log.setPlaceholderText("No errors")
        layout.addWidget(self.error_log)
        
        # Collapse indicator (visible when collapsed) - clickable to expand
        self.collapse_indicator = QPushButton("◀", self)
        self.collapse_indicator.setStyleSheet(f"""
            QPushButton {{
                font-size: 18px;
                color: {IBM_BLUE_60};
                background-color: {IBM_GREY_90};
                border: 2px solid {IBM_BLUE_60};
                border-right: none;
                border-radius: 5px 0 0 5px;
                padding: 5px 3px;
            }}
            QPushButton:hover {{
                background-color: #262626;
            }}
        """)
        self.collapse_indicator.setFixedSize(28, 50)
        self.collapse_indicator.move(0, self.popup_height // 2 - 25)
        self.collapse_indicator.clicked.connect(self.expand)
        
    def _button_style(self, bg_color, hover_color):
        """Generate button stylesheet with IBM design"""
        return f"""
            QPushButton {{
                background-color: {bg_color};
                color: #F4F4F4;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                padding: 8px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {hover_color};
            }}
            QPushButton:pressed {{
                background-color: {hover_color};
            }}
            QPushButton:disabled {{
                background-color: #525252;
                color: #8D8D8D;
            }}
        """
    
    def log_error(self, message):
        """Append an error message to the error display (keeps last 10)"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.error_log.append(f"⚠ [{timestamp}] {message}")
        # Keep only last 10 errors to prevent memory bloat
        if self.error_log.document().blockCount() > 10:
            cursor = self.error_log.textCursor()
            cursor.movePosition(cursor.Start)
            cursor.movePosition(cursor.Down, cursor.KeepAnchor)
            cursor.removeSelectedText()
        # Scroll to bottom to show latest error
        self.error_log.verticalScrollBar().setValue(
            self.error_log.verticalScrollBar().maximum()
        )
    
    def set_mode(self, mode_title):
        """Update the header for different modes (New Cases / Case Reviewer)"""
        self.header.setText(mode_title)
    
    def position_on_screen(self):
        """Position the popup on the right edge of the screen (collapsed)"""
        screen = QApplication.primaryScreen().geometry()
        x = screen.width() - self.collapsed_width
        y = screen.height() // 2 - self.popup_height // 2
        self.setGeometry(x, y, self.expanded_width, self.popup_height)
        self.collapsed_x = x
        self.expanded_x = screen.width() - self.expanded_width
    
    def expand(self):
        """Animate expansion from right edge"""
        if self.is_expanded:
            return
        self.is_expanded = True
        self.collapse_timer.stop()
        
        current_geo = self.geometry()
        target_geo = QRect(self.expanded_x, current_geo.y(), 
                          self.expanded_width, self.popup_height)
        
        self.slide_animation.setStartValue(current_geo)
        self.slide_animation.setEndValue(target_geo)
        self.slide_animation.start()
        
        self.collapse_indicator.hide()
        self.container.show()
    
    def collapse(self):
        """Animate collapse to right edge"""
        if not self.is_expanded:
            return
        self.is_expanded = False
        
        current_geo = self.geometry()
        target_geo = QRect(self.collapsed_x, current_geo.y(),
                          self.expanded_width, self.popup_height)
        
        self.slide_animation.setStartValue(current_geo)
        self.slide_animation.setEndValue(target_geo)
        self.slide_animation.start()
        
        self.collapse_indicator.show()
    
    def enterEvent(self, event):
        """Mouse entered - expand"""
        self.collapse_timer.stop()
        self.expand()
        super().enterEvent(event)
    
    def leaveEvent(self, event):
        """Mouse left - start collapse timer (but not during case review to avoid sticky behavior)"""
        # Only auto-collapse if not paused and not on break - keeps popup accessible
        if not self.is_paused and not self.take_a_break:
            self.collapse_timer.start(2000)  # 2 second delay
        super().leaveEvent(event)
    
    def mousePressEvent(self, event):
        """Enable dragging"""
        if event.button() == Qt.LeftButton:
            self.drag_position = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()
    
    def mouseMoveEvent(self, event):
        """Handle dragging"""
        if event.buttons() == Qt.LeftButton and self.drag_position:
            new_pos = event.globalPos() - self.drag_position
            self.move(new_pos)
            # Update collapsed/expanded positions based on new location
            screen = QApplication.primaryScreen().geometry()
            self.collapsed_x = screen.width() - self.collapsed_width
            self.expanded_x = screen.width() - self.expanded_width
            event.accept()
    
    def mouseReleaseEvent(self, event):
        """End dragging"""
        self.drag_position = None
    
    def update_progress(self, cases_done, total_cases=None):
        """Update the progress display"""
        self.cases_done = cases_done
        if total_cases is not None:
            self.total_cases = total_cases
        
        if self.total_cases > 0:
            percentage = int((self.cases_done / self.total_cases) * 100)
        else:
            percentage = 0
        
        self.progress_label.setText(f"{self.cases_done} of {self.total_cases} ({percentage}%)")
        self.progress_bar.setValue(percentage)
        self.progress_bar.setFormat(f"{self.cases_done}/{self.total_cases}")
        
        # Update status based on state
        if self.is_paused:
            self.status_label.setText("⏸ Paused - waiting...")
            self.status_label.setStyleSheet("font-size: 11px; color: #FFB74D; padding: 3px;")
        elif self.cases_done >= self.total_cases:
            self.status_label.setText("✅ All cases completed!")
            self.status_label.setStyleSheet("font-size: 11px; color: #81C784; padding: 3px;")
        else:
            self.status_label.setText("Processing...")
            self.status_label.setStyleSheet("font-size: 11px; color: #90CAF9; padding: 3px;")
    
    def on_pause_clicked(self):
        """Handle pause button click"""
        self.is_paused = True
        self.pause_btn.hide()
        self.continue_btn.show()
        self.status_label.setText("⏸ Pausing after current case...")
        self.status_label.setStyleSheet("font-size: 11px; color: #FFB74D; padding: 3px;")
        self.pause_signal.emit()
    
    def on_continue_clicked(self):
        """Handle continue button click"""
        self.is_paused = False
        self.continue_btn.hide()
        self.pause_btn.show()
        self.status_label.setText("Processing...")
        self.status_label.setStyleSheet("font-size: 11px; color: #90CAF9; padding: 3px;")
        self.continue_signal.emit()
    
    def on_quit_after_current(self):
        """Handle quit after current case button"""
        self.quit_after_current = True
        self.quit_after_btn.setText("🚪 Exiting after current...")
        self.quit_after_btn.setEnabled(False)
        self.status_label.setText("🚪 Will exit after current case...")
        self.status_label.setStyleSheet("font-size: 11px; color: #CE93D8; padding: 3px;")
        self.quit_after_current_signal.emit()
    
    def on_abort_now(self):
        """Handle abort now button"""
        self.abort_now = True
        self.status_label.setText("⛔ Aborting...")
        self.status_label.setStyleSheet("font-size: 11px; color: #EF5350; padding: 3px;")
        self.abort_now_signal.emit()
    
    def on_take_a_break(self):
        """Handle take a break button - pauses after current case and sets dialer to break"""
        self.take_a_break = True
        self.is_paused = True  # Also pause the processing
        self.break_btn.setText("☕ On Break")
        self.break_btn.setEnabled(False)
        self.pause_btn.hide()
        self.continue_btn.hide()
        self.back_on_duty_btn.show()  # Show the Back on Duty button
        self.status_label.setText("☕ On break - click Back on Duty when ready")
        self.status_label.setStyleSheet("font-size: 11px; color: #A1887F; padding: 3px;")
        self.take_a_break_signal.emit()
    
    def on_back_on_duty(self):
        """Handle back on duty button - sets dialer to Not Ready (Chat) and resumes processing"""
        self.take_a_break = False
        self.is_paused = False
        self.back_on_duty_btn.hide()
        self.pause_btn.show()
        self.break_btn.setText("☕ Take a Break")
        self.break_btn.setEnabled(True)
        self.status_label.setText("🟢 Back on duty - resuming...")
        self.status_label.setStyleSheet("font-size: 11px; color: #4CAF50; padding: 3px;")
        self.back_on_duty_signal.emit()


# ============================================================================
# CASE REVIEWER POPUP - Simple floating widget for case review with error log
# ============================================================================
class CaseReviewerPopup(QWidget):
    """
    A simple floating popup for Case Reviewer mode.
    Features: Error log display, Take a Break button.
    Always visible (no collapse), positioned in bottom-right corner.
    """
    take_a_break_signal = pyqtSignal()
    back_on_duty_signal = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_on_break = False
        
        # Window flags: frameless, always on top, tool window
        self.setWindowFlags(
            Qt.FramelessWindowHint | 
            Qt.WindowStaysOnTopHint | 
            Qt.Tool
        )
        self.setAttribute(Qt.WA_TranslucentBackground, False)
        
        # Dimensions - compact size
        self.popup_width = 260
        self.popup_height = 180
        
        self.setup_ui()
        self.position_on_screen()
    
    def setup_ui(self):
        """Setup the simple UI"""
        self.setFixedSize(self.popup_width, self.popup_height)
        
        # IBM Design Colors
        IBM_GREY_90 = "#161616"
        IBM_BLUE_60 = "#0F62FE"
        IBM_WHITE = "#F4F4F4"
        IBM_RED_50 = "#FA4D56"
        IBM_GREEN_50 = "#24A148"
        
        # Main container
        self.container = QFrame(self)
        self.container.setGeometry(0, 0, self.popup_width, self.popup_height)
        self.container.setStyleSheet(f"""
            QFrame {{
                background-color: {IBM_GREY_90};
                border: 2px solid {IBM_BLUE_60};
                border-radius: 8px;
            }}
        """)
        
        layout = QVBoxLayout(self.container)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)
        
        # Header
        header = QLabel("📋 Case Reviewer")
        header.setStyleSheet(f"font-weight: bold; font-size: 14px; color: {IBM_BLUE_60}; padding: 2px;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        # Error log display
        from PyQt5.QtWidgets import QTextEdit
        self.error_log = QTextEdit()
        self.error_log.setReadOnly(True)
        self.error_log.setMaximumHeight(80)
        self.error_log.setStyleSheet(f"""
            QTextEdit {{
                background-color: #1C1C1C;
                border: 1px solid #525252;
                border-radius: 4px;
                color: {IBM_RED_50};
                font-size: 11px;
                padding: 4px;
            }}
        """)
        self.error_log.setPlaceholderText("No errors")
        layout.addWidget(self.error_log)
        
        # Take a Break button
        self.break_btn = QPushButton("☕ Take a Break")
        self.break_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #795548;
                color: {IBM_WHITE};
                border: none;
                border-radius: 5px;
                font-weight: bold;
                padding: 10px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: #5D4037;
            }}
            QPushButton:disabled {{
                background-color: #525252;
                color: #8D8D8D;
            }}
        """)
        self.break_btn.clicked.connect(self.on_take_a_break)
        layout.addWidget(self.break_btn)
        
        # Back on Duty button (hidden initially)
        self.back_on_duty_btn = QPushButton("🟢 Back on Duty")
        self.back_on_duty_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {IBM_GREEN_50};
                color: {IBM_WHITE};
                border: none;
                border-radius: 5px;
                font-weight: bold;
                padding: 10px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: #1E8E3E;
            }}
        """)
        self.back_on_duty_btn.clicked.connect(self.on_back_on_duty)
        self.back_on_duty_btn.hide()
        layout.addWidget(self.back_on_duty_btn)
    
    def position_on_screen(self):
        """Position in bottom-right corner"""
        screen = QApplication.primaryScreen().geometry()
        x = screen.width() - self.popup_width - 20
        y = screen.height() - self.popup_height - 80  # Above taskbar
        self.setGeometry(x, y, self.popup_width, self.popup_height)
    
    def log_error(self, message):
        """Append an error message to the error display (keeps last 10)"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.error_log.append(f"⚠ [{timestamp}] {message}")
        # Keep only last 10 errors
        if self.error_log.document().blockCount() > 10:
            cursor = self.error_log.textCursor()
            cursor.movePosition(cursor.Start)
            cursor.movePosition(cursor.Down, cursor.KeepAnchor)
            cursor.removeSelectedText()
        # Scroll to bottom
        self.error_log.verticalScrollBar().setValue(
            self.error_log.verticalScrollBar().maximum()
        )
    
    def on_take_a_break(self):
        """Handle take a break button"""
        self.is_on_break = True
        self.break_btn.hide()
        self.back_on_duty_btn.show()
        self.take_a_break_signal.emit()
    
    def on_back_on_duty(self):
        """Handle back on duty button"""
        self.is_on_break = False
        self.back_on_duty_btn.hide()
        self.break_btn.show()
        self.back_on_duty_signal.emit()


def set_dialer_break_status(driver):
    """
    Switch to dialer tab and set status to Not Ready (Break).
    Used by the Take a Break feature.
    """
    try:
        # Get all window handles
        handles = driver.window_handles
        if len(handles) > 0:
            # Dialer is typically at index 0
            dialer_handle = handles[0]
            try:
                driver.switch_to.window(dialer_handle)
                print("[INFO] Switched to dialer window for break status")
            except Exception as e:
                print(f"[WARN] Failed to switch to dialer: {e}")
                return False
        else:
            print("[ERROR] No dialer window found")
            return False
        
        # Set the Status to Not Ready - Option 2 (Break)
        click_safe(driver, By.XPATH, '//span[contains(@class,"wwe-sprite-mark-done")]', timeout=1, retries=1)
        time.sleep(0.5)
        click_safe(driver, By.XPATH, "//table[@id='DataTables_Table_0']/tbody/tr/td[2]", timeout=1, retries=3)
        time.sleep(0.5)
        click_safe(driver, By.ID, "wweAgentSetNotReadyReason2Item_MyChannelsView", timeout=1, retries=3)
        
        print("[INFO] Dialer status set to Not Ready (Break)")
        return True
    except Exception as e:
        print(f"[WARN] set_dialer_break_status encountered an error: {e}")
        traceback.print_exc()
        return False


def set_dialer_chat_status(driver):
    """
    Switch to dialer tab and set status to Not Ready (Chat).
    Used by the Back on Duty feature to resume from break.
    """
    try:
        # Get all window handles
        handles = driver.window_handles
        if len(handles) > 0:
            # Dialer is typically at index 0
            dialer_handle = handles[0]
            try:
                driver.switch_to.window(dialer_handle)
                print("[INFO] Switched to dialer window for chat status")
            except Exception as e:
                print(f"[WARN] Failed to switch to dialer: {e}")
                return False
        else:
            print("[ERROR] No dialer window found")
            return False
        
        # Set the Status to Not Ready - Option 4 (Chat)
        click_safe(driver, By.XPATH, '//span[contains(@class,"wwe-sprite-mark-done")]', timeout=1, retries=1)
        time.sleep(0.5)
        click_safe(driver, By.XPATH, "//table[@id='DataTables_Table_0']/tbody/tr/td[2]", timeout=1, retries=3)
        time.sleep(0.5)
        click_safe(driver, By.ID, "wweAgentSetNotReadyReason4Item_MyChannelsView", timeout=1, retries=3)
        
        print("[INFO] Dialer status set to Not Ready (Chat)")
        return True
    except Exception as e:
        print(f"[WARN] set_dialer_chat_status encountered an error: {e}")
        traceback.print_exc()
        return False

def show_new_cases_popup_preview():
    """
    Show a preview of the New Cases Progress Popup for testing.
    This simulates the popup without actual case processing.
    """
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    popup = NewCasesProgressPopup(total_cases=10)
    popup.show()
    
    # Simulate progress updates
    current_case = [0]
    
    def simulate_progress():
        if current_case[0] < 10:
            current_case[0] += 1
            popup.update_progress(current_case[0], 10)
    
    # Create timer to simulate progress
    timer = QTimer()
    timer.timeout.connect(simulate_progress)
    timer.start(2000)  # Update every 2 seconds
    
    print("[INFO] New Cases Progress Popup Preview shown. Close the window to exit.")
    app.exec_()


def get_call_closing_code():
    """
    Opens a call outcome dialog for the current call.
    Returns the selected closing code and whether to add a note.
    """
    closing_code = {"value": None, "add_note": False}

    class CallOutcomeDialog(QDialog):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Call Closing Code")
            self.resize(400, 450)
            
            # Main layout
            main_layout = QVBoxLayout(self)
            main_layout.setContentsMargins(10, 10, 10, 10)
            main_layout.setSpacing(10)
            
            label = QLabel("Select Call Closing Code:")
            label.setStyleSheet("font-weight: bold; font-size: 12px;")
            main_layout.addWidget(label)
            
            # Button grid
            btn_frame = QWidget()
            grid = QGridLayout(btn_frame)
            grid.setSpacing(8)
            grid.setColumnStretch(0, 1)
            grid.setColumnStretch(1, 1)
            
            buttons = [
                ("Issue Resolved", "Called - Answered: Issue Resolved"),
                ("Issue Not Fixed", "Called - Answered: Issue Not Resolved"),
                ("Not Reached", "Customer Not Reached"),
                ("Not Yet Tested", "Customer Claims that the Machine Not Yet Tested"),
                ("Left Voicemail", "Called: Not Reached // left Voicemail"),
                ("Ext Not Found", "Called - Company NO. Extension Found: Not Reached"),
            ]
            
            def set_code_and_close(code):
                closing_code["value"] = code
                try:
                    closing_code["add_note"] = bool(self.add_note_checkbox.isChecked())
                except Exception:
                    closing_code["add_note"] = False
                self.accept()
            
            def ask_other_code():
                try:
                    text, ok = QInputDialog.getText(self, "Custom Closing Code", "Enter custom closing code:")
                    if ok and text.strip():
                        set_code_and_close(text.strip())
                except Exception as e:
                    print(f"[WARN] Custom code input failed: {e}")
            
            # Add buttons to grid
            for idx, (label_text, code) in enumerate(buttons):
                r = idx // 2
                c = idx % 2
                btn = QPushButton(label_text)
                btn.setMinimumHeight(40)
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: #E3F2FD;
                        border: 1px solid #CCCCCC;
                        border-radius: 4px;
                        font-weight: bold;
                        color: #333;
                        padding: 6px;
                        font-size: 15px;
                    }
                    QPushButton:hover {
                        border: 2px solid #1976D2;
                    }
                    QPushButton:pressed {
                        background-color: #90CAF9;
                    }
                """)
                btn.clicked.connect(lambda checked=False, c=code: set_code_and_close(c))
                grid.addWidget(btn, r, c)
            
            main_layout.addWidget(btn_frame)
            
            # Custom code button
            other_btn = QPushButton("Custom Code")
            other_btn.setMinimumHeight(40)
            other_btn.setStyleSheet("""
                QPushButton {
                    background-color: #FFF3E0;
                    border: 1px solid #CCCCCC;
                    border-radius: 4px;
                    font-weight: bold;
                    color: #333;
                    padding: 6px;
                }
                QPushButton:hover {
                    border: 2px solid #FF9800;
                }
            """)
            other_btn.clicked.connect(ask_other_code)
            main_layout.addWidget(other_btn)
            
            # Add note checkbox
            self.add_note_checkbox = QCheckBox("✓ Add Case Note")
            self.add_note_checkbox.setStyleSheet("font-size: 12px; padding: 5px;")
            main_layout.addWidget(self.add_note_checkbox)
            
            main_layout.addStretch()
            self.setLayout(main_layout)
    
    # Ensure QApplication exists
    app = QApplication.instance()
    created_app = False
    if app is None:
        app = QApplication(sys.argv)
        created_app = True
    
    # Show dialog
    dialog = CallOutcomeDialog()
    print(f"[INFO] Waiting for call outcome selection...")
    dialog.exec_()
    
    if created_app:
        app.quit()
    
    # Return tuple: (code, add_note)
    result_code = closing_code.get("value")
    result_note = bool(closing_code.get("add_note", False))
    print(f"[INFO] Call closing code selected: {result_code}, add_note={result_note}")
    
    return result_code, result_note

    
def cases_needing_processing_by_closing(df, case_col='case_number', closing_col='CaseClosingCode'):

    try:
        if case_col not in df.columns:
            return []
        if closing_col not in df.columns:
            # No closing code column yet — process all
            return df[case_col].dropna().astype(str).tolist()

        def _emptyish(v):
            try:
                s = str(v).strip()
                return (s == '' or s.lower() == 'nan')
            except Exception:
                return True

        mask = df[closing_col].apply(_emptyish)
        return df.loc[mask, case_col].dropna().astype(str).tolist()
    except Exception as e:
        print(f"[WARN] cases_needing_processing_by_closing failed: ")
        return []

def DND_CX(driver, case_number):
            
    # Contact Card
    contact_xpath = "//a[contains(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'sec_tab_contact-associatedEntityRecordName')]"
    if not click_safe(driver, By.XPATH, contact_xpath, timeout=1, retries=5):
        print(f"[WARN] Contact card not found for {case_number}")
    else:
        # click on edit button
        edit_xpath = "//button[contains(@id,'contact|NoRelationship|Form|lvidg.contact.TimeTrackingCheckOut.Command') and contains(@id,'-button')]"
        click_safe(driver, By.XPATH, edit_xpath, timeout=1, retries=5)

        # click on Details Tab
        details_xpath = "//li[contains(@id,'tab4') and @title='Details']"
        click_safe(driver, By.XPATH, details_xpath, timeout=1, retries=5)
        # click on Communication Preferences (Do Not Disturb)
        #xpath=//div[@id='id-cc760212-3eab-4009-b5c3-ea9cf6f3e141-86-lvidg_hasdonotdisturbf9a8a302-114e-466a-b582-6771b2ae0d92']/div[3]/div/div
        comm_pref_xpath = "//div[contains(@id,'lvidg_hasdonotdisturbf9a8a302')]/div[3]/div/div"
        click_safe(driver, By.XPATH, comm_pref_xpath, timeout=1, retries=5)

        time.sleep(2)

        # Save and Close
        save_xpath = "//button[contains(@id,'contact|NoRelationship|Form|lvidg.contact.TimeTrackingCheckIn.Command') and contains(@id,'button')]"
        click_safe(driver, By.XPATH, save_xpath, timeout=1, retries=5)    

def Chrome_ART_Profile():
    chrome_options = Options()
    user_data_dir = os.environ.get('CHROME_USER_DATA_DIR') or os.path.join(os.path.expanduser('~'), '@')
    profile_dir = os.environ.get('CHROME_PROFILE_DIR') or 'Default'
    if user_data_dir and os.path.exists(user_data_dir):
        try:
            chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
            chrome_options.add_argument(f'--profile-directory={profile_dir}')
            print(f"[INFO] Using Chrome profile: {user_data_dir}")
        except Exception as e:
            print(f"[WARN] Failed to set Chrome profile: ")
    else:
        # Create profile directory if it doesn't exist
        try:
            os.makedirs(user_data_dir, exist_ok=True)
            chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
            chrome_options.add_argument(f'--profile-directory={profile_dir}')
            print(f"[INFO] Created new Chrome profile at: {user_data_dir}")
        except Exception as e:
            print(f"[WARN] Failed to create Chrome profile: ")
    return chrome_options

def send_SMS(driver, sms_text):

    sms_sent = False

    # Wait for plus sign and open menu (Send SMS)
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
        timeout=2,
        retries=5,
    )

    # SMS Button
    click_SMS_button = click_safe(
        driver,
        By.XPATH,
        "//li[contains(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_lvidg_sms')]",
        timeout=2,
        retries=5,
    )

    # If SMS Button not found, try again once
    if not click_SMS_button:
        # open menu (Send SMS)
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
            timeout=1,
            retries=5,
        )

        # SMS Button
        click_SMS_button = click_safe(
            driver,
            By.XPATH,
            "//li[contains(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_lvidg_sms')]",
            timeout=2,
            retries=5,
        )
        
        # USFC discard dialog (if appears) - best-effort
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
            timeout=1,
            retries=5,
        )

        #clicking OK Button if avail
        click_safe(
            driver,
            By.XPATH,
            "//button[contains(@id,'okButton_')]/div",
            timeout=2,
            retries=2,
        )

        #sleep timer
        time.sleep(3)
    
    else:
        pass    


    # USFC discard dialog - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    # SMS TextBox -> enter text
    send_keys_safe(
        driver,
        By.XPATH,
        "//input[starts-with(@id,'id-3145bfd3-91e7-4364-92ed-5ca0cf0d65b8') and contains(@id,'subject.fieldControl-text-input-component')]",
        sms_text,
        timeout=2,
        retries=5,
    )

    # SMS Save & Close
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'lvidg_sms|NoRelationship|Form|Mscrm.SaveAndClosePrimary') and contains(@id, '-button')]",
        timeout=2,
        retries=5,
    )

    #SMS sent failed
    SMS_error_el = safe_find(driver, By.XPATH, "//div[contains(@id,'message') and contains(@id,'+lvidg_phonenumber')]/div/span", timeout=3, retries=2)
    if SMS_error_el:
        click_safe(
            driver,
            By.XPATH,
            "//button[@id='navigateBackButtontab-id-0']/span",
            timeout=1,
            retries=5,
        )
        sms_sent = False
    else:
        sms_sent = True



    # USFC discard dialog - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    return sms_sent

def add_Case_Note(driver, CaseNote):
    # To Create a Case Note: open add menu
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
        timeout=1,
        retries=5,
    )

    # Case Notes button
    note_button = click_safe(
        driver,
        By.XPATH,
        "//li[starts-with(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'notescontrol-createNewRecord_flyoutMenuItem_notes')]",
        timeout=1,
        retries=5,
    )
    if not note_button:
        # To Create a Case Note: open add menu
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
            timeout=1,
            retries=5,
        )

        # Case Notes button
        note_button = click_safe(
            driver,
            By.XPATH,
            "//li[starts-with(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'notescontrol-createNewRecord_flyoutMenuItem_notes')]",
            timeout=1,
            retries=5,
        )

        
        # USFC discard dialog (if appears) - best-effort
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
            timeout=1,
            retries=5,
        )

        #clicking OK Button if avail
        click_safe(
            driver,
            By.XPATH,
            "//button[contains(@id,'okButton_')]/div",
            timeout=2,
            retries=2,
        )

        #sleep timer
        time.sleep(3)
        
    else:
        pass

    # USFC discard dialog - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    # Case Notes editor -> enter notes
    send_ok = send_keys_safe(driver, By.XPATH, "//*[contains(@id, 'rtev')]", CaseNote, timeout=2, retries=5)
    time.sleep(2)  # small wait for text to register

    # Try to save the note by clicking 'Add note and close', fallback to Ctrl+S
    note_saved = False
    try:
        saved = click_safe(driver, By.XPATH, "//span[contains(@class,'ms-Button-label') and text()='Add note and close']", timeout=2, retries=4)
        time.sleep(2)  # small wait for save to complete
        if saved:
            note_saved = True
        else:
            try:
                # Press Tab, Tab, Enter to save the note (navigate to save button and activate)
                pyautogui.press(['tab', 'tab', 'enter'])
                time.sleep(2)  # small wait for save to complete
                note_saved = True
            except Exception as e:
                print(f"[WARN] Failed to send Tab/Tab/Enter for {case_number}: ")
    except Exception as e:
        print(f"[WARN] Save note step failed for {case_number}: ")
    return note_saved

def perform_dialer_login(driver):
        # Navigate to dialer
        print("[INFO] Opening dialer...")
        driver.get(DIALER_URL)
        time.sleep(3)  # Wait for page to load
        
        # Enter username
        print("[INFO] Entering username...")
        send_keys_safe(driver, By.ID, 'wweLoginUserNameField', DIALER_USERNAME, timeout=2, retries=120)
        
        # Enter Password
        print("[INFO] Entering password...")
        send_keys_safe(driver, By.ID, 'wweLoginPasswordField', DIALER_PASSWORD, timeout=2, retries=120, enter=True)
        # Enter PlaceID
        print("[INFO] Entering placeholder...")
        send_keys_safe(driver, By.ID, 'wweLoginPlaceInput', DIALER_PLACE_ID, timeout=2, retries=120, enter=True)
        time.sleep(10)  # wait for authentication

        handles = driver.window_handles
        dialer_handle = handles[0]
        driver.switch_to.window(dialer_handle)
         #Set the Status to Not Ready - Option 4 (Chat)
        click_safe(driver, By.XPATH, '//span[contains(@class,"wwe-sprite-mark-done")]', timeout=1, retries=1) 
        click_safe(driver, By.XPATH, "//table[@id='DataTables_Table_0']/tbody/tr/td[2]", timeout=1, retries=3)
        click_safe(driver, By.ID, "wweAgentSetNotReadyReason4Item_MyChannelsView", timeout=1, retries=3)

        # Robust CRM window switching with multiple strategies
        crm_switched = switch_to_crm_window(driver)
        
        if not crm_switched:
            print(f"[ERROR] Failed to switch to CRM window after multiple attempts. Program may be unstable.")
            return False
        
        return True

def switch_to_crm_window(driver, max_retries=5):
    """
    Robustly switch to CRM window using multiple strategies.
    Returns True if successful, False otherwise.
    """
    # Strategy 1: Identify CRM window by URL pattern
    crm_url_patterns = ["dynamics.microsoft.com", "crm", "power"]
    
    for attempt in range(1, max_retries + 1):
        # Exponential backoff: wait longer with each attempt
        wait_time = 2 ** (attempt - 1)  # 1, 2, 4, 8, 16 seconds
        if attempt > 1:
            print(f"[WARN] Switch attempt {attempt}/{max_retries}, waiting {wait_time}s before retry...")
            time.sleep(wait_time)
        
        # Get fresh window handles
        handles = driver.window_handles
        print(f"[DEBUG] Attempt {attempt}: Found {len(handles)} window handles")
        
        if len(handles) < 2:
            print(f"[ERROR] Expected 2+ windows, but found {len(handles)}")
            continue
        
        # Strategy 1: Try to find CRM by URL
        crm_handle = None
        for handle in handles:
            try:
                driver.switch_to.window(handle)
                current_url = driver.current_url.lower()
                print(f"[DEBUG] Handle URL: {current_url}")
                
                # Check if this window contains CRM URL patterns
                if any(pattern in current_url for pattern in crm_url_patterns):
                    crm_handle = handle
                    print(f"[INFO] Found CRM window by URL pattern")
                    break
            except Exception as e:
                print(f"[WARN] Error checking handle URL: {e}")
                continue
        
        # Strategy 2: If URL method fails, try verifying with CRM-specific elements
        if crm_handle is None and len(handles) >= 2:
            print(f"[DEBUG] URL matching failed, trying element-based detection...")
            for handle in handles:
                try:
                    driver.switch_to.window(handle)
                    # Check for a CRM-specific element (GlobalSearchBox is CRM-only)
                    crm_el = driver.find_elements(By.ID, "GlobalSearchBox")
                    if crm_el:
                        crm_handle = handle
                        print(f"[INFO] Found CRM window by GlobalSearchBox element")
                        break
                except Exception as e:
                    continue
        
        # Strategy 3: If both fail, use index-based approach (fallback)
        if crm_handle is None:
            print(f"[DEBUG] Element detection failed, trying index-based fallback...")
            for idx, handle in enumerate(handles):
                if idx != 0:  # Skip first (likely dialer)
                    try:
                        driver.switch_to.window(handle)
                        crm_handle = handle
                        print(f"[INFO] Switched to handle index {idx} (fallback)")
                        break
                    except Exception as e:
                        continue
        
        # If we found a CRM handle, verify the switch was successful
        if crm_handle is not None:
            try:
                driver.switch_to.window(crm_handle)
                # Extra verification: check if GlobalSearchBox is accessible
                try:
                    WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.ID, "GlobalSearchBox"))
                    )
                    print(f"[SUCCESS] ✓ Successfully switched to CRM window and verified GlobalSearchBox")
                    return True
                except TimeoutException:
                    print(f"[WARN] Switched to handle but GlobalSearchBox not found yet, may still be loading")
                    # Give it one more moment to load
                    time.sleep(2)
                    try:
                        driver.find_element(By.ID, "GlobalSearchBox")
                        print(f"[SUCCESS] ✓ CRM window verified after additional wait")
                        return True
                    except:
                        print(f"[WARN] GlobalSearchBox still not found, continuing retry loop")
                        continue
            except Exception as e:
                print(f"[ERROR] Failed to switch to identified CRM handle: {e}")
                continue
    
    print(f"[CRITICAL] Failed to switch to CRM window after {max_retries} attempts")
    return False

def verify_email_domain_selected(driver, expected_title="NA Think Care", expected_text="NA Think Care", timeout=3, poll=0.5, retries= 3):
    xpath = "//*[@data-id='from.fieldControl-LookupResultsDropdown_from_selected_tag_text']"
    def _check(d):
        try:
            el = d.find_element(By.XPATH, xpath)
            title = (el.get_attribute('title') or '').strip()
            text = (el.text or '').strip()
            return title == expected_title and text == expected_text
        except Exception:
            return False
    return _check(driver)



    sms_sent = False

    # Wait for plus sign and open menu (Send SMS)
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
        timeout=2,
        retries=5,
    )

    # SMS Button
    click_SMS_button = click_safe(
        driver,
        By.XPATH,
        "//li[contains(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_lvidg_sms')]",
        timeout=2,
        retries=5,
    )

    # If SMS Button not found, try again once
    if not click_SMS_button:
        # open menu (Send SMS)
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
            timeout=1,
            retries=5,
        )

        # SMS Button
        click_SMS_button = click_safe(
            driver,
            By.XPATH,
            "//li[contains(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_lvidg_sms')]",
            timeout=2,
            retries=5,
        )
        
        # USFC discard dialog (if appears) - best-effort
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
            timeout=1,
            retries=5,
        )

        #clicking OK Button if avail
        click_safe(
            driver,
            By.XPATH,
            "//button[contains(@id,'okButton_')]/div",
            timeout=2,
            retries=2,
        )

        #sleep timer
        time.sleep(3)
    
    else:
        pass    


    # USFC discard dialog - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    # SMS TextBox -> enter text
    send_keys_safe(
        driver,
        By.XPATH,
        "//input[starts-with(@id,'id-3145bfd3-91e7-4364-92ed-5ca0cf0d65b8') and contains(@id,'subject.fieldControl-text-input-component')]",
        sms_text,
        timeout=2,
        retries=5,
    )

    # SMS Save & Close
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'lvidg_sms|NoRelationship|Form|Mscrm.SaveAndClosePrimary') and contains(@id, '-button')]",
        timeout=2,
        retries=5,
    )

    #SMS sent failed
    SMS_error_el = safe_find(driver, By.XPATH, "//div[contains(@id,'message') and contains(@id,'+lvidg_phonenumber')]/div/span", timeout=3, retries=2)
    if SMS_error_el:
        click_safe(
            driver,
            By.XPATH,
            "//button[@id='navigateBackButtontab-id-0']/span",
            timeout=1,
            retries=5,
        )
        sms_sent = False
    else:
        sms_sent = True



    # USFC discard dialog - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    return sms_sent

# Define locators centrally for easier maintenance
LOCATORS = {
    # Add Menu Button - Using more stable text/aria attributes if possible, but keeping original structure as a base
    "ADD_MENU_BUTTON": (
        By.XPATH, 
        "//button[contains(@id, 'notescontrol-action_bar_add_command')]"
    ),
    # Email Button in Flyout - Prioritize text content if reliable
    "EMAIL_FLYOUT_BUTTON": (
        By.XPATH, 
        "//li[contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_email')]"
    ),
    # Discard Changes Dialog Button (USFC) - More stable text-based locator
    "DISCARD_CHANGES_BUTTON": (
        By.XPATH, 
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]"
    ),
    # Clear From Email (x) Button - Using text/title if available, otherwise shortening ID
    "CLEAR_FROM_BUTTON": (
        By.XPATH, 
        "//*[contains(@id, 'from.fieldControl-LookupResultsDropdown_from') and contains(@id, 'microsoftIcon_cancelButton_')]"
    ),
    # From Email Input Field
    "FROM_EMAIL_INPUT": (
        By.XPATH, 
        "//*[contains(@id, 'from.fieldControl-LookupResultsDropdown_from') and contains(@id, 'textInputBox_with_filter_new')]"
    ),
    # Pick Email Domain Entry (NA Think Care) - Highly unstable ID, consider using text or position instead.
    # **WARNING:** This ID is hardcoded and extremely brittle. You should find a text-based or list-item based locator.
    "PICK_EMAIL_ENTRY": (
        By.ID, 
        "id-1d5ad078-3edb-4edc-98c6-c0c21e3125e3-45-fromcbfb742c-14e7-4a17-96bb-1a13f7f64aa2-from.fieldControl-name0_0_0" 
    ),
    # Email Body (Rich Text Editor)
    "EMAIL_BODY_INPUT": (
        By.XPATH, 
        "//div[starts-with(@id, 'rtev')]/p"
    ),
    # Send Email Button
    "SEND_BUTTON": (
        By.XPATH, 
        "//button[starts-with(@id, 'email|NoRelationship|Form|Mscrm.Form.email.Send')]"
    ),
    # Navigate Back Button (used in failure path)
    "NAVIGATE_BACK_BUTTON": (
        By.XPATH, 
        "//button[@id='navigateBackButtontab-id-0']/span"
    )
}

def select_from_email(driver):
    """
    Handles clearing the existing 'From' address, entering the new one, and selecting it.
    Uses EMAIL_DOMAIN and EMAIL_DOMAIN_NAME from config.
    Returns True on success, False otherwise.
    """
    # 1. Clear existing 'From' address (best-effort click)
    click_safe(driver, *LOCATORS["CLEAR_FROM_BUTTON"], timeout=1, retries=3)

    # 2. Enter email address
    send_keys_safe(driver, *LOCATORS["FROM_EMAIL_INPUT"], EMAIL_DOMAIN, timeout=2, retries=5)

    # 3. Pick email domain entry (This part is highly unstable due to the brittle locator)
    if not click_safe(driver, *LOCATORS["PICK_EMAIL_ENTRY"], timeout=5, retries=5):
         # If the click failed, or if the selection didn't register:
         return False

    time.sleep(2) # Small wait for selection to register

    # 4. Verify selection
    return verify_email_domain_selected(
        driver, 
        expected_title=EMAIL_DOMAIN_NAME, 
        expected_text=EMAIL_DOMAIN_NAME,
        timeout=5, 
        poll=0.5, 
        retries=3
    )

def send_Email(driver, email_text):
    """
    Attempts to compose and send an email, returning True or False.
    """
    # 1. Open 'Add' Menu
    if not click_safe(driver, *LOCATORS["ADD_MENU_BUTTON"], timeout=5, retries=3):
        print("Error: Could not open Add Menu.")
        return False
    
    # 2. Click 'Email' Button
    if not click_safe(driver, *LOCATORS["EMAIL_FLYOUT_BUTTON"], timeout=5, retries=5):
        # Retry the entire open-click sequence once if the flyout button failed (robustness for slow UIs)
        print("Warning: Email button failed, retrying menu open and click.")
        click_safe(driver, *LOCATORS["ADD_MENU_BUTTON"], timeout=5, retries=3)
        if not click_safe(driver, *LOCATORS["EMAIL_FLYOUT_BUTTON"], timeout=5, retries=5):
             print("Error: Could not click Email Button after retry.")
             return False

    # 3. Handle Discard Dialog (best-effort, non-blocking)
    # This should happen only if a previous email window was partially open.
    click_safe(driver, *LOCATORS["DISCARD_CHANGES_BUTTON"], timeout=1, retries=3) 

    time.sleep(2)  # Small wait to ensure dialog is handled

    # 4. Select the correct 'From' email address
    if not select_from_email(driver):
        print("Error: Failed to select 'From' email address after multiple attempts.")
        
        # If selection fails, attempt to close the email form to clean up
        click_safe(driver, *LOCATORS["NAVIGATE_BACK_BUTTON"], timeout=3, retries=3)
        click_safe(driver, *LOCATORS["DISCARD_CHANGES_BUTTON"], timeout=1, retries=3)
        return False
    
    # 5. Type Email Body
    if not send_keys_safe(driver, *LOCATORS["EMAIL_BODY_INPUT"], email_text, timeout=5, retries=5):
        print("Error: Failed to enter email body text.")
        return False

    # 6. Send Email
    if not click_safe(driver, *LOCATORS["SEND_BUTTON"], timeout=10, retries=5):
        print("Error: Failed to click Send Email button.")
        return False
    time.sleep(3)  # Small wait to ensure send action is processed
    return True # Email assumed sent successfully


    # To Create a Case Note: open add menu
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
        timeout=1,
        retries=5,
    )

    # Case Notes button
    note_button = click_safe(
        driver,
        By.XPATH,
        "//li[starts-with(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'notescontrol-createNewRecord_flyoutMenuItem_notes')]",
        timeout=1,
        retries=5,
    )
    if not note_button:
        # To Create a Case Note: open add menu
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
            timeout=1,
            retries=5,
        )

        # Case Notes button
        note_button = click_safe(
            driver,
            By.XPATH,
            "//li[starts-with(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'notescontrol-createNewRecord_flyoutMenuItem_notes')]",
            timeout=1,
            retries=5,
        )

        
        # USFC discard dialog (if appears) - best-effort
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
            timeout=1,
            retries=5,
        )

        #clicking OK Button if avail
        click_safe(
            driver,
            By.XPATH,
            "//button[contains(@id,'okButton_')]/div",
            timeout=2,
            retries=2,
        )

        #sleep timer
        time.sleep(3)
        
    else:
        pass

    # USFC discard dialog - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    # Case Notes editor -> enter notes
    send_ok = send_keys_safe(driver, By.XPATH, "//*[contains(@id, 'rtev')]", CaseNote, timeout=2, retries=5)
    time.sleep(2)  # small wait for text to register

    # Try to save the note by clicking 'Add note and close', fallback to Ctrl+S
    note_saved = False
    try:
        saved = click_safe(driver, By.XPATH, "//span[contains(@class,'ms-Button-label') and text()='Add note and close']", timeout=2, retries=4)
        time.sleep(2)  # small wait for save to complete
        if saved:
            note_saved = True
        else:
            try:
                # Press Tab, Tab, Enter to save the note (navigate to save button and activate)
                pyautogui.press(['tab', 'tab', 'enter'])
                time.sleep(2)  # small wait for save to complete
                note_saved = True
            except Exception as e:
                print(f"[WARN] Failed to send Tab/Tab/Enter for {case_number}: ")
    except Exception as e:
        print(f"[WARN] Save note step failed for {case_number}: ")
    return note_saved

def case_search_and_open(driver, case_number):
    
    #Clicking Save Button for the previous case (will fail on first case, that's OK)
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckIn.Command') and contains(@id,'-button')]",
        timeout=1,
        retries=2,
    )
    
    #sleep timer
    time.sleep(5)

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    #Search Box - wait for it to be present
    safe_find(driver, By.ID, "GlobalSearchBox", timeout=3, retries=3)


    #Search for the case
    send_keys_safe(driver, By.ID, "GlobalSearchBox", case_number, timeout=3, retries=3)
    

    #click the first result's button to open the case
    click_safe(
        driver,
        By.XPATH,
        "//div[@id='id-globalSearchFlyout-1']/div/div/div/div/div/div[2]/div/button/span",
        timeout=7,
        retries=2,
    )

    # USFC discard dialog (if appears) - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=2,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    #Clicking Edit Button before starting
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckOut.Command') and contains(@id,'-button')]",
        timeout=1,
        retries=3,
    )

def case_search_and_open_no_edit(driver, case_number):
    """
    Clone of case_search_and_open WITHOUT pressing the Edit button.
    Used for checking case status before deciding to process.
    """
    #Clicking Save Button for the previous case (will fail on first case, that's OK)
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckIn.Command') and contains(@id,'-button')]",
        timeout=1,
        retries=2,
    )
    
    #sleep timer
    time.sleep(3)

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    #Search Box - wait for it to be present
    safe_find(driver, By.ID, "GlobalSearchBox", timeout=3, retries=3)

    #Search for the case
    send_keys_safe(driver, By.ID, "GlobalSearchBox", case_number, timeout=3, retries=3)
    
    #click the first result's button to open the case
    click_safe(
        driver,
        By.XPATH,
        "//div[@id='id-globalSearchFlyout-1']/div/div/div/div/div/div[2]/div/button/span",
        timeout=7,
        retries=2,
    )

    # USFC discard dialog (if appears) - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=2,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)
    
    # NOTE: No Edit button click - this is the key difference from case_search_and_open
    
def solution_provided_check_and_skip(driver, case_number, df, excel_path):

    solutionProv_case = True
    case_status_xpath = "//div[contains(@id,'headerControlsList')]/div[3]/div/div"
    case_status_el = safe_find(driver, By.XPATH, case_status_xpath, timeout=2, retries=6)
    case_status = case_status_el.text.strip()
    if not case_status.lower() == "solution provided":
        solutionProv_case = False
        
    else:
        solutionProv_case = True

    return solutionProv_case

def eticket_check_and_skip(driver, case_number, df, excel_path):
    # Verify the case has 'eticketing' in the header. The headerControlsList_<n> id suffix is variable,
    eticket_case = True
    case_channel_xpath = "//div[contains(@id,'headerControlsList')]/div[7]/div/div"
    case_channel_el = safe_find(driver, By.XPATH, case_channel_xpath, timeout=2, retries=6)
    if case_channel_el:
        case_channel = case_channel_el.text.strip()
        if not case_channel.lower() == "e-ticketing":
            eticket_case = False
        else:
            eticket_case = True
            # 1. Go to inprogress flyout menu
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id, 'MscrmControls.Containers.ProcessBreadCrumb-stageIndicatorContainer2d827664-559a-4937-beb1-52b3f836b967')]/div",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 2. Check the flyout menu of the aria-label = "Case Reason"
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id, 'header_process_casetypecode-header_process_casetypecode') and contains(@id, 'header_process_casetypecode.fieldControl-pcf-container-id')]/div/div/div/div/div/button",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 3. Choose "Service Call"
            click_safe(
                driver,
                By.XPATH,
                "//*[contains(@id, 'fluent-option') and text()='Service Call']",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 4. Check the flyout menu of aria-label "Contact Reason"
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id, 'header_process_lvidg_callreasoncode-header_process_lvidg_callreasoncode') and contains(@id, 'header_process_lvidg_callreasoncode.fieldControl-pcf-container-id')]/div/div/div/div/div/button",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 5. Choose "Hardware Repair"
            click_safe(
                driver,
                By.XPATH,
                "//*[contains(@id, 'fluent-option') and text()='Hardware Repair']",
                timeout=5,
                retries=3
            )
            # 6. check the flyout menu of aria-label "Tech Savvy"
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id,'header_process_lvidg_techsavvy-header_process_lvidg_techsavvy') and contains(@id, 'header_process_lvidg_techsavvy.fieldControl-pcf-container-id')]/div/div/div/div/div/button",
                timeout=5,
                retries=3
            )
            # 7. Choose "Undetermined"
            click_safe(
                driver,
                By.XPATH,
                "//*[contains(@id, 'fluent-option') and text()='Undetermined']",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 8. close the flyout menu
            click_safe(
                driver,
                By.XPATH,
                "//button[contains(@id,'MscrmControls.Containers.ProcessStageControl-stageContentClose') and @aria-label='Close']",
                timeout=5,
                retries=3
            )
    else:
        eticket_case = False
    return eticket_case

def process_new_case(driver, case_number, df, excel_path, idx, cache_file, EXCEL_SHEET_NAME, AGENT_NAME):
    """
    Explicit function for processing new cases:
    1. Check if solution provided
    2. If solution provided, click edit button and check for e-ticketing
    3. Returns True if case should be processed, False if it should be skipped
    """
    # First check if case is Solution Provided (before clicking edit)
    if not solution_provided_check_and_skip(driver, case_number, df, excel_path):
        # Case is NOT solution provided - skip it
        df.at[idx, "Assigned To"] = AGENT_NAME.split()[0]
        df.at[idx, "Status"] = 'Skipped'
        update_cache_file(cache_file, df, EXCEL_SHEET_NAME)
        return False
    
    # Case IS solution provided - now click the Edit button
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckOut.Command') and contains(@id,'-button')]",
        timeout=1,
        retries=3,
    )
    time.sleep(3)  # Wait for edit mode to activate
    
    # Now check for e-ticketing and perform dropdown selections if applicable
    eticket_check_and_skip(driver, case_number, df, excel_path)
    
    return True  # Case should be processed


def serial_extraction(driver, case_number, df):
    # Getting Serial Number (if available)
    serial_val = ''
    try:
        serial_xpath = "//*[contains(@id,'productserialnumber.fieldControl-text-input-component')]"
        el_serial = safe_find(driver, By.XPATH, serial_xpath, timeout=2, retries=2)
        if el_serial:
            try:
                serial_val = el_serial.get_attribute('value') or el_serial.text or ''
            except Exception:
                try:
                    serial_val = el_serial.text or ''
                except Exception:
                    serial_val = ''
    except Exception as e:
        print(f"[WARN] Could not read serial number for {case_number}: ")


    # normalize serial to string
    try:
        serial_val = str(serial_val).strip()
    except Exception:
        serial_val = ''

    return serial_val

def customer_name_extraction(driver, case_number):
    # Update SMS text with serial number and include customer name (CX_Name)
    CX_Name = ''
    try:
        # Try to read customer name from the associated entity anchor
        name_el = safe_find(driver, By.XPATH, "//*[contains(@id,'sec_tab_contact-associatedEntityRecordName')]", timeout=2, retries=2)
        if name_el:
            try:
                raw_name = name_el.get_attribute('aria-label') or name_el.get_attribute('title') or name_el.text or ''
            except Exception:
                raw_name = name_el.text or ''
            # Capitalize first letter of each word
            try:
                CX_Name = ' '.join([w.capitalize() for w in str(raw_name).split() if w])
            except Exception:
                CX_Name = str(raw_name)
    except Exception as e:
        print(f"[WARN] Could not read customer name for {case_number}: ")


    # Fallback name if none found
    if not CX_Name:
        CX_Name = 'Our Valued Customer'

    return CX_Name

def formatting_texts_sms(CX_Name, serial_val, case_number, df):

    # Format SMS text for this case
    try:
        sms_text = SMSText.format(CX_Name=CX_Name, case_number=case_number)
    except Exception:
        sms_text = SMSText.format(CX_Name='Our Valued Customer', case_number=case_number)

    return sms_text
    # Determine which email template to use based on WorkOrderType (if present in the Excel)

def formatting_texts_email(CX_Name, serial_val, case_number, df):
    ## Default fallback template name
    chosen_template = None

    try:
        # Try to read WorkOrderType for this case_number
        mask = df["case_number"].astype(str) == str(case_number)

        if mask.any() and "Work Order Type" in df.columns:
            wot_val = df.loc[mask, "Work Order Type"].iloc[0]
            if pd.isna(wot_val) or str(wot_val).strip() == "":
                wot = ""
            else:
                wot = str(wot_val).strip()
        else:
            wot = ""
    except Exception:
        wot = ""

    # Select template based on WorkOrderType
    if wot.lower() in ("onsite", "depot"):
        chosen_template = CaseEmailOnSite_Depot
    elif wot.lower() == "cru":
        chosen_template = CaseEmailCRU
    else:
        # Fallback to onsite/depot template if unknown or missing
        chosen_template = CaseEmailOnSite_Depot

    # Format Email text for this case  
    email_text = chosen_template.format(
        CX_Name=CX_Name or "Our Valued Customer",
        serial_val=serial_val,
        AGENT_NAME=AGENT_NAME
    )

    return email_text

def find_column_case_insensitive(df, name):
    for c in df.columns:
        if c.lower() == name.lower():
            return c
    return None

def todays_excel_path():
    now = datetime.now()
    dd = now.strftime('%d')
    mm = now.strftime('%m')
    file_name = f"Active Cases PA {mm}-{dd}.xlsx"
    path = os.path.join(EXCEL_BASE_PATH, file_name)
    return path

def excelCaseClosingCode(CaseClosingCode):

    match CaseClosingCode:
        case "Issue Resolved":
            return "Fixed"
        case "Issue Not Fixed":
            return "Issue Not Fixed"
        case "Customer Not Reached":
            return "Not Reached"
        case "Customer Claims that the Machine Not Yet Tested":
            return "Not yet Tested"
        case "DND":
            return "DND"
        case "Escalated":
            return "Escalation"
        case "Called - Company NO. Extension Found: Not Reached":
            return "Not Reached"
        case "Called: Not Reached // left Voicemail":
            return "Not Reached"
        case "Called - Answered: Issue Resolved":
            return "Fixed"
        case "Called - Answered: Issue Not Resolved":
            return "Issue Not Fixed"
        case "Need Third Action":
            return "In Progress"
        case "New":
            return "New"
        case "Skipped":
            return "Skipped"
        case "Send SMS":
            return "Send SMS"
        case "Send Email":
            return "Send Email"
        case "Send SMS and Email":
            return "Send SMS and Email"
        case "Call the Customer":
            return "Called the Customer"
        case _:
            return ""

def wait_until_time(hour: int, minute: int, check_interval: int = 30):
    """Block until the next occurrence of the specified hour:minute.

    If the requested time for today has already passed, waits until the same
    time on the next day. Prints periodic progress messages every
    `check_interval` seconds.
    """
    now = datetime.now()
    target = datetime(now.year, now.month, now.day, hour, minute)
    if now >= target:
        # schedule for next day
        target = target + timedelta(days=1)

    total_seconds = int((target - now).total_seconds())
    print(f"[INFO] Current time {now.strftime('%Y-%m-%d %H:%M:%S')}. Waiting until {target.strftime('%Y-%m-%d %H:%M:%S')} ({total_seconds}s)")

    # Sleep in intervals so we can provide periodic updates
    while True:
        now = datetime.now()
        remaining = (target - now).total_seconds()
        if remaining <= 0:
            break
        # print an update at a coarse interval
        if remaining > check_interval:
            mins = int(remaining // 60)
            secs = int(remaining % 60)
            print(f"[INFO] Time until start/close: {mins} minute(s) {secs} second(s)")
            time.sleep(check_interval)
        else:
            # final short sleep
            time.sleep(1)

    print("[INFO] Target start time reached — continuing.")

def wait_for_excel_file(path, check_interval=900):
    """
    Wait for Excel file to become available, checking every check_interval seconds (default 15 min).
    Logs progress and returns True once file exists.
    """
    if os.path.exists(path):
        print(f"[INFO] Excel file found: {path}")
        return True
    
    print(f"[INFO] Excel file not available yet: {path}")
    print(f"[INFO] Starting retry loop: checking every {check_interval} seconds (15 minutes)")
    
    attempt = 0
    while True:
        attempt += 1
        print(f"[INFO] Retry attempt {attempt}: waiting {check_interval} seconds before next check...")
        time.sleep(check_interval)
        
        if os.path.exists(path):
            print(f"[INFO] Excel file found on attempt {attempt}: {path}")
            return True
        else:
            print(f"[INFO] Excel file still not available on attempt {attempt}, will retry")

def keep_driver_alive(driver, check_interval=60):
    try:
        driver.refresh()
        print(f"[DEBUG] Driver refreshed to keep alive")
        return True
    except Exception as e:
        print(f"[DEBUG] Could not refresh driver (may not be at valid URL yet): ")
        return False

# OS Inhibit constants
ES_CONTINUOUS = 0x80000000
ES_SYSTEM_REQUIRED = 0x00000001
ES_DISPLAY_REQUIRED = 0x00000002
_inhibit_active = False

# ============================================================================
# SHOW WELCOME DIALOG FIRST - BEFORE EVERYTHING ELSE
# ============================================================================
print("[INFO] Displaying welcome dialog...")
proceed, skip_timer, support_agent = show_welcome_dialog()

# If support mode is enabled, use supported agent's sheet
if support_agent:
    EXCEL_SHEET_NAME = f"{support_agent}'s Cases"
    print(f"[INFO] SUPPORT MODE: Using sheet '{EXCEL_SHEET_NAME}' for supported agent")

if proceed == "MAIN_MENU":
    print("[INFO] returning to main menu...")
    try:
        if getattr(sys, 'frozen', False):
            # In-process: the main menu window is already running — just exit this
            # tool's execution path so control returns to the Qt event loop.
            print("[INFO] Frozen mode — main menu already running, exiting tool.")
        else:
            import subprocess as _sp
            current_dir = os.path.dirname(os.path.abspath(__file__))
            src_dir = os.path.dirname(current_dir)
            main_menu_path = os.path.join(src_dir, 'main.py')
            if os.path.exists(main_menu_path):
                _sp.Popen([sys.executable, main_menu_path])
            else:
                print(f"[ERROR] Main menu not found at {main_menu_path}")
    except Exception as e:
        print(f"[ERROR] Failed to launch main menu: {e}")
    sys.exit(0)

if not proceed:
    print("[INFO] User chose to update configuration. Exiting.")
    sys.exit(0)

print("[INFO] User confirmed to start process. Proceeding...")

excel_path = todays_excel_path()
# Enable Windows sleep/display inhibit while waiting
try:
    enable_windows_inhibit()
except Exception:
    pass

# Wait for start time unless user chose to skip timer
if skip_timer:
    print(f"[INFO] User chose to skip timer. Starting immediately...")
else:
    try:
        now = datetime.now()
        target = datetime(now.year, now.month, now.day, START_HOUR, START_MINUTE)
        if now >= target:
            print(f"[INFO] Current time is already past {START_TIME_STR} — running immediately.")
        else:
            wait_until_time(START_HOUR, START_MINUTE)
    except Exception as e:
        print(f"[WARN] wait_until_time check failed or was interrupted: ")

# Check if Excel file exists; if not, show popup with options
if not os.path.exists(excel_path):
    print(f"[INFO] Excel file does not exist: {excel_path}")
    print(f"[INFO] Showing file search popup...")
    
    # Show file search popup with countdown and options
    action, new_path = show_file_search_popup(excel_path, retry_interval_seconds=10)
    
    if action == "ABORT":
        print("[INFO] User aborted file search. Returning to welcome dialog...")
        # Return to welcome dialog
        proceed, skip_timer, support_agent_new = show_welcome_dialog()
        if not proceed or proceed == "MAIN_MENU":
            print("[INFO] Exiting application.")
            sys.exit(0)
        # Update support mode if changed
        if support_agent_new:
            EXCEL_SHEET_NAME = f"{support_agent_new}'s Cases"
            print(f"[INFO] SUPPORT MODE: Using sheet '{EXCEL_SHEET_NAME}' for supported agent")
        # Restart the file check with new excel_path
        excel_path = todays_excel_path()
    elif action == "MANUAL":
        print(f"[INFO] User selected manual file (one-time use): {new_path}")
        excel_path = new_path  # Use this file for this run only
    elif action == "FOUND":
        print(f"[INFO] Excel file found at configured path: {excel_path}")
    
    # Set Chrome to use ART Profile
    chrome_options = Chrome_ART_Profile()
    
    # Start driver
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    driver.maximize_window()

    try:
        perform_dialer_login(driver)
    except Exception as e:
        print(f"[WARN] Dialer login failed (optional): ")
    
    print(f"[INFO] Driver started. Proceeding with Excel file...")
else:
    # Excel file exists; proceed normally
    print(f"[INFO] Excel file exists: {excel_path}")
    print(f"[INFO] Starting driver and proceeding with case processing...")
    
    # Set Chrome to use ART Profile
    chrome_options = Chrome_ART_Profile()
    
    # Start driver
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    driver.maximize_window()
    try:
        perform_dialer_login(driver)
    except Exception as e:
        print(f"[WARN] Dialer login failed (optional): ")
    


# Read case numbers from Excel sheet
print("[INFO] Loading main Excel sheet...")
df_main = pd.read_excel(excel_path, sheet_name=EXCEL_SHEET_NAME)
print(f"[INFO] Main sheet loaded: {len(df_main)} total rows")

status_col = find_column_case_insensitive(df_main, 'Status')
worktype_col = find_column_case_insensitive(df_main, 'Work Order Type')
case_col = find_column_case_insensitive(df_main, 'Case Number')
assigned_col = find_column_case_insensitive(df_main, 'Assigned To')

# Filter for working cases (new, in_progress, skipped) by agent
print(f"[INFO] Filtering cases for agent: {AGENT_NAME.split()[0]}")
df_filtered = df_main[
    (df_main[status_col].astype(str).str.strip().str.lower().isin(['new', 'in_progress', 'skipped'])) &
    (df_main[assigned_col].astype(str).str.strip() == AGENT_NAME.split()[0])
].copy()

print(f"[INFO] Found {len(df_filtered)} cases to process (new + in_progress + skipped)")

# ============================================================================
# WORKING SHEET LOGIC - Resume from crash if today's cache exists
# ============================================================================
agent_name_clean = AGENT_NAME.split()[0].replace(' ', '_')
today_date_str = datetime.now().strftime('%Y-%m-%d')  # e.g. "2026-01-02"
cache_file = os.path.join(CACHE_DIRECTORY, f'working_cases_{agent_name_clean}_{today_date_str}.xlsx')

# Ensure cache directory exists
os.makedirs(CACHE_DIRECTORY, exist_ok=True)

# Check if today's working sheet already exists (for crash recovery)
if os.path.exists(cache_file):
    print(f"[INFO] *** RESUMING FROM EXISTING WORKING SHEET ***")
    print(f"[INFO] Found today's working cache: {cache_file}")
    try:
        df = pd.read_excel(cache_file, sheet_name=EXCEL_SHEET_NAME)
        print(f"[INFO] Loaded {len(df)} cases from cached working sheet")
        print(f"[INFO] This will continue from where you left off!")
        
        # Count remaining cases
        remaining_new = len(df[df[status_col].astype(str).str.strip().str.lower() == 'new'])
        remaining_in_progress = len(df[df[status_col].astype(str).str.strip().str.lower().isin(['in_progress', 'skipped'])])
        completed = len(df[df[status_col].astype(str).str.strip().str.lower().isin(['closed', 'in progress today'])])
        print(f"[INFO] Status: {completed} completed, {remaining_new} new, {remaining_in_progress} in_progress/skipped remaining")
        
    except Exception as e:
        print(f"[WARN] Failed to read existing cache, creating fresh: {e}")
        df = df_filtered.copy()
        # Save with both handler sheet AND Companies sheet
        with pd.ExcelWriter(cache_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=EXCEL_SHEET_NAME, index=False)
            # Add Companies sheet if available from main file
            try:
                df_companies = pd.read_excel(excel_path, sheet_name='Companies')
                # Filter for handler's cases only
                comp_assigned_col = find_column_case_insensitive(df_companies, 'Assigned To')
                if comp_assigned_col:
                    df_companies_handler = df_companies[df_companies[comp_assigned_col].astype(str).str.strip() == AGENT_NAME.split()[0]].copy()
                    df_companies_handler.to_excel(writer, sheet_name='Companies', index=False)
                    print(f"[INFO] Added {len(df_companies_handler)} handler cases to Companies sheet in cache")
            except Exception as ce:
                print(f"[INFO] No Companies sheet to add: {ce}")
        print(f"[INFO] Created fresh working cache with {len(df)} cases")
else:
    # No existing cache - create new from filtered cases
    print(f"[INFO] No existing cache for today - creating new working sheet")
    print(f"[INFO] Creating working cache file: {cache_file}")
    try:
        df = df_filtered.copy()
        # Save with both handler sheet AND Companies sheet
        with pd.ExcelWriter(cache_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=EXCEL_SHEET_NAME, index=False)
            # Add Companies sheet if available from main file
            try:
                df_companies = pd.read_excel(excel_path, sheet_name='Companies')
                # Filter for handler's cases only
                comp_assigned_col = find_column_case_insensitive(df_companies, 'Assigned To')
                if comp_assigned_col:
                    df_companies_handler = df_companies[df_companies[comp_assigned_col].astype(str).str.strip() == AGENT_NAME.split()[0]].copy()
                    df_companies_handler.to_excel(writer, sheet_name='Companies', index=False)
                    print(f"[INFO] Added {len(df_companies_handler)} handler cases to Companies sheet in cache")
            except Exception as ce:
                print(f"[INFO] No Companies sheet to add: {ce}")
        print(f"[INFO] Working cache created with {len(df)} cases")
    except Exception as e:
        print(f"[ERROR] Failed to create cache file: {e}")
        traceback.print_exc()
        df = df_filtered.copy()

# Store original indices for merge back later
original_indices = df_main[
    (df_main[status_col].astype(str).str.strip().str.lower().isin(['new', 'in_progress'])) &
    (df_main[assigned_col].astype(str).str.strip() == AGENT_NAME.split()[0])
].index.tolist()

# Optional: Perform dialer login if needed (comment out if not using dialer)


time.sleep(10)  # wait for manual login if needed
today_str = datetime.now().strftime("%b %d, %Y")

# ============================================================================
# COMPANIES PROCESS SECTION - Process company cases before case reviewer
# ============================================================================
try:
    # Check if Companies sheet exists in cache
    df_companies_check = pd.read_excel(cache_file, sheet_name='Companies')
    companies_count = len(df_companies_check)
    
    if companies_count > 0:
        # Count distinct emails
        email_col = find_column_case_insensitive(df_companies_check, 'Email')
        distinct_emails = df_companies_check[email_col].nunique() if email_col else 0
        
        print(f"[INFO] Found {companies_count} company cases ({distinct_emails} distinct companies)")
        
        # Show Companies Process prompt
        app_companies = QApplication.instance()
        if app_companies is None:
            app_companies = QApplication(sys.argv)
        
        companies_dialog = CompaniesProcessDialog(companies_count, distinct_emails)
        companies_dialog.exec_()
        
        if companies_dialog.result == "YES":
            print("[INFO] User chose to process company cases")
            run_companies_process(driver, cache_file, AGENT_NAME, 'Companies')
        else:
            print("[INFO] User skipped to Case Reviewer")
    else:
        print("[INFO] No company cases found in cache")
except Exception as ce:
    print(f"[INFO] Companies sheet not available or empty: {ce}")

# Track break state globally so it persists between dialog instances
is_on_break = False

# Filter out rows where Status == 'closed' (case-insensitive)
try:
    filtered_df = df[~df[status_col].astype(str).str.strip().str.lower().eq('closed')]
except Exception:
    filtered_df = df.copy()

def get_case_counts ():
    new_case_count = 0
    in_progress_case_count = 0
    for idx, row in filtered_df.iterrows():
        status = str(row.get(status_col, '')).strip().lower()
        case_number = row.get(case_col)
        if pd.isna(case_number) or not str(case_number).strip():
            continue
        # Convert to int first to remove .0, then to string
        try:
            case_number = str(int(float(case_number))).strip()
        except (ValueError, TypeError):
            case_number = str(case_number).strip()
        # Count cases by status
        if status not in ('new', 'in_progress', 'skipped'):
            continue

        if status == 'new':
            new_case_count += 1
        elif status in ('in_progress', 'skipped'):
            in_progress_case_count += 1
    return in_progress_case_count, new_case_count

in_progress_case_count,new_case_count = get_case_counts()

# Store total counts for progress tracking
total_in_progress_count = in_progress_case_count
total_new_count = new_case_count

print(f"[DEBUG] Total in_progress cases: {total_in_progress_count}, Total new cases: {total_new_count}")

# Get refresh interval from config
case_counter = 0
in_progress_cases_completed = 0  # Track completed in_progress cases for progress bar

# Build list of indices to process (pointer-based loop allows going backward)
to_process_indices = list(filtered_df.index)
pointer = 0

# ============================================================================
# CREATE NEW CASES PROGRESS POPUP
# ============================================================================
new_cases_popup = None
new_cases_done = 0  # Track how many new cases have been completed
case_reviewer_popup = None  # Separate popup for case reviewer mode

if total_new_count > 0:
    print(f"[INFO] Creating New Cases Progress Popup for {total_new_count} new cases...")
    # Ensure QApplication exists
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    new_cases_popup = NewCasesProgressPopup(total_cases=total_new_count)
    new_cases_popup.show()
    new_cases_popup.expand()  # Start expanded for visibility
    
    # Process events to ensure popup is displayed
    app.processEvents()
    print("[INFO] New Cases Progress Popup created and shown")

while pointer < len(to_process_indices):
    try:
        idx = to_process_indices[pointer]
        row = filtered_df.loc[idx]
        
        status = str(row.get(status_col, '')).strip().lower()
        case_number = row.get(case_col)
        if pd.isna(case_number) or not str(case_number).strip():
            pointer += 1
            continue
        # Convert to int first to remove .0, then to string
        try:
            case_number = str(int(float(case_number))).strip()
        except (ValueError, TypeError):
            case_number = str(case_number).strip()
        # Search and Open Case
        if status not in ('new', 'in_progress', 'skipped'):
            pointer += 1
            continue
        
        case_search_and_open(driver, case_number)
        
        # Extract Serial Number
        serial_val = serial_extraction(driver, case_number, df)         
        # Extract Customer Name
        CX_Name = customer_name_extraction(driver, case_number)  

        # Format SMS Text
        sms_text = formatting_texts_sms(CX_Name, serial_val, case_number, df)

        # Format Email Text
        email_text = formatting_texts_email(CX_Name, serial_val, case_number, df)

        if status == 'new':
            # ============ POPUP EVENT PROCESSING FOR NEW CASES ============
            if new_cases_popup:
                app = QApplication.instance()
                if app:
                    app.processEvents()  # Process button clicks
                
                # Check for Abort Now
                if new_cases_popup.abort_now:
                    print("[INFO] Abort Now clicked - stopping immediately...")
                    new_cases_popup.close()
                    break
                
                # Check for Quit After Current (will be checked after case completes)
                
                # Check for Take a Break - set dialer to break status
                if new_cases_popup.take_a_break and not getattr(new_cases_popup, '_break_status_set', False):
                    print("[INFO] Take a Break - setting dialer to Not Ready (Break)...")
                    set_dialer_break_status(driver)
                    switch_to_crm_window(driver)  # Switch back to CRM
                    new_cases_popup._break_status_set = True
                
                # Check for paused state - wait until unpaused
                while new_cases_popup.is_paused:
                    if app:
                        app.processEvents()
                    time.sleep(0.5)
                    
                    # Check if Back on Duty was clicked
                    if not new_cases_popup.is_paused:
                        if getattr(new_cases_popup, '_break_status_set', False):
                            print("[INFO] Back on Duty - setting dialer to Not Ready (Chat)...")
                            set_dialer_chat_status(driver)
                            switch_to_crm_window(driver)  # Switch back to CRM
                            new_cases_popup._break_status_set = False
                        break
                    
                    # Allow abort even while paused
                    if new_cases_popup.abort_now:
                        print("[INFO] Abort Now clicked while paused - stopping...")
                        new_cases_popup.close()
                        break
                
                # If aborted while paused, break outer loop too
                if new_cases_popup.abort_now:
                    break
            # ============================================================
            
            # Process new case: check solution provided, click edit, check e-ticketing
            if not process_new_case(driver, case_number, df, excel_path, idx, cache_file, EXCEL_SHEET_NAME, AGENT_NAME):
                pointer += 1
                continue

            # Check if each column is empty (NaN or blank string) and call helpers with correct args
            if pd.isna(row["Action 1"]) or str(row["Action 1"]).strip() == "":
                sms_sent = send_SMS(driver, sms_text)
            else:
                sms_sent = True  # Already sent

            if pd.isna(row["Action 2"]) or str(row["Action 2"]).strip() == "":
                email_sent = send_Email(driver, email_text)
            else:
                email_sent = True  # Already sent

            note_saved = add_Case_Note(driver, CaseNote)
            

            if sms_sent:
                df.at[idx, "Action 1"] = 'Sent SMS'
            if email_sent:
                df.at[idx, "Action 2"] = 'Sent Email'
            df.at[idx, "Action 3"] = ''
            df.at[idx, "Final Action"] = 'Sent Email'
            df.at[idx, "Assigned To"] = AGENT_NAME.split()[0]
            if sms_sent and email_sent and note_saved:
                df.at[idx, "Status"] = 'In Progress Today'      
                
            update_cache_file(cache_file, df, EXCEL_SHEET_NAME)

            new_case_count -= 1
            print(f"[INFO] New cases remaining: {new_case_count}, In-Progress cases remaining: {in_progress_case_count}")           
            case_counter += 1
            
            # ============ UPDATE POPUP PROGRESS ============
            if new_cases_popup:
                new_cases_done += 1
                new_cases_popup.update_progress(new_cases_done, total_new_count)
                app = QApplication.instance()
                if app:
                    app.processEvents()
                
                # Check for Quit After Current
                if new_cases_popup.quit_after_current:
                    print("[INFO] Quit After Current clicked - exiting after this case...")
                    new_cases_popup.close()
                    break
                
                # Close popup when all new cases are done
                if new_cases_done >= total_new_count:
                    print("[INFO] All new cases completed - closing popup")
                    new_cases_popup.close()
                    new_cases_popup = None  # Clear reference
            # ==============================================
            
            if case_counter % REFRESH_INTERVAL == 0:
                keep_driver_alive(driver)
                time.sleep(5)  # wait for refresh to complete

        elif status == 'in_progress' or status == 'skipped':
            # Note: Error log and Take a Break button are now integrated into CaseReviewerDialog
            print(f"[INFO] Processing in_progress case: {case_number}")
            print(f"[DEBUG] Cases completed before dialog: {in_progress_cases_completed}")
            try:
                CaseClosingCode, add_note = get_case_closing_code(case_number, in_progress_cases_completed, total_in_progress_count, case_status=status)
                
                # Check if user clicked Close & Exit
                if CaseClosingCode == "CLOSE_APPLICATION":
                    print(f"[INFO] User clicked Close & Exit - shutting down...")
                    break  # Exit the main loop
                
                # Check if user clicked Take a Break
                if CaseClosingCode == "TAKE_A_BREAK":
                    is_on_break = True  # Track state globally
                    print("[INFO] Take a Break - setting dialer to Not Ready (Break)...")
                    set_dialer_break_status(driver)
                    switch_to_crm_window(driver)
                    time.sleep(2)  # Brief pause before reopening
                    continue  # Re-show dialog for same case (button will show "Back on Duty")
                
                # Check if user clicked Back on Duty
                if CaseClosingCode == "BACK_ON_DUTY":
                    is_on_break = False  # Track state globally
                    print("[INFO] Back on Duty - setting dialer to Not Ready (Chat)...")
                    set_dialer_chat_status(driver)
                    switch_to_crm_window(driver)
                    time.sleep(2)  # Brief pause before reopening
                    continue  # Re-show dialog for same case (button will show "Take a Break")
                
                # Check if user clicked Previous Case
                if CaseClosingCode == "PREVIOUS_CASE":
                    print(f"[INFO] User clicked Previous Case - going back...")
                    if pointer > 0:
                        pointer -= 1  # Go back to previous case
                        # Decrement the completed count if we're going back
                        if in_progress_cases_completed > 0:
                            in_progress_cases_completed -= 1
                            print(f"[DEBUG] Decremented in_progress_cases_completed to: {in_progress_cases_completed}")
                    else:
                        print(f"[WARN] Already at first case, cannot go back further")
                    continue  # Don't increment pointer, just re-loop
                
                # Handle Skipped - set Status to Skipped and continue
                if CaseClosingCode == "Skipped":
                    print(f"[DEBUG] Skipping case {case_number} - Status set to Skipped")
                    df.at[idx, "Status"] = 'Skipped'
                    df.at[idx, "Assigned To"] = AGENT_NAME.split()[0]
                    update_cache_file(cache_file, df, EXCEL_SHEET_NAME)
                    in_progress_cases_completed += 1
                    pointer += 1
                    continue
                
                # Handle CUSTOM|custom_text|final_action|status format
                if CaseClosingCode.startswith("CUSTOM|"):
                    parts = CaseClosingCode.split("|")
                    custom_text = parts[1] if len(parts) > 1 else ""
                    custom_final_action = parts[2] if len(parts) > 2 else "Reviewed"
                    custom_status = parts[3] if len(parts) > 3 else "Skipped"
                    
                    print(f"[DEBUG] Custom code: text='{custom_text}', final_action='{custom_final_action}', status='{custom_status}'")
                    
                    # Add note with custom text if provided
                    if custom_text:
                        CaseNote = f"Date: {today_str}\nQueue: ART Project - Follow up \nAgent: {AGENT_NAME} \nAction: {custom_text}\n \n ------------------------"
                        try:
                            add_Case_Note(driver, CaseNote=CaseNote)
                        except Exception as e:
                            print(f"[WARN] add_Case_Note failed for {case_number}: ")
                    
                    # Set Excel values from radio button selections
                    df.at[idx, "Final Action"] = custom_final_action
                    df.at[idx, "Status"] = custom_status
                    df.at[idx, "Assigned To"] = AGENT_NAME.split()[0]
                    update_cache_file(cache_file, df, EXCEL_SHEET_NAME)
                    in_progress_cases_completed += 1
                    pointer += 1
                    continue
                
                in_progress_cases_completed += 1  # Increment after getting the code
                print(f"[DEBUG] Cases completed after increment: {in_progress_cases_completed}")
                print(f"[INFO] Received closing code: {CaseClosingCode}, add_note: {add_note}")
            except Exception as dialog_error:
                print(f"[ERROR] Dialog error for case {case_number}: {type(dialog_error).__name__}: {str(dialog_error)}")
                traceback.print_exc()
                pointer += 1
                continue  # Skip this case and move to next
            
            CaseNote = f"Date: {today_str}\nQueue: ART Project - Follow up \nAgent: {AGENT_NAME} \nAction: Case is Reviewed with closing code {CaseClosingCode}\n \n ------------------------"

            # If user requested adding a Case Note via the popup, create a small note and add it
            if add_note:
                try:
                    add_Case_Note(driver, CaseNote=CaseNote)
                except Exception as e:
                    print(f"[WARN] add_Case_Note failed for {case_number}: ")
            

            # If DND selected, update contact preferences
            if CaseClosingCode == "DND":
                DND_CX(driver, case_number)


            if CaseClosingCode == "Call the Customer":
                perform_call_flow(driver)
                # Set Action 3 immediately after call (before closing code dialog)
                df.at[idx, "Action 3"] = 'Called the Customer'
                # after call, re-open closing code dialog to capture outcome
                CaseClosingCode, add_note = get_call_closing_code()
                # recompute case note to reflect any new CaseClosingCode outcome
                try:
                    case_note_text = f"Date: {today_str}\nQueue: ART Project - Follow up\nAgent: {AGENT_NAME}\nAction: Called the Customer //  {CaseClosingCode}"
                except Exception:
                    case_note_text = f"Date: {today_str}\nQueue: ART Project - Follow up\nAgent: {AGENT_NAME}\nAction: Called the Customer //  {CaseClosingCode}"
                
                try:
                    add_Case_Note(driver, CaseNote=case_note_text)
                except Exception as e:
                    print(f"[WARN] add_Case_Note failed for {case_number} after call: ")

            # Prepare sms/email texts if any of those actions are requested
            if CaseClosingCode in ("Send SMS", "Send Email", "Send SMS and Email"):
                try:
                    serial_val = serial_extraction(driver, case_number, df)
                except Exception:
                    serial_val = ''
                try:
                    CX_Name = customer_name_extraction(driver, case_number)
                except Exception:
                    CX_Name = 'Our Valued Customer'

                try:
                    sms_text = formatting_texts_sms(CX_Name, serial_val, case_number, df)
                except Exception:
                    sms_text = SMSText.format(CX_Name=CX_Name, case_number=case_number)

                try:
                    email_text = formatting_texts_email(CX_Name, serial_val, case_number, df)
                except Exception:
                    email_text = ""

            if CaseClosingCode == "Send SMS":
                try:
                    send_SMS(driver, sms_text)
                    add_Case_Note(driver, CaseNote=CaseNote)
                except Exception as e:
                    print(f"[WARN] send_SMS failed for {case_number}: ")
                

            if CaseClosingCode == "Send Email":
                try:
                    send_Email(driver, email_text)
                    add_Case_Note(driver, CaseNote=CaseNote)  
                except Exception as e:
                    print(f"[WARN] send_Email failed for {case_number}: ")

            if CaseClosingCode == "Send SMS and Email":
                try:
                    send_SMS(driver, sms_text)
                except Exception as e:
                    print(f"[WARN] send_SMS (part of combined) failed for {case_number}: ")
                try:
                    send_Email(driver, email_text)
                except Exception as e:
                    print(f"[WARN] send_Email (part of combined) failed for {case_number}: ")
                try:
                    add_Case_Note(driver, CaseNote=CaseNote)
                except Exception as e:
                    print(f"[WARN] add_Case_Note failed after combined send for {case_number}: ")

            if CaseClosingCode == "Need Third Action":
                print(f"[DEBUG] Need Third Action case - in_progress_cases_completed is now: {in_progress_cases_completed}")
                pointer += 1
                continue
            
            if 'called' in CaseClosingCode.lower():
                df.at[idx, "Action 3"] = 'Called the Customer'
            

            df.at[idx, "Final Action"] = excelCaseClosingCode(CaseClosingCode)
            df.at[idx, "Assigned To"] = AGENT_NAME.split()[0]
            df.at[idx, "Status"] = 'Closed'
            update_cache_file(cache_file, df, EXCEL_SHEET_NAME)
            
            print(f"[INFO] In-Progress cases remaining: {in_progress_case_count}")
            case_counter += 1
            
            if case_counter % REFRESH_INTERVAL == 0:
                keep_driver_alive(driver)
                time.sleep(5)  # wait for refresh to complete

        # Move to next case
        pointer += 1

    except Exception as e:
        print(f"[ERROR] Exception in loop for case {case_number if 'case_number' in locals() else 'UNKNOWN'}: {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        pointer += 1
        continue       


print("All done.")

# ============================================================================
# NOTE: MERGE CACHE BACK TO MAIN EXCEL - WILL BE IMPLEMENTED LATER
# ============================================================================
# The working cases are in: cache_file
# The original main sheet is in: df_main with original_indices mapping
# To merge back later, we will:
# 1. Load updated cache_file
# 2. Update corresponding rows in df_main using original_indices
# 3. Write df_main back to excel_path (main file)
# This step is deferred to maintain safety and allow verification
# ============================================================================

print(f"[INFO] Working cases cache saved at: {cache_file}")
print(f"[INFO] Original data preserved in: {excel_path}")
print(f"[NOTE] Cache contains {len(df)} processed cases ready for merge")

# Close the persistent case reviewer dialog and cleanup
try:
    close_case_reviewer()
except Exception as e:
    print(f"[ERROR] Error closing case reviewer: {e}")
    traceback.print_exc()

# Check if user closed the app via the Close button
closed_by_user = 'CaseClosingCode' in locals() and CaseClosingCode == "CLOSE_APPLICATION"

if closed_by_user:
    print("[INFO] Application was closed by user via Close button")
    # Close Chrome driver immediately when user explicitly closes
    try:
        if 'driver' in locals() and driver is not None:
            print("[INFO] Closing Chrome driver...")
            driver.quit()
            print("[INFO] Chrome driver closed successfully.")
    except Exception as e:
        print(f"[WARN] Error closing Chrome driver: {e}")
        traceback.print_exc()
    
    # Disable Windows sleep inhibit
    try:
        disable_windows_inhibit()
    except Exception as e:
        print(f"[WARN] Error disabling Windows inhibit: {e}")
    
    # Show welcome dialog to ask if they want to run again
    try:
        from PyQt5.QtWidgets import QMessageBox
        app = QApplication.instance()
        if app is None:
            app = QApplication(sys.argv)
        
        reply = QMessageBox.question(
            None, "Exit Verification", 
            "The application was closed. Would you like to start it again?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.No:
            print("[INFO] User chose not to run again. Exiting...")
            sys.exit(0)
        else:
            print("[INFO] User wants to run again - restarting...")
            if not getattr(sys, 'frozen', False):
                # Dev mode only — in a frozen .exe __file__ is not a runnable .py
                import subprocess as _sp
                _sp.Popen([sys.executable, __file__])
            sys.exit(0)
    except Exception as e:
        print(f"[WARN] Could not show restart dialog: {e}")
        # Just exit if dialog fails
        sys.exit(0)
else:
    # Normal completion - all cases done, keep Chrome open until end time
    print("[INFO] All cases processed. Chrome will remain open until scheduled end time.")
    print(f"[INFO] Scheduled end time: {END_HOUR:02d}:{END_MINUTE:02d}")
    
    try:
        # Keep driver alive while waiting for end time
        while True:
            now = datetime.now()
            target = datetime(now.year, now.month, now.day, END_HOUR, END_MINUTE)
            if now >= target:
                print("[INFO] End time reached. Closing down...")
                break
            
            remaining = (target - now).total_seconds()
            mins = int(remaining // 60)
            secs = int(remaining % 60)
            print(f"[INFO] Waiting for end time. Remaining: {mins} minute(s) {secs} second(s)")
            
            # Keep driver alive by refreshing periodically
            try:
                keep_driver_alive(driver)
            except Exception:
                pass
            
            # Sleep for 60 seconds between checks
            time.sleep(60)
            
    except Exception as e:
        print(f"[WARN] Wait until end time failed or was interrupted: {e}")
    
    # NOW close Chrome driver after end time
    try:
        if 'driver' in locals() and driver is not None:
            print("[INFO] End time reached. Closing Chrome driver...")
            driver.quit()
            print("[INFO] Chrome driver closed successfully.")
    except Exception as e:
        print(f"[WARN] Error closing Chrome driver: {e}")
        traceback.print_exc()
    
    # Disable Windows sleep inhibit
    try:
        disable_windows_inhibit()
    except Exception as e:
        print(f"[WARN] Error disabling Windows inhibit: {e}")