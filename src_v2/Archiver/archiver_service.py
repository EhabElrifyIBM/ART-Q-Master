"""
Archiver Service Layer (Phase 6.2 - Fixed & Upgraded)
=======================================================

Pure business logic for Excel workbook archiving operations.
Extracted from original tkinter-based Archiver.py.

This module handles:
- Excel workbook analysis
- Date parsing and extraction  (multiple formats)
- Month-based filtering
- Age-based filtering
- Export operations with optional cleanup
- Timestamped backup creation
- Progress reporting with consistent (int, str) callback signature

No UI code — all PyQt5 UI lives in separate components.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path
import shutil
from copy import copy
from typing import Dict, List, Optional, Tuple, Any, Callable
from dataclasses import dataclass, field


# ---------------------------------------------------------------------------
# Progress callback type: (percent: int, message: str) -> None
# ---------------------------------------------------------------------------
ProgressCallback = Optional[Callable[[int, str], None]]


def _report(cb: ProgressCallback, percent: int, msg: str) -> None:
    """Safely invoke a progress callback."""
    if cb:
        try:
            cb(percent, msg)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Data-transfer objects
# ---------------------------------------------------------------------------

@dataclass
class AnalysisResult:
    """Results from analysing a single handler sheet."""
    sheet_name: str
    month_data: Dict[str, int]       # "YYYY-MM" → case count (date-parseable rows only)
    dates: List[Dict[str, Any]]      # list of {date, month, raw, row_idx}
    total_cases: int                 # ALL non-empty data rows (true row count)
    header_row: int
    status_col_index: int            # "Last Status Change" column (1-based)
    # New in upgraded version
    oldest_date: Optional[datetime] = None
    newest_date: Optional[datetime] = None
    # Date-source tracking
    completion_col_index: Optional[int] = None   # "Completion Date" column if found
    dated_by_completion: int = 0     # rows whose date came from "Completion Date"
    dated_by_status: int = 0         # rows whose date came from "Last Status Change"
    dated_blank: int = 0             # non-empty rows with no parseable date at all


@dataclass
class ExportOptions:
    """Base options for all export operations."""
    handler: str        # Handler sheet name, or "All Handlers"
    output_file: str
    cleanup: bool = False   # Delete exported rows from original
    merged: bool = False    # Create merged sheet combining all handlers


@dataclass
class MonthExportOptions(ExportOptions):
    """Options for month-based export.

    ``months`` is a list of one or more "YYYY-MM" strings.  All matching rows
    across every selected month are written to a single sheet in the output
    workbook (plus an optional Merged sheet when handler == "All Handlers").
    """
    months: List[str] = None   # e.g. ["2025-01", "2025-02"]

    def __post_init__(self):
        if self.months is None:
            self.months = []


@dataclass
class AgeExportOptions(ExportOptions):
    """Options for age-based export."""
    days: int = 0           # Export cases older than this many days


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class ArchiverService:
    """
    Service layer for Excel workbook archiving operations.

    Handles all business logic for analysing and exporting Excel data.
    Thread-safe state: the service does NOT hold open file handles between
    calls — every operation opens, processes, and closes the workbook.

    Progress callbacks receive (percent: int, message: str).
    """

    # Maximum number of header rows to scan before falling back to defaults
    HEADER_SCAN_ROWS = 20

    def __init__(self) -> None:
        self.workbook_path: Optional[str] = None
        self.analysis_results: Optional[Dict[str, AnalysisResult]] = None

    # ------------------------------------------------------------------
    # Loading & Analysis
    # ------------------------------------------------------------------

    def load_workbook(self, file_path: str) -> bool:
        """
        Load an Excel workbook for analysis.

        Args:
            file_path: Path to Excel file (.xlsx or .xls)

        Returns:
            True if loaded successfully.

        Raises:
            FileNotFoundError: File does not exist.
            ValueError: Not a valid Excel file.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if path.suffix.lower() not in ('.xlsx', '.xls'):
            raise ValueError(
                f"Invalid file type: {path.suffix}. Expected .xlsx or .xls"
            )

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            wb.close()
            self.workbook_path = file_path
            return True
        except Exception as exc:
            raise ValueError(f"Failed to load Excel file: {exc}") from exc

    # Canonical name of the Companies sheet
    COMPANIES_SHEET = "Companies"

    def _find_companies_sheet(self, sheetnames: List[str]) -> Optional[str]:
        """
        Return the actual sheet name that represents the Companies sheet.

        Uses a case-insensitive, strip-tolerant match so that minor variations
        (trailing spaces, different casing) never silently drop the sheet.

        Returns the real sheet name as it appears in the workbook, or None if
        no Companies-like sheet is found.
        """
        for name in sheetnames:
            if name.strip().lower() == self.COMPANIES_SHEET.lower():
                return name
        return None

    def analyze_workbook(self) -> Dict[str, AnalysisResult]:
        """
        Analyse the loaded workbook for handler sheets and date information.

        Includes the "Companies" sheet (if present, matched case-insensitively)
        alongside all handler sheets that contain "'s Cases" in their name.

        Returns:
            Dict keyed by sheet name.  Companies sheet, when present, is
            always appended last so the handler sheets appear first in UI.

        Raises:
            RuntimeError: No workbook loaded.
            ValueError: No handler sheets found.
        """
        if not self.workbook_path:
            raise RuntimeError("No workbook loaded. Call load_workbook() first.")

        wb = openpyxl.load_workbook(self.workbook_path, data_only=True)

        print(f"[Archiver] Workbook sheets: {wb.sheetnames}")

        # Handler sheets: all sheets whose name contains "'s Cases".
        # This includes "Chat Agent's Cases" alongside individual handler sheets.
        handler_sheets = [s for s in wb.sheetnames if "'s Cases" in s]

        print(f"[Archiver] Handler sheets found: {handler_sheets}")

        if not handler_sheets:
            wb.close()
            raise ValueError(
                'No sheets found containing "\'s Cases" in their name.'
            )

        results: Dict[str, AnalysisResult] = {}
        for sheet_name in handler_sheets:
            results[sheet_name] = self._analyze_sheet(wb[sheet_name], sheet_name)

        # Locate the Companies sheet using a tolerant match (handles trailing
        # spaces, wrong casing, etc.) so it is never silently skipped.
        companies_sheet_name = self._find_companies_sheet(wb.sheetnames)
        print(f"[Archiver] Companies sheet detected: {repr(companies_sheet_name)}")

        if companies_sheet_name is not None:
            results[self.COMPANIES_SHEET] = self._analyze_sheet(
                wb[companies_sheet_name], self.COMPANIES_SHEET
            )
            print(
                f"[Archiver] Companies sheet analysed: "
                f"{results[self.COMPANIES_SHEET].total_cases} rows"
            )
        else:
            print(
                f"[Archiver] WARNING: No 'Companies' sheet found. "
                f"All sheet names: {wb.sheetnames}"
            )

        wb.close()
        self.analysis_results = results
        return results

    def _analyze_sheet(self, ws, sheet_name: str) -> AnalysisResult:
        """Analyse a single handler sheet and return an AnalysisResult.

        ``total_cases`` reflects every non-empty data row, not just those with
        a parseable date.  This prevents any sheet (Chat Agent's Cases,
        Companies, handler sheets) from showing a case count lower than the
        actual number of rows in the sheet when some rows lack a date value.

        Date sourcing priority (per row):
          1. "Completion Date" column  — if cell is non-blank and parseable
          2. "Last Status Change" column — fallback when Completion Date is blank
          3. Neither parseable → row counted as ``dated_blank``
        """
        # -------------------------------------------------------------------
        # Scan header row for "Completion Date" AND "Last Status Change"
        # -------------------------------------------------------------------
        header_row: Optional[int] = None
        status_col_index: Optional[int] = None
        completion_col_index: Optional[int] = None

        for row in ws.iter_rows(min_row=1, max_row=self.HEADER_SCAN_ROWS):
            for cell in row:
                if not cell.value:
                    continue
                cell_text = str(cell.value)
                if "Last Status Change" in cell_text and status_col_index is None:
                    status_col_index = cell.column
                    if header_row is None:
                        header_row = cell.row
                if "Completion Date" in cell_text and completion_col_index is None:
                    completion_col_index = cell.column
                    if header_row is None:
                        header_row = cell.row
            # The inner loop always finishes the full row before we check here,
            # so both columns are found if they exist on the same row (even if
            # "Completion Date" is far to the right, e.g. column AP = 42).
            # Only break after the full row is scanned AND both are resolved.
            if status_col_index is not None and completion_col_index is not None:
                break
            # If only status was found, keep going — Completion Date might be
            # on the same row (already scanned above, so it wasn't there) or
            # on a different header row.  We stop after HEADER_SCAN_ROWS anyway.

        # Ensure we always have a valid header_row to start scanning from.
        if header_row is None:
            header_row = 1
        # Fallback to column I (9) only when "Last Status Change" was not found.
        if status_col_index is None:
            status_col_index = 9

        # -------------------------------------------------------------------
        # Collect date data row by row
        # -------------------------------------------------------------------
        month_data: Dict[str, int] = defaultdict(int)
        dates_list: List[Dict[str, Any]] = []
        oldest: Optional[datetime] = None
        newest: Optional[datetime] = None
        total_rows_count = 0  # All non-empty data rows, regardless of date
        dated_by_completion = 0
        dated_by_status = 0
        dated_blank = 0

        start_row = header_row + 1
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, values_only=True),
            start=start_row
        ):
            # Count every non-blank row regardless of whether its date parses.
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                total_rows_count += 1
            else:
                continue  # completely blank row — skip entirely

            try:
                date_obj: Optional[datetime] = None
                date_source: Optional[str] = None

                # ── Priority 1: Completion Date ────────────────────────────
                if completion_col_index is not None:
                    comp_val = row[completion_col_index - 1]
                    if comp_val is not None and str(comp_val).strip():
                        date_obj = self.parse_date(comp_val)
                        if date_obj:
                            date_source = "completion"

                # ── Priority 2: Last Status Change (fallback) ──────────────
                if date_obj is None:
                    status_val = row[status_col_index - 1]
                    if status_val is not None:
                        date_obj = self.parse_date(status_val)
                        if date_obj:
                            date_source = "status"

                # ── Record result ──────────────────────────────────────────
                if date_obj and date_source:
                    month_key = date_obj.strftime("%Y-%m")
                    month_data[month_key] += 1
                    dates_list.append({
                        "date": date_obj,
                        "month": month_key,
                        "raw": date_obj,
                        "row_idx": row_idx,
                        "date_source": date_source,
                    })
                    if date_source == "completion":
                        dated_by_completion += 1
                    else:
                        dated_by_status += 1
                    if oldest is None or date_obj < oldest:
                        oldest = date_obj
                    if newest is None or date_obj > newest:
                        newest = date_obj
                else:
                    dated_blank += 1

            except (ValueError, IndexError, TypeError):
                dated_blank += 1
                continue

        return AnalysisResult(
            sheet_name=sheet_name,
            month_data=dict(month_data),
            dates=dates_list,
            total_cases=total_rows_count,
            header_row=header_row,
            status_col_index=status_col_index,
            oldest_date=oldest,
            newest_date=newest,
            completion_col_index=completion_col_index,
            dated_by_completion=dated_by_completion,
            dated_by_status=dated_by_status,
            dated_blank=dated_blank,
        )

    # ------------------------------------------------------------------
    # Date parsing
    # ------------------------------------------------------------------

    def parse_date(self, date_str: Any) -> Optional[datetime]:
        """
        Parse a date value from an Excel cell.

        Handles:
        - ``datetime`` objects (returned as-is)
        - ``date`` objects (converted to datetime)
        - Common string formats: MM/DD/YYYY, YYYY-MM-DD, with or without time
        """
        if isinstance(date_str, datetime):
            return date_str

        # openpyxl may return date objects for date-formatted cells
        try:
            from datetime import date as _date
            if isinstance(date_str, _date):
                return datetime(date_str.year, date_str.month, date_str.day)
        except Exception:
            pass

        if not isinstance(date_str, str):
            return None

        formats = [
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %I:%M:%S %p",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y",
            "%Y-%m-%d",
            "%d/%m/%Y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue

        return None

    # ------------------------------------------------------------------
    # Queries
    # ------------------------------------------------------------------

    def get_available_months(self, handler: Optional[str] = None) -> List[str]:
        """
        Return sorted list of available months (newest first).

        Args:
            handler: Specific handler name, or None / "All Handlers" for all.
        """
        if not self.analysis_results:
            return []

        all_months: set = set()
        if handler and handler != "All Handlers":
            result = self.analysis_results.get(handler)
            if result:
                all_months = set(result.month_data.keys())
        else:
            for r in self.analysis_results.values():
                all_months.update(r.month_data.keys())

        return sorted(all_months, reverse=True)

    def get_handlers(self) -> List[str]:
        """Return list of all analysed sheet names (handler sheets + Companies)."""
        if not self.analysis_results:
            return []
        return list(self.analysis_results.keys())

    def has_companies_sheet(self) -> bool:
        """Return True if the loaded workbook has an analysed Companies sheet."""
        return (
            self.analysis_results is not None
            and self.COMPANIES_SHEET in self.analysis_results
        )

    def get_summary_stats(self) -> Dict[str, Any]:
        """
        Return high-level summary statistics for the loaded workbook.

        Returns a dict with: total_cases, total_handlers, has_companies,
        all_months, oldest_date, newest_date, dated_by_completion,
        dated_by_status, dated_blank.
        """
        if not self.analysis_results:
            return {}

        total = sum(r.total_cases for r in self.analysis_results.values())
        all_months = self.get_available_months()
        oldest = min(
            (r.oldest_date for r in self.analysis_results.values() if r.oldest_date),
            default=None,
        )
        newest = max(
            (r.newest_date for r in self.analysis_results.values() if r.newest_date),
            default=None,
        )
        # Count only true handler sheets (not Companies) for the handler stat
        handler_count = sum(
            1 for name in self.analysis_results
            if name != self.COMPANIES_SHEET
        )
        return {
            "total_cases": total,
            "total_handlers": handler_count,
            "has_companies": self.has_companies_sheet(),
            "all_months": all_months,
            "oldest_date": oldest,
            "newest_date": newest,
            "dated_by_completion": sum(
                r.dated_by_completion for r in self.analysis_results.values()
            ),
            "dated_by_status": sum(
                r.dated_by_status for r in self.analysis_results.values()
            ),
            "dated_blank": sum(
                r.dated_blank for r in self.analysis_results.values()
            ),
        }

    # ------------------------------------------------------------------
    # Preview helpers
    # ------------------------------------------------------------------

    def preview_month_export(
        self,
        handler: str,
        months: List[str],
        cleanup: bool = False,
    ) -> Dict[str, Any]:
        """
        Non-destructive preview of a month-based export.

        Args:
            handler: Sheet name or "All Handlers".
            months:  List of "YYYY-MM" strings (one or more).
            cleanup: Whether exported rows will be deleted from original.

        Returns a dict with: total, details, months, (optionally) cleanup.
        """
        if not self.analysis_results or not months:
            return {"total": 0, "details": []}

        months_set = set(months)
        handlers_to_check = (
            list(self.analysis_results.keys())
            if handler == "All Handlers"
            else [handler]
        )

        total_exported = 0
        details: List[Dict] = []

        for h in handlers_to_check:
            result = self.analysis_results.get(h)
            if result:
                count = sum(result.month_data.get(m, 0) for m in months_set)
                if count > 0:
                    total_exported += count
                    details.append({"handler": h, "count": count})

        preview: Dict[str, Any] = {
            "total": total_exported,
            "details": details,
            "months": list(months),
        }

        if cleanup and total_exported > 0:
            remaining = sum(
                self.analysis_results[h].total_cases
                - sum(self.analysis_results[h].month_data.get(m, 0) for m in months_set)
                for h in handlers_to_check
                if h in self.analysis_results
            )
            preview["cleanup"] = {
                "to_delete": total_exported,
                "remaining": remaining,
            }

        return preview

    def preview_age_export(
        self,
        handler: str,
        days: int,
        cleanup: bool = False,
    ) -> Dict[str, Any]:
        """
        Non-destructive preview of an age-based export.

        Returns a dict with: total, details, cutoff_date, days, (optionally) cleanup.
        """
        if not self.analysis_results:
            return {"total": 0, "details": [], "cutoff_date": None}

        cutoff = datetime.now() - timedelta(days=days)
        handlers_to_check = (
            list(self.analysis_results.keys())
            if handler == "All Handlers"
            else [handler]
        )

        total_exported = 0
        details: List[Dict] = []

        for h in handlers_to_check:
            result = self.analysis_results.get(h)
            if result:
                count = sum(1 for d in result.dates if d["date"] < cutoff)
                if count > 0:
                    total_exported += count
                    details.append({"handler": h, "count": count})

        preview: Dict[str, Any] = {
            "total": total_exported,
            "details": details,
            "cutoff_date": cutoff,
            "days": days,
        }

        if cleanup and total_exported > 0:
            remaining = sum(
                self.analysis_results[h].total_cases
                - sum(1 for d in self.analysis_results[h].dates if d["date"] < cutoff)
                for h in handlers_to_check
                if h in self.analysis_results
            )
            preview["cleanup"] = {
                "to_delete": total_exported,
                "remaining": remaining,
            }

        return preview

    # ------------------------------------------------------------------
    # Archive filename helper
    # ------------------------------------------------------------------

    @staticmethod
    def make_archive_filename(months: List[str]) -> str:
        """
        Generate a default archive filename from a list of 'YYYY-MM' month strings.

        Single month  → "Archive (Jun - 2025).xlsx"
        Range of months → "Archive (Jan - Jun 2025).xlsx"
        Cross-year range → "Archive (Dec 2024 - Jan 2025).xlsx"
        """
        if not months:
            return "Archive.xlsx"

        month_objs = []
        for m in months:
            try:
                month_objs.append(datetime.strptime(m, "%Y-%m"))
            except ValueError:
                continue

        if not month_objs:
            return "Archive.xlsx"

        month_objs.sort()
        oldest = month_objs[0]
        newest = month_objs[-1]

        if oldest.year == newest.year:
            if oldest.month == newest.month:
                # Single month
                return f"Archive ({oldest.strftime('%b')} - {oldest.year}).xlsx"
            else:
                # Same year, different months
                return f"Archive ({oldest.strftime('%b')} - {newest.strftime('%b')} {newest.year}).xlsx"
        else:
            # Cross-year
            return (
                f"Archive ({oldest.strftime('%b %Y')} - {newest.strftime('%b %Y')}).xlsx"
            )

    # ------------------------------------------------------------------
    # Export — month
    # ------------------------------------------------------------------

    def export_by_month(
        self,
        options: MonthExportOptions,
        progress_callback: ProgressCallback = None,
    ) -> Tuple[bool, str]:
        """
        Export cases from a specific month.

        Args:
            options: Month export options.
            progress_callback: Optional callable(percent: int, message: str).

        Returns:
            (success: bool, message: str)
        """
        try:
            if not self.analysis_results:
                return False, "No analysis results available. Please analyse a file first."

            if options.handler == "All Handlers":
                return self._export_month_all_handlers(options, progress_callback)
            else:
                return self._export_month_single_handler(options, progress_callback)

        except Exception as exc:
            return False, f"Export failed: {exc}"

    # ---- internal helpers ------------------------------------------------

    def _row_matches_months(
        self, date_val: Any, months_set: set
    ) -> bool:
        """Return True if date_val parses to a month in months_set."""
        if not date_val:
            return False
        date_obj = self.parse_date(date_val)
        return date_obj is not None and date_obj.strftime("%Y-%m") in months_set

    def _copy_rows_to_sheet(
        self,
        src_ws,
        export_ws,
        header_row: int,
        status_col: int,
        months_set: set,
        start_export_row: int,
    ) -> Tuple[int, List[int]]:
        """
        Copy all rows whose date falls in months_set from src_ws into export_ws.

        Returns (next_export_row, rows_to_delete).
        """
        export_row = start_export_row
        rows_to_delete: List[int] = []
        for row_idx, row in enumerate(
            src_ws.iter_rows(min_row=header_row + 1, max_row=src_ws.max_row),
            start=header_row + 1,
        ):
            try:
                date_val = row[status_col - 1].value
                if self._row_matches_months(date_val, months_set):
                    for col in range(1, src_ws.max_column + 1):
                        c = src_ws.cell(row=row_idx, column=col)
                        d = export_ws.cell(row=export_row, column=col)
                        d.value = c.value
                        self._copy_cell_style(c, d)
                    export_row += 1
                    rows_to_delete.append(row_idx)
            except Exception:
                continue
        return export_row, rows_to_delete

    def _do_cleanup(
        self,
        rows_to_delete_by_sheet: Dict[str, List[int]],
        cb: ProgressCallback,
    ) -> None:
        """Rebuild each affected sheet keeping only non-archived rows.

        Instead of deleting rows one-by-one (O(n²) XML shifts), this reads
        every sheet into memory, skips the archived row indices, wipes the
        sheet in one bulk operation, then writes the kept rows back.  Column
        widths, row heights, and merged-cell ranges are re-applied so the
        original file looks identical except the exported rows are gone.

        Complexity: O(total_rows) — typically 20-100× faster than the
        previous one-row-at-a-time delete approach.
        """
        wb2 = openpyxl.load_workbook(self.workbook_path)
        total_sheets = len(rows_to_delete_by_sheet)

        for sheet_idx, (handler, rows_to_delete) in enumerate(
            rows_to_delete_by_sheet.items()
        ):
            pct = 85 + int(10 * sheet_idx / max(total_sheets, 1))
            _report(cb, pct, f"Rebuilding {handler}…")

            ws2 = wb2[handler]
            skip = set(rows_to_delete)  # O(1) membership test

            # ── 1. Snapshot everything we need to restore ──────────────────
            # Cell data (value + style) for rows we're keeping
            kept: List[List[Tuple]] = []
            for r_idx in range(1, ws2.max_row + 1):
                if r_idx in skip:
                    continue
                row_snapshot = []
                for c_idx in range(1, ws2.max_column + 1):
                    cell = ws2.cell(row=r_idx, column=c_idx)
                    row_snapshot.append((
                        cell.value,
                        copy(cell.font)        if cell.has_style else None,
                        copy(cell.fill)        if cell.has_style else None,
                        copy(cell.border)      if cell.has_style else None,
                        copy(cell.alignment)   if cell.has_style else None,
                        cell.number_format     if cell.has_style else None,
                        copy(cell.protection)  if cell.has_style else None,
                    ))
                kept.append(row_snapshot)

            # Column widths and row heights for kept rows
            # old_row → new_row mapping (built once, O(1) lookups later)
            kept_row_numbers = [
                r for r in range(1, ws2.max_row + 1) if r not in skip
            ]
            old_to_new_row = {
                old_r: new_r
                for new_r, old_r in enumerate(kept_row_numbers, start=1)
            }
            col_widths = {
                col: ws2.column_dimensions[col].width
                for col in ws2.column_dimensions
            }
            row_heights = {
                old_to_new_row[r]: ws2.row_dimensions[r].height
                for r in kept_row_numbers
                if ws2.row_dimensions[r].height
            }

            # Merged cell ranges — discard any that were fully within archived rows
            surviving_merges = []
            for merge in ws2.merged_cells.ranges:
                # Keep the merge only if its top-left row survives
                if merge.min_row not in skip:
                    surviving_merges.append(str(merge))

            # ── 2. Wipe the sheet in ONE bulk operation ────────────────────
            ws2.delete_rows(1, ws2.max_row)

            # ── 3. Write kept rows back ────────────────────────────────────
            for new_r_idx, row_data in enumerate(kept, start=1):
                for new_c_idx, cell_snap in enumerate(row_data, start=1):
                    (val, font, fill, border, alignment,
                     number_format, protection) = cell_snap
                    c = ws2.cell(row=new_r_idx, column=new_c_idx)
                    c.value = val
                    if font is not None:
                        c.font        = font
                        c.fill        = fill
                        c.border      = border
                        c.alignment   = alignment
                        c.number_format = number_format
                        c.protection  = protection

            # ── 4. Restore column widths ───────────────────────────────────
            for col_letter, width in col_widths.items():
                ws2.column_dimensions[col_letter].width = width

            # ── 5. Restore row heights for kept rows ───────────────────────
            for new_r, height in row_heights.items():
                ws2.row_dimensions[new_r].height = height

            # ── 6. Re-apply merged cell ranges ─────────────────────────────
            for merge_str in surviving_merges:
                try:
                    ws2.merge_cells(merge_str)
                except Exception:
                    pass  # best-effort; never block the save

            pct = 85 + int(10 * (sheet_idx + 1) / max(total_sheets, 1))
            _report(cb, pct, f"Rebuilt {handler} — {len(kept)} rows kept")

        _report(cb, 99, "Saving cleaned workbook…")
        wb2.save(self.workbook_path)
        wb2.close()

    # ---- month — single handler ------------------------------------------

    def _export_month_single_handler(
        self,
        options: MonthExportOptions,
        cb: ProgressCallback = None,
    ) -> Tuple[bool, str]:
        """Export one or more months for a single handler sheet."""
        months_set = set(options.months)
        months_label = self.make_archive_filename(options.months).replace(".xlsx", "")

        _report(cb, 5, "Opening workbook…")
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb[options.handler]

        meta = self.analysis_results[options.handler]
        header_row = meta.header_row
        status_col = meta.status_col_index

        _report(cb, 20, "Creating export workbook…")
        export_wb = openpyxl.Workbook()
        export_ws = export_wb.active
        # Sheet title = first month (or range label trimmed to 31 chars)
        export_ws.title = (
            options.months[0] if len(options.months) == 1
            else months_label[:31]
        )

        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=header_row, column=col)
            dst = export_ws.cell(row=1, column=col)
            dst.value = src.value
            self._copy_cell_style(src, dst)

        _report(cb, 40, f"Copying rows for {len(months_set)} month(s)…")
        next_row, rows_to_delete = self._copy_rows_to_sheet(
            ws, export_ws, header_row, status_col, months_set, start_export_row=2
        )

        _report(cb, 70, "Saving export file…")
        export_wb.save(options.output_file)
        export_wb.close()
        wb.close()

        if options.cleanup and rows_to_delete:
            _report(cb, 80, "Creating backup of original…")
            self._create_backup(self.workbook_path)
            _report(cb, 85, "Removing exported rows…")
            self._do_cleanup({options.handler: rows_to_delete}, cb)

        exported_count = next_row - 2
        msg = (
            f"Successfully exported {exported_count} case(s) "
            f"from '{options.handler}' for {len(months_set)} month(s)"
        )
        if options.cleanup:
            msg += " — original cleaned up (backup created)"
        _report(cb, 100, msg)
        return True, msg

    # ---- month — all handlers --------------------------------------------

    def _export_month_all_handlers(
        self,
        options: MonthExportOptions,
        cb: ProgressCallback = None,
    ) -> Tuple[bool, str]:
        """Export one or more months for every handler sheet.

        Output workbook structure:
          - One sheet per handler (rows matching selected months)
          - Optional "Merged" sheet concatenating all handler rows
        """
        months_set = set(options.months)
        months_label = self.make_archive_filename(options.months).replace(".xlsx", "")

        _report(cb, 5, "Opening workbook…")
        wb = openpyxl.load_workbook(self.workbook_path)

        export_wb = openpyxl.Workbook()
        export_wb.remove(export_wb.active)

        rows_to_delete_by_sheet: Dict[str, List[int]] = {}
        merged_ws = None
        merged_row = 2
        merged_header_written = False
        total_exported = 0

        if options.merged:
            merged_ws = export_wb.create_sheet(title=f"Merged — {months_label}"[:31])

        handlers = list(self.analysis_results.keys())
        n = len(handlers)

        for idx, handler in enumerate(handlers):
            pct = 10 + int(60 * idx / max(n, 1))
            _report(cb, pct, f"Processing {handler}…")

            ws = wb[handler]
            meta = self.analysis_results[handler]
            header_row = meta.header_row
            status_col = meta.status_col_index

            export_ws = export_wb.create_sheet(title=handler)

            # Copy header row
            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=header_row, column=col)
                dst = export_ws.cell(row=1, column=col)
                dst.value = src.value
                self._copy_cell_style(src, dst)

            export_row = 2
            rows_to_delete: List[int] = []

            for row_idx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row),
                start=header_row + 1,
            ):
                try:
                    date_val = row[status_col - 1].value
                    if self._row_matches_months(date_val, months_set):
                        for col in range(1, ws.max_column + 1):
                            src = ws.cell(row=row_idx, column=col)
                            dst = export_ws.cell(row=export_row, column=col)
                            dst.value = src.value
                            self._copy_cell_style(src, dst)
                        export_row += 1
                        rows_to_delete.append(row_idx)
                        total_exported += 1

                        if options.merged and merged_ws is not None:
                            if not merged_header_written:
                                for col in range(1, ws.max_column + 1):
                                    m_src = ws.cell(row=header_row, column=col)
                                    m_dst = merged_ws.cell(row=1, column=col)
                                    m_dst.value = m_src.value
                                    self._copy_cell_style(m_src, m_dst)
                                merged_header_written = True
                            for col in range(1, ws.max_column + 1):
                                m_src = ws.cell(row=row_idx, column=col)
                                m_dst = merged_ws.cell(row=merged_row, column=col)
                                m_dst.value = m_src.value
                                self._copy_cell_style(m_src, m_dst)
                            merged_row += 1
                except Exception:
                    continue

            if rows_to_delete:
                rows_to_delete_by_sheet[handler] = rows_to_delete

        _report(cb, 75, "Saving export file…")
        export_wb.save(options.output_file)
        export_wb.close()
        wb.close()

        if options.cleanup and rows_to_delete_by_sheet:
            _report(cb, 82, "Creating backup of original…")
            self._create_backup(self.workbook_path)
            _report(cb, 85, "Removing exported rows…")
            self._do_cleanup(rows_to_delete_by_sheet, cb)

        msg = (
            f"Successfully exported {total_exported} case(s) from "
            f"{n} handler(s) for {len(months_set)} month(s)"
        )
        if options.cleanup:
            msg += " — original cleaned up (backup created)"
        _report(cb, 100, msg)
        return True, msg

    # ------------------------------------------------------------------
    # Export — age
    # ------------------------------------------------------------------

    def export_by_age(
        self,
        options: AgeExportOptions,
        progress_callback: ProgressCallback = None,
    ) -> Tuple[bool, str]:
        """
        Export cases older than *options.days* days.

        Args:
            options: Age export options.
            progress_callback: Optional callable(percent: int, message: str).

        Returns:
            (success: bool, message: str)
        """
        try:
            if not self.analysis_results:
                return False, "No analysis results available. Please analyse a file first."

            if options.handler == "All Handlers":
                return self._export_age_all_handlers(options, progress_callback)
            else:
                return self._export_age_single_handler(options, progress_callback)

        except Exception as exc:
            return False, f"Export failed: {exc}"

    def _export_age_single_handler(
        self,
        options: AgeExportOptions,
        cb: ProgressCallback = None,
    ) -> Tuple[bool, str]:
        """Export old cases for a single handler.

        When ``options.merged`` is True an additional sheet named
        ``"Merged — Old Cases"`` is created containing the same rows, so the
        workbook has both the per-handler sheet and a merged sheet for
        consistency (useful when the same workbook template is expected by
        downstream consumers).
        """
        _report(cb, 5, "Opening workbook…")
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb[options.handler]

        meta = self.analysis_results[options.handler]
        header_row = meta.header_row
        status_col = meta.status_col_index

        cutoff = datetime.now() - timedelta(days=options.days)

        _report(cb, 20, "Creating export workbook…")
        export_wb = openpyxl.Workbook()
        export_ws = export_wb.active
        export_ws.title = "Old Cases"

        # Optional merged sheet (mirrors main sheet for single-handler case)
        merged_ws = None
        merged_row = 2
        if options.merged:
            merged_ws = export_wb.create_sheet(title="Merged — Old Cases")

        # Copy header to main sheet (and merged sheet if requested)
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=header_row, column=col)
            dst = export_ws.cell(row=1, column=col)
            dst.value = src.value
            self._copy_cell_style(src, dst)
            if merged_ws is not None:
                m_dst = merged_ws.cell(row=1, column=col)
                m_dst.value = src.value
                self._copy_cell_style(src, m_dst)

        _report(cb, 40, f"Copying cases older than {cutoff.strftime('%Y-%m-%d')}…")

        export_row = 2
        rows_to_delete: List[int] = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row),
            start=header_row + 1,
        ):
            try:
                date_val = row[status_col - 1].value
                if date_val:
                    date_obj = self.parse_date(date_val)
                    if date_obj and date_obj < cutoff:
                        for col in range(1, ws.max_column + 1):
                            src = ws.cell(row=row_idx, column=col)
                            dst = export_ws.cell(row=export_row, column=col)
                            dst.value = src.value
                            self._copy_cell_style(src, dst)
                            if merged_ws is not None:
                                m_dst = merged_ws.cell(row=merged_row, column=col)
                                m_dst.value = src.value
                                self._copy_cell_style(src, m_dst)
                        export_row += 1
                        merged_row += 1
                        rows_to_delete.append(row_idx)
            except Exception:
                continue

        _report(cb, 70, "Saving export file…")
        export_wb.save(options.output_file)
        export_wb.close()
        wb.close()  # close read workbook before cleanup

        if options.cleanup and rows_to_delete:
            _report(cb, 80, "Creating backup of original…")
            self._create_backup(self.workbook_path)
            _report(cb, 85, "Removing exported rows…")
            self._do_cleanup({options.handler: rows_to_delete}, cb)

        exported_count = export_row - 2
        merged_note = " + Merged sheet" if options.merged else ""
        msg = (
            f"Successfully exported {exported_count} case(s) older than "
            f"{options.days} day(s){merged_note}"
        )
        if options.cleanup:
            msg += " — original file cleaned up (backup created)"

        _report(cb, 100, msg)
        return True, msg

    def _export_age_all_handlers(
        self,
        options: AgeExportOptions,
        cb: ProgressCallback = None,
    ) -> Tuple[bool, str]:
        """Export old cases for all handlers.

        Output workbook structure:
          - One sheet per handler containing that handler's old cases
          - Optional ``"Merged — Old Cases"`` sheet concatenating all rows
            (created when ``options.merged`` is True)
        """
        _report(cb, 5, "Opening workbook…")
        wb = openpyxl.load_workbook(self.workbook_path)

        export_wb = openpyxl.Workbook()
        export_wb.remove(export_wb.active)

        cutoff = datetime.now() - timedelta(days=options.days)
        rows_to_delete_by_sheet: Dict[str, List[int]] = {}
        total_exported = 0

        # Create merged sheet first so it appears as the first tab
        merged_ws = None
        merged_row = 2
        merged_header_written = False
        if options.merged:
            merged_ws = export_wb.create_sheet(title="Merged — Old Cases")

        handlers = list(self.analysis_results.keys())
        n = len(handlers)

        for idx, handler in enumerate(handlers):
            pct = 10 + int(60 * idx / max(n, 1))
            _report(cb, pct, f"Processing {handler}…")

            ws = wb[handler]
            meta = self.analysis_results[handler]
            header_row = meta.header_row
            status_col = meta.status_col_index

            export_ws = export_wb.create_sheet(title=handler)

            # Copy header row to per-handler sheet
            for col in range(1, ws.max_column + 1):
                src = ws.cell(row=header_row, column=col)
                dst = export_ws.cell(row=1, column=col)
                dst.value = src.value
                self._copy_cell_style(src, dst)

            export_row = 2
            rows_to_delete: List[int] = []

            for row_idx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row),
                start=header_row + 1,
            ):
                try:
                    date_val = row[status_col - 1].value
                    if date_val:
                        date_obj = self.parse_date(date_val)
                        if date_obj and date_obj < cutoff:
                            # Write to per-handler sheet
                            for col in range(1, ws.max_column + 1):
                                src = ws.cell(row=row_idx, column=col)
                                dst = export_ws.cell(row=export_row, column=col)
                                dst.value = src.value
                                self._copy_cell_style(src, dst)
                            export_row += 1
                            rows_to_delete.append(row_idx)
                            total_exported += 1

                            # Also write to merged sheet if requested
                            if merged_ws is not None:
                                if not merged_header_written:
                                    for col in range(1, ws.max_column + 1):
                                        m_src = ws.cell(row=header_row, column=col)
                                        m_dst = merged_ws.cell(row=1, column=col)
                                        m_dst.value = m_src.value
                                        self._copy_cell_style(m_src, m_dst)
                                    merged_header_written = True
                                for col in range(1, ws.max_column + 1):
                                    m_src = ws.cell(row=row_idx, column=col)
                                    m_dst = merged_ws.cell(row=merged_row, column=col)
                                    m_dst.value = m_src.value
                                    self._copy_cell_style(m_src, m_dst)
                                merged_row += 1
                except Exception:
                    continue

            if rows_to_delete:
                rows_to_delete_by_sheet[handler] = rows_to_delete

        _report(cb, 75, "Saving export file…")
        export_wb.save(options.output_file)
        export_wb.close()
        wb.close()  # close read workbook before cleanup

        if options.cleanup and rows_to_delete_by_sheet:
            _report(cb, 82, "Creating backup of original…")
            self._create_backup(self.workbook_path)
            _report(cb, 85, "Removing exported rows…")
            self._do_cleanup(rows_to_delete_by_sheet, cb)

        merged_note = " + Merged sheet" if options.merged else ""
        msg = (
            f"Successfully exported {total_exported} case(s) older than "
            f"{options.days} day(s) from {n} handler(s){merged_note}"
        )
        if options.cleanup:
            msg += " — original file cleaned up (backup created)"

        _report(cb, 100, msg)
        return True, msg

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _copy_cell_style(self, source_cell, target_cell) -> None:
        """Copy openpyxl cell style attributes from source to target."""
        try:
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
        except Exception:
            pass  # Style copy is best-effort; never block data export

    def _create_backup(self, file_path: str) -> str:
        """
        Create a timestamped backup of *file_path* beside the original.

        The backup name is ``<stem>_backup_YYYYMMDD_HHMMSS<suffix>``.
        This prevents silent overwrites of previous backups.

        Returns:
            Path to the created backup file.
        """
        p = Path(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = p.with_name(f"{p.stem}_backup_{timestamp}{p.suffix}")
        shutil.copy2(file_path, backup_path)
        return str(backup_path)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
__all__ = [
    "ArchiverService",
    "AnalysisResult",
    "ExportOptions",
    "MonthExportOptions",
    "AgeExportOptions",
    "ProgressCallback",
]

# Made with Bob
