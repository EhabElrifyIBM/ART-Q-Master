import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import scrolledtext
from tkinter import font as tkfont
from datetime import datetime, timedelta
import pandas as pd
from collections import defaultdict
import os
from pathlib import Path
import shutil
from copy import copy
import threading
import sys

from ui.responsive import calculate_responsive_font_size


class ExcelWorkbookAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("ART Q Master V2 - Archiver")
        self.root.geometry("1100x760")
        
        # Set icon
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir)) # src/Archiver -> src -> root
            icon_path = os.path.join(project_root, 'assets', 'ibm_logo.png')
            if os.path.exists(icon_path):
                img = tk.PhotoImage(file=icon_path)
                self.root.iconphoto(False, img)
        except Exception:
            pass
        
        self.workbook_path = None
        self.workbook_data = None
        self.analysis_results = None
        self._font_size = 18
        self._style = ttk.Style()
        self._last_applied_font_size = None
        self._is_rebuilding = False
        self._configure_theme()

        self.root.bind("<Configure>", self._on_window_resize)
        
        self.create_main_ui()
    
    def _configure_theme(self):
        self.root.configure(bg="#eef4ff")
        try:
            self._style.theme_use("clam")
        except Exception:
            pass

        base = self._font_size
        title = max(18, base + 4)
        small = max(11, base - 1)
        button_pad_x = max(8, int(base * 0.65))
        button_pad_y = max(6, int(base * 0.45))

        self._fonts = {
            "base": tkfont.Font(family="Segoe UI", size=base),
            "bold": tkfont.Font(family="Segoe UI", size=base, weight="bold"),
            "title": tkfont.Font(family="Segoe UI", size=title, weight="bold"),
            "small": tkfont.Font(family="Segoe UI", size=small),
            "mono": tkfont.Font(family="Consolas", size=max(10, base - 1)),
        }

        self._style.configure("V2.TFrame", background="#eef4ff")
        self._style.configure("V2Card.TFrame", background="#ffffff")
        self._style.configure("V2.TLabel", background="#eef4ff", foreground="#161616", font=self._fonts["base"])
        self._style.configure("V2Title.TLabel", background="#eef4ff", foreground="#0f62fe", font=self._fonts["title"])
        self._style.configure("V2Muted.TLabel", background="#eef4ff", foreground="#525252", font=self._fonts["small"])
        self._style.configure("V2Labelframe.TLabelframe", background="#ffffff", borderwidth=1, relief="solid")
        self._style.configure("V2Labelframe.TLabelframe.Label", background="#ffffff", foreground="#0f62fe", font=self._fonts["bold"])
        self._style.configure(
            "V2.TButton",
            font=self._fonts["bold"],
            padding=(button_pad_x, button_pad_y),
            background="#0f62fe",
            foreground="#ffffff",
            borderwidth=0,
        )
        self._style.map("V2.TButton", background=[("active", "#0353e9")], foreground=[("active", "#ffffff")])
        self._style.configure(
            "V2Secondary.TButton",
            font=self._fonts["bold"],
            padding=(button_pad_x, button_pad_y),
            background="#dbeafe",
            foreground="#161616",
            borderwidth=0,
        )
        self._style.map("V2Secondary.TButton", background=[("active", "#bfdbfe")])

        if hasattr(self, "results_text"):
            self.results_text.configure(font=self._fonts["mono"])

    def _calculate_window_font_size(self) -> int:
        width = max(760, self.root.winfo_width())
        height = max(560, self.root.winfo_height())
        return calculate_responsive_font_size(width, height, base_size=18)

    def _on_window_resize(self, event):
        if event.widget is not self.root or self._is_rebuilding:
            return
        new_size = self._calculate_window_font_size()
        if new_size != self._last_applied_font_size and new_size != self._font_size:
            self._font_size = new_size
            self._last_applied_font_size = new_size
            self._configure_theme()
            self._rebuild_main_ui()

    def _rebuild_main_ui(self):
        self._is_rebuilding = True
        try:
            self.create_main_ui()
        finally:
            self._is_rebuilding = False

    def create_main_ui(self):
        """Create the main user interface"""
        for widget in self.root.winfo_children():
            widget.destroy()

        main_frame = ttk.Frame(self.root, padding=str(max(12, self._font_size)), style="V2.TFrame")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # File selection
        ttk.Label(main_frame, text="Excel Workbook:", style="V2.TLabel").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.file_label = ttk.Label(main_frame, text="No file selected", style="V2Muted.TLabel")
        self.file_label.grid(row=0, column=1, sticky=tk.W, padx=10)
        
        ttk.Button(main_frame, text="Browse File", command=self.browse_file, style="V2Secondary.TButton").grid(row=0, column=2, padx=5)
        ttk.Button(main_frame, text="Analyze", command=self.analyze_workbook, style="V2.TButton").grid(row=0, column=3, padx=5)
        
        # Status
        ttk.Label(main_frame, text="Status:", style="V2.TLabel").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.status_label = ttk.Label(main_frame, text="Ready", style="V2Muted.TLabel")
        self.status_label.grid(row=1, column=1, columnspan=3, sticky=tk.W, padx=10)
        
        # Results display
        ttk.Label(main_frame, text="Analysis Results:", style="V2Title.TLabel").grid(row=2, column=0, columnspan=4, sticky=tk.W, pady=(20, 5))
        
        self.results_text = scrolledtext.ScrolledText(
            main_frame,
            height=15,
            width=96,
            state=tk.DISABLED,
            bg="#ffffff",
            fg="#161616",
            font=self._fonts["mono"],
            relief="solid",
            borderwidth=1,
        )
        self.results_text.grid(row=3, column=0, columnspan=4, sticky="nsew", pady=5)
        
        # Action buttons
        button_frame = ttk.Frame(main_frame, style="V2.TFrame")
        button_frame.grid(row=4, column=0, columnspan=4, sticky="ew", pady=10)
        
        ttk.Button(button_frame, text="Export by Specific Month", command=self.export_by_month, style="V2.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export Older Than Days", command=self.export_old_cases, style="V2.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Main Menu", command=self.return_to_main_menu, style="V2Secondary.TButton").pack(side=tk.RIGHT, padx=5)

        # Footer
        footer_text = "Developed by: Ehab Elrify | Adam Maged\nEmail: ehab.elrify@ibm.com | abdelrahman.maged@ibm.com\nAssurance Resolution Team"
        footer_label = ttk.Label(
             main_frame,
             text=footer_text,
             justify=tk.CENTER,
             style="V2Muted.TLabel"
         )
        footer_label.grid(row=5, column=0, columnspan=4, pady=10)

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)

    def return_to_main_menu(self):
        """Return to the main menu.

        In a frozen .exe the main menu is already running in the same process —
        just destroy this Tk root.  In development, re-launch main.py in a
        subprocess as before.
        """
        try:
            if not getattr(sys, 'frozen', False):
                import subprocess as _sp
                current_dir = os.path.dirname(os.path.abspath(__file__))
                src_dir = os.path.dirname(current_dir)
                main_script = os.path.join(src_dir, 'main.py')
                if os.path.exists(main_script):
                    _sp.Popen([sys.executable, main_script])
                else:
                    messagebox.showerror("Error", f"Main menu script not found: {main_script}")
            # Frozen path: main menu window is still open — just close this one
            self.root.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to return to Main Menu: {e}")
    
    def browse_file(self):
        """Browse and select an Excel file"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.workbook_path = file_path
            self.file_label.config(text=Path(file_path).name, foreground="black")
            self.status_label.config(text="File selected. Click 'Analyze' to proceed.", foreground="blue")
    
    def analyze_workbook(self):
        """Analyze the Excel workbook"""
        if not self.workbook_path:
            messagebox.showerror("Error", "Please select an Excel file first.")
            return
        
        try:
            self.status_label.config(text="Analyzing... Please wait.", foreground="orange")
            self.root.update()
            
            # Load workbook
            wb = openpyxl.load_workbook(self.workbook_path, data_only=True)
            
            # Find all sheets containing "'s Cases"
            handler_sheets = [sheet for sheet in wb.sheetnames if "'s Cases" in sheet]
            
            if not handler_sheets:
                messagebox.showwarning("Warning", "No sheets found containing \"'s Cases\" in the name.")
                self.status_label.config(text="No handler sheets found.", foreground="red")
                return
            
            # Analyze data
            self.analysis_results = self.analyze_sheets(wb, handler_sheets)
            
            # Display results
            self.display_results(handler_sheets)
            
            self.status_label.config(text=f"Analysis complete. Found {len(handler_sheets)} handler sheet(s).", foreground="green")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.status_label.config(text="Error during analysis.", foreground="red")
    
    def analyze_sheets(self, wb, handler_sheets):
        """Analyze all handler sheets and extract date information"""
        results = {}
        
        for sheet_name in handler_sheets:
            ws = wb[sheet_name]
            
            # Find "Last Status Change" column
            header_row = None
            status_col_index = None
            
            for row in ws.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    if cell.value and "Last Status Change" in str(cell.value):
                        status_col_index = cell.column
                        header_row = cell.row
                        break
                if status_col_index:
                    break
            
            if not status_col_index:
                # Try column I by default
                status_col_index = 9  # Column I
                header_row = 1
            
            # Extract dates
            month_data = defaultdict(int)
            dates_list = []
            
            for row in ws.iter_rows(min_row=header_row + 1 if header_row else 2, values_only=True):
                try:
                    date_str = row[status_col_index - 1]
                    if date_str:
                        # Parse date
                        date_obj = self.parse_date(date_str)
                        if date_obj:
                            month_key = date_obj.strftime("%Y-%m")
                            month_data[month_key] += 1
                            dates_list.append({
                                'date': date_obj,
                                'month': month_key,
                                'raw': date_str
                            })
                except (ValueError, IndexError, TypeError):
                    continue
            
            results[sheet_name] = {
                'month_data': dict(month_data),
                'dates': dates_list,
                'total_cases': len(dates_list),
                'header_row': header_row,
                'status_col_index': status_col_index
            }
        
        return results
    
    def parse_date(self, date_str):
        """Parse date string in format YYYY-MM-DD HH:MM:SS"""
        try:
            if isinstance(date_str, datetime):
                return date_str
            
            date_str = str(date_str).strip()
            
            # Try different date formats
            formats = [
                "%Y-%m-%d %H:%M:%S",
                "%Y/%m/%d %H:%M:%S",
                "%d-%m-%Y %H:%M:%S",
                "%Y-%m-%d",
            ]
            
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
            
            return None
        except Exception:
            return None
    
    def copy_cell_style(self, source_cell, target_cell):
        """Copy formatting from source cell to target cell"""
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

    def _create_progress_dialog(self, title="Progress"):
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.geometry("700x460")
        dlg.transient(self.root)
        dlg.configure(bg="#eef4ff")

        pb = ttk.Progressbar(dlg, mode='indeterminate')
        pb.pack(fill=tk.X, padx=10, pady=6)

        log_text = tk.Text(
            dlg,
            height=18,
            bg="#ffffff",
            fg="#161616",
            font=("Consolas", max(14, self._font_size - 2)),
            relief="solid",
            borderwidth=1,
        )
        log_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        log_text.config(state=tk.DISABLED)

        # Footer
        footer_text = "Developed by: Ehab Elrify | Adam Maged\nEmail: ehab.elrify@ibm.com | abdelrahman.maged@ibm.com\nAssurance Resolution Team"
        footer_label = ttk.Label(
             dlg,
             text=footer_text,
             justify=tk.CENTER,
             style="V2Muted.TLabel"
         )
        footer_label.pack(side="bottom", pady=10)

        return dlg, pb, log_text

    def _append_progress_log(self, log_text_widget, msg):
        def _append():
            try:
                log_text_widget.config(state=tk.NORMAL)
                log_text_widget.insert(tk.END, msg + "\n")
                log_text_widget.see(tk.END)
                log_text_widget.config(state=tk.DISABLED)
            except Exception:
                pass
        self.root.after(0, _append)

    def _start_progressbar(self, pb):
        self.root.after(0, lambda: pb.start(50))

    def _stop_progressbar(self, pb):
        self.root.after(0, lambda: pb.stop())

    def _export_month_thread(self, handler, month, save_path, cleanup, merged, progress_dialog, pb, log_text):
        """Background worker for exporting by month. Calls GUI updates on main thread and logs progress."""
        try:
            self._append_progress_log(log_text, f"Starting export: month={month}, handler={handler}")
            self._start_progressbar(pb)

            if handler == "All Handlers":
                self._append_progress_log(log_text, "Exporting all handlers...")
                self.export_data_by_month_all(month, save_path, cleanup, merged)
            else:
                self._append_progress_log(log_text, f"Exporting handler: {handler}...")
                self.export_data_by_month(handler, month, save_path, cleanup)

            self._append_progress_log(log_text, "Export completed successfully")
        except Exception as e:
            self._append_progress_log(log_text, f"Export failed: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"Export failed: {str(e)}"))
            self._stop_progressbar(pb)
            return

        def on_done():
            self._stop_progressbar(pb)
            messagebox.showinfo("Success", f"Data exported to {save_path}")
            if cleanup:
                messagebox.showinfo("Info", "Original file cleaned and backup created.")
            try:
                progress_dialog.destroy()
            except Exception:
                pass

        self.root.after(0, on_done)

    def _export_old_thread(self, handler, days, save_path, cleanup, merged, progress_dialog, pb, log_text):
        """Background worker for exporting old cases. Calls GUI updates on main thread and logs progress."""
        try:
            self._append_progress_log(log_text, f"Starting export of cases older than {days} days for handler={handler}")
            self._start_progressbar(pb)

            if handler == "All Handlers":
                self._append_progress_log(log_text, "Exporting all handlers...")
                self.export_old_cases_data_all(days, save_path, cleanup, merged)
            else:
                self._append_progress_log(log_text, f"Exporting handler: {handler}...")
                self.export_old_cases_data(handler, days, save_path, cleanup)

            self._append_progress_log(log_text, "Export completed successfully")
        except Exception as e:
            self._append_progress_log(log_text, f"Export failed: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"Export failed: {str(e)}"))
            self._stop_progressbar(pb)
            return

        def on_done():
            self._stop_progressbar(pb)
            messagebox.showinfo("Success", f"Data exported to {save_path}")
            if cleanup:
                messagebox.showinfo("Info", "Original file cleaned and backup created.")
            try:
                progress_dialog.destroy()
            except Exception:
                pass

        self.root.after(0, on_done)
    
    def display_results(self, handler_sheets):
        """Display analysis results in the text widget"""
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        
        for sheet_name in handler_sheets:
            data = self.analysis_results[sheet_name]
            self.results_text.insert(tk.END, f"\n{'='*100}\n")
            self.results_text.insert(tk.END, f"Handler: {sheet_name}\n")
            self.results_text.insert(tk.END, f"Total Cases: {data['total_cases']}\n")
            self.results_text.insert(tk.END, f"{'-'*100}\n")
            self.results_text.insert(tk.END, "Month\t\t\tNumber of Cases\n")
            self.results_text.insert(tk.END, f"{'-'*100}\n")
            
            for month in sorted(data['month_data'].keys(), reverse=True):
                count = data['month_data'][month]
                self.results_text.insert(tk.END, f"{month}\t\t\t{count}\n")
        
        self.results_text.config(state=tk.DISABLED)
    
    def export_by_month(self):
        """Export cases from a specific month"""
        if not self.analysis_results:
            messagebox.showerror("Error", "Please analyze a workbook first.")
            return
        
        # Create selection dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Export by Specific Month")
        dialog.geometry("600x600")
        
        ttk.Label(dialog, text="Export Cases from Specific Month", font=("Arial", 11, "bold")).pack(pady=15)
        
        # Handler selection
        ttk.Label(dialog, text="Select Handler(s):", font=("Arial", 10)).pack(anchor=tk.W, padx=15)
        handler_frame = ttk.Frame(dialog)
        handler_frame.pack(fill=tk.X, padx=15, pady=5)
        
        handler_var = tk.StringVar()
        handlers_list = list(self.analysis_results.keys())
        handler_combo = ttk.Combobox(handler_frame, textvariable=handler_var, state="readonly",
                                     values=["All Handlers"] + handlers_list, width=50)
        handler_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        handler_combo.set("All Handlers")
        
        # Month selection
        ttk.Label(dialog, text="Select Month:", font=("Arial", 10)).pack(anchor=tk.W, padx=15)
        month_var = tk.StringVar()
        month_combo = ttk.Combobox(dialog, textvariable=month_var, state="readonly", width=50)
        month_combo.pack(fill=tk.X, padx=15, pady=5)
        
        # Output location
        ttk.Label(dialog, text="Output Filename:", font=("Arial", 10)).pack(anchor=tk.W, padx=15)
        output_var = tk.StringVar(value="exported_cases.xlsx")
        ttk.Entry(dialog, textvariable=output_var, width=50).pack(fill=tk.X, padx=15, pady=5)
        
        # Cleanup checkbox - Define EARLY so update_preview can use it
        cleanup_var = tk.BooleanVar(value=False)
        # Merged sheet option
        merged_var = tk.BooleanVar(value=False)
        # Merged sheet option
        merged_var = tk.BooleanVar(value=False)
        
        # Preview area
        ttk.Label(dialog, text="Preview:", font=("Arial", 10, "bold")).pack(anchor=tk.W, padx=15, pady=(10, 5))
        preview_text = tk.Text(dialog, height=8, width=70, state=tk.DISABLED)
        preview_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
        
        # Define update_preview FIRST before using it (after cleanup_var is defined)
        def update_preview(*args):
            handler = handler_var.get()
            month = month_var.get()
            cleanup = cleanup_var.get()
            
            preview_text.config(state=tk.NORMAL)
            preview_text.delete(1.0, tk.END)
            
            if not month:
                preview_text.insert(tk.END, "Select a month to see preview...")
                preview_text.config(state=tk.DISABLED)
                return
            
            handlers_to_process = [handler] if handler != "All Handlers" else list(self.analysis_results.keys())
            total_exported = 0
            details = []
            
            for h in handlers_to_process:
                count = self.analysis_results[h]['month_data'].get(month, 0)
                if count > 0:
                    total_exported += count
                    details.append(f"  • {h}: {count} cases")
            
            preview_text.insert(tk.END, f"Total cases to export: {total_exported}\n\n")
            if details:
                preview_text.insert(tk.END, "Cases per handler:\n")
                for detail in details:
                    preview_text.insert(tk.END, detail + "\n")
            
            if cleanup and total_exported > 0:
                preview_text.insert(tk.END, f"\n--- CLEANUP IMPACT ---\n")
                preview_text.insert(tk.END, f"Cases to delete from original: {total_exported}\n")
                
                remaining = 0
                for h in handlers_to_process:
                    total = self.analysis_results[h]['total_cases']
                    count = self.analysis_results[h]['month_data'].get(month, 0)
                    remaining += (total - count)
                
                preview_text.insert(tk.END, f"Cases remaining in original: {remaining}\n")
            
            preview_text.config(state=tk.DISABLED)
        
        # Cleanup checkbox
        cleanup_check = ttk.Checkbutton(dialog, text="Delete exported data from original file (creates cleaned backup)", 
                       variable=cleanup_var, command=update_preview)
        cleanup_check.pack(anchor=tk.W, padx=15, pady=10)
        # Merged sheet checkbox
        merged_check = ttk.Checkbutton(dialog, text="Also create merged sheet with all handlers concatenated", 
                variable=merged_var)
        merged_check.pack(anchor=tk.W, padx=15, pady=(0,10))
        # Merged sheet checkbox
        merged_check = ttk.Checkbutton(dialog, text="Also create merged sheet with all handlers concatenated", 
                variable=merged_var)
        merged_check.pack(anchor=tk.W, padx=15, pady=(0,10))
        
        def update_months(*args):
            handler = handler_var.get()
            all_months = set()
            if handler == "All Handlers":
                for h_data in self.analysis_results.values():
                    all_months.update(h_data['month_data'].keys())
            elif handler in self.analysis_results:
                all_months = set(self.analysis_results[handler]['month_data'].keys())
            
            months = sorted(list(all_months), reverse=True)
            month_combo['values'] = months
            update_preview()
        
        handler_var.trace_add('write', update_months)
        month_var.trace_add('write', update_preview)
        
        def export():
            handler = handler_var.get()
            month = month_var.get()
            output_file = output_var.get()
            cleanup = cleanup_var.get()
            merged = merged_var.get()
            merged = merged_var.get()
            
            if not handler or not month or not output_file:
                messagebox.showerror("Error", "Please fill all fields.")
                return
            
            # Ask user where to save the file
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_file
            )
            
            if not save_path:
                return  # User cancelled
            
            # Create progress dialog and run export in a background thread to keep UI responsive
            progress_dialog, pb, log_text = self._create_progress_dialog(title="Exporting by Month")
            t = threading.Thread(target=self._export_month_thread, args=(handler, month, save_path, cleanup, merged, progress_dialog, pb, log_text), daemon=True)
            t.start()
            try:
                dialog.destroy()
            except Exception:
                pass
        
        # Button frame
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=15)
        ttk.Button(button_frame, text="Export", command=export, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy, width=20).pack(side=tk.LEFT, padx=5)
    
    def export_old_cases(self):
        """Export cases older than a specified number of days"""
        if not self.analysis_results:
            messagebox.showerror("Error", "Please analyze a workbook first.")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Export Cases Older Than Specific Days")
        dialog.geometry("600x600")
        
        ttk.Label(dialog, text="Export Cases Older Than Specific Days", font=("Arial", 11, "bold")).pack(pady=15)
        
        # Handler selection
        ttk.Label(dialog, text="Select Handler(s):", font=("Arial", 10)).pack(anchor=tk.W, padx=15)
        handler_frame = ttk.Frame(dialog)
        handler_frame.pack(fill=tk.X, padx=15, pady=5)
        
        handler_var = tk.StringVar()
        handlers_list = list(self.analysis_results.keys())
        handler_combo = ttk.Combobox(handler_frame, textvariable=handler_var, state="readonly",
                                     values=["All Handlers"] + handlers_list, width=50)
        handler_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        handler_combo.set("All Handlers")
        
        # Days input
        ttk.Label(dialog, text="Number of Days (cases older than this):", font=("Arial", 10)).pack(anchor=tk.W, padx=15)
        days_var = tk.StringVar(value="30")
        ttk.Entry(dialog, textvariable=days_var, width=50).pack(fill=tk.X, padx=15, pady=5)
        
        # Output location
        ttk.Label(dialog, text="Output Filename:", font=("Arial", 10)).pack(anchor=tk.W, padx=15)
        output_var = tk.StringVar(value="old_cases.xlsx")
        ttk.Entry(dialog, textvariable=output_var, width=50).pack(fill=tk.X, padx=15, pady=5)
        
        # Cleanup checkbox - Define EARLY so update_preview can use it
        cleanup_var = tk.BooleanVar(value=False)
        
        # Preview area
        ttk.Label(dialog, text="Preview:", font=("Arial", 10, "bold")).pack(anchor=tk.W, padx=15, pady=(10, 5))
        preview_text = tk.Text(dialog, height=8, width=70, state=tk.DISABLED)
        preview_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
        
        # Define update_preview FIRST before using it (after cleanup_var is defined)
        def update_preview(*args):
            handler = handler_var.get()
            cleanup = cleanup_var.get()
            
            preview_text.config(state=tk.NORMAL)
            preview_text.delete(1.0, tk.END)
            
            try:
                days = int(days_var.get()) if days_var.get() else 0
            except ValueError:
                preview_text.insert(tk.END, "Enter a valid number of days...")
                preview_text.config(state=tk.DISABLED)
                return
            
            if days <= 0:
                preview_text.insert(tk.END, "Enter a valid number of days (> 0)...")
                preview_text.config(state=tk.DISABLED)
                return
            
            handlers_to_process = [handler] if handler != "All Handlers" else list(self.analysis_results.keys())
            cutoff_date = datetime.now() - timedelta(days=days)
            total_exported = 0
            details = []
            
            for h in handlers_to_process:
                h_data = self.analysis_results[h]
                count = sum(1 for d in h_data['dates'] if d['date'] < cutoff_date)
                if count > 0:
                    total_exported += count
                    details.append(f"  • {h}: {count} cases")
            
            preview_text.insert(tk.END, f"Cutoff date: {cutoff_date.strftime('%Y-%m-%d %H:%M:%S')}\n")
            preview_text.insert(tk.END, f"Total cases to export: {total_exported}\n\n")
            if details:
                preview_text.insert(tk.END, "Cases per handler:\n")
                for detail in details:
                    preview_text.insert(tk.END, detail + "\n")
            
            if cleanup and total_exported > 0:
                preview_text.insert(tk.END, f"\n--- CLEANUP IMPACT ---\n")
                preview_text.insert(tk.END, f"Cases to delete from original: {total_exported}\n")
                
                remaining = 0
                for h in handlers_to_process:
                    h_data = self.analysis_results[h]
                    total = h_data['total_cases']
                    count = sum(1 for d in h_data['dates'] if d['date'] < cutoff_date)
                    remaining += (total - count)
                
                preview_text.insert(tk.END, f"Cases remaining in original: {remaining}\n")
            
            preview_text.config(state=tk.DISABLED)
        
        # Cleanup checkbox
        cleanup_check = ttk.Checkbutton(dialog, text="Delete exported data from original file (creates cleaned backup)", 
                       variable=cleanup_var, command=update_preview)
        cleanup_check.pack(anchor=tk.W, padx=15, pady=10)
        
        days_var.trace_add('write', update_preview)
        handler_var.trace_add('write', update_preview)
        
        def export():
            handler = handler_var.get()
            output_file = output_var.get()
            cleanup = cleanup_var.get()
            merged = False  # merged option is not available in export_old_cases dialog
            
            if not handler or not output_file or not days_var.get():
                messagebox.showerror("Error", "Please fill all fields.")
                return
            
            # Ask user where to save the file
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=output_file
            )
            
            if not save_path:
                return  # User cancelled
            
            try:
                days = int(days_var.get())
            except ValueError:
                messagebox.showerror("Error", "Days must be a valid number.")
                return

            # Create progress dialog and run export in a background thread to keep UI responsive
            progress_dialog, pb, log_text = self._create_progress_dialog(title="Exporting Old Cases")
            t = threading.Thread(target=self._export_old_thread, args=(handler, days, save_path, cleanup, merged, progress_dialog, pb, log_text), daemon=True)
            t.start()
            try:
                dialog.destroy()
            except Exception:
                pass
        
        # Button frame
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=15)
        ttk.Button(button_frame, text="Export", command=export, width=20).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy, width=20).pack(side=tk.LEFT, padx=5)
    
    def export_data_by_month(self, handler, month, output_file, cleanup):
        """Export specific month data to Excel with all columns, optionally clean original"""
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb[handler]
        
        # Get metadata from analysis
        metadata = self.analysis_results[handler]
        header_row = metadata['header_row']
        status_col_index = metadata['status_col_index']
        
        # Create new workbook for export
        export_wb = openpyxl.Workbook()
        export_ws = export_wb.active
        export_ws.title = month
        
        # Copy entire header row with all columns
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            export_cell = export_ws.cell(row=1, column=col_idx)
            export_cell.value = cell.value
            self.copy_cell_style(cell, export_cell)
        
        # Copy matching rows with all columns
        export_row = 2
        rows_to_delete = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=header_row + 1):
            try:
                date_str = row[status_col_index - 1].value
                if date_str:
                    date_obj = self.parse_date(date_str)
                    if date_obj and date_obj.strftime("%Y-%m") == month:
                        # Copy all columns
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            export_cell = export_ws.cell(row=export_row, column=col_idx)
                            export_cell.value = cell.value
                            self.copy_cell_style(cell, export_cell)
                        export_row += 1
                        rows_to_delete.append(row_idx)
            except:
                continue
        
        # Save export file
        export_wb.save(output_file)
        
        # If cleanup is requested, remove exported rows from original and save
        if cleanup:
            # Delete rows in reverse order to maintain row indices
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx, 1)
            
            # Create backup of original file
            original_file = self.workbook_path
            backup_file = original_file.replace('.xlsx', '_backup.xlsx').replace('.xls', '_backup.xls')
            shutil.copy2(original_file, backup_file)
            
            # Save cleaned workbook
            wb.save(original_file)
    
    def export_data_by_month_all(self, month, output_file, cleanup, merged=False):
        """Export specific month data from all handlers with all columns, optionally clean original.
        If merged=True, also create an extra sheet that concatenates all handlers' exported rows.
        """
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Create new workbook for export with multiple sheets
        export_wb = openpyxl.Workbook()
        export_wb.remove(export_wb.active)  # Remove default sheet
        
        rows_to_delete_by_sheet = {}  # Track rows to delete per sheet
        merged_ws = None
        merged_export_row = 2
        merged_header_written = False
        if merged:
            merged_ws = export_wb.create_sheet(title=f"Merged - {month}")
        
        for handler in self.analysis_results.keys():
            ws = wb[handler]
            metadata = self.analysis_results[handler]
            header_row = metadata['header_row']
            status_col_index = metadata['status_col_index']
            
            # Create sheet in export workbook
            export_ws = export_wb.create_sheet(title=handler)
            
            # Copy entire header row with all columns
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                export_cell = export_ws.cell(row=1, column=col_idx)
                export_cell.value = cell.value
                self.copy_cell_style(cell, export_cell)
            
            # Copy matching rows with all columns
            export_row = 2
            rows_to_delete = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=header_row + 1):
                try:
                    date_str = row[status_col_index - 1].value
                    if date_str:
                        date_obj = self.parse_date(date_str)
                        if date_obj and date_obj.strftime("%Y-%m") == month:
                            # Copy all columns
                            for col_idx in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                export_cell = export_ws.cell(row=export_row, column=col_idx)
                                export_cell.value = cell.value
                                self.copy_cell_style(cell, export_cell)
                            export_row += 1
                            rows_to_delete.append(row_idx)
                            # Also append into merged sheet if requested
                            if merged and merged_ws is not None:
                                # Ensure header copied for merged sheet
                                if not merged_header_written:
                                    for m_col_idx in range(1, ws.max_column + 1):
                                        m_cell = ws.cell(row=header_row, column=m_col_idx)
                                        m_export_cell = merged_ws.cell(row=1, column=m_col_idx)
                                        m_export_cell.value = m_cell.value
                                        self.copy_cell_style(m_cell, m_export_cell)
                                    merged_header_written = True
                                    merged_export_row = 2
                                for m_col_idx in range(1, ws.max_column + 1):
                                    m_cell = ws.cell(row=row_idx, column=m_col_idx)
                                    m_export_cell = merged_ws.cell(row=merged_export_row, column=m_col_idx)
                                    m_export_cell.value = m_cell.value
                                    self.copy_cell_style(m_cell, m_export_cell)
                                merged_export_row += 1
                except:
                    continue
            
            if rows_to_delete:
                rows_to_delete_by_sheet[handler] = rows_to_delete
        
        # Save export file
        export_wb.save(output_file)
        
        # If cleanup is requested, remove exported rows from original and save
        if cleanup:
            for handler, rows_to_delete in rows_to_delete_by_sheet.items():
                ws = wb[handler]
                # Delete rows in reverse order to maintain row indices
                for row_idx in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row_idx, 1)
            
            # Create backup of original file
            original_file = self.workbook_path
            backup_file = original_file.replace('.xlsx', '_backup.xlsx').replace('.xls', '_backup.xls')
            shutil.copy2(original_file, backup_file)
            
            # Save cleaned workbook
            wb.save(original_file)
    
    def export_old_cases_data(self, handler, days, output_file, cleanup):
        """Export cases older than specified days with all columns, optionally clean original"""
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb[handler]
        
        # Get metadata from analysis
        metadata = self.analysis_results[handler]
        header_row = metadata['header_row']
        status_col_index = metadata['status_col_index']
        
        # Create new workbook for export
        export_wb = openpyxl.Workbook()
        export_ws = export_wb.active
        export_ws.title = "Old Cases"
        
        # Copy entire header row with all columns
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            export_cell = export_ws.cell(row=1, column=col_idx)
            export_cell.value = cell.value
            self.copy_cell_style(cell, export_cell)
        
        # Calculate cutoff date
        cutoff_date = datetime.now() - timedelta(days=days)
        
        # Copy matching rows with all columns
        export_row = 2
        rows_to_delete = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=header_row + 1):
            try:
                date_str = row[status_col_index - 1].value
                if date_str:
                    date_obj = self.parse_date(date_str)
                    if date_obj and date_obj < cutoff_date:
                        # Copy all columns
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            export_cell = export_ws.cell(row=export_row, column=col_idx)
                            export_cell.value = cell.value
                            self.copy_cell_style(cell, export_cell)
                        export_row += 1
                        rows_to_delete.append(row_idx)
            except:
                continue
        
        # Save export file
        export_wb.save(output_file)
        
        # If cleanup is requested, remove exported rows from original and save
        if cleanup:
            # Delete rows in reverse order to maintain row indices
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx, 1)
            
            # Create backup of original file
            original_file = self.workbook_path
            backup_file = original_file.replace('.xlsx', '_backup.xlsx').replace('.xls', '_backup.xls')
            shutil.copy2(original_file, backup_file)
            
            # Save cleaned workbook
            wb.save(original_file)
    
    def export_old_cases_data_all(self, days, output_file, cleanup, merged=False):
        """Export cases older than specified days from all handlers with all columns, optionally clean original.
        If merged=True, also create an extra sheet that concatenates all handlers' exported rows.
        """
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Create new workbook for export with multiple sheets
        export_wb = openpyxl.Workbook()
        export_wb.remove(export_wb.active)  # Remove default sheet
        
        cutoff_date = datetime.now() - timedelta(days=days)
        rows_to_delete_by_sheet = {}  # Track rows to delete per sheet
        merged_ws = None
        merged_export_row = 2
        merged_header_written = False
        if merged:
            merged_ws = export_wb.create_sheet(title=f"Merged - OlderThan-{days}d")
        
        for handler in self.analysis_results.keys():
            ws = wb[handler]
            metadata = self.analysis_results[handler]
            header_row = metadata['header_row']
            status_col_index = metadata['status_col_index']
            
            # Create sheet in export workbook
            export_ws = export_wb.create_sheet(title=handler)
            
            # Copy entire header row with all columns
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                export_cell = export_ws.cell(row=1, column=col_idx)
                export_cell.value = cell.value
                self.copy_cell_style(cell, export_cell)
            
            # Copy matching rows with all columns
            export_row = 2
            rows_to_delete = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=header_row + 1):
                try:
                    date_str = row[status_col_index - 1].value
                    if date_str:
                        date_obj = self.parse_date(date_str)
                        if date_obj and date_obj < cutoff_date:
                            # Copy all columns
                            for col_idx in range(1, ws.max_column + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                export_cell = export_ws.cell(row=export_row, column=col_idx)
                                export_cell.value = cell.value
                                self.copy_cell_style(cell, export_cell)
                            export_row += 1
                            rows_to_delete.append(row_idx)
                            if merged and merged_ws is not None:
                                # Ensure header copied for merged sheet
                                if not merged_header_written:
                                    for m_col_idx in range(1, ws.max_column + 1):
                                        m_cell = ws.cell(row=header_row, column=m_col_idx)
                                        m_export_cell = merged_ws.cell(row=1, column=m_col_idx)
                                        m_export_cell.value = m_cell.value
                                        self.copy_cell_style(m_cell, m_export_cell)
                                    merged_header_written = True
                                    merged_export_row = 2
                                for m_col_idx in range(1, ws.max_column + 1):
                                    m_cell = ws.cell(row=row_idx, column=m_col_idx)
                                    m_export_cell = merged_ws.cell(row=merged_export_row, column=m_col_idx)
                                    m_export_cell.value = m_cell.value
                                    self.copy_cell_style(m_cell, m_export_cell)
                                merged_export_row += 1
                except:
                    continue
            
            if rows_to_delete:
                rows_to_delete_by_sheet[handler] = rows_to_delete
        
        # Save export file
        export_wb.save(output_file)
        
        # If cleanup is requested, remove exported rows from original and save
        if cleanup:
            for handler, rows_to_delete in rows_to_delete_by_sheet.items():
                ws = wb[handler]
                # Delete rows in reverse order to maintain row indices
                for row_idx in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row_idx, 1)
            
            # Create backup of original file
            original_file = self.workbook_path
            backup_file = original_file.replace('.xlsx', '_backup.xlsx').replace('.xls', '_backup.xls')
            shutil.copy2(original_file, backup_file)
            
            # Save cleaned workbook
            wb.save(original_file)


def main():
    import sys as _sys
    _v2_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _v2_root not in _sys.path:
        _sys.path.insert(0, _v2_root)
    from utils.crash_handler import install_crash_handler, attach_tkinter_sigint_guard
    install_crash_handler()

    root = tk.Tk()
    attach_tkinter_sigint_guard(root)
    app = ExcelWorkbookAnalyzer(root)
    root.mainloop()

    # mainloop() returned — window was closed.
    # Exit immediately so background threads don't keep the process alive.
    os._exit(0)


if __name__ == "__main__":
    main()
