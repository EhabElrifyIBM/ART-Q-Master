import pandas as pd
import logging
from datetime import datetime
import os
import math
import re

# Helper: safe missing-value check to avoid pandas type-checker overload issues
def _is_missing(x):
    """Return True if x should be considered missing/empty.
    Avoid direct pd.isna(x) to keep static type-checkers happy when x may be Hashable.
    """
    if x is None:
        return True
    try:
        # float NaN
        if isinstance(x, float) and math.isnan(x):
            return True
    except Exception:
        pass
    if isinstance(x, str):
        return x.strip().lower() in ('', 'nan', 'none')
    try:
        # Detect pandas NA sentinel
        if x is pd.NA:
            return True
    except Exception:
        pass
    return False

try:
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Create dummy classes for when openpyxl is not available
    class DummyDataValidation:
        def __init__(self, *args, **kwargs):
            pass
        def add(self, *args, **kwargs):
            pass
        error = ""
        errorTitle = ""
        prompt = ""
        promptTitle = ""
    
    class DummyOpenpyxl:
        def load_workbook(self, *args, **kwargs):
            raise ImportError("openpyxl not available")
    
    DataValidation = DummyDataValidation
    openpyxl = DummyOpenpyxl()
    get_column_letter = lambda x: chr(64 + x)  # Simple fallback

class FinalProcessor:
    def __init__(self):
        # Configure logger (use global helpers.setup_logging where possible)
        self.logger = logging.getLogger(__name__)
        self.validation_lists = {
            "Action 1": ["Sent SMS", "Handled by Case Owner", "Handled by DRO", "Handled by Chat", "Bank/Sutherland", "DND"],
            "Action 2": ["Sent Email", "Sent Email in another Case", "Bank/Sutherland", "DND"],
            "Action 3": ["Called the Customer", "Called in another Case", "Bank/Sutherland", "DND"],
            "Final Action": ["Fixed", "Refused Callback", "Issue Not Fixed", "Not yet Tested", "Escalation", "Sent SMS", "Sent Email", "Left VM", "Reviewed", "Bank/Sutherland", "Not Reached", "DND"],
            "Status": ["in_progress", "new", "closed", "Skipped", "In Progress Today"]
        }
        
        # Column order for Companies sheet - as specified by user
        self.companies_column_order = [
            'Case Number', 'Customer Name', 'Company Name', 'DND (Do Not Disturb)', 
            'Email', 'Phone Number', 'Work Order Type', 'Incoming Channel', 
            'Last Status Change', 'Country', 'State/Province', 'Local Time', 
            'Action 1', 'Action 2', 'Action 3', 'Final Action', 'Assigned To', 
            'Status', 'Unnamed: 18', 'Case', 'Problem Description', 
            'Case Status Updated', 'Case Reason', 'Work Order ID', 'Work Order Status', 
            'Order Type', 'Work Order Priority', 'Product ID (MTM)', 'Machine Type', 
            'Product Description', 'Serial Number', 'Created On', 'Survey Preference', 
            'Survey Fatigue', 'No survey reason', 'Program', 'Repeat Frequency', 
            'Repeat Repair', 'Closing Code', 'Reported Symptom', 'Case Status'
        ]

    def _resolve_col(self, df, preferred_name):
        """Resolve a preferred column name to an existing column in df.
        - If exact match (case-insensitive) exists, return it.
        - Otherwise, return the first column that contains the main token (e.g. 'customer') in its name.
        - If nothing found, return None.
        """
        if df is None:
            return None
        pref = preferred_name.strip().lower()
        # Exact match (case-insensitive)
        for col in df.columns:
            try:
                if str(col).strip().lower() == pref:
                    return col
            except Exception:
                continue
        # Token match (use the first word as token)
        token = pref.split()[0]
        for col in df.columns:
            try:
                if token in str(col).strip().lower():
                    return col
            except Exception:
                continue
        return None

    def process_final_output(self, output_df, output_file, processing_stats=None, sms_file=None, email_file=None, prev_file=None, duplicate_company_cases=None, selected_handlers=None):
        try:
            self.logger.info("\n=== Starting Final Processing ===")
            self.logger.info(f"Processing final output with {len(output_df)} records")
            
            # Use the previous input file (prev_file) to load Issue Not Fixed and DND Emails
            prev_output_file = None
            
            self.logger.info(f"=== Using prev_file to load Issue Not Fixed and DND Emails ===")
            self.logger.info(f"Previous file provided: {prev_file}")
            
            if prev_file and os.path.exists(prev_file):
                # Check if this file has the expected sheets
                self.logger.info(f"Checking prev_file for Issue Not Fixed and DND Emails sheets: {prev_file}")
                try:
                    with pd.ExcelFile(prev_file) as excel:
                        sheet_names = excel.sheet_names
                        self.logger.info(f"prev_file has these sheets: {sheet_names}")
                        
                        if 'Issue Not Fixed' in sheet_names or 'DND Emails' in sheet_names:
                            # This file contains the sheets we need
                            prev_output_file = prev_file
                            self.logger.info(f" Setting prev_output_file to: {prev_output_file}")
                            
                            # Verify what will be loaded and log the counts
                            if 'Issue Not Fixed' in sheet_names:
                                prev_issue = pd.read_excel(prev_file, sheet_name='Issue Not Fixed')
                                self.logger.info(f"  Issue Not Fixed: {len(prev_issue)} cases will be loaded")
                                if len(prev_issue) > 0:
                                    self.logger.info(f"    Sample cases: {prev_issue[['Case Number']].head(3).to_dict('records')}")
                            if 'DND Emails' in sheet_names:
                                prev_dnd = pd.read_excel(prev_file, sheet_name='DND Emails')
                                self.logger.info(f"  DND Emails: {len(prev_dnd)} emails will be loaded")
                        else:
                            self.logger.info(f"prev_file does not contain Issue Not Fixed or DND Emails sheets")
                            self.logger.info(f"This is likely a temp handler file - will search for actual output file")
                except Exception as e:
                    self.logger.warning(f"Could not check prev_file: {str(e)}")
            
            # If prev_file doesn't have the sheets, search for actual output file in the same directory
            if not prev_output_file:
                self.logger.info(f"Searching for actual previous output file in directory...")
                output_dir = os.path.dirname(output_file) or os.path.dirname(prev_file) if prev_file else None
                
                if output_dir and os.path.exists(output_dir):
                    try:
                        # Look for files in the same directory
                        possible_files = []
                        for file in os.listdir(output_dir):
                            if file.endswith('.xlsx') and file != os.path.basename(output_file) if output_file else file:
                                file_path = os.path.join(output_dir, file)
                                # Skip temp handler files
                                if '_prev_handlers' not in file:
                                    try:
                                        with pd.ExcelFile(file_path) as excel:
                                            if 'Issue Not Fixed' in excel.sheet_names or 'DND Emails' in excel.sheet_names:
                                                mod_time = os.path.getmtime(file_path)
                                                possible_files.append((file_path, mod_time, file))
                                                self.logger.info(f"Found candidate: {file}")
                                    except Exception:
                                        continue
                        
                        if possible_files:
                            # Sort by modification time and use the most recent
                            possible_files.sort(key=lambda x: x[1], reverse=True)
                            prev_output_file = possible_files[0][0]
                            self.logger.info(f"✓ Selected previous output file: {prev_output_file}")
                            
                            # Load and log counts
                            try:
                                with pd.ExcelFile(prev_output_file) as excel:
                                    if 'Issue Not Fixed' in excel.sheet_names:
                                        prev_issue = pd.read_excel(prev_output_file, sheet_name='Issue Not Fixed')
                                        self.logger.info(f"  Issue Not Fixed: {len(prev_issue)} cases will be loaded")
                                        if len(prev_issue) > 0:
                                            self.logger.info(f"    Sample: {prev_issue['Case Number'].head(3).tolist()}")
                                    if 'DND Emails' in excel.sheet_names:
                                        prev_dnd = pd.read_excel(prev_output_file, sheet_name='DND Emails')
                                        self.logger.info(f"  DND Emails: {len(prev_dnd)} emails will be loaded")
                            except Exception as e:
                                self.logger.warning(f"Could not verify contents: {str(e)}")
                        else:
                            self.logger.info("No previous output files found in directory")
                    except Exception as e:
                        self.logger.warning(f"Could not search directory: {str(e)}")
            
            # Final log
            if prev_output_file:
                self.logger.info(f"=== Will load Issue Not Fixed and DND Emails from: {prev_output_file} ===")
            else:
                self.logger.info(f"=== No previous file with Issue Not Fixed/DND Emails found - starting fresh ===")
            
            # STEP 1: Load and consolidate ALL previous handler sheets FIRST
            consolidated_prev_data, issue_not_resolved_cases = None, []
            if prev_file and os.path.exists(prev_file):
                try:
                    self.logger.info("=== STEP 1: LOADING AND CONSOLIDATING PREVIOUS HANDLER SHEETS ===")
                    consolidated_prev_data, issue_not_resolved_cases = self.load_and_consolidate_all_previous_handler_sheets(prev_file)
                    if consolidated_prev_data is not None and not consolidated_prev_data.empty:
                        self.logger.info(f"Successfully consolidated {len(consolidated_prev_data)} previous cases from all handler sheets")
                        self.logger.info(f"  Previous handlers found: {consolidated_prev_data['Assigned To'].unique().tolist()}")
                    else:
                        self.logger.info("No previous handler data found - starting fresh")
                except Exception as e:
                    self.logger.error(f"Error loading previous handler data: {str(e)}")
                    self.logger.warning("Continuing without previous data")
                    consolidated_prev_data = None

            # STEP 2: Merge current data with consolidated previous data
            if consolidated_prev_data is not None and not consolidated_prev_data.empty:
                self.logger.info("=== STEP 2: MERGING CURRENT DATA WITH CONSOLIDATED PREVIOUS DATA ===")
                # Merge current output_df with consolidated previous data
                merged_output_df = self.merge_current_with_consolidated_previous(output_df, consolidated_prev_data)
                self.logger.info(f"Merged data: {len(output_df)} current + {len(consolidated_prev_data)} previous = {len(merged_output_df)} total")
                output_df = merged_output_df  # Use merged data for all subsequent processing
            else:
                self.logger.info("No previous data to merge - using current data only")

            # STEP 3: Apply business rules (only specific updates, preserve handler work)
            self.logger.info("=== STEP 3: APPLYING BUSINESS RULES (PRESERVING HANDLER WORK) ===")
            output_df = self.apply_business_rules_preserving_handler_work(output_df)

            # STEP 3.5: Sort the main output DataFrame by status
            self.logger.info("=== STEP 3.5: SORTING MAIN OUTPUT BY STATUS ===")
            output_df = self.sort_cases_by_status(output_df)
            
            # STEP 3.6: Add preservation columns to main output for case tracking
            # Removed adding preservation columns; preservation will be handled during per-handler merging
            
            # Create sorting summary report
            self.create_sorting_summary_report(output_df, "PA Cases Main Sheet")

            # Create Excel file with xlsxwriter
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                # STEP 4: Create sheets in the CORRECT ORDER
                self.logger.info("=== STEP 4: CREATING SHEETS IN CORRECT ORDER ===")
                
                # 4.1: Create individual handler sheets FIRST (each handler gets their own sheet)
                self.logger.info("=== 4.1: CREATING INDIVIDUAL HANDLER SHEETS ===")
                self.create_handler_sheets_from_processed_data(writer, output_df, prev_file=prev_file)

                # 4.1.5: Create Companies sheet (email duplicates - preserved across runs)
                # Positioned right after handler sheets as requested
                self.logger.info("=== 4.1.5: CREATING COMPANIES SHEET ===")
                companies_df = self.create_companies_sheet_with_preservation(
                    duplicate_company_cases, 
                    prev_file, 
                    output_df,
                    selected_handlers=selected_handlers
                )
                
                if companies_df is not None and not companies_df.empty:
                    companies_df.to_excel(writer, sheet_name='Companies', index=False)
                    self.auto_adjust_columns(writer, companies_df, 'Companies')
                    # Lock the sheet
                    self.protect_worksheet(writer, 'Companies', password='artadmin')
                    self.logger.info(f"✓ Created Companies sheet with {len(companies_df)} total cases")
                    
                    # CRITICAL: Sync Companies assignments back to PA Cases output_df
                    # This ensures company cases show their assigned handler in PA Cases sheet too
                    if 'Case Number' in companies_df.columns and 'Assigned To' in companies_df.columns:
                        self.logger.info("Syncing Companies assignments to PA Cases...")
                        companies_assignments = {}
                        for _, row in companies_df.iterrows():
                            case_num = row['Case Number']
                            handler = row.get('Assigned To', '')
                            if pd.notna(case_num) and handler and str(handler).strip():
                                # Normalize case number for matching
                                try:
                                    norm_case = int(float(str(case_num).strip()))
                                    companies_assignments[norm_case] = handler
                                except:
                                    pass
                        
                        # Update output_df with Companies assignments
                        if companies_assignments:
                            output_df['_temp_case_num'] = output_df['Case Number'].apply(
                                lambda x: int(float(str(x).strip())) if pd.notna(x) and str(x).strip() else None
                            )
                            updated_count = 0
                            for case_num, handler in companies_assignments.items():
                                mask = output_df['_temp_case_num'] == case_num
                                if mask.any():
                                    output_df.loc[mask, 'Assigned To'] = handler
                                    updated_count += 1
                            output_df = output_df.drop(columns=['_temp_case_num'])
                            self.logger.info(f"✓ Synced {updated_count} company case assignments to PA Cases")
                else:
                    self.logger.info("No email duplicate cases found - skipping Companies sheet")
                
                # 4.2: Create Counter sheet
                self.logger.info("=== 4.2: CREATING COUNTER SHEET ===")
                self.create_counters_sheet(writer, output_df, prev_file=prev_file)
                
                # 4.3: Create Summary sheet
                self.logger.info("=== 4.3: CREATING SUMMARY SHEET ===")
                self.create_summary_sheet(writer, output_df, processing_stats)
                
                # 4.4: Create Issue Not Fixed sheet
                self.logger.info("=== 4.4: CREATING ISSUE NOT FIXED SHEET ===")
                self.logger.info(f"Passing prev_output_file to Issue Not Fixed: {prev_output_file}")
                if prev_output_file:
                    self.logger.info(f"  ✓ prev_output_file is set and will be used to load previous data")
                else:
                    self.logger.warning(f"  ✗ prev_output_file is None - no previous data will be loaded!")
                self.create_sms_replies_sheet(writer, output_df, sms_file, processing_stats, prev_output_file, issue_not_resolved_cases)
                
                # 4.4.1: Create Skipped SMS Cases sheet
                self.logger.info("=== 4.4.1: CREATING SKIPPED SMS CASES SHEET ===")
                self.create_skipped_sms_sheet(writer, processing_stats)
                
                # 4.4.2: Create Skipped Email Replies sheet
                self.logger.info("=== 4.4.2: CREATING SKIPPED EMAIL REPLIES SHEET ===")
                # Pass email_file and current output_df so the skipped-email sheet can be built
                self.create_skipped_email_sheet(writer, processing_stats, email_file=email_file, output_df=output_df)
                
                # 4.5: Create DND Emails Database sheet
                self.logger.info("=== 4.5: CREATING DND EMAILS DATABASE SHEET ===")
                self.logger.info(f"Passing prev_output_file to DND Emails: {prev_output_file}")
                if prev_output_file:
                    self.logger.info(f"  ✓ prev_output_file is set and will be used to load previous data")
                else:
                    self.logger.warning(f"  ✗ prev_output_file is None - no previous data will be loaded!")
                self.create_dnd_emails_database_sheet(writer, output_df, email_file, processing_stats, prev_file=prev_output_file)
                
                # 4.6: Create Validation sheet (empty for now, will be populated later)
                self.logger.info("=== 4.6: CREATING VALIDATION SHEET ===")
                validation_df = pd.DataFrame(columns=['Validation'])
                validation_df.to_excel(writer, sheet_name='Validation', index=False)
                
                # Auto-adjust columns for Validation sheet
                self.auto_adjust_columns(writer, validation_df, 'Validation')
                

                
                # 4.7: Create PA Cases main sheet LAST (already sorted by status)
                self.logger.info("=== 4.7: CREATING PA CASES MAIN SHEET ===")
                
                # Format date columns to show only dates (remove timestamps)
                # Convert to string format with date only
                date_columns = ['Last Status Change', 'Created On', 'Completion Date']
                for col in date_columns:
                    if col in output_df.columns:
                        try:
                            # Debug: show before
                            sample_before = output_df[col].iloc[0] if len(output_df) > 0 else None
                            before_str = str(sample_before)
                            print(f"CONSOLE DEBUG {col} BEFORE: {repr(sample_before)} -> str: {before_str}")
                            self.logger.info(f"DEBUG {col} BEFORE: {repr(sample_before)}")
                            
                            def extract_date_str(val):
                                """Extract date part as clean string - handles both strings and datetime objects"""
                                if pd.isna(val) or val == '' or val == 'NaT':
                                    return ''
                                
                                # Handle pandas Timestamp and datetime objects
                                if hasattr(val, 'date'):  # datetime, Timestamp, date objects have .date() method
                                    return str(val.date())  # Returns YYYY-MM-DD
                                
                                # Handle strings
                                val_str = str(val).strip()
                                if not val_str or val_str.lower() in ('nan', 'nat', 'none'):
                                    return ''
                                # If it contains a space, extract date part only
                                if ' ' in val_str:
                                    date_only = val_str.split(' ')[0]
                                    return date_only
                                return val_str
                            
                            output_df[col] = output_df[col].apply(extract_date_str)
                            
                            # Debug: show after
                            sample_after = output_df[col].iloc[0] if len(output_df) > 0 else None
                            self.logger.info(f"DEBUG {col} AFTER: {repr(sample_after)}")
                            self.logger.info(f"✓ Converted {col} to date-only strings")
                        except Exception as e:
                            self.logger.warning(f"Could not format {col}: {str(e)}")
                
                output_df.to_excel(writer, sheet_name='PA Cases', index=False)
                
                # Auto-adjust columns for PA Cases sheet
                self.auto_adjust_columns(writer, output_df, 'PA Cases')
            
            # STEP 5: Add validation dropdowns using openpyxl (this will populate the 'Validation' sheet)
            self.add_validation_dropdowns(output_file)
            
            # STEP 6: Final date formatting after validation dropdowns - ALL SHEETS
            self.logger.info("=== STEP 6: FINAL DATE FORMATTING AFTER VALIDATION (ALL SHEETS) ===")
            try:
                from openpyxl import load_workbook
                from datetime import datetime
                
                wb = load_workbook(output_file)
                
                date_columns_to_fix = ['Last Status Change', 'Created On', 'Completion Date']
                
                # Apply formatting to all sheets
                for sheet_name in wb.sheetnames:
                    try:
                        ws = wb[sheet_name]
                        
                        # Skip non-data sheets
                        if sheet_name in ['Validation', 'Counter', 'Summary']:
                            continue
                        
                        # Get header row
                        header = [cell.value for cell in ws[1]]
                        
                        # Format each date column that exists in this sheet
                        for col_name in date_columns_to_fix:
                            if col_name in header:
                                col_idx = header.index(col_name) + 1
                                cell_count = 0
                                # Process each cell - convert datetime to date-only string
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row, col_idx)
                                    if cell.value:
                                        try:
                                            # Convert to datetime first
                                            if isinstance(cell.value, str):
                                                # Parse the string
                                                dt = pd.to_datetime(cell.value, errors='coerce')
                                                if pd.notna(dt):
                                                    date_str = dt.strftime('%Y-%m-%d')
                                                    cell.value = date_str
                                                    cell.number_format = '@'  # Text format
                                                    cell_count += 1
                                            elif hasattr(cell.value, 'date'):  # datetime object
                                                # Extract just the date
                                                date_str = cell.value.date().isoformat()  # Returns YYYY-MM-DD
                                                cell.value = date_str
                                                cell.number_format = '@'  # Text format
                                                cell_count += 1
                                        except:
                                            pass  # Skip cells that can't be converted
                                
                                if cell_count > 0:
                                    self.logger.info(f"✓ Sheet '{sheet_name}': Fixed {cell_count} cells in '{col_name}'")
                    except Exception as e:
                        self.logger.warning(f"Could not format sheet '{sheet_name}': {str(e)}")
                        continue
                
                wb.save(output_file)
                self.logger.info("✓ Saved Excel file with timestamp-free dates across all sheets")
            except Exception as e:
                self.logger.warning(f"Could not apply date formatting to all sheets: {str(e)}")
            
            # STEP 7: Delete temporary _prev_handlers file if it exists
            if prev_file and '_prev_handlers' in prev_file and os.path.exists(prev_file):
                try:
                    os.remove(prev_file)
                    self.logger.info(f"✓ Deleted temporary handler file: {prev_file}")
                except Exception as e:
                    self.logger.warning(f"Could not delete temporary file {prev_file}: {str(e)}")
            
            self.logger.info(f"Final processing completed. File saved to: {output_file}")
            return True, "Final processing completed successfully"
        except Exception as e:
            self.logger.error(f"Error in final processing: {str(e)}")
            return False, f"Error during final processing: {str(e)}"

    def load_and_consolidate_all_previous_handler_sheets(self, prev_file):
        """Load and consolidate ALL previous handler sheets into one dataset for processing"""
        try:
            self.logger.info(f"Loading and consolidating all handler sheets from: {prev_file}")
            
            # Read the previous file
            prev_excel = pd.ExcelFile(prev_file)
            all_sheets = prev_excel.sheet_names
            self.logger.info(f"Found sheets in previous file: {all_sheets}")
            
            # Look for handler sheets (sheets ending with "'s Cases") - guard non-string sheet names
            handler_sheets = [sheet for sheet in all_sheets if isinstance(sheet, str) and sheet.endswith("'s Cases")]
            self.logger.info(f"Found handler sheets: {handler_sheets}")
            
            if not handler_sheets:
                self.logger.info("No handler sheets found in previous file")
                return None, []
            
            # Consolidate all handler sheets into one DataFrame
            consolidated_data = []
            total_cases = 0
            issue_not_resolved_cases = []
            
            for sheet_name in handler_sheets:
                try:
                    # Extract handler name from sheet name
                    handler_name = sheet_name.replace("'s Cases", "")
                    
                    # Read the handler sheet
                    handler_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                    
                    if not handler_df.empty and 'Case Number' in handler_df.columns:
                        # Optional: any health checks can remain here
                        # Check for cases marked as "Issue Not Resolved" in Final Action
                        if 'Final Action' in handler_df.columns:
                            issue_mask = handler_df['Final Action'].astype(str).str.strip().str.lower().isin([
                                'issue not resolved', 'issue not fixed', 'not resolved', 'unresolved'
                            ])
                            issue_cases = handler_df[issue_mask]
                            
                            if not issue_cases.empty:
                                for idx, row in issue_cases.iterrows():
                                    issue_case = {
                                        'Case Number': row['Case Number'],
                                        'Handler': handler_name,
                                        'Source Sheet': sheet_name,
                                        'Final Action': row['Final Action'],
                                        'Status': row.get('Status', ''),
                                        'Assigned To': row.get('Assigned To', ''),
                                        'Date Added': pd.Timestamp.now().strftime('%Y-%m-%d'),
                                        'Source': 'Handler Manual Update'
                                    }
                                    issue_not_resolved_cases.append(issue_case)
                                
                                self.logger.info(f"Found {len(issue_cases)} cases marked as 'Issue Not Resolved' in {sheet_name}")
                        
                        consolidated_data.append(handler_df)
                        total_cases += len(handler_df)
                        self.logger.info(f"Loaded handler sheet '{sheet_name}': {len(handler_df)} cases")
                        
                        # Log some sample case numbers for verification
                        if len(handler_df) > 0:
                            sample_cases = handler_df['Case Number'].head(3).tolist()
                            self.logger.info(f"  Sample case numbers: {sample_cases}")
                    else:
                        self.logger.warning(f"Handler sheet '{sheet_name}' is empty or missing Case Number column")
                        
                except Exception as e:
                    self.logger.warning(f"Error loading handler sheet '{sheet_name}': {str(e)}")
                    continue
            
            if not consolidated_data:
                self.logger.warning("No valid handler data could be loaded")
                return None, []
            
            # Combine all handler data into one DataFrame
            if len(consolidated_data) == 1:
                final_consolidated = consolidated_data[0]
            else:
                final_consolidated = pd.concat(consolidated_data, ignore_index=True)
            
            self.logger.info(f"Successfully consolidated {len(handler_sheets)} handler sheets with {total_cases} total cases")
            self.logger.info(f"Found {len(issue_not_resolved_cases)} cases marked as 'Issue Not Resolved' across all handler sheets")
            self.logger.info(f"Consolidated data columns: {final_consolidated.columns.tolist()}")
            
            # Additional verification logging
            self.logger.info(f"Final consolidated DataFrame shape: {final_consolidated.shape}")
            self.logger.info(f"Case Number column unique values: {final_consolidated['Case Number'].nunique()}")
            
            return final_consolidated, issue_not_resolved_cases
            
        except Exception as e:
            self.logger.error(f"Error consolidating previous handler sheets: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None, []

    def create_companies_sheet_with_preservation(self, new_duplicates, prev_file, current_df, selected_handlers=None):
        """
        Create- [/] **Companies Sheet Logic** <!-- id: 11 -->
    - [x] Update `processor.py` to pass `selected_handlers` <!-- id: 12 -->
    - [x] Implement Sorting Logic in `final_processor.py` (Group matching cases) <!-- id: 13 -->
    - [x] Implement Rounds Logic (Assignment by email) in `final_processor.py` <!-- id: 14 -->
    - [x] Implement Locking Logic (Password: 'artadmin', View-only) in `final_processor.py` <!-- id: 15 -->Email groups
        
        Args:
            new_duplicates: DataFrame of new cases with duplicate emails
            prev_file: Path to previous output file (for loading existing Companies data)
            current_df: Current processed DataFrame (for applying business rules)
            selected_handlers: List of handlers to assign cases to
        
        Returns:
            DataFrame ready for Companies sheet with proper column order
        """
        try:
            self.logger.info("Creating Companies sheet with preservation...")
            
            # Step 1: Load previous Companies sheet
            prev_companies = self._load_previous_companies_sheet(prev_file)
            
            # Step 2: Merge previous with new duplicates
            if new_duplicates is None or new_duplicates.empty:
                new_duplicates = pd.DataFrame()
            
            companies_df = self._merge_companies_data(new_duplicates, prev_companies, prev_file)
            
            if companies_df.empty:
                self.logger.info("No company cases to include in Companies sheet")
                return pd.DataFrame()
            
            # Step 3: Apply business rules (same as handler processing) + DND rules
            companies_df = self._apply_business_rules_to_companies(companies_df, prev_file)
            
            # Step 4: Assign cases to handlers (Round Robin by Email)
            companies_df = self._assign_companies_cases(companies_df, selected_handlers)
            
            # Step 5: Reorder columns to match specified order
            companies_df = self._reorder_companies_columns(companies_df)
            
            # Step 6: Sort: New first, then others
            companies_df = self._sort_companies_sheet(companies_df)
            
            self.logger.info(f"Companies sheet prepared with {len(companies_df)} total cases")
            return companies_df
            
        except Exception as e:
            self.logger.error(f"Error creating Companies sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            # Return new duplicates if available, otherwise empty
            return new_duplicates if new_duplicates is not None else pd.DataFrame()

    def _load_previous_companies_sheet(self, prev_file):
        """Load previous Companies sheet data for preservation"""
        try:
            if not prev_file or not os.path.exists(prev_file):
                self.logger.info("No previous file provided - starting fresh Companies sheet")
                return pd.DataFrame()
            
            with pd.ExcelFile(prev_file) as excel:
                if 'Companies' in excel.sheet_names:
                    prev_companies = pd.read_excel(prev_file, sheet_name='Companies')
                    self.logger.info(f"Loaded {len(prev_companies)} cases from previous Companies sheet")
                    
                    # Log sample case numbers for verification
                    if len(prev_companies) > 0 and 'Case Number' in prev_companies.columns:
                        sample = prev_companies['Case Number'].head(3).tolist()
                        self.logger.info(f"  Sample preserved cases: {sample}")
                    
                    return prev_companies
                else:
                    self.logger.info("No Companies sheet in previous file - starting fresh")
                    return pd.DataFrame()
                    
        except Exception as e:
            self.logger.warning(f"Could not load previous Companies sheet: {str(e)}")
            return pd.DataFrame()

    def _merge_companies_data(self, new_duplicates, prev_companies, prev_file=None):
        """
        Merge new email duplicate cases with preserved Companies data.
        
        CRITICAL: PREVIOUS DATA WINS FOR OVERLAPPING CASES
        - Overlapping cases (same Case Number): KEEP previous data (actions, status, handler preserved)
        - Non-overlapping previous cases: PRESERVE them as-is
        - Truly new cases (not in previous): ADD them for assignment
        
        Adds '_is_preserved' flag to identify which cases should skip business rules.
        """
        try:
            self.logger.info("=== MERGING COMPANIES DATA (Previous Data Preserved) ===")
            
            # Normalize case numbers for comparison
            def normalize_case_num(val):
                if pd.isna(val) or str(val).strip() == '':
                    return None
                try:
                    return int(float(str(val).strip()))
                except:
                    return None
            
            # Handle empty cases
            if prev_companies.empty and (new_duplicates is None or new_duplicates.empty):
                self.logger.info("No previous Companies data and no new duplicates - returning empty")
                return pd.DataFrame()
            
            if new_duplicates is None or new_duplicates.empty:
                self.logger.info(f"No new duplicates - preserving ALL {len(prev_companies)} previous Companies cases")
                result = prev_companies.copy()
                result['_is_preserved'] = True  # Mark all as preserved
                return result
            
            if prev_companies.empty:
                self.logger.info(f"No previous Companies data - using {len(new_duplicates)} new cases only")
                result = new_duplicates.copy()
                result['_is_preserved'] = False  # Mark all as new
                return result
            
            # STEP 1: Normalize case numbers in BOTH dataframes
            prev_companies_copy = prev_companies.copy()
            new_duplicates_copy = new_duplicates.copy()
            
            prev_companies_copy['_normalized_cn'] = prev_companies_copy['Case Number'].apply(normalize_case_num)
            new_duplicates_copy['_normalized_cn'] = new_duplicates_copy['Case Number'].apply(normalize_case_num)
            
            # STEP 2: Get case number sets
            prev_case_numbers = set(prev_companies_copy['_normalized_cn'].dropna())
            new_case_numbers = set(new_duplicates_copy['_normalized_cn'].dropna())
            
            self.logger.info(f"Previous Companies cases: {len(prev_case_numbers)}")
            self.logger.info(f"New duplicate cases: {len(new_case_numbers)}")
            
            # STEP 3: Identify overlapping and truly new
            overlapping = prev_case_numbers & new_case_numbers
            truly_new = new_case_numbers - prev_case_numbers
            
            self.logger.info(f"  Overlapping (KEEP PREVIOUS data): {len(overlapping)}")
            self.logger.info(f"  Truly new (add from new): {len(truly_new)}")
            
            # STEP 4: Build the merged result - PREVIOUS DATA WINS
            # Keep ALL previous cases (they have actions, status, handler preserved)
            preserved_cases = prev_companies_copy.copy()
            preserved_cases['_is_preserved'] = True  # Mark as preserved - skip business rules
            
            # Get only TRULY NEW cases (not in previous at all)
            truly_new_cases = new_duplicates_copy[~new_duplicates_copy['_normalized_cn'].isin(prev_case_numbers)].copy()
            truly_new_cases['_is_preserved'] = False  # Mark as new - apply business rules
            
            self.logger.info(f"Final counts:")
            self.logger.info(f"  Preserved cases (ALL previous): {len(preserved_cases)}")
            self.logger.info(f"  Truly new cases (to add): {len(truly_new_cases)}")
            
            # Drop temp columns
            if '_normalized_cn' in preserved_cases.columns:
                preserved_cases = preserved_cases.drop(columns=['_normalized_cn'])
            if '_normalized_cn' in truly_new_cases.columns:
                truly_new_cases = truly_new_cases.drop(columns=['_normalized_cn'])
            
            # STEP 5: Combine - PREVIOUS FIRST (preserved with all data), then truly new
            if not truly_new_cases.empty:
                combined = pd.concat([preserved_cases, truly_new_cases], ignore_index=True)
                self.logger.info(f"Merged Companies: {len(preserved_cases)} preserved + {len(truly_new_cases)} truly new = {len(combined)} total")
            else:
                combined = preserved_cases
                self.logger.info(f"No truly new cases - keeping {len(combined)} preserved cases only")
            
            # STEP 6: Remove any duplicates (by Case Number) - keep first (previous has priority)
            if 'Case Number' in combined.columns:
                pre_dedup = len(combined)
                combined = combined.drop_duplicates(subset=['Case Number'], keep='first')
                post_dedup = len(combined)
                if pre_dedup > post_dedup:
                    self.logger.info(f"Removed {pre_dedup - post_dedup} duplicate entries")
            
            self.logger.info(f"=== COMPANIES MERGE COMPLETE: {len(combined)} total cases ===")
            return combined
            
        except Exception as e:
            self.logger.error(f"Error merging companies data: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            # Return non-empty data if available
            if prev_companies is not None and not prev_companies.empty:
                return prev_companies
            return new_duplicates if new_duplicates is not None and not new_duplicates.empty else pd.DataFrame()

    def _apply_business_rules_to_companies(self, df, prev_file=None):
        """Apply business rules to Companies sheet data, including DND checks
        
        IMPORTANT: Skip business rules for preserved cases (_is_preserved=True)
        Only apply rules to truly new cases to avoid overwriting existing data.
        """
        try:
            if df.empty:
                return df
            
            df = df.copy()
            self.logger.info("Applying business rules to Companies sheet...")
            
            # Check for _is_preserved flag
            has_preserved_flag = '_is_preserved' in df.columns
            if has_preserved_flag:
                preserved_count = df['_is_preserved'].sum() if df['_is_preserved'].dtype == bool else (df['_is_preserved'] == True).sum()
                new_count = len(df) - preserved_count
                self.logger.info(f"  Preserved cases (skipping rules): {preserved_count}, New cases (applying rules): {new_count}")
                # Create mask for NEW cases only (where we should apply rules)
                new_cases_mask = ~df['_is_preserved'].fillna(False).astype(bool)
            else:
                new_cases_mask = pd.Series([True] * len(df), index=df.index)
                self.logger.info("  No preservation flag found - applying rules to all cases")
            
            # Rule 1: Convert "in progress today" to "in_progress" (apply to ALL)
            if 'Status' in df.columns:
                in_progress_mask = df['Status'].astype(str).str.strip().str.lower() == 'in progress today'
                count = in_progress_mask.sum()
                if count > 0:
                    df.loc[in_progress_mask, 'Status'] = 'in_progress'
                    self.logger.info(f"  Updated {count} cases from 'in progress today' to 'in_progress'")
            
            # Rule 2: Apply Bank/Sutherland rule ONLY TO NEW CASES
            if 'Status' in df.columns and 'Action 1' in df.columns:
                bank_mask = (
                    new_cases_mask &  # Only new cases!
                    (df['Status'].astype(str).str.lower() == 'closed') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                count = bank_mask.sum()
                if count > 0:
                    df.loc[bank_mask, 'Action 1'] = 'Bank/Sutherland'
                    self.logger.info(f"  Applied Bank/Sutherland rule to {count} NEW cases")
            
            # Rule 3: Load DND Emails from DND Emails sheet and apply rules
            dnd_emails = set()
            try:
                # Load from previous file's DND Emails sheet
                if prev_file and os.path.exists(prev_file):
                    with pd.ExcelFile(prev_file) as excel:
                        if 'DND Emails' in excel.sheet_names:
                            dnd_sheet = pd.read_excel(prev_file, sheet_name='DND Emails')
                            if 'Email' in dnd_sheet.columns:
                                dnd_emails = set(dnd_sheet['Email'].dropna().astype(str).str.strip().str.lower())
                                self.logger.info(f"  Loaded {len(dnd_emails)} DND emails from previous DND Emails sheet")
            except Exception as e:
                self.logger.warning(f"Could not load DND Emails for application: {str(e)}")
            
            # Apply DND rule using loaded emails - ONLY TO NEW CASES
            if dnd_emails and 'Email' in df.columns:
                # Normalize emails in current df
                df['_email_check'] = df['Email'].astype(str).str.strip().str.lower()
                
                # Identify NEW cases with DND emails (skip preserved)
                dnd_matches = new_cases_mask & df['_email_check'].isin(dnd_emails)
                match_count = dnd_matches.sum()
                
                if match_count > 0:
                    self.logger.info(f"  Found {match_count} NEW cases matching DND emails")
                    
                    # Update fields
                    if 'Action 1' in df.columns:
                        df.loc[dnd_matches, 'Action 1'] = 'DND'
                    if 'Action 2' in df.columns:
                        df.loc[dnd_matches, 'Action 2'] = 'DND'
                    if 'Action 3' in df.columns:
                        df.loc[dnd_matches, 'Action 3'] = 'DND'
                    if 'Final Action' in df.columns:
                        df.loc[dnd_matches, 'Final Action'] = 'DND'
                    if 'Status' in df.columns:
                        df.loc[dnd_matches, 'Status'] = 'closed'
                
                # Clean up temp column
                if '_email_check' in df.columns:
                    df = df.drop(columns=['_email_check'])
            
            # Application of "DND" status check (if manually set to DND status) - ONLY NEW CASES
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_mask = (
                    new_cases_mask &  # Only new cases!
                    (df['Status'].astype(str).str.lower() == 'dnd') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                count = dnd_mask.sum()
                if count > 0:
                    df.loc[dnd_mask, 'Action 1'] = 'DND'
                    self.logger.info(f"  Applied DND action to {count} NEW cases with Status='DND'")
            
            # Clean up _is_preserved flag - don't include in final output
            if '_is_preserved' in df.columns:
                df = df.drop(columns=['_is_preserved'])
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error applying business rules to Companies: {str(e)}")
            if '_email_check' in df.columns:
                df = df.drop(columns=['_email_check'])
            if '_is_preserved' in df.columns:
                df = df.drop(columns=['_is_preserved'])
            return df

    def _sort_companies_sheet(self, df):
        """Sort Companies sheet: New first, then by Email to group cases"""
        try:
            if df.empty or 'Status' not in df.columns:
                return df
            
            self.logger.info("Sorting Companies sheet (New first, then by Email)...")
            
            # Create a sort key
            # 0 for New/new, 1 for others
            df['_sort_key'] = df['Status'].astype(str).str.strip().str.lower().apply(
                lambda x: 0 if x == 'new' else 1
            )
            
            # Primary sort: Status (New first)
            # Secondary sort: Assigned To (Group by handler)
            # Tertiary sort: Email (Group same emails within handler)
            # Quaternary sort: Case Number
            sort_cols = ['_sort_key']
            ascending = [True]
            
            if 'Assigned To' in df.columns:
                sort_cols.append('Assigned To')
                ascending.append(True)
            
            if 'Email' in df.columns:
                sort_cols.append('Email')
                ascending.append(True)
            
            if 'Case Number' in df.columns:
                sort_cols.append('Case Number')
                ascending.append(True)
                
            df = df.sort_values(by=sort_cols, ascending=ascending)
            
            # Drop sort key
            df = df.drop(columns=['_sort_key'])
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error sorting Companies sheet: {str(e)}")
            return df

    def _reorder_companies_columns(self, df):
        """Reorder Companies sheet columns to match the specified order"""
        try:
            if df.empty:
                return df
            
            self.logger.info("Reordering Companies sheet columns...")
            
            # Get existing columns
            existing_cols = list(df.columns)
            
            # Build ordered columns list: include only columns that exist in df
            ordered_cols = []
            for col in self.companies_column_order:
                if col in existing_cols:
                    ordered_cols.append(col)
            
            # Add any remaining columns that weren't in the specified order (at the end)
            for col in existing_cols:
                if col not in ordered_cols:
                    ordered_cols.append(col)
            
            # Reorder the DataFrame
            df = df[ordered_cols]
            
            self.logger.info(f"Columns reordered. Final column count: {len(df.columns)}")
            return df
            
        except Exception as e:
            self.logger.error(f"Error reordering Companies columns: {str(e)}")
            return df

    def merge_current_with_consolidated_previous(self, current_df, consolidated_prev_df):
        """Merge current data with consolidated previous handler data, preserving handler work"""
        try:
            self.logger.info("Merging current data with consolidated previous handler data...")
            
            if consolidated_prev_df is None or consolidated_prev_df.empty:
                self.logger.info("No previous data to merge - returning current data")
                return current_df
            
            # Ensure Case Number column exists in both DataFrames
            if 'Case Number' not in current_df.columns:
                self.logger.warning("No 'Case Number' column in current data - returning current data")
                return current_df
            
            if 'Case Number' not in consolidated_prev_df.columns:
                self.logger.warning("No 'Case Number' column in previous data - returning current data")
                return current_df
            
            # Normalize case numbers for comparison
            current_df = current_df.copy()
            consolidated_prev_df = consolidated_prev_df.copy()
            
            # Normalize to integer for robust comparisons
            current_df['Case Number'] = current_df['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            consolidated_prev_df['Case Number'] = consolidated_prev_df['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            
            # Get case numbers
            current_case_numbers = set(current_df['Case Number'])
            prev_case_numbers = set(consolidated_prev_df['Case Number'])
            
            # Remove placeholder values
            current_case_numbers.discard('UNKNOWN')
            prev_case_numbers.discard('UNKNOWN')
            
            # Identify cases
            existing_cases = current_case_numbers & prev_case_numbers  # Cases in both
            new_cases = current_case_numbers - prev_case_numbers      # New cases
            removed_cases = prev_case_numbers - current_case_numbers  # Cases only in previous file
            
            self.logger.info(f"Case analysis:")
            self.logger.info(f"  Existing cases to update: {len(existing_cases)}")
            self.logger.info(f"  New cases to add: {len(new_cases)}")
            self.logger.info(f"  Cases only in previous file: {len(removed_cases)}")
            
            # CRITICAL FIX: Start with ALL previous cases (don't drop any!)
            final_cases = consolidated_prev_df.copy()
            self.logger.info(f"Started with {len(final_cases)} cases from previous file")
            
            # Update existing cases with new data from raw file
            updated_count = 0
            for case_num in existing_cases:
                try:
                    # Get current case data (from new raw file)
                    current_case = current_df[current_df['Case Number'] == case_num].iloc[0]
                    
                    # Find the case in final_cases and update it
                    case_mask = final_cases['Case Number'] == case_num
                    if case_mask.any():
                        case_idx = case_mask.idxmax()
                        
                        # Update basic fields with new data from raw file
                        basic_fields = [
                            'Customer Name', 'Company Name', 'Email', 'Phone Number', 
                            'Work Order Type', 'Incoming Channel', 'Last Status Change',
                            'Country', 'State/Province', 'Assigned To'
                        ]

                        # Use flexible column resolution when updating fields (handles variants like 'Customer')
                        for field in basic_fields:
                            col_current = self._resolve_col(current_df, field)
                            col_final = self._resolve_col(final_cases, field)
                            if col_current and col_final and col_current in current_case.index and col_final in final_cases.columns:
                                if pd.notna(current_case[col_current]) and str(current_case[col_current]).strip() != '':
                                    final_cases.loc[case_idx, col_final] = current_case[col_current]
                        
                        # Handle Status field specially - only change "in progress today" to "in_progress"
                        if 'Status' in current_case.index and 'Status' in final_cases.columns:
                            current_status = str(current_case['Status']).strip().lower()
                            if current_status == 'in progress today':
                                final_cases.loc[case_idx, 'Status'] = 'in_progress'
                                self.logger.debug(f"Case {case_num}: Status updated from 'in progress today' to 'in_progress'")
                            else:
                                # Only update if current status is more recent/meaningful
                                prev_status = str(final_cases.loc[case_idx, 'Status']).strip().lower()
                                if prev_status in ['new', ''] and current_status not in ['new', '']:
                                    final_cases.loc[case_idx, 'Status'] = current_case['Status']
                                    self.logger.debug(f"Case {case_num}: Status updated from '{prev_status}' to '{current_case['Status']}'")
                        
                        # Handle Action fields - preserve handler work unless current is more specific
                        action_fields = ['Action 1', 'Action 2', 'Action 3', 'Final Action']
                        for field in action_fields:
                            if field in current_case.index and field in final_cases.columns:
                                current_value = current_case[field]
                                prev_value = final_cases.loc[case_idx, field]
                                
                                # Only update if current value exists and previous is empty
                                if (pd.notna(current_value) and 
                                    str(current_value).strip() != '' and 
                                    str(current_value).lower() not in ['nan', 'none', ''] and
                                        (_is_missing(prev_value))):
                                    
                                    final_cases.loc[case_idx, field] = current_value
                                    self.logger.debug(f"Case {case_num}: Updated {field} from empty to '{current_value}'")
                        
                        updated_count += 1
                        
                except Exception as e:
                    self.logger.warning(f"Error updating case '{case_num}': {str(e)}")
                    continue
            
            self.logger.info(f"Updated {updated_count} existing cases with new data from raw file")
            
            # Add new cases from raw file
            new_cases_df = current_df[current_df['Case Number'].isin(new_cases)].copy()
            if not new_cases_df.empty:
                # Clean possible internal columns if they existed in legacy files
                internal_cols = ['_Source_Sheet', '_Handler_Name']
                for col in internal_cols:
                    if col in new_cases_df.columns:
                        new_cases_df = new_cases_df.drop(columns=[col])
                
                final_cases = pd.concat([final_cases, new_cases_df], ignore_index=True)
                self.logger.info(f"Added {len(new_cases_df)} new cases from raw file")
            
            # CRITICAL FIX: Remove duplicates after merging
            self.logger.info("=== REMOVING DUPLICATES AFTER MERGE ===")
            pre_dedup_count = len(final_cases)
            
            # Sort by Created On date (newest first) before removing duplicates
            if 'Created On' in final_cases.columns:
                self.logger.info("   Sorting by 'Created On' to keep newest duplicate cases")
                try:
                    # Parse Created On to datetime for proper sorting
                    final_cases['_created_on_parsed'] = pd.to_datetime(final_cases['Created On'].astype(str).str.replace(r'\s+', ' ', regex=True), 
                                                                        format='%m/%d/%Y %I:%M:%S %p', errors='coerce')
                    
                    # If parsing fails, try alternative format
                    if final_cases['_created_on_parsed'].isna().any():
                        final_cases['_created_on_parsed'] = pd.to_datetime(final_cases['Created On'], errors='coerce')
                    
                    # Sort by the parsed date in descending order (newest first)
                    final_cases = final_cases.sort_values('_created_on_parsed', ascending=False, na_position='last')
                    self.logger.info("   Successfully sorted by Created On date")
                except Exception as e:
                    self.logger.warning(f"   Could not parse Created On dates for sorting: {str(e)}")
                    # Fallback: sort by original column as string
                    final_cases = final_cases.sort_values('Created On', ascending=False, na_position='last')
            
            final_cases = final_cases.drop_duplicates(subset=['Case Number'], keep='first')
            
            # Drop the temporary sorting column if it exists
            if '_created_on_parsed' in final_cases.columns:
                final_cases = final_cases.drop(columns=['_created_on_parsed'])
            
            post_dedup_count = len(final_cases)
            duplicates_removed = pre_dedup_count - post_dedup_count
            
            if duplicates_removed > 0:
                self.logger.warning(f"Removed {duplicates_removed} duplicate case entries")
                self.logger.warning(f"   This was causing cases to appear multiple times in the database")
            else:
                self.logger.info("No duplicates found after merge")
            
            # Sort by case number for consistency
            if not final_cases.empty:
                final_cases = final_cases.sort_values('Case Number').reset_index(drop=True)
            
            self.logger.info(f"Final merged dataset: {len(final_cases)} total cases")
            self.logger.info(f"  - Previous cases preserved: {len(consolidated_prev_df)}")
            self.logger.info(f"  - New cases added: {len(new_cases_df)}")
            self.logger.info(f"  - Duplicates removed: {duplicates_removed}")
            self.logger.info(f"  - Total after merge and deduplication: {len(final_cases)}")
            
            return final_cases
            
        except Exception as e:
            self.logger.error(f"Error merging current with consolidated previous: {str(e)}")
            # Fallback: return current data only
            return current_df

    def apply_business_rules_preserving_handler_work(self, df):
        """Apply business rules while preserving handler work - only update specific fields"""
        try:
            self.logger.info("Applying business rules while preserving handler work...")
            df = df.copy()
            
            # Rule 1: Convert "in progress today" to "in_progress" (only this specific change)
            if 'Status' in df.columns:
                in_progress_today_mask = df['Status'].astype(str).str.strip().str.lower() == 'in progress today'
                in_progress_today_count = in_progress_today_mask.sum()
                if in_progress_today_count > 0:
                    df.loc[in_progress_today_mask, 'Status'] = 'in_progress'
                    self.logger.info(f"Applied rule: {in_progress_today_count} cases changed from 'in progress today' to 'in_progress'")
            
            # Rule 2: Apply Bank/Sutherland rule (only if Status is 'closed' and Action 1 is empty)
            if 'Status' in df.columns and 'Action 1' in df.columns and 'Company Name' in df.columns:
                bank_sutherland_mask = (
                    (df['Status'].str.lower() == 'closed') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                bank_sutherland_count = bank_sutherland_mask.sum()
                if bank_sutherland_count > 0:
                    df.loc[bank_sutherland_mask, 'Action 1'] = 'Bank/Sutherland'
                    self.logger.info(f"Applied Bank/Sutherland rule: {bank_sutherland_count} cases updated")
            
            # Rule 3: Apply DND rule (only if specific conditions are met)
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_mask = (
                    (df['Status'].str.lower() == 'dnd') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                dnd_count = dnd_mask.sum()
                if dnd_count > 0:
                    df.loc[dnd_mask, 'Action 1'] = 'DND'
                    self.logger.info(f"Applied DND rule: {dnd_count} cases updated")
            
            self.logger.info("Business rules applied while preserving handler work")
            return df
            
        except Exception as e:
            self.logger.error(f"Error applying business rules: {str(e)}")
            return df

    def sort_cases_by_status(self, df):
        """Sort cases by status in the order: New → In Progress → Skipped → Closed"""
        try:
            if 'Status' not in df.columns:
                self.logger.warning("No 'Status' column found - cannot sort by status")
                return df
            
            # Create a custom sorting order for status
            # Order: new → in_progress → skipped → closed
            status_order = {
                'new': 1,
                'in_progress': 2,
                'in progress': 2,
                'in progress today': 2,
                'skipped': 3,
                'Skipped': 3,
                'closed': 4,
                'Closed': 4,
                'needs follow up': 5,
                'pending': 6,
                'escalation': 7,
                'bank/sutherland': 8,
                'dnd': 9,
                'DND': 9
            }
            
            # Create a temporary column for sorting
            df_sorted = df.copy()
            df_sorted['_status_sort_order'] = df_sorted['Status'].astype(str).str.lower().map(
                lambda x: status_order.get(x.strip(), 999)  # Default to 999 for unknown statuses
            )
            
            # Sort by status order, then by case number for consistency
            df_sorted = df_sorted.sort_values(['_status_sort_order', 'Case Number']).reset_index(drop=True)
            
            # Remove the temporary sorting column
            df_sorted = df_sorted.drop(columns=['_status_sort_order'])
            
            # Log the sorting results
            if 'Status' in df_sorted.columns:
                status_counts = df_sorted['Status'].value_counts()
                self.logger.info(f"Cases sorted by status. Status distribution:")
                for status, count in status_counts.items():
                    if pd.notna(status) and str(status).strip():
                        self.logger.info(f"  {status}: {count} cases")
            
            return df_sorted
            
        except Exception as e:
            self.logger.error(f"Error sorting cases by status: {str(e)}")
            return df

    def create_handler_sheets_from_processed_data(self, writer, processed_df, prev_file=None):
        """Create individual handler sheets from the fully processed data"""
        try:
            self.logger.info("Creating individual handler sheets from processed data...")
            
            # Extract unique handlers from the processed data
            if 'Assigned To' not in processed_df.columns:
                self.logger.warning("No 'Assigned To' column found - skipping handler sheet creation")
                return
            
            # CRITICAL: Load previous Companies case numbers to exclude them from handler sheets
            # Cases in Companies sheet should NOT appear in handler sheets
            prev_companies_case_nums = set()
            if prev_file and os.path.exists(prev_file):
                try:
                    with pd.ExcelFile(prev_file) as excel:
                        # Look for Companies sheet
                        companies_sheets = [s for s in excel.sheet_names if isinstance(s, str) and 'companies' in s.lower()]
                        for sheet_name in companies_sheets:
                            try:
                                companies_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                                if 'Case Number' in companies_df.columns:
                                    for cn in companies_df['Case Number'].dropna():
                                        try:
                                            normalized = int(float(str(cn).strip()))
                                            prev_companies_case_nums.add(normalized)
                                        except:
                                            pass
                            except Exception:
                                continue
                    if prev_companies_case_nums:
                        self.logger.info(f"Loaded {len(prev_companies_case_nums)} case numbers from previous Companies sheet - these will be EXCLUDED from handler sheets")
                except Exception as e:
                    self.logger.warning(f"Could not load previous Companies for exclusion: {e}")
            
            # Exclude Companies cases from processed_df for handler sheet creation
            working_df = processed_df.copy()
            if prev_companies_case_nums and 'Case Number' in working_df.columns:
                working_df['_normalized_cn'] = working_df['Case Number'].apply(
                    lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() and str(v).strip().replace('.','',1).isdigit() else None
                )
                excluded_mask = working_df['_normalized_cn'].isin(prev_companies_case_nums)
                excluded_count = excluded_mask.sum()
                if excluded_count > 0:
                    self.logger.info(f"Excluding {excluded_count} cases from handler sheets (they belong to Companies sheet)")
                    working_df = working_df[~excluded_mask].copy()
                working_df = working_df.drop(columns=['_normalized_cn'], errors='ignore')
            
            # Get unique handlers (excluding empty/NaN values)
            handlers = working_df['Assigned To'].dropna().astype(str).str.strip().unique()
            handlers = [h for h in handlers if h and h.lower() not in ['nan', 'none', '']]

            # Load previous handler data if provided
            prev_handler_data = {}
            if prev_file and os.path.exists(prev_file):
                try:
                    prev_handler_data = self.load_previous_handler_sheets(prev_file)
                    self.logger.info(f"Loaded previous handler data for {len(prev_handler_data)} handlers (for preservation)")
                except Exception as e:
                    self.logger.warning(f"Could not load previous handler data: {str(e)}")
                    prev_handler_data = {}

            # Union current handlers with previous to ensure preservation
            all_handlers = set(handlers)
            all_handlers.update(prev_handler_data.keys())
            handlers = list(all_handlers)
            
            # Additional safety check for handler names
            valid_handlers = []
            for handler in handlers:
                if (len(handler) <= 50 and 
                    not handler.startswith('_') and 
                    not handler.startswith('.') and
                    not any(char in handler for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|'])):
                    valid_handlers.append(handler)
                else:
                    self.logger.warning(f"Skipping invalid handler name: '{handler}'")
            
            handlers = valid_handlers
            
            if not handlers:
                self.logger.warning("No valid handlers found in processed data - skipping handler sheet creation")
                return
            
            self.logger.info(f"Found {len(handlers)} valid handlers: {handlers}")
            
            # Create sheets for each handler
            for handler in handlers:
                try:
                    self.logger.info(f"Processing handler: {handler}")
                    
                    # New cases (current output for this handler) - use working_df which excludes Companies cases
                    new_cases = working_df[working_df['Assigned To'] == handler].copy()
                    if 'Case Number' in new_cases.columns:
                        new_cases['Case Number'] = new_cases['Case Number'].apply(
                            lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
                        ).astype('Int64')
                        new_cases = new_cases.dropna(subset=['Case Number'])

                    # Preserved cases from previous handler sheets (also exclude Companies cases)
                    prev_cases = prev_handler_data.get(handler, pd.DataFrame()).copy()
                    if not prev_cases.empty and 'Case Number' in prev_cases.columns:
                        prev_cases['Case Number'] = prev_cases['Case Number'].apply(
                            lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
                        ).astype('Int64')
                        prev_cases = prev_cases.dropna(subset=['Case Number'])
                        # Exclude Companies cases from preserved handler data too
                        if prev_companies_case_nums:
                            prev_cases = prev_cases[~prev_cases['Case Number'].isin(prev_companies_case_nums)].copy()
                        prev_cases['Assigned To'] = handler
                    else:
                        prev_cases = pd.DataFrame()

                    # Drop overlap: keep new first, then preserved not in new
                    new_case_nums = set(new_cases['Case Number'].dropna().tolist()) if not new_cases.empty else set()
                    if not prev_cases.empty:
                        prev_cases = prev_cases[~prev_cases['Case Number'].isin(new_case_nums)].copy()

                    # Combine with ordering: new first, then preserved
                    if not new_cases.empty or not prev_cases.empty:
                        handler_cases = pd.concat([new_cases, prev_cases], ignore_index=True)
                    else:
                        self.logger.warning(f"No cases found for handler: {handler}")
                        continue
                    
                    # Sort cases by status (New → In Progress → Closed)
                    self.logger.info(f"Sorting cases for handler '{handler}' by status...")
                    handler_cases = self.sort_cases_by_status(handler_cases)
                    
                    # Create handler sheet name (ensure it's Excel-safe)
                    sheet_name = f"{handler}'s Cases"
                    if len(sheet_name) > 31:
                        sheet_name = f"{handler[:25]}'s Cases"
                    
                    # Write to Excel
                    handler_cases.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust columns for this handler sheet
                    self.auto_adjust_columns(writer, handler_cases, sheet_name)
                    
                    # Apply handler-specific formatting
                    self.format_handler_sheet(writer, sheet_name, handler_cases)
                    
                    self.logger.info(f"Created sheet '{sheet_name}' with {len(handler_cases)} cases (sorted by status)")
                    
                except Exception as e:
                    self.logger.error(f"Error processing handler '{handler}': {str(e)}")
                    import traceback
                    self.logger.error(f"Handler '{handler}' traceback: {traceback.format_exc()}")
                    continue
            
            self.logger.info(f"Handler sheet creation completed for {len(handlers)} handlers")
            
        except Exception as e:
            self.logger.error(f"Error creating handler sheets from processed data: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def create_handler_sheets(self, writer, output_df, prev_file=None):
        """Create individual sheets for each handler with their assigned cases.
        FIXED: Now preserves ALL handler sheets regardless of selection status.
        This is called AFTER all SMS/Email processing to ensure statuses are current."""
        try:
            self.logger.info("Creating individual handler sheets...")
            
            # Extract unique handlers from the output data
            if 'Assigned To' not in output_df.columns:
                self.logger.warning("No 'Assigned To' column found - skipping handler sheet creation")
                return
            
            # Get unique handlers (excluding empty/NaN values)
            current_handlers = output_df['Assigned To'].dropna().astype(str).str.strip().unique()
            current_handlers = [h for h in current_handlers if h and h.lower() not in ['nan', 'none', '']]
            
            # Load previous handler sheets if they exist
            prev_handler_data = {}
            if prev_file and os.path.exists(prev_file):
                try:
                    prev_handler_data = self.load_previous_handler_sheets(prev_file)
                    self.logger.info(f"Loaded previous handler data for {len(prev_handler_data)} handlers")
                except Exception as e:
                    self.logger.warning(f"Could not load previous handler data: {str(e)}")
                    prev_handler_data = {}
            
            # FIXED: Combine current handlers with previous handlers to preserve ALL work
            all_handlers = set(current_handlers)
            all_handlers.update(prev_handler_data.keys())
            handlers = list(all_handlers)
            
            self.logger.info(f"Current handlers in output: {current_handlers}")
            self.logger.info(f"Previous handlers with work: {list(prev_handler_data.keys())}")
            self.logger.info(f"All handlers to process: {handlers}")
            
            # Additional safety check for handler names
            valid_handlers = []
            for handler in handlers:
                # Check if handler name is reasonable (not too long, contains valid characters)
                if (len(handler) <= 50 and 
                    not handler.startswith('_') and 
                    not handler.startswith('.') and
                    not any(char in handler for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|'])):
                    valid_handlers.append(handler)
                else:
                    self.logger.warning(f"Skipping invalid handler name: '{handler}'")
            
            handlers = valid_handlers
            
            if not handlers:
                self.logger.warning("No valid handlers found in data - skipping handler sheet creation")
                return
            
            self.logger.info(f"Found {len(handlers)} valid handlers to process: {handlers}")
            
            # Create/update sheets for each handler
            for handler in handlers:
                try:
                    self.logger.info(f"Processing handler: {handler}")
                    
                    # Filter cases for this handler from current output
                    handler_cases = output_df[output_df['Assigned To'] == handler].copy()
                    
                    # Get previous handler data if available
                    prev_cases = prev_handler_data.get(handler, pd.DataFrame())
                    
                    # FIXED: Handle handlers with previous work but no current cases
                    if handler_cases.empty and not prev_cases.empty:
                        self.logger.info(f"Handler '{handler}' has no current cases but has previous work - preserving previous work")
                        final_handler_cases = prev_cases.copy()
                    elif not handler_cases.empty and prev_cases.empty:
                        self.logger.info(f"Handler '{handler}' has current cases but no previous work - using current cases")
                        final_handler_cases = handler_cases.copy()
                    elif not handler_cases.empty and not prev_cases.empty:
                        self.logger.info(f"Handler '{handler}' has both current and previous cases - merging")
                        # Merge with previous data (update existing, append new)
                        final_handler_cases = self.merge_handler_cases(handler_cases, prev_cases, handler)
                    else:
                        self.logger.warning(f"Handler '{handler}' has no current or previous cases - skipping")
                        continue
                    
                    # Create handler sheet name (ensure it's Excel-safe)
                    sheet_name = f"{handler}'s Cases"
                    # Excel sheet names cannot exceed 31 characters
                    if len(sheet_name) > 31:
                        sheet_name = f"{handler[:25]}'s Cases"
                    
                    # Write to Excel
                    final_handler_cases.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust columns for this handler sheet
                    self.auto_adjust_columns(writer, final_handler_cases, sheet_name)
                    
                    # Apply handler-specific formatting (simplified for now)
                    self.format_handler_sheet(writer, sheet_name, final_handler_cases)
                    
                    self.logger.info(f"Created/updated sheet '{sheet_name}' with {len(final_handler_cases)} cases")
                    
                    # Do not add internal tracking columns; preservation handled by merging prev handler DataFrames
                    
                except Exception as e:
                    self.logger.error(f"Error processing handler '{handler}': {str(e)}")
                    import traceback
                    self.logger.error(f"Handler '{handler}' traceback: {traceback.format_exc()}")
                    continue  # Continue with next handler
            
            self.logger.info(f"Handler sheet creation completed for {len(handlers)} handlers")
            
        except Exception as e:
            self.logger.error(f"Error creating handler sheets: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def load_previous_handler_sheets(self, prev_file):
        """Load previous handler sheets data for smart updates"""
        prev_handler_data = {}

        # Validate file existence
        if not prev_file or not os.path.exists(prev_file):
            self.logger.info("No previous file provided or file doesn't exist")
            return prev_handler_data

        try:
            # Read the previous file safely
            prev_excel = pd.ExcelFile(prev_file)

            # Look for handler sheets (sheets ending with "'s Cases")
            handler_sheets = [s for s in prev_excel.sheet_names if isinstance(s, str) and s.endswith("'s Cases")]

            for sheet_name in handler_sheets:
                try:
                    # Normalize handler name and read the sheet
                    handler_name = str(sheet_name).replace("'s Cases", "").strip()
                    handler_df = pd.read_excel(prev_file, sheet_name=sheet_name)

                    if isinstance(handler_df, pd.DataFrame) and not handler_df.empty and 'Case Number' in handler_df.columns:
                        prev_handler_data[handler_name] = handler_df
                        self.logger.info(f"Loaded previous data for handler '{handler_name}': {len(handler_df)} cases")

                except Exception as e:
                    self.logger.warning(f"Error loading previous handler sheet '{sheet_name}': {str(e)}")
                    continue

            return prev_handler_data

        except Exception as e:
            self.logger.error(f"Error loading previous handler sheets: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return {}

    def merge_handler_cases(self, current_cases, prev_cases, handler_name):
        """Smart merge of current and previous handler cases:
        - Update existing cases with new data
        - Append new cases
        - Preserve handler-specific notes and custom fields
        """
        try:
            if prev_cases.empty:
                self.logger.info(f"No previous data for handler '{handler_name}' - using current cases only")
                return current_cases
            
            self.logger.info(f"Merging cases for handler '{handler_name}': {len(current_cases)} current, {len(prev_cases)} previous")
            
            # Ensure Case Number column exists
            if 'Case Number' not in current_cases.columns:
                self.logger.warning(f"No 'Case Number' column in current cases for handler '{handler_name}' - using current cases only")
                return current_cases
            
            if 'Case Number' not in prev_cases.columns:
                self.logger.warning(f"No 'Case Number' column in previous cases for handler '{handler_name}' - using current cases only")
                return current_cases
            
            # Normalize case numbers for comparison as integers (nullable)
            current_cases = current_cases.copy()
            prev_cases = prev_cases.copy()
            current_cases['Case Number'] = current_cases['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            prev_cases['Case Number'] = prev_cases['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            
            # Get case numbers
            current_case_numbers = set(current_cases['Case Number'].dropna().tolist())
            prev_case_numbers = set(prev_cases['Case Number'].dropna().tolist())
            
            # Identify cases
            existing_cases = current_case_numbers & prev_case_numbers  # Cases in both
            new_cases = current_case_numbers - prev_case_numbers      # New cases
            removed_cases = prev_case_numbers - current_case_numbers  # Cases no longer assigned to this handler
            
            self.logger.info(f"Handler '{handler_name}' case analysis:")
            self.logger.info(f"  Existing cases to update: {len(existing_cases)}")
            self.logger.info(f"  New cases to add: {len(new_cases)}")
            self.logger.info(f"  Cases no longer assigned: {len(removed_cases)}")
            
            # Start with new cases (these are current and complete)
            final_cases = current_cases[current_cases['Case Number'].isin(new_cases)].copy()
            
            # Add existing cases (merge current data with previous custom fields)
            for case_num in existing_cases:
                try:
                    # Get current case data
                    current_case = current_cases[current_cases['Case Number'] == case_num].iloc[0]
                    
                    # Get previous case data
                    prev_case = prev_cases[prev_cases['Case Number'] == case_num].iloc[0]
                    
                    # Create merged case row
                    merged_case = current_case.copy()
                    
                    # Preserve handler-specific custom fields from previous data
                    custom_fields = ['Handler Notes', 'Custom Status', 'Priority Level', 'Follow-up Date']
                    for field in custom_fields:
                        if field in prev_case.index and field not in merged_case.index:
                            merged_case[field] = prev_case[field]
                        elif field in prev_case.index and field in merged_case.index:
                            # Keep previous custom data if current is empty
                            if _is_missing(merged_case[field]):
                                merged_case[field] = prev_case[field]
                    
                    # Add to final cases
                    final_cases = pd.concat([final_cases, pd.DataFrame([merged_case])], ignore_index=True)
                    
                except Exception as e:
                    self.logger.warning(f"Error merging case '{case_num}' for handler '{handler_name}': {str(e)}")
                    # Add current case as fallback
                    current_case = current_cases[current_cases['Case Number'] == case_num].iloc[0]
                    final_cases = pd.concat([final_cases, pd.DataFrame([current_case])], ignore_index=True)
                    continue
            
            # Sort by case number for consistency
            if not final_cases.empty:
                final_cases = final_cases.sort_values('Case Number').reset_index(drop=True)
            
            self.logger.info(f"Final merged cases for handler '{handler_name}': {len(final_cases)} total")
            
            return final_cases
            
        except Exception as e:
            self.logger.error(f"Error merging handler cases for '{handler_name}': {str(e)}")
            # Fallback: return current cases only
            return current_cases

    def format_handler_sheet(self, writer, sheet_name, handler_cases):
        """Apply handler-specific formatting to the handler sheet"""
        try:
            # For now, just log that we would format the sheet
            # The actual formatting will be handled by pandas/xlsxwriter automatically
            self.logger.info(f"Handler sheet '{sheet_name}' created with {len(handler_cases)} cases")
            
            # Note: Column widths and formatting are handled automatically by pandas/xlsxwriter
            # We can add custom formatting later if needed using openpyxl after the file is created
            
        except Exception as e:
            self.logger.warning(f"Could not apply formatting to handler sheet '{sheet_name}': {str(e)}")

    def add_validation_dropdowns(self, output_file):
        """Add validation dropdown lists to the PA Cases sheet and all handler sheets based on validation lists, and append region/offset table after two empty columns."""
        try:
            self.logger.info("Adding validation dropdown lists to PA Cases sheet and handler sheets...")
            
            # Check if openpyxl is available
            if not OPENPYXL_AVAILABLE:
                self.logger.warning("openpyxl not available - skipping validation dropdowns")
                return
            
            # Load the workbook
            wb = openpyxl.load_workbook(output_file)

            # Get all sheets
            all_sheets = wb.sheetnames
            self.logger.info(f"Found sheets: {all_sheets}")
            
            # Identify handler sheets
            handler_sheets = [sheet for sheet in all_sheets if isinstance(sheet, str) and sheet.endswith("'s Cases")]
            main_sheets = ['PA Cases'] + handler_sheets
            if 'Companies' in all_sheets:
                main_sheets.append('Companies')
            
            self.logger.info(f"Will add validation to main sheets: {main_sheets}")

            # Process each main sheet (PA Cases + handler sheets)
            for sheet_name in main_sheets:
                if sheet_name not in wb.sheetnames:
                    self.logger.warning(f"Sheet '{sheet_name}' not found, skipping")
                    continue
                    
                self.logger.info(f"Adding validation dropdowns to sheet: {sheet_name}")
                data_sheet = wb[sheet_name]

                # Add filters to all header columns
                header_row = 1  # Headers are in row 1
                max_column = data_sheet.max_column
                data_sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_column)}{header_row}"

                # Create data validation for each specified column
                header_map = {}
                for cell in data_sheet[header_row]:
                    if cell.value:
                        header_map[cell.value] = cell.column

                for col_name, options in self.validation_lists.items():
                    if col_name in header_map:
                        col_letter = get_column_letter(header_map[col_name])
                        # Create data validation with explicit range
                        options_range = f'Validation!{get_column_letter(1 + list(self.validation_lists.keys()).index(col_name))}$2:{get_column_letter(1 + list(self.validation_lists.keys()).index(col_name))}${1 + len(options)}'
                        
                        # Check if we have the real DataValidation class (not dummy)
                        if OPENPYXL_AVAILABLE and hasattr(DataValidation, 'add'):
                            # Import the real DataValidation class for this use
                            from openpyxl.worksheet.datavalidation import DataValidation as RealDataValidation
                            dv = RealDataValidation(type="list", formula1=f'={options_range}')
                            dv.error = 'Your entry is not in the list'
                            dv.errorTitle = 'Invalid Entry'
                            dv.prompt = 'Please select from the list'
                            dv.promptTitle = col_name
                            # Apply to all cells in column except header
                            data_sheet.add_data_validation(dv)
                            dv.add(f'{col_letter}2:{col_letter}1048576')  # Applies to entire column
                            self.logger.info(f"Added validation dropdown for column: {col_name} in sheet: {sheet_name}")
                        else:
                            self.logger.warning(f"DataValidation not available - skipping validation for column: {col_name}")

            # Create or clear the Validation sheet
            if "Validation" in wb.sheetnames:
                validation_sheet = wb["Validation"]
                validation_sheet.delete_rows(1, validation_sheet.max_row)
            else:
                validation_sheet = wb.create_sheet("Validation")

            # Write validation options to the Validation sheet (horizontal, starting at A1)
            headers = list(self.validation_lists.keys())
            max_len = max(len(v) for v in self.validation_lists.values())
            for col_idx, col_name in enumerate(headers, start=1):
                validation_sheet.cell(row=1, column=col_idx, value=col_name)
                options = self.validation_lists[col_name]
                for row_idx, option in enumerate(options, start=2):
                    validation_sheet.cell(row=row_idx, column=col_idx, value=option)
                # Fill empty cells to max_len
                for row_idx in range(len(options)+2, max_len+2):
                    validation_sheet.cell(row=row_idx, column=col_idx, value="")

            # Leave two empty columns after the validation lists
            region_start_col = len(headers) + 3  # +1 for next col, +2 for two empty columns

            # Region/Offset data from screenshots
            regions = [
                "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado", "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota", "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada", "New Hampshire", "New Jersey", "New Mexico", "New York", "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon", "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington", "West Virginia", "Wisconsin", "Wyoming", "Alberta", "British Columbia", "Manitoba", "New Brunswick", "Newfoundland and Labrador", "Nova Scotia", "Ontario", "Prince Edward Island", "Quebec", "Saskatchewan", "Northwest Territories", "Nunavut", "Yukon"
            ]
            offsets = [
                "5:00:00 AM", "4:00:00 AM", "8:00:00 AM", "6:00:00 AM", "5:00:00 AM", "7:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "4:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "5:00:00 AM", "5:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM"
            ]
            region_table_headers = ["Region", "Offset", "=NOW()-K1", "3:00"]
            region_table_row_start = 1
            # Write headers
            for i, header in enumerate(region_table_headers):
                validation_sheet.cell(row=region_table_row_start, column=region_start_col + i, value=header)
            # Write region/offset data
            for idx, (region, offset) in enumerate(zip(regions, offsets), start=2):
                validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col, value=region)
                # Convert offset string to Excel time format
                from datetime import datetime
                import re
                time_match = re.match(r"(\d+):(\d+):(\d+) (AM|PM)", offset)
                if time_match:
                    hour, minute, second, ampm = time_match.groups()
                    hour = int(hour)
                    minute = int(minute)
                    second = int(second)
                    if ampm == 'PM' and hour != 12:
                        hour += 12
                    if ampm == 'AM' and hour == 12:
                        hour = 0
                    excel_time = (hour * 3600 + minute * 60 + second) / 86400
                    validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 1, value=excel_time)
                    validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 1).number_format = 'h:mm:ss AM/PM'
                else:
                    validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 1, value=offset)
                # Formula: =$J$1-I2 (I2 increments)
                formula = f"=$J$1-I{region_table_row_start + idx - 1}"
                validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 2, value=formula)
                # 3:00 column left blank except header

            # --- Add VLOOKUP formula to PA Cases sheet, column K ---
            # Only add to PA Cases sheet (not handler sheets) to avoid confusion
            if 'PA Cases' in wb.sheetnames:
                data_sheet = wb['PA Cases']
                pa_cases_max_row = data_sheet.max_row
                
                # Check if Local Time column already exists, if not create it
                local_time_col = None
                for col_idx, col_name in enumerate(data_sheet[1], 1):  # Check header row
                    if col_name.value == 'Local Time':
                        local_time_col = col_idx
                        break
                
                if not local_time_col:
                    # Create new Local Time column at the end
                    local_time_col = data_sheet.max_column + 1
                    # Set header for Local Time column
                    data_sheet.cell(row=1, column=local_time_col, value='Local Time')
                    self.logger.info(f"Created new Local Time column at column {local_time_col}")
                else:
                    self.logger.info(f"Local Time column already exists at column {local_time_col}")
                
                    # Add formulas to all data rows (overwrite existing data)
                    for row in range(2, pa_cases_max_row + 1):
                        # =VLOOKUP(k2,Validation!$H$1:$J$63,3) - Fixed range reference
                        # Column k is State, looking up in Validation sheet H:J range
                        formula = f'=VLOOKUP(k{row},Validation!{get_column_letter(region_start_col)}1:{get_column_letter(region_start_col+2)}{len(regions)+1},3)'
                        data_sheet.cell(row=row, column=local_time_col, value=formula)
                    self.logger.info(f"Added/Updated Local Time formulas in column {get_column_letter(local_time_col)} for {pa_cases_max_row-1} rows")

            # --- Add VLOOKUP formula to ALL handler sheets AND Companies ---
            # Get all handler sheet names (sheets ending with "'s Cases") + Companies
            handler_sheets = [sheet for sheet in wb.sheetnames if isinstance(sheet, str) and (sheet.endswith("'s Cases") or sheet == 'Companies')]
            self.logger.info(f"Adding Local Time formulas to {len(handler_sheets)} sheets (Handlers + Companies): {handler_sheets}")
            
            for handler_sheet_name in handler_sheets:
                try:
                    handler_sheet = wb[handler_sheet_name]
                    handler_max_row = handler_sheet.max_row
                    
                    # Check if Local Time column already exists, if not create it
                    local_time_col = None
                    for col_idx, col_name in enumerate(handler_sheet[1], 1):  # Check header row
                        if col_name.value == 'Local Time':
                            local_time_col = col_idx
                            break
                    
                    if not local_time_col:
                        # Create new Local Time column at the end
                        local_time_col = handler_sheet.max_column + 1
                        # Set header for Local Time column
                        handler_sheet.cell(row=1, column=local_time_col, value='Local Time')
                        self.logger.info(f"Created new Local Time column at column {local_time_col} in {handler_sheet_name}")
                    else:
                        self.logger.info(f"Local Time column already exists at column {local_time_col} in {handler_sheet_name}")
                    
                    # Add formulas to all data rows (overwrite existing data)
                    for row in range(2, handler_max_row + 1):
                        # =VLOOKUP(J2,Validation!$H$1:$J$63,3) - Same formula as PA Cases
                        # Column J is State/Province, looking up in Validation sheet H:J range
                        formula = f'=VLOOKUP(k{row},Validation!{get_column_letter(region_start_col)}1:{get_column_letter(region_start_col+2)}{len(regions)+1},3)'
                        handler_sheet.cell(row=row, column=local_time_col, value=formula)
                    
                    self.logger.info(f"Added/Updated Local Time formulas in column {get_column_letter(local_time_col)} for {handler_max_row-1} rows in {handler_sheet_name}")
                    
                except Exception as e:
                    self.logger.error(f"Error adding Local Time formulas to {handler_sheet_name}: {str(e)}")
                    continue

            # Save the workbook
            wb.save(output_file)
            self.logger.info("Validation dropdowns and region/offset table added successfully to all sheets")

        except Exception as e:
            self.logger.error(f"Error adding validation dropdowns: {str(e)}")

    def create_skipped_sms_sheet(self, writer, processing_stats=None):
        """Create a sheet showing skipped SMS cases"""
        try:
            self.logger.info("\n=== Collecting Skipped SMS Entries ===")
            
            # Initialize collection
            sms_entries = []
            
            if processing_stats:
                self.logger.info("Processing stats keys available:")
                for key in processing_stats.keys():
                    self.logger.info(f"- {key}")
                
                # Check SMS Processing Stats
                if 'SMS Processing Stats' in processing_stats:
                    sms_stats = processing_stats.get('SMS Processing Stats', {})
                    self.logger.info("SMS Processing Stats keys:")
                    for key in sms_stats.keys():
                        self.logger.info(f"- {key}")
                    
                    # Try to collect entries from both possible locations
                    skipped_updated = list(sms_stats.get('Skipped/Updated Entries', []))
                    skipped_cases = list(sms_stats.get('Skipped Cases', []))
                    
                    self.logger.info(f"Found {len(skipped_updated)} entries in Skipped/Updated Entries")
                    self.logger.info(f"Found {len(skipped_cases)} entries in Skipped Cases")
                    
                    # Combine unique entries
                    if skipped_updated:
                        sms_entries.extend(skipped_updated)
                    if skipped_cases and not skipped_updated:  # Only use legacy if no new format
                        sms_entries.extend(skipped_cases)
            else:
                self.logger.info("No processing_stats provided - cannot collect skipped SMS entries")
                return

            # Store collected entries in both locations for compatibility
            if processing_stats and sms_entries:
                try:
                    # Store in designated location for final unified sheet
                    processing_stats['_skipped_sms_entries'] = sms_entries
                    
                    # Also ensure it's in SMS Processing Stats
                    if 'SMS Processing Stats' not in processing_stats:
                        processing_stats['SMS Processing Stats'] = {}
                    processing_stats['SMS Processing Stats']['Skipped/Updated Entries'] = sms_entries
                    
                    self.logger.info(f"Successfully stored {len(sms_entries)} SMS entries for unified sheet")
                    
                    # Log some sample entries
                    if sms_entries:
                        self.logger.info("\nSample SMS entries collected:")
                        for entry in sms_entries[:3]:  # Show first 3
                            case = entry.get('Case Number', 'N/A')
                            reply = entry.get('SMS Text', entry.get('Reply Text', 'N/A'))
                            reason = entry.get('Reason', 'N/A')
                            self.logger.info(f"Case {case}: Reply='{reply}', Reason='{reason}'")
                except Exception as e:
                    self.logger.warning(f"Failed to store SMS entries: {str(e)}")
            return
        except Exception as e:
            self.logger.error(f"Error creating Skipped SMS Cases sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def _is_missing(self, value):
        """Helper function to check if a value should be considered missing"""
        if value is None:
            return True
        if isinstance(value, str) and value.strip() in ['', 'nan', 'None', 'N/A']:
            return True
        try:
            if pd.isna(value):
                return True
        except:
            pass
        return False

    def create_unified_skipped_sheet(self, writer, processing_stats=None):
        """Create a unified sheet showing all skipped cases (both SMS and Email)"""
        try:
            self.logger.info("\n=== Creating Unified Skipped Sheet ===")
            
            # Initialize storage for entries
            all_entries = []
            sms_entries = []
            email_entries = []
            
            # Get SMS entries - try both new and legacy locations
            if processing_stats:
                # First try direct storage
                sms_entries = list(processing_stats.get('_skipped_sms_entries', []))
                
                # If not found, try SMS Processing Stats
                if not sms_entries and 'SMS Processing Stats' in processing_stats:
                    sms_stats = processing_stats['SMS Processing Stats']
                    if sms_stats:
                        sms_entries = list(sms_stats.get('Skipped/Updated Entries', [])) or list(sms_stats.get('Skipped Cases', []))
                
                self.logger.info(f"Found {len(sms_entries)} SMS entries")
                
                # Get Email entries - try both new and legacy locations
                # First try direct storage
                email_entries = list(processing_stats.get('_skipped_email_entries', []))
                
                # If not found, try Email Processing Stats
                if not email_entries and 'Email Processing Stats' in processing_stats:
                    email_stats = processing_stats['Email Processing Stats']
                    if email_stats:
                        email_entries = list(email_stats.get('Skipped/Updated Entries', [])) or list(email_stats.get('Skipped Emails', []))
                
                self.logger.info(f"Found {len(email_entries)} Email entries")
            
            def normalize_entry(entry, source_type):
                """Helper to normalize entry format"""
                if not isinstance(entry, dict):
                    return None
                
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Base normalized entry
                normalized = {
                    'Timestamp': entry.get('Timestamp', now),
                    'Source': source_type,
                    'Case Number': entry.get('Case Number', 'N/A'),
                    'Reply Text': '',
                    'Reason': entry.get('Reason', 'Unknown'),
                    'Action': entry.get('Action', 'Skipped')
                }
                
                # Extract reply text based on source
                if source_type == 'SMS':
                    normalized['Reply Text'] = entry.get('SMS Text', entry.get('Reply Text', ''))
                else:  # Email
                    normalized['Reply Text'] = entry.get('Reply Text', '')
                
                # Clean up values
                for key in normalized:
                    if _is_missing(normalized[key]):
                        normalized[key] = 'N/A'
                    elif isinstance(normalized[key], str):
                        normalized[key] = normalized[key].strip()
                
                return normalized

            # Process SMS entries
            for entry in sms_entries:
                normalized = normalize_entry(entry, 'SMS')
                if normalized:
                    all_entries.append(normalized)
            
            # Process Email entries
            for entry in email_entries:
                normalized = normalize_entry(entry, 'Email')
                if normalized:
                    all_entries.append(normalized)
            
            if not all_entries:
                self.logger.info("No skipped entries to write to unified sheet")
                return

            # Create DataFrame
            df = pd.DataFrame(all_entries)
            
            # Ensure required columns exist and are in the right order
            required_columns = ['Timestamp', 'Source', 'Case Number', 'Reply Text', 'Reason', 'Action']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = 'N/A'
            
            # Reorder columns
            df = df.reindex(columns=required_columns)

            # Write to Excel
            sheet_name = 'Skipped SMS and Email Replies'
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Auto-adjust columns
            worksheet = writer.sheets[sheet_name]
            worksheet.autofit()
            
            self.logger.info(f"\nSuccessfully wrote {len(all_entries)} entries to unified skipped sheet")
            self.logger.info("Source distribution:")
            source_counts = df['Source'].value_counts()
            for source, count in source_counts.items():
                self.logger.info(f"  {source}: {count} entries")
            
            # Log some sample entries for verification
            self.logger.info("\nSample entries from unified sheet:")
            sample_size = min(5, len(df))
            for i, (_, row) in enumerate(df.head(sample_size).iterrows()):
                self.logger.info(f"Row {i+1}: {row['Source']} - Case {row['Case Number']} - {row['Action']}")
                reply_preview = str(row['Reply Text'])[:50] + ('...' if len(str(row['Reply Text'])) > 50 else '')
                self.logger.info(f"  Reply: '{reply_preview}'")
                self.logger.info(f"  Reason: {row['Reason']}")

        except Exception as e:
            self.logger.error(f"Error creating unified skipped sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def create_skipped_email_sheet(self, writer, processing_stats=None, email_file=None, output_df=None):
        """Create a sheet showing skipped Email replies.

        Fallback behavior: if processing_stats does not include skipped emails but an
        email_file is provided, attempt to parse the file and build skipped-email
        entries based on simple heuristics (missing case number, case not found in
        output_df, or non-actionable replies).
        """
        # Initialize the collections we'll use
        skipped_emails = []
        combined_entries = []
        sms_entries = []
        stats = processing_stats if processing_stats is not None else {}

        self.logger.info("\n=== Creating Skipped SMS / Email Replies Sheet ===")

        try:
            # Log what we find in stats for debugging
            if stats:
                self.logger.info("Processing stats keys found:")
                for key in stats.keys():
                    self.logger.info(f"- {key}")
                if 'SMS Processing Stats' in stats:
                    self.logger.info("SMS Processing Stats keys:")
                    for key in stats['SMS Processing Stats'].keys():
                        self.logger.info(f"- {key}")
                if 'Email Processing Stats' in stats:
                    self.logger.info("Email Processing Stats keys:")
                    for key in stats['Email Processing Stats'].keys():
                        self.logger.info(f"- {key}")

            # Get SMS entries from both places they might be stored
            sms_entries = []
            if stats:
                # Try the direct storage location
                try:
                    stored_sms = stats.get('_skipped_sms_entries', [])
                    if stored_sms:
                        sms_entries.extend(stored_sms)
                        self.logger.info(f"Found {len(stored_sms)} SMS entries in _skipped_sms_entries")
                except Exception:
                    pass
                
                        # Also check SMS Processing Stats
                try:
                    sms_stats = stats.get('SMS Processing Stats', {})
                    if sms_stats:
                        skipped_sms = list(sms_stats.get('Skipped/Updated Entries', [])) or list(sms_stats.get('Skipped Cases', []))
                        if skipped_sms:
                            sms_entries.extend(skipped_sms)
                            self.logger.info(f"Found {len(skipped_sms)} SMS entries in SMS Processing Stats")
                except Exception:
                    pass
                    
            # Initialize skipped_emails list
            skipped_emails = []

            # Get Email entries from processing stats
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                skipped_emails.extend(list(email_stats.get('Skipped/Updated Entries', [])) or list(email_stats.get('Skipped Emails', [])))
                self.logger.info(f"Found {len(skipped_emails)} email entries in Email Processing Stats")

            # Fallback: attempt to parse email_file to build minimal uniform entries
            input_email_file = email_file
            if not skipped_emails and input_email_file and os.path.exists(input_email_file):
                try:
                    self.logger.info(f"No skipped emails in processing_stats - attempting to parse email file as fallback: {input_email_file}")
                    email_df = pd.read_excel(input_email_file)

                    # Helper to find likely column by name or index (F/H)
                    def find_col_by_name_or_index(df, names, idx):
                        for n in names:
                            for col in df.columns:
                                if isinstance(col, str) and col.strip().lower() == n.strip().lower():
                                    return col
                        if len(df.columns) > idx:
                            return df.columns[idx]
                        return None

                    case_number_col = find_col_by_name_or_index(email_df, ['Case Number (Object) (Email)', 'Case Number'], 5)
                    reply_col = find_col_by_name_or_index(email_df, ['Action', 'Reply', 'Description'], 7)

                    if case_number_col and reply_col:
                        for idx, row in email_df.iterrows():
                            raw_case = row.get(case_number_col, '')
                            case_str = str(raw_case).strip()
                            reply_raw = row.get(reply_col, '')
                            reply_text = str(reply_raw).strip() if reply_raw else ''
                            # Normalize case
                            import re
                            m = re.search(r"(\d+)", case_str)
                            norm_case = m.group(1) if m else case_str
                            reason = 'Missing case number' if not norm_case or norm_case.lower() in ['nan', 'none', ''] else 'Non-actionable or not found'
                            if not norm_case or norm_case.lower() in ['nan', 'none', '']:
                                skipped_emails.append({'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Source': 'Email', 'Case Number': 'N/A', 'Reply Text': reply_text, 'Result': 'Skipped', 'Reason': 'Missing case number'})
                            else:
                                skipped_emails.append({'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Source': 'Email', 'Case Number': str(norm_case), 'Reply Text': reply_text, 'Result': 'Skipped', 'Reason': reason})
                    else:
                        self.logger.info("Could not find necessary columns in email file for fallback skipped-email parsing")
                except Exception as e:
                    self.logger.warning(f"Failed fallback parsing of email file for skipped emails: {e}")

            # Normalize entries: ensure they have the unified keys
            def normalize_entry(e, default_source=''):
                # Expecting already in unified format or legacy short dicts
                if not isinstance(e, dict):
                    return None
                # If it already has Timestamp/Source/Case Number/Reply Text/Result/Reason, use as-is
                keys = ['Timestamp', 'Source', 'Case Number', 'Reply Text', 'Result', 'Reason']
                if all(k in e for k in keys):
                    return {k: e.get(k, '') for k in keys}
                # Legacy SMS format
                if 'SMS Text' in e or 'Reply Text' in e:
                    ts = e.get('Timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    source = e.get('Source', default_source or 'SMS')
                    case = e.get('Case Number', e.get('case_number', ''))
                    reply = e.get('SMS Text', e.get('Reply Text', ''))
                    reason = e.get('Reason', e.get('Update Made', 'Skipped'))
                    result = 'Updated' if 'Update Made' in e or e.get('Result') == 'Updated' else 'Skipped'
                    return {'Timestamp': ts, 'Source': source, 'Case Number': case, 'Reply Text': reply, 'Result': result, 'Reason': reason}
                # Fallback minimal mapping
                ts = e.get('Timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                source = e.get('Source', default_source or 'Email')
                case = e.get('Case Number', e.get('case_number', ''))
                reply = e.get('Reply Text', e.get('reply_text', ''))
                reason = e.get('Reason', '')
                result = e.get('Result', 'Skipped')
                return {'Timestamp': ts, 'Source': source, 'Case Number': case, 'Reply Text': reply, 'Result': result, 'Reason': reason}

            # Combine entries
            combined = []
            
            # Add SMS entries
            sms_added = 0
            for s in sms_entries:
                ne = normalize_entry(s, default_source='SMS')
                if ne:
                    combined.append(ne)
                    sms_added += 1
            self.logger.info(f"Added {sms_added} normalized SMS entries")
            
            # Add Email entries
            email_added = 0
            for e in skipped_emails:
                ne = normalize_entry(e, default_source='Email')
                if ne:
                    combined.append(ne)
                    email_added += 1
            self.logger.info(f"Added {email_added} normalized Email entries")

            if not combined:
                self.logger.info("No skipped SMS/Email entries to report")
                return

            # Create the unified sheet
            self.logger.info(f"Creating unified Skipped SMS / Email Replies sheet with {len(combined)} total rows")
            skipped_df = pd.DataFrame(combined)
            # Ensure column order
            cols = ['Timestamp', 'Source', 'Case Number', 'Reply Text', 'Result', 'Reason']
            for c in cols:
                if c not in skipped_df.columns:
                    skipped_df[c] = ''
            skipped_df = skipped_df[cols]
            
            # Save and format
            try:
                skipped_df.to_excel(writer, sheet_name='Skipped SMS and Email Replies', index=False)
                self.auto_adjust_columns(writer, skipped_df, 'Skipped SMS and Email Replies')
                self.logger.info(f"Successfully created Skipped SMS and Email Replies sheet with {len(skipped_df)} rows")
                
                # Log some sample rows for verification
                self.logger.info("\nSample rows from unified sheet:")
                sample_size = min(5, len(skipped_df))
                sample_rows = skipped_df.head(sample_size)
                for idx, row in sample_rows.iterrows():
                    row_num = idx if isinstance(idx, int) else 0
                    self.logger.info(f"Row {row_num+1}: Source={row['Source']}, Case={row['Case Number']}, Result={row['Result']}")
            except Exception as write_error:
                self.logger.error(f"Error writing unified sheet: {str(write_error)}")
                raise
                
        except Exception as e:
            self.logger.error(f"Error creating Skipped Email Replies sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            try:
                raise
            except Exception as e:
                self.logger.error(f"Error processing email file: {str(e)}")
                return

    def create_dnd_emails_database_sheet(self, writer, output_df, email_file, processing_stats=None, prev_file=None):
        """Create DND emails sheet by preserving ALL previous data and only appending new emails"""
        try:
            self.logger.info("Creating DND Emails sheet...")
            
            # Use a set to store emails to easily handle uniqueness
            dnd_emails_set = set()
            dnd_list = []
            
            # STEP 1: Load and preserve ALL previous DND Emails data FIRST
            self.logger.info(f"Attempting to load DND Emails from prev_file: {prev_file}")
            if prev_file and os.path.exists(prev_file):
                self.logger.info(f"prev_file exists for DND Emails: {prev_file}")
                try:
                    self.logger.info(f"Loading previous DND Emails sheet from: {prev_file}")
                    prev_excel = pd.ExcelFile(prev_file)
                    
                    self.logger.info(f"Found sheets in DND prev_file: {prev_excel.sheet_names}")
                    
                    if 'DND Emails' in prev_excel.sheet_names:
                        prev_dnd = pd.read_excel(prev_file, sheet_name='DND Emails')
                        self.logger.info(f"Previous DND Emails sheet loaded with {len(prev_dnd)} records")
                        
                        if 'Email' in prev_dnd.columns:
                            # Add ALL previous emails to our set
                            for idx, row in prev_dnd.iterrows():
                                email = str(row.get('Email', '')).strip()
                                if email and email.lower() not in ['', 'nan', 'none']:
                                    if email not in dnd_emails_set:
                                        dnd_emails_set.add(email)
                                        dnd_list.append({'Email': email, 'DND': 'Yes'})
                            self.logger.info(f"Preserved {len(dnd_list)} emails from previous DND Emails sheet")
                        else:
                            self.logger.warning("Previous file does not contain 'Email' column")
                    else:
                        self.logger.info("Previous output file does not contain 'DND Emails' tab")
                        
                except Exception as e:
                    self.logger.error(f"Error loading previous DND Emails sheet: {str(e)}")
                    self.logger.warning("Continuing without previous file data")
            else:
                self.logger.info("No previous output file found - starting with empty DND Emails sheet")
            
            # STEP 2: Add new DND emails from PA Cases output_df (if not already in set)
            new_emails_count = 0
            if output_df is not None and 'Email' in output_df.columns:
                if 'DND (Do Not Disturb)' in output_df.columns:
                    mask = output_df['DND (Do Not Disturb)'].astype(str).str.strip().str.lower() == 'yes'
                    df_dnd = output_df[mask]
                    self.logger.info(f"Found {len(df_dnd)} new DND cases in PA Cases")
                else:
                    df_dnd = pd.DataFrame()

                for _, row in df_dnd.iterrows():
                    email = str(row.get('Email', '')).strip()
                    if email and email.lower() not in ['', 'nan', 'none']:
                        if email not in dnd_emails_set:
                            dnd_emails_set.add(email)
                            dnd_list.append({'Email': email, 'DND': 'Yes'})
                            new_emails_count += 1
                            self.logger.debug(f"Added NEW DND email from PA Cases: {email}")
            
            # STEP 3: Also include any raw-file-extracted DND emails (if not already in set)
            if processing_stats and isinstance(processing_stats, dict):
                raw_dnd = processing_stats.get('Raw File DND Emails') or []
                for item in raw_dnd:
                    if isinstance(item, dict):
                        email = str(item.get('Email', '')).strip()
                    else:
                        email = str(item).strip()
                    if email and email.lower() not in ['', 'nan', 'none']:
                        if email not in dnd_emails_set:
                            dnd_emails_set.add(email)
                            dnd_list.append({'Email': email, 'DND': 'Yes'})
                            new_emails_count += 1
                            self.logger.debug(f"Added NEW DND email from processing stats: {email}")
            
            # Create DataFrame
            if dnd_list:
                dnd_df = pd.DataFrame(dnd_list)
            else:
                dnd_df = pd.DataFrame(columns=['Email', 'DND'])
            
            # Save to sheet named 'DND Emails'
            try:
                dnd_df.to_excel(writer, sheet_name='DND Emails', index=False)
                self.auto_adjust_columns(writer, dnd_df, 'DND Emails')
                self.logger.info(f"DND Emails sheet created with {len(dnd_df)} total entries (added {new_emails_count} new emails)")
            except Exception as e:
                self.logger.warning(f"Could not write DND Emails sheet: {e}")

        except Exception as e:
            self.logger.error(f"Error creating DND Emails sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def extract_dnd_emails_from_file(self, email_file):
        """Extract DND emails from the email file - DISABLED (not used)"""
        # This method is not called and requires BeautifulSoup which may not be available
        # For future use, implement without external HTML parsing library
        return []

    def create_sms_replies_sheet(self, writer, df, sms_file=None, processing_stats=None, prev_file=None, issue_not_resolved_cases=None):
        """Create Issue Not Fixed sheet by preserving ALL previous data and only appending new cases"""
        try:
            self.logger.info("Creating Issue Not Fixed sheet...")
            
            # Use a dictionary to store cases, keyed by Case Number, to easily handle updates and ensure uniqueness
            # This will prioritize previous data if a case number exists in both previous and new data
            final_issue_not_fixed_cases_map = {}

            # STEP 1: Load and preserve ALL previous Issue Not Fixed tab data FIRST
            self.logger.info(f"Attempting to load from prev_file: {prev_file}")
            if prev_file and os.path.exists(prev_file):
                self.logger.info(f"prev_file exists: {prev_file}")
                try:
                    self.logger.info(f"Loading previous Issue Not Fixed tab from: {prev_file}")
                    # Read the previous output file to get the Issue Not Fixed tab
                    prev_excel = pd.ExcelFile(prev_file)
                    
                    self.logger.info(f"Found sheets in prev_file: {prev_excel.sheet_names}")
                    
                    if 'Issue Not Fixed' in prev_excel.sheet_names:
                        prev_sms = pd.read_excel(prev_file, sheet_name='Issue Not Fixed')
                        self.logger.info(f"Previous Issue Not Fixed tab loaded with {len(prev_sms)} records")
                        
                        if 'Case Number' in prev_sms.columns:
                            # Add ALL previous cases to our map (preserve everything exactly as it was)
                            for idx, row in prev_sms.iterrows():
                                case_num = str(row['Case Number']).strip()
                                if case_num and case_num.lower() not in ['nan', 'none', '']:
                                    final_issue_not_fixed_cases_map[case_num] = {
                                        'Case Number': case_num,
                                        'Incoming Channel': row.get('Incoming Channel', ''),
                                        'Date Added': row.get('Date Added', ''), # Keep original date
                                        'Actioned (Y/N)': row.get('Actioned (Y/N)', ''),
                                        'Action Note': row.get('Action Note', ''),
                                        'Source': 'Previous File' # Internal tracking
                                    }
                            self.logger.info(f"Preserved {len(final_issue_not_fixed_cases_map)} cases from previous Issue Not Fixed tab")
                        else:
                            self.logger.warning("Previous file does not contain 'Case Number' column")
                    else:
                        self.logger.info("Previous output file does not contain 'Issue Not Fixed' tab")

                except Exception as e:
                    self.logger.error(f"Error loading previous Issue Not Fixed tab: {str(e)}")
                    self.logger.warning("Continuing without previous file data")
            else:
                self.logger.info("No previous output file found - starting with empty Issue Not Fixed tab")

            # Get today's date for new cases
            today_date_str = datetime.now().strftime('%Y-%m-%d')
            
            # STEP 2: Add new cases from SMS/Email processing (if not already in map)
            new_cases_from_sms_email = 0
            
            # Get SMS Issue Not Fixed cases
            if processing_stats and 'SMS Processing Stats' in processing_stats:
                sms_stats = processing_stats['SMS Processing Stats']
                if 'Issue Not Fixed' in sms_stats:
                    sms_issue_not_fixed_cases = sms_stats['Issue Not Fixed']
                    self.logger.info(f"Found {len(sms_issue_not_fixed_cases)} cases marked as 'Issue Not Fixed' via SMS processing")

                    for case_data in sms_issue_not_fixed_cases:
                        # Handle both old format (string) and new format (dict)
                        if isinstance(case_data, dict):
                            case_num_str = str(case_data.get('case_number', '')).strip()
                        else:
                            case_num_str = str(case_data).strip()
                            
                        if case_num_str and case_num_str not in final_issue_not_fixed_cases_map:
                            # Get Incoming Channel from the main df if available
                            incoming_channel = ''
                            if 'Incoming Channel' in df.columns:
                                case_row = df[df['Case Number'] == case_num_str]
                                if not case_row.empty:
                                    incoming_channel = case_row.iloc[0].get('Incoming Channel', '')

                            # This is a truly NEW case - add it with today's date
                            final_issue_not_fixed_cases_map[case_num_str] = {
                                'Case Number': case_num_str,
                                'Incoming Channel': incoming_channel,
                                'Date Added': today_date_str,
                                'Actioned (Y/N)': '',
                                'Action Note': '',
                                'Source': 'SMS Processing' # Internal tracking
                            }
                            new_cases_from_sms_email += 1
                            self.logger.info(f"Added NEW case from SMS processing: {case_num_str} with today's date")
                        else:
                            self.logger.debug(f"Skipped existing case from SMS processing: {case_num_str} (already in tab)")

            # Get Email Issue Not Fixed cases
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                if 'Issue Not Fixed' in email_stats:
                    email_issue_not_fixed_cases = email_stats['Issue Not Fixed']
                    self.logger.info(f"Found {len(email_issue_not_fixed_cases)} cases marked as 'Issue Not Fixed' via Email processing")

                    for case_data in email_issue_not_fixed_cases:
                        # Handle both old format (string) and new format (dict)
                        if isinstance(case_data, dict):
                            case_num_str = str(case_data.get('case_number', '')).strip()
                        else:
                            case_num_str = str(case_data).strip()
                            
                        if case_num_str and case_num_str not in final_issue_not_fixed_cases_map:
                            # Get Incoming Channel from the main df if available
                            incoming_channel = ''
                            if 'Incoming Channel' in df.columns:
                                case_row = df[df['Case Number'] == case_num_str]
                                if not case_row.empty:
                                    incoming_channel = case_row.iloc[0].get('Incoming Channel', '')

                            # This is a truly NEW case - add it with today's date
                            final_issue_not_fixed_cases_map[case_num_str] = {
                                'Case Number': case_num_str,
                                'Incoming Channel': incoming_channel,
                                'Date Added': today_date_str,
                                'Actioned (Y/N)': '',
                                'Action Note': '',
                                'Source': 'Email Processing' # Internal tracking
                            }
                            new_cases_from_sms_email += 1
                            self.logger.info(f"Added NEW case from Email processing: {case_num_str} with today's date")
                        else:
                            self.logger.debug(f"Skipped existing case from Email processing: {case_num_str} (already in tab)")

            # STEP 3: Add new cases from PA Cases (only if not already in map)
            new_cases_from_pa = 0
            issue_not_fixed_in_pa = df[df['Final Action'].str.lower() == 'issue not fixed'.lower()]
            self.logger.info(f"Found {len(issue_not_fixed_in_pa)} cases with Final Action = 'Issue Not Fixed' in PA Cases")

            for idx, row in issue_not_fixed_in_pa.iterrows():
                case_num = str(row['Case Number']).strip()
                if case_num and case_num not in final_issue_not_fixed_cases_map:
                    # This is a truly NEW case - add it with today's date
                    final_issue_not_fixed_cases_map[case_num] = {
                        'Case Number': case_num,
                        'Incoming Channel': row.get('Incoming Channel', ''),
                        'Date Added': today_date_str,
                        'Actioned (Y/N)': '',
                        'Action Note': '',
                        'Source': 'PA Cases' # Internal tracking
                    }
                    new_cases_from_pa += 1
                    self.logger.info(f"Added NEW case from PA Cases: {case_num} with today's date")
                else:
                    self.logger.debug(f"Skipped existing case from PA Cases: {case_num} (already in tab)")
            
            # STEP 3.5: Add cases marked as "Issue Not Resolved" by handlers in their individual sheets
            new_cases_from_handlers = 0
            if issue_not_resolved_cases:
                self.logger.info(f"Found {len(issue_not_resolved_cases)} cases marked as 'Issue Not Resolved' by handlers in their individual sheets")
                
                for case_data in issue_not_resolved_cases:
                    case_num = str(case_data.get('Case Number', '')).strip()
                    if case_num and case_num not in final_issue_not_fixed_cases_map:
                        # This is a NEW case from handler manual updates - add it with today's date
                        final_issue_not_fixed_cases_map[case_num] = {
                            'Case Number': case_num,
                            'Incoming Channel': 'Handler Manual Update',
                            'Date Added': today_date_str,
                            'Actioned (Y/N)': '',
                            'Action Note': f"Handler: {case_data.get('Handler', 'Unknown')} - {case_data.get('Final Action', '')}",
                            'Source': 'Handler Manual Update' # Internal tracking
                        }
                        new_cases_from_handlers += 1
                        self.logger.info(f"Added NEW case from Handler Manual Update: {case_num} (Handler: {case_data.get('Handler', 'Unknown')})")
                    else:
                        self.logger.debug(f"Skipped existing case from Handler Manual Update: {case_num} (already in tab)")
            else:
                self.logger.info("No cases marked as 'Issue Not Resolved' by handlers found")
            
            # STEP 4: Create final DataFrame
            if final_issue_not_fixed_cases_map:
                # Convert map values to a list of dictionaries
                all_issue_not_fixed_cases_list = list(final_issue_not_fixed_cases_map.values())
                # Create DataFrame with ALL fields including Source
                sms_df = pd.DataFrame(all_issue_not_fixed_cases_list)
                
                # Ensure proper column order: Case Number, Incoming Channel, Date Added, Actioned (Y/N), Action Note
                # Note: Source field is kept for internal processing but not shown in final output
                column_order = ['Case Number', 'Incoming Channel', 'Date Added', 'Actioned (Y/N)', 'Action Note']
                sms_df = sms_df[column_order]

                # Log comprehensive summary
                self.logger.info(f"\n=== Issue Not Fixed Sheet Summary ===")
                self.logger.info(f"Total cases in sheet: {len(sms_df)}")

                # Count cases by source using the 'Source' field from the original map values
                prev_file_count = sum(1 for case_data in final_issue_not_fixed_cases_map.values() if case_data.get('Source') == 'Previous File')
                new_cases_count = sum(1 for case_data in final_issue_not_fixed_cases_map.values() if case_data.get('Source') != 'Previous File')
                
                self.logger.info(f"Cases preserved from previous file: {prev_file_count}")
                self.logger.info(f"NEW cases added (today's date): {new_cases_count}")
                self.logger.info(f"  - From SMS/Email processing: {new_cases_from_sms_email}")
                self.logger.info(f"  - From PA Cases: {new_cases_from_pa}")
                self.logger.info(f"  - From Handler Manual Updates: {new_cases_from_handlers}")
                
                # Log date distribution
                if 'Date Added' in sms_df.columns:
                    date_counts = sms_df['Date Added'].value_counts()
                    self.logger.info(f"Date distribution:")
                    for date, count in date_counts.items():
                        self.logger.info(f"  {date}: {count} cases")
                
                # Log action status
                if 'Actioned (Y/N)' in sms_df.columns:
                    action_counts = sms_df['Actioned (Y/N)'].value_counts()
                    self.logger.info(f"Action status distribution:")
                    for action, count in action_counts.items():
                        if _is_missing(action):
                            self.logger.info(f"  Not Actioned: {count} cases")
                        else:
                            self.logger.info(f"  {action}: {count} cases")
                
                # Save to sheet
                sms_df.to_excel(writer, sheet_name='Issue Not Fixed', index=False)
                
                # Auto-adjust columns for Issue Not Fixed sheet
                self.auto_adjust_columns(writer, sms_df, 'Issue Not Fixed')
                
                self.logger.info(f"Issue Not Fixed sheet created successfully with {len(sms_df)} cases")
                
                # Log sample cases for verification
                sample_df = sms_df.head(5)[['Case Number', 'Incoming Channel', 'Date Added', 'Actioned (Y/N)']]
                self.logger.info(f"Sample cases in sheet:")
                for idx, row in sample_df.iterrows():
                    self.logger.info(f"  Case {row['Case Number']}: Channel {row['Incoming Channel']}, Date {row['Date Added']}, Actioned {row['Actioned (Y/N)']}")
                
            else:
                # Create empty sheet with headers
                empty_df = pd.DataFrame(columns=['Case Number', 'Incoming Channel', 'Date Added', 'Actioned (Y/N)', 'Action Note'])
                empty_df.to_excel(writer, sheet_name='Issue Not Fixed', index=False)
                
                # Auto-adjust columns for empty Issue Not Fixed sheet
                self.auto_adjust_columns(writer, empty_df, 'Issue Not Fixed')
                
                self.logger.info("Created empty Issue Not Fixed sheet (no cases found)")
            
        except Exception as e:
            self.logger.error(f"Error creating Issue Not Fixed sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def count_invalid_emails(self, df):
        try:
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if 'Email' not in df.columns:
                return 0
            invalid_emails = df[(df['Email'] != '') & (~df['Email'].str.match(email_pattern, na=False))].shape[0]
            return invalid_emails
        except Exception as e:
            self.logger.error(f"Error counting invalid emails: {str(e)}")
            return 0

    def create_counters_sheet(self, writer, df, prev_df=None, prev_file=None):
        try:
            self.logger.info("Creating counters sheet (reading from PA Cases main sheet)...")
            # Gather all unique handlers from current output
            handler_set = set(df['Assigned To'].dropna().astype(str).str.strip())
            # If previous file is available, include its handlers too (from main sheet)
            if prev_df is not None:
                prev_handlers = set(prev_df['Assigned To'].dropna().astype(str).str.strip())
                handler_set.update(prev_handlers)
            # Also include handlers from previous individual handler sheets regardless of selection
            if prev_file and os.path.exists(prev_file):
                try:
                    prev_excel = pd.ExcelFile(prev_file)
                    handler_sheets = [sheet for sheet in prev_excel.sheet_names if isinstance(sheet, str) and sheet.endswith("'s Cases")]
                    sheet_handlers = [str(sheet).replace("'s Cases", '').strip() for sheet in handler_sheets]
                    handler_set.update(sheet_handlers)
                except Exception as e:
                    self.logger.warning(f"Could not read previous handler sheets for counter: {str(e)}")
            # Remove empty handler names
            handler_list = sorted([h for h in handler_set if h and h.lower() != 'nan'])
            
            # Check if we have handlers to process
            if not handler_list:
                self.logger.warning("No handlers found to create counter sheet")
                return
                
            progress_columns = [
                ("closed",    ("$R$", "$B$2")),
                ("in_progress", ("$R$", "$C$2")),
                ("new",        ("$R$", "$D$2")),
                ("Skipped", ("$Q$", "$E$2")),
                ("In Progress Today", ("$Q$", "$G$2")),
                ("Requires a Call", None),  # Special: Final Action='Sent Email' AND Status='in_progress'
                ("Total",      None),
            ]
            worksheet = writer.book.add_worksheet('Counter')
            writer.sheets['Counter'] = worksheet
            worksheet.write(0, 0, 'Progress Counter')
            # Write header
            worksheet.write(1, 0, 'Handler')
            for col_idx, (col_name, _) in enumerate(progress_columns, 1):
                worksheet.write(1, col_idx, col_name)
            
            # Initialize row_idx to ensure it's always defined
            row_idx = 1
            
            # Write handler rows with formulas
            for row_idx, handler in enumerate(handler_list, 2):
                worksheet.write(row_idx, 0, handler)
                for col_idx, (col_name, col_formula) in enumerate(progress_columns, 1):
                    if col_name == "Total":
                        worksheet.write_formula(row_idx, col_idx, f"=SUM(B{row_idx+1}:G{row_idx+1})")
                    elif col_name == "Requires a Call":
                        # Count cases with Final Action = 'Sent Email' AND Status = 'in_progress'
                        handler_sheet_name = f"{handler}'s Cases"
                        escaped_sheet_name = handler_sheet_name.replace("'", "''")
                        excel_sheet_name = f"'{escaped_sheet_name}'"
                        # P = Final Action, R = Status
                        formula = f"=COUNTIFS({excel_sheet_name}!P:P,\"Sent Email\",{excel_sheet_name}!R:R,\"in_progress\")"
                        worksheet.write_formula(row_idx, col_idx, formula)
                    else:
                        # Create dynamic handler sheet name with proper Excel escaping
                        handler_sheet_name = f"{handler}'s Cases"
                        # Excel requires double single quotes to escape single quotes in sheet names
                        escaped_sheet_name = handler_sheet_name.replace("'", "''")
                        excel_sheet_name = f"'{escaped_sheet_name}'"
                        
                        if col_name == "closed":
                            # Updated formulas using new column indices
                            # Status is column J (10th column) in new layout -> R
                            # Assigned To is column I (9th column) in new layout -> Q 
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$B$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "in_progress":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$C$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "new":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$D$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "Skipped":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$E$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "In Progress Today":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$F$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        else:
                            formula = ""
                        worksheet.write_formula(row_idx, col_idx, formula)


            # Write totals row
            worksheet.write(row_idx+1, 0, 'Total')
            for col_idx in range(1, len(progress_columns)+1):
                worksheet.write_formula(row_idx+1, col_idx, f"=SUM({chr(66+col_idx-1)}3:{chr(66+col_idx-1)}{row_idx+1})")
            self.logger.info("Progress Counter table created with Excel formulas (reading from individual handler sheets)")

            # --- Final Action Counter Table ---
            final_action_options = [
                            "Fixed", "Refused Callback", "Issue Not Fixed", "Not yet Tested", "Escalation", "Sent SMS", "Sent Email", "Left VM", "Reviewed", "Bank/Sutherland", "Not Reached", "DND"

            ]
            start_row = row_idx + 3
            worksheet.write(start_row, 0, 'Final Action Counter')
            worksheet.write(start_row+1, 0, 'Handler')
            for col_idx, action in enumerate(final_action_options, 1):
                worksheet.write(start_row+1, col_idx, action)
            
            # Initialize h_idx to ensure it's always defined
            h_idx = start_row + 1
            
            for h_idx, handler in enumerate(handler_list, start_row+2):
                worksheet.write(h_idx, 0, handler)
                for a_idx, action in enumerate(final_action_options, 1):
                    # Create dynamic handler sheet name with proper Excel escaping
                    handler_sheet_name = f"{handler}'s Cases"
                    # Excel requires double single quotes to escape single quotes in sheet names
                    escaped_sheet_name = handler_sheet_name.replace("'", "''")
                    excel_sheet_name = f"'{escaped_sheet_name}'"
                    # Read from individual handler sheet with proper COUNTIFS format
                    formula = f"=COUNTIFS({excel_sheet_name}!P:P,Counter!{chr(65+a_idx)}{start_row+2},{excel_sheet_name}!Q:Q,Counter!A{h_idx+1})"
                    worksheet.write_formula(h_idx, a_idx, formula)
            
            # Write totals row for final action counter
            worksheet.write(h_idx+1, 0, 'Total')
            for a_idx in range(1, len(final_action_options)+1):
                worksheet.write_formula(h_idx+1, a_idx, f"=SUM({chr(65+a_idx)}{start_row+3}:{chr(65+a_idx)}{h_idx+1})")
            self.logger.info("Final Action Counter table created with Excel formulas (reading from individual handler sheets)")
            
            # --- Company Cases Table ---
            # Shows handler progress specifically in Companies sheet
            company_start_row = h_idx + 4
            worksheet.write(company_start_row, 0, 'Company Cases')
            worksheet.write(company_start_row+1, 0, 'Handler')
            worksheet.write(company_start_row+1, 1, 'New')
            worksheet.write(company_start_row+1, 2, 'Closed')
            worksheet.write(company_start_row+1, 3, 'Total')
            
            # Initialize company row counter
            co_idx = company_start_row + 1
            
            # Only create this table if Companies sheet exists
            if 'Companies' in writer.sheets:
                for co_idx, handler in enumerate(handler_list, company_start_row+2):
                    worksheet.write(co_idx, 0, handler)
                    # New: Count where Status='new' AND Assigned To = handler in Companies
                    # Assigned To is column Q (index 16), Status is column R (index 17)
                    worksheet.write_formula(co_idx, 1, f"=COUNTIFS('Companies'!R:R,\"new\",'Companies'!Q:Q,Counter!A{co_idx+1})")
                    # Closed: Count where Status='closed' AND Assigned To = handler in Companies
                    worksheet.write_formula(co_idx, 2, f"=COUNTIFS('Companies'!R:R,\"closed\",'Companies'!Q:Q,Counter!A{co_idx+1})")
                    # Total: Sum of New + Closed for this handler
                    worksheet.write_formula(co_idx, 3, f"=SUM(B{co_idx+1}:C{co_idx+1})")
                
                # Write totals row for Company Cases
                worksheet.write(co_idx+1, 0, 'Total')
                worksheet.write_formula(co_idx+1, 1, f"=SUM(B{company_start_row+3}:B{co_idx+1})")
                worksheet.write_formula(co_idx+1, 2, f"=SUM(C{company_start_row+3}:C{co_idx+1})")
                worksheet.write_formula(co_idx+1, 3, f"=SUM(D{company_start_row+3}:D{co_idx+1})")
                self.logger.info("Company Cases table created with Excel formulas (reading from Companies sheet)")
        except Exception as e:
            self.logger.error(f"Error creating counters sheet: {str(e)}")

    def create_summary_sheet(self, writer, df, processing_stats=None):
        try:
            self.logger.info("Creating summary sheet...")

            
            summary_data = []
            
            # --- Processing Statistics Section ---
            if processing_stats:
                summary_data.extend([
                    {'Category': '--- Processing Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
                    {'Category': 'Raw File', 'Metric': 'Initial Count', 'Count': processing_stats.get('Initial Count', 0), 'Description': 'Total records in raw input file'},
                    {'Category': 'Work Order Status Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Work Order Status', 0), 'Description': 'Records remaining after excluding cancelled work orders'},
                    {'Category': 'Case Reason Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Case Reason', 0), 'Description': 'Records remaining after excluding escalation/complaint cases'},
                    {'Category': 'Closing Code Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Closing Code', 0), 'Description': 'Records remaining after excluding specific closing codes'},
                    {'Category': 'CID/DMR Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After CID/DMR Filter', 0), 'Description': 'Records remaining after CID/DMR filtering'},
                    {'Category': 'Duplicate Removal', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Duplicate Removal', 0), 'Description': 'Records remaining after removing duplicate case numbers'},
                    {'Category': 'Raw File', 'Metric': 'Final Count After Cleaning', 'Count': processing_stats.get('Raw File Final Count After Cleaning', ''), 'Description': 'Final count of raw file after all cleaning'},
                    {'Category': 'Final Output', 'Metric': 'Final Count', 'Count': processing_stats.get('Final Count', 0), 'Description': 'Final record count in output file'},
                    {'Category': 'Total Removed', 'Metric': 'Records Removed', 'Count': processing_stats.get('Total Removed', 0), 'Description': 'Total records removed during all filtering steps'},
                ])
            
            # --- Previous File Statistics Section ---
            prev_stats = processing_stats.get('Previous File Stats', {}) if processing_stats else {}
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Previous File Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': 'Previous File', 'Metric': 'Initial Count', 'Count': prev_stats.get('Initial Count', ''), 'Description': 'Total records in previous file'},
                {'Category': 'Matching Cases', 'Metric': 'Count', 'Count': prev_stats.get('Matching Cases', ''), 'Description': 'Cases that exist in both previous and current files'},
                {'Category': 'Updated Cases', 'Metric': 'Count', 'Count': prev_stats.get('Updated Cases', ''), 'Description': 'Cases updated with new information from raw file'},
                {'Category': 'New Cases', 'Metric': 'Count', 'Count': prev_stats.get('New Cases', ''), 'Description': 'New cases added from raw file'},
            ])
            
            # --- Dropped Cases Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Dropped Cases from Raw File Filtering ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            
            # Add dropped cases statistics if available
            if processing_stats:
                # Work Order Status Filter dropped cases
                initial_count = processing_stats.get('Initial Count', 0)
                after_wo_status = processing_stats.get('After Work Order Status', 0)
                if initial_count and after_wo_status:
                    dropped_wo_status = initial_count - after_wo_status
                    summary_data.append({
                        'Category': 'Work Order Status Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_wo_status,
                        'Description': 'Cases dropped due to cancelled work orders'
                    })
                
                # Case Reason Filter dropped cases
                after_case_reason = processing_stats.get('After Case Reason', 0)
                if after_wo_status and after_case_reason:
                    dropped_case_reason = after_wo_status - after_case_reason
                    summary_data.append({
                        'Category': 'Case Reason Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_case_reason,
                        'Description': 'Cases dropped due to escalation/complaint reasons'
                    })
                
                # Closing Code Filter dropped cases
                after_closing_code = processing_stats.get('After Closing Code', 0)
                if after_case_reason and after_closing_code:
                    dropped_closing_code = after_case_reason - after_closing_code
                    summary_data.append({
                        'Category': 'Closing Code Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_closing_code,
                        'Description': 'Cases dropped due to specific closing codes'
                    })
                
                # CID/DMR Filter dropped cases
                after_cid_dmr = processing_stats.get('After CID/DMR Filter', 0)
                if after_closing_code and after_cid_dmr:
                    dropped_cid_dmr = after_closing_code - after_cid_dmr
                    summary_data.append({
                        'Category': 'CID/DMR Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_cid_dmr,
                        'Description': 'Cases dropped due to CID/DMR filtering'
                    })
                
                # Duplicate Removal dropped cases
                after_duplicate = processing_stats.get('After Duplicate Removal', 0)
                if after_cid_dmr and after_duplicate:
                    dropped_duplicate = after_cid_dmr - after_duplicate
                    summary_data.append({
                        'Category': 'Duplicate Removal',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_duplicate,
                        'Description': 'Cases dropped due to duplicate case numbers'
                    })
            
            # --- Business Rules Updates Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Business Rules Updates ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            
            # Bank/Sutherland Rule Updates
            if 'Status' in df.columns and 'Action 1' in df.columns:
                bank_sutherland_updated = df[(df['Status'].str.lower() == 'closed') & (df['Action 1'] == 'Bank/Sutherland')].shape[0]
                summary_data.append({
                    'Category': 'Bank/Sutherland Rule',
                    'Metric': 'Cases Updated',
                    'Count': bank_sutherland_updated,
                    'Description': 'Cases updated with Status=closed and Action 1=Bank/Sutherland'
                })
            
            # DND Rule Updates
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_updated = df[(df['Status'] == 'Closed') | (df['Action 1'] == 'DND') | (df['Action 2'] == 'DND') | (df['Action 3'] == 'DND') | (df['Final Action'] == 'DND')].shape[0]
                summary_data.append({
                    'Category': 'DND Rule',
                    'Metric': 'Cases Updated',
                    'Count': dnd_updated,
                    'Description': 'Cases updated with DND status or actions'
                })
            
            # --- SMS Processing Updates Section ---
            if processing_stats and 'SMS Processing Stats' in processing_stats:
                sms_stats = processing_stats['SMS Processing Stats']
                if sms_stats:
                    summary_data.extend([
                        {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': '--- SMS Processing Updates ---', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': 'SMS Fixed', 'Metric': 'Cases Updated', 'Count': len(sms_stats.get('Fixed', [])), 'Description': 'Cases updated to Fixed status via SMS'},
                        {'Category': 'SMS Issue Not Fixed', 'Metric': 'Cases Updated', 'Count': len(sms_stats.get('Issue Not Fixed', [])), 'Description': 'Cases updated to Issue Not Fixed status via SMS'},
                        {'Category': 'SMS Refused Callback', 'Metric': 'Cases Updated', 'Count': len(sms_stats.get('Refused Callback', [])), 'Description': 'Cases updated to Refused Callback status via SMS'},
                        {'Category': 'SMS Skipped Cases', 'Metric': 'Cases Skipped', 'Count': len(sms_stats.get('Skipped Cases', [])), 'Description': 'SMS cases that were skipped (see Skipped SMS Cases sheet)'},
                        {'Category': 'SMS Ambiguous Cases', 'Metric': 'Cases for Review', 'Count': len(sms_stats.get('Ambiguous', [])), 'Description': 'SMS cases with unclear responses needing human review'},
                    ])
            
            # --- Email Processing Updates Section ---
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                if email_stats:
                    summary_data.extend([
                        {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': '--- Email Processing Updates ---', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': 'Email Fixed', 'Metric': 'Cases Updated', 'Count': len(email_stats.get('Fixed', [])), 'Description': 'Cases updated to Fixed status via Email'},
                        {'Category': 'Email Issue Not Fixed', 'Metric': 'Cases Updated', 'Count': len(email_stats.get('Issue Not Fixed', [])), 'Description': 'Cases updated to Issue Not Fixed status via Email'},
                        {'Category': 'Email DND', 'Metric': 'Cases Updated', 'Count': len(email_stats.get('DND', [])), 'Description': 'Cases updated to DND status via Email'},
                        {'Category': 'Email Skipped Cases', 'Metric': 'Cases Skipped', 'Count': len(email_stats.get('Skipped Emails', [])), 'Description': 'Email replies that were skipped (see Skipped Email Replies sheet)'},
                        {'Category': 'Email Ambiguous Cases', 'Metric': 'Cases for Review', 'Count': len(email_stats.get('Ambiguous', [])), 'Description': 'Email replies with unclear responses needing human review'},
                    ])
            
            # --- Data Quality Checks Section ---
            # Flexible customer name handling: detect customer column variants
            cust_col = self._resolve_col(df, 'Customer Name')
            missing_customer_count = df[df[cust_col].isna() | (df[cust_col] == '')].shape[0] if cust_col and cust_col in df.columns else 0
            
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Data Quality Checks ---', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': 'Missing Case Numbers', 'Metric': '', 'Count': df[df['Case Number'].isna() | (df['Case Number'] == '')].shape[0], 'Description': 'Cases without case numbers'},
                {'Category': 'Missing Customer Names', 'Metric': '', 'Count': missing_customer_count, 'Description': 'Cases without customer names'},
                {'Category': 'Missing Handler Assignments', 'Metric': '', 'Count': df[df['Assigned To'].isna() | (df['Assigned To'] == '')].shape[0], 'Description': 'Cases without handler assignments'},
                {'Category': 'Duplicate Case Numbers', 'Metric': '', 'Count': df[df.duplicated(subset=['Case Number'], keep=False)].shape[0], 'Description': 'Duplicate case numbers found'},
                {'Category': 'Invalid Email Formats', 'Metric': '', 'Count': self.count_invalid_emails(df), 'Description': 'Invalid email address formats'},
                {'Category': 'Empty Rows', 'Metric': '', 'Count': df[df.apply(lambda row: all(str(x).strip() == '' for x in row), axis=1)].shape[0], 'Description': 'Completely empty rows'},
            ])
            
            # --- Handler Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Handler Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            if 'Assigned To' in df.columns:
                handler_stats = df['Assigned To'].value_counts()
                for handler, count in handler_stats.items():
                    if handler and str(handler).strip():
                        summary_data.append({
                            'Category': 'Handler',
                            'Metric': handler,
                            'Count': count,
                            'Description': f'Total cases assigned to {handler}'
                        })
            
            # --- Bank/Sutherland Rule Closed Cases Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Bank/Sutherland Rule Closed Cases ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            # Use the same patterns as in the processor
            import re
            patterns = [
                ("Citi", r"\\bcitibank\\b|\\bcitigroup\\b|\\bciti group\\b|\\bciti bank\\b|\\bciti corp\\b|\\bciti\\b"),
                ("Amgen", r"amgen"),
                ("Sanofi", r"sanofi"),
                ("Scotiabank", r"bns|scotiabank|scotia bank|scotiabank canada|scotiaBank|bank of nova scotia|bank nova scotia"),
                ("Royal Bank", r"rbc|royal bank|royal bank of canada|rbc wealth management"),
                ("HSBC", r"hsbc"),
                ("TD Bank", r"td bank"),
            ]
            def matches_company(name, regex):
                if not isinstance(name, str):
                    return False
                name_lc = name.lower()
                if 'citizen' in name_lc and 'citi' in regex:
                    return False
                return re.search(regex, name_lc) is not None
            # Only count cases where Status == closed and Action 1 == Bank/Sutherland
            if 'Status' in df.columns and 'Action 1' in df.columns and 'Company Name' in df.columns:
                closed_mask = (df['Status'].str.lower() == 'closed') & (df['Action 1'] == 'Bank/Sutherland')
                for label, regex in patterns:
                    count = df[closed_mask & df['Company Name'].apply(lambda x: matches_company(x, regex))].shape[0]
                    summary_data.append({
                        'Category': 'Bank/Sutherland Rule',
                        'Metric': label,
                        'Count': count,
                        'Description': f'Closed cases due to Bank/Sutherland rule for {label}'
                    })
            
            # --- Status Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Status Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            if 'Status' in df.columns:
                status_stats = df['Status'].value_counts()
                for status, count in status_stats.items():
                    if status and str(status).strip():
                        summary_data.append({
                            'Category': 'Status',
                            'Metric': status,
                            'Count': count,
                            'Description': f'Cases with status: {status}'
                        })
            
            # --- Final Action Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Final Action Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            if 'Final Action' in df.columns:
                final_action_stats = df['Final Action'].value_counts()
                for action, count in final_action_stats.items():
                    if action and str(action).strip():
                        summary_data.append({
                            'Category': 'Final Action',
                            'Metric': action,
                            'Count': count,
                            'Description': f'Cases with final action: {action}'
                        })
            
            # Main summary data is complete - will add detailed case table separately
            
            # Dropped cases will be added to detailed table
            
            # SMS processing cases will be added to detailed table
            
            # Email processing cases will be added to detailed table
            
            # Add detailed case information for Bank/Sutherland updates
            if 'Status' in df.columns and 'Action 1' in df.columns and 'Company Name' in df.columns:
                bank_sutherland_cases = df[(df['Status'].str.lower() == 'closed') & (df['Action 1'] == 'Bank/Sutherland')]
                if not bank_sutherland_cases.empty:
                    case_numbers = bank_sutherland_cases['Case Number'].astype(str).tolist()
                    summary_data.append({
                        'Category': 'Bank/Sutherland Updates - Case Numbers',
                        'Metric': 'Cases',
                        'Count': len(case_numbers),
                        'Description': f'Cases updated with Bank/Sutherland rule: {", ".join(case_numbers[:10])}{"..." if len(case_numbers) > 10 else ""}'
                    })
            
            # Add detailed case information for DND updates
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_cases = df[(df['Status'] == 'DND') | (df['Action 1'] == 'DND') | (df['Action 2'] == 'DND') | (df['Action 3'] == 'DND') | (df['Final Action'] == 'DND')]
                if not dnd_cases.empty:
                    case_numbers = dnd_cases['Case Number'].astype(str).tolist()
                    summary_data.append({
                        'Category': 'DND Updates - Case Numbers',
                        'Metric': 'Cases',
                        'Count': len(case_numbers),
                        'Description': f'Cases updated with DND status/actions: {", ".join(case_numbers[:10])}{"..." if len(case_numbers) > 10 else ""}'
                    })
            
            # Create the summary DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # Add the detailed case table data to the summary DataFrame
            detailed_cases = []
            
            # Add dropped cases
            if processing_stats and 'Dropped Cases Details' in processing_stats:
                dropped_details = processing_stats['Dropped Cases Details']
                for filter_name, cases in dropped_details.items():
                    if cases:
                        for case_num in cases:
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'dropped',
                                'Reason': f"case reason: {filter_name}",
                                'Updated to': ''
                            })
            
            # Add updated cases from SMS processing
            if processing_stats and 'SMS Processing Stats' in processing_stats:
                sms_stats = processing_stats['SMS Processing Stats']
                self.logger.info(f"SMS stats found: {list(sms_stats.keys()) if sms_stats else 'None'}")
                for action_type in ['Fixed', 'Issue Not Fixed', 'Refused Callback', 'DND']:
                    if action_type in sms_stats and sms_stats[action_type]:
                        cases = sms_stats[action_type]
                        self.logger.info(f"SMS {action_type} cases found: {len(cases)}")
                        for case_data in cases:
                            # Handle both old format (string) and new format (dict)
                            if isinstance(case_data, dict):
                                case_num = case_data.get('case_number', '')
                                sms_text = case_data.get('sms_text', '')
                            else:
                                case_num = str(case_data)
                                sms_text = ''
                            
                            if case_num:
                                # Reason - include SMS text if available
                                if sms_text:
                                    reason = f"SMS: '{action_type}' - '{sms_text[:100]}{'...' if len(sms_text) > 100 else ''}'"
                                else:
                                    reason = f"SMS: '{action_type}'"
                                
                                detailed_cases.append({
                                    'Case Number': str(case_num),
                                    'Action': 'updated',
                                    'Reason': reason,
                                    'Updated to': action_type
                                })
                                self.logger.info(f"Added SMS case to summary: {case_num} - {action_type}")
                    else:
                        self.logger.info(f"SMS {action_type} cases not found or empty")
                
                # Add Ambiguous SMS cases for review
                if 'Ambiguous' in sms_stats and sms_stats['Ambiguous']:
                    ambiguous_cases = sms_stats['Ambiguous']
                    self.logger.info(f"SMS Ambiguous cases found: {len(ambiguous_cases)}")
                    for case_data in ambiguous_cases:
                        if isinstance(case_data, dict):
                            case_num = case_data.get('case_number', '')
                            sms_text = case_data.get('sms_text', '')
                        else:
                            case_num = str(case_data)
                            sms_text = ''
                        
                        if case_num:
                            # Reason - include SMS text if available
                            if sms_text:
                                reason = f"SMS: Ambiguous - needs review - '{sms_text[:100]}{'...' if len(sms_text) > 100 else ''}'"
                            else:
                                reason = f"SMS: Ambiguous - needs review"
                            
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'needs_review',
                                'Reason': reason,
                                'Updated to': 'Ambiguous'
                            })
                            self.logger.info(f"Added SMS Ambiguous case to summary: {case_num}")
                else:
                    self.logger.info("SMS Ambiguous cases not found or empty")
            
            # Add updated cases from Email processing
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                self.logger.info(f"Email stats found: {list(email_stats.keys()) if email_stats else 'None'}")
                for action_type in ['Fixed', 'Issue Not Fixed', 'DND']:
                    if action_type in email_stats and email_stats[action_type]:
                        cases = email_stats[action_type]
                        self.logger.info(f"Email {action_type} cases found: {len(cases)}")
                        for case_data in cases:
                            # Handle both old format (string) and new format (dict)
                            if isinstance(case_data, dict):
                                case_num = case_data.get('case_number', '')
                                reply_text = case_data.get('reply_text', '')
                            else:
                                case_num = str(case_data)
                                reply_text = ''
                            
                            if case_num:
                                # Reason - include email reply text if available
                                if reply_text:
                                    reason = f"Email: '{action_type}' - '{reply_text[:100]}{'...' if len(reply_text) > 100 else ''}'"
                                else:
                                    reason = f"Email: '{action_type}'"
                                
                                detailed_cases.append({
                                    'Case Number': str(case_num),
                                    'Action': 'updated',
                                    'Reason': reason,
                                    'Updated to': action_type
                                })
                                self.logger.info(f"Added Email case to summary: {case_num} - {action_type}")
                    else:
                        self.logger.info(f"Email {action_type} cases not found or empty")
                
                # Add Ambiguous Email cases for review
                if 'Ambiguous' in email_stats and email_stats['Ambiguous']:
                    ambiguous_cases = email_stats['Ambiguous']
                    self.logger.info(f"Email Ambiguous cases found: {len(ambiguous_cases)}")
                    for case_data in ambiguous_cases:
                        if isinstance(case_data, dict):
                            case_num = case_data.get('case_number', '')
                            reply_text = case_data.get('reply_text', '')
                        else:
                            case_num = str(case_data)
                            reply_text = ''
                        
                        if case_num:
                            # Reason - include email reply text if available
                            if reply_text:
                                reason = f"Email: Ambiguous - needs review - '{reply_text[:100]}{'...' if len(reply_text) > 100 else ''}'"
                            else:
                                reason = f"Email: Ambiguous - needs review"
                            
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'needs_review',
                                'Reason': reason,
                                'Updated to': 'Ambiguous'
                            })
                            self.logger.info(f"Added Email Ambiguous case to summary: {case_num}")
                else:
                    self.logger.info("Email Ambiguous cases not found or empty")
            else:
                self.logger.info("Email Processing Stats not found in processing stats")
            
            # Add Bank/Sutherland updated cases
            if processing_stats and 'Bank/Sutherland Updated Cases' in processing_stats:
                bank_sutherland_cases = processing_stats['Bank/Sutherland Updated Cases']
                self.logger.info(f"Bank/Sutherland cases found: {len(bank_sutherland_cases)}")
                if bank_sutherland_cases:
                    for case_data in bank_sutherland_cases:
                        case_num = str(case_data.get('case_number', ''))
                        company_name = str(case_data.get('company_name', ''))
                        if case_num:
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'updated',
                                'Reason': f"Bank/Sutherland rule applied - Bank: {company_name}",
                                'Updated to': 'Closed'
                            })
                            self.logger.info(f"Added Bank/Sutherland case to summary: {case_num} - {company_name}")
                else:
                    self.logger.info("No Bank/Sutherland cases to add")
            else:
                self.logger.info("Bank/Sutherland cases not found in processing stats")
            
            # Add DND updated cases
            if processing_stats and 'DND Updated Cases' in processing_stats:
                dnd_cases = processing_stats['DND Updated Cases']
                self.logger.info(f"DND cases found: {len(dnd_cases)}")
                if dnd_cases:
                    for case_data in dnd_cases:
                        case_num = str(case_data.get('case_number', ''))
                        source = str(case_data.get('source', 'DND rule'))
                        if case_num:
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'updated',
                                'Reason': f"{source} applied",
                                'Updated to': 'DND'
                            })
                            self.logger.info(f"Added DND case to summary: {case_num} - {source}")
                else:
                    self.logger.info("No DND cases to add")
            # Now create the final summary DataFrame with the detailed cases table
            # First, create the main summary DataFrame
            self.logger.info("Creating main summary DataFrame...")
            main_summary_df = pd.DataFrame(summary_data)
            self.logger.info(f"Main summary DataFrame created with shape: {main_summary_df.shape}")
            
            # Create the detailed cases DataFrame
            if detailed_cases:
                # Remove duplicates based on Case Number and Action
                unique_cases = {}
                for case in detailed_cases:
                    key = (case['Case Number'], case['Action'])
                    if key not in unique_cases:
                        unique_cases[key] = case
                    else:
                        # If duplicate found, keep the one with more detailed reason
                        if len(case['Reason']) > len(unique_cases[key]['Reason']):
                            unique_cases[key] = case
                
                # Convert back to list
                detailed_cases = list(unique_cases.values())
                self.logger.info(f"Removed duplicates from detailed cases. Original: {len(detailed_cases) + len(unique_cases) - len(detailed_cases)}, Final: {len(detailed_cases)}")
                
                detailed_df = pd.DataFrame(detailed_cases)
                
                # Create a combined DataFrame with detailed cases FIRST, then main summary
                # Add empty columns to main summary to align with detailed cases
                for col in ['Case Number', 'Action', 'Reason', 'Updated to']:
                    if col not in main_summary_df.columns:
                        main_summary_df[col] = ''
                
                # Combine the DataFrames - detailed cases FIRST, then main summary
                combined_df = pd.concat([detailed_df, main_summary_df], ignore_index=True)
                
                # Write the combined DataFrame to Excel
                self.logger.info(f"About to write combined DataFrame to Excel with {len(combined_df)} rows")
                self.logger.info(f"Combined DataFrame columns: {list(combined_df.columns)}")
                self.logger.info(f"First few rows of combined DataFrame:")
                for i, row in combined_df.head(3).iterrows():
                    self.logger.info(f"  Row {i}: {dict(row)}")
                
                # Create DataFrame and save to sheet
                if combined_df is not None and not combined_df.empty:
                    combined_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Auto-adjust columns for Summary sheet
                    self.auto_adjust_columns(writer, combined_df, 'Summary')
                    
                    self.logger.info(f"Summary sheet created with {len(combined_df)} rows")
                else:
                    # Create empty summary sheet
                    main_summary_df = pd.DataFrame(columns=['Category', 'Metric', 'Count', 'Description'])
                    main_summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Auto-adjust columns for empty Summary sheet
                    self.auto_adjust_columns(writer, main_summary_df, 'Summary')
                    
                    self.logger.info("Created empty Summary sheet")
            else:
                # No detailed cases, just write the main summary
                main_summary_df.to_excel(writer, sheet_name='Summary', index=False)
                self.logger.info(f"Summary sheet created with {len(main_summary_df)} summary items (no detailed cases)")
            
            self.logger.info(f"Detailed case table added to Summary sheet starting from column J, row 1 (at the TOP)")
            
            self.logger.info("=== SUMMARY SHEET CREATION COMPLETED SUCCESSFULLY ===")
        except Exception as e:
            self.logger.error(f"=== ERROR in create_summary_sheet: {str(e)} ===")
            self.logger.error(f"Exception type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def auto_adjust_columns(self, writer, df, sheet_name):
        """Auto-adjust column widths for a specific sheet based on content"""
        try:
            if not hasattr(writer, 'book') or sheet_name not in writer.sheets:
                return
            
            worksheet = writer.sheets[sheet_name]
            
            # Get the maximum length for each column
            for col_idx, column in enumerate(df.columns):
                # Get the maximum length of the column header
                header_length = len(str(column))
                
                # Get the maximum length of the data in this column
                if not df.empty:
                    # Convert all values to string and get max length
                    data_length = df[column].astype(str).str.len().max()
                    # Add some padding (minimum 10 characters)
                    max_length = max(header_length, data_length, 10) + 2
                else:
                    max_length = header_length + 2
                
                # Set column width (Excel column width is approximately 0.9 * character count)
                column_width = min(max_length * 0.9, 50)  # Cap at 50 to avoid extremely wide columns
                worksheet.set_column(col_idx, col_idx, column_width)
            
            self.logger.info(f"Auto-adjusted column widths for sheet: {sheet_name}")
            
        except Exception as e:
            self.logger.warning(f"Could not auto-adjust columns for {sheet_name}: {str(e)}")

    def validate_output_data(self, df):
        try:
            self.logger.info("Validating output data...")
            validation_results = {
                'total_records': len(df),
                'missing_case_numbers': df[df['Case Number'].isna() | (df['Case Number'] == '')].shape[0],
                'missing_handlers': df[df['Assigned To'].isna() | (df['Assigned To'] == '')].shape[0],
                'duplicate_cases': df[df.duplicated(subset=['Case Number'], keep=False)].shape[0],
                'empty_rows': df[df.apply(lambda row: all(str(x).strip() == '' for x in row), axis=1)].shape[0]
            }
            self.logger.info("Validation Results:")
            for key, value in validation_results.items():
                self.logger.info(f"  {key}: {value}")
            critical_issues = []
            if validation_results['missing_case_numbers'] > 0:
                critical_issues.append(f"Missing case numbers: {validation_results['missing_case_numbers']}")
            if validation_results['missing_handlers'] > 0:
                critical_issues.append(f"Missing handler assignments: {validation_results['missing_handlers']}")
            if validation_results['duplicate_cases'] > 0:
                critical_issues.append(f"Duplicate case numbers: {validation_results['duplicate_cases']}")
            if validation_results['empty_rows'] > 0:
                critical_issues.append(f"Empty rows: {validation_results['empty_rows']}")
            if critical_issues:
                self.logger.warning("Critical validation issues found:")
                for issue in critical_issues:
                    self.logger.warning(f"  - {issue}")
            else:
                self.logger.info("No critical validation issues found")
            return validation_results, critical_issues
        except Exception as e:
            self.logger.error(f"Error validating output data: {str(e)}")
            return {}, [f"Validation error: {str(e)}"] 

    def create_sorting_summary_report(self, df, sheet_name="Main Output"):
        """Create a summary report of case sorting results"""
        try:
            if 'Status' not in df.columns:
                return
            
            self.logger.info(f"\n=== {sheet_name} - CASE SORTING SUMMARY ===")
            
            # Get status distribution
            status_counts = df['Status'].value_counts()
            total_cases = len(df)
            
            # Calculate percentages
            status_summary = []
            for status, count in status_counts.items():
                if pd.notna(status) and str(status).strip():
                    percentage = (count / total_cases) * 100
                    status_summary.append({
                        'Status': status,
                        'Count': count,
                        'Percentage': f"{percentage:.1f}%"
                    })
            
            # Sort by our custom order
            status_order = {
                'new': 1, 'in_progress': 2, 'in progress': 2, 'in progress today': 2,
                'needs follow up': 3, 'pending': 4, 'escalation': 5, 'closed': 6,
                'Closed': 6, 'bank/sutherland': 7, 'dnd': 8, 'DND': 8
            }
            
            status_summary.sort(key=lambda x: status_order.get(str(x['Status']).lower(), 999))
            
            # Log the summary
            self.logger.info(f"Total cases: {total_cases}")
            self.logger.info("Status distribution (sorted by priority):")
            for item in status_summary:
                self.logger.info(f"  {item['Status']}: {item['Count']} cases ({item['Percentage']})")
            
            # Show first few cases of each status for verification
            self.logger.info("\nFirst few cases by status:")
            for item in status_summary[:5]:  # Show top 5 statuses
                status = item['Status']
                status_cases = df[df['Status'] == status].head(3)
                if not status_cases.empty:
                    case_numbers = status_cases['Case Number'].astype(str).tolist()
                    self.logger.info(f"  {status}: {case_numbers}")
            
            self.logger.info(f"=== END {sheet_name} SORTING SUMMARY ===\n")
            
        except Exception as e:
            self.logger.error(f"Error creating sorting summary report: {str(e)}")

    def _assign_companies_cases(self, df, selected_handlers):
        """
        Assign cases to handlers with round-robin logic based on Email groups.
        Ensures all cases from the same email are assigned to the same handler.
        
        CRITICAL: Only assigns NEW cases. Cases with existing Assigned To from
        previous file are preserved (similar to handler sheet preservation logic).
        """
        try:
            if df.empty or not selected_handlers:
                self.logger.info("No cases or no handlers selected - skipping assignment")
                return df
            
            self.logger.info(f"Processing {len(df)} cases for assignment to {len(selected_handlers)} handlers...")
            
            # Ensure Assigned To column exists
            if 'Assigned To' not in df.columns:
                df['Assigned To'] = ''
            
            # Get valid emails (exclude empty/NaN)
            if 'Email' not in df.columns:
                self.logger.warning("No 'Email' column - cannot group by email")
                return df
            
            # Convert emails to lowercase string for grouping
            df['_email_group'] = df['Email'].astype(str).str.strip().str.lower()
            
            # CRITICAL: Identify which cases need assignment (those without a handler)
            # A case needs assignment if Assigned To is empty
            # Status is NOT checked - preserved cases may have various statuses but still need their handler kept
            needs_new_assignment = (
                df['Assigned To'].isna() | 
                (df['Assigned To'].astype(str).str.strip() == '') |
                (df['Assigned To'].astype(str).str.strip().str.lower().isin(['nan', 'none']))
            )
            
            new_cases_count = needs_new_assignment.sum()
            preserved_cases_count = len(df) - new_cases_count
            
            self.logger.info(f"Companies assignment analysis:")
            self.logger.info(f"  Preserved cases (already have handler): {preserved_cases_count}")
            self.logger.info(f"  New cases needing handler assignment: {new_cases_count}")
            
            if new_cases_count == 0:
                self.logger.info("No new cases need assignment - all handlers preserved from previous file")
                if '_email_group' in df.columns:
                    df = df.drop(columns=['_email_group'])
                return df
            
            # Get only the NEW cases that need assignment
            new_cases_df = df[needs_new_assignment].copy()
            
            # Identify unique email groups in NEW cases
            unique_emails = new_cases_df['_email_group'].unique()
            unique_emails = [e for e in unique_emails if e and e not in ['nan', 'none', '']]
            
            self.logger.info(f"Found {len(unique_emails)} unique email groups in new cases")
            
            # Create a map for assignments
            email_handler_map = {}
            
            # First pass: Check if any email group already has a handler assigned 
            # (from preserved cases with same email)
            for email in unique_emails:
                # Check if this email exists in preserved cases with a handler
                preserved_mask = (~needs_new_assignment) & (df['_email_group'] == email)
                existing_handlers = df.loc[preserved_mask, 'Assigned To'].dropna().unique()
                existing_handlers = [h for h in existing_handlers if h and str(h).strip() != '' and str(h).lower() not in ['nan', 'none']]
                
                if existing_handlers:
                    # Use the existing handler for this email group
                    email_handler_map[email] = existing_handlers[0]
                    self.logger.info(f"  Email '{email[:30]}...' - using preserved handler: {existing_handlers[0]}")
            
            # Second pass: Assign unassigned emails via round-robin
            handler_idx = 0
            assignments_made = 0
            
            for email in unique_emails:
                if email not in email_handler_map:
                    # Get next handler
                    handler = selected_handlers[handler_idx % len(selected_handlers)]
                    email_handler_map[email] = handler
                    handler_idx += 1
                    assignments_made += 1
            
            self.logger.info(f"Made {assignments_made} new email group assignments (round-robin)")
            
            # Apply assignments ONLY to NEW cases
            for email, handler in email_handler_map.items():
                mask = needs_new_assignment & (df['_email_group'] == email)
                df.loc[mask, 'Assigned To'] = handler
                
            # Handle NEW cases with no email (simple round robin)
            no_email_mask = needs_new_assignment & (
                (df['_email_group'].isin(['nan', 'none', ''])) | (df['_email_group'].isna())
            )
            no_email_count = no_email_mask.sum()
            
            if no_email_count > 0:
                self.logger.info(f"Assigning {no_email_count} new cases without email via simple round robin")
                for idx in df[no_email_mask].index:
                    handler = selected_handlers[handler_idx % len(selected_handlers)]
                    df.loc[idx, 'Assigned To'] = handler
                    handler_idx += 1
            
            # Log final assignment distribution
            self.logger.info("Final Companies assignment distribution:")
            for handler in selected_handlers:
                count = (df['Assigned To'] == handler).sum()
                self.logger.info(f"  {handler}: {count} cases")
            
            # Clean up
            if '_email_group' in df.columns:
                df = df.drop(columns=['_email_group'])
                
            return df
            
        except Exception as e:
            self.logger.error(f"Error assigning companies cases: {str(e)}")
            if '_email_group' in df.columns:
                df = df.drop(columns=['_email_group'])
            return df

    def protect_worksheet(self, writer, sheet_name, password='artadmin'):
        """Protect a worksheet with a password, allowing filtering and selection"""
        try:
            if not hasattr(writer, 'book') or sheet_name not in writer.sheets:
                self.logger.warning(f"Cannot protect sheet '{sheet_name}' - sheet or book not found")
                return
            
            worksheet = writer.sheets[sheet_name]
            
            # xlsxwriter worksheet protection
            options = {
                'select_locked_cells': True,
                'select_unlocked_cells': True,
                'format_cells': False,
                'format_columns': False,
                'format_rows': False,
                'insert_columns': False,
                'insert_rows': False,
                'insert_hyperlinks': False,
                'delete_columns': False,
                'delete_rows': False,
                'sort': True, # Allow sorting
                'autofilter': True, # Allow filtering
                'pivot_tables': False,
                'scenarios': False,
                'objects': False,
            }
            
            worksheet.protect(password, options)
            self.logger.info(f"Sheet '{sheet_name}' protected with password")
            
        except Exception as e:
            self.logger.error(f"Error protecting sheet '{sheet_name}': {str(e)}")