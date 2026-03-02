# ============================================================================
# SharedFunctions.py - Shared Utilities for ART Q Control
# ============================================================================
# This module contains shared functions, configuration, and constants used by
# both AutoSender and CaseReviewer modules.
# ============================================================================

import platform
import subprocess
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from datetime import datetime, timedelta
import tkinter as tk
from openpyxl import load_workbook
import time
import json
import os
import pandas as pd
import pyautogui
import traceback
import ctypes
from PyQt5.QtWidgets import (
    QApplication, QDialog, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QGridLayout, QInputDialog, QWidget, QCheckBox, QScrollArea, QProgressBar,
    QFrame, QFileDialog, QTextEdit
)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QIcon
import sys
from config_loader import init_config
from pathlib import Path

# ============================================================================
# CONFIGURATION INITIALIZATION
# ============================================================================
print("[INFO] Initializing configuration system...")
CONFIG_MANAGER = init_config()

# Load configuration values - NO FALLBACKS, strict validation
try:
    AGENT_NAME = CONFIG_MANAGER.get_value('agent_settings', 'agent_name')
    DIALER_USERNAME = CONFIG_MANAGER.get_value('agent_settings', 'user_id')
    DIALER_PASSWORD = CONFIG_MANAGER.get_value('agent_settings', 'password')
    DIALER_PLACE_ID = CONFIG_MANAGER.get_value('agent_settings', 'place_id')
    
    EXCEL_BASE_PATH = CONFIG_MANAGER.get_value('file_paths', 'excel_base_path')
    CACHE_DIRECTORY = CONFIG_MANAGER.get_value('file_paths', 'cache_directory')
    
    EXCEL_SHEET_NAME = CONFIG_MANAGER.get_value('crm_settings', 'excel_sheet_name')
    
    REFRESH_INTERVAL = CONFIG_MANAGER.get_value('execution_settings', 'refresh_interval')
    
    print(f"[INFO] Configuration loaded for agent: {AGENT_NAME}")
    
except Exception as e:
    print(f"[CRITICAL ERROR] Failed to load configuration: {e}")
    sys.exit(1)

# ============================================================================
# HARDCODED SETTINGS (Do not change)
# ============================================================================
DIALER_URL = "https://104.232.254.43/ui/ad/v1/index.html"
EMAIL_DOMAIN = "na_thinkcare@lenovo.com"
EMAIL_DOMAIN_NAME = "NA Think Care"

# ============================================================================
# OS Inhibit constants
# ============================================================================
ES_CONTINUOUS = 0x80000000
ES_SYSTEM_REQUIRED = 0x00000001
ES_DISPLAY_REQUIRED = 0x00000002
_inhibit_active = False

# ============================================================================
# ART Team Variables - Templates
# ============================================================================
SMSText = ("Hello {CX_Name} , \n\nPlease reply with 1, 2, or 3 regarding your recent Lenovo Service for serial number: {case_number} . \n\n1. Issue Resolved\n2. Need Assistance\n3. Stop Text Messages\n\nIf there is no reply, we will contact you by phone within 24 hours between 8:00 AM - 6:00 PM local time.")

CaseEmailOnSite_Depot = (
    " Hello {CX_Name}\n\n"
    "I hope you are doing well.\n\n"
    "I hope your device of Serial Number: {serial_val} is performing well after its recent repair.\n\n"
    "To ensure everything is working as expected, please reply with one of the following options:\n\n"
    "1 for Issue Resolved — Everything is working properly.\n"
    "2 for Need Assistance — Please include details of any ongoing issues.\n"
    "3 for STOP — To stop receiving follow-up messages.\n\n"
    "If we do not receive a response, a member of our team may follow up with a phone call on the next business day (Monday—Friday) between\n"
    "8:00 AM and 6:00 PM local time.\n\n"
    "Thank you for choosing Lenovo Services.\n\n"
    "We appreciate your business and are here to support you.\n\n"
    "Thanks and Regards,\n"
    "{AGENT_NAME}\n"
    "NA Lenovo PC Assurance Resolution Team"
)

CaseEmailCRU = (
    " Hello {CX_Name}\n\n"
    "I hope you are doing well.\n\n"
    "Regarding the device of Serial number:  {serial_val}\n\n"
    "I am contacting you to verify that the Lenovo replacement part is functioning properly and performing to your expectations.\n\n"
    "To ensure everything is working as expected, please reply with one of the following options:\n\n"
    "1 for Issue Resolved — Everything is working properly.\n"
    "2 for Need Assistance — Please include details of any ongoing issues.\n"
    "3 for STOP — To stop receiving follow-up messages.\n\n"
    "If we do not receive a response, a member of our team may follow up with a phone call on the next business day (Monday—Friday) between\n"
    "8:00 AM and 6:00 PM local time.\n\n"
    "Thank you for choosing Lenovo Services.\n\n"
    "We appreciate your business and are here to support you.\n\n"
    "Thanks and Regards,\n"
    "{AGENT_NAME}\n"
    "NA Lenovo PC Assurance Resolution Team"
)

# Dynamic CaseNote - will be formatted when used
def get_case_note(action="Sent SMS  // Sent Email"):
    today_str = datetime.now().strftime("%b %d, %Y")
    return f"Date: {today_str}\nQueue: ART Project - Follow up \nAgent: {AGENT_NAME} \nAction: {action}\n \n ------------------------"

# ============================================================================
# LOCATORS - Centralized for easier maintenance
# ============================================================================
LOCATORS = {
    "ADD_MENU_BUTTON": (
        By.XPATH, 
        "//button[contains(@id, 'notescontrol-action_bar_add_command')]"
    ),
    "EMAIL_FLYOUT_BUTTON": (
        By.XPATH, 
        "//li[contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_email')]"
    ),
    "DISCARD_CHANGES_BUTTON": (
        By.XPATH, 
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]"
    ),
    "CLEAR_FROM_BUTTON": (
        By.XPATH, 
        "//*[contains(@id, 'from.fieldControl-LookupResultsDropdown_from') and contains(@id, 'microsoftIcon_cancelButton_')]"
    ),
    "FROM_EMAIL_INPUT": (
        By.XPATH, 
        "//*[contains(@id, 'from.fieldControl-LookupResultsDropdown_from') and contains(@id, 'textInputBox_with_filter_new')]"
    ),
    "PICK_EMAIL_ENTRY": (
        By.ID, 
        "id-1d5ad078-3edb-4edc-98c6-c0c21e3125e3-45-fromcbfb742c-14e7-4a17-96bb-1a13f7f64aa2-from.fieldControl-name0_0_0" 
    ),
    "EMAIL_BODY_INPUT": (
        By.XPATH, 
        "//div[starts-with(@id, 'rtev')]/p"
    ),
    "SEND_BUTTON": (
        By.XPATH, 
        "//button[starts-with(@id, 'email|NoRelationship|Form|Mscrm.Form.email.Send')]"
    ),
    "NAVIGATE_BACK_BUTTON": (
        By.XPATH, 
        "//button[@id='navigateBackButtontab-id-0']/span"
    )
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def _ensure_dir(path):
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        pass

def update_excel_sheet_safely(excel_path, dataframe, sheet_name):
    """
    Update a specific sheet in an Excel workbook while preserving all other sheets.
    """
    try:
        workbook = load_workbook(excel_path)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        workbook.close()
        
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"[INFO] Excel sheet '{sheet_name}' updated successfully. Other sheets preserved.")
        
    except FileNotFoundError:
        dataframe.to_excel(excel_path, sheet_name=sheet_name, index=False)
        print(f"[INFO] Excel file created: {excel_path}")
    except Exception as e:
        print(f"[ERROR] Failed to update Excel sheet: {e}")
        dataframe.to_excel(excel_path, sheet_name=sheet_name, index=False)

def update_cache_file(cache_path, dataframe, sheet_name):
    """
    Fast update to cache file (much smaller, no other sheets to preserve)
    """
    try:
        dataframe.to_excel(cache_path, sheet_name=sheet_name, index=False)
    except Exception as e:
        print(f"[ERROR] Failed to update cache file: {e}")

def safe_find(driver, by, locator, timeout=20, clickable=False, retries=3, poll=0.5):
    """Find element with retries. Silent retries, clean error output."""
    for attempt in range(1, retries + 1):
        try:
            wait = WebDriverWait(driver, timeout)
            if clickable:
                el = wait.until(EC.element_to_be_clickable((by, locator)))
            else:
                el = wait.until(EC.presence_of_element_located((by, locator)))
            return el
        except Exception:
            time.sleep(poll)

    # All retries failed - show clean error (no stacktrace)
    short_locator = locator[:80] + "..." if len(locator) > 80 else locator
    print(f"[ERROR] Element not found: {short_locator} (after {retries} attempts)")
    
    # Save screenshot silently
    _ensure_dir('errors')
    safe_name = locator.replace('/', '_').replace(' ', '_')[:60]
    img_path = os.path.join('errors', f'failure_{safe_name}.png')
    try:
        driver.save_screenshot(img_path)
    except Exception:
        pass
    
    return None

def click_safe(driver, by, locator, timeout=20, retries=3, poll=0.5):
    el = safe_find(driver, by, locator, timeout=timeout, clickable=True, retries=retries, poll=poll)
    if not el:
        return False
    try:
        el.click()
        return True
    except Exception as e:
        print(f"[ERROR] click_safe failed on {locator}: ")
        return False

def send_keys_safe(driver, by, locator, value, timeout=20, retries=3, poll=0.5, enter=False):
    el = safe_find(driver, by, locator, timeout=timeout, clickable=True, retries=retries, poll=poll)
    if not el:
        return False
    try:
        try:
            el.clear()
        except Exception:
            pass
        el.send_keys(value)
        if enter:
            el.send_keys(Keys.ENTER)
        return True
    except Exception as e:
        print(f"[ERROR] send_keys_safe failed on {locator}: ")
        return False

def enable_windows_inhibit():
    global _inhibit_active
    try:
        if platform.system().lower().startswith('win'):
            ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS | ES_SYSTEM_REQUIRED | ES_DISPLAY_REQUIRED)
            _inhibit_active = True
            print("[INFO] Windows sleep/display inhibit enabled")
    except Exception as e:
        print(f"[WARN] Failed to enable Windows inhibit: ")

def disable_windows_inhibit():
    global _inhibit_active
    try:
        if _inhibit_active and platform.system().lower().startswith('win'):
            ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS)
            _inhibit_active = False
            print("[INFO] Windows sleep/display inhibit disabled")
    except Exception as e:
        print(f"[WARN] Failed to disable Windows inhibit: ")

def Chrome_ART_Profile():
    chrome_options = Options()
    user_data_dir = os.environ.get('CHROME_USER_DATA_DIR') or os.path.join(os.path.expanduser('~'), '@')
    profile_dir = os.environ.get('CHROME_PROFILE_DIR') or 'Default'
    if user_data_dir and os.path.exists(user_data_dir):
        try:
            chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
            chrome_options.add_argument(f'--profile-directory={profile_dir}')
            print(f"[INFO] Using Chrome profile: {user_data_dir}")
        except Exception as e:
            print(f"[WARN] Failed to set Chrome profile: ")
    else:
        try:
            os.makedirs(user_data_dir, exist_ok=True)
            chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
            chrome_options.add_argument(f'--profile-directory={profile_dir}')
            print(f"[INFO] Created new Chrome profile at: {user_data_dir}")
        except Exception as e:
            print(f"[WARN] Failed to create Chrome profile: ")
    return chrome_options

def todays_excel_path():
    now = datetime.now()
    dd = now.strftime('%d')
    mm = now.strftime('%m')
    file_name = f"Active Cases PA {mm}-{dd}.xlsx"
    path = os.path.join(EXCEL_BASE_PATH, file_name)
    return path

def find_column_case_insensitive(df, name):
    for c in df.columns:
        if c.lower() == name.lower():
            return c
    return None

def keep_driver_alive(driver, check_interval=60):
    """
    Keep the Chrome driver alive by refreshing the page.
    This prevents Chrome from going into 'Oh Snap' error state due to inactivity.
    
    Args:
        driver: Selenium WebDriver instance
        check_interval: Not used, kept for backward compatibility
    
    Returns:
        bool: True if refresh succeeded, False otherwise
    """
    try:
        driver.refresh()

        return True
    except Exception as e:

        return False

def wait_for_excel_file(path, check_interval=900        
):
    """
    Wait for Excel file to become available, checking every check_interval seconds.
    """
    while not os.path.exists(path):
        print(f"[INFO] Waiting for Excel file: {path}")
        time.sleep(check_interval)
    print(f"[INFO] Excel file found: {path}")
    return True

def get_todays_cache_path(agent_name, mode="autosender"):
    """
    Get today's working cache file path with date-based naming.
    Format: working_cases_{agent}_{mode}_{MMDD}.xlsx
    
    Args:
        agent_name: Agent's first name (used for file naming)
        mode: "autosender" or "casereviewer"
    
    Returns:
        Full path to today's cache file
    """
    today = datetime.now()
    date_str = today.strftime("%m%d")  # MMDD format
    agent_clean = agent_name.split()[0].replace(' ', '_')
    filename = f"working_cases_{agent_clean}_{mode}_{date_str}.xlsx"
    return os.path.join(CACHE_DIRECTORY, filename)

def check_existing_cache_and_ask(cache_path, mode_name="Auto Sender"):
    """
    Check if today's cache file exists and ask user if they want to resume.
    
    Args:
        cache_path: Path to today's cache file
        mode_name: Display name for the mode ("Auto Sender" or "Case Reviewer")
    
    Returns:
        - "RESUME" = Use existing cache file to continue where left off
        - "NEW" = Create new cache from main Excel file
    """
    if not os.path.exists(cache_path):
        return "NEW"
    
    # File exists, ask user
    class ResumeDialog(QDialog):
        def __init__(self):
            super().__init__()
            self.setWindowTitle(f"Resume {mode_name}?")
            self.setFixedSize(450, 180)
            self.result = "NEW"
            
            layout = QVBoxLayout(self)
            layout.setSpacing(15)
            
            # Message
            msg = QLabel(f"📋 Found existing work from today.\n\nWould you like to resume where you left off?")
            msg.setStyleSheet("font-size: 15px; color: #161616;")
            msg.setWordWrap(True)
            layout.addWidget(msg)
            
            # Buttons
            btn_layout = QHBoxLayout()
            
            resume_btn = QPushButton("✅ Resume")
            resume_btn.setStyleSheet("""
                QPushButton {
                    background-color: #0f62fe;
                    color: white;
                    font-weight: bold;
                    padding: 12px 24px;
                    border-radius: 5px;
                    font-size: 15px;
                }
                QPushButton:hover { background-color: #0353e9; }
            """)
            resume_btn.clicked.connect(self.on_resume)
            btn_layout.addWidget(resume_btn)
            
            new_btn = QPushButton("🔄 Start Fresh")
            new_btn.setStyleSheet("""
                QPushButton {
                    background-color: #e0e0e0;
                    color: #161616;
                    font-weight: bold;
                    padding: 12px 24px;
                    border-radius: 5px;
                    font-size: 15px;
                }
                QPushButton:hover { background-color: #cacaca; }
            """)
            new_btn.clicked.connect(self.on_new)
            btn_layout.addWidget(new_btn)
            
            layout.addLayout(btn_layout)
        
        def on_resume(self):
            self.result = "RESUME"
            self.accept()
        
        def on_new(self):
            self.result = "NEW"
            self.accept()
    
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    dialog = ResumeDialog()
    dialog.exec_()
    return dialog.result

# ============================================================================
# CRM / DYNAMICS FUNCTIONS
# ============================================================================

def case_search_and_open(driver, case_number):
    """Search for and open a case in Dynamics CRM"""
    
    #Clicking Save Button for the previous case (will fail on first case, that's OK)
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckIn.Command') and contains(@id,'-button')]",
        timeout=1,
        retries=2,
    )
    
    #sleep timer
    time.sleep(5)

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(1.5)

    #Search Box - wait for it to be present
    safe_find(driver, By.ID, "GlobalSearchBox", timeout=3, retries=3)

    #Search for the case
    send_keys_safe(driver, By.ID, "GlobalSearchBox", case_number, timeout=3, retries=3)
    
    #click the first result's button to open the case
    click_safe(
        driver,
        By.XPATH,
        "//div[@id='id-globalSearchFlyout-1']/div/div/div/div/div/div[2]/div/button/span",
        timeout=7,
        retries=2,
    )

    # USFC discard dialog (if appears) - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=2,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #Wait timer
    time.sleep(3)

    #Clicking Edit Button before starting
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckOut.Command') and contains(@id,'-button')]",
        timeout=1,
        retries=3,
    )

def solution_provided_check_and_skip(driver, case_number, df, excel_path):
    """Check if case has 'Solution Provided' status - skip if not"""
    solutionProv_case = True
    case_status_xpath = "//div[contains(@id,'headerControlsList')]/div[3]/div/div"
    case_status_el = safe_find(driver, By.XPATH, case_status_xpath, timeout=2, retries=6)
    case_status = case_status_el.text.strip()
    if not case_status.lower() == "solution provided":
        solutionProv_case = False
    return solutionProv_case

def eticket_check_and_skip(driver, case_number, df, excel_path):
    """
    Check if case channel is 'e-ticketing'. If so, update case reason and contact reason fields.
    Returns True if it's an e-ticketing case, False otherwise.
    """
    eticket_case = True
    case_channel_xpath = "//div[contains(@id,'headerControlsList')]/div[7]/div/div"
    case_channel_el = safe_find(driver, By.XPATH, case_channel_xpath, timeout=2, retries=6)
    if case_channel_el:
        case_channel = case_channel_el.text.strip()
        if not case_channel.lower() == "e-ticketing":
            eticket_case = False
        else:
            eticket_case = True
            # 1. Go to inprogress flyout menu
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id, 'MscrmControls.Containers.ProcessBreadCrumb-stageIndicatorContainer2d827664-559a-4937-beb1-52b3f836b967')]/div",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 2. Check the flyout menu of the aria-label = "Case Reason"
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id, 'header_process_casetypecode-header_process_casetypecode') and contains(@id, 'header_process_casetypecode.fieldControl-pcf-container-id')]/div/div/div/div/div/button",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 3. Choose "Service Call"
            click_safe(
                driver,
                By.XPATH,
                "//*[contains(@id, 'fluent-option') and text()='Service Call']",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 4. Check the flyout menu of aria-label "Contact Reason"
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id, 'header_process_lvidg_callreasoncode-header_process_lvidg_callreasoncode') and contains(@id, 'header_process_lvidg_callreasoncode.fieldControl-pcf-container-id')]/div/div/div/div/div/button",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 5. Choose "Hardware Repair"
            click_safe(
                driver,
                By.XPATH,
                "//*[contains(@id, 'fluent-option') and text()='Hardware Repair']",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 6. check the flyout menu of aria-label "Tech Savvy"
            click_safe(
                driver,
                By.XPATH,
                "//div[contains(@id,'header_process_lvidg_techsavvy-header_process_lvidg_techsavvy') and contains(@id, 'header_process_lvidg_techsavvy.fieldControl-pcf-container-id')]/div/div/div/div/div/button",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 7. Choose "Undetermined"
            click_safe(
                driver,
                By.XPATH,
                "//*[contains(@id, 'fluent-option') and text()='Undetermined']",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 8. close the flyout menu
            click_safe(
                driver,
                By.XPATH,
                "//button[contains(@id,'MscrmControls.Containers.ProcessStageControl-stageContentClose') and @aria-label='Close']",
                timeout=5,
                retries=3
            )
            time.sleep(1)
            # 9. Clicking Save Button
            click_safe(
                driver,
                By.XPATH,
                "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckIn.Command') and contains(@id,'-button')]",
                timeout=1,
                retries=2,
            )
            time.sleep(2.5)
            # 10. Clicking Edit Button before resuming
            click_safe(
                driver,
                By.XPATH,
                "//button[contains(@id,'incident|NoRelationship|Form|lvdcg.incident.TimeTrackingCheckOut.Command') and contains(@id,'-button')]",
                timeout=1,
                retries=3,
            )
            time.sleep(2)

    else:
        eticket_case = False
    return eticket_case

def solution_provided_check_and_skip_companies(driver, case_number, df, excel_path):
    """
    Check if case has 'Solution Provided' OR 'Closed' status - for Companies Process.
    Returns True if case should be added to batch (Solution Provided or Closed), False otherwise.
    """
    solution_or_closed_case = False
    case_status_xpath = "//div[contains(@id,'headerControlsList')]/div[3]/div/div"
    case_status_el = safe_find(driver, By.XPATH, case_status_xpath, timeout=2, retries=6)
    
    if case_status_el:
        case_status = case_status_el.text.strip().lower()
        # Check if case status is either "Solution Provided" or "Closed"
        if case_status in ["solution provided", "closed"]:
            solution_or_closed_case = True
            print(f"[INFO] ✓ Case {case_number}: Status '{case_status}' - Added to batch")
        else:
            solution_or_closed_case = False
            print(f"[INFO] ✗ Case {case_number}: Status '{case_status}' - NOT added to batch")
    else:
        print(f"[WARN] Could not find case status element for {case_number}")
        solution_or_closed_case = False
    
    return solution_or_closed_case

def serial_extraction(driver, case_number, df):
    """Extract serial number from case"""
    serial_val = ''
    try:
        serial_xpath = "//*[contains(@id,'productserialnumber.fieldControl-text-input-component')]"
        el_serial = safe_find(driver, By.XPATH, serial_xpath, timeout=2, retries=2)
        if el_serial:
            try:
                serial_val = el_serial.get_attribute('value') or el_serial.text or ''
            except Exception:
                try:
                    serial_val = el_serial.text or ''
                except Exception:
                    serial_val = ''
    except Exception as e:
        print(f"[WARN] Could not read serial number for {case_number}: ")

    try:
        serial_val = str(serial_val).strip()
    except Exception:
        serial_val = ''

    return serial_val

def customer_name_extraction(driver, case_number):
    """Extract customer name from case"""
    CX_Name = ''
    try:
        name_el = safe_find(driver, By.XPATH, "//*[contains(@id,'sec_tab_contact-associatedEntityRecordName')]", timeout=2, retries=2)
        if name_el:
            try:
                raw_name = name_el.get_attribute('aria-label') or name_el.get_attribute('title') or name_el.text or ''
            except Exception:
                raw_name = name_el.text or ''
            try:
                CX_Name = ' '.join([w.capitalize() for w in str(raw_name).split() if w])
            except Exception:
                CX_Name = str(raw_name)
    except Exception as e:
        print(f"[WARN] Could not read customer name for {case_number}: ")

    if not CX_Name:
        CX_Name = 'Our Valued Customer'

    return CX_Name

def formatting_texts_sms(CX_Name, serial_val, case_number, df):
    """Format SMS text for the case"""
    try:
        sms_text = SMSText.format(CX_Name=CX_Name, case_number=case_number)
    except Exception:
        sms_text = SMSText.format(CX_Name='Our Valued Customer', case_number=case_number)
    return sms_text

def formatting_texts_email(CX_Name, serial_val, case_number, df):
    """Format email text based on Work Order Type"""
    chosen_template = None

    try:
        mask = df["case_number"].astype(str) == str(case_number)
        if mask.any() and "Work Order Type" in df.columns:
            wot_val = df.loc[mask, "Work Order Type"].iloc[0]
            if pd.isna(wot_val) or str(wot_val).strip() == "":
                wot = ""
            else:
                wot = str(wot_val).strip()
        else:
            wot = ""
    except Exception:
        wot = ""

    if wot.lower() in ("onsite", "depot"):
        chosen_template = CaseEmailOnSite_Depot
    elif wot.lower() == "cru":
        chosen_template = CaseEmailCRU
    else:
        chosen_template = CaseEmailOnSite_Depot

    email_text = chosen_template.format(
        CX_Name=CX_Name or "Our Valued Customer",
        serial_val=serial_val,
        AGENT_NAME=AGENT_NAME
    )

    return email_text

def send_SMS(driver, sms_text):
    """Send SMS via Dynamics CRM"""
    sms_sent = False

    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
        timeout=2,
        retries=5,
    )

    click_SMS_button = click_safe(
        driver,
        By.XPATH,
        "//li[contains(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_lvidg_sms')]",
        timeout=2,
        retries=5,
    )

    if not click_SMS_button:
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
            timeout=1,
            retries=5,
        )

        click_SMS_button = click_safe(
            driver,
            By.XPATH,
            "//li[contains(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-createNewRecord_flyoutMenuItem_lvidg_sms')]",
            timeout=2,
            retries=5,
        )
        
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
            timeout=1,
            retries=5,
        )

        #clicking OK Button if avail
        click_safe(
            driver,
            By.XPATH,
            "//button[contains(@id,'okButton_')]/div",
            timeout=2,
            retries=2,
        )

        #sleep timer
        time.sleep(3)

    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    send_keys_safe(
        driver,
        By.XPATH,
        "//input[starts-with(@id,'id-3145bfd3-91e7-4364-92ed-5ca0cf0d65b8') and contains(@id,'subject.fieldControl-text-input-component')]",
        sms_text,
        timeout=2,
        retries=5,
    )

    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'lvidg_sms|NoRelationship|Form|Mscrm.SaveAndClosePrimary') and contains(@id, '-button')]",
        timeout=2,
        retries=5,
    )

    SMS_error_el = safe_find(driver, By.XPATH, "//div[contains(@id,'message') and contains(@id,'+lvidg_phonenumber')]/div/span", timeout=3, retries=2)
    if SMS_error_el:
        click_safe(
            driver,
            By.XPATH,
            "//button[@id='navigateBackButtontab-id-0']/span",
            timeout=1,
            retries=5,
        )
        sms_sent = False
    else:
        sms_sent = True

    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    return sms_sent

def verify_email_domain_selected(driver, expected_title="NA Think Care", expected_text="NA Think Care", timeout=3, poll=0.5, retries=3):
    xpath = "//*[@data-id='from.fieldControl-LookupResultsDropdown_from_selected_tag_text']"
    def _check(d):
        try:
            el = d.find_element(By.XPATH, xpath)
            title = (el.get_attribute('title') or '').strip()
            text = (el.text or '').strip()
            return title == expected_title and text == expected_text
        except Exception:
            return False
    return _check(driver)

def select_from_email(driver):
    """Handle selecting the 'From' email address"""
    click_safe(driver, *LOCATORS["CLEAR_FROM_BUTTON"], timeout=1, retries=3)
    send_keys_safe(driver, *LOCATORS["FROM_EMAIL_INPUT"], EMAIL_DOMAIN, timeout=2, retries=5)

    if not click_safe(driver, *LOCATORS["PICK_EMAIL_ENTRY"], timeout=5, retries=5):
        return False

    time.sleep(2)

    return verify_email_domain_selected(
        driver, 
        expected_title=EMAIL_DOMAIN_NAME, 
        expected_text=EMAIL_DOMAIN_NAME,
        timeout=5, 
        poll=0.5, 
        retries=3
    )

def send_Email(driver, email_text):
    """Send email via Dynamics CRM"""
    if not click_safe(driver, *LOCATORS["ADD_MENU_BUTTON"], timeout=5, retries=3):
        print("Error: Could not open Add Menu.")
        return False
    
    if not click_safe(driver, *LOCATORS["EMAIL_FLYOUT_BUTTON"], timeout=5, retries=5):
        print("Warning: Email button failed, retrying menu open and click.")
        click_safe(driver, *LOCATORS["ADD_MENU_BUTTON"], timeout=5, retries=3)
        if not click_safe(driver, *LOCATORS["EMAIL_FLYOUT_BUTTON"], timeout=5, retries=5):
            print("Error: Could not click Email Button after retry.")
            return False

    click_safe(driver, *LOCATORS["DISCARD_CHANGES_BUTTON"], timeout=1, retries=3) 
    time.sleep(2)

    if not select_from_email(driver):
        print("Error: Failed to select 'From' email address after multiple attempts.")
        click_safe(driver, *LOCATORS["NAVIGATE_BACK_BUTTON"], timeout=3, retries=3)
        click_safe(driver, *LOCATORS["DISCARD_CHANGES_BUTTON"], timeout=1, retries=3)
        return False
    
    if not send_keys_safe(driver, *LOCATORS["EMAIL_BODY_INPUT"], email_text, timeout=5, retries=5):
        print("Error: Failed to enter email body text.")
        return False

    if not click_safe(driver, *LOCATORS["SEND_BUTTON"], timeout=10, retries=5):
        print("Error: Failed to click Send Email button.")
        return False
    
    time.sleep(3)
    return True

def add_Case_Note(driver, CaseNote):
    """Add a case note in Dynamics CRM"""
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
        timeout=1,
        retries=5,
    )

    note_button = click_safe(
        driver,
        By.XPATH,
        "//li[starts-with(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'notescontrol-createNewRecord_flyoutMenuItem_notes')]",
        timeout=1,
        retries=5,
    )
    
    if not note_button:
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id, 'notescontrol-action_bar_add_command')]",
            timeout=1,
            retries=5,
        )

        note_button = click_safe(
            driver,
            By.XPATH,
            "//li[starts-with(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'notescontrol-createNewRecord_flyoutMenuItem_notes')]",
            timeout=1,
            retries=5,
        )
        
        click_safe(
            driver,
            By.XPATH,
            "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
            timeout=1,
            retries=5,
        )

        #clicking OK Button if avail
        click_safe(
            driver,
            By.XPATH,
            "//button[contains(@id,'okButton_')]/div",
            timeout=2,
            retries=2,
        )

        #sleep timer
        time.sleep(3)

    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=5,
    )

    #clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    #sleep timer
    time.sleep(3)

    send_ok = send_keys_safe(driver, By.XPATH, "//*[contains(@id, 'rtev')]", CaseNote, timeout=2, retries=5)
    time.sleep(2)

    note_saved = False
    try:
        saved = click_safe(driver, By.XPATH, "//span[contains(@class,'ms-Button-label') and text()='Add note and close']", timeout=2, retries=4)
        time.sleep(2)
        if saved:
            note_saved = True
        else:
            try:
                pyautogui.press(['tab', 'tab', 'enter'])
                time.sleep(2)
                note_saved = True
            except Exception as e:
                print(f"[WARN] Failed to send Tab/Tab/Enter: ")
    except Exception as e:
        print(f"[WARN] Save note step failed: ")
    
    return note_saved

def DND_CX(driver, case_number):
    """Mark contact as DND (Do Not Disturb)"""
    contact_xpath = "//a[contains(@id,'id-915f6055-2e07-4276-ae08-2b96c8d02c57') and contains(@id,'sec_tab_contact-associatedEntityRecordName')]"
    if not click_safe(driver, By.XPATH, contact_xpath, timeout=1, retries=5):
        print(f"[WARN] Contact card not found for {case_number}")
    else:
        edit_xpath = "//button[contains(@id,'contact|NoRelationship|Form|lvidg.contact.TimeTrackingCheckOut.Command') and contains(@id,'-button')]"
        click_safe(driver, By.XPATH, edit_xpath, timeout=1, retries=5)

        details_xpath = "//li[contains(@id,'tab4') and @title='Details']"
        click_safe(driver, By.XPATH, details_xpath, timeout=1, retries=5)
        
        comm_pref_xpath = "//div[contains(@id,'lvidg_hasdonotdisturbf9a8a302')]/div[3]/div/div"
        click_safe(driver, By.XPATH, comm_pref_xpath, timeout=1, retries=5)

        time.sleep(2)

        save_xpath = "//button[contains(@id,'contact|NoRelationship|Form|lvidg.contact.TimeTrackingCheckIn.Command') and contains(@id,'button')]"
        click_safe(driver, By.XPATH, save_xpath, timeout=1, retries=5)

def excelCaseClosingCode(CaseClosingCode):
    """Convert closing code to Excel-friendly format"""
    match CaseClosingCode:
        case "Issue Resolved":
            return "Fixed"
        case "Issue Not Fixed":
            return "Issue Not Fixed"
        case "Customer Not Reached":
            return "Not Reached"
        case "Customer Claims that the Machine Not Yet Tested":
            return "Not yet Tested"
        case "Not Yet Tested":
            return "Not yet Tested"
        case "DND":
            return "DND"
        case "Escalated":
            return "Escalation"
        case "Called - Company NO. Extension Found: Not Reached":
            return "Not Reached"
        case "Called: Not Reached // left Voicemail":
            return "Not Reached"
        case "Called - Answered: Issue Resolved":
            return "Fixed"
        case "Called - Answered: Issue Not Resolved":
            return "Issue Not Fixed"
        case "Need Third Action":
            return "In Progress"
        case "New":
            return "New"
        case "Skipped":
            return "Skipped"
        case "Send SMS":
            return "Send SMS"
        case "Send Email":
            return "Send Email"
        case "Send SMS and Email":
            return "Send SMS and Email"
        case "Call the Customer":
            return "Called the Customer"
        case "Left Voicemail":
            return "Not Reached"
        case _:
            return ""

# ============================================================================
# DIALER FUNCTIONS
# ============================================================================

def switch_to_crm_window(driver, max_retries=5):
    """
    Robustly switch to CRM window using multiple strategies.
    Returns True if successful, False otherwise.
    """
    crm_url_patterns = ["dynamics.microsoft.com", "crm", "power"]
    
    for attempt in range(1, max_retries + 1):
        wait_time = 2 ** (attempt - 1)
        if attempt > 1:
            print(f"[WARN] Switch attempt {attempt}/{max_retries}, waiting {wait_time}s before retry...")
            time.sleep(wait_time)
        
        handles = driver.window_handles

        if len(handles) < 2:
            print(f"[ERROR] Expected 2+ windows, but found {len(handles)}")
            continue
        
        crm_handle = None
        for handle in handles:
            try:
                driver.switch_to.window(handle)
                current_url = driver.current_url.lower()

                if any(pattern in current_url for pattern in crm_url_patterns):
                    crm_handle = handle
                    print(f"[INFO] Found CRM window by URL pattern")
                    break
            except Exception as e:
                print(f"[WARN] Error checking handle URL: {e}")
                continue
        
        if crm_handle is None and len(handles) >= 2:

            for handle in handles:
                try:
                    driver.switch_to.window(handle)
                    crm_el = driver.find_elements(By.ID, "GlobalSearchBox")
                    if crm_el:
                        crm_handle = handle
                        print(f"[INFO] Found CRM window by GlobalSearchBox element")
                        break
                except Exception as e:
                    continue
        
        if crm_handle is None:

            for idx, handle in enumerate(handles):
                if idx != 0:
                    try:
                        driver.switch_to.window(handle)
                        crm_handle = handle
                        print(f"[INFO] Switched to handle index {idx} (fallback)")
                        break
                    except Exception as e:
                        continue
        
        if crm_handle is not None:
            try:
                driver.switch_to.window(crm_handle)
                try:
                    WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.ID, "GlobalSearchBox"))
                    )
                    print(f"[SUCCESS] ✓ Successfully switched to CRM window and verified GlobalSearchBox")
                    return True
                except TimeoutException:
                    print(f"[WARN] Switched to handle but GlobalSearchBox not found yet, may still be loading")
                    time.sleep(2)
                    try:
                        driver.find_element(By.ID, "GlobalSearchBox")
                        print(f"[SUCCESS] ✓ CRM window verified after additional wait")
                        return True
                    except:
                        print(f"[WARN] GlobalSearchBox still not found, continuing retry loop")
                        continue
            except Exception as e:
                print(f"[ERROR] Failed to switch to identified CRM handle: {e}")
                continue
    
    print(f"[CRITICAL] Failed to switch to CRM window after {max_retries} attempts")
    return False

def perform_dialer_login(driver):
    """
    Login to dialer.
    Args:
        driver: Selenium WebDriver
     
    """
    print("[INFO] Opening dialer...")
    driver.get(DIALER_URL)
    time.sleep(3)
    
    print("[INFO] Entering username...")
    send_keys_safe(driver, By.ID, 'wweLoginUserNameField', DIALER_USERNAME, timeout=2, retries=120)
    
    print("[INFO] Entering password...")
    send_keys_safe(driver, By.ID, 'wweLoginPasswordField', DIALER_PASSWORD, timeout=2, retries=120, enter=True)
    
    print("[INFO] Entering placeholder...")
    send_keys_safe(driver, By.ID, 'wweLoginPlaceInput', DIALER_PLACE_ID, timeout=2, retries=120, enter=True)
    time.sleep(10)

    handles = driver.window_handles
    dialer_handle = handles[0]
    driver.switch_to.window(dialer_handle)
    
    crm_switched = switch_to_crm_window(driver)
    
    if not crm_switched:
        print(f"[ERROR] Failed to switch to CRM window after multiple attempts. Program may be unstable.")
        return False
    
    return True

def perform_call_flow(driver):
    """Perform call flow via dialer"""
    try:
        mobile_input = safe_find(driver, By.XPATH, "//input[@type='tel' and @aria-label='Mobile Phone']", timeout=1, retries=7)
        if mobile_input:
            mobile_value = (mobile_input.get_attribute('value') or '').strip()
            if not mobile_value:
                mobile_value = (mobile_input.get_attribute('title') or '').strip()
        else:
            mobile_value = ''

        if mobile_value:
            number_core = mobile_value.lstrip('+')
            dialed_number = f"9{number_core}"
        else:
            print("[WARN] No mobile phone value found to dial")
            return False
        
        handles = driver.window_handles
        if len(handles) > 0:
            dialer_handle = handles[0]
            try:
                driver.switch_to.window(dialer_handle)
                print("[INFO] Switched to dialer window for call")
            except Exception as e:
                print(f"[WARN] Failed to switch to dialer: {e}")
                return False
        else:
            print("[ERROR] No dialer window found")
            return False

        click_safe(driver, By.XPATH, '//span[contains(@class,"wwe-sprite-mark-done")]', timeout=1, retries=1)
        send_keys_safe(driver, By.ID, 'wweTeamCommunicatorDialerField', dialed_number, timeout=3, retries=5, poll=0.5, enter=True)
        
        print(f"[INFO] Dialed number: {dialed_number}")
        
        time.sleep(5)
        
        print("[INFO] Switching back to CRM window after call...")
        success = switch_to_crm_window(driver, max_retries=5)
        
        if not success:
            print("[ERROR] Failed to switch back to CRM window after call")
            return False
        
        print("[SUCCESS] Successfully returned to CRM window")
        return True
        
    except Exception as e:
        print(f"[WARN] perform_call_flow encountered an error: {e}")
        traceback.print_exc()
        return False

# ============================================================================
# FILE SEARCH POPUP - Shows when file is not found with retry countdown
# ============================================================================
def show_file_search_popup(excel_path, retry_interval_seconds=10):
    """
    Show a popup when the Excel file is not found.
    Features: live countdown, status indicator, Abort button, Manual file button.
    
    Returns:
        - ("FOUND", excel_path) = File was found at configured path
        - ("MANUAL", manual_path) = User selected a manual file (one-time use)
        - ("ABORT", None) = User clicked abort
    """
    result = {"action": None, "path": None}
    
    class FileSearchDialog(QDialog):
        def __init__(self, file_path, retry_seconds):
            super().__init__()
            self.file_path = file_path
            self.retry_seconds = retry_seconds
            self.countdown = retry_seconds
            self.found = False
            
            self.setWindowTitle("Searching for File")
            self.resize(500, 200)
            self.setModal(True)
            
            layout = QVBoxLayout(self)
            layout.setContentsMargins(20, 20, 20, 20)
            layout.setSpacing(15)
            
            # Status label
            self.status_label = QLabel("🔍 Searching for Excel file...")
            self.status_label.setStyleSheet("font-weight: bold; font-size: 15px; color: #0F62FE;")
            self.status_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(self.status_label)
            
            # File path label
            path_label = QLabel(f"Path: {file_path}")
            path_label.setStyleSheet("font-size: 12px; color: #525252;")
            path_label.setWordWrap(True)
            layout.addWidget(path_label)
            
            # Countdown label
            self.countdown_label = QLabel(f"Next check in: {self.countdown} seconds")
            self.countdown_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #161616;")
            self.countdown_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(self.countdown_label)
            
            # Status indicator
            self.found_label = QLabel("Status: ❌ Not Yet Found")
            self.found_label.setStyleSheet("font-size: 15px; color: #DA1E28;")
            self.found_label.setAlignment(Qt.AlignCenter)
            layout.addWidget(self.found_label)
            
            # Buttons
            btn_layout = QHBoxLayout()
            
            abort_btn = QPushButton("⛔ Abort")
            abort_btn.setStyleSheet("""
                QPushButton {
                    background-color: #DA1E28;
                    color: white;
                    font-weight: bold;
                    padding: 12px 24px;
                    border-radius: 5px;
                    font-size: 15px;
                }
                QPushButton:hover { background-color: #A2191F; }
            """)
            abort_btn.clicked.connect(self.on_abort)
            btn_layout.addWidget(abort_btn)
            
            manual_btn = QPushButton("📂 Load File Manually")
            manual_btn.setStyleSheet("""
                QPushButton {
                    background-color: #0F62FE;
                    color: white;
                    font-weight: bold;
                    padding: 12px 24px;
                    border-radius: 5px;
                    font-size: 15px;
                }
                QPushButton:hover { background-color: #0043CE; }
            """)
            manual_btn.clicked.connect(self.on_manual)
            btn_layout.addWidget(manual_btn)
            
            layout.addLayout(btn_layout)
            
            # Timer for countdown and file check
            self.timer = QTimer(self)
            self.timer.timeout.connect(self.tick)
            self.timer.start(1000)  # Every 1 second
            
            # Initial check
            self.check_file()
        
        def check_file(self):
            if os.path.exists(self.file_path):
                self.found = True
                self.found_label.setText("Status: ✅ Found!")
                self.found_label.setStyleSheet("font-size: 15px; color: #24A148; font-weight: bold;")
                self.status_label.setText("✅ Excel file found!")
                self.status_label.setStyleSheet("font-weight: bold; font-size: 15px; color: #24A148;")
                self.countdown_label.setText("Proceeding in 2 seconds...")
                self.timer.stop()
                # Wait 2 seconds then proceed
                QTimer.singleShot(2000, self.on_found)
        
        def tick(self):
            self.countdown -= 1
            if self.countdown <= 0:
                self.countdown = self.retry_seconds
                self.check_file()
            self.countdown_label.setText(f"Next check in: {self.countdown} seconds")
        
        def on_found(self):
            result["action"] = "FOUND"
            result["path"] = self.file_path
            self.accept()
        
        def on_abort(self):
            self.timer.stop()
            result["action"] = "ABORT"
            result["path"] = None
            self.accept()
        
        def on_manual(self):
            self.timer.stop()
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)"
            )
            if file_path:
                result["action"] = "MANUAL"
                result["path"] = file_path
                self.accept()
            else:
                # User cancelled file dialog, resume countdown
                self.timer.start(1000)
    
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    dialog = FileSearchDialog(excel_path, retry_interval_seconds)
    dialog.exec_()
    
    return result["action"], result["path"]

# ============================================================================
# COMPANIES PROCESS - Functions for processing company cases
# ============================================================================

class CompaniesProcessDialog(QDialog):
    """Dialog asking if user wants to process company cases first"""
    def __init__(self, total_cases, distinct_emails):
        super().__init__()
        self.setWindowTitle("Companies Process")
        self.setFixedSize(400, 180)
        self.result = "SKIP"
        
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        
        # Title
        title = QLabel("🏢 Company Cases Available")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #0f62fe;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Stats
        stats = QLabel(f"Total Cases: {total_cases}  |  Companies (Emails): {distinct_emails}")
        stats.setStyleSheet("font-size: 15px; color: #161616;")
        stats.setAlignment(Qt.AlignCenter)
        layout.addWidget(stats)
        
        # Description
        desc = QLabel("Process company cases grouped by email before reviewing individual cases?")
        desc.setStyleSheet("font-size: 15px; color: #525252;")
        desc.setWordWrap(True)
        desc.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        yes_btn = QPushButton("✅ Yes, Process Companies")
        yes_btn.setStyleSheet("""
            QPushButton {
                background-color: #0f62fe;
                color: white;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 15px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #0353e9; }
        """)
        yes_btn.clicked.connect(self.on_yes)
        btn_layout.addWidget(yes_btn)
        
        skip_btn = QPushButton("⏭️ Skip to Case Reviewer")
        skip_btn.setStyleSheet("""
            QPushButton {
                background-color: #e0e0e0;
                color: #161616;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 15px;
            }
            QPushButton:hover { background-color: #cacaca; }
        """)
        skip_btn.clicked.connect(self.on_skip)
        btn_layout.addWidget(skip_btn)
        
        layout.addLayout(btn_layout)
    
    def on_yes(self):
        self.result = "YES"
        self.accept()
    
    def on_skip(self):
        self.result = "SKIP"
        self.accept()

def load_companies_for_handler(cache_file, handler_name, sheet_name="Companies"):
    """Load Companies sheet cases assigned to handler, grouped by email"""
    try:
        df = pd.read_excel(cache_file, sheet_name=sheet_name)
        
        # Find columns
        email_col = find_column_case_insensitive(df, 'Email')
        assigned_col = find_column_case_insensitive(df, 'Assigned To')
        case_col = find_column_case_insensitive(df, 'Case Number')
        serial_col = find_column_case_insensitive(df, 'Serial Number')
        mtm_col = find_column_case_insensitive(df, 'Product ID (MTM)') or find_column_case_insensitive(df, 'Product ID')
        status_col = find_column_case_insensitive(df, 'Status')
        company_name_col = find_column_case_insensitive(df, 'Company Name')
        state_province_col = find_column_case_insensitive(df, 'State/Province')
        phone_col = find_column_case_insensitive(df, 'Phone')
        
        if not email_col or not case_col:
            print("[WARN] Companies sheet missing required columns")
            return {}, df
        
        # Process ALL cases in Companies sheet (filtering was already done when cache was created)
        # This ensures the process proceeds even if Assigned To format differs
        df_handler = df.copy()
        
        # Group by email
        grouped = {}
        for idx, row in df_handler.iterrows():
            email = str(row.get(email_col, '')).strip().lower()
            if not email or email == 'nan':
                continue
            
            case_num = str(row.get(case_col, '')).strip()
            serial = str(row.get(serial_col, '')).strip() if serial_col else ''
            mtm = str(row.get(mtm_col, '')).strip() if mtm_col else ''
            status = str(row.get(status_col, '')).strip().lower() if status_col else ''
            company_name = str(row.get(company_name_col, 'Unknown Company')).strip() if company_name_col else 'Unknown Company'
            state_province = str(row.get(state_province_col, '')).strip() if state_province_col else ''
            phone = str(row.get(phone_col, '')).strip() if phone_col else ''
            
            if email not in grouped:
                grouped[email] = {'cases': [], 'df_indices': []}
            
            grouped[email]['cases'].append({
                'case_number': case_num,
                'serial': serial if serial and serial.lower() != 'nan' else '',
                'mtm': mtm if mtm and mtm.lower() != 'nan' else '',
                'status': status,
                'company_name': company_name if company_name != 'nan' else 'Unknown Company',
                'state_province': state_province if state_province != 'nan' else '',
                'phone': phone if phone != 'nan' else '',
                'row_idx': idx
            })
            grouped[email]['df_indices'].append(idx)
        
        return grouped, df
    except Exception as e:
        print(f"[ERROR] Failed to load companies for handler: {e}")
        return {}, pd.DataFrame()

def build_companies_email_body(cases_info, agent_name):
    """Build email body with Case Number | Serial/MTM format matching CRU/Onsite template"""
    devices_list = []
    for case in cases_info:
        case_num = case.get('case_number', '')
        serial = case.get('serial', '')
        mtm = case.get('mtm', '')
        
        if serial:
            devices_list.append(f"{case_num} | {serial}")
        elif mtm:
            devices_list.append(f"{case_num} | {mtm}")
        else:
            devices_list.append(f"{case_num} | (No Serial/MTM)")
    
    devices_str = "\n".join(devices_list)
    
    email_body = (
        " Hello All\n\n"
        "I hope you are doing well.\n\n"
        f"I am contacting you regarding the devices:\n\n"
        f"{devices_str}\n\n"
        "I am contacting you to verify that the devices are functioning properly and performing to your expectations.\n\n"
        "To ensure everything is working as expected, please reply with one of the following options:\n\n"
        "1 for Issue Resolved — Everything is working properly.\n"
        "2 for Need Assistance — Please include details of any ongoing issues.\n"
        "3 for STOP — To stop receiving follow-up messages.\n\n"
        "You can reply to this email if you missed our call.\n\n"
        "Thank you for choosing Lenovo Services.\n\n"
        "Thanks and Regards,\n"
        f"{agent_name}\n"
        "NA Lenovo PC Assurance Resolution Team"
    )
    return email_body

def show_companies_email_confirmation(email, cases_info, email_body, cases_left=0, email_groups_left=0):
    """Show email preview with confirmation - IBM Carbon Design"""

    class EmailConfirmDialog(QDialog):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Confirm Email  —  Company Batch")
            self.setMinimumWidth(580)
            self.resize(640, 580)
            self.confirmed = False

            # ── IBM tokens ──────────────────────────────────────────────────
            try:
                from ibm_theme import IBM, get_qss, _read_font_size
                _c = IBM.LIGHT
                _fs = _read_font_size()
                self.setStyleSheet(get_qss('light', _fs))
            except Exception:
                _c = {
                    'bg': '#f4f4f4', 'layer_01': '#ffffff', 'layer_02': '#f4f4f4',
                    'text_primary': '#161616', 'text_secondary': '#525252',
                    'text_on_color': '#ffffff', 'interactive': '#0f62fe',
                    'interactive_hover': '#0353e9', 'border_subtle': '#e0e0e0',
                    'teal': '#005d5d', 'teal_hover': '#004144',
                    'success': '#198038', 'danger': '#da1e28',
                    'disabled_bg': '#c6c6c6', 'text_disabled': '#a8a8a8',
                }
                _fs = 13

            from PyQt5.QtGui import QFont
            self.setFont(QFont('IBM Plex Sans', _fs))

            layout = QVBoxLayout(self)
            layout.setContentsMargins(24, 20, 24, 20)
            layout.setSpacing(12)


            # ── EMAIL HEADER ─────────────────────────────────────────────────
            hdr = QLabel(f"Sending to:  {email}")
            hdr.setFont(QFont('IBM Plex Sans', _fs, QFont.Bold))
            hdr.setStyleSheet(
                f"font-weight: 700; color: {_c['interactive']};"
                f" background: transparent; border: none;"
            )
            layout.addWidget(hdr)

            # ── EMAIL BODY PREVIEW ────────────────────────────────────────────
            preview = QTextEdit()
            preview.setPlainText(email_body)
            preview.setReadOnly(True)
            preview.setFont(QFont('IBM Plex Mono', _fs - 2))
            preview.setStyleSheet(
                f"QTextEdit {{ background: {_c['layer_02']};"
                f" border: 1px solid {_c['border_subtle']};"
                f" border-radius: 4px;"
                f" font-family: 'IBM Plex Mono','Courier New',monospace;"
                f" font-size: {_fs - 2}pt;"
                f" color: {_c['text_primary']};"
                f" padding: 8px; }}"
            )
            layout.addWidget(preview)

            # ── CONFIRM BUTTON ────────────────────────────────────────────────
            confirm_btn = QPushButton("I Understood  —  Confirm and Send Email")
            confirm_btn.setFont(QFont('IBM Plex Sans', _fs, QFont.Bold))
            confirm_btn.setMinimumHeight(48)
            confirm_btn.setStyleSheet(
                f"QPushButton {{ background-color: {_c['interactive']};"
                f" color: #ffffff; border: none; border-radius: 8px;"
                f" font-weight: 700; font-size: {_fs}pt;"
                f" padding: 12px 28px; letter-spacing: 0.3px; }}"
                f"QPushButton:hover {{ background-color: {_c['interactive_hover']}; }}"
                f"QPushButton:pressed {{ background-color: #002d9c; }}"
            )
            confirm_btn.clicked.connect(self.on_confirm)
            layout.addWidget(confirm_btn)

        def on_confirm(self):
            self.confirmed = True
            self.accept()

    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)

    dialog = EmailConfirmDialog()
    dialog.exec_()
    return dialog.confirmed

def case_search_and_open_no_edit(driver, case_number):
    """
    Search for and open a case in Dynamics CRM WITHOUT clicking Edit button.
    Used in AutoSender to check Solution Provided before editing.
    """

    # clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    time.sleep(1)

    # Search Box - wait for it to be present
    safe_find(driver, By.ID, "GlobalSearchBox", timeout=3, retries=3)

    # Search for the case
    send_keys_safe(driver, By.ID, "GlobalSearchBox", case_number, timeout=3, retries=3)
    
    # click the first result's button to open the case
    click_safe(
        driver,
        By.XPATH,
        "//div[@id='id-globalSearchFlyout-1']/div/div/div/div/div/div[2]/div/button/span",
        timeout=7,
        retries=2,
    )

    # USFC discard dialog (if appears) - best-effort
    click_safe(
        driver,
        By.XPATH,
        "//button[starts-with(@id, 'cancelButton_') and .//div[text()='Discard changes']]",
        timeout=1,
        retries=2,
    )

    # clicking OK Button if avail
    click_safe(
        driver,
        By.XPATH,
        "//button[contains(@id,'okButton_')]/div",
        timeout=2,
        retries=2,
    )

    time.sleep(1)
    # NOTE: Edit button is NOT clicked here - done after solution check

