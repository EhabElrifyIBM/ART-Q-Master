from datetime import datetime
import os
import json
import logging
import pandas as pd
import itertools
import pytz
import re
import math
from collections import defaultdict

import pandas as pd
import logging
from datetime import datetime
import os
import math
import re

# Helper: safe missing-value check to avoid pandas type-checker overload issues
def _is_missing(x):
    """Return True if x should be considered missing/empty.
    Avoid direct pd.isna(x) to keep static type-checkers happy when x may be Hashable.
    """
    if x is None:
        return True
    try:
        # float NaN
        if isinstance(x, float) and math.isnan(x):
            return True
    except Exception:
        pass
    if isinstance(x, str):
        return x.strip().lower() in ('', 'nan', 'none')
    try:
        # Detect pandas NA sentinel
        if x is pd.NA:
            return True
    except Exception:
        pass
    return False

try:
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Create dummy classes for when openpyxl is not available
    class DummyDataValidation:
        def __init__(self, *args, **kwargs):
            pass
        def add(self, *args, **kwargs):
            pass
        error = ""
        errorTitle = ""
        prompt = ""
        promptTitle = ""
    
    class DummyOpenpyxl:
        def load_workbook(self, *args, **kwargs):
            raise ImportError("openpyxl not available")
    
    DataValidation = DummyDataValidation
    openpyxl = DummyOpenpyxl()
    get_column_letter = lambda x: chr(64 + x)  # Simple fallback


class FileProcessor:
    def __init__(self):
        self.setup_logging()
        self.file_path = None
        self.output_path = None
        self.settings = None
        self.excluded_answers = None
        self.include_files = None
        self.previous_assignments = {}
        
        # Initialize tracking attributes for summary
        self.bank_sutherland_updated_cases = []
        self.dnd_updated_cases = []
        self.duplicate_company_cases = None
        
        # Initialize fair share calculation info
        self.fair_share_info = None
        
        # Define filtering criteria
        self.excluded_work_order_status = [
            'Cancelled',
            'Cancelled by lenovo'
        ]
        
        self.excluded_case_reasons = [
            'Escalation/Complaint'
        ]
        
        self.excluded_closing_codes = [
            'Customer Induced Damage',
            'Cancelled by Customer',
            'Cancelled by Lenovo',
            'Customer Not available',
            'Machine Not Found',
            'Returned but Un-repaired'
        ]
        
        # Define canonical field mapping - the single source of truth for column names
        self.canonical_fields = {
            'Customer Name': [' Contact Name (Contact) (Contact)', 'Contact Name', 'Contact Name (Contact)'],
            'Company Name': ['Company Name'],
            'Country': ['Country/Region (Contact) (Contact)', 'Country'],
            'State/Province': ['State/Province (Case) (Case)', 'State/Province'],
            'Phone Number': ['Contact Mobile Phone', 'Phone Number'],
            'Email': ['Primary Email (Contact) (Contact)', 'Email'],
            'Last Status Change': ['Last Status Change (Workflow Use Only) (Case) (Case)', 'Last Status Change'],
            'Case Status': ['Case Status (Case) (Case)', 'Case Status'],  # Specific to Case
            'Status': ['Status'],  # Independent status field
            'Work Order Status': ['Work Order Status'],  # Specific to Work Order
            'DND (Do Not Disturb)': ['Do Not Disturb (Contact) (Contact)', 'DND (Do Not Disturb)'],
            'Problem Description': ['Problem Description (Case) (Case)', 'Problem Description'],
            'Incoming Channel': ['Incoming Channel (Case) (Case)', 'Incoming Channel'],
            'Case Reason': ['Case Reason (Case) (Case)', 'Case Reason'],
            'Work Order ID': ['Work Order ID'],
            'Product ID (MTM)': ['Product ID (MTM) (Case) (Case)', 'Product ID (MTM)'],
            'Machine Type': ['Machine Type'],
            'Product Description': ['Product Description'],
            'Serial Number': ['Serial Number (Case) (Case)', 'Serial Number'],
            'Survey Preference': ['Survey Preference (Case) (Case)', 'Survey Preference'],
            'Survey Fatigue': ['Survey Fatigue (Case) (Case)', 'Survey Fatigue'],
            'No survey reason': ['No survey reason (Case) (Case)', 'No survey reason'],
            'Program': ['Program (Case) (Case)', 'Program'],
            'Repeat Frequency': ['Repeat Frequency (Case) (Case)', 'Repeat Frequency'],
            'Repeat Repair': ['Repeat Repair'],
            'Closing Code': ['Closing Code']
        }
        
        # Define output columns and raw-to-output mapping
        self.output_columns = [
            'Case Number', 'Work Order Type', 'State/Province', 'Local Time', 'Action 1', 'Action 2', 'Action 3', 'Final Action', 'Assigned To', 'Status',
            'Company Name', 'DND (Do Not Disturb)', 'Email', 'Phone Number', 'Incoming Channel', 'Last Status Change', 'Country', 'Customer Name',
            'Case', 'Problem Description', 'Case Status', 'Case Status Updated', 'Case Reason', 'Work Order ID', 'Work Order Status', 'Order Type', 'Work Order Priority',
            'Product ID (MTM)', 'Machine Type', 'Product Description', 'Serial Number', 'Created On', 'Survey Preference', 'Survey Fatigue', 'No survey reason',
            'Program', 'Repeat Frequency', 'Repeat Repair', 'Closing Code', 'Reported Symptom', 'Completion Date'
        ]

        # Define output columns and raw-to-output mapping (aligned with canonical_fields)
        self.raw_to_output = {
            # Core identifiers and workflow fields
            'Case Number': ['Case Number'],
            'Work Order Type': ['Work Order Type'],
            'Local Time': ['Local Time'],
            'Action 1': ['Action 1'],
            'Action 2': ['Action 2'],
            'Action 3': ['Action 3'],
            'Final Action': ['Final Action'],
            'Assigned To': ['Assigned To'],
            
            # Status fields (kept separate)
            'Status': ['Status'],  # Independent status
            'Case Status': ['Case Status (Case) (Case)', 'Case Status'],  # Case-specific status
            'Work Order Status': ['Work Order Status'],  # Work Order-specific status
            
            # Contact and location information
            'Customer Name': [' Contact Name (Contact) (Contact)', 'Contact Name', 'Contact Name (Contact)'],
            'Company Name': ['Company Name'],
            'Email': ['Primary Email (Contact) (Contact)', 'Email'],
            'Phone Number': ['Contact Mobile Phone', 'Phone Number'],
            'Country': ['Country/Region (Contact) (Contact)', 'Country'],
            'State/Province': ['State/Province (Case) (Case)', 'State/Province'],
            
            # Case details
            'Case': ['Case'],
            'Problem Description': ['Problem Description (Case) (Case)', 'Problem Description'],
            'Case Status Updated': ['Case status Updated (Case) (Case)', 'Case Status Updated'],
            'Case Reason': ['Case Reason (Case) (Case)', 'Case Reason'],
            'Reported Symptom': ['Reported Symptom'],
            'Last Status Change': ['Last Status Change (Workflow Use Only) (Case) (Case)', 'Last Status Change'],
            
            # Work Order details
            'Work Order ID': ['Work Order ID'],
            'Order Type': ['Order Type'],
            'Work Order Priority': ['Work Order Priority'],
            
            # Product information
            'Product ID (MTM)': ['Product ID (MTM) (Case) (Case)', 'Product ID (MTM)'],
            'Machine Type': ['Machine Type'],
            'Product Description': ['Product Description'],
            'Serial Number': ['Serial Number (Case) (Case)', 'Serial Number'],
            
            # Additional metadata
            'Created On': ['Created On'],
            'Incoming Channel': ['Incoming Channel (Case) (Case)', 'Incoming Channel'],
            'DND (Do Not Disturb)': ['Do Not Disturb (Contact) (Contact)', 'DND (Do Not Disturb)'],
            'Survey Preference': ['Survey Preference (Case) (Case)', 'Survey Preference'],
            'Survey Fatigue': ['Survey Fatigue (Case) (Case)', 'Survey Fatigue'],
            'No survey reason': ['No survey reason (Case) (Case)', 'No survey reason'],
            'Program': ['Program (Case) (Case)', 'Program'],
            'Repeat Frequency': ['Repeat Frequency (Case) (Case)', 'Repeat Frequency'],
            'Repeat Repair': ['Repeat Repair'],
            'Closing Code': ['Closing Code'],
            'Completion Date': ['Completion Date']
        }
        # Default canonical column name mapping used across the processor.
        # This provides a central place for code that references self.columns[...] to
        # look up the expected column name in incoming dataframes. Values here are
        # defaults and may be overridden later if different variants are detected.
        self.columns = {
            # Core identifiers and workflow
            'case_number': 'Case Number',
            'local_time': 'Local Time',
            'action_1': 'Action 1',
            'action_2': 'Action 2',
            'action_3': 'Action 3',
            'final_action': 'Final Action',
            'assigned_to': 'Assigned To',
            
            # Status fields (kept separate)
            'status': 'Status',  # Independent status
            'case_status': 'Case Status (Case) (Case)',  # Case-specific status
            'wo_status': 'Work Order Status',  # Work Order-specific status
            
            # Contact and location information
            'customer_name': 'Contact Name (Contact) (Contact)',
            'company': 'Company Name',
            'email': 'Primary Email (Contact) (Contact)',
            'phone': 'Contact Mobile Phone',
            'country': 'Country/Region (Contact) (Contact)',
            'state': 'State/Province (Case) (Case)',
            
            # Case details
            'case': 'Case',
            'problem_description': 'Problem Description (Case) (Case)',
            'case_status_updated': 'Case status Updated (Case) (Case)',
            'case_reason': 'Case Reason (Case) (Case)',
            'last_status_change': 'Last Status Change (Workflow Use Only) (Case) (Case)',
            'incoming_channel': 'Incoming Channel (Case) (Case)',
            
            # Work Order details
            'wo_id': 'Work Order ID',
            'wo_type': 'Work Order Type',
            'order_type': 'Order Type',
            'wo_priority': 'Work Order Priority',
            
            # Product information
            'product_id': 'Product ID (MTM) (Case) (Case)',
            'machine_type': 'Machine Type',
            'product_desc': 'Product Description',
            'serial': 'Serial Number (Case) (Case)',
            
            # Additional metadata
            'created_on': 'Created On',
            'dnd': 'DND (Do Not Disturb)',
            'survey_pref': 'Survey Preference (Case) (Case)',
            'survey_fatigue': 'Survey Fatigue (Case) (Case)',
            'no_survey_reason': 'No survey reason (Case) (Case)',
            'program': 'Program (Case) (Case)',
            'repeat_freq': 'Repeat Frequency (Case) (Case)',
            'repeat_repair': 'Repeat Repair',
            'closing_code': 'Closing Code',
            'reported_symptom': 'Reported Symptom',
            'completion_date': 'Completion Date'
        }
        
    def setup_logging(self, log_callback=None):
        """Set up logging configuration with optional GUI callback"""
        class GUIHandler(logging.Handler):
            def __init__(self, callback):
                super().__init__()
                self.callback = callback

            def emit(self, record):
                if self.callback:
                    msg = self.format(record)
                    self.callback(msg)

        # Set up basic logging format
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        
        # File handler with detailed logging using UTF-8 encoding 
        file_handler = logging.FileHandler('file_processing.log', encoding='utf-8')
        file_handler.setFormatter(formatter)
        file_handler.setLevel(logging.DEBUG)  # Log everything to file
        
        # Console handler — do not reassign sys.stdout.buffer (can fail in some environments)
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        console_handler.setLevel(logging.INFO)
        
        # GUI handler (if callback provided) - only important messages
        handlers = [file_handler, console_handler]
        if log_callback is not None:
            gui_handler = GUIHandler(log_callback)
            gui_handler.setFormatter(formatter)
            gui_handler.setLevel(logging.INFO)  # Only INFO and above to GUI
            handlers.append(gui_handler)
        
        # Configure logger
        logging.basicConfig(
            level=logging.DEBUG,  # Allow all levels for file logging
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=handlers,
            force=True
        )
        
        self.logger = logging.getLogger(__name__)
        
        # Create a separate logger for dropped cases tracking
        self.dropped_cases_logger = logging.getLogger('dropped_cases')
        self.dropped_cases_logger.setLevel(logging.INFO)
        
        # Create dropped cases log file with UTF-8 encoding
        dropped_handler = logging.FileHandler('dropped_cases.log', encoding='utf-8')
        dropped_handler.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
        self.dropped_cases_logger.addHandler(dropped_handler)
        self.dropped_cases_logger.propagate = False  # Don't propagate to root logger
        
    def get_raw_col(self, df, output_col):
        """Return the best matching raw column name for a given output column.
        Uses the mapping stored in self.raw_to_output. Performs matches in order:
        1. Exact case-insensitive match
        2. Normalized string match (removes spaces, special chars)
        3. Partial substring match
        4. Fuzzy match on column name components
        """
        try:
            if output_col not in getattr(self, 'raw_to_output', {}):
                self.logger.warning(f"No mapping defined for output column: {output_col}")
                return None
                
            def normalize_colname(col):
                """Normalize column name for comparison"""
                return re.sub(r'[^a-z0-9]', '', str(col).lower())
            
            candidates = self.raw_to_output[output_col]
            self.logger.debug(f"\nLooking for raw column matching {output_col}")
            self.logger.debug(f"Candidates: {candidates}")
            
            # 1. Try exact case-insensitive match
            for raw_candidate in candidates:
                for col in df.columns:
                    try:
                        if col.strip().lower() == raw_candidate.strip().lower():
                            self.logger.debug(f"Found exact match: {col}")
                            return col
                    except Exception:
                        continue
            
            # 2. Try normalized string match
            for raw_candidate in candidates:
                norm_candidate = normalize_colname(raw_candidate)
                for col in df.columns:
                    try:
                        if normalize_colname(col) == norm_candidate:
                            self.logger.debug(f"Found normalized match: {col}")
                            return col
                    except Exception:
                        continue
            
            # 3. Try partial match
            for raw_candidate in candidates:
                norm_candidate = normalize_colname(raw_candidate)
                for col in df.columns:
                    try:
                        if norm_candidate in normalize_colname(col):
                            self.logger.debug(f"Found partial match: {col}")
                            return col
                    except Exception:
                        continue
            
            # No match found
            self.logger.warning(f"No matching raw column found for {output_col}")
            self.logger.debug("Available columns:")
            for col in df.columns:
                self.logger.debug(f"  {col}")
                
            return None
            
        except Exception as e:
            self.logger.error(f"Error in get_raw_col for {output_col}: {str(e)}")
            return None
        
    def log_dropped_case(self, case_number, reason, details=None, status=None):
        """Log a dropped case with its reason and details"""
        try:
            log_message = f"DROPPED CASE: {case_number} - Reason: {reason}"
            if status:
                log_message += f" - Status: {status}"
            if details:
                log_message += f" - Details: {details}"
            
            self.dropped_cases_logger.info(log_message)
            self.logger.debug(log_message)  # Also log to main log file for debugging
        except Exception as e:
            self.logger.error(f"Error logging dropped case: {str(e)}")
    
    def log_case_status_change(self, case_number, old_status, new_status, reason=None):
        """Log when a case status changes (e.g., to fixed, not fixed, DND)"""
        try:
            log_message = f"STATUS CHANGE: {case_number} - {old_status} -> {new_status}"
            if reason:
                log_message += f" - Reason: {reason}"
            
            self.dropped_cases_logger.info(log_message)
            self.logger.info(log_message)  # Show status changes in main log
        except Exception as e:
            self.logger.error(f"Error logging status change: {str(e)}")
    
    def log_filtering_summary(self, filter_name, initial_count, final_count, dropped_reasons):
        """Log a summary of filtering results with dropped case details"""
        try:
            dropped_count = initial_count - final_count
            self.logger.info(f"\n{filter_name} Filter Summary:")
            self.logger.info(f"  Initial count: {initial_count}")
            self.logger.info(f"  Final count: {final_count}")
            self.logger.info(f"  Dropped count: {dropped_count}")
            
            if dropped_reasons:
                self.logger.info(f"  Dropped reasons:")
                for reason, cases in dropped_reasons.items():
                    self.logger.info(f"    {reason}: {len(cases)} cases")
                    # Log individual dropped cases to dropped cases log
                    for case in cases:
                        self.log_dropped_case(case, reason, filter_name)
            
            # Also log to dropped cases log
            self.dropped_cases_logger.info(f"FILTER SUMMARY: {filter_name} - Dropped {dropped_count} out of {initial_count} cases")
            
        except Exception as e:
            self.logger.error(f"Error logging filtering summary: {str(e)}")
    
    def log_issue_not_fixed_case(self, case_number, source, details=None):
        """Log when a case is added to the Issue Not Fixed sheet"""
        try:
            log_message = f"ISSUE NOT FIXED CASE: {case_number} - Source: {source}"
            if details:
                log_message += f" - Details: {details}"
            
            self.dropped_cases_logger.info(log_message)
            self.logger.info(log_message)  # Show in main log as well
        except Exception as e:
            self.logger.error(f"Error logging issue not fixed case: {str(e)}")
    
    def log_issue_not_fixed_summary(self, total_cases, sms_cases, email_cases, new_cases):
        """Log a summary of Issue Not Fixed cases processing"""
        try:
            self.logger.info(f"\n=== Issue Not Fixed Processing Summary ===")
            self.logger.info(f"Total cases in Issue Not Fixed sheet: {total_cases}")
            self.logger.info(f"Cases from SMS processing: {sms_cases}")
            self.logger.info(f"Cases from Email processing: {email_cases}")
            self.logger.info(f"New cases added: {new_cases}")
            
            # Also log to dropped cases log
            self.dropped_cases_logger.info(f"ISSUE NOT FIXED SUMMARY: Total {total_cases}, SMS {sms_cases}, Email {email_cases}, New {new_cases}")
            
        except Exception as e:
            self.logger.error(f"Error logging issue not fixed summary: {str(e)}")

    def process_raw_file(self, file_path):
        """Process the raw input file"""
        self.logger.info(f"Processing file: {file_path}")
        # Implementation will be added based on specific requirements
        
    def update_case_status(self, case_id, status):
        """Update status for a specific case"""
        self.logger.info(f"Updating status for case {case_id} to {status}")
        # Implementation will be added based on specific requirements
        
    def assign_cases_to_handlers(self, cases, handlers):
        """Assign cases to handlers based on rules"""
        self.logger.info(f"Assigning {len(cases)} cases to {len(handlers)} handlers")
        # Implementation will be added based on specific requirements

    def validate_files(self, raw_file, prev_file=None, sms_file=None):
        """Validate input files"""
        try:
            # Check if raw file exists
            if not os.path.exists(raw_file):
                return False, "Raw file does not exist"
            
            # Check previous file if provided
            if prev_file and not os.path.exists(prev_file):
                return False, "Previous file does not exist"
            
            # Check SMS file if provided
            if sms_file and not os.path.exists(sms_file):
                return False, "SMS file does not exist"
            
            return True, "All files validated successfully"
        except Exception as e:
            self.logger.error(f"File validation error: {str(e)}")
            return False, f"Error during file validation: {str(e)}"

    def clean_string_series(self, series):
        """Clean string series by removing extra whitespace and converting to lowercase"""
        # First convert to string, handling NaN values
        series = series.fillna('').astype(str)
        return series.str.strip().str.lower()

    def filter_case(self, df):
        """Filter cases based on CID and DMR rules"""
        try:
            original_len = len(df)
            
            self.logger.info("\nStarting CID/DMR filtering:")
            self.logger.info(f"Initial count: {original_len}")
            
            # Debug: Print sample of cases before filtering
            self.logger.debug("\nSample of cases before filtering:")
            case_col_name = self.get_raw_col(df, 'Case')
            sample_cases = df[case_col_name].head() if case_col_name else []
            for case in sample_cases:
                self.logger.debug(f"Case value: {case}")
            
            # Create a mask for cases to keep
            case_col = df[case_col_name] if case_col_name else pd.Series(dtype=str)
            
            # 1. Find "not cid" cases (these should be kept)
            not_cid_mask = case_col.str.contains(r'not\s*[\[\(]?cid[\]\)]?', case=False, regex=True, na=False)
            not_cid_count = not_cid_mask.sum()
            self.logger.info(f"\nFound {not_cid_count} cases with 'not cid'")
            
            # 2. Find CID cases
            cid_mask = case_col.str.contains(r'[\[\(]?cid[\]\)]?', case=False, regex=True, na=False)
            cid_count = cid_mask.sum()
            self.logger.info(f"Found {cid_count} cases with 'cid'")
            
            # 3. Find DMR cases
            dmr_mask = case_col.str.contains(r'[\[\(]?dmr[\]\)]?', case=False, regex=True, na=False)
            dmr_count = dmr_mask.sum()
            self.logger.info(f"Found {dmr_count} cases with 'dmr'")
            
            # Create the final mask
            # Keep rows that either:
            # 1. Contain "not cid", OR
            # 2. Don't contain either "cid" or "dmr"
            mask = not_cid_mask | (~cid_mask & ~dmr_mask)
            
            # Store removed cases for analysis
            removed_cases = df[~mask].copy()
            
            # Apply the filter
            df_filtered = df[mask].copy()
            
            # Log detailed results
            self.logger.info("\nCID/DMR filtering results:")
            self.logger.info(f"Original count: {original_len}")
            self.logger.info(f"After filtering: {len(df_filtered)}")
            self.logger.info(f"Removed {len(removed_cases)} cases")
            
            # Log sample of removed cases with their full case text
            if not removed_cases.empty:
                self.logger.debug("\nSample of removed cases:")
                sample_size = min(10, len(removed_cases))
                for idx, row in removed_cases.head(sample_size).iterrows():
                    case_text = row[case_col_name] if case_col_name in row else ''
                    case_num_col = self.get_raw_col(df, 'Case Number')
                    case_num = row[case_num_col] if case_num_col in row else ''
                    self.logger.debug(f"Case Number: {case_num}")
                    self.logger.debug(f"Case Text: {case_text}")
                    self.logger.debug("---")
            
            # Use new logging system for dropped cases
            case_num_col = self.get_raw_col(df, 'Case Number')
            dropped_reasons = {
                'CID Case': removed_cases[cid_mask & ~not_cid_mask][case_num_col].tolist() if case_num_col else [],
                'DMR Case': removed_cases[dmr_mask][case_num_col].tolist() if case_num_col else []
            }
            
            self.log_filtering_summary("CID/DMR", original_len, len(df_filtered), dropped_reasons)
            
            return df_filtered, removed_cases
            
        except Exception as e:
            self.logger.error(f"Error in filter_case: {str(e)}")
            self.logger.error("Stack trace:", exc_info=True)
            raise

    def validate_columns(self, df):
        """Validate that all required columns are present in the dataframe"""
        missing_columns = []
        found_columns = []
        
        self.logger.info("\n=== Column Validation Analysis ===")
        self.logger.info("\nAvailable columns in file:")
        for col in df.columns:
            # Check column contents
            null_count = df[col].isnull().sum()
            empty_count = (df[col].fillna('').astype(str).str.strip() == '').sum()
            total_rows = len(df)
            self.logger.info(f"'{col}': {null_count} null, {empty_count} empty out of {total_rows} total rows")
            
            # Sample non-empty values if they exist
            non_empty_values = df[df[col].notna() & (df[col].astype(str).str.strip() != '')][col].head(3).tolist()
            if non_empty_values:
                self.logger.info(f"  Sample values: {non_empty_values}")
        
        self.logger.info("\nExpected output columns:")
        for out_col in self.output_columns:
            self.logger.info(f"'{out_col}'")
            # Track potential sources for this column
            potential_sources = self.raw_to_output.get(out_col, [])
            self.logger.info(f"  Potential source columns: {potential_sources}")
            
            # Check if any potential sources exist in df
            found_sources = [src for src in potential_sources if src in df.columns]
            if found_sources:
                self.logger.info(f"  Found these source columns: {found_sources}")
                for src in found_sources:
                    null_count = df[src].isnull().sum()
                    empty_count = (df[src].fillna('').astype(str).str.strip() == '').sum()
                    self.logger.info(f"    '{src}': {null_count} null, {empty_count} empty values")
            else:
                self.logger.warning(f"  No source columns found for {out_col}")
        
        # Column mapping analysis
        self.logger.info("\n=== Column Mapping Analysis ===")
        for out_col in self.output_columns:
            raw_col = self.get_raw_col(df, out_col)
            if raw_col:
                found_columns.append(f"{out_col}: Found as '{raw_col}'")
                # Analyze content of found column
                null_count = df[raw_col].isnull().sum()
                empty_count = (df[raw_col].fillna('').astype(str).str.strip() == '').sum()
                total_rows = len(df)
                self.logger.info(f"\nAnalysis for {out_col} -> {raw_col}:")
                self.logger.info(f"  Total rows: {total_rows}")
                self.logger.info(f"  Null values: {null_count} ({(null_count/total_rows)*100:.1f}%)")
                self.logger.info(f"  Empty strings: {empty_count} ({(empty_count/total_rows)*100:.1f}%)")
                
                # Sample values
                sample = df[df[raw_col].notna() & (df[raw_col].astype(str).str.strip() != '')][raw_col].head()
                if not sample.empty:
                    self.logger.info(f"  Sample values: {sample.tolist()}")
            else:
                missing_columns.append(f"{out_col}")
                self.logger.error(f"\nNo mapping found for {out_col}")
                self.logger.error(f"Expected sources: {self.raw_to_output.get(out_col, [])}")
        
        if missing_columns:
            self.logger.error("\n=== Missing Columns Analysis ===")
            for col in missing_columns:
                self.logger.error(f"Missing: {col}")
                expected_sources = self.raw_to_output.get(col, [])
                self.logger.error(f"  Expected source columns: {expected_sources}")
                # Check if any of these sources exist with different casing/spacing
                for src in expected_sources:
                    close_matches = [c for c in df.columns if src.lower() in c.lower()]
                    if close_matches:
                        self.logger.error(f"  Similar columns found: {close_matches}")
            
            self.logger.info("\n=== Found Columns ===")
            for col in found_columns:
                self.logger.info(f"Found: {col}")
        
        return len(missing_columns) == 0

    def ensure_output_columns(self, df, prev_df=None):
        """Ensure that every expected output column exists in `df`.
        Maps fields correctly from raw file to standardized output columns,
        particularly for Email, Phone Number, Last Status Change, Country, State/Province, Case Status.
        Uses canonical_fields mapping to find the correct source columns.
        """
        created = []
        try:
            self.logger.info("Ensuring output columns with proper mapping...")

            # IMPORTANT: Convert Completion Date Timestamps to strings early
            if 'Completion Date' in df.columns:
                self.logger.info("Converting Completion Date Timestamps to strings...")
                df['Completion Date'] = df['Completion Date'].astype(str)
                self.logger.info("Completion Date converted to string format")
                
                # DIAGNOSTIC: Check if conversion actually worked
                non_empty_after_early_conversion = (df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                self.logger.info(f"DIAGNOSTIC ensure_output_columns: After early conversion - {non_empty_after_early_conversion}/{len(df)} non-empty")
                if non_empty_after_early_conversion > 0:
                    samples_early = df[df['Completion Date'].fillna('').astype(str).str.strip() != '']['Completion Date'].head(3).tolist()
                    self.logger.info(f"DIAGNOSTIC ensure_output_columns: Sample values after early conversion: {samples_early}")
                else:
                    all_samples_early = df['Completion Date'].head(3).tolist()
                    self.logger.info(f"DIAGNOSTIC ensure_output_columns: ALL values after early conversion: {all_samples_early}")

            # Ensure df has a normalized Case Number column for lookups
            if 'Case Number' not in df.columns:
                for col in df.columns:
                    if 'case' in str(col).lower() and 'number' in str(col).lower():
                        df['Case Number'] = df[col]
                        self.logger.info(f"Found Case Number column: {col}")
                        break

            # Define critical field mappings - strict priority order
            critical_fields = {
                'Email': [
                    'Primary Email (Contact) (Contact)',  # First priority
                    'Email Address',  # Second priority
                    'Email',  # Third priority
                    'Contact Email'  # Fourth priority
                ],
                'Phone Number': [
                    'Contact Mobile Phone',  # First priority
                    'Phone Number',  # Second priority
                    'Mobile Phone',  # Third priority
                    'Contact Phone'  # Fourth priority
                ],
                'Last Status Change': [
                    'Last Status Change (Workflow Use Only) (Case) (Case)',  # First priority
                    'Last Status Change (Case) (Case)',  # Second priority
                    'Last Status Change'  # Third priority
                ],
                'Country': [
                    'Country/Region (Contact) (Contact)',  # First priority
                    'Country/Region',  # Second priority
                    'Country (Contact)',  # Third priority
                    'Country'  # Fourth priority
                ],
                'State/Province': [
                    'State/Province (Case) (Case)',  # First priority
                    'State/Province (Contact) (Contact)',  # Second priority
                    'State/Province (Contact)',  # Third priority
                    'State/Province'  # Fourth priority
                ],
                'Case Status': [
                    'Case Status (Case) (Case)',  # First priority
                    'Case Status (Case)',  # Second priority
                    'Case Status'  # Third priority
                ],
                'Customer Name': [
                    ' Contact Name (Contact) (Contact)',  # First priority
                    'Contact Name',  # Second priority
                    'Contact Name (Contact) (Contact)',  # Third priority
                    'Customer Name'  # Fourth priority
                ],
                'Completion Date': [
                    'Completion Date'  # Direct mapping from raw file
                ]
            }

                # Process each output column
            for out_col in self.output_columns:
                # Log the current column being processed
                self.logger.info(f"\nProcessing output column: {out_col}")
                
                # Check current state of column
                if out_col in df.columns:
                    null_count = df[out_col].isnull().sum()
                    empty_count = (df[out_col].fillna('').astype(str).str.strip() == '').sum()
                    self.logger.info(f"Column '{out_col}' exists: {null_count} null values, {empty_count} empty strings")
                else:
                    self.logger.info(f"Column '{out_col}' does not exist yet")
                
                # Check if this is a critical field needing special handling
                # Critical fields should ALWAYS be processed to fill in missing values
                if out_col in critical_fields:
                    source_cols = critical_fields[out_col]
                    found_data = False
                    
                    # Initialize the output column if it doesn't exist
                    if out_col not in df.columns:
                        df[out_col] = ''
                        created.append(out_col)
                        self.logger.info(f"Created column {out_col}")
                    
                    # Try each source column in strict priority order
                    for source_col in source_cols:
                        self.logger.info(f"\nTrying source column '{source_col}' for {out_col}")
                        
                        # Skip if source and target are the same column - column already has its data
                        if source_col == out_col:
                            self.logger.info(f"Source and target are the same column - skipping update (data already present)")
                            found_data = True
                            continue
                        
                        if source_col in df.columns:
                            # Log the source column's current state
                            source_null = df[source_col].isnull().sum()
                            source_empty = (df[source_col].fillna('').astype(str).str.strip() == '').sum()
                            self.logger.info(f"Source column exists with {source_null} null values, {source_empty} empty strings")
                            
                            # Only copy to empty cells or cells that need updating
                            empty_mask = (df[out_col].fillna('').astype(str).str.strip() == '')
                            empty_count = empty_mask.sum()
                            self.logger.info(f"Target column has {empty_count} empty cells to potentially fill")
                            
                            source_values = df[source_col].fillna('').astype(str).str.strip()
                            valid_source_mask = (source_values != '')
                            valid_count = valid_source_mask.sum()
                            self.logger.info(f"Source column has {valid_count} non-empty values available")
                            
                            # Update only where we have valid source values and target is empty
                            update_mask = empty_mask & valid_source_mask
                            updates_possible = update_mask.sum()
                            self.logger.info(f"Found {updates_possible} cells that can be updated")
                            
                            if update_mask.any():
                                # For Completion Date, convert Timestamp to string if needed
                                if out_col == 'Completion Date':
                                    # Convert Timestamp objects to strings in the proper format
                                    df.loc[update_mask, out_col] = df.loc[update_mask, source_col].astype(str)
                                    self.logger.info(f"Special handling: Converted Timestamps to strings for Completion Date")
                                else:
                                    df.loc[update_mask, out_col] = df.loc[update_mask, source_col]
                                found_data = True
                                self.logger.info(f"Successfully mapped {source_col} -> {out_col} ({update_mask.sum()} rows)")
                                
                                # Debug logging for ALL updated values
                                updates = df[update_mask]
                                self.logger.info("\nDetailed update log:")
                                for idx, row in updates.iterrows():
                                    self.logger.info(f"Updated {out_col} for case {row['Case Number']}: '{row[source_col]}' -> '{df.loc[idx, out_col]}'")
                                
                                # Log sample of cells that couldn't be updated
                                still_empty = df[empty_mask & ~update_mask]
                                if not still_empty.empty:
                                    self.logger.warning(f"\nSample of cases where update failed from this source:")
                                    for _, row in still_empty.head().iterrows():
                                        self.logger.warning(f"Case {row['Case Number']}: Source value: '{row.get(source_col, 'N/A')}', Current target value: '{row.get(out_col, 'N/A')}'")
                            else:
                                self.logger.warning(f"No updates possible from {source_col} to {out_col}")
                    
                    # Double-check if we still have empty values after trying all sources
                    still_empty = (df[out_col].fillna('').astype(str).str.strip() == '').sum()
                    if still_empty > 0:
                        self.logger.warning(f"{still_empty} rows still missing {out_col} after checking all source columns")
                        # Log some examples of cases with missing data
                        empty_examples = df[df[out_col].fillna('').astype(str).str.strip() == '']['Case Number'].head(5).tolist()
                        self.logger.warning(f"Example cases missing {out_col}: {empty_examples}")
                    else:
                        self.logger.info(f"Successfully populated all rows for {out_col}")
                else:
                    # Standard column handling - skip if already exists and has data
                    if out_col in df.columns and not df[out_col].isnull().all():
                        # Check if column has any meaningful data
                        has_data = (df[out_col].fillna('').astype(str).str.strip() != '').any()
                        if has_data:
                            self.logger.info(f"Column '{out_col}' already has data - skipping")
                            continue
                    
                    raw_col = self.get_raw_col(df, out_col)
                    if raw_col and raw_col in df.columns:
                        # Check data in raw column before mapping
                        non_empty = df[raw_col].fillna('').astype(str).str.strip() != ''
                        if non_empty.any():
                            df[out_col] = df[raw_col]
                            created.append(out_col)
                            self.logger.info(f"Standard mapping: {raw_col} -> {out_col} ({non_empty.sum()} non-empty values)")
                            # Log sample of mapped values
                            sample = df[non_empty][raw_col].head(3)
                            self.logger.info(f"Sample values mapped: {sample.tolist()}")
                        else:
                            self.logger.warning(f"Raw column {raw_col} found but contains no non-empty values")
                    elif out_col not in df.columns:
                        df[out_col] = ''
                        created.append(out_col)
                        self.logger.info(f"Created empty column: {out_col}")

            # Fill in from previous file for existing cases
            if prev_df is not None and not prev_df.empty and 'Case Number' in prev_df.columns:
                self.logger.info("Checking previous file for existing cases...")
                
                # First normalize case numbers in previous file
                prev_df['Case Number'] = prev_df['Case Number'].apply(self.normalize_case_number)
                prev_case_nums = set(prev_df['Case Number'].dropna())
                self.logger.info(f"Found {len(prev_case_nums)} unique case numbers in previous file")
                
                # Track which cases and columns are populated from previous file
                prev_file_updates = defaultdict(list)
                raw_file_values = defaultdict(list)
                
                for out_col in self.output_columns:
                    # SKIP Completion Date - it should NOT be overwritten by previous file data
                    # It was just properly populated from the raw file
                    if out_col == 'Completion Date':
                        self.logger.info(f"Skipping {out_col} - preserving values from raw file")
                        continue
                    
                    if out_col in prev_df.columns:
                        try:
                            # Log column state before population
                            self.logger.info(f"\nPopulating {out_col} from previous file:")
                            self.logger.info("Current state:")
                            non_empty_current = (df[out_col].fillna('').astype(str).str.strip() != '').sum() if out_col in df.columns else 0
                            self.logger.info(f"  Non-empty values in current data: {non_empty_current}")
                            
                            updates = 0
                            for idx, row in df.iterrows():
                                case_num = self.normalize_case_number(row['Case Number'])
                                if case_num in prev_case_nums:
                                    # Store original value before update
                                    orig_val = str(row.get(out_col, '')).strip()
                                    if orig_val:
                                        raw_file_values[case_num].append(f"{out_col}={orig_val}")
                                        
                                    # Get value from previous file
                                    prev_val = prev_df[prev_df['Case Number'] == case_num][out_col].iloc[0]
                                    if pd.notna(prev_val) and str(prev_val).strip():
                                        df.at[idx, out_col] = prev_val
                                        prev_file_updates[case_num].append(f"{out_col}={prev_val}")
                                        updates += 1
                            
                            self.logger.info(f"  Updated {updates} cases from previous file")
                        except Exception as e:
                            self.logger.warning(f"Error copying {out_col} from previous data: {str(e)}")

                # Verify and fix critical fields for new cases
            if prev_df is not None:
                self.logger.info("\n=== Beginning New Cases Verification ===")
                
                # Normalize case numbers for comparison
                df['_normalized_case'] = df['Case Number'].apply(self.normalize_case_number)
                prev_df['_normalized_case'] = prev_df['Case Number'].apply(self.normalize_case_number)
                
                # Identify new cases
                current_cases = set(df['_normalized_case'])
                previous_cases = set(prev_df['_normalized_case'])
                new_case_nums = current_cases - previous_cases
                
                self.logger.info(f"Total cases in current file: {len(current_cases)}")
                self.logger.info(f"Total cases in previous file: {len(previous_cases)}")
                self.logger.info(f"Number of new cases identified: {len(new_case_nums)}")
                
                # Log the first few new cases for verification
                sample_new_cases = list(new_case_nums)[:5]
                self.logger.info(f"Sample of new cases: {sample_new_cases}")
                
                if new_case_nums:
                    self.logger.info(f"\n=== Verifying Critical Fields for {len(new_case_nums)} New Cases ===")
                    
                    for field, source_cols in critical_fields.items():
                        if field not in df.columns:
                            df[field] = ''  # Ensure the column exists
                            
                        # Check which new cases are missing this field
                        new_cases_mask = df['_normalized_case'].isin(new_case_nums)
                        empty_mask = (df[field].fillna('').astype(str).str.strip() == '')
                        missing_mask = new_cases_mask & empty_mask
                        
                        if missing_mask.any():
                            missing_count = missing_mask.sum()
                            self.logger.warning(f"\nFound {missing_count} new cases missing {field}")
                            
                            # Try to fill missing values from each source column
                            for source_col in source_cols:
                                if source_col in df.columns:
                                    source_values = df[source_col].fillna('').astype(str).str.strip()
                                    valid_source = (source_values != '')
                                    update_mask = missing_mask & valid_source
                                    
                                    if update_mask.any():
                                        df.loc[update_mask, field] = df.loc[update_mask, source_col]
                                        self.logger.info(f"Filled {update_mask.sum()} missing {field} values from {source_col}")
                                        
                                        # Update the missing mask
                                        missing_mask = new_cases_mask & (df[field].fillna('').astype(str).str.strip() == '')
                                        
                                        if not missing_mask.any():
                                            self.logger.info(f"Successfully filled all missing {field} values")
                                            break
                            
                            # Final check for any remaining missing values
                            still_missing = missing_mask.sum()
                            if still_missing > 0:
                                missing_cases = df[missing_mask]['Case Number'].tolist()[:5]
                                self.logger.error(f"CRITICAL: Still missing {field} for {still_missing} new cases")
                                self.logger.error(f"Example cases missing {field}: {missing_cases}")
                        else:
                            self.logger.info(f"All new cases have {field} populated")
                    
                    # Clean up temporary column
                    df = df.drop(columns=['_normalized_case'])
                    if '_normalized_case' in prev_df.columns:
                        prev_df = prev_df.drop(columns=['_normalized_case'])

            return created

        except Exception as e:
            self.logger.error(f"Error ensuring output columns: {str(e)}")
            return created
        

    def _normalize_id_columns(self, df):
        for col in df.columns:
            if 'case number' in col.lower():
                df[col] = df[col].apply(self.normalize_case_number)
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')
            elif 'phone' in col.lower():
                def norm_phone(val):
                    if pd.isna(val):
                        return ''
                    if isinstance(val, float) and val.is_integer():
                        return str(int(val))
                    return str(val).strip()
                df[col] = df[col].apply(norm_phone)
        return df

    def process_files(self, raw_file, prev_file, sms_file, email_file, output_file, selected_handlers, chat_agent_info=None):
        """Process the input files"""
        try:
            self.logger.info("\n=== Starting file processing ===")
            self.logger.info(f"Selected handlers: {selected_handlers}")
            if chat_agent_info:
                self.logger.info(f"Chat Agent enabled: {chat_agent_info['supporter_name']} (15% capacity)")
            
            # Read the raw file
            self.logger.info(f"\nReading raw file: {raw_file}")
            df = pd.read_excel(raw_file)
            
            # Log file loaded (removed verbose per-column sample logging)
            self.logger.info(f"Loaded {len(df)} records with {len(df.columns)} columns")
            
            # Clean and normalize all string columns
            string_cols = df.select_dtypes(include=['object']).columns
            for col in string_cols:
                df[col] = df[col].fillna('').astype(str).str.strip()
                
            # Normalize ID columns
            df = self._normalize_id_columns(df)
            initial_count = len(df)
            
            self.logger.info(f"Initial record count after normalization: {initial_count}")
            
            # Load previous file if it exists
            prev_df = None
            if prev_file and os.path.exists(prev_file):
                self.logger.info(f"\nLoading previous file: {prev_file}")
                try:
                    prev_df = self.load_previous_file(prev_file)
                    self.logger.info(f"Previous file loaded successfully with {len(prev_df)} records")
                except Exception as e:
                    self.logger.error(f"Error loading previous file: {str(e)}")
                    self.logger.warning("Continuing without previous file data")
                    prev_df = None
            
            # Validate columns before processing
            valid = self.validate_columns(df)
            if not valid:
                # Instead of failing immediately, attempt to standardize column names
                self.logger.info("Standardizing column names for consistency...")
                created_cols = self.ensure_output_columns(df, prev_df)
                self.logger.info(f"Standardized column names to: {created_cols}")
                # Re-run validation
                if not self.validate_columns(df):
                    raise ValueError("Missing required columns in input file after attempted standardization")
            
            # Clean and normalize the relevant columns
            self.logger.info("\nCleaning and normalizing columns...")
            
            # CRITICAL: Save Completion Date before filtering, as it might be lost during operations
            completion_date_backup = None
            if 'Completion Date' in df.columns:
                completion_date_backup = df['Completion Date'].copy()
                self.logger.info(f"DIAGNOSTIC: Saved Completion Date backup with {(completion_date_backup.fillna('').astype(str).str.strip() != '').sum()} non-empty values")
            
            for out_col in ['Work Order Status', 'Case Reason', 'Closing Code', 'Case', 'Case Number']:
                raw_col = self.get_raw_col(df, out_col)
                if raw_col:
                    df[raw_col] = self.clean_string_series(df[raw_col])
            
            # Apply filters with detailed tracking
            self.logger.info("\n=== Applying Filters ===")
            
            # Track all dropped cases for comprehensive logging
            all_dropped_cases = []
            dropped_cases_details = {}
            
            # 1. Work Order Status filter
            if not isinstance(df, pd.DataFrame):
                df = pd.DataFrame(df)
            excluded_wo_status = [x.strip().lower() for x in self.excluded_work_order_status]
            wo_status_col = self.get_raw_col(df, 'Work Order Status')
            mask_wo = ~df[wo_status_col].isin(excluded_wo_status) if wo_status_col else pd.Series([True]*len(df))
            removed_wo = df[~mask_wo]
            if not isinstance(removed_wo, pd.DataFrame):
                removed_wo = pd.DataFrame(removed_wo)
            
            # DIAGNOSTIC: Check Completion Date before Work Order Status filter
            if 'Completion Date' in df.columns:
                completion_date_non_empty = (df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                self.logger.info(f"DIAGNOSTIC: Before Work Order Status filter - Completion Date has {completion_date_non_empty}/{len(df)} non-empty")
            
            df = df[mask_wo]
            
            # DIAGNOSTIC: Check Completion Date after Work Order Status filter
            if 'Completion Date' in df.columns:
                completion_date_non_empty = (df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                self.logger.info(f"DIAGNOSTIC: After Work Order Status filter - Completion Date has {completion_date_non_empty}/{len(df)} non-empty")
            wo_status_count = len(df)
            self.logger.info(f"\n1. After Work Order Status filter: {wo_status_count} records remaining")
            # Log dropped cases from Work Order Status filter
            if not removed_wo.empty and wo_status_col:
                dropped_reasons = {}
                case_num_col = self.get_raw_col(df, 'Case Number')
                for status in excluded_wo_status:
                    status_cases = removed_wo[removed_wo[wo_status_col] == status]
                    if not status_cases.empty and case_num_col:
                        dropped_reasons[f"Status: {status}"] = status_cases[case_num_col].tolist()
                self.log_filtering_summary("Work Order Status", initial_count, wo_status_count, dropped_reasons)
                # Add to dropped cases details for summary table
                for status in excluded_wo_status:
                    status_cases = removed_wo[removed_wo[wo_status_col] == status]
                    if not status_cases.empty and case_num_col:
                        dropped_cases_details[f"Work Order Status: {status}"] = status_cases[case_num_col].tolist()
            
            # 2. Case Reason filter
            if not isinstance(df, pd.DataFrame):
                df = pd.DataFrame(df)
            excluded_case_reasons = [x.strip().lower() for x in self.excluded_case_reasons]
            case_reason_col = self.get_raw_col(df, 'Case Reason')
            mask_case = ~df[case_reason_col].isin(excluded_case_reasons) if case_reason_col else pd.Series([True]*len(df))
            removed_case = df[~mask_case]
            if not isinstance(removed_case, pd.DataFrame):
                removed_case = pd.DataFrame(removed_case)
            df = df[mask_case]
            case_reason_count = len(df)
            self.logger.info(f"\n2. After Case Reason filter: {case_reason_count} records remaining")
            # Log dropped cases from Case Reason filter
            if not removed_case.empty and case_reason_col:
                dropped_reasons = {}
                case_num_col = self.get_raw_col(df, 'Case Number')
                for reason in excluded_case_reasons:
                    reason_cases = removed_case[removed_case[case_reason_col] == reason]
                    if not reason_cases.empty and case_num_col:
                        dropped_reasons[f"Reason: {reason}"] = reason_cases[case_num_col].tolist()
                self.log_filtering_summary("Case Reason", wo_status_count, case_reason_count, dropped_reasons)
                # Add to dropped cases details for summary table
                for reason in excluded_case_reasons:
                    reason_cases = removed_case[removed_case[case_reason_col] == reason]
                    if not reason_cases.empty and case_num_col:
                        dropped_cases_details[f"Case Reason: {reason}"] = reason_cases[case_num_col].tolist()
            
            # 3. Closing Code filter
            if not isinstance(df, pd.DataFrame):
                df = pd.DataFrame(df)
            excluded_closing_codes = [x.strip().lower() for x in self.excluded_closing_codes]
            closing_code_col = self.get_raw_col(df, 'Closing Code')
            mask_closing = ~df[closing_code_col].isin(excluded_closing_codes) if closing_code_col else pd.Series([True]*len(df))
            removed_closing = df[~mask_closing]
            if not isinstance(removed_closing, pd.DataFrame):
                removed_closing = pd.DataFrame(removed_closing)
            df = df[mask_closing]
            closing_code_count = len(df)
            self.logger.info(f"\n3. After Closing Code filter: {closing_code_count} records remaining")
            
            # CRITICAL: Restore Completion Date after filtering
            if completion_date_backup is not None and 'Completion Date' in df.columns:
                # Only restore for rows that are still in df
                df.loc[df.index, 'Completion Date'] = completion_date_backup.loc[df.index]
                restored_non_empty = (df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                self.logger.info(f"DIAGNOSTIC: Restored Completion Date after filters - {restored_non_empty}/{len(df)} non-empty")
            # Log dropped cases from Closing Code filter
            if not removed_closing.empty and closing_code_col:
                dropped_reasons = {}
                case_num_col = self.get_raw_col(df, 'Case Number')
                for code in excluded_closing_codes:
                    code_cases = removed_closing[removed_closing[closing_code_col] == code]
                    if not code_cases.empty and case_num_col:
                        dropped_reasons[f"Closing Code: {code}"] = code_cases[case_num_col].tolist()
                self.log_filtering_summary("Closing Code", case_reason_count, closing_code_count, dropped_reasons)
                # Add to dropped cases details for summary table
                for code in excluded_closing_codes:
                    code_cases = removed_closing[removed_closing[closing_code_col] == code]
                    if not code_cases.empty and case_num_col:
                        dropped_cases_details[f"Closing Code: {code}"] = code_cases[case_num_col].tolist()
            
            # 4. Apply CID/DMR filter with detailed logging
            self.logger.info(f"\n4. Applying CID/DMR filter:")
            pre_cid_count = len(df)
            df, removed_cases = self.filter_case(df)
            if not isinstance(removed_cases, pd.DataFrame):
                removed_cases = pd.DataFrame(removed_cases)
            post_cid_count = len(df)
            self.logger.info(f"   Records remaining: {post_cid_count}")
            
            # Add CID/DMR dropped cases to details
            if not removed_cases.empty:
                # Check if there are CID and DMR cases
                case_col = self.columns['case']
                cid_mask = removed_cases[case_col].astype(str).str.contains('cid', case=False, na=False)
                dmr_mask = removed_cases[case_col].astype(str).str.contains('dmr', case=False, na=False)
                
                if not cid_mask.empty:
                    cid_cases = removed_cases[cid_mask][self.columns['case_number']].tolist()
                    if cid_cases:
                        dropped_cases_details['CID Case'] = cid_cases
                
                if not dmr_mask.empty:
                    dmr_cases = removed_cases[dmr_mask][self.columns['case_number']].tolist()
                    if dmr_cases:
                        dropped_cases_details['DMR Case'] = dmr_cases
            
            # 5. Remove duplicate cases based on Case Number
            self.logger.info(f"\n5. Removing duplicate cases:")
            pre_duplicate_count = len(df)
            # Normalize to integer case numbers for accurate deduplication
            if self.columns['case_number'] in df.columns:
                df[self.columns['case_number']] = df[self.columns['case_number']].apply(self.normalize_case_number)
                df[self.columns['case_number']] = pd.Series(pd.to_numeric(df[self.columns['case_number']], errors='coerce')).astype('Int64')
            
            # Sort by Created On date (newest first) before removing duplicates
            created_on_col = self.columns['created_on']
            if created_on_col in df.columns:
                self.logger.info(f"   Sorting by '{created_on_col}' to keep newest duplicate cases")
                # Parse Created On to datetime for proper sorting
                try:
                    # Remove extra spaces and parse the date
                    df['_created_on_parsed'] = pd.to_datetime(df[created_on_col].astype(str).str.replace(r'\s+', ' ', regex=True), 
                                                               format='%m/%d/%Y %I:%M:%S %p', errors='coerce')
                    
                    # If parsing fails, try alternative format
                    if df['_created_on_parsed'].isna().any():
                        df['_created_on_parsed'] = pd.to_datetime(df[created_on_col], errors='coerce')
                    
                    # Sort by the parsed date in descending order (newest first)
                    df = df.sort_values('_created_on_parsed', ascending=False, na_position='last')
                    self.logger.info(f"   Successfully sorted by Created On date")
                except Exception as e:
                    self.logger.warning(f"   Could not parse Created On dates for sorting: {str(e)}")
                    # Fallback: sort by original column as string
                    df = df.sort_values(created_on_col, ascending=False, na_position='last')
            
            duplicates = df[df.duplicated(subset=[self.columns['case_number']], keep='first')]
            df = df.dropna(subset=[self.columns['case_number']]).drop_duplicates(subset=[self.columns['case_number']], keep='first')
            
            # Drop the temporary sorting column if it exists
            if '_created_on_parsed' in df.columns:
                df = df.drop(columns=['_created_on_parsed'])
            
            post_duplicate_count = len(df)
            self.logger.info(f"   Records remaining: {post_duplicate_count}")
            
            # Log dropped cases from Duplicate filter
            if not duplicates.empty:
                dropped_reasons = {"Duplicate Case Number": duplicates[self.columns['case_number']].tolist()}
                self.log_filtering_summary("Duplicate Removal", pre_duplicate_count, post_duplicate_count, dropped_reasons)
                
                # Add to dropped cases details for summary table
                dropped_cases_details['Duplicate Case Number'] = duplicates[self.columns['case_number']].tolist()

            # --- Log cases present in both raw and previous file but missing from output ---
            if prev_df is not None and not prev_df.empty:
                # Integer-normalize for reliable comparisons
                prev_cn = pd.Series(prev_df['Case Number']).apply(self.normalize_case_number).astype('Int64')
                raw_cn = pd.Series(df['Case Number']).apply(self.normalize_case_number).astype('Int64')
                out_cn = pd.Series(df['Case Number']).apply(self.normalize_case_number).astype('Int64')
                prev_case_numbers = set(prev_cn.dropna().tolist())
                raw_case_numbers = set(raw_cn.dropna().tolist())
                output_case_numbers = set(out_cn.dropna().tolist())
                missing_in_output = (prev_case_numbers & raw_case_numbers) - output_case_numbers
                if missing_in_output:
                    self.logger.warning(f"Cases present in both raw and previous file but missing from output: {missing_in_output}")
                    for case_num in missing_in_output:
                        self.log_dropped_case(case_num, "Missing in output after filtering", "Present in both raw and previous file")

            # Final count after all filters
            filtered_count = len(df)
            self.logger.info(f"\n=== Filtering Summary ===")
            self.logger.info(f"Initial count: {initial_count}")
            self.logger.info(f"After Work Order Status filter: {wo_status_count}")
            self.logger.info(f"After Case Reason filter: {case_reason_count}")
            self.logger.info(f"After Closing Code filter: {closing_code_count}")
            self.logger.info(f"After CID/DMR filter: {post_duplicate_count}")
            self.logger.info(f"After Duplicate Removal: {filtered_count}")
            self.logger.info(f"Total removed: {initial_count - filtered_count}")

            # Calculate previous file statistics if prev_df is available
            prev_stats = {}
            if prev_df is not None and not prev_df.empty:
                prev_cn = pd.Series(prev_df['Case Number']).apply(self.normalize_case_number).astype('Int64')
                raw_cn = pd.Series(df['Case Number']).apply(self.normalize_case_number).astype('Int64')
                prev_case_numbers = set(prev_cn.dropna().tolist())
                raw_case_numbers = set(raw_cn.dropna().tolist())
                prev_stats['Initial Count'] = len(prev_df)
                prev_stats['Matching Cases'] = len(prev_case_numbers & raw_case_numbers)
                prev_stats['New Cases'] = len(raw_case_numbers - prev_case_numbers)
                # Updated cases: cases in both, but with any field different (simplified: count all matches as updated)
                updated_cases = 0
                for case_num in (prev_case_numbers & raw_case_numbers):
                    prev_row = prev_df[prev_df['Case Number'] == case_num]
                    raw_row = df[df['Case Number'] == case_num]
                    if not prev_row.equals(raw_row):
                        updated_cases += 1
                prev_stats['Updated Cases'] = updated_cases
            else:
                prev_stats['Initial Count'] = ''
                prev_stats['Matching Cases'] = ''
                prev_stats['New Cases'] = ''
                prev_stats['Updated Cases'] = ''

            # Format the output data and merge with previous file
            self.logger.info("\nFormatting output data and merging with previous file...")
            output_df = self.format_output_data(df, prev_df)

            # Assign handlers (all cases processed normally)
            self.logger.info("\nAssigning handlers...")
            output_df = self.assign_handlers(output_df, selected_handlers, prev_df, prev_file)

            # Calculate fair share distribution AFTER handler assignment (so we can see current workload)
            self.logger.info("\n=== Calculating Fair Share Distribution (AFTER handler assignment) ===")
            self.fair_share_info = self.calculate_fair_share(output_df, selected_handlers, chat_agent_info)

            # Redistribute cases to Chat Agent if enabled (AFTER fair share is calculated)
            if chat_agent_info and chat_agent_info.get('enabled'):
                self.logger.info("\n=== Redistributing Cases for Chat Agent Support ===")
                chat_agent_supporter_name = chat_agent_info.get('supporter_name', 'Chat Agent')
                output_df = self.redistribute_cases_to_chat_agent(output_df, chat_agent_info, chat_agent_supporter_name)

            # AFTER handler assignment: Identify email duplicates in NEW cases only
            self.logger.info("\n=== Identifying email duplicates in new cases for Companies sheet ===")
            output_df, self.duplicate_company_cases = self.identify_email_duplicates_new_cases(output_df, prev_df, prev_file)

            # Final cleanup - ensure no completely empty rows
            output_df = self.clean_empty_rows(output_df)
            final_count = len(output_df)
            self.logger.info(f"\nFinal record count: {final_count}")

            # Process DND Emails Database and update PA Cases
            self.logger.info("\nProcessing DND Emails Database...")
            output_df = self.process_dnd_emails_database(output_df, prev_file)

            # Prepare a place to collect handler sheet updates from SMS/Email processing
            updated_handler_sheets = {}
            # tentative prev file for final (may be replaced if we write updated handler sheets)
            prev_file_for_final = prev_file

            # Process SMS replies if provided
            sms_processing_stats = {}
            if sms_file and os.path.exists(sms_file):
                self.logger.info(f"\nProcessing SMS replies file: {sms_file}")
                output_df, sms_processing_stats, updated_handler_sheets = self.process_sms_replies(output_df, sms_file, prev_file)
                self.logger.info("SMS processing completed")
                # If handler sheets were updated in-memory, persist them to a temporary prev_file
                # so FinalProcessor will pick up the updated handler sheets when creating the final workbook.
                prev_file_for_final = prev_file
                if updated_handler_sheets:
                    try:
                        # Create a small excel file containing updated handler sheets + Issue Not Fixed + DND Emails
                        tmp_dir = os.path.dirname(output_file) or os.getcwd()
                        tmp_prev_path = os.path.join(tmp_dir, f"{os.path.splitext(os.path.basename(output_file))[0]}_prev_handlers.xlsx")
                        
                        self.logger.info(f"Creating comprehensive temp file with handlers, Issue Not Fixed, and DND Emails: {tmp_prev_path}")
                        
                        with pd.ExcelWriter(tmp_prev_path, engine='xlsxwriter') as tmp_writer:
                            # Write handler sheets
                            for sheet_name, sheet_df in updated_handler_sheets.items():
                                try:
                                    # Ensure DataFrame has proper columns and is safe to write
                                    if isinstance(sheet_df, pd.DataFrame) and not sheet_df.empty:
                                        sheet_df.to_excel(tmp_writer, sheet_name=sheet_name, index=False)
                                    else:
                                        # Write empty frame with Case Number header if nothing else
                                        pd.DataFrame(columns=['Case Number']).to_excel(tmp_writer, sheet_name=sheet_name, index=False)
                                except Exception:
                                    # Skip problematic sheets
                                    continue
                            
                            # ALSO load Issue Not Fixed, DND Emails, and Companies from the ORIGINAL prev_file if it exists
                            if prev_file and os.path.exists(prev_file):
                                try:
                                    prev_excel = pd.ExcelFile(prev_file)
                                    prev_sheets = prev_excel.sheet_names
                                    self.logger.info(f"Original prev_file sheets: {prev_sheets}")
                                    
                                    # Load and add Issue Not Fixed sheet
                                    if 'Issue Not Fixed' in prev_sheets:
                                        try:
                                            issue_not_fixed_df = pd.read_excel(prev_file, sheet_name='Issue Not Fixed')
                                            self.logger.info(f"Loading {len(issue_not_fixed_df)} Issue Not Fixed cases from original prev_file")
                                            issue_not_fixed_df.to_excel(tmp_writer, sheet_name='Issue Not Fixed', index=False)
                                            self.logger.info(f"✓ Added Issue Not Fixed sheet with {len(issue_not_fixed_df)} cases")
                                        except Exception as e:
                                            self.logger.warning(f"Could not load Issue Not Fixed from prev_file: {str(e)}")
                                    
                                    # Load and add DND Emails sheet
                                    if 'DND Emails' in prev_sheets:
                                        try:
                                            dnd_emails_df = pd.read_excel(prev_file, sheet_name='DND Emails')
                                            self.logger.info(f"Loading {len(dnd_emails_df)} DND Emails from original prev_file")
                                            dnd_emails_df.to_excel(tmp_writer, sheet_name='DND Emails', index=False)
                                            self.logger.info(f"✓ Added DND Emails sheet with {len(dnd_emails_df)} emails")
                                        except Exception as e:
                                            self.logger.warning(f"Could not load DND Emails from prev_file: {str(e)}")
                                    
                                    # Load and add Companies sheet (for preservation across runs)
                                    if 'Companies' in prev_sheets:
                                        try:
                                            companies_df = pd.read_excel(prev_file, sheet_name='Companies')
                                            self.logger.info(f"Loading {len(companies_df)} Companies cases from original prev_file")
                                            companies_df.to_excel(tmp_writer, sheet_name='Companies', index=False)
                                            self.logger.info(f"✓ Added Companies sheet with {len(companies_df)} cases")
                                        except Exception as e:
                                            self.logger.warning(f"Could not load Companies from prev_file: {str(e)}")
                                except Exception as e:
                                    self.logger.warning(f"Could not load Issue Not Fixed/DND Emails/Companies from prev_file: {str(e)}")
                        
                        prev_file_for_final = tmp_prev_path
                        self.logger.info(f" Wrote comprehensive temp file with handlers + Issue Not Fixed + DND Emails: {prev_file_for_final}")
                    except Exception as e:
                        self.logger.error(f"Could not write temporary prev handler file: {str(e)}")
                        import traceback
                        self.logger.error(f"Traceback: {traceback.format_exc()}")
                else:
                    prev_file_for_final = prev_file

            # Process Email replies if provided
            email_processing_stats = {}
            self.logger.info(f"\nEmail file path: {email_file}")
            if email_file:
                self.logger.info(f"Email file exists: {os.path.exists(email_file)}")
                if os.path.exists(email_file):
                    self.logger.info(f"Processing Email replies file: {email_file}")
                    # Collect handler sheet updates from email processing as well
                    output_df, email_processing_stats, email_updated_handler_sheets = self.process_email_replies(output_df, email_file, prev_file)
                    self.logger.info("Email processing completed")
                    # Merge any handler sheet updates from email processing into the updated_handler_sheets map
                    try:
                        if email_updated_handler_sheets:
                            if not updated_handler_sheets:
                                updated_handler_sheets = {}
                            for k, v in email_updated_handler_sheets.items():
                                updated_handler_sheets[k] = v
                    except Exception:
                        pass
                else:
                    self.logger.warning(f"Email file does not exist: {email_file}")
            else:
                self.logger.info("No email file provided")

            # Extract DND emails from raw file
            raw_dnd_emails = []
            if self.columns['dnd'] in df.columns and self.columns['email'] in df.columns:
                dnd_mask = df[self.columns['dnd']].astype(str).str.strip().str.lower() == 'yes'
                dnd_cases = df[dnd_mask]
                
                for idx, row in dnd_cases.iterrows():
                    email = str(row[self.columns['email']]).strip()
                    if email and email.lower() not in ['', 'nan', 'none']:
                        raw_dnd_emails.append({
                            'Email': email,
                            'DND': 'Yes'
                        })
                
                self.logger.info(f"Found {len(raw_dnd_emails)} DND emails from raw file")

            # Prepare processing statistics for final processor
            processing_stats = {
                'Initial Count': initial_count,
                'After Work Order Status': wo_status_count,
                'After Case Reason': case_reason_count,
                'After Closing Code': closing_code_count,
                'After CID/DMR Filter': post_cid_count,
                'After Duplicate Removal': filtered_count,
                'Final Count': final_count,
                'Total Removed': initial_count - filtered_count,
                'Raw File Final Count After Cleaning': filtered_count,
                'Previous File Stats': prev_stats,
                'SMS Processing Stats': sms_processing_stats,
                'Email Processing Stats': email_processing_stats,
                'Raw File DND Emails': raw_dnd_emails,
                'DND Database': getattr(self, 'dnd_database', []),
                'Dropped Cases Details': dropped_cases_details,
                'Bank/Sutherland Updated Cases': getattr(self, 'bank_sutherland_updated_cases', []),
                'DND Updated Cases': getattr(self, 'dnd_updated_cases', [])
            }
            
            # Log processing summary
            self.logger.info(f"Processing complete: {initial_count} → {final_count} records ({initial_count - final_count} removed)")
            
            # Use final processor for additional sheets and validations
            try:
                final_processor = FinalProcessor()
                
                # Validate output data before final processing
                validation_results, critical_issues = final_processor.validate_output_data(output_df)
                
                if critical_issues:
                    self.logger.warning("Critical issues found during validation:")
                    for issue in critical_issues:
                        self.logger.warning(f"  - {issue}")
                
                # Process final output with additional sheets
                # Use prev_file_for_final if we created one with updated handler sheets
                try:
                    prev_for_final = prev_file_for_final if 'prev_file_for_final' in locals() else prev_file
                except Exception:
                    prev_for_final = prev_file

                success, message = final_processor.process_final_output(
                    output_df, 
                    output_file, 
                    processing_stats, 
                    sms_file, 
                    email_file, 
                    prev_for_final, 
                    duplicate_company_cases=self.duplicate_company_cases,
                    selected_handlers=selected_handlers,
                    chat_agent_info=chat_agent_info
                )
                
                if not success:
                    raise Exception(f"Final processing failed: {message}")
            except ImportError as e:
                self.logger.error(f"Error importing FinalProcessor: {str(e)}")
                # Fallback: save basic output without additional sheets
                output_df.to_excel(output_file, index=False)
                self.logger.warning("Saved basic output without additional sheets due to import error")
            except Exception as e:
                self.logger.error(f"Error in final processing: {str(e)}")
                # Fallback: save basic output
                output_df.to_excel(output_file, index=False)
                self.logger.warning("Saved basic output due to final processing error")
            
            # Verify Chat Agent persistence if enabled
            if chat_agent_info and chat_agent_info.get('enabled'):
                persistence_info = self.verify_chat_agent_persistence(output_df, chat_agent_info, prev_file)
                self.logger.info(f"Chat Agent persistence info: {persistence_info}")
            
            return True, f"Processing completed successfully. Final record count: {final_count}"
            
        except Exception as e:
            self.logger.error(f"File processing error: {str(e)}")
            self.logger.error("Stack trace:", exc_info=True)
            return False, f"Error during file processing: {str(e)}"
    
    def load_previous_file(self, prev_file):
        """Load and process the previous file, allowing for flexible customer name column detection."""
        if not prev_file or not os.path.exists(prev_file):
            self.logger.info("No previous file provided or file doesn't exist")
            return pd.DataFrame()
        try:
            self.logger.info(f"Loading previous file: {prev_file}")
            prev_df = pd.read_excel(prev_file)
            prev_df = self._normalize_id_columns(prev_df)
            # --- Flexible customer name column detection ---
            customer_col = None
            for col in prev_df.columns:
                if '(customer)' in col.lower() or 'customer name' in col.lower():
                    customer_col = col
                    break
            if customer_col and customer_col != 'Customer Name':
                prev_df['Customer Name'] = prev_df[customer_col]
            # Store previous assignments in a dictionary
            if 'Case Number' in prev_df.columns and 'Assigned To' in prev_df.columns:
                self.previous_assignments = dict(zip(prev_df['Case Number'], prev_df['Assigned To']))
                self.logger.info(f"Loaded {len(self.previous_assignments)} previous assignments")
            else:
                self.previous_assignments = {}
            return prev_df
        except Exception as e:
            self.logger.error(f"Error loading previous file: {str(e)}")
            return pd.DataFrame()

    def process_sms_replies(self, output_df, sms_file, prev_file=None):
        """Process SMS replies file and update PA cases based on the latest SMS text values per case.

        Matching is performed against handler sheets ("<handler>'s Cases") inside `prev_file` when available.
        Only exact replies '1', '2', or '3' are considered actionable. Other replies are skipped.
        """

        try:
            self.logger.info("Starting SMS replies processing...")
            sms_df = pd.read_excel(sms_file)
            sms_df = self._normalize_id_columns(sms_df)
            self.logger.info(f"SMS file loaded with {len(sms_df)} records")
            self.logger.info(f"SMS file columns: {list(sms_df.columns)}")

            # Load handler sheets from previous file (sheet names ending with "'s Cases")
            handler_sheets_map = {}
            if prev_file and os.path.exists(prev_file):
                try:
                    excel = pd.ExcelFile(prev_file)
                    for sheet_name in excel.sheet_names:
                        try:
                            if not isinstance(sheet_name, str):
                                continue
                            if not sheet_name.endswith("'s Cases"):
                                continue
                            sheet_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                            # Normalize Case Number in the handler sheet if present
                            if 'Case Number' in sheet_df.columns:
                                sheet_df['Case Number'] = sheet_df['Case Number'].apply(self.normalize_case_number)
                                sheet_df['Case Number'] = pd.to_numeric(sheet_df['Case Number'], errors='coerce').astype('Int64')
                            handler_sheets_map[sheet_name] = sheet_df
                        except Exception:
                            # Skip problematic sheets
                            continue
                except Exception as e:
                    self.logger.warning(f"Could not read previous file for handler sheets: {str(e)}")

            # Robust column detection
            def find_col(df, target):
                for col in df.columns:
                    if col.strip().lower() == target.strip().lower():
                        return col
                for col in df.columns:
                    if target.strip().lower() in col.strip().lower():
                        return col
                return None

            sms_text_col = find_col(sms_df, 'SMS Text')
            case_number_col = find_col(sms_df, 'Case Number (Regarding) (Case)')
            # Try to find a timestamp/date column
            timestamp_col = None
            for col in sms_df.columns:
                col_lc = col.lower()
                if 'date' in col_lc or 'time' in col_lc or 'created' in col_lc:
                    timestamp_col = col
                    break
            if not sms_text_col or not case_number_col or not timestamp_col:
                self.logger.warning(f"Could not find required columns in SMS file. Columns found: {list(sms_df.columns)}")
                self.logger.warning(f"Expected: 'SMS Text', 'Case Number (Regarding) (Case)', and a date/time column.")
                return output_df, {'Skipped Cases': [{'Case Number': 'N/A', 'SMS Text': '', 'Reason': 'Required columns not found in SMS file'}]}, {}

            self.logger.info(f"Using columns - SMS Text: {sms_text_col}, Case Number: {case_number_col}, Timestamp: {timestamp_col}")

            # Convert timestamp to datetime
            sms_df[timestamp_col] = pd.to_datetime(sms_df[timestamp_col], errors='coerce')
            # Drop rows with missing case number or timestamp
            sms_df = sms_df.dropna(subset=[case_number_col, timestamp_col])
            # Sort and get latest SMS per case
            sms_df = sms_df.sort_values([case_number_col, timestamp_col])
            latest_sms = sms_df.groupby(case_number_col, as_index=False).last()

            summary = {'Fixed': [], 'Issue Not Fixed': [], 'Refused Callback': [], 'DND': [], 'Ambiguous': [], 'Updated SMS': []}
            skipped_cases = []

            # Only accept exact '1','2','3' as actionable replies
            def interpret_sms_text(sms_text):
                text = str(sms_text).strip()
                if text == '1':
                    return 'Fixed'
                if text == '2':
                    return 'Issue Not Fixed'
                if text == '3':
                    return 'DND'
                return None

            # --- HANDLER SHEET UPDATE LOGIC ---
            # After processing, update handler sheets if available
            def update_handler_sheets_from_output(output_df, handler_sheets_dict):
                # handler_sheets_dict: {handler_name: handler_df}
                if not handler_sheets_dict:
                    return
                for handler, handler_df in handler_sheets_dict.items():
                    # Normalize case numbers for matching
                    handler_df['Case Number'] = handler_df['Case Number'].apply(self.normalize_case_number).astype('Int64')
                    output_df['Case Number'] = output_df['Case Number'].apply(self.normalize_case_number).astype('Int64')
                    # Update rows in handler_df with values from output_df
                    for idx, row in handler_df.iterrows():
                        case_num = row['Case Number']
                        match = output_df[output_df['Case Number'] == case_num]
                        if not match.empty:
                            # Update status and final action
                            handler_df.at[idx, 'Status'] = match.iloc[0].get('Status', row.get('Status', ''))
                            handler_df.at[idx, 'Final Action'] = match.iloc[0].get('Final Action', row.get('Final Action', ''))
                return handler_sheets_dict
            # Usage: update_handler_sheets_from_output(output_df, handler_sheets_dict)

            for idx, row in latest_sms.iterrows():
                sms_text = str(row.get(sms_text_col, '')).strip()
                case_number = self.normalize_case_number(row.get(case_number_col, None))
                action = interpret_sms_text(sms_text)
                if case_number is None:
                    skipped_cases.append({
                        'Case Number': 'N/A',
                        'SMS Text': sms_text,
                        'Reason': 'Missing case number'
                    })
                    continue
                if action in ['Fixed', 'Issue Not Fixed', 'DND']:
                    # Match only against handler sheets (previous file) rather than PA Cases
                    found_in_handler = False
                    matched_sheet_name = None
                    matched_sheet_indices = []
                    for sheet_name, sheet_df in handler_sheets_map.items():
                        try:
                            if 'Case Number' not in sheet_df.columns:
                                continue
                            mask = sheet_df['Case Number'] == case_number
                            if mask.any():
                                found_in_handler = True
                                matched_sheet_name = sheet_name
                                matched_sheet_indices = sheet_df[mask].index.tolist()
                                # Update the handler sheet rows in-memory
                                for mi in matched_sheet_indices:
                                    if action == 'Fixed':
                                        sheet_df.at[mi, 'Final Action'] = 'Fixed'
                                        sheet_df.at[mi, 'Status'] = 'closed'
                                    elif action == 'Issue Not Fixed':
                                        sheet_df.at[mi, 'Final Action'] = 'Issue Not Fixed'
                                        sheet_df.at[mi, 'Status'] = 'closed'
                                    elif action == 'DND':
                                        sheet_df.at[mi, 'Final Action'] = 'DND'
                                        sheet_df.at[mi, 'Status'] = 'closed'
                                # store back updated sheet
                                handler_sheets_map[sheet_name] = sheet_df
                                break
                        except Exception:
                            continue

                    if found_in_handler:
                        # Per user's request: do NOT search/update PA Cases for SMS replies anymore.
                        # Always treat this as a handler-sheet update and record it for persistence.
                        update_made = f"Final Action set to {action}, Status=closed"
                        skipped_entry = {
                            'Case Number': case_number,
                            'SMS Text': sms_text,
                            'Reason': f'Handler-only update: {matched_sheet_name}',
                            'Source': 'SMS',
                            'Update Made': update_made,
                            'Handler Sheet': matched_sheet_name
                        }
                        skipped_cases.append(skipped_entry)
                        try:
                            summary.setdefault('Updated SMS', []).append({
                                'case_number': case_number,
                                'sms_text': sms_text,
                                'action': action,
                                'handler_sheet': matched_sheet_name,
                                'pa_updated': False,
                                'update_made': update_made
                            })
                        except Exception:
                            pass
                        self.logger.info(f"Handler-only update for SMS case {case_number}: {update_made} (handler: {matched_sheet_name}); PA Cases not checked/updated per configuration")
                    else:
                        reason = 'Case not found in handler sheets'
                        skipped_cases.append({
                            'Case Number': case_number,
                            'SMS Text': sms_text,
                            'Reason': reason
                        })
                        self.logger.info(f"Skipped SMS case {case_number}: {reason}")
                # All other replies are skipped
                else:
                    reason = f"Invalid or irrelevant SMS text: '{sms_text}'"
                    skipped_cases.append({
                        'Case Number': case_number,
                        'SMS Text': sms_text,
                        'Reason': reason
                    })
                    self.logger.info(f"Skipped SMS case {case_number}: {reason}")

            summary['Skipped Cases'] = skipped_cases
            # Log updated/skipped SMS summary
            try:
                self.logger.info(f"SMS updated count: {len(summary.get('Updated SMS', []))}")
                if summary.get('Updated SMS'):
                    self.logger.info(f"Updated SMS cases: {summary.get('Updated SMS')}")
                self.logger.info(f"SMS skipped count: {len(summary.get('Skipped Cases', []))}")
            except Exception:
                pass

            self.logger.info(f"SMS Processing Summary:")
            self.logger.info(f"Fixed: {len(summary['Fixed'])} cases")
            self.logger.info(f"Issue Not Fixed: {len(summary['Issue Not Fixed'])} cases")
            self.logger.info(f"Refused Callback: {len(summary['Refused Callback'])} cases")
            self.logger.info(f"DND: {len(summary['DND'])} cases")
            self.logger.info(f"Ambiguous (needs review): {len(summary['Ambiguous'])} cases")
            self.logger.info(f"Skipped: {len(skipped_cases)} cases")
            
            # Log comprehensive summary for Issue Not Fixed cases
            if summary['Issue Not Fixed']:
                self.log_issue_not_fixed_summary(
                    total_cases=len(summary['Issue Not Fixed']),
                    sms_cases=len(summary['Issue Not Fixed']),
                    email_cases=0,
                    new_cases=len(summary['Issue Not Fixed'])
                )

            return output_df, summary, handler_sheets_map
        except Exception as e:
            self.logger.error(f"Error processing SMS replies: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return output_df, {}, {}

    def process_email_replies(self, output_df, email_file, prev_file=None):
        """Process Email replies in the same strict way as SMS replies.

        - Accept only exact replies '1', '2', '3' (after trimming).
        - Case numbers: prefer header 'Case Number (Object) (Email)', else column F (index 5).
        - Replies: prefer header 'Action', else column H (index 7).
        - Only update handler sheets (sheets ending with "'s Cases") from prev_file; do not update PA Cases.
        - Return a summary containing a unified 'Skipped/Updated Entries' list of dicts with keys:
          Timestamp, Source, Case Number, Reply Text, Result, Reason
        """

        try:
            self.logger.info("Starting Email replies processing (SMS-style)...")
            email_df = pd.read_excel(email_file)
            email_df = self._normalize_id_columns(email_df)
            self.logger.info(f"Email file loaded with {len(email_df)} records")

            # Load handler sheets from previous file (sheet names ending with "'s Cases")
            handler_sheets_map = {}
            if prev_file and os.path.exists(prev_file):
                try:
                    excel = pd.ExcelFile(prev_file)
                    for sheet_name in excel.sheet_names:
                        try:
                            if not isinstance(sheet_name, str):
                                continue
                            if not sheet_name.endswith("'s Cases"):
                                continue
                            sheet_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                            if 'Case Number' in sheet_df.columns:
                                sheet_df['Case Number'] = sheet_df['Case Number'].apply(self.normalize_case_number)
                                sheet_df['Case Number'] = pd.to_numeric(sheet_df['Case Number'], errors='coerce').astype('Int64')
                            handler_sheets_map[sheet_name] = sheet_df
                        except Exception:
                            continue
                except Exception as e:
                    self.logger.warning(f"Could not read previous file for handler sheets (emails): {str(e)}")

            # Column resolution: prefer named headers, otherwise use Excel column indices (F=5, H=7)
            def col_by_name_or_index(df, name, idx):
                if name in df.columns:
                    return name
                if len(df.columns) > idx:
                    return df.columns[idx]
                return None

            case_col = col_by_name_or_index(email_df, 'Case Number (Object) (Email)', 5)
            reply_col = col_by_name_or_index(email_df, 'Action', 7)

            # Timestamp column: try to find a date/time-like column; else None
            timestamp_col = None
            for col in email_df.columns:
                col_lc = str(col).lower()
                if 'date' in col_lc or 'time' in col_lc or 'created' in col_lc:
                    timestamp_col = col
                    break

            if not case_col or not reply_col:
                self.logger.warning(f"Could not find required email columns. Found columns: {list(email_df.columns)}")
                return output_df, {'Skipped/Updated Entries': [{'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Source': 'Email', 'Case Number': 'N/A', 'Reply Text': '', 'Result': 'Skipped', 'Reason': 'Required columns not found in email file'}]}, {}

            self.logger.info(f"Using columns - Case Number: {case_col}, Reply: {reply_col}, Timestamp: {timestamp_col}")

            # Prepare dataframe: convert timestamp if present
            if timestamp_col:
                email_df[timestamp_col] = pd.to_datetime(email_df[timestamp_col], errors='coerce')
                email_df = email_df.dropna(subset=[case_col])
                email_df = email_df.sort_values([case_col, timestamp_col])
                latest = email_df.groupby(case_col, as_index=False).last()
            else:
                # No timestamp: use last occurrence in file order
                latest = email_df.dropna(subset=[case_col]).copy()
                latest = latest.groupby(case_col, as_index=False).last()

            # Interpret replies: exact matches only
            def interpret_reply(text):
                t = str(text).strip()
                if t == '1':
                    return 'Fixed'
                if t == '2':
                    return 'Issue Not Fixed'
                if t == '3':
                    return 'DND'
                return None

            skipped_or_updated_entries = []
            summary = {'Fixed': [], 'Issue Not Fixed': [], 'DND': [], 'Ambiguous': [], 'Updated Emails': [], 'Skipped/Updated Entries': skipped_or_updated_entries}

            for idx, row in latest.iterrows():
                raw_case = row.get(case_col, '')
                case_number = self.normalize_case_number(raw_case)
                reply_text = str(row.get(reply_col, '')).strip()
                ts_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                if timestamp_col:
                    val = row.get(timestamp_col)
                    try:
                        ts = pd.to_datetime(str(val), errors='coerce')
                        if isinstance(ts, pd.Timestamp) and not pd.isna(ts):
                            ts_str = ts.strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        ts_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                action = interpret_reply(reply_text)

                # Normalize case existence check
                if case_number is None:
                    skipped_or_updated_entries.append({'Timestamp': ts_str, 'Source': 'Email', 'Case Number': 'N/A', 'Reply Text': reply_text, 'Result': 'Skipped', 'Reason': 'Missing or invalid case number'})
                    continue

                if action in ['Fixed', 'Issue Not Fixed', 'DND']:
                    # Attempt to find and update in handler sheets only
                    found_in_handler = False
                    matched_sheet = None
                    for sheet_name, sheet_df in handler_sheets_map.items():
                        try:
                            if 'Case Number' not in sheet_df.columns:
                                continue
                            mask = sheet_df['Case Number'] == case_number
                            if mask.any():
                                found_in_handler = True
                                matched_sheet = sheet_name
                                indices = sheet_df[mask].index.tolist()
                                for mi in indices:
                                    if action == 'Fixed':
                                        sheet_df.at[mi, 'Final Action'] = 'Fixed'
                                        sheet_df.at[mi, 'Status'] = 'closed'
                                    elif action == 'Issue Not Fixed':
                                        sheet_df.at[mi, 'Final Action'] = 'Issue Not Fixed'
                                        sheet_df.at[mi, 'Status'] = 'closed'
                                    elif action == 'DND':
                                        sheet_df.at[mi, 'Final Action'] = 'DND'
                                        sheet_df.at[mi, 'Status'] = 'closed'
                                handler_sheets_map[sheet_name] = sheet_df
                                break
                        except Exception:
                            continue

                    if found_in_handler:
                        skipped_or_updated_entries.append({'Timestamp': ts_str, 'Source': 'Email', 'Case Number': str(case_number), 'Reply Text': reply_text, 'Result': 'Updated', 'Reason': f'Handler-only update: {matched_sheet}'})
                        summary.setdefault('Updated Emails', []).append({'case_number': case_number, 'reply_text': reply_text, 'action': action, 'handler_sheet': matched_sheet, 'pa_updated': False})
                        self.logger.info(f"Handler-only update for Email case {case_number} in {matched_sheet}")
                    else:
                        skipped_or_updated_entries.append({'Timestamp': ts_str, 'Source': 'Email', 'Case Number': str(case_number), 'Reply Text': reply_text, 'Result': 'Skipped', 'Reason': 'Case not found in handler sheets'})
                        self.logger.info(f"Skipped Email case {case_number}: not found in handler sheets")
                else:
                    skipped_or_updated_entries.append({'Timestamp': ts_str, 'Source': 'Email', 'Case Number': str(case_number), 'Reply Text': reply_text, 'Result': 'Skipped', 'Reason': 'Invalid or non-actionable reply'})

            summary['Skipped/Updated Entries'] = skipped_or_updated_entries
            self.logger.info(f"Email processing completed: {len(skipped_or_updated_entries)} skipped/updated entries recorded")
            return output_df, summary, handler_sheets_map
        except Exception as e:
            self.logger.error(f"Error processing Email replies: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return output_df, {}, {}

    def _interpret_sms_text_for_test(self, sms_text):
        """Simplified SMS interpreter for runtime: accept only exact '1','2','3'.

        Returns:
            'Fixed' | 'Issue Not Fixed' | 'DND' | None
        """
        text = str(sms_text).strip()
        if text == '1':
            return 'Fixed'
        if text == '2':
            return 'Issue Not Fixed'
        if text == '3':
            return 'DND'
        return None

    def _interpret_email_reply_for_test(self, reply_text):
        """Test method for email interpretation - extracts the logic for testing"""
        text = reply_text.lower().strip()
        
        # HIGH CONFIDENCE patterns (definite matches)
        high_confidence_fixed = [
            r'^\s*1\s*$',                           # Just "1"
            r'\b1\b.*issue resolved',               # "1 issue resolved"
            r'\bissue resolved\b',                  # "issue resolved"
            r'\bfixed\b',                           # "fixed"
            r'\bresolved\b',                        # "resolved"
            r'\bsorted\b',                          # "sorted"
            r'\bdone\b',                            # "done"
            r'\bcomplete\b',                        # "complete"
            r'\bfinished\b',                        # "finished"
            r'\bworking\b.*\bnow\b',               # "working now"
            r'\bproblem\b.*\bsolved\b',             # "problem solved"
            r'\ball\b.*\bgood\b',                   # "all good"
        ]
        
        high_confidence_not_fixed = [
            r'^\s*2\s*$',                           # Just "2"
            r'\b2\b.*need assistance',               # "2 need assistance"
            r'\bnot fixed\b',                       # "not fixed"
            r'\bissue not fixed\b',                 # "issue not fixed"
            r'\bstill broken\b',                    # "still broken"
            r'\bproblem persists\b',                # "problem persists"
            r'\bnot working\b',                     # "not working"
            r'\bneed\b.*\bhelp\b',                  # "need help"
            r'\bstill\b.*\bissue\b',                # "still issue"
        ]
        
        high_confidence_dnd = [
            r'^\s*2\s*$',                           # Just "2"
            r'\b3\b.*stop',                         # "3 stop"
            r'\bstop text messages\b',               # "stop text messages"
            r'\bdnd\b',                             # "dnd"
            r'\bdo not disturb\b',                   # "do not disturb"
            r'\bopt out\b',                         # "opt out"
            r'\bunsubscribe\b',                     # "unsubscribe"
        ]
        
        # MEDIUM CONFIDENCE patterns (likely matches)
        medium_confidence_fixed = [
            r'\bissue\b.*\bresolved\b',             # "issue resolved"
            r'\bresolved\b.*\bissue\b',             # "resolved issue"
            r'\bsorted\b.*\bout\b',                 # "sorted out"
            r'\bworking\b.*\bproperly\b',           # "working properly"
            r'\bissue\b.*\bfixed\b',                # "issue fixed"
            r'\bfixed\b.*\bissue\b',                # "fixed issue"
        ]
        
        medium_confidence_not_fixed = [
            r'\bassistance\b.*\bneeded\b',          # "assistance needed"
            r'\bproblem\b.*\bcontinues\b',          # "problem continues"
            r'\bissue\b.*\bremains\b',              # "issue remains"
            r'\bnot\b.*\bresolved\b',               # "not resolved"
        ]
        
        # Check HIGH CONFIDENCE patterns first (definite matches)
        for pat in high_confidence_fixed:
            if re.search(pat, text):
                return 'Fixed'
        for pat in high_confidence_not_fixed:
            if re.search(pat, text):
                return 'Issue Not Fixed'
        for pat in high_confidence_dnd:
            if re.search(pat, text):
                return 'DND'
        
        # Check MEDIUM CONFIDENCE patterns (likely matches)
        for pat in medium_confidence_fixed:
            if re.search(pat, text):
                return 'Fixed'
        for pat in medium_confidence_not_fixed:
            if re.search(pat, text):
                return 'Issue Not Fixed'
        
        # If we get here, the text is unclear - log it for review
        # This prevents losing potentially important responses
        return 'Ambiguous'  # New status for unclear responses

    def process_dnd_emails_database(self, output_df, prev_file):
        """Process DND Emails Database and update PA Cases accordingly"""
        try:
            self.logger.info("Starting DND Emails Database processing...")
            
            # Load DND emails from previous file's database
            prev_dnd_emails = self.load_previous_dnd_database(prev_file)
            self.logger.info(f"Loaded {len(prev_dnd_emails)} DND emails from previous file database")
            
            # Extract current DND emails from PA Cases
            current_dnd_emails = []
            if 'DND (Do Not Disturb)' in output_df.columns and 'Email' in output_df.columns:
                dnd_mask = output_df['DND (Do Not Disturb)'].astype(str).str.strip().str.lower() == 'yes'
                dnd_cases = output_df[dnd_mask]
                
                for idx, row in dnd_cases.iterrows():
                    email = str(row['Email']).strip()
                    if email and email.lower() not in ['', 'nan', 'none']:
                        current_dnd_emails.append({
                            'Email': email,
                            'DND': 'Yes'
                        })
                
                self.logger.info(f"Found {len(current_dnd_emails)} DND emails from current PA Cases")
            
            # Combine DND emails (previous + current)
            all_dnd_emails = prev_dnd_emails.copy()
            existing_emails = {email['Email'] for email in prev_dnd_emails}
            
            # Add new DND emails from current PA Cases
            new_dnd_emails = []
            for dnd_email in current_dnd_emails:
                if dnd_email['Email'] not in existing_emails:
                    all_dnd_emails.append(dnd_email)
                    existing_emails.add(dnd_email['Email'])
                    new_dnd_emails.append(dnd_email['Email'])
            
            self.logger.info(f"Added {len(new_dnd_emails)} new DND emails to database")
            
            # Store the combined DND database for later use
            self.dnd_database = all_dnd_emails
            
            # Update PA Cases based on DND database
            updated_cases = self.update_pa_cases_with_dnd_database(output_df, all_dnd_emails)
            
            self.logger.info(f"Updated {updated_cases} PA Cases with DND status")
            
            return output_df
            
        except Exception as e:
            self.logger.error(f"Error processing DND Emails Database: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return output_df

    def load_previous_dnd_database(self, prev_file):
        """Load DND emails from previous file's DND Emails Database sheet"""
        try:
            if not prev_file or not os.path.exists(prev_file):
                self.logger.info("No previous file provided or file doesn't exist - starting with empty DND database")
                return []
            
            # Try to read the previous file
            try:
                prev_df = pd.read_excel(prev_file, sheet_name='DND Emails Database')
                self.logger.info(f"Successfully loaded DND Emails Database from previous file with {len(prev_df)} emails")
                
                # Convert to list of dictionaries
                dnd_emails = []
                if 'Email' in prev_df.columns and 'DND' in prev_df.columns:
                    for idx, row in prev_df.iterrows():
                        email = str(row['Email']).strip()
                        if email and email.lower() not in ['', 'nan', 'none']:
                            dnd_emails.append({
                                'Email': email,
                                'DND': 'Yes'
                            })
                
                return dnd_emails
                
            except Exception as e:
                self.logger.warning(f"Could not read DND Emails Database from previous file: {str(e)}")
                self.logger.info("Starting with empty DND database")
                return []
                
        except Exception as e:
            self.logger.error(f"Error loading previous DND database: {str(e)}")
            return []

    def update_pa_cases_with_dnd_database(self, output_df, dnd_database):
        """Update PA Cases based on DND database matches"""
        try:
            if not dnd_database or 'Email' not in output_df.columns:
                self.logger.info("No DND database or Email column found - skipping PA Cases update")
                return 0
            
            # Create set of DND emails for faster lookup
            dnd_emails_set = {email['Email'].lower() for email in dnd_database}
            
            # Find matching cases
            updated_count = 0
            
            for idx, row in output_df.iterrows():
                email = str(row['Email']).strip().lower()
                if email and email in dnd_emails_set:
                    # Update the case to DND status
                    output_df.at[idx, 'DND (Do Not Disturb)'] = 'Yes'
                    output_df.at[idx, 'Action 1'] = 'DND'
                    output_df.at[idx, 'Action 2'] = 'DND'
                    output_df.at[idx, 'Action 3'] = 'DND'
                    output_df.at[idx, 'Final Action'] = 'DND'
                    output_df.at[idx, 'Status'] = 'Closed'
                    updated_count += 1
                    
                    # Log the update
                    case_number = str(row.get('Case Number', 'N/A'))
                    self.logger.info(f"Updated case {case_number} to DND status (email: {email})")
            
            self.logger.info(f"Total PA Cases updated to DND status: {updated_count}")
            return updated_count
            
        except Exception as e:
            self.logger.error(f"Error updating PA Cases with DND database: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return 0

    def clean_handler_name(self, handler):
        """Clean handler name by removing any extra formatting, booleans, or special characters"""
        if pd.isna(handler) or not str(handler).strip():
            return ''
        
        handler = str(handler).strip()
        
        # If it's a tuple string like "(name, False)", extract just the name
        if handler.startswith('(') and handler.endswith(')'):
            handler = handler[1:-1].split(',')[0]
            
        # Remove common artifacts
        handler = handler.replace("'", "").replace('"', '').replace('(', '').replace(')', '')
        handler = handler.split(',')[0]  # Remove everything after a comma
        handler = handler.replace('False', '').replace('True', '')  # Remove boolean values
        
        # Remove any remaining whitespace
        handler = handler.strip()
        
        return handler

    def normalize_case_number(self, case_number):
        """Normalize case number to an integer where possible to avoid '.0' string artifacts.

        Returns:
            int | pandas._libs.missing.NAType | None: Integer case number when numeric, NA/None otherwise
        """
        if case_number is None or (isinstance(case_number, float) and pd.isna(case_number)):
            return None

        text_value = str(case_number).strip()
        if text_value == "" or text_value.lower() in {"nan", "none"}:
            return None

        # If it is an integer-like number possibly with a trailing .0, coerce to int
        if re.fullmatch(r"\d+(?:\.0+)?", text_value):
            try:
                return int(float(text_value))
            except Exception:
                pass

        # Try general numeric coercion then keep only integers
        try:
            numeric_value = pd.to_numeric(text_value, errors='coerce')
            if pd.notna(numeric_value) and float(numeric_value).is_integer():
                return int(numeric_value)
        except Exception:
            pass

        # If still not numeric, try to extract a contiguous digit sequence (common when values contain notes)
        try:
            m = re.search(r"(\d+)", text_value)
            if m:
                return int(m.group(1))
        except Exception:
            pass

        # Non-numeric or non-integer-like case numbers are treated as missing for dedup/compare
        return None

    def identify_email_duplicates_new_cases(self, df, prev_df=None, prev_file=None):
        """
        Identify NEW cases that have duplicate emails (2+ occurrences).
        Only checks cases that are NOT in the previous file.
        These cases will be added to the Companies sheet and preserved across runs.
        
        CRITICAL: Checks against ALL sheets in previous file (PA Cases + handler sheets + Companies)
        to correctly identify truly new cases.
        
        Args:
            df: Current processed DataFrame
            prev_df: Previous file DataFrame (for identifying new cases)
            prev_file: Path to previous file (for loading all sheets)
        
        Returns: 
            tuple (df_with_cleared_assigned_to, duplicate_cases_df)
            - df_with_cleared_assigned_to: Original df with Assigned To cleared for duplicate cases
            - duplicate_cases_df: DataFrame of cases with duplicate emails (for Companies sheet)
        """
        try:
            self.logger.info("=== Analyzing new cases for email duplicates ===")
            
            # Make a copy to avoid modifying original
            df = df.copy()
            
            # Build comprehensive set of ALL previous case numbers from ALL sheets
            prev_case_numbers = set()
            
            # Add cases from prev_df (PA Cases - first sheet)
            if prev_df is not None and not prev_df.empty and 'Case Number' in prev_df.columns:
                for cn in prev_df['Case Number'].dropna():
                    normalized = self.normalize_case_number(cn)
                    if normalized is not None:
                        prev_case_numbers.add(normalized)
                self.logger.info(f"Loaded {len(prev_case_numbers)} case numbers from PA Cases")
            
            # ALSO load case numbers from ALL handler sheets and Companies in the Excel file
            # This is critical - otherwise cases from handler sheets would be treated as "new"
            import os
            if prev_file and os.path.exists(prev_file):
                try:
                    with pd.ExcelFile(prev_file) as excel:
                        for sheet_name in excel.sheet_names:
                            if not isinstance(sheet_name, str):
                                continue
                            # Include handler sheets and Companies sheet
                            if sheet_name.endswith("'s Cases") or 'companies' in sheet_name.lower():
                                try:
                                    sheet_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                                    if 'Case Number' in sheet_df.columns:
                                        for cn in sheet_df['Case Number'].dropna():
                                            normalized = self.normalize_case_number(cn)
                                            if normalized is not None:
                                                prev_case_numbers.add(normalized)
                                except Exception:
                                    continue
                    self.logger.info(f"After loading all sheets: {len(prev_case_numbers)} total previous case numbers")
                except Exception as e:
                    self.logger.warning(f"Could not load additional sheets from prev file: {e}")
            
            # Identify new cases (not in ANY previous sheet)
            if prev_case_numbers:
                df['_normalized_case_num'] = df['Case Number'].apply(self.normalize_case_number)
                new_cases_mask = ~df['_normalized_case_num'].isin(prev_case_numbers)
                new_cases = df[new_cases_mask].copy()
                
                self.logger.info(f"Total cases: {len(df)}, Previous cases (all sheets): {len(prev_case_numbers)}, New cases: {len(new_cases)}")
            else:
                new_cases = df.copy()
                new_cases['_normalized_case_num'] = new_cases['Case Number'].apply(self.normalize_case_number)
                self.logger.info(f"No previous file - checking all {len(new_cases)} cases for email duplicates")
            
            if new_cases.empty:
                self.logger.info("No new cases to check for email duplicates")
                # Drop temp column if added
                if '_normalized_case_num' in df.columns:
                    df = df.drop(columns=['_normalized_case_num'])
                return df, pd.DataFrame()
            
            # Check for Email column
            if 'Email' not in new_cases.columns:
                self.logger.warning("No Email column found - cannot detect email duplicates")
                if '_normalized_case_num' in df.columns:
                    df = df.drop(columns=['_normalized_case_num'])
                return df, pd.DataFrame()
            
            # Normalize emails for comparison (case-insensitive, trimmed)
            new_cases['_email_normalized'] = new_cases['Email'].astype(str).str.strip().str.lower()
            
            # CRITICAL: Only count NEW/OPEN cases (not closed) for duplicate detection
            # This prevents adding cases where email has 1 open + several closed
            if 'Status' in new_cases.columns:
                open_statuses = ['new', 'in_progress', 'in progress', 'in progress today', 'skipped', '']
                status_normalized = new_cases['Status'].astype(str).str.strip().str.lower()
                open_cases_mask = status_normalized.isin(open_statuses) | status_normalized.isna()
                new_open_cases = new_cases[open_cases_mask].copy()
                self.logger.info(f"Filtered to {len(new_open_cases)} NEW/OPEN cases (from {len(new_cases)} new cases) for duplicate check")
            else:
                new_open_cases = new_cases.copy()
                self.logger.info("No Status column - using all new cases for duplicate check")
            
            # Find emails that appear 2+ times in NEW/OPEN cases only
            email_counts = new_open_cases['_email_normalized'].value_counts()
            duplicate_emails = email_counts[email_counts >= 2].index.tolist()
            
            # Remove blank/invalid emails from duplicate list
            invalid_emails = ['', 'nan', 'none', 'null', 'n/a', 'na']
            duplicate_emails = [e for e in duplicate_emails if e and e not in invalid_emails]
            
            if not duplicate_emails:
                self.logger.info("No email duplicates found in new cases")
                # Drop temp columns
                if '_normalized_case_num' in df.columns:
                    df = df.drop(columns=['_normalized_case_num'])
                return df, pd.DataFrame()
            
            self.logger.info(f"Found {len(duplicate_emails)} duplicate email addresses in new cases:")
            for email in duplicate_emails[:10]:  # Log first 10 for brevity
                count = email_counts[email]
                self.logger.info(f"  '{email}': {count} occurrences")
            if len(duplicate_emails) > 10:
                self.logger.info(f"  ... and {len(duplicate_emails) - 10} more")
            
            # FIXED: Get only OPEN cases with duplicate emails (not closed cases for same email)
            # This ensures only 2+ OPEN cases go to Companies sheet
            duplicate_cases = new_open_cases[new_open_cases['_email_normalized'].isin(duplicate_emails)].copy()
            
            # Drop temporary columns from duplicate_cases
            temp_cols = ['_email_normalized', '_normalized_case_num']
            for col in temp_cols:
                if col in duplicate_cases.columns:
                    duplicate_cases = duplicate_cases.drop(columns=[col])
            
            # Clear Assigned To column for these cases (will be assigned in Companies sheet)
            duplicate_cases['Assigned To'] = ''
            
            # Get case numbers of duplicate cases
            duplicate_case_numbers = set(duplicate_cases['Case Number'].dropna().apply(self.normalize_case_number))
            
            # Clear Assigned To in main df for duplicate cases
            df['_normalized_case_num'] = df['Case Number'].apply(self.normalize_case_number)
            df.loc[df['_normalized_case_num'].isin(duplicate_case_numbers), 'Assigned To'] = ''
            
            # Drop temp column from main df
            df = df.drop(columns=['_normalized_case_num'])
            self.logger.info(f"✓ Identified {len(duplicate_cases)} OPEN cases with duplicate emails")
            self.logger.info(f"  These cases will be added to the Companies sheet")
            
            return df, duplicate_cases
            
        except Exception as e:
            self.logger.warning(f"Error identifying email duplicates: {str(e)}")
            import traceback
            self.logger.warning(f"Traceback: {traceback.format_exc()}")
            # Clean up any temp columns
            if '_normalized_case_num' in df.columns:
                df = df.drop(columns=['_normalized_case_num'])
            return df, pd.DataFrame()

    def assign_handlers(self, df, selected_handlers, prev_df=None, prev_file=None):
        """Assign handlers for the current dataset only (no preservation from previous sheets).
        - Do NOT read or rely on previous handler sheets or special preservation columns
        - Distribute cases fairly across SELECTED handlers based on company grouping
        - Preservation is handled later during per-handler sheet merging
        """
        try:
            self.logger.info("=== SEPARATED APPROACH: Preserve ALL Previous Work + Assign NEW Cases to SELECTED Handlers Only ===")
            self.logger.info(f"RAW selected_handlers received: {selected_handlers}")
            cleaned_handlers = [self.clean_handler_name(h) for h in selected_handlers if self.clean_handler_name(h)]
            self.logger.info(f"SELECTED handlers for NEW case assignment: {cleaned_handlers}")
            
            if 'Case Number' not in df.columns:
                raise ValueError("Case Number column not found in dataframe")
            if 'Assigned To' not in df.columns:
                df['Assigned To'] = ''
            
            selected_handlers_set = set(cleaned_handlers)

            # STEP 1: NORMALIZE ALL CASE NUMBERS FIRST
            self.logger.info("=== STEP 1: NORMALIZING ALL CASE NUMBERS ===")
            df['Case Number'] = df['Case Number'].apply(self.normalize_case_number)
            # Enforce integer dtype with nullable Int64 to preserve NA
            df['Case Number'] = pd.to_numeric(df['Case Number'], errors='coerce').astype('Int64')
            self.logger.info(f"Normalized all case numbers in current data (stored as integers)")
            
            # STEP 2: Preserve existing assignments from previous main file (LOCK old cases)
            self.logger.info("=== STEP 2: PRESERVING previous-file assignments (lock old cases) ===")
            prev_assignments = set()
            prev_case_to_handler = {}
            if prev_df is not None and not prev_df.empty:
                try:
                    prev_df_local = prev_df.copy()
                    # Normalize case numbers in previous file
                    prev_df_local['Case Number'] = prev_df_local['Case Number'].apply(self.normalize_case_number)
                    prev_df_local['Case Number'] = pd.to_numeric(prev_df_local['Case Number'], errors='coerce').astype('Int64')
                    # Build mapping and set for quick lookup
                    for _, row in prev_df_local.dropna(subset=['Case Number']).iterrows():
                        case_num = int(row['Case Number'])
                        handler = self.clean_handler_name(row.get('Assigned To', ''))
                        if handler:
                            prev_case_to_handler[case_num] = handler
                            prev_assignments.add(case_num)
                    # Apply preserved assignments to current df where case exists
                    for case_num, handler in prev_case_to_handler.items():
                        mask = df['Case Number'] == case_num
                        if mask.any():
                            idx = mask.idxmax()
                            df.at[idx, 'Assigned To'] = handler
                    self.logger.info(f"Preserved {len(prev_assignments)} existing cases from previous file")
                except Exception as preserve_err:
                    self.logger.warning(f"Could not preserve previous assignments from main file: {str(preserve_err)}")

            # Also preserve assignments from previous individual handler sheets (by case number → handler)
            try:
                if prev_file and os.path.exists(prev_file):
                    excel_file = pd.ExcelFile(prev_file)
                    handler_sheets = [sheet for sheet in excel_file.sheet_names if isinstance(sheet, str) and sheet.endswith("'s Cases")]
                    preserved_from_sheets = 0
                    for sheet_name in handler_sheets:
                        handler_name = sheet_name.replace("'s Cases", '').strip()
                        try:
                            sheet_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                            if 'Case Number' not in sheet_df.columns:
                                continue
                            # Normalize and iterate
                            sheet_df_local = sheet_df.copy()
                            sheet_df_local['Case Number'] = sheet_df_local['Case Number'].apply(self.normalize_case_number)
                            sheet_df_local['Case Number'] = pd.to_numeric(sheet_df_local['Case Number'], errors='coerce').astype('Int64')
                            for _, row in sheet_df_local.dropna(subset=['Case Number']).iterrows():
                                case_num = int(row['Case Number'])
                                prev_case_to_handler[case_num] = handler_name
                                if case_num not in prev_assignments:
                                    prev_assignments.add(case_num)
                                    preserved_from_sheets += 1
                        except Exception:
                            continue
                    # Apply preserved assignments from sheets to current df
                    applied = 0
                    for case_num, handler in prev_case_to_handler.items():
                        mask = df['Case Number'] == case_num
                        if mask.any():
                            idx = mask.idxmax()
                            df.at[idx, 'Assigned To'] = handler
                            applied += 1
                    self.logger.info(f"Preserved {preserved_from_sheets} new cases from previous handler sheets; applied to {applied} rows in current data")
            except Exception as e:
                self.logger.warning(f"Could not preserve from previous handler sheets: {str(e)}")

            # STEP 3: Calculate current workload after preservation
            self.logger.info("=== STEP 3: Calculating Current Workload After Preservation ===")
            current_workload = {}
            for handler in cleaned_handlers:
                count = len(df[df['Assigned To'] == handler])
                current_workload[handler] = count
            
            self.logger.info(f"Current workload after preservation: {current_workload}")
            
            # STEP 4: Assign ONLY NEW cases to SELECTED handlers (do not touch preserved ones)
            self.logger.info("=== STEP 4: Assigning ONLY NEW cases to SELECTED handlers ===")
            if prev_assignments:
                new_cases_df = df[~df['Case Number'].isin(prev_assignments)].copy()
            else:
                new_cases_df = df.copy()
            
            if not new_cases_df.empty:
                self.logger.info(f"Processing {len(new_cases_df)} NEW cases for assignment")
                self.logger.info(f"Cases will ONLY be assigned to SELECTED handlers: {cleaned_handlers}")
                # Extra log: preserved vs new per handler before assignment
                try:
                    preserved_counts = {}
                    for handler in cleaned_handlers:
                        preserved_counts[handler] = len(df[(df['Assigned To'] == handler) & (df['Case Number'].isin(prev_assignments))])
                    self.logger.info(f"Preserved counts (locked from previous): {preserved_counts}")
                except Exception:
                    pass
                
                # Get company column
                company_col = self._get_company_column(df)
                if company_col is None:
                    self.logger.warning("No company column found - cannot do company-based assignment")
                    return df
                
                # FAIR DISTRIBUTION LOGIC: Only assign to SELECTED handlers
                self.logger.info("=== FAIR DISTRIBUTION: Assigning cases only to SELECTED handlers ===")
                
                # Group new cases by company
                new_cases_df['Company_Lower'] = new_cases_df[company_col].fillna('').astype(str).str.strip().str.lower()
                company_groups = new_cases_df.groupby('Company_Lower')
                
                self.logger.info(f"Found {len(company_groups)} unique companies in NEW cases")
                
                # Separate companies with multiple cases from single cases and no-company cases
                multi_case_companies = {}
                single_case_companies = {}
                no_company_cases = []
                
                for company, group in company_groups:
                    case_count = len(group)
                    if company == '' or company == 'nan' or company == 'none':
                        # No company cases
                        no_company_cases.extend(group['Case Number'].tolist())
                    elif case_count == 1:
                        # Single case companies
                        single_case_companies[company] = group['Case Number'].iloc[0]
                    else:
                        # Multi-case companies
                        multi_case_companies[company] = group['Case Number'].tolist()
                
                self.logger.info(f"Company breakdown:")
                self.logger.info(f"  Multi-case companies: {len(multi_case_companies)} (companies with 2+ cases)")
                self.logger.info(f"  Single-case companies: {len(single_case_companies)} (companies with 1 case)")
                self.logger.info(f"  No-company cases: {len(no_company_cases)} (cases without company)")
                
                # STEP 3A: Distribute multi-case companies fairly across SELECTED handlers only
                self.logger.info("=== STEP 3A: Distributing Multi-Case Companies Fairly to SELECTED Handlers ===")
                multi_case_assignments = {}
                
                if multi_case_companies:
                    # Sort companies by number of cases (largest first for better distribution)
                    sorted_companies = sorted(multi_case_companies.items(), key=lambda x: len(x[1]), reverse=True)
                    
                    # Distribute companies across SELECTED handlers in round-robin fashion
                    handler_index = 0
                    for company, case_numbers in sorted_companies:
                        assigned_handler = cleaned_handlers[handler_index % len(cleaned_handlers)]
                        
                        # Assign all cases from this company to the same SELECTED handler
                        for case_num in case_numbers:
                            multi_case_assignments[case_num] = assigned_handler
                        
                        self.logger.info(f"Company '{company}' ({len(case_numbers)} cases) -> {assigned_handler} (SELECTED)")
                        
                        # Move to next SELECTED handler for next company
                        handler_index += 1
                    
                    self.logger.info(f"Distributed {len(multi_case_companies)} multi-case companies across SELECTED handlers")
                else:
                    self.logger.info("No multi-case companies found")
                
                # STEP 3B: Distribute single-case companies and no-company cases to SELECTED handlers
                self.logger.info("=== STEP 3B: Distributing Single Cases to SELECTED Handlers ===")
                single_and_no_company_assignments = {}
                
                # Combine single cases and no-company cases
                all_single_cases = list(single_case_companies.values()) + no_company_cases
                
                if all_single_cases:
                    # Distribute these cases evenly across SELECTED handlers only
                    handler_index = 0
                    for case_num in all_single_cases:
                        assigned_handler = cleaned_handlers[handler_index % len(cleaned_handlers)]
                        single_and_no_company_assignments[case_num] = assigned_handler
                        
                        # Move to next SELECTED handler for next case
                        handler_index += 1
                    
                    self.logger.info(f"Distributed {len(all_single_cases)} single cases and no-company cases across SELECTED handlers")
                else:
                    self.logger.info("No single cases or no-company cases found")
                
                # STEP 3C: Apply all assignments to the dataframe
                self.logger.info("=== STEP 3C: Applying All Assignments ===")
                all_new_assignments = {**multi_case_assignments, **single_and_no_company_assignments}
                
                # VALIDATION: Ensure no unselected handlers get new cases
                unselected_handlers_with_new_cases = []
                for case_num, handler in all_new_assignments.items():
                    if handler not in selected_handlers_set:
                        unselected_handlers_with_new_cases.append((case_num, handler))
                
                if unselected_handlers_with_new_cases:
                    self.logger.error(f"❌ CRITICAL ERROR: Unselected handlers are getting new cases!")
                    for case_num, handler in unselected_handlers_with_new_cases:
                        self.logger.error(f"   Case {case_num} assigned to unselected handler: {handler}")
                    raise ValueError("Unselected handlers are getting new cases - this should not happen!")
                
                for case_num, handler in all_new_assignments.items():
                    case_mask = df['Case Number'] == case_num
                    if case_mask.any():
                        case_idx = case_mask.idxmax()
                        df.at[case_idx, 'Assigned To'] = handler
                        self.logger.debug(f"Assigned NEW case {case_num} -> {handler}")
                
                self.logger.info(f"Applied {len(all_new_assignments)} new case assignments")
                self.logger.info(f"VALIDATION PASSED: All cases assigned only to selected handlers")
                
                # STEP 3D: Distribution Summary
                self.logger.info("=== STEP 3D: Distribution Summary ===")
                handler_counts = {}
                for handler in cleaned_handlers:
                    count = len(df[df['Assigned To'] == handler])
                    handler_counts[handler] = count
                
                self.logger.info(f"Final distribution after NEW case assignment:")
                for handler, count in sorted(handler_counts.items()):
                    self.logger.info(f"  {handler}: {count} total cases")
                
                # Calculate distribution fairness
                if handler_counts:
                    min_cases = min(handler_counts.values())
                    max_cases = max(handler_counts.values())
                    fairness_ratio = min_cases / max_cases if max_cases > 0 else 1.0
                    self.logger.info(f"Distribution fairness ratio: {fairness_ratio:.2f} (1.0 = perfectly fair)")
                    
                    if fairness_ratio < 0.8:
                        self.logger.warning(f"⚠️  Distribution may be uneven (fairness ratio: {fairness_ratio:.2f})")
                    else:
                        self.logger.info(f"Distribution is fair (fairness ratio: {fairness_ratio:.2f})")
            else:
                self.logger.info("No new cases to assign")

            # STEP 5: Final deduplication check
            self.logger.info("=== STEP 5: Final Deduplication Check ===")
            pre_dedup_count = len(df)
            # Ensure integer dtype for deduplication
            df['Case Number'] = pd.Series(pd.to_numeric(df['Case Number'], errors='coerce')).astype('Int64')
            df = df.dropna(subset=['Case Number']).drop_duplicates(subset=['Case Number'], keep='last')
            post_dedup_count = len(df)
            duplicates_removed = pre_dedup_count - post_dedup_count
            
            if duplicates_removed > 0:
                self.logger.warning(f"⚠️  Removed {duplicates_removed} duplicate case entries during assignment")
                self.logger.warning(f"   This prevents cases from being assigned to multiple handlers")
            else:
                self.logger.info("No duplicates found in assigned cases")
            
            # STEP 6: Final workload summary
            self.logger.info("=== STEP 6: Final Workload Summary ===")
            
            # Get ALL handlers (including preserved ones)
            all_handlers = df['Assigned To'].dropna().unique()
            all_handlers = [h for h in all_handlers if h and str(h).strip()]
            
            final_workload = {}
            for handler in all_handlers:
                count = len(df[df['Assigned To'] == handler])
                final_workload[handler] = count
            
            self.logger.info(f"Final workload distribution (ALL handlers): {final_workload}")
            
            # Show separation between selected and preserved handlers
            selected_workload = {h: final_workload.get(h, 0) for h in cleaned_handlers}
            preserved_workload = {h: final_workload.get(h, 0) for h in all_handlers if h not in cleaned_handlers}
            
            self.logger.info(f"SELECTED handlers workload: {selected_workload}")
            if preserved_workload:
                self.logger.info(f"PRESERVED handlers workload (no new cases): {preserved_workload}")
            
            # Calculate distribution fairness for selected handlers only
            if selected_workload and len(selected_workload) > 1:
                min_cases = min(selected_workload.values())
                max_cases = max(selected_workload.values())
                fairness_ratio = min_cases / max_cases if max_cases > 0 else 1.0
                self.logger.info(f"Distribution fairness ratio (selected handlers only): {fairness_ratio:.2f} (1.0 = perfectly fair)")
                
                if fairness_ratio < 0.8:
                    self.logger.warning(f"⚠️  Distribution may be uneven (fairness ratio: {fairness_ratio:.2f})")
                else:
                    self.logger.info(f"Distribution is fair (fairness ratio: {fairness_ratio:.2f})")
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error in assign_handlers: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return df

    def calculate_fair_share(self, df, selected_handlers, chat_agent_info=None):
        """Calculate fair share of cases for equitable distribution
        
        CRITICAL: Only counts IN_PROGRESS cases for fair share calculation
        Chat Agent gets: fair_share × 1.15 from in_progress cases only
        
        Args:
            df: DataFrame with all cases
            selected_handlers: List of selected handler names
            chat_agent_info: Dict with Chat Agent info or None
            
        Returns:
            dict: {
                'total_cases': int (in_progress only),
                'total_handlers': int,
                'fair_share': float,
                'chat_agent_capacity': float (if enabled),
                'handler_list': list of all handlers including Chat Agent
            }
        """
        try:
            # CRITICAL: Count ONLY in_progress cases
            # Find all status columns and identify in_progress cases
            status_cols = [col for col in df.columns if 'status' in col.lower()]
            
            # Build a mask for in_progress cases
            in_progress_mask = pd.Series([False] * len(df), index=df.index)
            
            if status_cols:
                for status_col in status_cols:
                    in_progress_mask = in_progress_mask | (df[status_col].astype(str).str.lower().str.contains(r'in[_\s.]?progress', regex=True, na=False))
            
            # Count only in_progress cases
            total_cases = in_progress_mask.sum()
            
            self.logger.info(f"Fair Share Calculation (IN_PROGRESS CASES ONLY):")
            self.logger.info(f"  Total cases in dataframe: {len(df)}")
            self.logger.info(f"  Total IN_PROGRESS cases: {total_cases}")
            self.logger.info(f"  Status columns used: {status_cols}")
            
            # Calculate total handlers
            total_handlers = len(selected_handlers)
            if chat_agent_info and chat_agent_info.get('enabled'):
                total_handlers += 1  # Add Chat Agent as a handler
            
            # Calculate fair share based on in_progress cases only
            if total_handlers > 0:
                fair_share = total_cases / total_handlers
            else:
                fair_share = 0
            
            result = {
                'total_cases': total_cases,
                'total_handlers': total_handlers,
                'fair_share': fair_share,
                'handler_list': selected_handlers.copy()
            }
            
            # Add Chat Agent info if enabled
            if chat_agent_info and chat_agent_info.get('enabled'):
                import math
                # Calculate Chat Agent capacity: 15% more than fair share
                chat_agent_capacity_raw = fair_share * 1.15
                chat_agent_capacity = math.ceil(chat_agent_capacity_raw)  # Round UP to nearest integer
                
                result['chat_agent_capacity'] = chat_agent_capacity
                result['chat_agent_capacity_raw'] = chat_agent_capacity_raw
                result['chat_agent_name'] = chat_agent_info.get('supporter_name', 'Chat Agent')
                result['handler_list'].append('Chat Agent')
                
                self.logger.info(f"Chat Agent Capacity Rule (15% bonus):")
                self.logger.info(f"  Fair share (regular handlers): {fair_share:.2f}")
                self.logger.info(f"  Chat Agent capacity (raw): {chat_agent_capacity_raw:.2f}")
                self.logger.info(f"  Chat Agent capacity (rounded up): {chat_agent_capacity}")
            
            self.logger.info(f"Fair Share Calculation:")
            self.logger.info(f"  Total cases: {total_cases}")
            self.logger.info(f"  Total handlers: {total_handlers}")
            self.logger.info(f"  Fair share per handler: {fair_share:.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error in calculate_fair_share: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return {
                'total_cases': 0,
                'total_handlers': len(selected_handlers),
                'fair_share': 0,
                'handler_list': selected_handlers.copy()
            }

    def calculate_chat_agent_cases_needed(self, df, chat_agent_supporter_name='Chat Agent'):
        """Calculate how many cases Chat Agent needs to reach capacity
        
        Only counts in_progress cases.
        
        Args:
            df: DataFrame with current case assignments
            chat_agent_supporter_name: Name of Chat Agent supporter
            
        Returns:
            dict: {
                'can_receive': bool,
                'current_queue': int,
                'capacity': int,
                'cases_needed': int
            }
        """
        try:
            if not self.fair_share_info or 'chat_agent_capacity' not in self.fair_share_info:
                return {
                    'can_receive': False,
                    'current_queue': 0,
                    'capacity': 0,
                    'cases_needed': 0,
                    'reason': 'Chat Agent not enabled or capacity not calculated'
                }
            
            # Count current Chat Agent queue - ONLY in_progress cases
            current_queue = 0
            if 'Assigned To' in df.columns:
                chat_cases = df[df['Assigned To'] == chat_agent_supporter_name]
                # Filter for in_progress status (check Status or Case Status columns)
                if len(chat_cases) > 0:
                    # Look for in_progress status - OR all status columns
                    status_cols = [col for col in chat_cases.columns if 'status' in col.lower()]
                    in_progress_mask = pd.Series([False] * len(chat_cases), index=chat_cases.index)
                    for status_col in status_cols:
                        if status_col in chat_cases.columns:
                            in_progress_mask = in_progress_mask | (chat_cases[status_col].astype(str).str.lower().str.contains(r'in[_\s.]?progress', regex=True, na=False))
                    in_progress_cases = chat_cases[in_progress_mask]
                    current_queue = len(in_progress_cases)
            
            capacity = self.fair_share_info['chat_agent_capacity']
            cases_needed = capacity - current_queue
            can_receive = cases_needed > 0
            
            self.logger.info(f"Chat Agent Capacity Check:")
            self.logger.info(f"  Current in_progress queue: {current_queue}")
            self.logger.info(f"  Capacity (fair share × 1.15): {capacity}")
            self.logger.info(f"  Cases needed: {cases_needed}")
            self.logger.info(f"  Can receive more: {can_receive}")
            
            return {
                'can_receive': can_receive,
                'current_queue': current_queue,
                'capacity': capacity,
                'cases_needed': cases_needed if cases_needed > 0 else 0
            }
            
        except Exception as e:
            self.logger.error(f"Error in calculate_chat_agent_cases_needed: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return {
                'can_receive': False,
                'current_queue': 0,
                'capacity': 0,
                'cases_needed': 0,
                'error': str(e)
            }

    def redistribute_cases_to_chat_agent(self, df, chat_agent_info, chat_agent_supporter_name='Chat Agent'):
        """Redistribute IN_PROGRESS cases from eligible handlers to Chat Agent
        
        Redistribution rules:
        - Chat Agent capacity = Fair Share × 1.15 (already calculated)
        - Preserve previous Chat Agent cases
        - Pull ONLY in_progress cases from bottom of eligible handlers' queues (oldest first - FIFO)
        - Only from handlers whose queue is ABOVE fair share
        - Pull from BOTTOM of handler queue (oldest cases) - copy entire row with all data
        - Only update "Assigned To" to Chat Agent name
        
        Args:
            df: DataFrame with case assignments
            chat_agent_info: Chat Agent info dict
            chat_agent_supporter_name: Chat Agent supporter name (goes in Assigned To column)
            
        Returns:
            DataFrame with redistributed cases (ALL data preserved, only Assigned To changed)
        """
        try:
            if not chat_agent_info or not chat_agent_info.get('enabled'):
                self.logger.info("Chat Agent not enabled - skipping redistribution")
                return df
            
            if not self.fair_share_info:
                self.logger.warning("Fair share info not available - skipping redistribution")
                return df
            
            # Get target capacity for Chat Agent (115% of fair share)
            target_capacity = self.fair_share_info.get('chat_agent_capacity', 0)
            fair_share = self.fair_share_info.get('fair_share', 0)
            selected_handlers = [h for h in self.fair_share_info.get('handler_list', []) if h != 'Chat Agent']
            
            self.logger.info(f"\n=== CHAT AGENT REDISTRIBUTION (115% CAPACITY MODEL) ===")
            self.logger.info(f"Chat Agent: {chat_agent_supporter_name}")
            self.logger.info(f"Fair share per handler: {fair_share:.2f}")
            self.logger.info(f"Chat Agent target capacity (115% of fair share): {target_capacity}")
            
            # STEP 1: Identify all status columns and determine what 'in_progress' looks like
            status_cols = [col for col in df.columns if 'status' in col.lower()]
            self.logger.info(f"\nStatus columns identified: {status_cols}")
            
            if not status_cols or 'Assigned To' not in df.columns:
                self.logger.warning("Missing status columns or 'Assigned To' column - cannot redistribute")
                return df
            
            # Show sample status values for debugging
            all_status_values = set()
            for col in status_cols:
                unique_vals = df[col].dropna().unique()[:10]
                all_status_values.update([str(v).lower() for v in unique_vals])
            self.logger.info(f"Sample status values found: {all_status_values}")
            
            # STEP 2: Count current Chat Agent in_progress cases
            current_chat_in_progress = 0
            chat_agent_cases = df[df['Assigned To'] == chat_agent_supporter_name]
            
            if not chat_agent_cases.empty:
                # Build in_progress mask for Chat Agent's current cases
                in_progress_mask_chat = pd.Series([False] * len(chat_agent_cases), index=chat_agent_cases.index)
                for status_col in status_cols:
                    if status_col in chat_agent_cases.columns:
                        in_progress_mask_chat = in_progress_mask_chat | (
                            chat_agent_cases[status_col].astype(str).str.lower().str.contains(r'in[_\s.]?progress', regex=True, na=False)
                        )
                current_chat_in_progress = in_progress_mask_chat.sum()
            
            cases_needed = target_capacity - current_chat_in_progress
            self.logger.info(f"\nCurrent Chat Agent in_progress cases: {current_chat_in_progress}")
            self.logger.info(f"Cases needed to reach target capacity: {max(0, cases_needed)}")
            
            if cases_needed <= 0:
                self.logger.info("Chat Agent at or above capacity - no redistribution needed")
                return df
            
            # STEP 3: Identify eligible handlers (in_progress queue > fair share)
            self.logger.info(f"\n=== IDENTIFYING ELIGIBLE HANDLERS (in_progress queue > {fair_share:.2f}) ===")
            eligible_handlers_info = {}  # handler -> {queue_size, indices}
            
            for handler in selected_handlers:
                handler_all = df[df['Assigned To'] == handler]
                if handler_all.empty:
                    continue
                
                # Build in_progress mask for this handler
                in_progress_mask_handler = pd.Series([False] * len(handler_all), index=handler_all.index)
                for status_col in status_cols:
                    if status_col in handler_all.columns:
                        in_progress_mask_handler = in_progress_mask_handler | (
                            handler_all[status_col].astype(str).str.lower().str.contains(r'in[_\s.]?progress', regex=True, na=False)
                        )
                
                handler_in_progress_indices = handler_all[in_progress_mask_handler].index.tolist()
                queue_size = len(handler_in_progress_indices)
                
                if queue_size > fair_share:
                    eligible_status = f" → ELIGIBLE (excess: {queue_size - fair_share:.0f})"
                    eligible_handlers_info[handler] = {
                        'queue_size': queue_size,
                        'in_progress_indices': handler_in_progress_indices,
                        'excess': queue_size - fair_share
                    }
                else:
                    eligible_status = " → not eligible"
                
                self.logger.info(f"  {handler}: {queue_size} in_progress cases (fair share: {fair_share:.2f}){eligible_status}")
            
            if not eligible_handlers_info:
                self.logger.info("No eligible handlers found - all have in_progress queue at or below fair share")
                return df
            
            # STEP 4: Sort eligible handlers by excess workload (descending)
            sorted_handlers = sorted(eligible_handlers_info.items(), 
                                    key=lambda x: x[1]['excess'], reverse=True)
            
            self.logger.info(f"\n=== PULLING CASES TO REACH CHAT AGENT CAPACITY ({target_capacity} cases) ===")
            
            # STEP 5: Pull in_progress cases from eligible handlers (from bottom = oldest first)
            total_pulled = 0
            
            for handler, info in sorted_handlers:
                if cases_needed <= 0:
                    break
                
                # Get in_progress indices for this handler (already identified above)
                in_progress_indices = info['in_progress_indices']
                queue_size = info['queue_size']
                
                # Pull from bottom (start of list = oldest entries)
                cases_to_pull = min(cases_needed, len(in_progress_indices))
                indices_to_reassign = in_progress_indices[:cases_to_pull]
                
                # Get case numbers for logging
                case_num_col = None
                for col in df.columns:
                    if 'case number' in col.lower():
                        case_num_col = col
                        break
                
                case_numbers_pulled = []
                if case_num_col:
                    case_numbers_pulled = df.loc[indices_to_reassign, case_num_col].tolist()
                
                # Reassign these cases to Chat Agent
                df.loc[indices_to_reassign, 'Assigned To'] = chat_agent_supporter_name
                
                cases_needed -= cases_to_pull
                total_pulled += cases_to_pull
                
                remaining_after = queue_size - cases_to_pull  
                self.logger.info(f"  {handler}: Pulled {cases_to_pull} cases (queue: {queue_size} → {remaining_after}, still need: {max(0, cases_needed)})")
                if case_numbers_pulled:
                    self.logger.info(f"    Case #s: {case_numbers_pulled[:5]}{'...' if len(case_numbers_pulled) > 5 else ''}")
            
            self.logger.info(f"\n✓ REDISTRIBUTION COMPLETE")
            self.logger.info(f"  Total in_progress cases pulled to Chat Agent: {total_pulled}")
            self.logger.info(f"  Chat Agent will now have: {current_chat_in_progress + total_pulled} cases (target: {target_capacity})")
            
            # Verify results
            final_chat_cases = df[df['Assigned To'] == chat_agent_supporter_name]
            final_chat_in_progress = 0
            if not final_chat_cases.empty:
                in_progress_mask_final = pd.Series([False] * len(final_chat_cases), index=final_chat_cases.index)
                for status_col in status_cols:
                    if status_col in final_chat_cases.columns:
                        in_progress_mask_final = in_progress_mask_final | (
                            final_chat_cases[status_col].astype(str).str.lower().str.contains('in.?progress|inprogress', regex=True, na=False)
                        )
                final_chat_in_progress = in_progress_mask_final.sum()
            
            self.logger.info(f"  Verification: Chat Agent now has {final_chat_in_progress} in_progress cases (capacity: {target_capacity})")
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error in redistribute_cases_to_chat_agent: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return df

    def verify_chat_agent_persistence(self, df, chat_agent_info, prev_file=None):
        """Verify that Chat Agent cases are properly preserved from previous day
        
        Sheet name: Always "Chat Agent's Cases" (not using supporter name)
        Only counts in_progress cases
        
        Args:
            df: Current dataframe with assignments
            chat_agent_info: Chat Agent info dict
            prev_file: Path to previous file
            
        Returns:
            dict with persistence verification info
        """
        try:
            if not chat_agent_info or not chat_agent_info.get('enabled'):
                return {'enabled': False}
            
            chat_agent_supporter = chat_agent_info.get('supporter_name', 'Chat Agent')
            
            # Count in_progress Chat Agent cases in current dataframe
            current_chat_agent_cases = 0
            if 'Assigned To' in df.columns:
                chat_cases = df[df['Assigned To'] == chat_agent_supporter]
                # Filter for in_progress status only
                status_cols = [col for col in df.columns if 'status' in col.lower()]
                in_progress_cases = chat_cases
                for status_col in status_cols:
                    in_progress_cases = in_progress_cases[in_progress_cases[status_col].astype(str).str.lower().str.contains('in_progress', na=False)]
                current_chat_agent_cases = len(in_progress_cases)
            
            # Check if Chat Agent sheet exists in previous file
            prev_chat_agent_cases = 0
            chat_agent_sheet_name = "Chat Agent's Cases"  # Fixed sheet name
            
            if prev_file and os.path.exists(prev_file):
                try:
                    excel_file = pd.ExcelFile(prev_file)
                    if chat_agent_sheet_name in excel_file.sheet_names:
                        prev_df = pd.read_excel(prev_file, sheet_name=chat_agent_sheet_name)
                        # Only count in_progress cases from previous file
                        status_cols = [col for col in prev_df.columns if 'status' in col.lower()]
                        in_progress_prev = prev_df
                        for status_col in status_cols:
                            in_progress_prev = in_progress_prev[in_progress_prev[status_col].astype(str).str.lower().str.contains('in_progress', na=False)]
                        prev_chat_agent_cases = len(in_progress_prev)
                except Exception:
                    pass
            
            self.logger.info(f"\n=== Chat Agent Persistence Verification ===")
            self.logger.info(f"Chat Agent supporter name: {chat_agent_supporter}")
            self.logger.info(f"Sheet name: {chat_agent_sheet_name}")
            self.logger.info(f"Current in_progress cases assigned: {current_chat_agent_cases}")
            self.logger.info(f"Previous day in_progress cases: {prev_chat_agent_cases}")

            return {
                'enabled': True,
                'supporter_name': chat_agent_supporter,
                'current_cases': current_chat_agent_cases,
                'previous_cases': prev_chat_agent_cases,
                'sheet_name': chat_agent_sheet_name,
                'persistence_confirmed': current_chat_agent_cases >= prev_chat_agent_cases
            }
            
        except Exception as e:
            self.logger.error(f"Error verifying Chat Agent persistence: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return {'enabled': True, 'error': str(e)}

    def create_chat_agent_sheet_output(self, writer, output_df, chat_agent_info, prev_file=None):
        """Create and write the 'Chat Agent's Cases' sheet to the output Excel file
        
        Preservation Rules:
        - Load ALL cases from previous "Chat Agent's Cases" sheet (all statuses)
        - Load ALL current cases assigned to Chat Agent supporter
        - Merge: current cases first, then add any previous cases not in current
        - Sort by status
        
        Args:
            writer: pd.ExcelWriter object for writing to Excel
            output_df: Current processed DataFrame with assignments
            chat_agent_info: Chat Agent info dict with enabled status and supporter name
            prev_file: Path to previous file to load previous Chat Agent cases
        """
        try:
            if not chat_agent_info or not chat_agent_info.get('enabled'):
                self.logger.info("Chat Agent not enabled - skipping Chat Agent's Cases sheet")
                return
            
            chat_agent_supporter_name = chat_agent_info.get('supporter_name', 'Chat Agent')
            self.logger.info(f"\n=== CREATING CHAT AGENT'S CASES SHEET ===")
            self.logger.info(f"Chat Agent supporter: {chat_agent_supporter_name}")
            
            # STEP 1: Get current Chat Agent cases from output_df (ALL cases assigned to this supporter)
            current_chat_cases = pd.DataFrame()
            if 'Assigned To' in output_df.columns:
                current_chat_cases = output_df[output_df['Assigned To'] == chat_agent_supporter_name].copy()
                self.logger.info(f"  Current cases assigned to {chat_agent_supporter_name}: {len(current_chat_cases)}")
                if not current_chat_cases.empty:
                    status_cols = [col for col in current_chat_cases.columns if 'status' in col.lower()]
                    if status_cols:
                        status_summary = {}
                        for status_col in status_cols:
                            statuses = current_chat_cases[status_col].value_counts()
                            status_summary[status_col] = statuses.to_dict()
                        self.logger.info(f"  Status breakdown: {status_summary}")
            
            # STEP 2: Load previous Chat Agent cases from previous file (ALL cases, all statuses)
            prev_chat_cases = pd.DataFrame()
            chat_agent_sheet_name = "Chat Agent's Cases"  # Fixed sheet name
            
            if prev_file and os.path.exists(prev_file):
                try:
                    excel_file = pd.ExcelFile(prev_file)
                    if chat_agent_sheet_name in excel_file.sheet_names:
                        prev_chat_cases = pd.read_excel(prev_file, sheet_name=chat_agent_sheet_name)
                        self.logger.info(f"  ✓ Previous Chat Agent cases loaded from '{chat_agent_sheet_name}': {len(prev_chat_cases)}")
                        if not prev_chat_cases.empty:
                            status_cols = [col for col in prev_chat_cases.columns if 'status' in col.lower()]
                            if status_cols:
                                status_summary = {}
                                for status_col in status_cols:
                                    statuses = prev_chat_cases[status_col].value_counts()
                                    status_summary[status_col] = statuses.to_dict()
                                self.logger.info(f"  Previous status breakdown: {status_summary}")
                    else:
                        self.logger.info(f"  No '{chat_agent_sheet_name}' sheet found in previous file")
                except Exception as e:
                    self.logger.warning(f"Could not load previous Chat Agent cases: {str(e)}")
            
            # STEP 3: Normalize case numbers for merging
            if 'Case Number' in current_chat_cases.columns:
                current_chat_cases['Case Number'] = current_chat_cases['Case Number'].apply(
                    lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip().replace('.','',1).isdigit() else pd.NA
                ).astype('Int64')
            
            if 'Case Number' in prev_chat_cases.columns:
                prev_chat_cases['Case Number'] = prev_chat_cases['Case Number'].apply(
                    lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip().replace('.','',1).isdigit() else pd.NA
                ).astype('Int64')
            
            # STEP 4: Merge - keep current, add previous not in current
            current_case_nums = set(current_chat_cases['Case Number'].dropna().tolist()) if not current_chat_cases.empty else set()
            
            if not prev_chat_cases.empty:
                # Only keep previous cases not in current (to avoid duplicates)
                previously_preserved = prev_chat_cases[~prev_chat_cases['Case Number'].isin(current_case_nums)].copy()
                self.logger.info(f"  Preserved from previous days: {len(previously_preserved)} cases")
            else:
                previously_preserved = pd.DataFrame()
            
            # Combine: current cases first, then preserved cases
            if not current_chat_cases.empty or not previously_preserved.empty:
                chat_agent_cases = pd.concat([current_chat_cases, previously_preserved], ignore_index=True)
            else:
                chat_agent_cases = pd.DataFrame()
            
            # STEP 5: Sort by status
            if not chat_agent_cases.empty:
                total_cases = len(chat_agent_cases)
                self.logger.info(f"  ✓ Total cases for Chat Agent sheet:")
                self.logger.info(f"    - Current: {len(current_chat_cases)}")
                self.logger.info(f"    - Preserved from previous: {len(previously_preserved)}")
                self.logger.info(f"    - TOTAL: {total_cases}")
                
                chat_agent_cases = self.sort_cases_by_status(chat_agent_cases)
                
                # STEP 6: Write to Excel
                chat_agent_cases.to_excel(writer, sheet_name=chat_agent_sheet_name, index=False)
                self.auto_adjust_columns(writer, chat_agent_cases, chat_agent_sheet_name)
                # Lock the sheet
                self.protect_worksheet(writer, chat_agent_sheet_name, password='artadmin')
                self.logger.info(f"✓ Created '{chat_agent_sheet_name}' sheet with {total_cases} total cases")
            else:
                # CRITICAL FIX: Create empty sheet with proper headers even if no cases
                self.logger.info(f"No cases for Chat Agent - creating empty '{chat_agent_sheet_name}' sheet with headers")
                
                # Create empty DataFrame with output columns (same structure as other handler sheets)
                empty_sheet_df = pd.DataFrame(columns=self.output_columns)
                
                # Write empty sheet with headers
                empty_sheet_df.to_excel(writer, sheet_name=chat_agent_sheet_name, index=False)
                self.auto_adjust_columns(writer, empty_sheet_df, chat_agent_sheet_name)
                # Lock the sheet
                self.protect_worksheet(writer, chat_agent_sheet_name, password='artadmin')
                self.logger.info(f"✓ Created empty '{chat_agent_sheet_name}' sheet with headers (rows: 0, columns: {len(self.output_columns)})")

                
        except Exception as e:
            self.logger.error(f"Error creating Chat Agent's Cases sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")


    def _load_handler_sheet_assignments_v2(self, prev_file, selected_handlers_set):
        """SEPARATED METHOD: Load ALL handler sheet assignments for preservation
        - Preserves ALL handler work regardless of selection status
        - Selection status only affects NEW case assignment, not preservation
        - This ensures Ibrahim keeps his existing cases even when not selected"""
        try:
            self.logger.info(f"=== LOADING ALL HANDLER SHEET ASSIGNMENTS FOR PRESERVATION ===")
            self.logger.info(f"File: {prev_file}")
            self.logger.info(f"Selected handlers for NEW assignments: {selected_handlers_set}")
            
            if not os.path.exists(prev_file):
                self.logger.error(f"File does not exist: {prev_file}")
                return {}
            
            # Read the Excel file
            excel_file = pd.ExcelFile(prev_file)
            self.logger.info(f"Excel sheets found: {excel_file.sheet_names}")
            
            handler_sheet_assignments = {}
            total_cases_found = 0
            total_assignments_loaded = 0
            
            # Look for sheets ending with "'s Cases"
            handler_sheets = [sheet for sheet in excel_file.sheet_names if isinstance(sheet, str) and sheet.endswith("'s Cases")]
            self.logger.info(f"Handler sheets found: {handler_sheets}")
            
            for sheet_name in handler_sheets:
                try:
                    self.logger.info(f"\n--- Processing sheet: {sheet_name} ---")
                    
                    # Extract handler name from sheet name
                    handler_name = sheet_name.replace("'s Cases", "").strip()
                    self.logger.info(f"Extracted handler name: '{handler_name}'")
                    
                    # CRITICAL: Always process ALL handler sheets to preserve work
                    # Selection status only affects NEW case assignment, not preservation
                    cleaned_handler = self.clean_handler_name(handler_name)
                    if cleaned_handler not in selected_handlers_set:
                        self.logger.info(f"Handler '{cleaned_handler}' not selected for NEW cases, but PRESERVING their existing work")
                    else:
                        self.logger.info(f"Handler '{cleaned_handler}' is selected - preserving existing work AND allowing new assignments")
                    
                    # Read the sheet
                    sheet_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                    self.logger.info(f"Sheet columns: {list(sheet_df.columns)}")
                    self.logger.info(f"Sheet shape: {sheet_df.shape}")
                    
                    # Look for Case Number column (case insensitive)
                    case_col = None
                    for col in sheet_df.columns:
                        if 'case' in str(col).lower() and 'number' in str(col).lower():
                            case_col = col
                            break
                    
                    if case_col is None:
                        self.logger.warning(f"No Case Number column found in {sheet_name}")
                        continue
                    
                    self.logger.info(f"Using Case Number column: '{case_col}'")
                    
                    # Look for _Source_Sheet column (case insensitive) - CRITICAL FOR PRESERVATION
                    source_col = None
                    for col in sheet_df.columns:
                        if '_source_sheet' in str(col).lower() or '_Source_Sheet' in str(col):
                            source_col = col
                            break
                    
                    if source_col is None:
                        self.logger.warning(f"No _Source_Sheet column found in {sheet_name} - CRITICAL FOR PRESERVATION")
                        
                        # CRITICAL FIX: Only use fallback for selected handlers to prevent new case assignment
                        if cleaned_handler not in selected_handlers_set:
                            self.logger.warning(f"Handler '{cleaned_handler}' not selected - SKIPPING fallback assignment")
                            self.logger.warning(f"  This prevents unselected handlers from getting new cases through fallback logic")
                            self.logger.warning(f"  Ibrahim's existing cases will be preserved through other means")
                            continue
                        
                        # FALLBACK: Only for selected handlers - assume all cases in this sheet belong to this handler
                        self.logger.info(f"Using fallback for SELECTED handler: All cases in '{sheet_name}' will be assigned to '{cleaned_handler}'")
                        
                        # Process each row in the sheet using fallback logic
                        sheet_cases = 0
                        sheet_assignments = 0
                        
                        for idx, row in sheet_df.iterrows():
                            try:
                                # Get case number (case insensitive, handle NaN)
                                case_number_raw = row.get(case_col)
                                if pd.isna(case_number_raw):
                                    continue
                                
                                case_number = self.normalize_case_number(case_number_raw)
                                if not case_number:
                                    continue
                                
                                sheet_cases += 1
                                
                                # FALLBACK LOGIC: Case is in this sheet, so it belongs to this handler (selected handlers only)
                                # Store normalized case number to prevent .0 suffix issues
                                handler_sheet_assignments[case_number] = cleaned_handler
                                sheet_assignments += 1
                                total_assignments_loaded += 1
                                
                                self.logger.debug(f"  ✅ Case {case_number} -> {cleaned_handler} (fallback: sheet location)")
                                
                            except Exception as row_error:
                                self.logger.warning(f"Error processing row {idx} in {sheet_name}: {str(row_error)}")
                                continue
                        
                        self.logger.info(f"Sheet {sheet_name}: {sheet_cases} total cases, {sheet_assignments} assigned to {cleaned_handler} (fallback method)")
                        total_cases_found += sheet_cases
                        continue
                    
                    self.logger.info(f"Using _Source_Sheet column: '{source_col}'")
                    
                    # Process each row in the sheet
                    sheet_cases = 0
                    sheet_assignments = 0
                    sheet_skipped = 0
                    
                    for idx, row in sheet_df.iterrows():
                        try:
                            # Get case number (case insensitive, handle NaN)
                            case_number_raw = row.get(case_col)
                            if pd.isna(case_number_raw):
                                continue
                            
                            case_number = self.normalize_case_number(case_number_raw)
                            if not case_number:
                                continue
                            
                            # Get source sheet (case insensitive, handle NaN)
                            source_sheet_raw = row.get(source_col)
                            if pd.isna(source_sheet_raw):
                                continue
                            
                            source_sheet = str(source_sheet_raw).strip()
                            if not source_sheet:
                                continue
                            
                            sheet_cases += 1
                            
                            # CRITICAL LOGIC: Use _Source_Sheet to determine ownership
                            # If _Source_Sheet matches this sheet name, the case belongs to this handler
                            if source_sheet == sheet_name:
                                # This case belongs to this handler based on _Source_Sheet
                                handler_sheet_assignments[case_number] = cleaned_handler
                                sheet_assignments += 1
                                total_assignments_loaded += 1
                                
                                self.logger.debug(f"  ✅ Case {case_number} -> {cleaned_handler} (Source: {source_sheet})")
                            else:
                                # Case is in this sheet but _Source_Sheet says it belongs elsewhere
                                # This is important for debugging - cases shouldn't be in wrong sheets
                                sheet_skipped += 1
                                self.logger.warning(f"  ⚠️  Case {case_number} in sheet '{sheet_name}' but _Source_Sheet says '{source_sheet}'")
                                self.logger.warning(f"     This case will NOT be assigned to {cleaned_handler} - it belongs to {source_sheet}")
                                
                        except Exception as row_error:
                            self.logger.warning(f"Error processing row {idx} in {sheet_name}: {str(row_error)}")
                            continue
                    
                    self.logger.info(f"Sheet {sheet_name}: {sheet_cases} total cases, {sheet_assignments} assigned to {cleaned_handler}, {sheet_skipped} skipped")
                    total_cases_found += sheet_cases
                    
                except Exception as sheet_error:
                    self.logger.error(f"Error processing sheet {sheet_name}: {str(sheet_error)}")
                    continue
            
            self.logger.info(f"\n=== HANDLER SHEET ASSIGNMENTS V2 SUMMARY (IMPROVED) ===")
            self.logger.info(f"Total cases found across all sheets: {total_cases_found}")
            self.logger.info(f"Total assignments loaded: {total_assignments_loaded}")
            self.logger.info(f"Assignment success rate: {(total_assignments_loaded/total_cases_found*100):.1f}%" if total_cases_found > 0 else "N/A")
            
            # Log detailed assignment breakdown
            if handler_sheet_assignments:
                self.logger.info(f"\nDetailed assignments loaded:")
                for case_num, handler in sorted(handler_sheet_assignments.items())[:20]:  # Show first 20
                    self.logger.info(f"  {case_num} -> {handler}")
                if len(handler_sheet_assignments) > 20:
                    self.logger.info(f"  ... and {len(handler_sheet_assignments) - 20} more cases")
                
                # Log handler distribution
                handler_counts = {}
                for handler in handler_sheet_assignments.values():
                    handler_counts[handler] = handler_counts.get(handler, 0) + 1
                
                self.logger.info(f"\nHandler distribution from sheets:")
                for handler, count in sorted(handler_counts.items()):
                    self.logger.info(f"  {handler}: {count} cases")
            else:
                self.logger.warning("No assignments loaded from handler sheets!")
            
            return handler_sheet_assignments
            
        except Exception as e:
            self.logger.error(f"Error in _load_handler_sheet_assignments_v2: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return {}

    def _case_belongs_to_handler(self, row, handler_name, sheet_name):
        """Helper method to determine if a case belongs to a specific handler based on various indicators"""
        try:
            # Check if there are any other columns that might indicate handler assignment
            # This is a fallback method for cases where _Source_Sheet and _handler_name are not definitive
            
            # Look for any column containing the handler name
            for col_name, value in row.items():
                if pd.notna(value) and str(value).strip():
                    value_str = str(value).strip().lower()
                    handler_lower = handler_name.lower()
                    
                    # Check if the value contains the handler name
                    if handler_lower in value_str:
                        return True
                    
                    # Check if the value contains the sheet name (without "'s Cases")
                    sheet_base = sheet_name.replace("'s Cases", "").lower()
                    if sheet_base in value_str:
                        return True
            
            # If no clear indicators found, be conservative and return False
            return False
            
        except Exception as e:
            self.logger.debug(f"Error in _case_belongs_to_handler: {str(e)}")
            return False

    def _calculate_current_workload(self, df, handlers):
        """Calculate current workload (case count) per handler"""
        workload = {handler: 0 for handler in handlers}
        
        for handler in handlers:
            workload[handler] = len(df[df['Assigned To'] == handler])
        
        return workload

    def _get_company_column(self, df):
        """Helper method to find the company column"""
        for col in ['Company Name', 'Customer Name']:
            if col in df.columns:
                return col
        return None

    def clean_empty_rows(self, df):
        """Remove only completely empty rows while preserving all data"""
        try:
            before_count = len(df)
            
            # Check for completely empty rows (all columns empty or contain only whitespace/NaN/NaT)
            empty_mask = df.apply(lambda row: all(
                pd.isna(val) or str(val).strip() == '' or str(val).strip().lower() in ['nan', 'nat', 'none', 'null']
                for val in row
            ), axis=1)
            
            # Remove only completely empty rows
            df_cleaned = df[~empty_mask].copy()
            after_count = len(df_cleaned)
            
            removed_count = before_count - after_count
            if removed_count > 0:
                self.logger.info(f"Removed {removed_count} completely empty rows")
            else:
                self.logger.info("No completely empty rows found")
            
            # Final cleanup: replace any remaining NaN/NaT/None with empty strings
            df_cleaned = df_cleaned.fillna('').replace(['NaT', 'nan', 'None', 'null'], '')
            
            return df_cleaned
            
        except Exception as e:
            self.logger.error(f"Error cleaning empty rows: {str(e)}")
            return df

    def _ensure_fields_for_new_cases(self, df, new_cases):
        """Helper function to ensure required fields are populated for new cases"""
        try:
            new_cases_mask = df['Case Number'].isin(new_cases)
            
            # Define critical fields that must be populated for new cases
            critical_fields = {
                'Email': ['Primary Email (Contact) (Contact)', 'Primary Email', 'Contact Email', 'Email'],
                'Country': ['Country/Region (Contact) (Contact)', 'Country/Region', 'Country (Contact)', 'Country'],
                'State/Province': ['State/Province (Case) (Case)', 'State/Province (Case)', 'State/Province (Contact) (Contact)', 'State/Province (Contact)', 'State/Province'],
                'Case Status': ['Case Status (Case) (Case)', 'Case Status (Case)', 'Case Status'],
                'Last Status Change': ['Last Status Change (Workflow Use Only) (Case) (Case)', 'Last Status Change (Case)', 'Last Status Change'],
                'Customer Name': [' Contact Name (Contact) (Contact)', 'Contact Name', 'Contact Name (Contact)', 'Customer Name'],
                'Company Name': ['Company Name', 'Company Name (Contact) (Contact)', 'Company (Contact)'],
                'Status': ['Status']
            }
            
            # For each critical field
            for target_field, source_fields in critical_fields.items():
                # Ensure target field exists
                if target_field not in df.columns:
                    df[target_field] = ''
                
                # Check for empty values including NaN and whitespace
                empty_mask = df[target_field].fillna('').astype(str).str.strip() == ''
                update_mask = new_cases_mask & empty_mask
                
                if update_mask.any():
                    self.logger.info(f"Found {sum(update_mask)} empty {target_field} values in new cases")
                    # Try each source field in order until we find values
                    for source_field in source_fields:
                        if source_field in df.columns:
                            source_values = df.loc[update_mask, source_field].fillna('').astype(str).str.strip()
                            valid_values = source_values != ''
                            if valid_values.any():
                                # Update only where we have valid values
                                update_indices = update_mask[valid_values.index] & valid_values
                                df.loc[update_indices, target_field] = source_values[valid_values]
                                self.logger.info(f"Backfilled {sum(valid_values)} {target_field} values from {source_field}")
                                # Update the mask to exclude rows we just filled
                                update_mask = update_mask & ~valid_values
                                if not update_mask.any():
                                    self.logger.info(f"All {target_field} values have been populated")
                                    break
            
            return df
        except Exception as e:
            self.logger.error(f"Error in _ensure_fields_for_new_cases: {str(e)}")
            return df
    
    def merge_with_previous(self, new_df, prev_df):
        """Merge new data with previous file data while preserving all existing data and adding new cases.
        Additionally, backfill empty cells in previous/overlapping rows using values from the raw (new_df) by Case Number.
        Ensures critical fields (Email, Country, State/Province, etc.) are populated for new cases.
        """
        try:
            self.logger.info("\nMerging with previous file data...")
            
            # IMPORTANT: Convert Completion Date Timestamps to strings early
            if 'Completion Date' in new_df.columns:
                self.logger.info("Converting Completion Date Timestamps in new_df to strings...")
                
                # DIAGNOSTIC: Log BEFORE conversion
                completion_date_dtype = new_df['Completion Date'].dtype
                non_empty_before = (new_df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                self.logger.info(f"DIAGNOSTIC: Before conversion - dtype={completion_date_dtype}, non-empty={non_empty_before}/{len(new_df)}")
                samples_before = new_df['Completion Date'].head(3).tolist()
                self.logger.info(f"DIAGNOSTIC: Sample values before: {samples_before}")
                
                # Do the conversion
                new_df['Completion Date'] = new_df['Completion Date'].astype(str)
                self.logger.info("Completion Date in new_df converted to string format")
                
                # DIAGNOSTIC: Log AFTER conversion
                completion_date_dtype = new_df['Completion Date'].dtype
                non_empty_count = (new_df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                self.logger.info(f"DIAGNOSTIC: After conversion - dtype={completion_date_dtype}, non-empty={non_empty_count}/{len(new_df)}")
                if non_empty_count > 0:
                    samples = new_df[new_df['Completion Date'].fillna('').astype(str).str.strip() != '']['Completion Date'].head(3).tolist()
                    self.logger.info(f"DIAGNOSTIC: Sample values after (non-empty): {samples}")
                else:
                    samples_after = new_df['Completion Date'].head(3).tolist()
                    self.logger.info(f"DIAGNOSTIC: Sample values after (ALL): {samples_after}")
            
            if prev_df is None or prev_df.empty:
                self.logger.info("No previous file data to merge - using only new data")
                return new_df
            
            # Ensure Case Number columns are integer-normalized in both dataframes
            new_df = new_df.copy()
            prev_df = prev_df.copy()
            new_df['Case Number'] = new_df['Case Number'].apply(self.normalize_case_number)
            prev_df['Case Number'] = prev_df['Case Number'].apply(self.normalize_case_number)
            new_df['Case Number'] = pd.Series(pd.to_numeric(new_df['Case Number'], errors='coerce')).astype('Int64')
            prev_df['Case Number'] = pd.Series(pd.to_numeric(prev_df['Case Number'], errors='coerce')).astype('Int64')
            
            # Get case numbers from both dataframes
            new_case_numbers = set(new_df['Case Number'].dropna().tolist())
            prev_case_numbers = set(prev_df['Case Number'].dropna().tolist())
            
            # Identify overlapping and new cases
            overlapping_cases = new_case_numbers & prev_case_numbers
            new_cases = new_case_numbers - prev_case_numbers
            existing_only_cases = prev_case_numbers - new_case_numbers
            
            self.logger.info(f"Merge Summary:")
            self.logger.info(f"  New cases: {len(new_cases)}")
            self.logger.info(f"  Overlapping cases: {len(overlapping_cases)}")
            self.logger.info(f"  Existing only cases: {len(existing_only_cases)}")
            
            # Create a list to store rows in the correct order
            ordered_rows = []
            
            # Check if Completion Date exists in previous file
            completion_date_in_prev = 'Completion Date' in prev_df.columns
            self.logger.info(f"Completion Date in previous file: {completion_date_in_prev}")
            
            # Helper function to find matching column names for special fields
            def get_column_variants(field):
                if field not in self.canonical_fields:
                    return [field]
                return self.canonical_fields[field]
            
            # 1. Add NEW cases first (these are cases that don't exist in previous file)
            new_cases_df = new_df[new_df['Case Number'].isin(new_cases)]
            
            # DIAGNOSTIC: Log Completion Date state in new_cases_df
            if 'Completion Date' in new_cases_df.columns:
                non_empty_count = (new_cases_df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                self.logger.info(f"DIAGNOSTIC: new_cases_df has {non_empty_count}/{len(new_cases_df)} non-empty Completion Date values")
                if non_empty_count == 0:
                    self.logger.error(f"CRITICAL: All Completion Date values were lost when extracting new_cases_df!")
            
            if not new_cases_df.empty:
                # Handle special fields for new cases
                for field, variants in self.canonical_fields.items():
                    # SPECIAL HANDLING FOR COMPLETION DATE (NEW COLUMN ON FIRST RUN)
                    if field == 'Completion Date' and not completion_date_in_prev:
                        # Completion Date not in previous file - only populate new cases
                        self.logger.info("Completion Date column not in previous file - will populate only new cases")
                        target_col = field
                        if target_col not in new_cases_df.columns:
                            new_cases_df[target_col] = ''
                        
                        # Populate from variants (direct source columns)
                        for variant in variants:
                            if variant in new_cases_df.columns:
                                # Get non-empty values from raw file
                                non_empty_mask = (new_cases_df[variant].fillna('').astype(str).str.strip() != '')
                                if non_empty_mask.any():
                                    target_empty = (new_cases_df[target_col].fillna('').astype(str).str.strip() == '')
                                    update_mask = target_empty & non_empty_mask
                                    if update_mask.any():
                                        # Convert Timestamps to strings for Completion Date
                                        new_cases_df.loc[update_mask, target_col] = new_cases_df.loc[update_mask, variant].astype(str)
                                        self.logger.info(f"Completion Date: Populated {update_mask.sum()} new cases from raw file")
                                        sample_vals = new_cases_df.loc[update_mask, target_col].head(3).tolist()
                                        self.logger.info(f"Completion Date sample values: {sample_vals}")
                        continue  # Skip to next field
                    
                    target_col = field
                    
                    # Initialize the target column if it doesn't exist
                    if target_col not in new_cases_df.columns:
                        new_cases_df[target_col] = ''
                    
                    # Try to populate from variants, checking for actual non-empty data
                    for variant in variants:
                        if variant in new_cases_df.columns:
                            # Only copy non-empty values
                            non_empty_mask = (new_cases_df[variant].fillna('').astype(str).str.strip() != '')
                            if non_empty_mask.any():
                                # Copy only where target is empty and source has data
                                target_empty = (new_cases_df[target_col].fillna('').astype(str).str.strip() == '')
                                update_mask = target_empty & non_empty_mask
                                if update_mask.any():
                                    new_cases_df.loc[update_mask, target_col] = new_cases_df.loc[update_mask, variant]
                                    self.logger.debug(f"New cases: Populated {update_mask.sum()} {target_col} values from {variant}")
                
                # Ensure critical fields are populated for new cases
                new_cases_df = self._ensure_fields_for_new_cases(new_cases_df, new_cases)
                ordered_rows.append(new_cases_df)
                self.logger.info(f"Added {len(new_cases_df)} NEW cases at the top")
            
            # 2. Add OVERLAPPING cases (cases that exist in both files)
            # For overlapping cases, start from previous row, then fill any empty cells with new data values
            overlapping_rows = []
            for case_num in overlapping_cases:
                try:
                    new_row = new_df[new_df['Case Number'] == case_num].iloc[0]
                    prev_row = prev_df[prev_df['Case Number'] == case_num].iloc[0]
                    
                    # Start from previous row to preserve agent updates
                    merged_row = prev_row.copy()
                    
                    # Handle special fields that need merging
                    for field, variants in self.canonical_fields.items():
                        # SPECIAL HANDLING FOR COMPLETION DATE
                        if field == 'Completion Date':
                            # Initialize if not in previous row
                            if field not in merged_row.index:
                                merged_row[field] = ''
                            
                            # If Completion Date NOT in previous file, populate from new file for overlapping cases
                            if not completion_date_in_prev:
                                # This is an overlapping case where Completion Date column is new
                                # Populate from new file
                                for variant in variants:
                                    if variant in new_row.index:
                                        new_val = str(new_row[variant]).strip() if pd.notna(new_row[variant]) else ''
                                        if new_val:
                                            merged_row[field] = new_val
                                            self.logger.debug(f"Overlapping case {case_num}: Populated Completion Date from new file")
                                            break
                            else:
                                # Completion Date exists in previous file - preserve it
                                prev_val = str(merged_row[field]).strip() if pd.notna(merged_row[field]) else ''
                                # Don't overwrite - keep previous
                                merged_row[field] = prev_val
                            continue  # Skip to next field
                        
                        # Initialize if not in previous row (e.g., other new columns)
                        if field not in merged_row.index:
                            merged_row[field] = ''
                        
                        # Get value from previous file (preserve it)
                        prev_val = str(merged_row[field]).strip() if pd.notna(merged_row[field]) else ''
                        
                        # If previous doesn't have it, try to get from new file for new columns
                        if not prev_val:
                            for variant in variants:
                                if variant in new_row.index:
                                    new_val = str(new_row[variant]).strip() if pd.notna(new_row[variant]) else ''
                                    if new_val:
                                        merged_row[field] = new_val
                                        break
                        else:
                            merged_row[field] = prev_val
                    
                    # Handle all other columns
                    for col in new_df.columns:
                        if col not in merged_row.index and col not in self.canonical_fields:
                            merged_row[col] = ''
                        if pd.isna(merged_row[col]) or str(merged_row[col]).strip() == '':
                            new_val = new_row.get(col)
                            if pd.notna(new_val) and str(new_val).strip():
                                merged_row[col] = new_val
                    
                    overlapping_rows.append(merged_row)
                    
                except Exception as e:
                    self.logger.warning(f"Error processing overlapping case {case_num}: {str(e)}")
                    try:
                        fallback_prev = prev_df[prev_df['Case Number'] == case_num].iloc[0]
                        overlapping_rows.append(fallback_prev)
                    except Exception:
                        overlapping_rows.append(new_df[new_df['Case Number'] == case_num].iloc[0])
                    continue
            
            if overlapping_rows:
                overlapping_df = pd.DataFrame(overlapping_rows)
                # Ensure canonical fields are properly set and populated
                for field, variants in self.canonical_fields.items():
                    if field not in overlapping_df.columns:
                        overlapping_df[field] = ''
                    # Try to populate empty values from variants
                    empty_mask = overlapping_df[field].fillna('').astype(str).str.strip() == ''
                    if empty_mask.any():
                        for variant in variants:
                            if variant in overlapping_df.columns:
                                variant_values = overlapping_df.loc[empty_mask, variant].fillna('').astype(str).str.strip()
                                valid_values = variant_values != ''
                                if valid_values.any():
                                    overlapping_df.loc[empty_mask & valid_values.index.isin(overlapping_df.index), field] = variant_values[valid_values]
                                    empty_mask = empty_mask & ~valid_values
                                    if not empty_mask.any():
                                        break
                ordered_rows.append(overlapping_df)
                self.logger.info(f"Added {len(overlapping_df)} OVERLAPPING cases (updated with new data)")
            
            # 3. Add EXISTING cases (cases that only exist in previous file, in their original order)
            if existing_only_cases:
                # Get the original order from previous file
                prev_df_ordered = prev_df.copy()
                prev_df_ordered['_original_index'] = range(len(prev_df_ordered))
                
                # Filter to only include cases not in new data
                existing_only_df = prev_df_ordered[prev_df_ordered['Case Number'].isin(existing_only_cases)]
                
                # Sort by original index to maintain order
                existing_only_df = existing_only_df.sort_values('_original_index')
                existing_only_df = existing_only_df.drop(columns=['_original_index'])
                
                # Handle canonical fields for existing cases
                for field, variants in self.canonical_fields.items():
                    values_found = []
                    for variant in variants:
                        if variant in existing_only_df.columns:
                            non_empty_values = existing_only_df[variant].fillna('').astype(str).str.strip()
                            values_found.extend([(idx, val) for idx, val in non_empty_values.items() if val])
                    if values_found:
                        existing_only_df[field] = ''  # Initialize the canonical field
                        for idx, val in values_found:
                            existing_only_df.at[idx, field] = val
                
                # Ensure all new columns exist and are properly ordered
                for col in new_df.columns:
                    if col not in existing_only_df.columns and col not in self.canonical_fields:
                        existing_only_df[col] = ''
                
                ordered_rows.append(existing_only_df)
                self.logger.info(f"Added {len(existing_only_df)} EXISTING cases in their original order")
            
            # Combine all rows in the correct order
            if ordered_rows:
                merged_df = pd.concat(ordered_rows, ignore_index=True)
            else:
                merged_df = new_df.copy()
            
            # Ensure all canonical fields are present
            for field in self.canonical_fields.keys():
                if field not in merged_df.columns:
                    merged_df[field] = ''
            
            # Final data cleanup and verification
            for field, variants in self.canonical_fields.items():
                if field not in merged_df.columns:
                    continue
                    
                # Clean empty values
                merged_df[field] = merged_df[field].fillna('').astype(str).str.strip()
                
                # SPECIAL HANDLING FOR COMPLETION DATE - only populate new cases from raw data
                if field == 'Completion Date' and not completion_date_in_prev:
                    self.logger.info("Final cleanup: Completion Date not in previous file - only populating new cases")
                    new_cases_mask = merged_df['Case Number'].isin(new_cases)
                    empty_mask = merged_df[field] == ''
                    for variant in variants:
                        if variant in merged_df.columns:
                            update_mask = new_cases_mask & empty_mask
                            if update_mask.any():
                                variant_values = merged_df.loc[update_mask, variant].fillna('').astype(str).str.strip()
                                non_empty_mask = variant_values != ''
                                if non_empty_mask.any():
                                    merged_df.loc[update_mask & non_empty_mask, field] = variant_values[non_empty_mask]
                                    final_populated = (merged_df[field] != '').sum()
                                    self.logger.info(f"Final: Populated {sum(non_empty_mask)} empty Completion Date values from {variant} for new cases")
                                    self.logger.info(f"Total rows with Completion Date: {final_populated}/{len(merged_df)}")
                    continue  # Skip standard population logic for Completion Date
                
                # For new cases, ensure data is populated from variant columns if main field is empty
                new_cases_mask = merged_df['Case Number'].isin(new_cases)
                empty_mask = merged_df[field] == ''
                for variant in variants:
                    if variant in merged_df.columns:
                        update_mask = new_cases_mask & empty_mask
                        variant_values = merged_df.loc[update_mask, variant].fillna('').astype(str).str.strip()
                        non_empty_mask = variant_values != ''
                        if non_empty_mask.any():
                            merged_df.loc[update_mask & non_empty_mask, field] = variant_values[non_empty_mask]
                            self.logger.debug(f"Populated {sum(non_empty_mask)} empty {field} values from {variant}")
            
            # Drop duplicate variant columns and raw contact-related columns
            columns_to_drop = []
            
            # Drop variant columns from canonical fields
            for variants in self.canonical_fields.values():
                for variant in variants:
                    if variant in merged_df.columns and variant not in self.canonical_fields:
                        columns_to_drop.append(variant)
            
            # Drop any raw contact-related columns to prevent confusion
            contact_patterns = ['contact name', 'contact) (contact', 'Contact Name (Contact)', ' Contact Name (Contact) (Contact)']
            for col in merged_df.columns:
                if any(pattern in col.lower() for pattern in contact_patterns):
                    if col not in self.canonical_fields and col != 'Customer Name':
                        columns_to_drop.append(col)
            
            # Drop the columns and log what was removed
            if columns_to_drop:
                self.logger.info("Removing duplicate/raw columns:")
                for col in columns_to_drop:
                    self.logger.info(f"  - {col}")
                merged_df = merged_df.drop(columns=list(set(columns_to_drop)), errors='ignore')
            
            # Ensure proper column order and data types
            merged_df = merged_df.reset_index(drop=True)
            
            self.logger.info(f"\nMerge completed successfully:")
            self.logger.info(f"  Final merged DataFrame shape: {merged_df.shape}")
            self.logger.info(f"  New cases added first: {len(new_cases)}")
            self.logger.info(f"  Overlapping cases updated: {len(overlapping_cases)}")
            self.logger.info(f"  Existing cases preserved in order: {len(existing_only_cases)}")
            
            # Data quality checks
            self.logger.info("\nData Quality Summary:")
            for field in ['Customer Name', 'Email', 'Country', 'State/Province']:
                if field in merged_df.columns:
                    non_empty = (merged_df[field].fillna('').astype(str).str.strip() != '').sum()
                    self.logger.info(f"  {field}: {non_empty} non-empty values ({non_empty/len(merged_df)*100:.1f}%)")
            
            # Order verification
            if len(merged_df) > 0:
                first_new_case = merged_df.iloc[0]['Case Number']
                if first_new_case in new_cases:
                    self.logger.info(f"\nVERIFIED: First case is NEW: {first_new_case}")
                else:
                    self.logger.warning(f"⚠ WARNING: First case is NOT new: {first_new_case}")
                
                # Case distribution verification
                new_cases_in_output = merged_df[merged_df['Case Number'].isin(new_cases)]
                overlapping_cases_in_output = merged_df[merged_df['Case Number'].isin(overlapping_cases)]
                existing_cases_in_output = merged_df[merged_df['Case Number'].isin(existing_only_cases)]
                
                self.logger.info(f"\nCase Distribution Verification:")
                self.logger.info(f"  New cases in output: {len(new_cases_in_output)} (expected {len(new_cases)})")
                self.logger.info(f"  Overlapping cases: {len(overlapping_cases_in_output)} (expected {len(overlapping_cases)})")
                self.logger.info(f"  Existing cases: {len(existing_cases_in_output)} (expected {len(existing_only_cases)})")
                
                # Verify Completion Date population
                if 'Completion Date' in merged_df.columns:
                    completion_date_populated = (merged_df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                    self.logger.info(f"\nCompletion Date Verification:")
                    self.logger.info(f"  Total rows: {len(merged_df)}")
                    self.logger.info(f"  Populated Completion Dates: {completion_date_populated}")
                    self.logger.info(f"  Empty Completion Dates: {len(merged_df) - completion_date_populated}")
                    
                    # Check new cases specifically
                    new_cases_completion = (new_cases_in_output['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                    self.logger.info(f"  Completion Date in new cases: {new_cases_completion}/{len(new_cases_in_output)}")
                    if new_cases_completion > 0:
                        sample_completion = new_cases_in_output[new_cases_in_output['Completion Date'].fillna('').astype(str).str.strip() != '']['Completion Date'].head(3).tolist()
                        self.logger.info(f"  Sample Completion Date values: {sample_completion}")
            
            return merged_df
            
        except Exception as e:
            self.logger.error(f"Error in merge_with_previous: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            # Return new data if merge fails
            return new_df

    def _apply_bank_sutherland_rule(self, df, prev_case_numbers=None):
        """Set Action 1/2/3/Final Action/Status to Bank/Sutherland/closed for specific company names, but do NOT match 'citi' in 'citizen' or 'citizen group'.
        Only applies to new cases (not in prev_case_numbers) to preserve agent actions on previous cases."""
        # List of patterns (case-insensitive, partial match, with special handling for 'citi')
        patterns = [
            r"\bcitibank\b|\bcitigroup\b|\bciti group\b|\bciti bank\b|\bciti corp\b|\bciti\b",  # Citi, but not Citizen
            r"amgen",
            r"sanofi",
            r"bns|scotiabank|scotia bank|scotiabank canada|scotiaBank|bank of nova scotia|bank nova scotia",
            r"rbc|royal bank|royal bank of canada|rbc wealth management",
            r"hsbc",
            r"td bank"
        ]
        import re
        def matches_company(name):
            if not isinstance(name, str):
                return False
            name_lc = name.lower()
            # Special handling: skip if 'citizen' in name
            if 'citizen' in name_lc:
                return False
            return any(re.search(p, name_lc) for p in patterns)
        
        # Determine the company column in a flexible/canonical way
        company_col = self._get_company_column(df)
        # Additional fallbacks for variant names
        if company_col is None:
            for candidate in ['Company Name', 'Customer Name', 'company name', 'customer name']:
                if candidate in df.columns:
                    company_col = candidate
                    break

        # If still not found, create a canonical 'Company Name' column so downstream logic can rely on it
        if company_col is None:
            self.logger.warning("Company column not found; creating empty 'Company Name' column for Bank/Sutherland rule")
            df['Company Name'] = ''
            company_col = 'Company Name'

        # Ensure we operate on a clean string series for matching
        company_series = df[company_col].fillna('').astype(str)

        # Only apply to new cases (not in previous file)
        if prev_case_numbers is not None:
            # Create mask for new cases only using integer-normalized case numbers
            try:
                df['Case Number'] = df['Case Number'].apply(self.normalize_case_number).astype('Int64')
            except Exception:
                # If Case Number normalization fails, coerce to numeric as fallback
                df['Case Number'] = pd.Series(pd.to_numeric(df['Case Number'], errors='coerce')).astype('Int64')
            new_cases_mask = ~df['Case Number'].isin(prev_case_numbers)
            company_match_mask = company_series.apply(matches_company)
            # Combine masks: must be both a new case AND match company pattern
            mask = new_cases_mask & company_match_mask
            self.logger.info(f"Bank/Sutherland rule: Applied to {mask.sum()} new cases (out of {company_match_mask.sum()} total matching companies) using column '{company_col}'")
        else:
            # Fallback: apply to all cases if no previous case numbers provided
            mask = company_series.apply(matches_company)
            self.logger.info(f"Bank/Sutherland rule: Applied to {mask.sum()} cases (no previous case filter) using column '{company_col}'")
        
        # Store the updated cases for summary tracking
        if mask.sum() > 0:
            updated_cases = df[mask].copy()
            # Add to processing stats for summary
            if not hasattr(self, 'bank_sutherland_updated_cases'):
                self.bank_sutherland_updated_cases = []
            
            for idx, row in updated_cases.iterrows():
                case_number = str(row['Case Number']).strip()
                try:
                    company_name = str(row.get(company_col, '')).strip()
                except Exception:
                    company_name = ''
                self.bank_sutherland_updated_cases.append({
                    'case_number': case_number,
                    'company_name': company_name
                })
            
            # Debug logging for Bank/Sutherland tracking
            self.logger.info(f"Bank/Sutherland Debug - Updated {len(updated_cases)} cases")
            self.logger.info(f"Bank/Sutherland Debug - Tracking {len(self.bank_sutherland_updated_cases)} total cases")
            for case in self.bank_sutherland_updated_cases[-len(updated_cases):]:  # Show the latest ones
                self.logger.info(f"Bank/Sutherland Debug - Case: {case['case_number']}, Company: {case['company_name']}")
        
        for col in ['Action 1', 'Action 2', 'Action 3', 'Final Action']:
            df.loc[mask, col] = 'Bank/Sutherland'
        df.loc[mask, 'Status'] = 'closed'
        return df

    def format_output_data(self, df, prev_df=None):
        """Format the data according to the required output columns"""
        try:
            # Canonical column mapping for all input variants - MUST match self.output_columns names
            canonical_mapping = {
                'Case Number': ['Case Number'],
                'Work Order Type': ['Work Order Type'],
                'State/Province': ['State/Province (Case) (Case)', 'State/Province'],  # CRITICAL: Use exact output column name
                'Local Time': ['Local Time'],
                'Action 1': ['Action 1'],
                'Action 2': ['Action 2'],
                'Action 3': ['Action 3'],
                'Final Action': ['Final Action'],
                'Assigned To': ['Assigned To'],
                'Status': ['Status'],
                'Company Name': ['Company Name'],
                'DND (Do Not Disturb)': ['DND (Do Not Disturb)', 'Do Not Disturb (Contact) (Contact)'],
                'Email': ['Primary Email (Contact) (Contact)', 'Email'],  # CRITICAL: Use exact output column name
                'Phone Number': ['Contact Mobile Phone', 'Phone Number'],
                'Incoming Channel': ['Incoming Channel (Case) (Case)', 'Incoming Channel'],
                'Last Status Change': ['Last Status Change (Workflow Use Only) (Case) (Case)', 'Last Status Change'],
                'Country': ['Country/Region (Contact) (Contact)', 'Country'],  # CRITICAL: Use exact output column name
                'Customer Name': [' Contact Name (Contact) (Contact)', 'Contact Name', 'Contact Name (Contact)', 'Customer Name'],  # CRITICAL: Add Customer Name mapping
                'Case': ['Case'],
                'Problem Description': ['Problem Description (Case) (Case)', 'Problem Description'],
                'Case Status': ['Case Status (Case) (Case)', 'Case Status'],  # CRITICAL: Add Case Status mapping
                'Case Status Updated': ['Case status Updated (Case) (Case)', 'Case Status Updated'],
                'Case Reason': ['Case Reason (Case) (Case)', 'Case Reason'],
                'Work Order ID': ['Work Order ID'],
                'Work Order Status': ['Work Order Status'],
                'Order Type': ['Order Type'],
                'Work Order Priority': ['Work Order Priority'],
                'Product ID (MTM)': ['Product ID (MTM) (Case) (Case)', 'Product ID (MTM)'],
                'Machine Type': ['Machine Type'],
                'Product Description': ['Product Description'],
                'Serial Number': ['Serial Number (Case) (Case)', 'Serial Number'],
                'Created On': ['Created On'],
                'Survey Preference': ['Survey Preference (Case) (Case)', 'Survey Preference'],
                'Survey Fatigue': ['Survey Fatigue (Case) (Case)', 'Survey Fatigue'],
                'No survey reason': ['No survey reason (Case) (Case)', 'No survey reason'],
                'Program': ['Program (Case) (Case)', 'Program'],
                'Repeat Frequency': ['Repeat Frequency (Case) (Case)', 'Repeat Frequency'],
                'Repeat Repair': ['Repeat Repair'],
                'Closing Code': ['Closing Code'],
                'Reported Symptom': ['Reported Symptom'],
                'Completion Date': ['Completion Date']  # CRITICAL: Add Completion Date mapping
            }

            # Use self.output_columns as the desired output columns (defined at init)
            # This ensures consistency with expected output format
            desired_columns = list(self.output_columns) + ['_Source_Sheet']
            
            # Initialize input mapping (not really used but kept for backward compatibility)
            input_mapping = {
                # Fixed order columns first
                'Case Number': self.columns.get('case_number', 'Case Number'),
                'Work Order Type': 'Work Order Type',
                'State/Province': self.columns.get('state', 'State/Province (Case) (Case)'),
                
                # Then map all input columns
                'Country/Region (Contact) (Contact)': 'Country/Region (Contact) (Contact)',
                'Customer Name': [' Contact Name (Contact) (Contact)', 'Contact Name', 'Customer Name'],  # Map contact fields to Customer Name
                'Company Name': 'Company Name',
                'Primary Email (Contact) (Contact)': 'Primary Email (Contact) (Contact)',
                'Contact Mobile Phone': 'Contact Mobile Phone',
                'Case': 'Case',
                'Case Status (Case) (Case)': 'Case Status (Case) (Case)',
                'Last Status Change (Workflow Use Only) (Case) (Case)': 'Last Status Change (Workflow Use Only) (Case) (Case)',
                'Phone Number': self.columns.get('phone', 'Phone Number'),
                'Case Number': self.columns.get('case_number', 'Case Number'),
                'Case': self.columns.get('case', 'Case'),
                'Reported Symptom': self.columns.get('reported_symptom', 'Reported Symptom'),
                'Problem Description': self.columns.get('problem_description', 'Problem Description (Case) (Case)'),
                'Last Status Change': self.columns.get('last_status_change', 'Last Status Change'),
                'Case Status Updated': self.columns.get('case_status_updated', 'Case status Updated'),
                'Case Reason': self.columns.get('case_reason', 'Case Reason'),
                'Work Order ID': self.columns.get('wo_id', 'Work Order ID'),
                'Work Order Status': self.columns.get('wo_status', 'Work Order Status'),
                'Work Order Type': self.columns.get('wo_type', 'Work Order Type'),
                'Incoming Channel': self.columns.get('incoming_channel', 'Incoming Channel'),
                'Order Type': self.columns.get('order_type', 'Order Type'),
                'Work Order Priority': self.columns.get('wo_priority', 'Work Order Priority'),
                'Product ID (MTM)': self.columns.get('product_id', 'Product ID (MTM) (Case) (Case)'),
                'Machine Type': self.columns.get('machine_type', 'Machine Type'),
                'Product Description': self.columns.get('product_desc', 'Product Description'),
                'Serial Number': self.columns.get('serial', 'Serial Number'),
                'Created On': self.columns.get('created_on', 'Created On'),
                'Survey Preference': self.columns.get('survey_pref', 'Survey Preference'),
                'Survey Fatigue': self.columns.get('survey_fatigue', 'Survey Fatigue'),
                'No survey reason': self.columns.get('no_survey_reason', 'No survey reason'),
                'Program': self.columns.get('program', 'Program'),
                'Repeat Frequency': self.columns.get('repeat_freq', 'Repeat Frequency'),
                'Repeat Repair': self.columns.get('repeat_repair', 'Repeat Repair'),
                'Closing Code': self.columns.get('closing_code', 'Closing Code'),
                'DND (Do Not Disturb)': self.columns.get('dnd', 'Do Not Disturb'),
            }
            
            # Create a new DataFrame with the same index as the input DataFrame
            output_df = pd.DataFrame(index=df.index, columns=desired_columns)
            
            # Initialize all columns with empty strings
            for col in desired_columns:
                output_df[col] = ''
            
            # Copy data from source DataFrame to output DataFrame using canonical mapping
            for target_col, source_cols in canonical_mapping.items():
                found_col = None
                found_value = None
                
                # First pass: Look for exact matches in all source columns
                for source_col in source_cols:
                    if source_col in df.columns:
                        found_values = df[source_col].fillna('')
                        non_empty = found_values[found_values.astype(str).str.strip() != '']
                        if not non_empty.empty:
                            found_col = source_col
                            found_value = non_empty.iloc[0]  # Take first non-empty value
                            self.logger.debug(f"Found exact match for {target_col} in {source_col}")
                            break
                
                if found_col:
                    if target_col in ['Last Status Change', 'Created On', 'Completion Date'] and pd.notna(found_value) and hasattr(found_value, 'strftime'):
                        # Handle datetime values - date only format (no time)
                        output_df[target_col] = df[found_col].apply(
                            lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and hasattr(x, 'strftime') else str(x) if pd.notna(x) else ''
                        )
                    else:
                        output_df[target_col] = df[found_col].fillna('').astype(str)
                    
                    non_empty_count = (output_df[target_col] != '').sum()
                    self.logger.info(f"Mapped {found_col} -> {target_col}: {non_empty_count} non-empty values")
                    
                    # Special handling for DND and Problem Description in new cases
                    if target_col in ['DND (Do Not Disturb)', 'Problem Description']:
                        # Only update if we actually have data from raw input
                        raw_values = df[found_col].fillna('').astype(str).str.strip()
                        non_empty_raw = raw_values[raw_values != '']
                        if not non_empty_raw.empty:
                            # Log for each case that gets data from raw
                            for idx, val in non_empty_raw.items():
                                case_num = df.at[idx, 'Case Number']
                                if pd.notna(case_num):
                                    self.logger.debug(f"Set {target_col} for case {case_num} from raw input: '{val}'")
                    
                    # Add debug logging for merged fields
                    if found_col != target_col:  # Only log when a secondary name was used
                        sample_data = output_df[output_df[target_col] != '']
                        if not sample_data.empty:
                            sample_cases = sample_data['Case Number'].head(3).tolist()
                            self.logger.debug(f"Merged '{found_col}' into '{target_col}' for cases: {sample_cases}")
                    
                    # Add debug logging for merged fields
                    if found_col != target_col:  # Only log when a secondary name was used
                        sample_data = output_df[output_df[target_col] != '']
                        if not sample_data.empty:
                            sample_cases = sample_data['Case Number'].head(3).tolist()
                            self.logger.debug(f"Merged '{found_col}' into '{target_col}' for cases: {sample_cases}")
                    
                    # Special debugging for Incoming Channel
                    if target_col == 'Incoming Channel':
                        sample_values = output_df[target_col].head(5).tolist()
                        self.logger.info(f"Incoming Channel sample values: {sample_values}")
                    
                    # Special debugging for State/Province
                    if target_col == 'State/Province':
                        sample_values = output_df[target_col].head(5).tolist()
                        self.logger.info(f"State/Province sample values: {sample_values}")
                        # Additional debugging for State/Province
                        non_empty_count = (output_df[target_col].astype(str).str.strip() != '').sum()
                        self.logger.info(f"State/Province non-empty count: {non_empty_count} out of {len(output_df)}")
                        if non_empty_count > 0:
                            non_empty_values = output_df[target_col][output_df[target_col].astype(str).str.strip() != ''].head(10).tolist()
                            self.logger.info(f"State/Province non-empty values: {non_empty_values}")
                else:
                    self.logger.warning(f"None of the source columns {source_cols} found in input data")
                    # Special debugging for State/Province source column
                    if target_col == 'State/Province':
                        self.logger.error(f"State/Province source columns {source_cols} not found in input data")
                        self.logger.info(f"Available columns in input data: {df.columns.tolist()}")
                        # Check for similar column names
                        similar_cols = [col for col in df.columns if 'state' in col.lower() or 'province' in col.lower()]
                        if similar_cols:
                            self.logger.info(f"Found similar columns: {similar_cols}")
            
            # Add Local Time column (empty for later processing)
            output_df['Local Time'] = ''
            
            # Ensure _Source_Sheet exists for preservation logic
            if '_Source_Sheet' not in output_df.columns:
                output_df['_Source_Sheet'] = ''
                
            # Set initial Status while preserving previous statuses
            # Default all statuses to 'new'
            output_df['Status'] = 'new'
                
            # If we have previous data, preserve existing case statuses
            if prev_df is not None and not prev_df.empty and 'Status' in prev_df.columns:
                try:
                    # Normalize case numbers in both frames
                    output_df['_temp_case'] = output_df['Case Number'].apply(self.normalize_case_number)
                    prev_df['_temp_case'] = prev_df['Case Number'].apply(self.normalize_case_number)
                    
                    # Create mask for matching cases
                    for case_num in prev_df['_temp_case'].dropna().unique():
                        mask = (output_df['_temp_case'] == case_num)
                        if mask.any():
                            prev_status = prev_df.loc[prev_df['_temp_case'] == case_num, 'Status'].iloc[0]
                            output_df.loc[mask, 'Status'] = prev_status
                            self.logger.debug(f"Preserved Status '{prev_status}' for case {case_num}")
                    
                    # Clean up temp columns
                    output_df = output_df.drop(columns=['_temp_case'])
                    prev_df = prev_df.drop(columns=['_temp_case'])
                        
                except Exception as e:
                    self.logger.warning(f"Error preserving previous statuses: {str(e)}")
                    # Remove temp column if it exists
                    if '_temp_case' in output_df.columns:
                        output_df = output_df.drop(columns=['_temp_case'])
                    if '_temp_case' in prev_df.columns:
                        prev_df = prev_df.drop(columns=['_temp_case'])
            
            # Ensure all columns are strings and empty values are handled
            for col in output_df.columns:
                output_df[col] = output_df[col].fillna('').astype(str)
            
            # Debug: Show sample of formatted data
            self.logger.info(f"\nSample of formatted data (first 3 rows):")
            try:
                for idx, row in output_df.head(3).iterrows():
                    case_num = row.get('Case Number', 'N/A')
                    Customer_name = row.get(' Contact Name (Contact) (Contact)', 'N/A')
                    company = row.get('Company Name', 'N/A')
                    self.logger.info(f"Row {idx}: Case={case_num}, Contact={Customer_name}, Company={company}")
            except Exception as e:
                self.logger.warning(f"Error showing sample data: {str(e)}")

            # Add the previous data if available
            prev_case_numbers = set()
            if prev_df is not None and not prev_df.empty:
                # DIAGNOSTIC: Check Completion Date in output_df before filtering
                if 'Completion Date' in output_df.columns:
                    completion_date_non_empty = (output_df['Completion Date'].fillna('').astype(str).str.strip() != '').sum()
                    self.logger.info(f"DIAGNOSTIC format_output_data: Before filtering - Completion Date has {completion_date_non_empty}/{len(output_df)} non-empty values")
                    if completion_date_non_empty == 0:
                        self.logger.error(f"CRITICAL: Completion Date is empty in output_df BEFORE merge_with_previous!")
                else:
                    self.logger.error(f"CRITICAL: Completion Date column not in output_df at all!")
                
                for col in desired_columns:
                    if col not in prev_df.columns:
                        prev_df[col] = ''
                output_df = output_df[desired_columns]
                prev_df = prev_df[desired_columns]
                merged_df = self.merge_with_previous(output_df, prev_df)
                # Get previous case numbers for filtering rules
                prev_case_numbers = set(pd.Series(prev_df['Case Number']).apply(self.normalize_case_number).astype('Int64').dropna().tolist())
            else:
                merged_df = output_df

            # Create case number mask for new cases
            new_cases_mask = ~merged_df['Case Number'].apply(self.normalize_case_number).isin(prev_case_numbers)

            # Backfill important contact/company/location fields for NEW cases using values from the pre-merge output_df
            try:
                # Build a lookup from normalized Case Number -> value for each field from output_df
                out_lookup = output_df.copy()
                out_lookup['__norm_case'] = out_lookup['Case Number'].apply(self.normalize_case_number)
                out_lookup = out_lookup.set_index('__norm_case')

                fields_to_backfill = [
                    ('Email', ['Primary Email (Contact) (Contact)', 'Email', 'Contact Email']),
                    ('Phone Number', ['Contact Mobile Phone', 'Phone Number', 'Mobile Phone', 'Contact Phone']),
                    ('Country', ['Country/Region (Contact) (Contact)', 'Country/Region', 'Country (Contact)', 'Country']),
                    ('State/Province', ['State/Province (Case) (Case)', 'State/Province (Contact) (Contact)', 'State/Province (Case)', 'State/Province']),
                    ('Case Status', ['Case Status (Case) (Case)', 'Case Status (Case)', 'Case Status']),
                    ('Last Status Change', ['Last Status Change (Workflow Use Only) (Case) (Case)', 'Last Status Change (Case)', 'Last Status Change']),
                    ('Customer Name', [' Contact Name (Contact) (Contact)', 'Contact Name', 'Contact Name (Contact)', 'Customer Name']),
                    ('Company Name', ['Company Name', 'Company Name (Contact) (Contact)', 'Company (Contact)']),
                ]
                for field, source_fields in fields_to_backfill:
                    if field not in merged_df.columns:
                        merged_df[field] = ''  # Create the column if it doesn't exist
                    
                    # Rows that are new and currently empty for this field
                    empty_mask = merged_df[field].fillna('').astype(str).str.strip() == ''
                    to_fill_mask = new_cases_mask & empty_mask
                    if not to_fill_mask.any():
                        continue

                    # Try each source field in order
                    for source_field in source_fields:
                        if source_field in out_lookup.columns:
                            for idx in merged_df[to_fill_mask].index:
                                try:
                                    case_norm = self.normalize_case_number(merged_df.at[idx, 'Case Number'])
                                    if case_norm is not None and case_norm in out_lookup.index:
                                        val = out_lookup.at[case_norm, source_field]
                                        if pd.notna(val) and str(val).strip() != '':
                                            merged_df.at[idx, field] = str(val).strip()
                                            to_fill_mask.at[idx] = False  # Mark as filled
                                            self.logger.debug(f"Backfilled {field} for new case {case_norm} from {source_field}")
                                except Exception:
                                    continue
                            
                            if not to_fill_mask.any():  # Stop if all rows are filled
                                break

                    # For each row needing fill, try to lookup from out_lookup by normalized case
                    for idx in merged_df[to_fill_mask].index:
                        try:
                            case_norm = self.normalize_case_number(merged_df.at[idx, 'Case Number'])
                            if case_norm is None:
                                continue
                            if case_norm in out_lookup.index:
                                val = out_lookup.at[case_norm, field] if field in out_lookup.columns else ''
                                if pd.notna(val) and str(val).strip() != '':
                                    merged_df.at[idx, field] = str(val).strip()
                                    self.logger.debug(f"Backfilled {field} for new case {case_norm} from raw data")
                        except Exception:
                            continue
            except Exception as e:
                self.logger.debug(f"Error during backfill of important fields for new cases: {e}")
            
            # Set default DND value for new cases with empty DND
            if 'DND (Do Not Disturb)' in merged_df.columns:
                empty_dnd_mask = merged_df['DND (Do Not Disturb)'].fillna('').astype(str).str.strip() == ''
                new_empty_dnd_mask = new_cases_mask & empty_dnd_mask
                if new_empty_dnd_mask.any():
                    merged_df.loc[new_empty_dnd_mask, 'DND (Do Not Disturb)'] = 'No'
                    for case_num in merged_df.loc[new_empty_dnd_mask, 'Case Number']:
                        self.logger.debug(f"Set default DND='No' for new case {case_num}")

            # Check for empty Problem Description in new cases
            if 'Problem Description' in merged_df.columns:
                empty_desc_mask = merged_df['Problem Description'].fillna('').astype(str).str.strip() == ''
                new_empty_desc_mask = new_cases_mask & empty_desc_mask
                if new_empty_desc_mask.any():
                    for case_num in merged_df.loc[new_empty_desc_mask, 'Case Number']:
                        self.logger.debug(f"No Problem Description found for new case {case_num}")
            
            # Final verification of DND values
            if 'DND (Do Not Disturb)' in merged_df.columns:
                new_cases_no_dnd = merged_df[new_cases_mask & (merged_df['DND (Do Not Disturb)'].fillna('').astype(str).str.strip() == '')]
                if not new_cases_no_dnd.empty:
                    self.logger.warning(f"Found {len(new_cases_no_dnd)} new cases still missing DND value after setting defaults")
                    for _, row in new_cases_no_dnd.head(5).iterrows():
                        self.logger.warning(f"Case {row['Case Number']} missing DND value")
            # Apply Bank/Sutherland rule (only to new cases)
            self.logger.info("\n=== Applying Bank/Sutherland Rule ===")
            merged_df = self._apply_bank_sutherland_rule(merged_df, prev_case_numbers=prev_case_numbers)
            self.logger.info(f"Bank/Sutherland rule applied. Total cases tracked: {len(getattr(self, 'bank_sutherland_updated_cases', []))}")
            
            # Apply DND rule (only to new cases)
            self.logger.info("\n=== Applying DND Rule ===")
            merged_df = self._apply_dnd_actions_and_status(merged_df, prev_case_numbers=prev_case_numbers)
            self.logger.info(f"DND rule applied. Total cases tracked: {len(getattr(self, 'dnd_updated_cases', []))}")
            
            # CRITICAL DATA QUALITY CHECK for new cases
            if prev_df is not None and not prev_df.empty:
                self.logger.info("\n=== NEW CASES DATA QUALITY CHECK ===")
                new_cases_data = merged_df[new_cases_mask].copy()
                if not new_cases_data.empty:
                    for field in ['Customer Name', 'Email', 'Country', 'State/Province', 'Case Status']:
                        if field in new_cases_data.columns:
                            empty_count = (new_cases_data[field].fillna('').astype(str).str.strip() == '').sum()
                            if empty_count > 0:
                                self.logger.warning(f"WARNING: {empty_count} new cases missing {field}")
                                sample_missing = new_cases_data[new_cases_data[field].fillna('').astype(str).str.strip() == '']['Case Number'].head(5).tolist()
                                self.logger.warning(f"Sample cases missing {field}: {sample_missing}")
                else:
                    self.logger.info("No new cases found for data quality check")

            # FINAL VERIFICATION: Ensure the order is maintained after applying rules
            if prev_df is not None and not prev_df.empty:
                self.logger.info("\n=== FINAL ORDER VERIFICATION ===")
                
                # Get the case numbers in order from the final output
                final_case_numbers = pd.Series(merged_df['Case Number']).apply(self.normalize_case_number).astype('Int64').dropna().tolist()
                
                # Count how many new cases are at the top
                new_cases_at_top = 0
                for case_num in final_case_numbers:
                    if case_num not in prev_case_numbers:
                        new_cases_at_top += 1
                    else:
                        break  # Stop counting when we hit the first previous case
                
                self.logger.info(f"New cases at the top: {new_cases_at_top}")
                
                # Verify that new cases are actually at the top
                if new_cases_at_top > 0:
                    first_new_case = merged_df.iloc[0]['Case Number']
                    if first_new_case not in prev_case_numbers:
                        self.logger.info(f"SUCCESS: First case is NEW: {first_new_case}")
                    else:
                        self.logger.error(f"❌ ERROR: First case is NOT new: {first_new_case}")
                    
                    # Show the first few cases to verify order
                    first_10_cases = merged_df.head(10)['Case Number'].tolist()
                    self.logger.info(f"First 10 cases in final output: {first_10_cases}")
                    
                    # Count new vs previous cases in first 10
                    new_in_first_10 = sum(1 for case in first_10_cases if case not in prev_case_numbers)
                    prev_in_first_10 = sum(1 for case in first_10_cases if case in prev_case_numbers)
                    self.logger.info(f"First 10 breakdown: {new_in_first_10} new, {prev_in_first_10} previous")
                else:
                    self.logger.warning("⚠ WARNING: No new cases found at the top of the output")
            
            # Verify the output
            self.logger.info(f"\nOutput DataFrame created with {len(merged_df)} rows")
            self.logger.info(f"Columns in order: {merged_df.columns.tolist()}")
            
            # Debug: Check if State/Province column exists in final output
            if 'State/Province' in merged_df.columns:
                self.logger.info(f"State/Province column found at position {merged_df.columns.get_loc('State/Province') + 1}")
                state_count = (merged_df['State/Province'].astype(str).str.strip() != '').sum()
                self.logger.info(f"State/Province non-empty count in final output: {state_count} out of {len(merged_df)}")
            else:
                self.logger.error("CRITICAL ERROR: State/Province column is missing from final output!")
                self.logger.error(f"Available columns: {merged_df.columns.tolist()}")
            
                # CRITICAL FINAL VERIFICATION: Ensure State/Province column is present and has data
                state_province_col = 'State/Province (Case) (Case)'  # Use full column name
                if state_province_col in merged_df.columns:
                    final_state_count = (merged_df[state_province_col].astype(str).str.strip() != '').sum()
                    self.logger.info(f"Final State/Province non-empty count: {final_state_count} out of {len(merged_df)}")
                    if final_state_count > 0:
                        final_states = merged_df[state_province_col][merged_df[state_province_col].astype(str).str.strip() != ''].head(10).tolist()
                        self.logger.info(f"Final State/Province sample values: {final_states}")
                    else:
                        self.logger.warning("WARNING: No State/Province data found in final output!")
                else:
                    self.logger.error("CRITICAL ERROR: State/Province column missing from final output!")
                    # Emergency fix: add State/Province column after Country/Region
                    country_col = 'Country/Region (Contact) (Contact)'  # Use full column name
                    try:
                        # First try to add after Country/Region
                        merged_df.insert(merged_df.columns.get_loc(country_col) + 1, state_province_col, '')
                        self.logger.info("Emergency fix: Added State/Province column after Country/Region")
                    except (KeyError, ValueError):
                        # If Country/Region not found, add as third column (after Work Order Type)
                        merged_df.insert(2, state_province_col, '')
                        self.logger.info("Emergency fix: Added State/Province column as third column")
                
            # Filter to only include expected output columns (remove unwanted columns like "Unnamed: 18", "Contact Name", etc.)
            self.logger.info(f"\n=== FILTERING FINAL OUTPUT COLUMNS ===")
            self.logger.info(f"Before filtering: {len(merged_df.columns)} columns")
            self.logger.info(f"Columns before filtering: {merged_df.columns.tolist()}")
            
            final_columns = list(self.output_columns)
            if '_Source_Sheet' in merged_df.columns:
                final_columns.append('_Source_Sheet')
            
            # Get only columns that exist in merged_df
            available_final_columns = [col for col in final_columns if col in merged_df.columns]
            
            # Filter merged_df to only include these columns (in the correct order)
            merged_df = merged_df[available_final_columns]
            
            self.logger.info(f"After filtering: {len(merged_df.columns)} columns")
            self.logger.info(f"Final output columns: {merged_df.columns.tolist()}")
            
            # Check for any unwanted columns that might have sneaked in
            unwanted_columns = [col for col in merged_df.columns if col not in final_columns and col != '_Source_Sheet']
            if unwanted_columns:
                self.logger.warning(f"WARNING: Found unwanted columns in final output: {unwanted_columns}")
                # Remove them
                merged_df = merged_df[[col for col in merged_df.columns if col in final_columns or col == '_Source_Sheet']]
                self.logger.info(f"Removed unwanted columns. Final count: {len(merged_df.columns)}")
            
            return merged_df
        except Exception as e:
            self.logger.error(f"Error in format_output_data: {str(e)}")
            self.logger.error(f"Input DataFrame shape: {df.shape}")
            self.logger.error(f"Input DataFrame columns: {df.columns.tolist()}")
            raise

    def _apply_dnd_actions_and_status(self, df, prev_case_numbers=None):
        """After merging, if the 'DND (Do Not Disturb)' column is 'Yes' for any case, set Action 1, Action 2, Action 3, and Final Action to 'DND', and Status to 'closed' for that row.
        Only applies to new cases (not in prev_case_numbers) to preserve agent actions on previous cases."""
        dnd_col = 'DND (Do Not Disturb)'
        if dnd_col in df.columns:
            dnd_mask = df[dnd_col].astype(str).str.strip().str.lower() == 'yes'
            
            # Only apply to new cases (not in previous file)
            if prev_case_numbers is not None:
                # Create mask for new cases only using integer-normalized case numbers
                df['Case Number'] = df['Case Number'].apply(self.normalize_case_number).astype('Int64')
                new_cases_mask = ~df['Case Number'].isin(prev_case_numbers)
                # Combine masks: must be both a new case AND have DND = 'Yes'
                mask = new_cases_mask & dnd_mask
                self.logger.info(f"DND rule: Applied to {mask.sum()} new cases (out of {dnd_mask.sum()} total DND cases)")
            else:
                # Fallback: apply to all cases if no previous case numbers provided
                mask = dnd_mask
                self.logger.info(f"DND rule: Applied to {mask.sum()} cases (no previous case filter)")
            
            # Store the updated cases for summary tracking
            if mask.sum() > 0:
                updated_cases = df[mask].copy()
                # Add to processing stats for summary
                if not hasattr(self, 'dnd_updated_cases'):
                    self.dnd_updated_cases = []
                
                for idx, row in updated_cases.iterrows():
                    case_number = str(row['Case Number']).strip()
                    self.dnd_updated_cases.append({
                        'case_number': case_number,
                        'source': 'DND rule'
                    })
                
                # Debug logging for DND tracking
                self.logger.info(f"DND Debug - Updated {len(updated_cases)} cases")
                self.logger.info(f"DND Debug - Tracking {len(self.dnd_updated_cases)} total cases")
                for case in self.dnd_updated_cases[-len(updated_cases):]:  # Show the latest ones
                    self.logger.info(f"DND Debug - Case: {case['case_number']}, Source: {case['source']}")
            
            for action_col in ['Action 1', 'Action 2', 'Action 3', 'Final Action']:
                if action_col in df.columns:
                    df.loc[mask, action_col] = 'DND'
            if 'Status' in df.columns:
                df.loc[mask, 'Status'] = 'closed'
        return df

    def _force_additional_balancing(self, df, handlers, prev_assignments):
        """Force additional balancing to achieve much more even distribution"""
        try:
            self.logger.info("Starting additional balancing for fairer distribution...")
            
            # Work on a copy to avoid corrupting the original
            df_copy = df.copy()
            
            # Ensure preserved assignments are maintained
            for idx, row in df_copy.iterrows():
                case_num = self.normalize_case_number(row['Case Number'])
                if case_num is not None and case_num in prev_assignments:
                    df_copy.at[idx, 'Assigned To'] = prev_assignments[case_num]
            
            # Calculate current workload
            current_workload = {}
            for handler in handlers:
                count = len(df_copy[df_copy['Assigned To'] == handler])
                current_workload[handler] = count
            
            self.logger.info(f"Current workload before additional balancing: {current_workload}")
            
            # Find handlers that are significantly overloaded and underloaded
            avg_workload = sum(current_workload.values()) / len(current_workload)
            overloaded_threshold = avg_workload * 1.15  # 15% above average
            underloaded_threshold = avg_workload * 0.85  # 15% below average
            
            overloaded_handlers = [h for h, w in current_workload.items() if w > overloaded_threshold]
            underloaded_handlers = [h for h, w in current_workload.items() if w < underloaded_threshold]
            
            self.logger.info(f"Overloaded handlers (> {overloaded_threshold:.1f}): {overloaded_handlers}")
            self.logger.info(f"Underloaded handlers (< {underloaded_threshold:.1f}): {underloaded_handlers}")
            
            # Move cases from overloaded to underloaded handlers (ONLY NEW CASES)
            moved_count = 0
            
            for overloaded_handler in overloaded_handlers:
                if not underloaded_handlers:
                    break
                
                # Find ONLY NEW cases assigned to this overloaded handler
                new_cases_mask = ~df_copy['Case Number'].isin(prev_assignments)
                overloaded_cases = df_copy[(df_copy['Assigned To'] == overloaded_handler) & new_cases_mask]
                
                self.logger.info(f"Found {len(overloaded_cases)} new cases for {overloaded_handler} that can be moved")
                
                if not overloaded_cases.empty:
                    # Sort cases by company to maintain consistency
                    company_col = self._get_company_column(df_copy)
                    if company_col and company_col in df_copy.columns:
                        overloaded_cases['Company_Lower'] = overloaded_cases[company_col].fillna('').astype(str).str.strip().str.lower()
                        overloaded_cases = overloaded_cases.sort_values('Company_Lower')
                    
                    # Move cases to underloaded handlers
                    for idx, row in overloaded_cases.iterrows():
                        if not underloaded_handlers:
                            break
                        
                        # Find the most underloaded handler
                        target_handler = min(underloaded_handlers, key=lambda h: current_workload[h])
                        
                        # Only move if it's not in previous assignments
                        case_num = self.normalize_case_number(row['Case Number'])
                        if case_num is not None and case_num in prev_assignments:
                            continue
                        
                        # Move the case
                        old_handler = df_copy.at[idx, 'Assigned To']
                        df_copy.at[idx, 'Assigned To'] = target_handler
                        current_workload[overloaded_handler] -= 1
                        current_workload[target_handler] += 1
                        moved_count += 1
                        
                        self.logger.info(f"Moved case {case_num} from {old_handler} to {target_handler}")
                        
                        # Remove target handler if it's no longer underloaded
                        if current_workload[target_handler] >= underloaded_threshold:
                            underloaded_handlers.remove(target_handler)
                            self.logger.info(f"Handler {target_handler} no longer underloaded, removing from list")
                        
                        # Update the workload tracking
                        self.logger.debug(f"Updated workload: {current_workload}")
            
            self.logger.info(f"Additional balancing completed: {moved_count} cases moved")
            
            # Verify that no preserved assignments were corrupted
            preserved_corrupted = []
            for idx, row in df_copy.iterrows():
                case_num = self.normalize_case_number(row['Case Number'])
                if case_num is not None and case_num in prev_assignments:
                    expected_handler = prev_assignments[case_num]
                    actual_handler = row['Assigned To']
                    if expected_handler != actual_handler:
                        preserved_corrupted.append((case_num, expected_handler, actual_handler))
            
            if preserved_corrupted:
                self.logger.error(f"CRITICAL ERROR: {len(preserved_corrupted)} preserved assignments corrupted during additional balancing!")
                for case_num, expected, actual in preserved_corrupted:
                    self.logger.error(f"  {case_num}: expected {expected}, got {actual}")
                    # Restore immediately
                    case_idx = df_copy[df_copy['Case Number'] == case_num].index
                    if len(case_idx) > 0:
                        df_copy.at[case_idx[0], 'Assigned To'] = expected
                        self.logger.info(f"Immediately restored corrupted assignment: {case_num} -> {expected}")
            else:
                self.logger.info("SUCCESS: All preserved assignments maintained during additional balancing")
            
            # Final workload calculation
            final_workload = {}
            for handler in handlers:
                count = len(df_copy[df_copy['Assigned To'] == handler])
                final_workload[handler] = count
            
            self.logger.info(f"Final workload after additional balancing: {final_workload}")
            
            return df_copy
            
        except Exception as e:
            self.logger.error(f"Error in additional balancing: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return df

    def _force_gentle_rebalancing(self, df, handlers, prev_assignments):
        """Force additional balancing to achieve much more even distribution"""
        try:
            self.logger.info("Starting additional balancing for fairer distribution...")
            
            # Work on a copy to avoid corrupting the original
            df_copy = df.copy()
            
            # Ensure preserved assignments are maintained
            for idx, row in df_copy.iterrows():
                case_num = self.normalize_case_number(row['Case Number'])
                if case_num is not None and case_num in prev_assignments:
                    df_copy.at[idx, 'Assigned To'] = prev_assignments[case_num]
            
            # Calculate current workload
            current_workload = {}
            for handler in handlers:
                count = len(df_copy[df_copy['Assigned To'] == handler])
                current_workload[handler] = count
            
            self.logger.info(f"Current workload before additional balancing: {current_workload}")
            
            # Find handlers that are significantly overloaded and underloaded
            avg_workload = sum(current_workload.values()) / len(current_workload)
            overloaded_threshold = avg_workload * 1.15  # 15% above average
            underloaded_threshold = avg_workload * 0.85  # 15% below average
            
            overloaded_handlers = [h for h, w in current_workload.items() if w > overloaded_threshold]
            underloaded_handlers = [h for h, w in current_workload.items() if w < underloaded_threshold]
            
            self.logger.info(f"Overloaded handlers (> {overloaded_threshold:.1f}): {overloaded_handlers}")
            self.logger.info(f"Underloaded handlers (< {underloaded_threshold:.1f}): {underloaded_handlers}")
            
            # Move cases from overloaded to underloaded handlers (ONLY NEW CASES)
            moved_count = 0
            
            for overloaded_handler in overloaded_handlers:
                if not underloaded_handlers:
                    break
                
                # Find ONLY NEW cases assigned to this overloaded handler
                new_cases_mask = ~df_copy['Case Number'].isin(prev_assignments)
                overloaded_cases = df_copy[(df_copy['Assigned To'] == overloaded_handler) & new_cases_mask]
                
                self.logger.info(f"Found {len(overloaded_cases)} new cases for {overloaded_handler} that can be moved")
                
                if not overloaded_cases.empty:
                    # Sort cases by company to maintain consistency
                    company_col = self._get_company_column(df_copy)
                    if company_col and company_col in df_copy.columns:
                        overloaded_cases['Company_Lower'] = overloaded_cases[company_col].fillna('').astype(str).str.strip().str.lower()
                        overloaded_cases = overloaded_cases.sort_values('Company_Lower')
                    
                    # Move cases to underloaded handlers
                    for idx, row in overloaded_cases.iterrows():
                        if not underloaded_handlers:
                            break
                        
                        # Find the most underloaded handler
                        target_handler = min(underloaded_handlers, key=lambda h: current_workload[h])
                        
                        # Only move if it's not in previous assignments
                        case_num = self.normalize_case_number(row['Case Number'])
                        if case_num is not None and case_num in prev_assignments:
                            continue
                        
                        # Move the case
                        old_handler = df_copy.at[idx, 'Assigned To']
                        df_copy.at[idx, 'Assigned To'] = target_handler
                        current_workload[overloaded_handler] -= 1
                        current_workload[target_handler] += 1
                        moved_count += 1
                        
                        self.logger.info(f"Moved case {case_num} from {old_handler} to {target_handler}")
                        
                        # Remove target handler if it's no longer underloaded
                        if current_workload[target_handler] >= underloaded_threshold:
                            underloaded_handlers.remove(target_handler)
                            self.logger.info(f"Handler {target_handler} no longer underloaded, removing from list")
                        
                        # Update the workload tracking
                        self.logger.debug(f"Updated workload: {current_workload}")
            
            self.logger.info(f"Additional balancing completed: {moved_count} cases moved")
            
            # Verify that no preserved assignments were corrupted
            preserved_corrupted = []
            for idx, row in df_copy.iterrows():
                case_num = self.normalize_case_number(row['Case Number'])
                if case_num is not None and case_num in prev_assignments:
                    expected_handler = prev_assignments[case_num]
                    actual_handler = row['Assigned To']
                    if expected_handler != actual_handler:
                        preserved_corrupted.append((case_num, expected_handler, actual_handler))
            
            if preserved_corrupted:
                self.logger.error(f"CRITICAL ERROR: {len(preserved_corrupted)} preserved assignments corrupted during additional balancing!")
                for case_num, expected, actual in preserved_corrupted:
                    self.logger.error(f"  {case_num}: expected {expected}, got {actual}")
                    # Restore immediately
                    case_idx = df_copy[df_copy['Case Number'] == case_num].index
                    if len(case_idx) > 0:
                        df_copy.at[case_idx[0], 'Assigned To'] = expected
                        self.logger.info(f"Immediately restored corrupted assignment: {case_num} -> {expected}")
            else:
                self.logger.info("SUCCESS: All preserved assignments maintained during additional balancing")
            
            # Final workload calculation
            final_workload = {}
            for handler in handlers:
                count = len(df_copy[df_copy['Assigned To'] == handler])
                final_workload[handler] = count
            
            self.logger.info(f"Final workload after additional balancing: {final_workload}")
            
            return df_copy
            
        except Exception as e:
            self.logger.error(f"Error in additional balancing: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return df




class FinalProcessor:
    def __init__(self):
        # Configure logger (use global helpers.setup_logging where possible)
        self.logger = logging.getLogger(__name__)
        self.validation_lists = {
            "Action 1": ["Sent SMS", "Handled by Case Owner", "Handled by DRO", "Handled by Chat", "Bank/Sutherland", "DND"],
            "Action 2": ["Sent Email", "Sent Email in another Case", "Bank/Sutherland", "DND"],
            "Action 3": ["Called the Customer", "Called in another Case", "Bank/Sutherland", "DND"],
            "Final Action": ["Fixed", "Refused Callback", "Issue Not Fixed", "Not yet Tested", "Escalation", "Sent SMS", "Sent Email", "Left VM", "Reviewed", "Bank/Sutherland", "Not Reached", "DND"],
            "Status": ["in_progress", "new", "closed", "Skipped", "In Progress Today"]
        }
        
        # Column order for Companies sheet - as specified by user
        self.companies_column_order = [
            'Case Number', 'Customer Name', 'Company Name', 'DND (Do Not Disturb)', 
            'Email', 'Phone Number', 'Work Order Type', 'Incoming Channel', 
            'Last Status Change', 'Country', 'State/Province', 'Local Time', 
            'Action 1', 'Action 2', 'Action 3', 'Final Action', 'Assigned To', 
            'Status', 'Unnamed: 18', 'Case', 'Problem Description', 
            'Case Status Updated', 'Case Reason', 'Work Order ID', 'Work Order Status', 
            'Order Type', 'Work Order Priority', 'Product ID (MTM)', 'Machine Type', 
            'Product Description', 'Serial Number', 'Created On', 'Survey Preference', 
            'Survey Fatigue', 'No survey reason', 'Program', 'Repeat Frequency', 
            'Repeat Repair', 'Closing Code', 'Reported Symptom', 'Case Status'
        ]

    def _resolve_col(self, df, preferred_name):
        """Resolve a preferred column name to an existing column in df.
        - If exact match (case-insensitive) exists, return it.
        - Otherwise, return the first column that contains the main token (e.g. 'customer') in its name.
        - If nothing found, return None.
        """
        if df is None:
            return None
        pref = preferred_name.strip().lower()
        # Exact match (case-insensitive)
        for col in df.columns:
            try:
                if str(col).strip().lower() == pref:
                    return col
            except Exception:
                continue
        # Token match (use the first word as token)
        token = pref.split()[0]
        for col in df.columns:
            try:
                if token in str(col).strip().lower():
                    return col
            except Exception:
                continue
        return None

    def process_final_output(self, output_df, output_file, processing_stats=None, sms_file=None, email_file=None, prev_file=None, duplicate_company_cases=None, selected_handlers=None, chat_agent_info=None):
        try:
            self.logger.info("\n=== Starting Final Processing ===")
            self.logger.info(f"Processing final output with {len(output_df)} records")
            
            # Use the previous input file (prev_file) to load Issue Not Fixed and DND Emails
            prev_output_file = None
            
            self.logger.info(f"=== Using prev_file to load Issue Not Fixed and DND Emails ===")
            self.logger.info(f"Previous file provided: {prev_file}")
            
            if prev_file and os.path.exists(prev_file):
                # Check if this file has the expected sheets
                self.logger.info(f"Checking prev_file for Issue Not Fixed and DND Emails sheets: {prev_file}")
                try:
                    with pd.ExcelFile(prev_file) as excel:
                        sheet_names = excel.sheet_names
                        self.logger.info(f"prev_file has these sheets: {sheet_names}")
                        
                        if 'Issue Not Fixed' in sheet_names or 'DND Emails' in sheet_names:
                            # This file contains the sheets we need
                            prev_output_file = prev_file
                            self.logger.info(f" Setting prev_output_file to: {prev_output_file}")
                            
                            # Verify what will be loaded and log the counts
                            if 'Issue Not Fixed' in sheet_names:
                                prev_issue = pd.read_excel(prev_file, sheet_name='Issue Not Fixed')
                                self.logger.info(f"  Issue Not Fixed: {len(prev_issue)} cases will be loaded")
                                if len(prev_issue) > 0:
                                    self.logger.info(f"    Sample cases: {prev_issue[['Case Number']].head(3).to_dict('records')}")
                            if 'DND Emails' in sheet_names:
                                prev_dnd = pd.read_excel(prev_file, sheet_name='DND Emails')
                                self.logger.info(f"  DND Emails: {len(prev_dnd)} emails will be loaded")
                        else:
                            self.logger.info(f"prev_file does not contain Issue Not Fixed or DND Emails sheets")
                            self.logger.info(f"This is likely a temp handler file - will search for actual output file")
                except Exception as e:
                    self.logger.warning(f"Could not check prev_file: {str(e)}")
            
            # If prev_file doesn't have the sheets, search for actual output file in the same directory
            if not prev_output_file:
                self.logger.info(f"Searching for actual previous output file in directory...")
                output_dir = os.path.dirname(output_file) or os.path.dirname(prev_file) if prev_file else None
                
                if output_dir and os.path.exists(output_dir):
                    try:
                        # Look for files in the same directory
                        possible_files = []
                        for file in os.listdir(output_dir):
                            if file.endswith('.xlsx') and file != os.path.basename(output_file) if output_file else file:
                                file_path = os.path.join(output_dir, file)
                                # Skip temp handler files
                                if '_prev_handlers' not in file:
                                    try:
                                        with pd.ExcelFile(file_path) as excel:
                                            if 'Issue Not Fixed' in excel.sheet_names or 'DND Emails' in excel.sheet_names:
                                                mod_time = os.path.getmtime(file_path)
                                                possible_files.append((file_path, mod_time, file))
                                                self.logger.info(f"Found candidate: {file}")
                                    except Exception:
                                        continue
                        
                        if possible_files:
                            # Sort by modification time and use the most recent
                            possible_files.sort(key=lambda x: x[1], reverse=True)
                            prev_output_file = possible_files[0][0]
                            self.logger.info(f"✓ Selected previous output file: {prev_output_file}")
                            
                            # Load and log counts
                            try:
                                with pd.ExcelFile(prev_output_file) as excel:
                                    if 'Issue Not Fixed' in excel.sheet_names:
                                        prev_issue = pd.read_excel(prev_output_file, sheet_name='Issue Not Fixed')
                                        self.logger.info(f"  Issue Not Fixed: {len(prev_issue)} cases will be loaded")
                                        if len(prev_issue) > 0:
                                            self.logger.info(f"    Sample: {prev_issue['Case Number'].head(3).tolist()}")
                                    if 'DND Emails' in excel.sheet_names:
                                        prev_dnd = pd.read_excel(prev_output_file, sheet_name='DND Emails')
                                        self.logger.info(f"  DND Emails: {len(prev_dnd)} emails will be loaded")
                            except Exception as e:
                                self.logger.warning(f"Could not verify contents: {str(e)}")
                        else:
                            self.logger.info("No previous output files found in directory")
                    except Exception as e:
                        self.logger.warning(f"Could not search directory: {str(e)}")
            
            # Final log
            if prev_output_file:
                self.logger.info(f"=== Will load Issue Not Fixed and DND Emails from: {prev_output_file} ===")
            else:
                self.logger.info(f"=== No previous file with Issue Not Fixed/DND Emails found - starting fresh ===")
            
            # STEP 1: Load and consolidate ALL previous handler sheets FIRST
            consolidated_prev_data, issue_not_resolved_cases = None, []
            if prev_file and os.path.exists(prev_file):
                try:
                    self.logger.info("=== STEP 1: LOADING AND CONSOLIDATING PREVIOUS HANDLER SHEETS ===")
                    consolidated_prev_data, issue_not_resolved_cases = self.load_and_consolidate_all_previous_handler_sheets(prev_file)
                    if consolidated_prev_data is not None and not consolidated_prev_data.empty:
                        self.logger.info(f"Successfully consolidated {len(consolidated_prev_data)} previous cases from all handler sheets")
                        self.logger.info(f"  Previous handlers found: {consolidated_prev_data['Assigned To'].unique().tolist()}")
                    else:
                        self.logger.info("No previous handler data found - starting fresh")
                except Exception as e:
                    self.logger.error(f"Error loading previous handler data: {str(e)}")
                    self.logger.warning("Continuing without previous data")
                    consolidated_prev_data = None

            # STEP 2: Merge current data with consolidated previous data
            if consolidated_prev_data is not None and not consolidated_prev_data.empty:
                self.logger.info("=== STEP 2: MERGING CURRENT DATA WITH CONSOLIDATED PREVIOUS DATA ===")
                # Merge current output_df with consolidated previous data
                merged_output_df = self.merge_current_with_consolidated_previous(output_df, consolidated_prev_data)
                self.logger.info(f"Merged data: {len(output_df)} current + {len(consolidated_prev_data)} previous = {len(merged_output_df)} total")
                output_df = merged_output_df  # Use merged data for all subsequent processing
            else:
                self.logger.info("No previous data to merge - using current data only")

            # STEP 3: Apply business rules (only specific updates, preserve handler work)
            self.logger.info("=== STEP 3: APPLYING BUSINESS RULES (PRESERVING HANDLER WORK) ===")
            output_df = self.apply_business_rules_preserving_handler_work(output_df)

            # STEP 3.5: Sort the main output DataFrame by status
            self.logger.info("=== STEP 3.5: SORTING MAIN OUTPUT BY STATUS ===")
            output_df = self.sort_cases_by_status(output_df)
            
            # STEP 3.6: Add preservation columns to main output for case tracking
            # Removed adding preservation columns; preservation will be handled during per-handler merging
            
            # Create sorting summary report
            self.create_sorting_summary_report(output_df, "PA Cases Main Sheet")

            # Create Excel file with xlsxwriter
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                # STEP 4: Create sheets in the CORRECT ORDER
                self.logger.info("=== STEP 4: CREATING SHEETS IN CORRECT ORDER ===")
                
                # 4.1: Create individual handler sheets FIRST (each handler gets their own sheet)
                self.logger.info("=== 4.1: CREATING INDIVIDUAL HANDLER SHEETS ===")
                self.create_handler_sheets_from_processed_data(writer, output_df, prev_file=prev_file, chat_agent_info=chat_agent_info)

                # 4.1.2: Create Chat Agent's Cases sheet (if Chat Agent enabled)
                if chat_agent_info and chat_agent_info.get('enabled'):
                    self.logger.info("=== 4.1.2: CREATING CHAT AGENT'S CASES SHEET ===")
                    self.create_chat_agent_sheet_output(writer, output_df, chat_agent_info, prev_file=prev_file)

                # 4.1.5: Create Companies sheet (email duplicates - preserved across runs)
                # Positioned right after handler sheets as requested
                self.logger.info("=== 4.1.5: CREATING COMPANIES SHEET ===")
                companies_df = self.create_companies_sheet_with_preservation(
                    duplicate_company_cases, 
                    prev_file, 
                    output_df,
                    selected_handlers=selected_handlers
                )
                
                if companies_df is not None and not companies_df.empty:
                    companies_df.to_excel(writer, sheet_name='Companies', index=False)
                    self.auto_adjust_columns(writer, companies_df, 'Companies')
                    # Lock the sheet
                    self.protect_worksheet(writer, 'Companies', password='artadmin')
                    self.logger.info(f"✓ Created Companies sheet with {len(companies_df)} total cases")
                    
                    # CRITICAL: Sync Companies assignments back to PA Cases output_df
                    # This ensures company cases show their assigned handler in PA Cases sheet too
                    if 'Case Number' in companies_df.columns and 'Assigned To' in companies_df.columns:
                        self.logger.info("Syncing Companies assignments to PA Cases...")
                        companies_assignments = {}
                        for _, row in companies_df.iterrows():
                            case_num = row['Case Number']
                            handler = row.get('Assigned To', '')
                            if pd.notna(case_num) and handler and str(handler).strip():
                                # Normalize case number for matching
                                try:
                                    norm_case = int(float(str(case_num).strip()))
                                    companies_assignments[norm_case] = handler
                                except:
                                    pass
                        
                        # Update output_df with Companies assignments
                        if companies_assignments:
                            output_df['_temp_case_num'] = output_df['Case Number'].apply(
                                lambda x: int(float(str(x).strip())) if pd.notna(x) and str(x).strip() else None
                            )
                            updated_count = 0
                            for case_num, handler in companies_assignments.items():
                                mask = output_df['_temp_case_num'] == case_num
                                if mask.any():
                                    output_df.loc[mask, 'Assigned To'] = handler
                                    updated_count += 1
                            output_df = output_df.drop(columns=['_temp_case_num'])
                            self.logger.info(f"✓ Synced {updated_count} company case assignments to PA Cases")
                else:
                    self.logger.info("No email duplicate cases found - skipping Companies sheet")
                
                # 4.2: Create Counter sheet
                self.logger.info("=== 4.2: CREATING COUNTER SHEET ===")
                self.create_counters_sheet(writer, output_df, prev_file=prev_file)
                
                # 4.3: Create Summary sheet
                self.logger.info("=== 4.3: CREATING SUMMARY SHEET ===")
                self.create_summary_sheet(writer, output_df, processing_stats)
                
                # 4.4: Create Issue Not Fixed sheet
                self.logger.info("=== 4.4: CREATING ISSUE NOT FIXED SHEET ===")
                self.logger.info(f"Passing prev_output_file to Issue Not Fixed: {prev_output_file}")
                if prev_output_file:
                    self.logger.info(f"  ✓ prev_output_file is set and will be used to load previous data")
                else:
                    self.logger.warning(f"  ✗ prev_output_file is None - no previous data will be loaded!")
                self.create_sms_replies_sheet(writer, output_df, sms_file, processing_stats, prev_output_file, issue_not_resolved_cases)
                
                # 4.4.1: Create Skipped SMS Cases sheet
                self.logger.info("=== 4.4.1: CREATING SKIPPED SMS CASES SHEET ===")
                self.create_skipped_sms_sheet(writer, processing_stats)
                
                # 4.4.2: Create Skipped Email Replies sheet
                self.logger.info("=== 4.4.2: CREATING SKIPPED EMAIL REPLIES SHEET ===")
                # Pass email_file and current output_df so the skipped-email sheet can be built
                self.create_skipped_email_sheet(writer, processing_stats, email_file=email_file, output_df=output_df)
                
                # 4.5: Create DND Emails Database sheet
                self.logger.info("=== 4.5: CREATING DND EMAILS DATABASE SHEET ===")
                self.logger.info(f"Passing prev_output_file to DND Emails: {prev_output_file}")
                if prev_output_file:
                    self.logger.info(f"  ✓ prev_output_file is set and will be used to load previous data")
                else:
                    self.logger.warning(f"  ✗ prev_output_file is None - no previous data will be loaded!")
                self.create_dnd_emails_database_sheet(writer, output_df, email_file, processing_stats, prev_file=prev_output_file)
                
                # 4.6: Create Validation sheet (empty for now, will be populated later)
                self.logger.info("=== 4.6: CREATING VALIDATION SHEET ===")
                validation_df = pd.DataFrame(columns=['Validation'])
                validation_df.to_excel(writer, sheet_name='Validation', index=False)
                
                # Auto-adjust columns for Validation sheet
                self.auto_adjust_columns(writer, validation_df, 'Validation')
                

                
                # 4.7: Create PA Cases main sheet LAST (already sorted by status)
                self.logger.info("=== 4.7: CREATING PA CASES MAIN SHEET ===")
                
                # Format date columns to show only dates (remove timestamps)
                # Convert to string format with date only
                date_columns = ['Last Status Change', 'Created On', 'Completion Date']
                for col in date_columns:
                    if col in output_df.columns:
                        try:
                            # Debug: show before
                            sample_before = output_df[col].iloc[0] if len(output_df) > 0 else None
                            before_str = str(sample_before)
                            print(f"CONSOLE DEBUG {col} BEFORE: {repr(sample_before)} -> str: {before_str}")
                            self.logger.info(f"DEBUG {col} BEFORE: {repr(sample_before)}")
                            
                            def extract_date_str(val):
                                """Extract date part as clean string - handles both strings and datetime objects"""
                                if pd.isna(val) or val == '' or val == 'NaT':
                                    return ''
                                
                                # Handle pandas Timestamp and datetime objects
                                if hasattr(val, 'date'):  # datetime, Timestamp, date objects have .date() method
                                    return str(val.date())  # Returns YYYY-MM-DD
                                
                                # Handle strings
                                val_str = str(val).strip()
                                if not val_str or val_str.lower() in ('nan', 'nat', 'none'):
                                    return ''
                                # If it contains a space, extract date part only
                                if ' ' in val_str:
                                    date_only = val_str.split(' ')[0]
                                    return date_only
                                return val_str
                            
                            output_df[col] = output_df[col].apply(extract_date_str)
                            
                            # Debug: show after
                            sample_after = output_df[col].iloc[0] if len(output_df) > 0 else None
                            self.logger.info(f"DEBUG {col} AFTER: {repr(sample_after)}")
                            self.logger.info(f"✓ Converted {col} to date-only strings")
                        except Exception as e:
                            self.logger.warning(f"Could not format {col}: {str(e)}")
                
                output_df.to_excel(writer, sheet_name='PA Cases', index=False)
                
                # Auto-adjust columns for PA Cases sheet
                self.auto_adjust_columns(writer, output_df, 'PA Cases')
            
            # STEP 5: Add validation dropdowns using openpyxl (this will populate the 'Validation' sheet)
            self.add_validation_dropdowns(output_file)
            
            # STEP 6: Final date formatting after validation dropdowns - ALL SHEETS
            self.logger.info("=== STEP 6: FINAL DATE FORMATTING AFTER VALIDATION (ALL SHEETS) ===")
            try:
                from openpyxl import load_workbook
                from datetime import datetime
                
                wb = load_workbook(output_file)
                
                date_columns_to_fix = ['Last Status Change', 'Created On', 'Completion Date']
                
                # Apply formatting to all sheets
                for sheet_name in wb.sheetnames:
                    try:
                        ws = wb[sheet_name]
                        
                        # Skip non-data sheets
                        if sheet_name in ['Validation', 'Counter', 'Summary']:
                            continue
                        
                        # Get header row
                        header = [cell.value for cell in ws[1]]
                        
                        # Format each date column that exists in this sheet
                        for col_name in date_columns_to_fix:
                            if col_name in header:
                                col_idx = header.index(col_name) + 1
                                cell_count = 0
                                # Process each cell - convert datetime to date-only string
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row, col_idx)
                                    if cell.value:
                                        try:
                                            # Convert to datetime first
                                            if isinstance(cell.value, str):
                                                # Parse the string
                                                dt = pd.to_datetime(cell.value, errors='coerce')
                                                if pd.notna(dt):
                                                    date_str = dt.strftime('%Y-%m-%d')
                                                    cell.value = date_str
                                                    cell.number_format = '@'  # Text format
                                                    cell_count += 1
                                            elif hasattr(cell.value, 'date'):  # datetime object
                                                # Extract just the date
                                                date_str = cell.value.date().isoformat()  # Returns YYYY-MM-DD
                                                cell.value = date_str
                                                cell.number_format = '@'  # Text format
                                                cell_count += 1
                                        except:
                                            pass  # Skip cells that can't be converted
                                
                                if cell_count > 0:
                                    self.logger.info(f"✓ Sheet '{sheet_name}': Fixed {cell_count} cells in '{col_name}'")
                    except Exception as e:
                        self.logger.warning(f"Could not format sheet '{sheet_name}': {str(e)}")
                        continue
                
                wb.save(output_file)
                self.logger.info("✓ Saved Excel file with timestamp-free dates across all sheets")
            except Exception as e:
                self.logger.warning(f"Could not apply date formatting to all sheets: {str(e)}")
            
            # STEP 7: Delete temporary _prev_handlers file if it exists
            if prev_file and '_prev_handlers' in prev_file and os.path.exists(prev_file):
                try:
                    os.remove(prev_file)
                    self.logger.info(f"✓ Deleted temporary handler file: {prev_file}")
                except Exception as e:
                    self.logger.warning(f"Could not delete temporary file {prev_file}: {str(e)}")
            
            self.logger.info(f"Final processing completed. File saved to: {output_file}")
            return True, "Final processing completed successfully"
        except Exception as e:
            self.logger.error(f"Error in final processing: {str(e)}")
            return False, f"Error during final processing: {str(e)}"

    def load_and_consolidate_all_previous_handler_sheets(self, prev_file):
        """Load and consolidate ALL previous handler sheets into one dataset for processing"""
        try:
            self.logger.info(f"Loading and consolidating all handler sheets from: {prev_file}")
            
            # Read the previous file
            prev_excel = pd.ExcelFile(prev_file)
            all_sheets = prev_excel.sheet_names
            self.logger.info(f"Found sheets in previous file: {all_sheets}")
            
            # Look for handler sheets (sheets ending with "'s Cases") - guard non-string sheet names
            handler_sheets = [sheet for sheet in all_sheets if isinstance(sheet, str) and sheet.endswith("'s Cases")]
            self.logger.info(f"Found handler sheets: {handler_sheets}")
            
            if not handler_sheets:
                self.logger.info("No handler sheets found in previous file")
                return None, []
            
            # Consolidate all handler sheets into one DataFrame
            consolidated_data = []
            total_cases = 0
            issue_not_resolved_cases = []
            
            for sheet_name in handler_sheets:
                try:
                    # Extract handler name from sheet name
                    handler_name = sheet_name.replace("'s Cases", "")
                    
                    # Read the handler sheet
                    handler_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                    
                    if not handler_df.empty and 'Case Number' in handler_df.columns:
                        # Optional: any health checks can remain here
                        # Check for cases marked as "Issue Not Resolved" in Final Action
                        if 'Final Action' in handler_df.columns:
                            issue_mask = handler_df['Final Action'].astype(str).str.strip().str.lower().isin([
                                'issue not resolved', 'issue not fixed', 'not resolved', 'unresolved'
                            ])
                            issue_cases = handler_df[issue_mask]
                            
                            if not issue_cases.empty:
                                for idx, row in issue_cases.iterrows():
                                    issue_case = {
                                        'Case Number': row['Case Number'],
                                        'Handler': handler_name,
                                        'Source Sheet': sheet_name,
                                        'Final Action': row['Final Action'],
                                        'Status': row.get('Status', ''),
                                        'Assigned To': row.get('Assigned To', ''),
                                        'Date Added': pd.Timestamp.now().strftime('%Y-%m-%d'),
                                        'Source': 'Handler Manual Update'
                                    }
                                    issue_not_resolved_cases.append(issue_case)
                                
                                self.logger.info(f"Found {len(issue_cases)} cases marked as 'Issue Not Resolved' in {sheet_name}")
                        
                        consolidated_data.append(handler_df)
                        total_cases += len(handler_df)
                        self.logger.info(f"Loaded handler sheet '{sheet_name}': {len(handler_df)} cases")
                        
                        # Log some sample case numbers for verification
                        if len(handler_df) > 0:
                            sample_cases = handler_df['Case Number'].head(3).tolist()
                            self.logger.info(f"  Sample case numbers: {sample_cases}")
                    else:
                        self.logger.warning(f"Handler sheet '{sheet_name}' is empty or missing Case Number column")
                        
                except Exception as e:
                    self.logger.warning(f"Error loading handler sheet '{sheet_name}': {str(e)}")
                    continue
            
            if not consolidated_data:
                self.logger.warning("No valid handler data could be loaded")
                return None, []
            
            # Combine all handler data into one DataFrame
            if len(consolidated_data) == 1:
                final_consolidated = consolidated_data[0]
            else:
                final_consolidated = pd.concat(consolidated_data, ignore_index=True)
            
            self.logger.info(f"Successfully consolidated {len(handler_sheets)} handler sheets with {total_cases} total cases")
            self.logger.info(f"Found {len(issue_not_resolved_cases)} cases marked as 'Issue Not Resolved' across all handler sheets")
            self.logger.info(f"Consolidated data columns: {final_consolidated.columns.tolist()}")
            
            # Additional verification logging
            self.logger.info(f"Final consolidated DataFrame shape: {final_consolidated.shape}")
            self.logger.info(f"Case Number column unique values: {final_consolidated['Case Number'].nunique()}")
            
            return final_consolidated, issue_not_resolved_cases
            
        except Exception as e:
            self.logger.error(f"Error consolidating previous handler sheets: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return None, []

    def create_companies_sheet_with_preservation(self, new_duplicates, prev_file, current_df, selected_handlers=None):
        """
        Create- [/] **Companies Sheet Logic** <!-- id: 11 -->
    - [x] Update `processor.py` to pass `selected_handlers` <!-- id: 12 -->
    - [x] Implement Sorting Logic in `final_processor.py` (Group matching cases) <!-- id: 13 -->
    - [x] Implement Rounds Logic (Assignment by email) in `final_processor.py` <!-- id: 14 -->
    - [x] Implement Locking Logic (Password: 'artadmin', View-only) in `final_processor.py` <!-- id: 15 -->Email groups
        
        Args:
            new_duplicates: DataFrame of new cases with duplicate emails
            prev_file: Path to previous output file (for loading existing Companies data)
            current_df: Current processed DataFrame (for applying business rules)
            selected_handlers: List of handlers to assign cases to
        
        Returns:
            DataFrame ready for Companies sheet with proper column order
        """
        try:
            self.logger.info("Creating Companies sheet with preservation...")
            
            # Step 1: Load previous Companies sheet
            prev_companies = self._load_previous_companies_sheet(prev_file)
            
            # Step 2: Merge previous with new duplicates
            if new_duplicates is None or new_duplicates.empty:
                new_duplicates = pd.DataFrame()
            
            companies_df = self._merge_companies_data(new_duplicates, prev_companies, prev_file)
            
            if companies_df.empty:
                self.logger.info("No company cases to include in Companies sheet")
                return pd.DataFrame()
            
            # Step 3: Apply business rules (same as handler processing) + DND rules
            companies_df = self._apply_business_rules_to_companies(companies_df, prev_file)
            
            # Step 4: Assign cases to handlers (Round Robin by Email)
            companies_df = self._assign_companies_cases(companies_df, selected_handlers)
            
            # Step 5: Reorder columns to match specified order
            companies_df = self._reorder_companies_columns(companies_df)
            
            # Step 6: Sort: New first, then others
            companies_df = self._sort_companies_sheet(companies_df)
            
            self.logger.info(f"Companies sheet prepared with {len(companies_df)} total cases")
            return companies_df
            
        except Exception as e:
            self.logger.error(f"Error creating Companies sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            # Return new duplicates if available, otherwise empty
            return new_duplicates if new_duplicates is not None else pd.DataFrame()

    def _load_previous_companies_sheet(self, prev_file):
        """Load previous Companies sheet data for preservation"""
        try:
            if not prev_file or not os.path.exists(prev_file):
                self.logger.info("No previous file provided - starting fresh Companies sheet")
                return pd.DataFrame()
            
            with pd.ExcelFile(prev_file) as excel:
                if 'Companies' in excel.sheet_names:
                    prev_companies = pd.read_excel(prev_file, sheet_name='Companies')
                    self.logger.info(f"Loaded {len(prev_companies)} cases from previous Companies sheet")
                    
                    # Log sample case numbers for verification
                    if len(prev_companies) > 0 and 'Case Number' in prev_companies.columns:
                        sample = prev_companies['Case Number'].head(3).tolist()
                        self.logger.info(f"  Sample preserved cases: {sample}")
                    
                    return prev_companies
                else:
                    self.logger.info("No Companies sheet in previous file - starting fresh")
                    return pd.DataFrame()
                    
        except Exception as e:
            self.logger.warning(f"Could not load previous Companies sheet: {str(e)}")
            return pd.DataFrame()

    def _merge_companies_data(self, new_duplicates, prev_companies, prev_file=None):
        """
        Merge new email duplicate cases with preserved Companies data.
        
        CRITICAL: PREVIOUS DATA WINS FOR OVERLAPPING CASES
        - Overlapping cases (same Case Number): KEEP previous data (actions, status, handler preserved)
        - Non-overlapping previous cases: PRESERVE them as-is
        - Truly new cases (not in previous): ADD them for assignment
        
        Adds '_is_preserved' flag to identify which cases should skip business rules.
        """
        try:
            self.logger.info("=== MERGING COMPANIES DATA (Previous Data Preserved) ===")
            
            # Normalize case numbers for comparison
            def normalize_case_num(val):
                if pd.isna(val) or str(val).strip() == '':
                    return None
                try:
                    return int(float(str(val).strip()))
                except:
                    return None
            
            # Handle empty cases
            if prev_companies.empty and (new_duplicates is None or new_duplicates.empty):
                self.logger.info("No previous Companies data and no new duplicates - returning empty")
                return pd.DataFrame()
            
            if new_duplicates is None or new_duplicates.empty:
                self.logger.info(f"No new duplicates - preserving ALL {len(prev_companies)} previous Companies cases")
                result = prev_companies.copy()
                result['_is_preserved'] = True  # Mark all as preserved
                return result
            
            if prev_companies.empty:
                self.logger.info(f"No previous Companies data - using {len(new_duplicates)} new cases only")
                result = new_duplicates.copy()
                result['_is_preserved'] = False  # Mark all as new
                return result
            
            # STEP 1: Normalize case numbers in BOTH dataframes
            prev_companies_copy = prev_companies.copy()
            new_duplicates_copy = new_duplicates.copy()
            
            prev_companies_copy['_normalized_cn'] = prev_companies_copy['Case Number'].apply(normalize_case_num)
            new_duplicates_copy['_normalized_cn'] = new_duplicates_copy['Case Number'].apply(normalize_case_num)
            
            # STEP 2: Get case number sets
            prev_case_numbers = set(prev_companies_copy['_normalized_cn'].dropna())
            new_case_numbers = set(new_duplicates_copy['_normalized_cn'].dropna())
            
            self.logger.info(f"Previous Companies cases: {len(prev_case_numbers)}")
            self.logger.info(f"New duplicate cases: {len(new_case_numbers)}")
            
            # STEP 3: Identify overlapping and truly new
            overlapping = prev_case_numbers & new_case_numbers
            truly_new = new_case_numbers - prev_case_numbers
            
            self.logger.info(f"  Overlapping (KEEP PREVIOUS data): {len(overlapping)}")
            self.logger.info(f"  Truly new (add from new): {len(truly_new)}")
            
            # STEP 4: Build the merged result - PREVIOUS DATA WINS
            # Keep ALL previous cases (they have actions, status, handler preserved)
            preserved_cases = prev_companies_copy.copy()
            preserved_cases['_is_preserved'] = True  # Mark as preserved - skip business rules
            
            # Get only TRULY NEW cases (not in previous at all)
            truly_new_cases = new_duplicates_copy[~new_duplicates_copy['_normalized_cn'].isin(prev_case_numbers)].copy()
            truly_new_cases['_is_preserved'] = False  # Mark as new - apply business rules
            
            self.logger.info(f"Final counts:")
            self.logger.info(f"  Preserved cases (ALL previous): {len(preserved_cases)}")
            self.logger.info(f"  Truly new cases (to add): {len(truly_new_cases)}")
            
            # Drop temp columns
            if '_normalized_cn' in preserved_cases.columns:
                preserved_cases = preserved_cases.drop(columns=['_normalized_cn'])
            if '_normalized_cn' in truly_new_cases.columns:
                truly_new_cases = truly_new_cases.drop(columns=['_normalized_cn'])
            
            # STEP 5: Combine - PREVIOUS FIRST (preserved with all data), then truly new
            if not truly_new_cases.empty:
                combined = pd.concat([preserved_cases, truly_new_cases], ignore_index=True)
                self.logger.info(f"Merged Companies: {len(preserved_cases)} preserved + {len(truly_new_cases)} truly new = {len(combined)} total")
            else:
                combined = preserved_cases
                self.logger.info(f"No truly new cases - keeping {len(combined)} preserved cases only")
            
            # STEP 6: Remove any duplicates (by Case Number) - keep first (previous has priority)
            if 'Case Number' in combined.columns:
                pre_dedup = len(combined)
                combined = combined.drop_duplicates(subset=['Case Number'], keep='first')
                post_dedup = len(combined)
                if pre_dedup > post_dedup:
                    self.logger.info(f"Removed {pre_dedup - post_dedup} duplicate entries")
            
            self.logger.info(f"=== COMPANIES MERGE COMPLETE: {len(combined)} total cases ===")
            return combined
            
        except Exception as e:
            self.logger.error(f"Error merging companies data: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            # Return non-empty data if available
            if prev_companies is not None and not prev_companies.empty:
                return prev_companies
            return new_duplicates if new_duplicates is not None and not new_duplicates.empty else pd.DataFrame()

    def _apply_business_rules_to_companies(self, df, prev_file=None):
        """Apply business rules to Companies sheet data, including DND checks
        
        IMPORTANT: Skip business rules for preserved cases (_is_preserved=True)
        Only apply rules to truly new cases to avoid overwriting existing data.
        """
        try:
            if df.empty:
                return df
            
            df = df.copy()
            self.logger.info("Applying business rules to Companies sheet...")
            
            # Check for _is_preserved flag
            has_preserved_flag = '_is_preserved' in df.columns
            if has_preserved_flag:
                preserved_count = df['_is_preserved'].sum() if df['_is_preserved'].dtype == bool else (df['_is_preserved'] == True).sum()
                new_count = len(df) - preserved_count
                self.logger.info(f"  Preserved cases (skipping rules): {preserved_count}, New cases (applying rules): {new_count}")
                # Create mask for NEW cases only (where we should apply rules)
                new_cases_mask = ~df['_is_preserved'].fillna(False).astype(bool)
            else:
                new_cases_mask = pd.Series([True] * len(df), index=df.index)
                self.logger.info("  No preservation flag found - applying rules to all cases")
            
            # Rule 1: Convert "in progress today" to "in_progress" (apply to ALL)
            if 'Status' in df.columns:
                in_progress_mask = df['Status'].astype(str).str.strip().str.lower() == 'in progress today'
                count = in_progress_mask.sum()
                if count > 0:
                    df.loc[in_progress_mask, 'Status'] = 'in_progress'
                    self.logger.info(f"  Updated {count} cases from 'in progress today' to 'in_progress'")
            
            # Rule 2: Apply Bank/Sutherland rule ONLY TO NEW CASES
            if 'Status' in df.columns and 'Action 1' in df.columns:
                bank_mask = (
                    new_cases_mask &  # Only new cases!
                    (df['Status'].astype(str).str.lower() == 'closed') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                count = bank_mask.sum()
                if count > 0:
                    df.loc[bank_mask, 'Action 1'] = 'Bank/Sutherland'
                    self.logger.info(f"  Applied Bank/Sutherland rule to {count} NEW cases")
            
            # Rule 3: Load DND Emails from DND Emails sheet and apply rules
            dnd_emails = set()
            try:
                # Load from previous file's DND Emails sheet
                if prev_file and os.path.exists(prev_file):
                    with pd.ExcelFile(prev_file) as excel:
                        if 'DND Emails' in excel.sheet_names:
                            dnd_sheet = pd.read_excel(prev_file, sheet_name='DND Emails')
                            if 'Email' in dnd_sheet.columns:
                                dnd_emails = set(dnd_sheet['Email'].dropna().astype(str).str.strip().str.lower())
                                self.logger.info(f"  Loaded {len(dnd_emails)} DND emails from previous DND Emails sheet")
            except Exception as e:
                self.logger.warning(f"Could not load DND Emails for application: {str(e)}")
            
            # Apply DND rule using loaded emails - ONLY TO NEW CASES
            if dnd_emails and 'Email' in df.columns:
                # Normalize emails in current df
                df['_email_check'] = df['Email'].astype(str).str.strip().str.lower()
                
                # Identify NEW cases with DND emails (skip preserved)
                dnd_matches = new_cases_mask & df['_email_check'].isin(dnd_emails)
                match_count = dnd_matches.sum()
                
                if match_count > 0:
                    self.logger.info(f"  Found {match_count} NEW cases matching DND emails")
                    
                    # Update fields
                    if 'Action 1' in df.columns:
                        df.loc[dnd_matches, 'Action 1'] = 'DND'
                    if 'Action 2' in df.columns:
                        df.loc[dnd_matches, 'Action 2'] = 'DND'
                    if 'Action 3' in df.columns:
                        df.loc[dnd_matches, 'Action 3'] = 'DND'
                    if 'Final Action' in df.columns:
                        df.loc[dnd_matches, 'Final Action'] = 'DND'
                    if 'Status' in df.columns:
                        df.loc[dnd_matches, 'Status'] = 'closed'
                
                # Clean up temp column
                if '_email_check' in df.columns:
                    df = df.drop(columns=['_email_check'])
            
            # Application of "DND" status check (if manually set to DND status) - ONLY NEW CASES
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_mask = (
                    new_cases_mask &  # Only new cases!
                    (df['Status'].astype(str).str.lower() == 'dnd') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                count = dnd_mask.sum()
                if count > 0:
                    df.loc[dnd_mask, 'Action 1'] = 'DND'
                    self.logger.info(f"  Applied DND action to {count} NEW cases with Status='DND'")
            
            # Clean up _is_preserved flag - don't include in final output
            if '_is_preserved' in df.columns:
                df = df.drop(columns=['_is_preserved'])
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error applying business rules to Companies: {str(e)}")
            if '_email_check' in df.columns:
                df = df.drop(columns=['_email_check'])
            if '_is_preserved' in df.columns:
                df = df.drop(columns=['_is_preserved'])
            return df

    def _sort_companies_sheet(self, df):
        """Sort Companies sheet: New first, then by Email to group cases"""
        try:
            if df.empty or 'Status' not in df.columns:
                return df
            
            self.logger.info("Sorting Companies sheet (New first, then by Email)...")
            
            # Create a sort key
            # 0 for New/new, 1 for others
            df['_sort_key'] = df['Status'].astype(str).str.strip().str.lower().apply(
                lambda x: 0 if x == 'new' else 1
            )
            
            # Primary sort: Status (New first)
            # Secondary sort: Assigned To (Group by handler)
            # Tertiary sort: Email (Group same emails within handler)
            # Quaternary sort: Case Number
            sort_cols = ['_sort_key']
            ascending = [True]
            
            if 'Assigned To' in df.columns:
                sort_cols.append('Assigned To')
                ascending.append(True)
            
            if 'Email' in df.columns:
                sort_cols.append('Email')
                ascending.append(True)
            
            if 'Case Number' in df.columns:
                sort_cols.append('Case Number')
                ascending.append(True)
                
            df = df.sort_values(by=sort_cols, ascending=ascending)
            
            # Drop sort key
            df = df.drop(columns=['_sort_key'])
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error sorting Companies sheet: {str(e)}")
            return df

    def _reorder_companies_columns(self, df):
        """Reorder Companies sheet columns to match the specified order"""
        try:
            if df.empty:
                return df
            
            self.logger.info("Reordering Companies sheet columns...")
            
            # Get existing columns
            existing_cols = list(df.columns)
            
            # Build ordered columns list: include only columns that exist in df
            ordered_cols = []
            for col in self.companies_column_order:
                if col in existing_cols:
                    ordered_cols.append(col)
            
            # Add any remaining columns that weren't in the specified order (at the end)
            for col in existing_cols:
                if col not in ordered_cols:
                    ordered_cols.append(col)
            
            # Reorder the DataFrame
            df = df[ordered_cols]
            
            self.logger.info(f"Columns reordered. Final column count: {len(df.columns)}")
            return df
            
        except Exception as e:
            self.logger.error(f"Error reordering Companies columns: {str(e)}")
            return df

    def merge_current_with_consolidated_previous(self, current_df, consolidated_prev_df):
        """Merge current data with consolidated previous handler data, preserving handler work"""
        try:
            self.logger.info("Merging current data with consolidated previous handler data...")
            
            if consolidated_prev_df is None or consolidated_prev_df.empty:
                self.logger.info("No previous data to merge - returning current data")
                return current_df
            
            # Ensure Case Number column exists in both DataFrames
            if 'Case Number' not in current_df.columns:
                self.logger.warning("No 'Case Number' column in current data - returning current data")
                return current_df
            
            if 'Case Number' not in consolidated_prev_df.columns:
                self.logger.warning("No 'Case Number' column in previous data - returning current data")
                return current_df
            
            # Normalize case numbers for comparison
            current_df = current_df.copy()
            consolidated_prev_df = consolidated_prev_df.copy()
            
            # Normalize to integer for robust comparisons
            current_df['Case Number'] = current_df['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            consolidated_prev_df['Case Number'] = consolidated_prev_df['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            
            # Get case numbers
            current_case_numbers = set(current_df['Case Number'])
            prev_case_numbers = set(consolidated_prev_df['Case Number'])
            
            # Remove placeholder values
            current_case_numbers.discard('UNKNOWN')
            prev_case_numbers.discard('UNKNOWN')
            
            # Identify cases
            existing_cases = current_case_numbers & prev_case_numbers  # Cases in both
            new_cases = current_case_numbers - prev_case_numbers      # New cases
            removed_cases = prev_case_numbers - current_case_numbers  # Cases only in previous file
            
            self.logger.info(f"Case analysis:")
            self.logger.info(f"  Existing cases to update: {len(existing_cases)}")
            self.logger.info(f"  New cases to add: {len(new_cases)}")
            self.logger.info(f"  Cases only in previous file: {len(removed_cases)}")
            
            # CRITICAL FIX: Start with ALL previous cases (don't drop any!)
            final_cases = consolidated_prev_df.copy()
            self.logger.info(f"Started with {len(final_cases)} cases from previous file")
            
            # Update existing cases with new data from raw file
            updated_count = 0
            for case_num in existing_cases:
                try:
                    # Get current case data (from new raw file)
                    current_case = current_df[current_df['Case Number'] == case_num].iloc[0]
                    
                    # Find the case in final_cases and update it
                    case_mask = final_cases['Case Number'] == case_num
                    if case_mask.any():
                        case_idx = case_mask.idxmax()
                        
                        # Update basic fields with new data from raw file
                        basic_fields = [
                            'Customer Name', 'Company Name', 'Email', 'Phone Number', 
                            'Work Order Type', 'Incoming Channel', 'Last Status Change',
                            'Country', 'State/Province', 'Assigned To'
                        ]

                        # Use flexible column resolution when updating fields (handles variants like 'Customer')
                        for field in basic_fields:
                            col_current = self._resolve_col(current_df, field)
                            col_final = self._resolve_col(final_cases, field)
                            if col_current and col_final and col_current in current_case.index and col_final in final_cases.columns:
                                if pd.notna(current_case[col_current]) and str(current_case[col_current]).strip() != '':
                                    final_cases.loc[case_idx, col_final] = current_case[col_current]
                        
                        # Handle Status field specially - only change "in progress today" to "in_progress"
                        if 'Status' in current_case.index and 'Status' in final_cases.columns:
                            current_status = str(current_case['Status']).strip().lower()
                            if current_status == 'in progress today':
                                final_cases.loc[case_idx, 'Status'] = 'in_progress'
                                self.logger.debug(f"Case {case_num}: Status updated from 'in progress today' to 'in_progress'")
                            else:
                                # Only update if current status is more recent/meaningful
                                prev_status = str(final_cases.loc[case_idx, 'Status']).strip().lower()
                                if prev_status in ['new', ''] and current_status not in ['new', '']:
                                    final_cases.loc[case_idx, 'Status'] = current_case['Status']
                                    self.logger.debug(f"Case {case_num}: Status updated from '{prev_status}' to '{current_case['Status']}'")
                        
                        # Handle Action fields - preserve handler work unless current is more specific
                        action_fields = ['Action 1', 'Action 2', 'Action 3', 'Final Action']
                        for field in action_fields:
                            if field in current_case.index and field in final_cases.columns:
                                current_value = current_case[field]
                                prev_value = final_cases.loc[case_idx, field]
                                
                                # Only update if current value exists and previous is empty
                                if (pd.notna(current_value) and 
                                    str(current_value).strip() != '' and 
                                    str(current_value).lower() not in ['nan', 'none', ''] and
                                        (_is_missing(prev_value))):
                                    
                                    final_cases.loc[case_idx, field] = current_value
                                    self.logger.debug(f"Case {case_num}: Updated {field} from empty to '{current_value}'")
                        
                        updated_count += 1
                        
                except Exception as e:
                    self.logger.warning(f"Error updating case '{case_num}': {str(e)}")
                    continue
            
            self.logger.info(f"Updated {updated_count} existing cases with new data from raw file")
            
            # Add new cases from raw file
            new_cases_df = current_df[current_df['Case Number'].isin(new_cases)].copy()
            if not new_cases_df.empty:
                # Clean possible internal columns if they existed in legacy files
                internal_cols = ['_Source_Sheet', '_Handler_Name']
                for col in internal_cols:
                    if col in new_cases_df.columns:
                        new_cases_df = new_cases_df.drop(columns=[col])
                
                final_cases = pd.concat([final_cases, new_cases_df], ignore_index=True)
                self.logger.info(f"Added {len(new_cases_df)} new cases from raw file")
            
            # CRITICAL FIX: Remove duplicates after merging
            self.logger.info("=== REMOVING DUPLICATES AFTER MERGE ===")
            pre_dedup_count = len(final_cases)
            
            # Sort by Created On date (newest first) before removing duplicates
            if 'Created On' in final_cases.columns:
                self.logger.info("   Sorting by 'Created On' to keep newest duplicate cases")
                try:
                    # Parse Created On to datetime for proper sorting
                    final_cases['_created_on_parsed'] = pd.to_datetime(final_cases['Created On'].astype(str).str.replace(r'\s+', ' ', regex=True), 
                                                                        format='%m/%d/%Y %I:%M:%S %p', errors='coerce')
                    
                    # If parsing fails, try alternative format
                    if final_cases['_created_on_parsed'].isna().any():
                        final_cases['_created_on_parsed'] = pd.to_datetime(final_cases['Created On'], errors='coerce')
                    
                    # Sort by the parsed date in descending order (newest first)
                    final_cases = final_cases.sort_values('_created_on_parsed', ascending=False, na_position='last')
                    self.logger.info("   Successfully sorted by Created On date")
                except Exception as e:
                    self.logger.warning(f"   Could not parse Created On dates for sorting: {str(e)}")
                    # Fallback: sort by original column as string
                    final_cases = final_cases.sort_values('Created On', ascending=False, na_position='last')
            
            final_cases = final_cases.drop_duplicates(subset=['Case Number'], keep='first')
            
            # Drop the temporary sorting column if it exists
            if '_created_on_parsed' in final_cases.columns:
                final_cases = final_cases.drop(columns=['_created_on_parsed'])
            
            post_dedup_count = len(final_cases)
            duplicates_removed = pre_dedup_count - post_dedup_count
            
            if duplicates_removed > 0:
                self.logger.warning(f"Removed {duplicates_removed} duplicate case entries")
                self.logger.warning(f"   This was causing cases to appear multiple times in the database")
            else:
                self.logger.info("No duplicates found after merge")
            
            # Sort by case number for consistency
            if not final_cases.empty:
                final_cases = final_cases.sort_values('Case Number').reset_index(drop=True)
            
            self.logger.info(f"Final merged dataset: {len(final_cases)} total cases")
            self.logger.info(f"  - Previous cases preserved: {len(consolidated_prev_df)}")
            self.logger.info(f"  - New cases added: {len(new_cases_df)}")
            self.logger.info(f"  - Duplicates removed: {duplicates_removed}")
            self.logger.info(f"  - Total after merge and deduplication: {len(final_cases)}")
            
            return final_cases
            
        except Exception as e:
            self.logger.error(f"Error merging current with consolidated previous: {str(e)}")
            # Fallback: return current data only
            return current_df

    def apply_business_rules_preserving_handler_work(self, df):
        """Apply business rules while preserving handler work - only update specific fields"""
        try:
            self.logger.info("Applying business rules while preserving handler work...")
            df = df.copy()
            
            # Rule 1: Convert "in progress today" to "in_progress" (only this specific change)
            if 'Status' in df.columns:
                in_progress_today_mask = df['Status'].astype(str).str.strip().str.lower() == 'in progress today'
                in_progress_today_count = in_progress_today_mask.sum()
                if in_progress_today_count > 0:
                    df.loc[in_progress_today_mask, 'Status'] = 'in_progress'
                    self.logger.info(f"Applied rule: {in_progress_today_count} cases changed from 'in progress today' to 'in_progress'")
            
            # Rule 2: Apply Bank/Sutherland rule (only if Status is 'closed' and Action 1 is empty)
            if 'Status' in df.columns and 'Action 1' in df.columns and 'Company Name' in df.columns:
                bank_sutherland_mask = (
                    (df['Status'].str.lower() == 'closed') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                bank_sutherland_count = bank_sutherland_mask.sum()
                if bank_sutherland_count > 0:
                    df.loc[bank_sutherland_mask, 'Action 1'] = 'Bank/Sutherland'
                    self.logger.info(f"Applied Bank/Sutherland rule: {bank_sutherland_count} cases updated")
            
            # Rule 3: Apply DND rule (only if specific conditions are met)
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_mask = (
                    (df['Status'].str.lower() == 'dnd') & 
                    (df['Action 1'].isna() | (df['Action 1'].astype(str).str.strip() == ''))
                )
                dnd_count = dnd_mask.sum()
                if dnd_count > 0:
                    df.loc[dnd_mask, 'Action 1'] = 'DND'
                    self.logger.info(f"Applied DND rule: {dnd_count} cases updated")
            
            self.logger.info("Business rules applied while preserving handler work")
            return df
            
        except Exception as e:
            self.logger.error(f"Error applying business rules: {str(e)}")
            return df

    def create_chat_agent_sheet_output(self, writer, output_df, chat_agent_info, prev_file=None):
        """Create and write the 'Chat Agent's Cases' sheet to the output Excel file
        
        Preservation Rules:
        - Load ALL cases from previous "Chat Agent's Cases" sheet (all statuses)
        - Load ALL current cases assigned to Chat Agent supporter
        - Merge: current cases first, then add any previous cases not in current
        - Sort by status
        
        Args:
            writer: pd.ExcelWriter object for writing to Excel
            output_df: Current processed DataFrame with assignments
            chat_agent_info: Chat Agent info dict with enabled status and supporter name
            prev_file: Path to previous file to load previous Chat Agent cases
        """
        try:
            if not chat_agent_info or not chat_agent_info.get('enabled'):
                self.logger.info("Chat Agent not enabled - skipping Chat Agent's Cases sheet")
                return
            
            chat_agent_supporter_name = chat_agent_info.get('supporter_name', 'Chat Agent')
            self.logger.info(f"\n=== CREATING CHAT AGENT'S CASES SHEET ===")
            self.logger.info(f"Chat Agent supporter: {chat_agent_supporter_name}")
            
            # STEP 1: Get current Chat Agent cases from output_df (ALL cases assigned to this supporter)
            current_chat_cases = pd.DataFrame()
            if 'Assigned To' in output_df.columns:
                current_chat_cases = output_df[output_df['Assigned To'] == chat_agent_supporter_name].copy()
                self.logger.info(f"  Current cases assigned to {chat_agent_supporter_name}: {len(current_chat_cases)}")
                if not current_chat_cases.empty:
                    status_cols = [col for col in current_chat_cases.columns if 'status' in col.lower()]
                    if status_cols:
                        status_summary = {}
                        for status_col in status_cols:
                            statuses = current_chat_cases[status_col].value_counts()
                            status_summary[status_col] = statuses.to_dict()
                        self.logger.info(f"  Status breakdown: {status_summary}")
            
            # STEP 2: Load previous Chat Agent cases from previous file (ALL cases, all statuses)
            prev_chat_cases = pd.DataFrame()
            chat_agent_sheet_name = "Chat Agent's Cases"  # Fixed sheet name
            
            if prev_file and os.path.exists(prev_file):
                try:
                    excel_file = pd.ExcelFile(prev_file)
                    if chat_agent_sheet_name in excel_file.sheet_names:
                        prev_chat_cases = pd.read_excel(prev_file, sheet_name=chat_agent_sheet_name)
                        self.logger.info(f"  ✓ Previous Chat Agent cases loaded from '{chat_agent_sheet_name}': {len(prev_chat_cases)}")
                        if not prev_chat_cases.empty:
                            status_cols = [col for col in prev_chat_cases.columns if 'status' in col.lower()]
                            if status_cols:
                                status_summary = {}
                                for status_col in status_cols:
                                    statuses = prev_chat_cases[status_col].value_counts()
                                    status_summary[status_col] = statuses.to_dict()
                                self.logger.info(f"  Previous status breakdown: {status_summary}")
                    else:
                        self.logger.info(f"  No '{chat_agent_sheet_name}' sheet found in previous file")
                except Exception as e:
                    self.logger.warning(f"Could not load previous Chat Agent cases: {str(e)}")
            
            # STEP 3: Normalize case numbers for merging
            if 'Case Number' in current_chat_cases.columns:
                current_chat_cases['Case Number'] = current_chat_cases['Case Number'].apply(
                    lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip().replace('.','',1).isdigit() else pd.NA
                ).astype('Int64')
            
            if 'Case Number' in prev_chat_cases.columns:
                prev_chat_cases['Case Number'] = prev_chat_cases['Case Number'].apply(
                    lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip().replace('.','',1).isdigit() else pd.NA
                ).astype('Int64')
            
            # STEP 4: Merge - keep current, add previous not in current
            current_case_nums = set(current_chat_cases['Case Number'].dropna().tolist()) if not current_chat_cases.empty else set()
            
            if not prev_chat_cases.empty:
                # Only keep previous cases not in current (to avoid duplicates)
                previously_preserved = prev_chat_cases[~prev_chat_cases['Case Number'].isin(current_case_nums)].copy()
                self.logger.info(f"  Preserved from previous days: {len(previously_preserved)} cases")
            else:
                previously_preserved = pd.DataFrame()
            
            # Combine: current cases first, then preserved cases
            if not current_chat_cases.empty or not previously_preserved.empty:
                chat_agent_cases = pd.concat([current_chat_cases, previously_preserved], ignore_index=True)
            else:
                chat_agent_cases = pd.DataFrame()
            
            # STEP 5: Sort by status
            if not chat_agent_cases.empty:
                total_cases = len(chat_agent_cases)
                self.logger.info(f"  ✓ Total cases for Chat Agent sheet:")
                self.logger.info(f"    - Current: {len(current_chat_cases)}")
                self.logger.info(f"    - Preserved from previous: {len(previously_preserved)}")
                self.logger.info(f"    - TOTAL: {total_cases}")
                
                chat_agent_cases = self.sort_cases_by_status(chat_agent_cases)
                
                # STEP 6: Write to Excel
                chat_agent_cases.to_excel(writer, sheet_name=chat_agent_sheet_name, index=False)
                self.logger.info(f"✓ Created '{chat_agent_sheet_name}' sheet with {total_cases} total cases")
            else:
                # CRITICAL FIX: Create empty sheet with proper headers even if no cases
                self.logger.info(f"No cases for Chat Agent - creating empty '{chat_agent_sheet_name}' sheet with headers")
                
                # Create empty DataFrame with standard columns (same structure as handler sheets)
                empty_sheet_df = pd.DataFrame(columns=self.companies_column_order)
                
                # Write empty sheet with headers
                empty_sheet_df.to_excel(writer, sheet_name=chat_agent_sheet_name, index=False)
                self.logger.info(f"✓ Created empty '{chat_agent_sheet_name}' sheet with headers (rows: 0, columns: {len(self.companies_column_order)})")
                
        except Exception as e:
            self.logger.error(f"Error creating Chat Agent's Cases sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")


    def sort_cases_by_status(self, df):
        """Sort cases by status in the order: New → In Progress → Skipped → Closed"""
        try:
            if 'Status' not in df.columns:
                self.logger.warning("No 'Status' column found - cannot sort by status")
                return df
            
            # Create a custom sorting order for status
            # Order: new → in_progress → skipped → closed
            status_order = {
                'new': 1,
                'in_progress': 2,
                'in progress': 2,
                'in progress today': 2,
                'skipped': 3,
                'Skipped': 3,
                'closed': 4,
                'Closed': 4,
                'needs follow up': 5,
                'pending': 6,
                'escalation': 7,
                'bank/sutherland': 8,
                'dnd': 9,
                'DND': 9
            }
            
            # Create a temporary column for sorting
            df_sorted = df.copy()
            df_sorted['_status_sort_order'] = df_sorted['Status'].astype(str).str.lower().map(
                lambda x: status_order.get(x.strip(), 999)  # Default to 999 for unknown statuses
            )
            
            # Sort by status order, then by case number for consistency
            df_sorted = df_sorted.sort_values(['_status_sort_order', 'Case Number']).reset_index(drop=True)
            
            # Remove the temporary sorting column
            df_sorted = df_sorted.drop(columns=['_status_sort_order'])
            
            # Log the sorting results
            if 'Status' in df_sorted.columns:
                status_counts = df_sorted['Status'].value_counts()
                self.logger.info(f"Cases sorted by status. Status distribution:")
                for status, count in status_counts.items():
                    if pd.notna(status) and str(status).strip():
                        self.logger.info(f"  {status}: {count} cases")
            
            return df_sorted
            
        except Exception as e:
            self.logger.error(f"Error sorting cases by status: {str(e)}")
            return df

    def create_handler_sheets_from_processed_data(self, writer, processed_df, prev_file=None, chat_agent_info=None):
        """Create individual handler sheets from the fully processed data (excluding Chat Agent cases)"""
        try:
            self.logger.info("Creating individual handler sheets from processed data...")
            
            # Extract unique handlers from the processed data
            if 'Assigned To' not in processed_df.columns:
                self.logger.warning("No 'Assigned To' column found - skipping handler sheet creation")
                return
            
            # Get Chat Agent supporter name to exclude their cases from individual handler sheets
            chat_agent_supporter_name = None
            if chat_agent_info and chat_agent_info.get('enabled'):
                chat_agent_supporter_name = chat_agent_info.get('supporter_name', 'Chat Agent')
                self.logger.info(f"Chat Agent enabled - will exclude '{chat_agent_supporter_name}' from individual handler sheets")
            
            # CRITICAL: Load previous Companies case numbers to exclude them from handler sheets
            # Cases in Companies sheet should NOT appear in handler sheets
            prev_companies_case_nums = set()
            if prev_file and os.path.exists(prev_file):
                try:
                    with pd.ExcelFile(prev_file) as excel:
                        # Look for Companies sheet
                        companies_sheets = [s for s in excel.sheet_names if isinstance(s, str) and 'companies' in s.lower()]
                        for sheet_name in companies_sheets:
                            try:
                                companies_df = pd.read_excel(prev_file, sheet_name=sheet_name)
                                if 'Case Number' in companies_df.columns:
                                    for cn in companies_df['Case Number'].dropna():
                                        try:
                                            normalized = int(float(str(cn).strip()))
                                            prev_companies_case_nums.add(normalized)
                                        except:
                                            pass
                            except Exception:
                                continue
                    if prev_companies_case_nums:
                        self.logger.info(f"Loaded {len(prev_companies_case_nums)} case numbers from previous Companies sheet - these will be EXCLUDED from handler sheets")
                except Exception as e:
                    self.logger.warning(f"Could not load previous Companies for exclusion: {e}")
            
            # Exclude Companies cases from processed_df for handler sheet creation
            working_df = processed_df.copy()
            if prev_companies_case_nums and 'Case Number' in working_df.columns:
                working_df['_normalized_cn'] = working_df['Case Number'].apply(
                    lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() and str(v).strip().replace('.','',1).isdigit() else None
                )
                excluded_mask = working_df['_normalized_cn'].isin(prev_companies_case_nums)
                excluded_count = excluded_mask.sum()
                if excluded_count > 0:
                    self.logger.info(f"Excluding {excluded_count} cases from handler sheets (they belong to Companies sheet)")
                    working_df = working_df[~excluded_mask].copy()
                working_df = working_df.drop(columns=['_normalized_cn'], errors='ignore')
            
            # Get unique handlers (excluding empty/NaN values)
            handlers = working_df['Assigned To'].dropna().astype(str).str.strip().unique()
            handlers = [h for h in handlers if h and h.lower() not in ['nan', 'none', '']]

            # Load previous handler data if provided
            prev_handler_data = {}
            if prev_file and os.path.exists(prev_file):
                try:
                    prev_handler_data = self.load_previous_handler_sheets(prev_file)
                    self.logger.info(f"Loaded previous handler data for {len(prev_handler_data)} handlers (for preservation)")
                except Exception as e:
                    self.logger.warning(f"Could not load previous handler data: {str(e)}")
                    prev_handler_data = {}

            # Union current handlers with previous to ensure preservation
            all_handlers = set(handlers)
            all_handlers.update(prev_handler_data.keys())
            handlers = list(all_handlers)
            
            # Additional safety check for handler names
            valid_handlers = []
            for handler in handlers:
                if (len(handler) <= 50 and 
                    not handler.startswith('_') and 
                    not handler.startswith('.') and
                    not any(char in handler for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|'])):
                    valid_handlers.append(handler)
                else:
                    self.logger.warning(f"Skipping invalid handler name: '{handler}'")
            
            handlers = valid_handlers
            
            if not handlers:
                self.logger.warning("No valid handlers found in processed data - skipping handler sheet creation")
                return
            
            self.logger.info(f"Found {len(handlers)} valid handlers: {handlers}")
            
            # Create sheets for each handler
            for handler in handlers:
                try:
                    # SKIP: Don't create individual sheet for Chat Agent (they get their own "Chat Agent's Cases" sheet)
                    if chat_agent_supporter_name and handler == chat_agent_supporter_name:
                        self.logger.info(f"Skipping '{handler}' - Chat Agent cases go to 'Chat Agent's Cases' sheet instead")
                        continue
                    
                    self.logger.info(f"Processing handler: {handler}")
                    
                    # New cases (current output for this handler) - use working_df which excludes Companies cases
                    new_cases = working_df[working_df['Assigned To'] == handler].copy()
                    if 'Case Number' in new_cases.columns:
                        new_cases['Case Number'] = new_cases['Case Number'].apply(
                            lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
                        ).astype('Int64')
                        new_cases = new_cases.dropna(subset=['Case Number'])

                    # Preserved cases from previous handler sheets (also exclude Companies cases)
                    prev_cases = prev_handler_data.get(handler, pd.DataFrame()).copy()
                    if not prev_cases.empty and 'Case Number' in prev_cases.columns:
                        prev_cases['Case Number'] = prev_cases['Case Number'].apply(
                            lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
                        ).astype('Int64')
                        prev_cases = prev_cases.dropna(subset=['Case Number'])
                        # Exclude Companies cases from preserved handler data too
                        if prev_companies_case_nums:
                            prev_cases = prev_cases[~prev_cases['Case Number'].isin(prev_companies_case_nums)].copy()
                        prev_cases['Assigned To'] = handler
                    else:
                        prev_cases = pd.DataFrame()

                    # Drop overlap: keep new first, then preserved not in new
                    new_case_nums = set(new_cases['Case Number'].dropna().tolist()) if not new_cases.empty else set()
                    if not prev_cases.empty:
                        prev_cases = prev_cases[~prev_cases['Case Number'].isin(new_case_nums)].copy()

                    # Combine with ordering: new first, then preserved
                    if not new_cases.empty or not prev_cases.empty:
                        handler_cases = pd.concat([new_cases, prev_cases], ignore_index=True)
                    else:
                        self.logger.warning(f"No cases found for handler: {handler}")
                        continue
                    
                    # Sort cases by status (New → In Progress → Closed)
                    self.logger.info(f"Sorting cases for handler '{handler}' by status...")
                    handler_cases = self.sort_cases_by_status(handler_cases)
                    
                    # Create handler sheet name (ensure it's Excel-safe)
                    sheet_name = f"{handler}'s Cases"
                    if len(sheet_name) > 31:
                        sheet_name = f"{handler[:25]}'s Cases"
                    
                    # Write to Excel
                    handler_cases.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust columns for this handler sheet
                    self.auto_adjust_columns(writer, handler_cases, sheet_name)
                    
                    # Apply handler-specific formatting
                    self.format_handler_sheet(writer, sheet_name, handler_cases)
                    
                    self.logger.info(f"Created sheet '{sheet_name}' with {len(handler_cases)} cases (sorted by status)")
                    
                except Exception as e:
                    self.logger.error(f"Error processing handler '{handler}': {str(e)}")
                    import traceback
                    self.logger.error(f"Handler '{handler}' traceback: {traceback.format_exc()}")
                    continue
            
            self.logger.info(f"Handler sheet creation completed for {len(handlers)} handlers")
            
        except Exception as e:
            self.logger.error(f"Error creating handler sheets from processed data: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def create_handler_sheets(self, writer, output_df, prev_file=None):
        """Create individual sheets for each handler with their assigned cases.
        FIXED: Now preserves ALL handler sheets regardless of selection status.
        This is called AFTER all SMS/Email processing to ensure statuses are current."""
        try:
            self.logger.info("Creating individual handler sheets...")
            
            # Extract unique handlers from the output data
            if 'Assigned To' not in output_df.columns:
                self.logger.warning("No 'Assigned To' column found - skipping handler sheet creation")
                return
            
            # Get unique handlers (excluding empty/NaN values)
            current_handlers = output_df['Assigned To'].dropna().astype(str).str.strip().unique()
            current_handlers = [h for h in current_handlers if h and h.lower() not in ['nan', 'none', '']]
            
            # Load previous handler sheets if they exist
            prev_handler_data = {}
            if prev_file and os.path.exists(prev_file):
                try:
                    prev_handler_data = self.load_previous_handler_sheets(prev_file)
                    self.logger.info(f"Loaded previous handler data for {len(prev_handler_data)} handlers")
                except Exception as e:
                    self.logger.warning(f"Could not load previous handler data: {str(e)}")
                    prev_handler_data = {}
            
            # FIXED: Combine current handlers with previous handlers to preserve ALL work
            all_handlers = set(current_handlers)
            all_handlers.update(prev_handler_data.keys())
            handlers = list(all_handlers)
            
            self.logger.info(f"Current handlers in output: {current_handlers}")
            self.logger.info(f"Previous handlers with work: {list(prev_handler_data.keys())}")
            self.logger.info(f"All handlers to process: {handlers}")
            
            # Additional safety check for handler names
            valid_handlers = []
            for handler in handlers:
                # Check if handler name is reasonable (not too long, contains valid characters)
                if (len(handler) <= 50 and 
                    not handler.startswith('_') and 
                    not handler.startswith('.') and
                    not any(char in handler for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|'])):
                    valid_handlers.append(handler)
                else:
                    self.logger.warning(f"Skipping invalid handler name: '{handler}'")
            
            handlers = valid_handlers
            
            if not handlers:
                self.logger.warning("No valid handlers found in data - skipping handler sheet creation")
                return
            
            self.logger.info(f"Found {len(handlers)} valid handlers to process: {handlers}")
            
            # Create/update sheets for each handler
            for handler in handlers:
                try:
                    self.logger.info(f"Processing handler: {handler}")
                    
                    # Filter cases for this handler from current output
                    handler_cases = output_df[output_df['Assigned To'] == handler].copy()
                    
                    # Get previous handler data if available
                    prev_cases = prev_handler_data.get(handler, pd.DataFrame())
                    
                    # FIXED: Handle handlers with previous work but no current cases
                    if handler_cases.empty and not prev_cases.empty:
                        self.logger.info(f"Handler '{handler}' has no current cases but has previous work - preserving previous work")
                        final_handler_cases = prev_cases.copy()
                    elif not handler_cases.empty and prev_cases.empty:
                        self.logger.info(f"Handler '{handler}' has current cases but no previous work - using current cases")
                        final_handler_cases = handler_cases.copy()
                    elif not handler_cases.empty and not prev_cases.empty:
                        self.logger.info(f"Handler '{handler}' has both current and previous cases - merging")
                        # Merge with previous data (update existing, append new)
                        final_handler_cases = self.merge_handler_cases(handler_cases, prev_cases, handler)
                    else:
                        self.logger.warning(f"Handler '{handler}' has no current or previous cases - skipping")
                        continue
                    
                    # Create handler sheet name (ensure it's Excel-safe)
                    sheet_name = f"{handler}'s Cases"
                    # Excel sheet names cannot exceed 31 characters
                    if len(sheet_name) > 31:
                        sheet_name = f"{handler[:25]}'s Cases"
                    
                    # Write to Excel
                    final_handler_cases.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Auto-adjust columns for this handler sheet
                    self.auto_adjust_columns(writer, final_handler_cases, sheet_name)
                    
                    # Apply handler-specific formatting (simplified for now)
                    self.format_handler_sheet(writer, sheet_name, final_handler_cases)
                    
                    self.logger.info(f"Created/updated sheet '{sheet_name}' with {len(final_handler_cases)} cases")
                    
                    # Do not add internal tracking columns; preservation handled by merging prev handler DataFrames
                    
                except Exception as e:
                    self.logger.error(f"Error processing handler '{handler}': {str(e)}")
                    import traceback
                    self.logger.error(f"Handler '{handler}' traceback: {traceback.format_exc()}")
                    continue  # Continue with next handler
            
            self.logger.info(f"Handler sheet creation completed for {len(handlers)} handlers")
            
        except Exception as e:
            self.logger.error(f"Error creating handler sheets: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def load_previous_handler_sheets(self, prev_file):
        """Load previous handler sheets data for smart updates"""
        prev_handler_data = {}

        # Validate file existence
        if not prev_file or not os.path.exists(prev_file):
            self.logger.info("No previous file provided or file doesn't exist")
            return prev_handler_data

        try:
            # Read the previous file safely
            prev_excel = pd.ExcelFile(prev_file)

            # Look for handler sheets (sheets ending with "'s Cases")
            handler_sheets = [s for s in prev_excel.sheet_names if isinstance(s, str) and s.endswith("'s Cases")]

            for sheet_name in handler_sheets:
                try:
                    # Normalize handler name and read the sheet
                    handler_name = str(sheet_name).replace("'s Cases", "").strip()
                    handler_df = pd.read_excel(prev_file, sheet_name=sheet_name)

                    if isinstance(handler_df, pd.DataFrame) and not handler_df.empty and 'Case Number' in handler_df.columns:
                        prev_handler_data[handler_name] = handler_df
                        self.logger.info(f"Loaded previous data for handler '{handler_name}': {len(handler_df)} cases")

                except Exception as e:
                    self.logger.warning(f"Error loading previous handler sheet '{sheet_name}': {str(e)}")
                    continue

            return prev_handler_data

        except Exception as e:
            self.logger.error(f"Error loading previous handler sheets: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            return {}

    def merge_handler_cases(self, current_cases, prev_cases, handler_name):
        """Smart merge of current and previous handler cases:
        - Update existing cases with new data
        - Append new cases
        - Preserve handler-specific notes and custom fields
        """
        try:
            if prev_cases.empty:
                self.logger.info(f"No previous data for handler '{handler_name}' - using current cases only")
                return current_cases
            
            self.logger.info(f"Merging cases for handler '{handler_name}': {len(current_cases)} current, {len(prev_cases)} previous")
            
            # Ensure Case Number column exists
            if 'Case Number' not in current_cases.columns:
                self.logger.warning(f"No 'Case Number' column in current cases for handler '{handler_name}' - using current cases only")
                return current_cases
            
            if 'Case Number' not in prev_cases.columns:
                self.logger.warning(f"No 'Case Number' column in previous cases for handler '{handler_name}' - using current cases only")
                return current_cases
            
            # Normalize case numbers for comparison as integers (nullable)
            current_cases = current_cases.copy()
            prev_cases = prev_cases.copy()
            current_cases['Case Number'] = current_cases['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            prev_cases['Case Number'] = prev_cases['Case Number'].apply(
                lambda v: int(float(str(v).strip())) if pd.notna(v) and str(v).strip() != '' and str(v).strip().replace('.','',1).isdigit() else pd.NA
            ).astype('Int64')
            
            # Get case numbers
            current_case_numbers = set(current_cases['Case Number'].dropna().tolist())
            prev_case_numbers = set(prev_cases['Case Number'].dropna().tolist())
            
            # Identify cases
            existing_cases = current_case_numbers & prev_case_numbers  # Cases in both
            new_cases = current_case_numbers - prev_case_numbers      # New cases
            removed_cases = prev_case_numbers - current_case_numbers  # Cases no longer assigned to this handler
            
            self.logger.info(f"Handler '{handler_name}' case analysis:")
            self.logger.info(f"  Existing cases to update: {len(existing_cases)}")
            self.logger.info(f"  New cases to add: {len(new_cases)}")
            self.logger.info(f"  Cases no longer assigned: {len(removed_cases)}")
            
            # Start with new cases (these are current and complete)
            final_cases = current_cases[current_cases['Case Number'].isin(new_cases)].copy()
            
            # Add existing cases (merge current data with previous custom fields)
            for case_num in existing_cases:
                try:
                    # Get current case data
                    current_case = current_cases[current_cases['Case Number'] == case_num].iloc[0]
                    
                    # Get previous case data
                    prev_case = prev_cases[prev_cases['Case Number'] == case_num].iloc[0]
                    
                    # Create merged case row
                    merged_case = current_case.copy()
                    
                    # Preserve handler-specific custom fields from previous data
                    custom_fields = ['Handler Notes', 'Custom Status', 'Priority Level', 'Follow-up Date']
                    for field in custom_fields:
                        if field in prev_case.index and field not in merged_case.index:
                            merged_case[field] = prev_case[field]
                        elif field in prev_case.index and field in merged_case.index:
                            # Keep previous custom data if current is empty
                            if _is_missing(merged_case[field]):
                                merged_case[field] = prev_case[field]
                    
                    # Add to final cases
                    final_cases = pd.concat([final_cases, pd.DataFrame([merged_case])], ignore_index=True)
                    
                except Exception as e:
                    self.logger.warning(f"Error merging case '{case_num}' for handler '{handler_name}': {str(e)}")
                    # Add current case as fallback
                    current_case = current_cases[current_cases['Case Number'] == case_num].iloc[0]
                    final_cases = pd.concat([final_cases, pd.DataFrame([current_case])], ignore_index=True)
                    continue
            
            # Sort by case number for consistency
            if not final_cases.empty:
                final_cases = final_cases.sort_values('Case Number').reset_index(drop=True)
            
            self.logger.info(f"Final merged cases for handler '{handler_name}': {len(final_cases)} total")
            
            return final_cases
            
        except Exception as e:
            self.logger.error(f"Error merging handler cases for '{handler_name}': {str(e)}")
            # Fallback: return current cases only
            return current_cases

    def format_handler_sheet(self, writer, sheet_name, handler_cases):
        """Apply handler-specific formatting to the handler sheet"""
        try:
            # For now, just log that we would format the sheet
            # The actual formatting will be handled by pandas/xlsxwriter automatically
            self.logger.info(f"Handler sheet '{sheet_name}' created with {len(handler_cases)} cases")
            
            # Note: Column widths and formatting are handled automatically by pandas/xlsxwriter
            # We can add custom formatting later if needed using openpyxl after the file is created
            
        except Exception as e:
            self.logger.warning(f"Could not apply formatting to handler sheet '{sheet_name}': {str(e)}")

    def add_validation_dropdowns(self, output_file):
        """Add validation dropdown lists to the PA Cases sheet and all handler sheets based on validation lists, and append region/offset table after two empty columns."""
        try:
            self.logger.info("Adding validation dropdown lists to PA Cases sheet and handler sheets...")
            
            # Check if openpyxl is available
            if not OPENPYXL_AVAILABLE:
                self.logger.warning("openpyxl not available - skipping validation dropdowns")
                return
            
            # Load the workbook
            wb = openpyxl.load_workbook(output_file)

            # Get all sheets
            all_sheets = wb.sheetnames
            self.logger.info(f"Found sheets: {all_sheets}")
            
            # Identify handler sheets
            handler_sheets = [sheet for sheet in all_sheets if isinstance(sheet, str) and sheet.endswith("'s Cases")]
            main_sheets = ['PA Cases'] + handler_sheets
            if 'Companies' in all_sheets:
                main_sheets.append('Companies')
            
            self.logger.info(f"Will add validation to main sheets: {main_sheets}")

            # Process each main sheet (PA Cases + handler sheets)
            for sheet_name in main_sheets:
                if sheet_name not in wb.sheetnames:
                    self.logger.warning(f"Sheet '{sheet_name}' not found, skipping")
                    continue
                    
                self.logger.info(f"Adding validation dropdowns to sheet: {sheet_name}")
                data_sheet = wb[sheet_name]

                # Add filters to all header columns
                header_row = 1  # Headers are in row 1
                max_column = data_sheet.max_column
                data_sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_column)}{header_row}"

                # Create data validation for each specified column
                header_map = {}
                for cell in data_sheet[header_row]:
                    if cell.value:
                        header_map[cell.value] = cell.column

                for col_name, options in self.validation_lists.items():
                    if col_name in header_map:
                        col_letter = get_column_letter(header_map[col_name])
                        # Create data validation with explicit range
                        options_range = f'Validation!{get_column_letter(1 + list(self.validation_lists.keys()).index(col_name))}$2:{get_column_letter(1 + list(self.validation_lists.keys()).index(col_name))}${1 + len(options)}'
                        
                        # Check if we have the real DataValidation class (not dummy)
                        if OPENPYXL_AVAILABLE and hasattr(DataValidation, 'add'):
                            # Import the real DataValidation class for this use
                            from openpyxl.worksheet.datavalidation import DataValidation as RealDataValidation
                            dv = RealDataValidation(type="list", formula1=f'={options_range}')
                            dv.error = 'Your entry is not in the list'
                            dv.errorTitle = 'Invalid Entry'
                            dv.prompt = 'Please select from the list'
                            dv.promptTitle = col_name
                            # Apply to all cells in column except header
                            data_sheet.add_data_validation(dv)
                            dv.add(f'{col_letter}2:{col_letter}1048576')  # Applies to entire column
                            self.logger.info(f"Added validation dropdown for column: {col_name} in sheet: {sheet_name}")
                        else:
                            self.logger.warning(f"DataValidation not available - skipping validation for column: {col_name}")

            # Create or clear the Validation sheet
            if "Validation" in wb.sheetnames:
                validation_sheet = wb["Validation"]
                validation_sheet.delete_rows(1, validation_sheet.max_row)
            else:
                validation_sheet = wb.create_sheet("Validation")

            # Write validation options to the Validation sheet (horizontal, starting at A1)
            headers = list(self.validation_lists.keys())
            max_len = max(len(v) for v in self.validation_lists.values())
            for col_idx, col_name in enumerate(headers, start=1):
                validation_sheet.cell(row=1, column=col_idx, value=col_name)
                options = self.validation_lists[col_name]
                for row_idx, option in enumerate(options, start=2):
                    validation_sheet.cell(row=row_idx, column=col_idx, value=option)
                # Fill empty cells to max_len
                for row_idx in range(len(options)+2, max_len+2):
                    validation_sheet.cell(row=row_idx, column=col_idx, value="")

            # Leave two empty columns after the validation lists
            region_start_col = len(headers) + 3  # +1 for next col, +2 for two empty columns

            # Region/Offset data from screenshots
            regions = [
                "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado", "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota", "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada", "New Hampshire", "New Jersey", "New Mexico", "New York", "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon", "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington", "West Virginia", "Wisconsin", "Wyoming", "Alberta", "British Columbia", "Manitoba", "New Brunswick", "Newfoundland and Labrador", "Nova Scotia", "Ontario", "Prince Edward Island", "Quebec", "Saskatchewan", "Northwest Territories", "Nunavut", "Yukon"
            ]
            offsets = [
                "5:00:00 AM", "4:00:00 AM", "8:00:00 AM", "6:00:00 AM", "5:00:00 AM", "7:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "4:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "5:00:00 AM", "5:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM", "5:00:00 AM", "6:00:00 AM", "4:00:00 AM", "4:00:00 AM"
            ]
            region_table_headers = ["Region", "Offset", "=NOW()-K1", "3:00"]
            region_table_row_start = 1
            # Write headers
            for i, header in enumerate(region_table_headers):
                validation_sheet.cell(row=region_table_row_start, column=region_start_col + i, value=header)
            # Write region/offset data
            for idx, (region, offset) in enumerate(zip(regions, offsets), start=2):
                validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col, value=region)
                # Convert offset string to Excel time format
                from datetime import datetime
                import re
                time_match = re.match(r"(\d+):(\d+):(\d+) (AM|PM)", offset)
                if time_match:
                    hour, minute, second, ampm = time_match.groups()
                    hour = int(hour)
                    minute = int(minute)
                    second = int(second)
                    if ampm == 'PM' and hour != 12:
                        hour += 12
                    if ampm == 'AM' and hour == 12:
                        hour = 0
                    excel_time = (hour * 3600 + minute * 60 + second) / 86400
                    validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 1, value=excel_time)
                    validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 1).number_format = 'h:mm:ss AM/PM'
                else:
                    validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 1, value=offset)
                # Formula: =$J$1-I2 (I2 increments)
                formula = f"=$J$1-I{region_table_row_start + idx - 1}"
                validation_sheet.cell(row=region_table_row_start + idx - 1, column=region_start_col + 2, value=formula)
                # 3:00 column left blank except header

            # --- Add VLOOKUP formula to PA Cases sheet, column K ---
            # Only add to PA Cases sheet (not handler sheets) to avoid confusion
            if 'PA Cases' in wb.sheetnames:
                data_sheet = wb['PA Cases']
                pa_cases_max_row = data_sheet.max_row
                
                # Check if Local Time column already exists, if not create it
                local_time_col = None
                for col_idx, col_name in enumerate(data_sheet[1], 1):  # Check header row
                    if col_name.value == 'Local Time':
                        local_time_col = col_idx
                        break
                
                if not local_time_col:
                    # Create new Local Time column at the end
                    local_time_col = data_sheet.max_column + 1
                    # Set header for Local Time column
                    data_sheet.cell(row=1, column=local_time_col, value='Local Time')
                    self.logger.info(f"Created new Local Time column at column {local_time_col}")
                else:
                    self.logger.info(f"Local Time column already exists at column {local_time_col}")
                
                    # Add formulas to all data rows (overwrite existing data)
                    for row in range(2, pa_cases_max_row + 1):
                        # =VLOOKUP(k2,Validation!$H$1:$J$63,3) - Fixed range reference
                        # Column k is State, looking up in Validation sheet H:J range
                        formula = f'=VLOOKUP(k{row},Validation!{get_column_letter(region_start_col)}1:{get_column_letter(region_start_col+2)}{len(regions)+1},3)'
                        data_sheet.cell(row=row, column=local_time_col, value=formula)
                    self.logger.info(f"Added/Updated Local Time formulas in column {get_column_letter(local_time_col)} for {pa_cases_max_row-1} rows")

            # --- Add VLOOKUP formula to ALL handler sheets AND Companies ---
            # Get all handler sheet names (sheets ending with "'s Cases") + Companies
            handler_sheets = [sheet for sheet in wb.sheetnames if isinstance(sheet, str) and (sheet.endswith("'s Cases") or sheet == 'Companies')]
            self.logger.info(f"Adding Local Time formulas to {len(handler_sheets)} sheets (Handlers + Companies): {handler_sheets}")
            
            for handler_sheet_name in handler_sheets:
                try:
                    handler_sheet = wb[handler_sheet_name]
                    handler_max_row = handler_sheet.max_row
                    
                    # Check if Local Time column already exists, if not create it
                    local_time_col = None
                    for col_idx, col_name in enumerate(handler_sheet[1], 1):  # Check header row
                        if col_name.value == 'Local Time':
                            local_time_col = col_idx
                            break
                    
                    if not local_time_col:
                        # Create new Local Time column at the end
                        local_time_col = handler_sheet.max_column + 1
                        # Set header for Local Time column
                        handler_sheet.cell(row=1, column=local_time_col, value='Local Time')
                        self.logger.info(f"Created new Local Time column at column {local_time_col} in {handler_sheet_name}")
                    else:
                        self.logger.info(f"Local Time column already exists at column {local_time_col} in {handler_sheet_name}")
                    
                    # Add formulas to all data rows (overwrite existing data)
                    for row in range(2, handler_max_row + 1):
                        # =VLOOKUP(J2,Validation!$H$1:$J$63,3) - Same formula as PA Cases
                        # Column J is State/Province, looking up in Validation sheet H:J range
                        formula = f'=VLOOKUP(k{row},Validation!{get_column_letter(region_start_col)}1:{get_column_letter(region_start_col+2)}{len(regions)+1},3)'
                        handler_sheet.cell(row=row, column=local_time_col, value=formula)
                    
                    self.logger.info(f"Added/Updated Local Time formulas in column {get_column_letter(local_time_col)} for {handler_max_row-1} rows in {handler_sheet_name}")
                    
                except Exception as e:
                    self.logger.error(f"Error adding Local Time formulas to {handler_sheet_name}: {str(e)}")
                    continue

            # Save the workbook
            wb.save(output_file)
            self.logger.info("Validation dropdowns and region/offset table added successfully to all sheets")

        except Exception as e:
            self.logger.error(f"Error adding validation dropdowns: {str(e)}")

    def create_skipped_sms_sheet(self, writer, processing_stats=None):
        """Create a sheet showing skipped SMS cases"""
        try:
            self.logger.info("\n=== Collecting Skipped SMS Entries ===")
            
            # Initialize collection
            sms_entries = []
            
            if processing_stats:
                self.logger.info("Processing stats keys available:")
                for key in processing_stats.keys():
                    self.logger.info(f"- {key}")
                
                # Check SMS Processing Stats
                if 'SMS Processing Stats' in processing_stats:
                    sms_stats = processing_stats.get('SMS Processing Stats', {})
                    self.logger.info("SMS Processing Stats keys:")
                    for key in sms_stats.keys():
                        self.logger.info(f"- {key}")
                    
                    # Try to collect entries from both possible locations
                    skipped_updated = list(sms_stats.get('Skipped/Updated Entries', []))
                    skipped_cases = list(sms_stats.get('Skipped Cases', []))
                    
                    self.logger.info(f"Found {len(skipped_updated)} entries in Skipped/Updated Entries")
                    self.logger.info(f"Found {len(skipped_cases)} entries in Skipped Cases")
                    
                    # Combine unique entries
                    if skipped_updated:
                        sms_entries.extend(skipped_updated)
                    if skipped_cases and not skipped_updated:  # Only use legacy if no new format
                        sms_entries.extend(skipped_cases)
            else:
                self.logger.info("No processing_stats provided - cannot collect skipped SMS entries")
                return

            # Store collected entries in both locations for compatibility
            if processing_stats and sms_entries:
                try:
                    # Store in designated location for final unified sheet
                    processing_stats['_skipped_sms_entries'] = sms_entries
                    
                    # Also ensure it's in SMS Processing Stats
                    if 'SMS Processing Stats' not in processing_stats:
                        processing_stats['SMS Processing Stats'] = {}
                    processing_stats['SMS Processing Stats']['Skipped/Updated Entries'] = sms_entries
                    
                    self.logger.info(f"Successfully stored {len(sms_entries)} SMS entries for unified sheet")
                    
                    # Log some sample entries
                    if sms_entries:
                        self.logger.info("\nSample SMS entries collected:")
                        for entry in sms_entries[:3]:  # Show first 3
                            case = entry.get('Case Number', 'N/A')
                            reply = entry.get('SMS Text', entry.get('Reply Text', 'N/A'))
                            reason = entry.get('Reason', 'N/A')
                            self.logger.info(f"Case {case}: Reply='{reply}', Reason='{reason}'")
                except Exception as e:
                    self.logger.warning(f"Failed to store SMS entries: {str(e)}")
            return
        except Exception as e:
            self.logger.error(f"Error creating Skipped SMS Cases sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def _is_missing(self, value):
        """Helper function to check if a value should be considered missing"""
        if value is None:
            return True
        if isinstance(value, str) and value.strip() in ['', 'nan', 'None', 'N/A']:
            return True
        try:
            if pd.isna(value):
                return True
        except:
            pass
        return False

    def create_unified_skipped_sheet(self, writer, processing_stats=None):
        """Create a unified sheet showing all skipped cases (both SMS and Email)"""
        try:
            self.logger.info("\n=== Creating Unified Skipped Sheet ===")
            
            # Initialize storage for entries
            all_entries = []
            sms_entries = []
            email_entries = []
            
            # Get SMS entries - try both new and legacy locations
            if processing_stats:
                # First try direct storage
                sms_entries = list(processing_stats.get('_skipped_sms_entries', []))
                
                # If not found, try SMS Processing Stats
                if not sms_entries and 'SMS Processing Stats' in processing_stats:
                    sms_stats = processing_stats['SMS Processing Stats']
                    if sms_stats:
                        sms_entries = list(sms_stats.get('Skipped/Updated Entries', [])) or list(sms_stats.get('Skipped Cases', []))
                
                self.logger.info(f"Found {len(sms_entries)} SMS entries")
                
                # Get Email entries - try both new and legacy locations
                # First try direct storage
                email_entries = list(processing_stats.get('_skipped_email_entries', []))
                
                # If not found, try Email Processing Stats
                if not email_entries and 'Email Processing Stats' in processing_stats:
                    email_stats = processing_stats['Email Processing Stats']
                    if email_stats:
                        email_entries = list(email_stats.get('Skipped/Updated Entries', [])) or list(email_stats.get('Skipped Emails', []))
                
                self.logger.info(f"Found {len(email_entries)} Email entries")
            
            def normalize_entry(entry, source_type):
                """Helper to normalize entry format"""
                if not isinstance(entry, dict):
                    return None
                
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Base normalized entry
                normalized = {
                    'Timestamp': entry.get('Timestamp', now),
                    'Source': source_type,
                    'Case Number': entry.get('Case Number', 'N/A'),
                    'Reply Text': '',
                    'Reason': entry.get('Reason', 'Unknown'),
                    'Action': entry.get('Action', 'Skipped')
                }
                
                # Extract reply text based on source
                if source_type == 'SMS':
                    normalized['Reply Text'] = entry.get('SMS Text', entry.get('Reply Text', ''))
                else:  # Email
                    normalized['Reply Text'] = entry.get('Reply Text', '')
                
                # Clean up values
                for key in normalized:
                    if _is_missing(normalized[key]):
                        normalized[key] = 'N/A'
                    elif isinstance(normalized[key], str):
                        normalized[key] = normalized[key].strip()
                
                return normalized

            # Process SMS entries
            for entry in sms_entries:
                normalized = normalize_entry(entry, 'SMS')
                if normalized:
                    all_entries.append(normalized)
            
            # Process Email entries
            for entry in email_entries:
                normalized = normalize_entry(entry, 'Email')
                if normalized:
                    all_entries.append(normalized)
            
            if not all_entries:
                self.logger.info("No skipped entries to write to unified sheet")
                return

            # Create DataFrame
            df = pd.DataFrame(all_entries)
            
            # Ensure required columns exist and are in the right order
            required_columns = ['Timestamp', 'Source', 'Case Number', 'Reply Text', 'Reason', 'Action']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = 'N/A'
            
            # Reorder columns
            df = df.reindex(columns=required_columns)

            # Write to Excel
            sheet_name = 'Skipped SMS and Email Replies'
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Auto-adjust columns
            worksheet = writer.sheets[sheet_name]
            worksheet.autofit()
            
            self.logger.info(f"\nSuccessfully wrote {len(all_entries)} entries to unified skipped sheet")
            self.logger.info("Source distribution:")
            source_counts = df['Source'].value_counts()
            for source, count in source_counts.items():
                self.logger.info(f"  {source}: {count} entries")
            
            # Log some sample entries for verification
            self.logger.info("\nSample entries from unified sheet:")
            sample_size = min(5, len(df))
            for i, (_, row) in enumerate(df.head(sample_size).iterrows()):
                self.logger.info(f"Row {i+1}: {row['Source']} - Case {row['Case Number']} - {row['Action']}")
                reply_preview = str(row['Reply Text'])[:50] + ('...' if len(str(row['Reply Text'])) > 50 else '')
                self.logger.info(f"  Reply: '{reply_preview}'")
                self.logger.info(f"  Reason: {row['Reason']}")

        except Exception as e:
            self.logger.error(f"Error creating unified skipped sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def create_skipped_email_sheet(self, writer, processing_stats=None, email_file=None, output_df=None):
        """Create a sheet showing skipped Email replies.

        Fallback behavior: if processing_stats does not include skipped emails but an
        email_file is provided, attempt to parse the file and build skipped-email
        entries based on simple heuristics (missing case number, case not found in
        output_df, or non-actionable replies).
        """
        # Initialize the collections we'll use
        skipped_emails = []
        combined_entries = []
        sms_entries = []
        stats = processing_stats if processing_stats is not None else {}

        self.logger.info("\n=== Creating Skipped SMS / Email Replies Sheet ===")

        try:
            # Log what we find in stats for debugging
            if stats:
                self.logger.info("Processing stats keys found:")
                for key in stats.keys():
                    self.logger.info(f"- {key}")
                if 'SMS Processing Stats' in stats:
                    self.logger.info("SMS Processing Stats keys:")
                    for key in stats['SMS Processing Stats'].keys():
                        self.logger.info(f"- {key}")
                if 'Email Processing Stats' in stats:
                    self.logger.info("Email Processing Stats keys:")
                    for key in stats['Email Processing Stats'].keys():
                        self.logger.info(f"- {key}")

            # Get SMS entries from both places they might be stored
            sms_entries = []
            if stats:
                # Try the direct storage location
                try:
                    stored_sms = stats.get('_skipped_sms_entries', [])
                    if stored_sms:
                        sms_entries.extend(stored_sms)
                        self.logger.info(f"Found {len(stored_sms)} SMS entries in _skipped_sms_entries")
                except Exception:
                    pass
                
                        # Also check SMS Processing Stats
                try:
                    sms_stats = stats.get('SMS Processing Stats', {})
                    if sms_stats:
                        skipped_sms = list(sms_stats.get('Skipped/Updated Entries', [])) or list(sms_stats.get('Skipped Cases', []))
                        if skipped_sms:
                            sms_entries.extend(skipped_sms)
                            self.logger.info(f"Found {len(skipped_sms)} SMS entries in SMS Processing Stats")
                except Exception:
                    pass
                    
            # Initialize skipped_emails list
            skipped_emails = []

            # Get Email entries from processing stats
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                skipped_emails.extend(list(email_stats.get('Skipped/Updated Entries', [])) or list(email_stats.get('Skipped Emails', [])))
                self.logger.info(f"Found {len(skipped_emails)} email entries in Email Processing Stats")

            # Fallback: attempt to parse email_file to build minimal uniform entries
            input_email_file = email_file
            if not skipped_emails and input_email_file and os.path.exists(input_email_file):
                try:
                    self.logger.info(f"No skipped emails in processing_stats - attempting to parse email file as fallback: {input_email_file}")
                    email_df = pd.read_excel(input_email_file)

                    # Helper to find likely column by name or index (F/H)
                    def find_col_by_name_or_index(df, names, idx):
                        for n in names:
                            for col in df.columns:
                                if isinstance(col, str) and col.strip().lower() == n.strip().lower():
                                    return col
                        if len(df.columns) > idx:
                            return df.columns[idx]
                        return None

                    case_number_col = find_col_by_name_or_index(email_df, ['Case Number (Object) (Email)', 'Case Number'], 5)
                    reply_col = find_col_by_name_or_index(email_df, ['Action', 'Reply', 'Description'], 7)

                    if case_number_col and reply_col:
                        for idx, row in email_df.iterrows():
                            raw_case = row.get(case_number_col, '')
                            case_str = str(raw_case).strip()
                            reply_raw = row.get(reply_col, '')
                            reply_text = str(reply_raw).strip() if reply_raw else ''
                            # Normalize case
                            import re
                            m = re.search(r"(\d+)", case_str)
                            norm_case = m.group(1) if m else case_str
                            reason = 'Missing case number' if not norm_case or norm_case.lower() in ['nan', 'none', ''] else 'Non-actionable or not found'
                            if not norm_case or norm_case.lower() in ['nan', 'none', '']:
                                skipped_emails.append({'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Source': 'Email', 'Case Number': 'N/A', 'Reply Text': reply_text, 'Result': 'Skipped', 'Reason': 'Missing case number'})
                            else:
                                skipped_emails.append({'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Source': 'Email', 'Case Number': str(norm_case), 'Reply Text': reply_text, 'Result': 'Skipped', 'Reason': reason})
                    else:
                        self.logger.info("Could not find necessary columns in email file for fallback skipped-email parsing")
                except Exception as e:
                    self.logger.warning(f"Failed fallback parsing of email file for skipped emails: {e}")

            # Normalize entries: ensure they have the unified keys
            def normalize_entry(e, default_source=''):
                # Expecting already in unified format or legacy short dicts
                if not isinstance(e, dict):
                    return None
                # If it already has Timestamp/Source/Case Number/Reply Text/Result/Reason, use as-is
                keys = ['Timestamp', 'Source', 'Case Number', 'Reply Text', 'Result', 'Reason']
                if all(k in e for k in keys):
                    return {k: e.get(k, '') for k in keys}
                # Legacy SMS format
                if 'SMS Text' in e or 'Reply Text' in e:
                    ts = e.get('Timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    source = e.get('Source', default_source or 'SMS')
                    case = e.get('Case Number', e.get('case_number', ''))
                    reply = e.get('SMS Text', e.get('Reply Text', ''))
                    reason = e.get('Reason', e.get('Update Made', 'Skipped'))
                    result = 'Updated' if 'Update Made' in e or e.get('Result') == 'Updated' else 'Skipped'
                    return {'Timestamp': ts, 'Source': source, 'Case Number': case, 'Reply Text': reply, 'Result': result, 'Reason': reason}
                # Fallback minimal mapping
                ts = e.get('Timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                source = e.get('Source', default_source or 'Email')
                case = e.get('Case Number', e.get('case_number', ''))
                reply = e.get('Reply Text', e.get('reply_text', ''))
                reason = e.get('Reason', '')
                result = e.get('Result', 'Skipped')
                return {'Timestamp': ts, 'Source': source, 'Case Number': case, 'Reply Text': reply, 'Result': result, 'Reason': reason}

            # Combine entries
            combined = []
            
            # Add SMS entries
            sms_added = 0
            for s in sms_entries:
                ne = normalize_entry(s, default_source='SMS')
                if ne:
                    combined.append(ne)
                    sms_added += 1
            self.logger.info(f"Added {sms_added} normalized SMS entries")
            
            # Add Email entries
            email_added = 0
            for e in skipped_emails:
                ne = normalize_entry(e, default_source='Email')
                if ne:
                    combined.append(ne)
                    email_added += 1
            self.logger.info(f"Added {email_added} normalized Email entries")

            if not combined:
                self.logger.info("No skipped SMS/Email entries to report")
                return

            # Create the unified sheet
            self.logger.info(f"Creating unified Skipped SMS / Email Replies sheet with {len(combined)} total rows")
            skipped_df = pd.DataFrame(combined)
            # Ensure column order
            cols = ['Timestamp', 'Source', 'Case Number', 'Reply Text', 'Result', 'Reason']
            for c in cols:
                if c not in skipped_df.columns:
                    skipped_df[c] = ''
            skipped_df = skipped_df[cols]
            
            # Save and format
            try:
                skipped_df.to_excel(writer, sheet_name='Skipped SMS and Email Replies', index=False)
                self.auto_adjust_columns(writer, skipped_df, 'Skipped SMS and Email Replies')
                self.logger.info(f"Successfully created Skipped SMS and Email Replies sheet with {len(skipped_df)} rows")
                
                # Log some sample rows for verification
                self.logger.info("\nSample rows from unified sheet:")
                sample_size = min(5, len(skipped_df))
                sample_rows = skipped_df.head(sample_size)
                for idx, row in sample_rows.iterrows():
                    row_num = idx if isinstance(idx, int) else 0
                    self.logger.info(f"Row {row_num+1}: Source={row['Source']}, Case={row['Case Number']}, Result={row['Result']}")
            except Exception as write_error:
                self.logger.error(f"Error writing unified sheet: {str(write_error)}")
                raise
                
        except Exception as e:
            self.logger.error(f"Error creating Skipped Email Replies sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")
            try:
                raise
            except Exception as e:
                self.logger.error(f"Error processing email file: {str(e)}")
                return

    def create_dnd_emails_database_sheet(self, writer, output_df, email_file, processing_stats=None, prev_file=None):
        """Create DND emails sheet by preserving ALL previous data and only appending new emails"""
        try:
            self.logger.info("Creating DND Emails sheet...")
            
            # Use a set to store emails to easily handle uniqueness
            dnd_emails_set = set()
            dnd_list = []
            
            # STEP 1: Load and preserve ALL previous DND Emails data FIRST
            self.logger.info(f"Attempting to load DND Emails from prev_file: {prev_file}")
            if prev_file and os.path.exists(prev_file):
                self.logger.info(f"prev_file exists for DND Emails: {prev_file}")
                try:
                    self.logger.info(f"Loading previous DND Emails sheet from: {prev_file}")
                    prev_excel = pd.ExcelFile(prev_file)
                    
                    self.logger.info(f"Found sheets in DND prev_file: {prev_excel.sheet_names}")
                    
                    if 'DND Emails' in prev_excel.sheet_names:
                        prev_dnd = pd.read_excel(prev_file, sheet_name='DND Emails')
                        self.logger.info(f"Previous DND Emails sheet loaded with {len(prev_dnd)} records")
                        
                        if 'Email' in prev_dnd.columns:
                            # Add ALL previous emails to our set
                            for idx, row in prev_dnd.iterrows():
                                email = str(row.get('Email', '')).strip()
                                if email and email.lower() not in ['', 'nan', 'none']:
                                    if email not in dnd_emails_set:
                                        dnd_emails_set.add(email)
                                        dnd_list.append({'Email': email, 'DND': 'Yes'})
                            self.logger.info(f"Preserved {len(dnd_list)} emails from previous DND Emails sheet")
                        else:
                            self.logger.warning("Previous file does not contain 'Email' column")
                    else:
                        self.logger.info("Previous output file does not contain 'DND Emails' tab")
                        
                except Exception as e:
                    self.logger.error(f"Error loading previous DND Emails sheet: {str(e)}")
                    self.logger.warning("Continuing without previous file data")
            else:
                self.logger.info("No previous output file found - starting with empty DND Emails sheet")
            
            # STEP 2: Add new DND emails from PA Cases output_df (if not already in set)
            new_emails_count = 0
            if output_df is not None and 'Email' in output_df.columns:
                if 'DND (Do Not Disturb)' in output_df.columns:
                    mask = output_df['DND (Do Not Disturb)'].astype(str).str.strip().str.lower() == 'yes'
                    df_dnd = output_df[mask]
                    self.logger.info(f"Found {len(df_dnd)} new DND cases in PA Cases")
                else:
                    df_dnd = pd.DataFrame()

                for _, row in df_dnd.iterrows():
                    email = str(row.get('Email', '')).strip()
                    if email and email.lower() not in ['', 'nan', 'none']:
                        if email not in dnd_emails_set:
                            dnd_emails_set.add(email)
                            dnd_list.append({'Email': email, 'DND': 'Yes'})
                            new_emails_count += 1
                            self.logger.debug(f"Added NEW DND email from PA Cases: {email}")
            
            # STEP 3: Also include any raw-file-extracted DND emails (if not already in set)
            if processing_stats and isinstance(processing_stats, dict):
                raw_dnd = processing_stats.get('Raw File DND Emails') or []
                for item in raw_dnd:
                    if isinstance(item, dict):
                        email = str(item.get('Email', '')).strip()
                    else:
                        email = str(item).strip()
                    if email and email.lower() not in ['', 'nan', 'none']:
                        if email not in dnd_emails_set:
                            dnd_emails_set.add(email)
                            dnd_list.append({'Email': email, 'DND': 'Yes'})
                            new_emails_count += 1
                            self.logger.debug(f"Added NEW DND email from processing stats: {email}")
            
            # Create DataFrame
            if dnd_list:
                dnd_df = pd.DataFrame(dnd_list)
            else:
                dnd_df = pd.DataFrame(columns=['Email', 'DND'])
            
            # Save to sheet named 'DND Emails'
            try:
                dnd_df.to_excel(writer, sheet_name='DND Emails', index=False)
                self.auto_adjust_columns(writer, dnd_df, 'DND Emails')
                self.logger.info(f"DND Emails sheet created with {len(dnd_df)} total entries (added {new_emails_count} new emails)")
            except Exception as e:
                self.logger.warning(f"Could not write DND Emails sheet: {e}")

        except Exception as e:
            self.logger.error(f"Error creating DND Emails sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def extract_dnd_emails_from_file(self, email_file):
        """Extract DND emails from the email file - DISABLED (not used)"""
        # This method is not called and requires BeautifulSoup which may not be available
        # For future use, implement without external HTML parsing library
        return []

    def create_sms_replies_sheet(self, writer, df, sms_file=None, processing_stats=None, prev_file=None, issue_not_resolved_cases=None):
        """Create Issue Not Fixed sheet by preserving ALL previous data and only appending new cases"""
        try:
            self.logger.info("Creating Issue Not Fixed sheet...")
            
            # Use a dictionary to store cases, keyed by Case Number, to easily handle updates and ensure uniqueness
            # This will prioritize previous data if a case number exists in both previous and new data
            final_issue_not_fixed_cases_map = {}

            # STEP 1: Load and preserve ALL previous Issue Not Fixed tab data FIRST
            self.logger.info(f"Attempting to load from prev_file: {prev_file}")
            if prev_file and os.path.exists(prev_file):
                self.logger.info(f"prev_file exists: {prev_file}")
                try:
                    self.logger.info(f"Loading previous Issue Not Fixed tab from: {prev_file}")
                    # Read the previous output file to get the Issue Not Fixed tab
                    prev_excel = pd.ExcelFile(prev_file)
                    
                    self.logger.info(f"Found sheets in prev_file: {prev_excel.sheet_names}")
                    
                    if 'Issue Not Fixed' in prev_excel.sheet_names:
                        prev_sms = pd.read_excel(prev_file, sheet_name='Issue Not Fixed')
                        self.logger.info(f"Previous Issue Not Fixed tab loaded with {len(prev_sms)} records")
                        
                        if 'Case Number' in prev_sms.columns:
                            # Add ALL previous cases to our map (preserve everything exactly as it was)
                            for idx, row in prev_sms.iterrows():
                                case_num = str(row['Case Number']).strip()
                                if case_num and case_num.lower() not in ['nan', 'none', '']:
                                    final_issue_not_fixed_cases_map[case_num] = {
                                        'Case Number': case_num,
                                        'Incoming Channel': row.get('Incoming Channel', ''),
                                        'Date Added': row.get('Date Added', ''), # Keep original date
                                        'Actioned (Y/N)': row.get('Actioned (Y/N)', ''),
                                        'Action Note': row.get('Action Note', ''),
                                        'Source': 'Previous File' # Internal tracking
                                    }
                            self.logger.info(f"Preserved {len(final_issue_not_fixed_cases_map)} cases from previous Issue Not Fixed tab")
                        else:
                            self.logger.warning("Previous file does not contain 'Case Number' column")
                    else:
                        self.logger.info("Previous output file does not contain 'Issue Not Fixed' tab")

                except Exception as e:
                    self.logger.error(f"Error loading previous Issue Not Fixed tab: {str(e)}")
                    self.logger.warning("Continuing without previous file data")
            else:
                self.logger.info("No previous output file found - starting with empty Issue Not Fixed tab")

            # Get today's date for new cases
            today_date_str = datetime.now().strftime('%Y-%m-%d')
            
            # STEP 2: Add new cases from SMS/Email processing (if not already in map)
            new_cases_from_sms_email = 0
            
            # Get SMS Issue Not Fixed cases
            if processing_stats and 'SMS Processing Stats' in processing_stats:
                sms_stats = processing_stats['SMS Processing Stats']
                if 'Issue Not Fixed' in sms_stats:
                    sms_issue_not_fixed_cases = sms_stats['Issue Not Fixed']
                    self.logger.info(f"Found {len(sms_issue_not_fixed_cases)} cases marked as 'Issue Not Fixed' via SMS processing")

                    for case_data in sms_issue_not_fixed_cases:
                        # Handle both old format (string) and new format (dict)
                        if isinstance(case_data, dict):
                            case_num_str = str(case_data.get('case_number', '')).strip()
                        else:
                            case_num_str = str(case_data).strip()
                            
                        if case_num_str and case_num_str not in final_issue_not_fixed_cases_map:
                            # Get Incoming Channel from the main df if available
                            incoming_channel = ''
                            if 'Incoming Channel' in df.columns:
                                case_row = df[df['Case Number'] == case_num_str]
                                if not case_row.empty:
                                    incoming_channel = case_row.iloc[0].get('Incoming Channel', '')

                            # This is a truly NEW case - add it with today's date
                            final_issue_not_fixed_cases_map[case_num_str] = {
                                'Case Number': case_num_str,
                                'Incoming Channel': incoming_channel,
                                'Date Added': today_date_str,
                                'Actioned (Y/N)': '',
                                'Action Note': '',
                                'Source': 'SMS Processing' # Internal tracking
                            }
                            new_cases_from_sms_email += 1
                            self.logger.info(f"Added NEW case from SMS processing: {case_num_str} with today's date")
                        else:
                            self.logger.debug(f"Skipped existing case from SMS processing: {case_num_str} (already in tab)")

            # Get Email Issue Not Fixed cases
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                if 'Issue Not Fixed' in email_stats:
                    email_issue_not_fixed_cases = email_stats['Issue Not Fixed']
                    self.logger.info(f"Found {len(email_issue_not_fixed_cases)} cases marked as 'Issue Not Fixed' via Email processing")

                    for case_data in email_issue_not_fixed_cases:
                        # Handle both old format (string) and new format (dict)
                        if isinstance(case_data, dict):
                            case_num_str = str(case_data.get('case_number', '')).strip()
                        else:
                            case_num_str = str(case_data).strip()
                            
                        if case_num_str and case_num_str not in final_issue_not_fixed_cases_map:
                            # Get Incoming Channel from the main df if available
                            incoming_channel = ''
                            if 'Incoming Channel' in df.columns:
                                case_row = df[df['Case Number'] == case_num_str]
                                if not case_row.empty:
                                    incoming_channel = case_row.iloc[0].get('Incoming Channel', '')

                            # This is a truly NEW case - add it with today's date
                            final_issue_not_fixed_cases_map[case_num_str] = {
                                'Case Number': case_num_str,
                                'Incoming Channel': incoming_channel,
                                'Date Added': today_date_str,
                                'Actioned (Y/N)': '',
                                'Action Note': '',
                                'Source': 'Email Processing' # Internal tracking
                            }
                            new_cases_from_sms_email += 1
                            self.logger.info(f"Added NEW case from Email processing: {case_num_str} with today's date")
                        else:
                            self.logger.debug(f"Skipped existing case from Email processing: {case_num_str} (already in tab)")

            # STEP 3: Add new cases from PA Cases (only if not already in map)
            new_cases_from_pa = 0
            issue_not_fixed_in_pa = df[df['Final Action'].str.lower() == 'issue not fixed'.lower()]
            self.logger.info(f"Found {len(issue_not_fixed_in_pa)} cases with Final Action = 'Issue Not Fixed' in PA Cases")

            for idx, row in issue_not_fixed_in_pa.iterrows():
                case_num = str(row['Case Number']).strip()
                if case_num and case_num not in final_issue_not_fixed_cases_map:
                    # This is a truly NEW case - add it with today's date
                    final_issue_not_fixed_cases_map[case_num] = {
                        'Case Number': case_num,
                        'Incoming Channel': row.get('Incoming Channel', ''),
                        'Date Added': today_date_str,
                        'Actioned (Y/N)': '',
                        'Action Note': '',
                        'Source': 'PA Cases' # Internal tracking
                    }
                    new_cases_from_pa += 1
                    self.logger.info(f"Added NEW case from PA Cases: {case_num} with today's date")
                else:
                    self.logger.debug(f"Skipped existing case from PA Cases: {case_num} (already in tab)")
            
            # STEP 3.5: Add cases marked as "Issue Not Resolved" by handlers in their individual sheets
            new_cases_from_handlers = 0
            if issue_not_resolved_cases:
                self.logger.info(f"Found {len(issue_not_resolved_cases)} cases marked as 'Issue Not Resolved' by handlers in their individual sheets")
                
                for case_data in issue_not_resolved_cases:
                    case_num = str(case_data.get('Case Number', '')).strip()
                    if case_num and case_num not in final_issue_not_fixed_cases_map:
                        # This is a NEW case from handler manual updates - add it with today's date
                        final_issue_not_fixed_cases_map[case_num] = {
                            'Case Number': case_num,
                            'Incoming Channel': 'Handler Manual Update',
                            'Date Added': today_date_str,
                            'Actioned (Y/N)': '',
                            'Action Note': f"Handler: {case_data.get('Handler', 'Unknown')} - {case_data.get('Final Action', '')}",
                            'Source': 'Handler Manual Update' # Internal tracking
                        }
                        new_cases_from_handlers += 1
                        self.logger.info(f"Added NEW case from Handler Manual Update: {case_num} (Handler: {case_data.get('Handler', 'Unknown')})")
                    else:
                        self.logger.debug(f"Skipped existing case from Handler Manual Update: {case_num} (already in tab)")
            else:
                self.logger.info("No cases marked as 'Issue Not Resolved' by handlers found")
            
            # STEP 4: Create final DataFrame
            if final_issue_not_fixed_cases_map:
                # Convert map values to a list of dictionaries
                all_issue_not_fixed_cases_list = list(final_issue_not_fixed_cases_map.values())
                # Create DataFrame with ALL fields including Source
                sms_df = pd.DataFrame(all_issue_not_fixed_cases_list)
                
                # Ensure proper column order: Case Number, Incoming Channel, Date Added, Actioned (Y/N), Action Note
                # Note: Source field is kept for internal processing but not shown in final output
                column_order = ['Case Number', 'Incoming Channel', 'Date Added', 'Actioned (Y/N)', 'Action Note']
                sms_df = sms_df[column_order]

                # Log comprehensive summary
                self.logger.info(f"\n=== Issue Not Fixed Sheet Summary ===")
                self.logger.info(f"Total cases in sheet: {len(sms_df)}")

                # Count cases by source using the 'Source' field from the original map values
                prev_file_count = sum(1 for case_data in final_issue_not_fixed_cases_map.values() if case_data.get('Source') == 'Previous File')
                new_cases_count = sum(1 for case_data in final_issue_not_fixed_cases_map.values() if case_data.get('Source') != 'Previous File')
                
                self.logger.info(f"Cases preserved from previous file: {prev_file_count}")
                self.logger.info(f"NEW cases added (today's date): {new_cases_count}")
                self.logger.info(f"  - From SMS/Email processing: {new_cases_from_sms_email}")
                self.logger.info(f"  - From PA Cases: {new_cases_from_pa}")
                self.logger.info(f"  - From Handler Manual Updates: {new_cases_from_handlers}")
                
                # Log date distribution
                if 'Date Added' in sms_df.columns:
                    date_counts = sms_df['Date Added'].value_counts()
                    self.logger.info(f"Date distribution:")
                    for date, count in date_counts.items():
                        self.logger.info(f"  {date}: {count} cases")
                
                # Log action status
                if 'Actioned (Y/N)' in sms_df.columns:
                    action_counts = sms_df['Actioned (Y/N)'].value_counts()
                    self.logger.info(f"Action status distribution:")
                    for action, count in action_counts.items():
                        if _is_missing(action):
                            self.logger.info(f"  Not Actioned: {count} cases")
                        else:
                            self.logger.info(f"  {action}: {count} cases")
                
                # Save to sheet
                sms_df.to_excel(writer, sheet_name='Issue Not Fixed', index=False)
                
                # Auto-adjust columns for Issue Not Fixed sheet
                self.auto_adjust_columns(writer, sms_df, 'Issue Not Fixed')
                
                self.logger.info(f"Issue Not Fixed sheet created successfully with {len(sms_df)} cases")
                
                # Log sample cases for verification
                sample_df = sms_df.head(5)[['Case Number', 'Incoming Channel', 'Date Added', 'Actioned (Y/N)']]
                self.logger.info(f"Sample cases in sheet:")
                for idx, row in sample_df.iterrows():
                    self.logger.info(f"  Case {row['Case Number']}: Channel {row['Incoming Channel']}, Date {row['Date Added']}, Actioned {row['Actioned (Y/N)']}")
                
            else:
                # Create empty sheet with headers
                empty_df = pd.DataFrame(columns=['Case Number', 'Incoming Channel', 'Date Added', 'Actioned (Y/N)', 'Action Note'])
                empty_df.to_excel(writer, sheet_name='Issue Not Fixed', index=False)
                
                # Auto-adjust columns for empty Issue Not Fixed sheet
                self.auto_adjust_columns(writer, empty_df, 'Issue Not Fixed')
                
                self.logger.info("Created empty Issue Not Fixed sheet (no cases found)")
            
        except Exception as e:
            self.logger.error(f"Error creating Issue Not Fixed sheet: {str(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def count_invalid_emails(self, df):
        try:
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if 'Email' not in df.columns:
                return 0
            invalid_emails = df[(df['Email'] != '') & (~df['Email'].str.match(email_pattern, na=False))].shape[0]
            return invalid_emails
        except Exception as e:
            self.logger.error(f"Error counting invalid emails: {str(e)}")
            return 0

    def create_counters_sheet(self, writer, df, prev_df=None, prev_file=None):
        try:
            self.logger.info("Creating counters sheet (reading from PA Cases main sheet)...")
            # Gather all unique handlers from current output
            handler_set = set(df['Assigned To'].dropna().astype(str).str.strip())
            # If previous file is available, include its handlers too (from main sheet)
            if prev_df is not None:
                prev_handlers = set(prev_df['Assigned To'].dropna().astype(str).str.strip())
                handler_set.update(prev_handlers)
            # Also include handlers from previous individual handler sheets regardless of selection
            if prev_file and os.path.exists(prev_file):
                try:
                    prev_excel = pd.ExcelFile(prev_file)
                    handler_sheets = [sheet for sheet in prev_excel.sheet_names if isinstance(sheet, str) and sheet.endswith("'s Cases")]
                    sheet_handlers = [str(sheet).replace("'s Cases", '').strip() for sheet in handler_sheets]
                    handler_set.update(sheet_handlers)
                except Exception as e:
                    self.logger.warning(f"Could not read previous handler sheets for counter: {str(e)}")
            # Remove empty handler names
            handler_list = sorted([h for h in handler_set if h and h.lower() != 'nan'])
            
            # Check if we have handlers to process
            if not handler_list:
                self.logger.warning("No handlers found to create counter sheet")
                return
                
            progress_columns = [
                ("closed",    ("$R$", "$B$2")),
                ("in_progress", ("$R$", "$C$2")),
                ("new",        ("$R$", "$D$2")),
                ("Skipped", ("$Q$", "$E$2")),
                ("In Progress Today", ("$Q$", "$G$2")),
                ("Requires a Call", None),  # Special: Final Action='Sent Email' AND Status='in_progress'
                ("Total",      None),
            ]
            worksheet = writer.book.add_worksheet('Counter')
            writer.sheets['Counter'] = worksheet
            worksheet.write(0, 0, 'Progress Counter')
            # Write header
            worksheet.write(1, 0, 'Handler')
            for col_idx, (col_name, _) in enumerate(progress_columns, 1):
                worksheet.write(1, col_idx, col_name)
            
            # Initialize row_idx to ensure it's always defined
            row_idx = 1
            
            # Write handler rows with formulas
            for row_idx, handler in enumerate(handler_list, 2):
                worksheet.write(row_idx, 0, handler)
                for col_idx, (col_name, col_formula) in enumerate(progress_columns, 1):
                    if col_name == "Total":
                        worksheet.write_formula(row_idx, col_idx, f"=SUM(B{row_idx+1}:G{row_idx+1})")
                    elif col_name == "Requires a Call":
                        # Count cases with Final Action = 'Sent Email' AND Status = 'in_progress'
                        handler_sheet_name = f"{handler}'s Cases"
                        escaped_sheet_name = handler_sheet_name.replace("'", "''")
                        excel_sheet_name = f"'{escaped_sheet_name}'"
                        # P = Final Action, R = Status
                        formula = f"=COUNTIFS({excel_sheet_name}!P:P,\"Sent Email\",{excel_sheet_name}!R:R,\"in_progress\")"
                        worksheet.write_formula(row_idx, col_idx, formula)
                    else:
                        # Create dynamic handler sheet name with proper Excel escaping
                        handler_sheet_name = f"{handler}'s Cases"
                        # Excel requires double single quotes to escape single quotes in sheet names
                        escaped_sheet_name = handler_sheet_name.replace("'", "''")
                        excel_sheet_name = f"'{escaped_sheet_name}'"
                        
                        if col_name == "closed":
                            # Updated formulas using new column indices
                            # Status is column J (10th column) in new layout -> R
                            # Assigned To is column I (9th column) in new layout -> Q 
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$B$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "in_progress":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$C$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "new":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$D$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "Skipped":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$E$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        elif col_name == "In Progress Today":
                            formula = f"=COUNTIFS({excel_sheet_name}!R:R,Counter!$F$2,{excel_sheet_name}!Q:Q,Counter!A{row_idx+1})"
                        else:
                            formula = ""
                        worksheet.write_formula(row_idx, col_idx, formula)


            # Write totals row
            worksheet.write(row_idx+1, 0, 'Total')
            for col_idx in range(1, len(progress_columns)+1):
                worksheet.write_formula(row_idx+1, col_idx, f"=SUM({chr(66+col_idx-1)}3:{chr(66+col_idx-1)}{row_idx+1})")
            self.logger.info("Progress Counter table created with Excel formulas (reading from individual handler sheets)")

            # --- Final Action Counter Table ---
            final_action_options = [
                            "Fixed", "Refused Callback", "Issue Not Fixed", "Not yet Tested", "Escalation", "Sent SMS", "Sent Email", "Left VM", "Reviewed", "Bank/Sutherland", "Not Reached", "DND"

            ]
            start_row = row_idx + 3
            worksheet.write(start_row, 0, 'Final Action Counter')
            worksheet.write(start_row+1, 0, 'Handler')
            for col_idx, action in enumerate(final_action_options, 1):
                worksheet.write(start_row+1, col_idx, action)
            
            # Initialize h_idx to ensure it's always defined
            h_idx = start_row + 1
            
            for h_idx, handler in enumerate(handler_list, start_row+2):
                worksheet.write(h_idx, 0, handler)
                for a_idx, action in enumerate(final_action_options, 1):
                    # Create dynamic handler sheet name with proper Excel escaping
                    handler_sheet_name = f"{handler}'s Cases"
                    # Excel requires double single quotes to escape single quotes in sheet names
                    escaped_sheet_name = handler_sheet_name.replace("'", "''")
                    excel_sheet_name = f"'{escaped_sheet_name}'"
                    # Read from individual handler sheet with proper COUNTIFS format
                    formula = f"=COUNTIFS({excel_sheet_name}!P:P,Counter!{chr(65+a_idx)}{start_row+2},{excel_sheet_name}!Q:Q,Counter!A{h_idx+1})"
                    worksheet.write_formula(h_idx, a_idx, formula)
            
            # Write totals row for final action counter
            worksheet.write(h_idx+1, 0, 'Total')
            for a_idx in range(1, len(final_action_options)+1):
                worksheet.write_formula(h_idx+1, a_idx, f"=SUM({chr(65+a_idx)}{start_row+3}:{chr(65+a_idx)}{h_idx+1})")
            self.logger.info("Final Action Counter table created with Excel formulas (reading from individual handler sheets)")
            
            # --- Company Cases Table ---
            # Shows handler progress specifically in Companies sheet
            company_start_row = h_idx + 4
            worksheet.write(company_start_row, 0, 'Company Cases')
            worksheet.write(company_start_row+1, 0, 'Handler')
            worksheet.write(company_start_row+1, 1, 'New')
            worksheet.write(company_start_row+1, 2, 'Closed')
            worksheet.write(company_start_row+1, 3, 'Total')
            
            # Initialize company row counter
            co_idx = company_start_row + 1
            
            # Only create this table if Companies sheet exists
            if 'Companies' in writer.sheets:
                for co_idx, handler in enumerate(handler_list, company_start_row+2):
                    worksheet.write(co_idx, 0, handler)
                    # New: Count where Status='new' AND Assigned To = handler in Companies
                    # Assigned To is column Q (index 16), Status is column R (index 17)
                    worksheet.write_formula(co_idx, 1, f"=COUNTIFS('Companies'!R:R,\"new\",'Companies'!Q:Q,Counter!A{co_idx+1})")
                    # Closed: Count where Status='closed' AND Assigned To = handler in Companies
                    worksheet.write_formula(co_idx, 2, f"=COUNTIFS('Companies'!R:R,\"closed\",'Companies'!Q:Q,Counter!A{co_idx+1})")
                    # Total: Sum of New + Closed for this handler
                    worksheet.write_formula(co_idx, 3, f"=SUM(B{co_idx+1}:C{co_idx+1})")
                
                # Write totals row for Company Cases
                worksheet.write(co_idx+1, 0, 'Total')
                worksheet.write_formula(co_idx+1, 1, f"=SUM(B{company_start_row+3}:B{co_idx+1})")
                worksheet.write_formula(co_idx+1, 2, f"=SUM(C{company_start_row+3}:C{co_idx+1})")
                worksheet.write_formula(co_idx+1, 3, f"=SUM(D{company_start_row+3}:D{co_idx+1})")
                self.logger.info("Company Cases table created with Excel formulas (reading from Companies sheet)")
        except Exception as e:
            self.logger.error(f"Error creating counters sheet: {str(e)}")

    def create_summary_sheet(self, writer, df, processing_stats=None):
        try:
            self.logger.info("Creating summary sheet...")

            
            summary_data = []
            
            # --- Processing Statistics Section ---
            if processing_stats:
                summary_data.extend([
                    {'Category': '--- Processing Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
                    {'Category': 'Raw File', 'Metric': 'Initial Count', 'Count': processing_stats.get('Initial Count', 0), 'Description': 'Total records in raw input file'},
                    {'Category': 'Work Order Status Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Work Order Status', 0), 'Description': 'Records remaining after excluding cancelled work orders'},
                    {'Category': 'Case Reason Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Case Reason', 0), 'Description': 'Records remaining after excluding escalation/complaint cases'},
                    {'Category': 'Closing Code Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Closing Code', 0), 'Description': 'Records remaining after excluding specific closing codes'},
                    {'Category': 'CID/DMR Filter', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After CID/DMR Filter', 0), 'Description': 'Records remaining after CID/DMR filtering'},
                    {'Category': 'Duplicate Removal', 'Metric': 'Records After Filter', 'Count': processing_stats.get('After Duplicate Removal', 0), 'Description': 'Records remaining after removing duplicate case numbers'},
                    {'Category': 'Raw File', 'Metric': 'Final Count After Cleaning', 'Count': processing_stats.get('Raw File Final Count After Cleaning', ''), 'Description': 'Final count of raw file after all cleaning'},
                    {'Category': 'Final Output', 'Metric': 'Final Count', 'Count': processing_stats.get('Final Count', 0), 'Description': 'Final record count in output file'},
                    {'Category': 'Total Removed', 'Metric': 'Records Removed', 'Count': processing_stats.get('Total Removed', 0), 'Description': 'Total records removed during all filtering steps'},
                ])
            
            # --- Previous File Statistics Section ---
            prev_stats = processing_stats.get('Previous File Stats', {}) if processing_stats else {}
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Previous File Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': 'Previous File', 'Metric': 'Initial Count', 'Count': prev_stats.get('Initial Count', ''), 'Description': 'Total records in previous file'},
                {'Category': 'Matching Cases', 'Metric': 'Count', 'Count': prev_stats.get('Matching Cases', ''), 'Description': 'Cases that exist in both previous and current files'},
                {'Category': 'Updated Cases', 'Metric': 'Count', 'Count': prev_stats.get('Updated Cases', ''), 'Description': 'Cases updated with new information from raw file'},
                {'Category': 'New Cases', 'Metric': 'Count', 'Count': prev_stats.get('New Cases', ''), 'Description': 'New cases added from raw file'},
            ])
            
            # --- Dropped Cases Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Dropped Cases from Raw File Filtering ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            
            # Add dropped cases statistics if available
            if processing_stats:
                # Work Order Status Filter dropped cases
                initial_count = processing_stats.get('Initial Count', 0)
                after_wo_status = processing_stats.get('After Work Order Status', 0)
                if initial_count and after_wo_status:
                    dropped_wo_status = initial_count - after_wo_status
                    summary_data.append({
                        'Category': 'Work Order Status Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_wo_status,
                        'Description': 'Cases dropped due to cancelled work orders'
                    })
                
                # Case Reason Filter dropped cases
                after_case_reason = processing_stats.get('After Case Reason', 0)
                if after_wo_status and after_case_reason:
                    dropped_case_reason = after_wo_status - after_case_reason
                    summary_data.append({
                        'Category': 'Case Reason Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_case_reason,
                        'Description': 'Cases dropped due to escalation/complaint reasons'
                    })
                
                # Closing Code Filter dropped cases
                after_closing_code = processing_stats.get('After Closing Code', 0)
                if after_case_reason and after_closing_code:
                    dropped_closing_code = after_case_reason - after_closing_code
                    summary_data.append({
                        'Category': 'Closing Code Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_closing_code,
                        'Description': 'Cases dropped due to specific closing codes'
                    })
                
                # CID/DMR Filter dropped cases
                after_cid_dmr = processing_stats.get('After CID/DMR Filter', 0)
                if after_closing_code and after_cid_dmr:
                    dropped_cid_dmr = after_closing_code - after_cid_dmr
                    summary_data.append({
                        'Category': 'CID/DMR Filter',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_cid_dmr,
                        'Description': 'Cases dropped due to CID/DMR filtering'
                    })
                
                # Duplicate Removal dropped cases
                after_duplicate = processing_stats.get('After Duplicate Removal', 0)
                if after_cid_dmr and after_duplicate:
                    dropped_duplicate = after_cid_dmr - after_duplicate
                    summary_data.append({
                        'Category': 'Duplicate Removal',
                        'Metric': 'Dropped Cases',
                        'Count': dropped_duplicate,
                        'Description': 'Cases dropped due to duplicate case numbers'
                    })
            
            # --- Business Rules Updates Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Business Rules Updates ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            
            # Bank/Sutherland Rule Updates
            if 'Status' in df.columns and 'Action 1' in df.columns:
                bank_sutherland_updated = df[(df['Status'].str.lower() == 'closed') & (df['Action 1'] == 'Bank/Sutherland')].shape[0]
                summary_data.append({
                    'Category': 'Bank/Sutherland Rule',
                    'Metric': 'Cases Updated',
                    'Count': bank_sutherland_updated,
                    'Description': 'Cases updated with Status=closed and Action 1=Bank/Sutherland'
                })
            
            # DND Rule Updates
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_updated = df[(df['Status'] == 'Closed') | (df['Action 1'] == 'DND') | (df['Action 2'] == 'DND') | (df['Action 3'] == 'DND') | (df['Final Action'] == 'DND')].shape[0]
                summary_data.append({
                    'Category': 'DND Rule',
                    'Metric': 'Cases Updated',
                    'Count': dnd_updated,
                    'Description': 'Cases updated with DND status or actions'
                })
            
            # --- SMS Processing Updates Section ---
            if processing_stats and 'SMS Processing Stats' in processing_stats:
                sms_stats = processing_stats['SMS Processing Stats']
                if sms_stats:
                    summary_data.extend([
                        {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': '--- SMS Processing Updates ---', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': 'SMS Fixed', 'Metric': 'Cases Updated', 'Count': len(sms_stats.get('Fixed', [])), 'Description': 'Cases updated to Fixed status via SMS'},
                        {'Category': 'SMS Issue Not Fixed', 'Metric': 'Cases Updated', 'Count': len(sms_stats.get('Issue Not Fixed', [])), 'Description': 'Cases updated to Issue Not Fixed status via SMS'},
                        {'Category': 'SMS Refused Callback', 'Metric': 'Cases Updated', 'Count': len(sms_stats.get('Refused Callback', [])), 'Description': 'Cases updated to Refused Callback status via SMS'},
                        {'Category': 'SMS Skipped Cases', 'Metric': 'Cases Skipped', 'Count': len(sms_stats.get('Skipped Cases', [])), 'Description': 'SMS cases that were skipped (see Skipped SMS Cases sheet)'},
                        {'Category': 'SMS Ambiguous Cases', 'Metric': 'Cases for Review', 'Count': len(sms_stats.get('Ambiguous', [])), 'Description': 'SMS cases with unclear responses needing human review'},
                    ])
            
            # --- Email Processing Updates Section ---
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                if email_stats:
                    summary_data.extend([
                        {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': '--- Email Processing Updates ---', 'Metric': '', 'Count': '', 'Description': ''},
                        {'Category': 'Email Fixed', 'Metric': 'Cases Updated', 'Count': len(email_stats.get('Fixed', [])), 'Description': 'Cases updated to Fixed status via Email'},
                        {'Category': 'Email Issue Not Fixed', 'Metric': 'Cases Updated', 'Count': len(email_stats.get('Issue Not Fixed', [])), 'Description': 'Cases updated to Issue Not Fixed status via Email'},
                        {'Category': 'Email DND', 'Metric': 'Cases Updated', 'Count': len(email_stats.get('DND', [])), 'Description': 'Cases updated to DND status via Email'},
                        {'Category': 'Email Skipped Cases', 'Metric': 'Cases Skipped', 'Count': len(email_stats.get('Skipped Emails', [])), 'Description': 'Email replies that were skipped (see Skipped Email Replies sheet)'},
                        {'Category': 'Email Ambiguous Cases', 'Metric': 'Cases for Review', 'Count': len(email_stats.get('Ambiguous', [])), 'Description': 'Email replies with unclear responses needing human review'},
                    ])
            
            # --- Data Quality Checks Section ---
            # Flexible customer name handling: detect customer column variants
            cust_col = self._resolve_col(df, 'Customer Name')
            missing_customer_count = df[df[cust_col].isna() | (df[cust_col] == '')].shape[0] if cust_col and cust_col in df.columns else 0
            
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Data Quality Checks ---', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': 'Missing Case Numbers', 'Metric': '', 'Count': df[df['Case Number'].isna() | (df['Case Number'] == '')].shape[0], 'Description': 'Cases without case numbers'},
                {'Category': 'Missing Customer Names', 'Metric': '', 'Count': missing_customer_count, 'Description': 'Cases without customer names'},
                {'Category': 'Missing Handler Assignments', 'Metric': '', 'Count': df[df['Assigned To'].isna() | (df['Assigned To'] == '')].shape[0], 'Description': 'Cases without handler assignments'},
                {'Category': 'Duplicate Case Numbers', 'Metric': '', 'Count': df[df.duplicated(subset=['Case Number'], keep=False)].shape[0], 'Description': 'Duplicate case numbers found'},
                {'Category': 'Invalid Email Formats', 'Metric': '', 'Count': self.count_invalid_emails(df), 'Description': 'Invalid email address formats'},
                {'Category': 'Empty Rows', 'Metric': '', 'Count': df[df.apply(lambda row: all(str(x).strip() == '' for x in row), axis=1)].shape[0], 'Description': 'Completely empty rows'},
            ])
            
            # --- Handler Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Handler Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            if 'Assigned To' in df.columns:
                handler_stats = df['Assigned To'].value_counts()
                for handler, count in handler_stats.items():
                    if handler and str(handler).strip():
                        summary_data.append({
                            'Category': 'Handler',
                            'Metric': handler,
                            'Count': count,
                            'Description': f'Total cases assigned to {handler}'
                        })
            
            # --- Bank/Sutherland Rule Closed Cases Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Bank/Sutherland Rule Closed Cases ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            # Use the same patterns as in the processor
            import re
            patterns = [
                ("Citi", r"\\bcitibank\\b|\\bcitigroup\\b|\\bciti group\\b|\\bciti bank\\b|\\bciti corp\\b|\\bciti\\b"),
                ("Amgen", r"amgen"),
                ("Sanofi", r"sanofi"),
                ("Scotiabank", r"bns|scotiabank|scotia bank|scotiabank canada|scotiaBank|bank of nova scotia|bank nova scotia"),
                ("Royal Bank", r"rbc|royal bank|royal bank of canada|rbc wealth management"),
                ("HSBC", r"hsbc"),
                ("TD Bank", r"td bank"),
            ]
            def matches_company(name, regex):
                if not isinstance(name, str):
                    return False
                name_lc = name.lower()
                if 'citizen' in name_lc and 'citi' in regex:
                    return False
                return re.search(regex, name_lc) is not None
            # Only count cases where Status == closed and Action 1 == Bank/Sutherland
            if 'Status' in df.columns and 'Action 1' in df.columns and 'Company Name' in df.columns:
                closed_mask = (df['Status'].str.lower() == 'closed') & (df['Action 1'] == 'Bank/Sutherland')
                for label, regex in patterns:
                    count = df[closed_mask & df['Company Name'].apply(lambda x: matches_company(x, regex))].shape[0]
                    summary_data.append({
                        'Category': 'Bank/Sutherland Rule',
                        'Metric': label,
                        'Count': count,
                        'Description': f'Closed cases due to Bank/Sutherland rule for {label}'
                    })
            
            # --- Status Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Status Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            if 'Status' in df.columns:
                status_stats = df['Status'].value_counts()
                for status, count in status_stats.items():
                    if status and str(status).strip():
                        summary_data.append({
                            'Category': 'Status',
                            'Metric': status,
                            'Count': count,
                            'Description': f'Cases with status: {status}'
                        })
            
            # --- Final Action Statistics Section ---
            summary_data.extend([
                {'Category': '', 'Metric': '', 'Count': '', 'Description': ''},
                {'Category': '--- Final Action Statistics ---', 'Metric': '', 'Count': '', 'Description': ''},
            ])
            if 'Final Action' in df.columns:
                final_action_stats = df['Final Action'].value_counts()
                for action, count in final_action_stats.items():
                    if action and str(action).strip():
                        summary_data.append({
                            'Category': 'Final Action',
                            'Metric': action,
                            'Count': count,
                            'Description': f'Cases with final action: {action}'
                        })
            
            # Main summary data is complete - will add detailed case table separately
            
            # Dropped cases will be added to detailed table
            
            # SMS processing cases will be added to detailed table
            
            # Email processing cases will be added to detailed table
            
            # Add detailed case information for Bank/Sutherland updates
            if 'Status' in df.columns and 'Action 1' in df.columns and 'Company Name' in df.columns:
                bank_sutherland_cases = df[(df['Status'].str.lower() == 'closed') & (df['Action 1'] == 'Bank/Sutherland')]
                if not bank_sutherland_cases.empty:
                    case_numbers = bank_sutherland_cases['Case Number'].astype(str).tolist()
                    summary_data.append({
                        'Category': 'Bank/Sutherland Updates - Case Numbers',
                        'Metric': 'Cases',
                        'Count': len(case_numbers),
                        'Description': f'Cases updated with Bank/Sutherland rule: {", ".join(case_numbers[:10])}{"..." if len(case_numbers) > 10 else ""}'
                    })
            
            # Add detailed case information for DND updates
            if 'Status' in df.columns and 'Action 1' in df.columns:
                dnd_cases = df[(df['Status'] == 'DND') | (df['Action 1'] == 'DND') | (df['Action 2'] == 'DND') | (df['Action 3'] == 'DND') | (df['Final Action'] == 'DND')]
                if not dnd_cases.empty:
                    case_numbers = dnd_cases['Case Number'].astype(str).tolist()
                    summary_data.append({
                        'Category': 'DND Updates - Case Numbers',
                        'Metric': 'Cases',
                        'Count': len(case_numbers),
                        'Description': f'Cases updated with DND status/actions: {", ".join(case_numbers[:10])}{"..." if len(case_numbers) > 10 else ""}'
                    })
            
            # Create the summary DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # Add the detailed case table data to the summary DataFrame
            detailed_cases = []
            
            # Add dropped cases
            if processing_stats and 'Dropped Cases Details' in processing_stats:
                dropped_details = processing_stats['Dropped Cases Details']
                for filter_name, cases in dropped_details.items():
                    if cases:
                        for case_num in cases:
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'dropped',
                                'Reason': f"case reason: {filter_name}",
                                'Updated to': ''
                            })
            
            # Add updated cases from SMS processing
            if processing_stats and 'SMS Processing Stats' in processing_stats:
                sms_stats = processing_stats['SMS Processing Stats']
                self.logger.info(f"SMS stats found: {list(sms_stats.keys()) if sms_stats else 'None'}")
                for action_type in ['Fixed', 'Issue Not Fixed', 'Refused Callback', 'DND']:
                    if action_type in sms_stats and sms_stats[action_type]:
                        cases = sms_stats[action_type]
                        self.logger.info(f"SMS {action_type} cases found: {len(cases)}")
                        for case_data in cases:
                            # Handle both old format (string) and new format (dict)
                            if isinstance(case_data, dict):
                                case_num = case_data.get('case_number', '')
                                sms_text = case_data.get('sms_text', '')
                            else:
                                case_num = str(case_data)
                                sms_text = ''
                            
                            if case_num:
                                # Reason - include SMS text if available
                                if sms_text:
                                    reason = f"SMS: '{action_type}' - '{sms_text[:100]}{'...' if len(sms_text) > 100 else ''}'"
                                else:
                                    reason = f"SMS: '{action_type}'"
                                
                                detailed_cases.append({
                                    'Case Number': str(case_num),
                                    'Action': 'updated',
                                    'Reason': reason,
                                    'Updated to': action_type
                                })
                                self.logger.info(f"Added SMS case to summary: {case_num} - {action_type}")
                    else:
                        self.logger.info(f"SMS {action_type} cases not found or empty")
                
                # Add Ambiguous SMS cases for review
                if 'Ambiguous' in sms_stats and sms_stats['Ambiguous']:
                    ambiguous_cases = sms_stats['Ambiguous']
                    self.logger.info(f"SMS Ambiguous cases found: {len(ambiguous_cases)}")
                    for case_data in ambiguous_cases:
                        if isinstance(case_data, dict):
                            case_num = case_data.get('case_number', '')
                            sms_text = case_data.get('sms_text', '')
                        else:
                            case_num = str(case_data)
                            sms_text = ''
                        
                        if case_num:
                            # Reason - include SMS text if available
                            if sms_text:
                                reason = f"SMS: Ambiguous - needs review - '{sms_text[:100]}{'...' if len(sms_text) > 100 else ''}'"
                            else:
                                reason = f"SMS: Ambiguous - needs review"
                            
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'needs_review',
                                'Reason': reason,
                                'Updated to': 'Ambiguous'
                            })
                            self.logger.info(f"Added SMS Ambiguous case to summary: {case_num}")
                else:
                    self.logger.info("SMS Ambiguous cases not found or empty")
            
            # Add updated cases from Email processing
            if processing_stats and 'Email Processing Stats' in processing_stats:
                email_stats = processing_stats['Email Processing Stats']
                self.logger.info(f"Email stats found: {list(email_stats.keys()) if email_stats else 'None'}")
                for action_type in ['Fixed', 'Issue Not Fixed', 'DND']:
                    if action_type in email_stats and email_stats[action_type]:
                        cases = email_stats[action_type]
                        self.logger.info(f"Email {action_type} cases found: {len(cases)}")
                        for case_data in cases:
                            # Handle both old format (string) and new format (dict)
                            if isinstance(case_data, dict):
                                case_num = case_data.get('case_number', '')
                                reply_text = case_data.get('reply_text', '')
                            else:
                                case_num = str(case_data)
                                reply_text = ''
                            
                            if case_num:
                                # Reason - include email reply text if available
                                if reply_text:
                                    reason = f"Email: '{action_type}' - '{reply_text[:100]}{'...' if len(reply_text) > 100 else ''}'"
                                else:
                                    reason = f"Email: '{action_type}'"
                                
                                detailed_cases.append({
                                    'Case Number': str(case_num),
                                    'Action': 'updated',
                                    'Reason': reason,
                                    'Updated to': action_type
                                })
                                self.logger.info(f"Added Email case to summary: {case_num} - {action_type}")
                    else:
                        self.logger.info(f"Email {action_type} cases not found or empty")
                
                # Add Ambiguous Email cases for review
                if 'Ambiguous' in email_stats and email_stats['Ambiguous']:
                    ambiguous_cases = email_stats['Ambiguous']
                    self.logger.info(f"Email Ambiguous cases found: {len(ambiguous_cases)}")
                    for case_data in ambiguous_cases:
                        if isinstance(case_data, dict):
                            case_num = case_data.get('case_number', '')
                            reply_text = case_data.get('reply_text', '')
                        else:
                            case_num = str(case_data)
                            reply_text = ''
                        
                        if case_num:
                            # Reason - include email reply text if available
                            if reply_text:
                                reason = f"Email: Ambiguous - needs review - '{reply_text[:100]}{'...' if len(reply_text) > 100 else ''}'"
                            else:
                                reason = f"Email: Ambiguous - needs review"
                            
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'needs_review',
                                'Reason': reason,
                                'Updated to': 'Ambiguous'
                            })
                            self.logger.info(f"Added Email Ambiguous case to summary: {case_num}")
                else:
                    self.logger.info("Email Ambiguous cases not found or empty")
            else:
                self.logger.info("Email Processing Stats not found in processing stats")
            
            # Add Bank/Sutherland updated cases
            if processing_stats and 'Bank/Sutherland Updated Cases' in processing_stats:
                bank_sutherland_cases = processing_stats['Bank/Sutherland Updated Cases']
                self.logger.info(f"Bank/Sutherland cases found: {len(bank_sutherland_cases)}")
                if bank_sutherland_cases:
                    for case_data in bank_sutherland_cases:
                        case_num = str(case_data.get('case_number', ''))
                        company_name = str(case_data.get('company_name', ''))
                        if case_num:
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'updated',
                                'Reason': f"Bank/Sutherland rule applied - Bank: {company_name}",
                                'Updated to': 'Closed'
                            })
                            self.logger.info(f"Added Bank/Sutherland case to summary: {case_num} - {company_name}")
                else:
                    self.logger.info("No Bank/Sutherland cases to add")
            else:
                self.logger.info("Bank/Sutherland cases not found in processing stats")
            
            # Add DND updated cases
            if processing_stats and 'DND Updated Cases' in processing_stats:
                dnd_cases = processing_stats['DND Updated Cases']
                self.logger.info(f"DND cases found: {len(dnd_cases)}")
                if dnd_cases:
                    for case_data in dnd_cases:
                        case_num = str(case_data.get('case_number', ''))
                        source = str(case_data.get('source', 'DND rule'))
                        if case_num:
                            detailed_cases.append({
                                'Case Number': str(case_num),
                                'Action': 'updated',
                                'Reason': f"{source} applied",
                                'Updated to': 'DND'
                            })
                            self.logger.info(f"Added DND case to summary: {case_num} - {source}")
                else:
                    self.logger.info("No DND cases to add")
            # Now create the final summary DataFrame with the detailed cases table
            # First, create the main summary DataFrame
            self.logger.info("Creating main summary DataFrame...")
            main_summary_df = pd.DataFrame(summary_data)
            self.logger.info(f"Main summary DataFrame created with shape: {main_summary_df.shape}")
            
            # Create the detailed cases DataFrame
            if detailed_cases:
                # Remove duplicates based on Case Number and Action
                unique_cases = {}
                for case in detailed_cases:
                    key = (case['Case Number'], case['Action'])
                    if key not in unique_cases:
                        unique_cases[key] = case
                    else:
                        # If duplicate found, keep the one with more detailed reason
                        if len(case['Reason']) > len(unique_cases[key]['Reason']):
                            unique_cases[key] = case
                
                # Convert back to list
                detailed_cases = list(unique_cases.values())
                self.logger.info(f"Removed duplicates from detailed cases. Original: {len(detailed_cases) + len(unique_cases) - len(detailed_cases)}, Final: {len(detailed_cases)}")
                
                detailed_df = pd.DataFrame(detailed_cases)
                
                # Create a combined DataFrame with detailed cases FIRST, then main summary
                # Add empty columns to main summary to align with detailed cases
                for col in ['Case Number', 'Action', 'Reason', 'Updated to']:
                    if col not in main_summary_df.columns:
                        main_summary_df[col] = ''
                
                # Combine the DataFrames - detailed cases FIRST, then main summary
                combined_df = pd.concat([detailed_df, main_summary_df], ignore_index=True)
                
                # Write the combined DataFrame to Excel
                self.logger.info(f"About to write combined DataFrame to Excel with {len(combined_df)} rows")
                self.logger.info(f"Combined DataFrame columns: {list(combined_df.columns)}")
                self.logger.info(f"First few rows of combined DataFrame:")
                for i, row in combined_df.head(3).iterrows():
                    self.logger.info(f"  Row {i}: {dict(row)}")
                
                # Create DataFrame and save to sheet
                if combined_df is not None and not combined_df.empty:
                    combined_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Auto-adjust columns for Summary sheet
                    self.auto_adjust_columns(writer, combined_df, 'Summary')
                    
                    self.logger.info(f"Summary sheet created with {len(combined_df)} rows")
                else:
                    # Create empty summary sheet
                    main_summary_df = pd.DataFrame(columns=['Category', 'Metric', 'Count', 'Description'])
                    main_summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Auto-adjust columns for empty Summary sheet
                    self.auto_adjust_columns(writer, main_summary_df, 'Summary')
                    
                    self.logger.info("Created empty Summary sheet")
            else:
                # No detailed cases, just write the main summary
                main_summary_df.to_excel(writer, sheet_name='Summary', index=False)
                self.logger.info(f"Summary sheet created with {len(main_summary_df)} summary items (no detailed cases)")
            
            self.logger.info(f"Detailed case table added to Summary sheet starting from column J, row 1 (at the TOP)")
            
            self.logger.info("=== SUMMARY SHEET CREATION COMPLETED SUCCESSFULLY ===")
        except Exception as e:
            self.logger.error(f"=== ERROR in create_summary_sheet: {str(e)} ===")
            self.logger.error(f"Exception type: {type(e)}")
            import traceback
            self.logger.error(f"Traceback: {traceback.format_exc()}")

    def auto_adjust_columns(self, writer, df, sheet_name):
        """Auto-adjust column widths for a specific sheet based on content"""
        try:
            if not hasattr(writer, 'book') or sheet_name not in writer.sheets:
                return
            
            worksheet = writer.sheets[sheet_name]
            
            # Get the maximum length for each column
            for col_idx, column in enumerate(df.columns):
                # Get the maximum length of the column header
                header_length = len(str(column))
                
                # Get the maximum length of the data in this column
                if not df.empty:
                    # Convert all values to string and get max length
                    data_length = df[column].astype(str).str.len().max()
                    # Add some padding (minimum 10 characters)
                    max_length = max(header_length, data_length, 10) + 2
                else:
                    max_length = header_length + 2
                
                # Set column width (Excel column width is approximately 0.9 * character count)
                column_width = min(max_length * 0.9, 50)  # Cap at 50 to avoid extremely wide columns
                worksheet.set_column(col_idx, col_idx, column_width)
            
            self.logger.info(f"Auto-adjusted column widths for sheet: {sheet_name}")
            
        except Exception as e:
            self.logger.warning(f"Could not auto-adjust columns for {sheet_name}: {str(e)}")

    def validate_output_data(self, df):
        try:
            self.logger.info("Validating output data...")
            validation_results = {
                'total_records': len(df),
                'missing_case_numbers': df[df['Case Number'].isna() | (df['Case Number'] == '')].shape[0],
                'missing_handlers': df[df['Assigned To'].isna() | (df['Assigned To'] == '')].shape[0],
                'duplicate_cases': df[df.duplicated(subset=['Case Number'], keep=False)].shape[0],
                'empty_rows': df[df.apply(lambda row: all(str(x).strip() == '' for x in row), axis=1)].shape[0]
            }
            self.logger.info("Validation Results:")
            for key, value in validation_results.items():
                self.logger.info(f"  {key}: {value}")
            critical_issues = []
            if validation_results['missing_case_numbers'] > 0:
                critical_issues.append(f"Missing case numbers: {validation_results['missing_case_numbers']}")
            if validation_results['missing_handlers'] > 0:
                critical_issues.append(f"Missing handler assignments: {validation_results['missing_handlers']}")
            if validation_results['duplicate_cases'] > 0:
                critical_issues.append(f"Duplicate case numbers: {validation_results['duplicate_cases']}")
            if validation_results['empty_rows'] > 0:
                critical_issues.append(f"Empty rows: {validation_results['empty_rows']}")
            if critical_issues:
                self.logger.warning("Critical validation issues found:")
                for issue in critical_issues:
                    self.logger.warning(f"  - {issue}")
            else:
                self.logger.info("No critical validation issues found")
            return validation_results, critical_issues
        except Exception as e:
            self.logger.error(f"Error validating output data: {str(e)}")
            return {}, [f"Validation error: {str(e)}"] 

    def create_sorting_summary_report(self, df, sheet_name="Main Output"):
        """Create a summary report of case sorting results"""
        try:
            if 'Status' not in df.columns:
                return
            
            self.logger.info(f"\n=== {sheet_name} - CASE SORTING SUMMARY ===")
            
            # Get status distribution
            status_counts = df['Status'].value_counts()
            total_cases = len(df)
            
            # Calculate percentages
            status_summary = []
            for status, count in status_counts.items():
                if pd.notna(status) and str(status).strip():
                    percentage = (count / total_cases) * 100
                    status_summary.append({
                        'Status': status,
                        'Count': count,
                        'Percentage': f"{percentage:.1f}%"
                    })
            
            # Sort by our custom order
            status_order = {
                'new': 1, 'in_progress': 2, 'in progress': 2, 'in progress today': 2,
                'needs follow up': 3, 'pending': 4, 'escalation': 5, 'closed': 6,
                'Closed': 6, 'bank/sutherland': 7, 'dnd': 8, 'DND': 8
            }
            
            status_summary.sort(key=lambda x: status_order.get(str(x['Status']).lower(), 999))
            
            # Log the summary
            self.logger.info(f"Total cases: {total_cases}")
            self.logger.info("Status distribution (sorted by priority):")
            for item in status_summary:
                self.logger.info(f"  {item['Status']}: {item['Count']} cases ({item['Percentage']})")
            
            # Show first few cases of each status for verification
            self.logger.info("\nFirst few cases by status:")
            for item in status_summary[:5]:  # Show top 5 statuses
                status = item['Status']
                status_cases = df[df['Status'] == status].head(3)
                if not status_cases.empty:
                    case_numbers = status_cases['Case Number'].astype(str).tolist()
                    self.logger.info(f"  {status}: {case_numbers}")
            
            self.logger.info(f"=== END {sheet_name} SORTING SUMMARY ===\n")
            
        except Exception as e:
            self.logger.error(f"Error creating sorting summary report: {str(e)}")

    def _assign_companies_cases(self, df, selected_handlers):
        """
        Assign cases to handlers with round-robin logic based on Email groups.
        Ensures all cases from the same email are assigned to the same handler.
        
        CRITICAL: Only assigns NEW cases. Cases with existing Assigned To from
        previous file are preserved (similar to handler sheet preservation logic).
        """
        try:
            if df.empty or not selected_handlers:
                self.logger.info("No cases or no handlers selected - skipping assignment")
                return df
            
            self.logger.info(f"Processing {len(df)} cases for assignment to {len(selected_handlers)} handlers...")
            
            # Ensure Assigned To column exists
            if 'Assigned To' not in df.columns:
                df['Assigned To'] = ''
            
            # Get valid emails (exclude empty/NaN)
            if 'Email' not in df.columns:
                self.logger.warning("No 'Email' column - cannot group by email")
                return df
            
            # Convert emails to lowercase string for grouping
            df['_email_group'] = df['Email'].astype(str).str.strip().str.lower()
            
            # CRITICAL: Identify which cases need assignment (those without a handler)
            # A case needs assignment if Assigned To is empty
            # Status is NOT checked - preserved cases may have various statuses but still need their handler kept
            needs_new_assignment = (
                df['Assigned To'].isna() | 
                (df['Assigned To'].astype(str).str.strip() == '') |
                (df['Assigned To'].astype(str).str.strip().str.lower().isin(['nan', 'none']))
            )
            
            new_cases_count = needs_new_assignment.sum()
            preserved_cases_count = len(df) - new_cases_count
            
            self.logger.info(f"Companies assignment analysis:")
            self.logger.info(f"  Preserved cases (already have handler): {preserved_cases_count}")
            self.logger.info(f"  New cases needing handler assignment: {new_cases_count}")
            
            if new_cases_count == 0:
                self.logger.info("No new cases need assignment - all handlers preserved from previous file")
                if '_email_group' in df.columns:
                    df = df.drop(columns=['_email_group'])
                return df
            
            # Get only the NEW cases that need assignment
            new_cases_df = df[needs_new_assignment].copy()
            
            # Identify unique email groups in NEW cases
            unique_emails = new_cases_df['_email_group'].unique()
            unique_emails = [e for e in unique_emails if e and e not in ['nan', 'none', '']]
            
            self.logger.info(f"Found {len(unique_emails)} unique email groups in new cases")
            
            # Create a map for assignments
            email_handler_map = {}
            
            # First pass: Check if any email group already has a handler assigned 
            # (from preserved cases with same email)
            for email in unique_emails:
                # Check if this email exists in preserved cases with a handler
                preserved_mask = (~needs_new_assignment) & (df['_email_group'] == email)
                existing_handlers = df.loc[preserved_mask, 'Assigned To'].dropna().unique()
                existing_handlers = [h for h in existing_handlers if h and str(h).strip() != '' and str(h).lower() not in ['nan', 'none']]
                
                if existing_handlers:
                    # Use the existing handler for this email group
                    email_handler_map[email] = existing_handlers[0]
                    self.logger.info(f"  Email '{email[:30]}...' - using preserved handler: {existing_handlers[0]}")
            
            # Second pass: Assign unassigned emails via round-robin
            handler_idx = 0
            assignments_made = 0
            
            for email in unique_emails:
                if email not in email_handler_map:
                    # Get next handler
                    handler = selected_handlers[handler_idx % len(selected_handlers)]
                    email_handler_map[email] = handler
                    handler_idx += 1
                    assignments_made += 1
            
            self.logger.info(f"Made {assignments_made} new email group assignments (round-robin)")
            
            # Apply assignments ONLY to NEW cases
            for email, handler in email_handler_map.items():
                mask = needs_new_assignment & (df['_email_group'] == email)
                df.loc[mask, 'Assigned To'] = handler
                
            # Handle NEW cases with no email (simple round robin)
            no_email_mask = needs_new_assignment & (
                (df['_email_group'].isin(['nan', 'none', ''])) | (df['_email_group'].isna())
            )
            no_email_count = no_email_mask.sum()
            
            if no_email_count > 0:
                self.logger.info(f"Assigning {no_email_count} new cases without email via simple round robin")
                for idx in df[no_email_mask].index:
                    handler = selected_handlers[handler_idx % len(selected_handlers)]
                    df.loc[idx, 'Assigned To'] = handler
                    handler_idx += 1
            
            # Log final assignment distribution
            self.logger.info("Final Companies assignment distribution:")
            for handler in selected_handlers:
                count = (df['Assigned To'] == handler).sum()
                self.logger.info(f"  {handler}: {count} cases")
            
            # Clean up
            if '_email_group' in df.columns:
                df = df.drop(columns=['_email_group'])
                
            return df
            
        except Exception as e:
            self.logger.error(f"Error assigning companies cases: {str(e)}")
            if '_email_group' in df.columns:
                df = df.drop(columns=['_email_group'])
            return df

    def protect_worksheet(self, writer, sheet_name, password='artadmin'):
        """Protect a worksheet with a password, allowing filtering and selection"""
        try:
            if not hasattr(writer, 'book') or sheet_name not in writer.sheets:
                self.logger.warning(f"Cannot protect sheet '{sheet_name}' - sheet or book not found")
                return
            
            worksheet = writer.sheets[sheet_name]
            
            # xlsxwriter worksheet protection
            options = {
                'select_locked_cells': True,
                'select_unlocked_cells': True,
                'format_cells': False,
                'format_columns': False,
                'format_rows': False,
                'insert_columns': False,
                'insert_rows': False,
                'insert_hyperlinks': False,
                'delete_columns': False,
                'delete_rows': False,
                'sort': True, # Allow sorting
                'autofilter': True, # Allow filtering
                'pivot_tables': False,
                'scenarios': False,
                'objects': False,
            }
            
            worksheet.protect(password, options)
            self.logger.info(f"Sheet '{sheet_name}' protected with password")
            
        except Exception as e:
            self.logger.error(f"Error protecting sheet '{sheet_name}': {str(e)}")